
## 새로운 로직개발 헤밍구조에 대한 통합버전 개발 쪽쟘, 멍텅구리 통합
## 실전 사용하면서 수정사항 반영 B1, B2 크기 조정 pyinstaller 실행 후 반영되는 값 적용
## 작지 양식 일부 수정 라이트케이스 크기 강제입력 부분 추가 24/03/05
## LED는 CD값이 1030인 경우는 기본 -100 처리 후 930이면 950으로, 즉 50단위로 생성해야 한는데, 30이면 30을 버리는 것이 아니라 50으로 만든다.
## 032 부품 치수선 블럭 수정 (분리 : 치수선과 레이져선 분리작업)
## 20240922 031 1400 W 이하일때 조건식 일부 수정 1,2페이지 LC홀 위치 관련
## 032 프레임홀 위치 수정 2024/11/01 
## 031 조립도 폴리갈 중앙에 표시 수정 
## 032 2페이지 프레임 3.2파이에서 3.2x10장공으로 수정(2025/01/24)
## 008_A 모델개발시작 2025/02/03 (에이원엘리베이터)
## 026 높이수정 및 재질변경 개발 2025/02/19 (비용절감 프로젝트)
## 본천장 개발 시작(2025/03/10)
## 008_A 전체 폭 수정 (2025/04/)

import math
import ezdxf
from ezdxf.enums import TextEntityAlignment
from datetime import datetime
import openpyxl
import os
import glob
import time
import os
import sys
import io
from datetime import datetime

import json
from gooey import Gooey, GooeyParser
import warnings
import re
import requests
        
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')    

# 경고 메시지 필터링
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")

# 전역 변수 초기화
saved_DimXpos = 0
saved_DimYpos = 0
saved_Xpos = 0
saved_Ypos = 0
saved_direction = "up"
saved_text_height = 0.22
saved_text_gap = 0.05
dimdistance = 0
dim_horizontalbase = 0
dim_verticalbase = 0
distanceXpos = 0
distanceYpos = 0

start_time = 0

workplace = ""
drawdate = None
deadlinedate = None
lctype = ""
secondord = ""
CW_raw = 0
CD_raw = 0
panel_thickness = 0
LC_height = 0
su = 0
LCframeMaterial = ""
LCplateMaterial = ""
input_CD = 0
input_CW = 0
CD = 0
CW = 0
drawdate_str = ""
deadlinedate_str = ""
T5_is =""
frontPanel, rearPanel, sidePanel = None, None, None
slotposition = 0
EE_raw = 0
AS_raw = 0
ceilingdivision = 0
SplitYpos1, SplitYpos2, SplitYpos3 = 0, 0, 0
ELYpos = 0
footOP = 0
headertype = None
# entrytype, doortype, capacity, emergencytype, baseMaterial, LHtype, headertype, Sehwa_Cutfromcenter = None
# OP_raw, emergencyRow, emergencyCol, fan_su, fantop, fanside, ceilingdivision=0

# 폴더 내의 모든 .xlsm 파일을 검색
application_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
excel_saved_file = os.path.join(application_path, 'ceiling_excel')
xlsm_files = glob.glob(os.path.join(excel_saved_file, '*.xlsm'))
jamb_ini = os.path.join(application_path, 'data', 'jamb.json')
license_file_path = os.path.join(application_path, 'data', 'hdsettings.json') # 하드디스크 고유번호 인식

# DXF 파일 로드
doc = ezdxf.readfile(os.path.join(application_path, '', 'block.dxf'))
msp = doc.modelspace()

# TEXTSTYLE 정의
text_style_name = 'JKW'  # 원하는 텍스트 스타일 이름
if text_style_name not in doc.styles:
    text_style = doc.styles.new(
        name=text_style_name,
        dxfattribs={
            'font': 'Arial.ttf',  # TrueType 글꼴 파일명            
        }
    )
else:
    text_style = doc.styles.get(text_style_name)    

# 찾은 .xlsm 파일 목록 출력
# for xlsm_file in xlsm_files:
#     print(xlsm_file)

def log_login():
    # PHP 파일의 URL 서버에서는 아이피를 저장한다. 업체 아이피를 기록한다.
    url = f"http://8440.co.kr/autopanel/savelog.php?company=미래기업&content={workplace}_{lctype}_{su}EA"

    # HTTP 요청 보내기
    response = requests.get(url)
    
    # 요청이 성공했는지 확인
    if response.status_code == 200:
        # print("logged successfully.")
        print(response.json())
    else:
        print("Failed to log login time.")
        print(response.text)   
        exit(1)

def parse_arguments_settings():
    parser = GooeyParser()
    settings = parser.add_argument_group('설정')
    settings.add_argument('--config', action='store_true', default=True,  help='라이센스')
    settings.add_argument('--password', widget='PasswordField', gooey_options={'visible': True, 'show_label': True}, help='비밀번호')
    return parser.parse_args()

def parse_arguments():
    parser = GooeyParser()

    group1 = parser.add_argument_group('라이트케이스')
    # group1.add_argument('--opt1', action='store_true',  help='기본')
    # group1.add_argument('--opt2', action='store_true',  help='도장홀')
    # group1.add_argument('--opt3', action='store_true',  help='쪽쟘 상판끝 라운드(추영덕소장)')    
    group1.add_argument('--opt1', action='store_true', default=True, help='기본')

    settings = parser.add_argument_group('설정')
    settings.add_argument('--config', action='store_true', help='라이센스')
    settings.add_argument('--password', widget='PasswordField', gooey_options={'visible': True, 'show_label': True}, help='비밀번호')

    return parser.parse_args()
def display_message():
    message = program_message.format('\n'.join(sys.argv[1:])).split('\n')
    delay = 1.5 / len(message)

    for line in message:
        print(line)
        time.sleep(delay)
# 환경설정 가져오기(하드공유 번호)
def load_env_settings():
    try:
        with open(license_file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return data.get("DiskID")
    except FileNotFoundError:
        return None
def get_current_disk_id():
    return os.popen('wmic diskdrive get serialnumber').read().strip()
# None이면 0을 리턴하는 함수
def validate_or_default(value):
    if value is None:
        return 0
    return value
def find_intersection(start1, end1, start2, end2):
    # 교차점 계산 (두 직선이 수직 및 수평으로 만나는 경우에 대해서만)
    if start1[0] == end1[0]:
        return (start1[0], start2[1])
    else:
        return (start2[0], start1[1])
def calculate_fillet_point(center, point, radius):
    # 필렛 접점 계산
    dx = point[0] - center[0]
    dy = point[1] - center[1]
    if abs(dx) > abs(dy):
        return (center[0] + radius * (1 if dx > 0 else -1), center[1])
    else:
        return (center[0], center[1] + radius * (1 if dy > 0 else -1))
def calculate_angle(center, point):
    """함수 설명:
    math.atan2(y, x): 점 (x, y)에 대한 아크탄젠트(atan2) 값을 반환합니다.
    이는 (0,0)을 기준으로 (x, y)까지의 반시계 방향으로 회전한 각도를 라디안 단위로 반환합니다.
    math.degrees(): atan2의 결과(라디안)를 **각도(도)**로 변환합니다.
    (point[1] - center[1], point[0] - center[0]): center에서 point까지의 상대적인 y, x 차이를 구합니다."
    """
    # 각도 계산
    return math.degrees(math.atan2(point[1] - center[1], point[0] - center[0]))
def add_90_degree_fillet(doc, start1, end1, start2, end2, radius):
    msp = doc.modelspace()

    # 교차점 찾기
    intersection_point = find_intersection(start1, end1, start2, end2)

    # 필렛 접점 계산
    point1 = calculate_fillet_point(intersection_point, end1, radius)
    point2 = calculate_fillet_point(intersection_point, end2, radius)

    # 각도 계산
    start_angle = calculate_angle(intersection_point, point1)
    end_angle = calculate_angle(intersection_point, point2)

    # 필렛 원호 그리기
    msp.add_arc(
        center=intersection_point,
        radius=radius,
        start_angle=start_angle,
        end_angle=end_angle,
        dxfattribs={'layer': '레이져'},
    )    
    return msp
def calculate_midpoint(point1, point2):
    return ((point1[0] + point2[0]) / 2, (point1[1] + point2[1]) / 2)
def calculate_circle_center(midpoint, radius, point1, point2):
    dx = point2[0] - point1[0]
    dy = point2[1] - point1[1]
    dist = math.sqrt(dx**2 + dy**2)
    factor = math.sqrt(radius**2 - (dist / 2)**2) / dist
    return (midpoint[0] - factor * dy, midpoint[1] + factor * dx)
def add_arc_between_points(doc, point1, point2, radius):
    msp = doc.modelspace()
    
    # 원의 중심 계산 (반지름이 마이너스일때는 point1, point2 변경)
    if radius < 0 :
        # 중점 계산
        midpoint = calculate_midpoint(point1, point2)
        center = calculate_circle_center(midpoint, radius * - 1, point2, point1)
    else:
        midpoint = calculate_midpoint(point1, point2)
        center = calculate_circle_center(midpoint, radius, point1, point2)

    # 각도 계산
    start_angle = calculate_angle(center, point1)
    end_angle = calculate_angle(center, point2)

    # 아크 그리기
    msp.add_arc(
        center=center,
        radius=radius,
        start_angle=start_angle,
        end_angle=end_angle,
        dxfattribs={'layer': '레이져'},
    )    
    return msp
def dim_leader_line(doc, start_x, start_y, end_x, end_y, text, layer='JKW', style='0', text_height=22):
    msp = doc.modelspace()

    text_style_name = 'JKW'
        
    # override 설정
    override_settings = {
        'dimasz': 15
    }    

    # 지시선 추가
    leader = msp.add_leader(
        vertices=[(start_x, start_y), (end_x, end_y)],  # 시작점과 끝점
        dxfattribs={
            'dimstyle': layer,
            'layer': layer
        },
        override=override_settings
    )

    # 텍스트 추가 (선택적)
    if text:
        msp.add_mtext(text, dxfattribs={
            'insert': (end_x + 10, end_y + 5),  # 텍스트 위치 조정
            'layer': style,
            'char_height': text_height,
            'style': text_style_name,
            'attachment_point': 1  # 텍스트 정렬 방식 설정
        })

    return leader
def dim_leader(doc, start_x, start_y, end_x, end_y, text, text_height=20, direction=None, option=None, distance=None):
    """  distance=13 # 글자크기에 대한 정의임  """
    global saved_DimXpos, saved_DimYpos, saved_text_height, saved_text_gap, saved_direction, dimdistance, dim_horizontalbase, dim_verticalbase, distanceXpos, distanceYpos 
    
    msp = doc.modelspace()
    layer = '0'        
    text_style_name = 'JKW'            
    # text_style_name = selected_dimstyle
    # override 설정
    override_settings = {
        'dimasz': 15
    }

    # 텍스트 위치 조정 및 꺾이는 지점 설정
    if option is None:    
        text_offset_x = 20
        text_offset_y = 20
    else:
        text_offset_x = 0
        text_offset_y = 0        
    
    if direction == 'leftToright':
        mid_x = end_x - text_offset_x
        mid_y = end_y  # 텍스트 앞에서 꺾임
        text_position = (end_x, end_y)
    elif direction == 'rightToleft':
        mid_x = end_x + text_offset_x
        mid_y = end_y  # 텍스트 앞에서 꺾임
        if distance is None:
            text_position = (end_x - len(text) * 13, end_y)
        else:
            text_position = (end_x - len(text) * distance, end_y)
    else:
        mid_x = (start_x + end_x) / 2
        mid_y = (start_y + end_y) / 2
        text_position = (end_x + text_offset_x, end_y + text_offset_y)

    # 지시선 추가
    leader = msp.add_leader(
        vertices=[(start_x, start_y), (mid_x, mid_y-text_height/2), (end_x, end_y-text_height/2)],  # 시작점, 중간점(문자 앞에서 꺾임), 끝점
        dxfattribs={
            'dimstyle': text_style_name,
            'layer': layer,
            'color': 3  # 녹색 (AutoCAD 색상 인덱스에서 3번은 녹색)
        },
        override=override_settings
    )

    if option is None:
        # 텍스트 추가 (선택적)
        if text:
            msp.add_mtext(text, dxfattribs={
                'insert': text_position,
                'layer': layer,
                'char_height': text_height,
                'style': text_style_name,
                'attachment_point': 1,  # 텍스트 정렬 방식 설정
                'color': 0  # 노란색 (AutoCAD 색상 인덱스에서 2번은 노란색)
            })

    return leader
def line(doc, x1, y1, x2, y2, layer=None):
    global saved_Xpos, saved_Ypos  # 전역 변수로 사용할 것임을 명시
    
    # 선 추가
    start_point = (x1, y1)
    end_point = (x2, y2)
    if layer:
        # 절곡선 22 layer는 ltscale을 조정한다
        if(layer=="22"):
            msp.add_line(start=start_point, end=end_point, dxfattribs={'layer': layer, 'ltscale' : 30})
        else:
            msp.add_line(start=start_point, end=end_point, dxfattribs={'layer': layer })
    else:
        msp.add_line(start=start_point, end=end_point)

    # 다음 선분의 시작점을 업데이트
    saved_Xpos = x2
    saved_Ypos = y2        
def lineto(doc, x, y, layer=None):
    global saved_Xpos, saved_Ypos  # 전역 변수로 사용할 것임을 명시
    
    # 현재 위치를 시작점으로 설정
    start_x = saved_Xpos
    start_y = saved_Ypos

    # 끝점 좌표 계산
    end_x = x
    end_y = y

    # 모델 공간 (2D)을 가져옴
    msp = doc.modelspace()

    # 선 추가
    start_point = (start_x, start_y)
    end_point = (end_x, end_y)
    if layer:
        msp.add_line(start=start_point, end=end_point, dxfattribs={'layer': layer})
    else:
        msp.add_line(start=start_point, end=end_point)

    # 다음 선분의 시작점을 업데이트
    saved_Xpos = end_x
    saved_Ypos = end_y
def lineclose(doc, start_index, end_index, layer='레이져'):    

    firstX, firstY = globals()[f'X{start_index}'], globals()[f'Y{start_index}']
    prev_x, prev_y = globals()[f'X{start_index}'], globals()[f'Y{start_index}']

    # start_index+1부터 end_index까지 반복합니다.
    for i in range(start_index + 1, end_index + 1):
        # 현재 인덱스의 좌표를 가져옵니다.
        curr_x, curr_y = globals()[f'X{i}'], globals()[f'Y{i}']

        # 이전 좌표에서 현재 좌표까지 선을 그립니다.
        line(doc, prev_x, prev_y, curr_x, curr_y, layer)

        # 이전 좌표 업데이트
        prev_x, prev_y = curr_x, curr_y

        print(f"prev_x {prev_x}" )
        print(f"prev_y {prev_y}" )
    
    # 마지막으로 첫번째 점과 연결
    line(doc, prev_x, prev_y, firstX , firstY, layer)        
def rectangle(doc, x1, y1, dx, dy, layer=None, offset=None):
    if offset is not None:
        # 네 개의 선분으로 직사각형 그리기 offset 추가
        line(doc, x1+offset, y1+offset, dx-offset, y1+offset, layer=layer)   
        lineto(doc, dx - offset, dy - offset, layer=layer)  
        lineto(doc, x1 + offset, dy - offset, layer=layer)  
        lineto(doc, x1 + offset, y1 + offset, layer=layer)  
    else:        
        # 네 개의 선분으로 직사각형 그리기
        line(doc, x1, y1, dx, y1, layer=layer)   
        line(doc, dx, y1, dx, dy, layer=layer)   
        line(doc, dx, dy, x1, dy, layer=layer)   
        line(doc, x1, dy, x1, y1, layer=layer)   
def add_angular_dim_2l(drawing, base, line1, line2, location=None, text=None, text_rotation=None, dimstyle='0', override=None, dxfattribs=None):
    # Create a new dimension line
    dim = drawing.dimstyle.add(
        'EZ_ANGULAR_2L',
        dxfattribs={
            'dimstyle': dimstyle,
            'dimtad': 1,  # Place text above dimension line
            'dimtih': False,  # Align text horizontally to dimension line
            'dimtoh': False,  # Align text outside horizontal
            'dimdec': 1
        }
    )
    
    # If override is provided, update the dimension style
    if override:
        dim.update(override)
    
    # Create angular dimension
    angular_dim = drawing.add_angular_dim_2l(
        base=base,
        line1=line1,
        line2=line2,
        location=location,
        text=text,
        text_rotation=text_rotation,
        dimstyle=dim.name,
        override=override,
        dxfattribs=dxfattribs,
    )
    
    # Render the dimension
    angular_dim.render()
    return angular_dim
def dim_angular(doc, x1, y1, x2, y2, x3, y3, x4, y4, distance=80, direction="left", dimstyle="JKW"):
    msp = doc.modelspace()

    # 두 선분의 좌표
    p1 = (x1, y1)
    p2 = (x2, y2)
    p3 = (x3, y3)
    p4 = (x4, y4)

    # 치수선의 위치 계산
    base_x = (p1[0] + p2[0] + p3[0] + p4[0]) / 4
    base_y = (p1[1] + p2[1] + p3[1] + p4[1]) / 4

    if direction=="left":
        base = (base_x - distance , base_y)
        # 각도 치수선 추가
        dimension = msp.add_angular_dim_2l(
            base=base,  # 치수선 위치
            line1=(p1, p2),  # 첫 번째 선의 시작점과 끝점
            line2=(p3, p4),  # 두 번째 선의 시작점과 끝점
            dimstyle=dimstyle,  # 치수 스타일        
            override={'dimtxt': 0.27, 'dimgap': 0.02, 'dimscl' : 1, 'dimlfac' : 1, 'dimclrt': 7, 'dimdsep': 46, 'dimdec': 0 } # 선색 백색 소수점 . 표기
        )

    elif direction=="up":
        base = (base_x , base_y + distance )          
        # "dimtad": 1, # 0=center; 1=above; 4=below;  
        # 각도 치수선 추가
        dimension = msp.add_angular_dim_2l(
            base=base,  # 치수선 위치
            line1=(p1, p2),  # 첫 번째 선의 시작점과 끝점
            line2=(p3, p4),  # 두 번째 선의 시작점과 끝점
            dimstyle=dimstyle,  # 치수 스타일        
            override={'dimtxt': 0.27, 'dimgap': 0.02, 'dimscl' : 1, 'dimlfac' : 1, 'dimclrt': 7, 'dimdsep': 46, 'dimdec': 0 } # 선색 백색 소수점 . 표기
        )        
    elif direction=="down":
        base = (base_x , base_y - distance )                    
    else:
        base = (base_x + distance , base_y)    
        # 각도 치수선 추가
        dimension = msp.add_angular_dim_2l(
            base=base,  # 치수선 위치
            line1=(p1, p2),  # 첫 번째 선의 시작점과 끝점
            line2=(p3, p4),  # 두 번째 선의 시작점과 끝점
            dimstyle=dimstyle,  # 치수 스타일        
            override={'dimtxt': 0.27, 'dimgap': 0.02, 'dimscl' : 1, 'dimlfac' : 1, 'dimclrt': 7, 'dimdsep': 46, 'dimdec': 0 } # 선색 백색 소수점 . 표기
        )        

    # 치수선의 기하학적 형태 생성
    dimension.render()
    return dimension
def dim_diameter(doc, center, diameter, angle, dimstyle="JKW", override=None):
    msp = doc.modelspace()
    
    # 기본 지름 치수선 추가
    dimension = msp.add_diameter_dim(
        center=center,  # 원의 중심점
        radius=diameter/2,  # 반지름
        angle=angle,    # 치수선 각도
        dimstyle=dimstyle,  # 치수 스타일
        override={"dimtoh": 1}    # 추가 스타일 설정 (옵션) 지시선이 한번 꺾여서 글자각도가 표준형으로 나오는 옵션
    )
    
    # 치수선의 기하학적 형태 생성
    dimension.render()
def dim_string(doc, x1, y1, x2, y2, dis,  textstr,  text_height=0.22, text_gap=0.05, direction="up"):
    global saved_DimXpos, saved_DimYpos, saved_text_height, saved_text_gap, saved_direction, dimdistance, dim_horizontalbase, dim_verticalbase

    msp = doc.modelspace()
    dim_style = 'JKW'
    layer = "JKW"

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점이 있는지 확인
    if dimension_value % 1 != 0:
        dimdec = 1  # 소수점이 있는 경우 소수점 첫째 자리까지 표시
    else:
        dimdec = 0  # 소수점이 없는 경우 소수점 표시 없음

    # override 설정
    override_settings = {
        'dimtxt': text_height,
        'dimgap': text_gap,
        'dimscl': 1,
        'dimlfac': 1,
        'dimclrt': 7,
        'dimdsep': 46,
        'dimdec': dimdec,
        'dimtix': 1,  # 치수선 내부에 텍스트를 표시 (필요에 따라)
        'dimtad': 1 # 치수선 상단에 텍스트를 표시 (필요에 따라)               
    }

    # 방향에 따른 치수선 추가
    if direction == "up":
        dimension = msp.add_linear_dim(
            base=(x1, y1 + dis),
            dimstyle=dim_style,
            p1=(x1, y1),
            p2=(x2, y2),
            dxfattribs={'layer': layer},
            text = textstr,
            override=override_settings
        )
    elif direction == "down":
        dimension = msp.add_linear_dim(
            dimstyle=dim_style,
            base=(x1, y1 - dis),
            p1=(x1, y1),
            p2=(x2, y2),
            dxfattribs={'layer': layer},
            text = textstr,            
            override=override_settings
        )
    elif direction == "left":        
        dimension =  msp.add_linear_dim(
            dimstyle=dim_style,
            base=(x1 - dis, y1),
            angle=90,
            p1=(x1, y1),
            p2=(x2, y2),
            dxfattribs={'layer': layer},
            text = textstr,  
            override=override_settings
        )     
    elif direction == "right":        
        dimension =  msp.add_linear_dim(
            dimstyle=dim_style,
            base=(x1 + dis, y1),
            p1=(x1, y1),
            p2=(x2, y2),
            dxfattribs={'layer': layer},
            text = textstr,  
            angle=270,            
            override=override_settings
        )
    else:
        raise ValueError("Invalid direction. Use 'up', 'down', or 'aligned'.")

    dimension.render()
    return dimension

def d(doc, x1, y1, x2, y2, dis, text_height=0.22, text_gap=0.05, direction="up", option=None, startoption=None, text=None) :
    global saved_DimXpos, saved_DimYpos, saved_text_height, saved_text_gap, saved_direction, dimdistance, dim_horizontalbase, dim_verticalbase, distanceXpos, distanceYpos

    # Option 처리
    if option == 'reverse':
        x1, x2 = x2, x1
        y1, y2 = y2, y1     
        saved_DimXpos, saved_DimYpos = x2, y2
    else:
        saved_DimXpos, saved_DimYpos = x2, y2

    dimdistance = dis
    saved_text_height = text_height
    saved_text_gap = text_gap
    saved_direction = direction

    # 연속선 구현을 위한 구문
    if startoption is None:        
        if direction == "left":             
            distance = min(x1, x2) - dis            
            dim_horizontalbase = distance
        elif direction == "right":   
            distance = max(x1, x2) + dis                       
            dim_horizontalbase = distance 
        elif direction == "up":   
            distance = max(y1, y2) + dis            
            dim_verticalbase = distance
        elif direction == "down":     
            distance = min(y1, y2) - dis                                  
            dim_horizontalbase = distance
    else:
        if direction == "left":             
            distance = distanceXpos
        elif direction == "right":                        
            distance = distanceXpos
        elif direction == "up":            
            distance = distanceYpos
        elif direction == "down":     
            distance = distanceYpos
     

    # flip을 선언하면 치수선의 시작과 끝을 바꾼다. 저장된 좌표는 지장없다.
    # 치수선의 시작점 끝점에 따라 치수선이 나오는 것을 만들기 위함이다. 
    # 연속치수선때는 좌표가 바뀌면 안되기때문에 고려한 부분이다.

    msp = doc.modelspace()
    dim_style = 'JKW'
    layer = "JKW"

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점 확인
    dimdec = 1 if dimension_value % 1 != 0 else 0

    # override 설정
    override_settings = {      
        'dimtxt': text_height,  
        'dimgap': text_gap if text_gap is not None else 0.05,  # 여기에서 dimgap에 기본값 설정
        'dimscl': 1,
        'dimlfac': 1,
        'dimclrt': 7,
        'dimdsep': 46,
        'dimdec': dimdec,
        'dimtix': 1, 
        'dimtad': 1  
    }

    # 방향에 따른 치수선 추가
    if direction == "up":
        add_dim_args = {
            'base': ((x1+x2)/2, distance),
            'dimstyle': dim_style,
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        # 절대값으로 거리를 저장
        if startoption is None:
            distanceYpos = max(y1, y2) + dis           
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "down":
        add_dim_args = {
            'dimstyle': dim_style,
            'base': ((x1+x2)/2, distance),
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        # 절대값으로 거리를 저장
        if startoption is None:
            distanceYpos = min(y1, y2) - dis         
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "left":
        if option == 'reverse':   
            add_dim_args = {
                'dimstyle': dim_style,
                'base': (distance, (y1+y2)/2),
                'angle': 90,
                'p1': (x1, y1),
                'p2': (x2, y2),
                'override': override_settings
            }
            # 절대값으로 거리를 저장
            if startoption is None:
                distanceXpos = min(x1, x2) - dis 
        else:
            add_dim_args = {
                'dimstyle': dim_style,
                'base': (distance, (y1+y2)/2),
                'angle': 90,
                'p1': (x2, y2),
                'p2': (x1, y1),
                'override': override_settings
            }       
            # 절대값으로 거리를 저장
            if startoption is None:
                distanceXpos = min(x1, x2) - dis 
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "right":
        add_dim_args = {
            'dimstyle': dim_style,
            'base': (distance, (y1+y2)/2),
            'angle': 90,
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        # 절대값으로 거리를 저장
        if startoption is None:
            distanceXpos = max(x1, x2) + dis         
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "aligned":
        add_dim_args = {
            'dimstyle': dim_style,
            'points': [(x1, y1), (x2, y2)],
            'distance': dis,
            'dxfattribs': {'layer': layer},
            'override': override_settings
        }
        if text is not None :
            add_dim_args['text'] = text
        return msp.add_multi_point_linear_dim(**add_dim_args)

    else:
        raise ValueError("Invalid direction. Use 'up', 'down', 'left', 'right', or 'aligned'.")

def dc(doc, x, y, distance=None, option=None, text=None) :    
    global saved_DimXpos, saved_DimYpos, saved_text_height, saved_text_gap, saved_direction, dimdistance, dim_horizontalbase, dim_verticalbase, distanceXpos, distanceYpos

    x1 = saved_DimXpos
    y1 = saved_DimYpos
    x2 = x
    y2 = y        

    if distance is not None :
        dimdistance = distance    

    # reverse 옵션 처리
    if option == 'reverse':
        x1, x2 = x2, x1
        y1, y2 = y2, y1                

    d(doc, x1, y1, x2, y2, dimdistance, text_height=saved_text_height, text_gap=saved_text_gap, direction=saved_direction, startoption='continue', text=text)

def dim(doc, x1, y1, x2, y2, dis, text_height=0.22, text_gap=0.05, direction="up", option=None, startoption=None, text=None, localdimstyle=None):
    global saved_DimXpos, saved_DimYpos, saved_text_height, saved_text_gap, saved_direction, dimdistance, dim_horizontalbase, dim_verticalbase, dim_style

    # Option 처리
    if option == 'reverse':
        saved_DimXpos, saved_DimYpos = x1, y1
    else:
        saved_DimXpos, saved_DimYpos = x2, y2

    dimdistance = dis
    saved_text_height = text_height
    saved_text_gap = text_gap
    saved_direction = direction

    # 연속선 구현을 위한 구문
    if startoption is None:
        if direction == "left":             
            dim_horizontalbase = dis - (x1 - x2)
        elif direction == "right":                        
            dim_horizontalbase = dis 
        elif direction == "up":            
            dim_verticalbase = dis
        elif direction == "down":                        
            dim_verticalbase = dis   

    # flip을 선언하면 치수선의 시작과 끝을 바꾼다. 저장된 좌표는 지장없다.
    # 치수선의 시작점 끝점에 따라 치수선이 나오는 것을 만들기 위함이다. 
    # 연속치수선때는 좌표가 바뀌면 안되기때문에 고려한 부분이다.

    msp = doc.modelspace()
    if localdimstyle is None :
        dim_style = 'JKW'
        layer = "JKW"
    elif localdimstyle == '0.2 JKW':
        dim_style = '0.2 JKW'
    elif localdimstyle == '0.5 JKW':
        dim_style = '0.5 JKW'        

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점 확인
    dimdec = 1 if dimension_value % 1 != 0 else 0

    # override 설정
    override_settings = {      
        'dimtxt': text_height,  
        'dimgap': text_gap if text_gap is not None else 0.05,  # 여기에서 dimgap에 기본값 설정
        'dimscl': 1,
        'dimlfac': 1,
        'dimclrt': 7,
        'dimdsep': 46,
        'dimdec': dimdec,
        'dimtix': 1, 
        'dimtad': 1  
    }

    # 방향에 따른 치수선 추가
    if direction == "up":
        add_dim_args = {
            'base': (x1, y1 + dis),
            'dimstyle': dim_style,
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "down":
        add_dim_args = {
            'dimstyle': dim_style,
            'base': (x1, y1 - dis),
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "left":
        if option == 'reverse':   
            add_dim_args = {
                'dimstyle': dim_style,
                'base': (x2 - dis, y2),
                'angle': 90,
                'p1': (x1, y1),
                'p2': (x2, y2),
                'override': override_settings
            }
        else:
            add_dim_args = {
                'dimstyle': dim_style,
                'base': (x1 - dis, y1),
                'angle': 90,
                'p1': (x2, y2),
                'p2': (x1, y1),
                'override': override_settings
            }        
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "right":
        add_dim_args = {
            'dimstyle': dim_style,
            'base': (x1 + dis, y1),
            'angle': 90,
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "aligned":
        add_dim_args = {
            'dimstyle': dim_style,
            'points': [(x1, y1), (x2, y2)],
            'distance': dis,
            'dxfattribs': {'layer': layer},
            'override': override_settings
        }
        if text is not None :
            add_dim_args['text'] = text
        return msp.add_multi_point_linear_dim(**add_dim_args)

    else:
        raise ValueError("Invalid direction. Use 'up', 'down', 'left', 'right', or 'aligned'.")

def dimcontinue(doc, x, y, distance=None, option=None) :    
    global saved_DimXpos
    global saved_DimYpos
    global saved_text_height
    global saved_text_gap
    global saved_direction
    global dimdistance
    global dim_horizontalbase
    global dim_verticalbase

    x1 = saved_DimXpos
    y1 = saved_DimYpos
    x2 = x
    y2 = y        



# 방향에 대한 정의를 하고 연속적으로 차이를 감지해서 처리하는 것이다.
    if saved_direction=="left" :
        dimdistance = dim_horizontalbase
        # 재계산해야 함.        
        dim_horizontalbase = dimdistance - (x1 - x2)        
    if saved_direction=="right" :
        dimdistance = dim_horizontalbase 
        # 재계산해야 함.
        dim_horizontalbase = dimdistance - (x2 - x1)
    if saved_direction=="up" :
        dimdistance = dim_verticalbase
        # 재계산해야 함.
        dim_verticalbase = dimdistance - (y2 - y1)
    if saved_direction=="down" :
        dimdistance = dim_verticalbase 
        # 재계산해야 함.
        dim_verticalbase = dimdistance - (y1 - y2)

    if distance is not None :
        dimdistance = distance    

    # reverse 옵션 처리
    if option == 'reverse':
        x1, x2 = x2, x1
        y1, y2 = y2, y1                

    dim(doc, x1, y1, x2, y2, dimdistance, text_height=saved_text_height, text_gap=saved_text_gap, direction=saved_direction, startoption='continue', localdimstyle = dim_style)

def dimto(doc, x2, y2, dis, text_height=0.20, text_gap=None, option=None) :    
    global saved_DimXpos
    global saved_DimYpos
    global saved_text_height
    global saved_text_gap
    global saved_direction

    if text_gap is not None :
        text_gap = saved_text_gap
        saved_text_gap = text_gap

    if text_height is not None :
        text_height = saved_text_height
        saved_text_height = text_height

    dim(doc, saved_DimXpos, saved_DimYpos, x2, y2, dis, text_height=text_height, text_gap=text_gap, direction=saved_direction, option=option)

def create_vertical_dim(doc, x1, y1, x2, y2, dis, angle, layer=None, text_height=0.22, text_gap=0.05):
    msp = doc.modelspace()
    dim_style = layer  # 치수 스타일 이름
    points = [(x1, y1), (x2, y2)]

    if angle==None :
        angle = 270

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점이 있는지 확인
    if dimension_value % 1 != 0:
        dimdec = 1  # 소수점이 있는 경우 소수점 첫째 자리까지 표시
    else:
        dimdec = 1  # 소수점이 없는 경우 소수점 표시 없음    

    return msp.add_multi_point_linear_dim(
        base=(x1 + dis if angle == 270 else x1 - dis , y1),  #40은 보정
        points = points,
        angle = angle,
        dimstyle = dim_style,
        discard = True,
        dxfattribs = {'layer': layer},
        # 치수 문자 위치 조정 (0: 치수선 위, 1: 치수선 옆) 'dimtmove': 1 
        # override={'dimtxt': text_height, 'dimscl': 1, 'dimlfac': 1, 'dimclrt': 7, 'dimdec': dimdec, 'dimdsep': 46, 'dimtmove': 3  }
        override = {'dimtxt': text_height, 'dimgap': text_gap, 'dimscl': 1, 'dimlfac': 1, 'dimclrt': 7, 'dimdec': dimdec, 'dimdsep': 46, 'dimtmove': 3  }
    )

def dim_vertical_right(doc, x1, y1, x2, y2, dis, layer="JKW", text_height=0.22,  text_gap=0.07, angle=None):
    return create_vertical_dim(doc, x1, y1, x2, y2, dis, angle, layer, text_height, text_gap)

def dim_vertical_left(doc, x1, y1, x2, y2, dis, layer=None, text_height=0.22,  text_gap=0.07):
    return create_vertical_dim(doc, x1, y1, x2, y2, dis, 90, layer, text_height, text_gap)
  
def create_vertical_dim_string(doc, x1, y1, x2, y2, dis, angle, textstr, text_height=0.22, text_gap=0.07):    
    msp = doc.modelspace()
    dim_style = 'JKW'
    layer = "JKW"
    points = [(x1, y1), (x2, y2)]

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점이 있는지 확인
    if dimension_value % 1 != 0:
        dimdec = 1  # 소수점이 있는 경우 소수점 첫째 자리까지 표시
    else:
        dimdec = 1  # 소수점이 없는 경우 소수점 표시 없음    

    return msp.add_multi_point_linear_dim(
        base=(x1 + dis if angle == 270 else x1 - dis , y1),  #40은 보정
        points=points,
        angle=angle,
        dimstyle=dim_style,
        discard=True,
        dxfattribs={'layer': layer},           
        # 치수 문자 위치 조정 (0: 치수선 위, 1: 치수선 옆) 'dimtmove': 1 
        override={'dimtxt': text_height, 'dimgap': text_gap, 'dimscl': 1, 'dimlfac': 1, 'dimclrt': 7, 'dimdec': dimdec, 'dimdsep': 46, 'dimtmove': 3 , 'text' : textstr      }
    )

def dim_vertical_right_string(doc, x1, y1, x2, y2, dis, textstr, text_height=0.22,  text_gap=0.07):
    return create_vertical_dim_string(doc, x1, y1, x2, y2, dis, 270, textstr, text_height, text_gap)

def dim_vertical_left_string(doc, x1, y1, x2, y2, dis, textstr, text_height=0.22,  text_gap=0.07):
    return create_vertical_dim_string(doc, x1, y1, x2, y2, dis, 90, textstr, text_height, text_gap)

def draw_Text_direction(doc, x, y, size, text, layer=None, rotation = 90):
    # 모델 공간 (2D)을 가져옴
    msp = doc.modelspace()

    # MText 객체 생성
    mtext = msp.add_mtext(
        text,  # 텍스트 내용
        dxfattribs={
            'layer': layer,  # 레이어 지정
            'style': text_style_name,  # 텍스트 스타일 지정
            'char_height': size,  # 문자 높이 (크기) 지정
        }
    )

    # MText 위치와 회전 설정
    mtext.set_location(insert=(x, y), attachment_point=1, rotation=rotation)

    return mtext

def draw_Text(doc, x, y, size, text, layer=None):
    # 모델 공간 (2D)을 가져옴
    msp = doc.modelspace()

    # 텍스트 추가 및 생성된 Text 객체 가져오기
    text_entity = msp.add_text(
        text,  # 텍스트 내용
        dxfattribs={
            'layer': layer,  # 레이어 지정
            'style': text_style_name,  # 텍스트 스타일 지정
            'height': size,  # 텍스트 높이 (크기) 지정
        }
    )

    # Text 객체의 위치 설정
    # text_entity.set_placement((x, y), align=TextEntityAlignment.CENTER)  # 텍스트 위치 및 정렬 설정
    # Text 객체의 위치 설정 (가로 및 세로 중앙 정렬)
    text_entity.set_placement((x, y), align=TextEntityAlignment.MIDDLE_LEFT)

def draw_circle(doc, center_x, center_y, radius, layer='0', color='7'):
    """
    DXF 문서에 원을 그리는 함수
    :param doc: ezdxf 문서 객체
    :param center_x: 원의 중심 x 좌표
    :param center_y: 원의 중심 y 좌표
    :param radius: 원의 반지름
    :param layer: 원을 추가할 레이어 이름 (기본값은 '0')
    지름으로 수정 /2 적용
    """
    msp = doc.modelspace()
    circle = msp.add_circle(center=(center_x, center_y), radius=radius/2, dxfattribs={'layer': layer, 'color' : color})
    return circle
def draw_arc_Global(doc, x1, y1, x2, y2, radius, direction, layer='레이져'): # radius는 반지름을 넣는다. 지름이 아님
    msp = doc.modelspace()
    
    # 점들과 반지름을 이용하여 중점과 원의 중심 계산
    midpoint = calculate_midpoint((x1, y1), (x2, y2))
    center = calculate_circle_center(midpoint, radius, (x1, y1), (x2, y2))

    # 시작 각도와 끝 각도 계산
    start_angle = calculate_angle(center, (x1, y1))
    end_angle = calculate_angle(center, (x2, y2))

    # 방향에 따라 각도 조정
    if direction == 'up' or direction == 'down':
        if start_angle > end_angle:
            start_angle, end_angle = end_angle, start_angle
        if direction == 'down':
            start_angle += 180
            end_angle += 180
    elif direction == 'left' or direction == 'right':
        if start_angle > end_angle:
            start_angle, end_angle = end_angle, start_angle
        if direction == 'right':
            start_angle += 180
            end_angle += 180

    # 아크 그리기
    msp.add_arc(
        center=center,
        radius=radius,
        start_angle=start_angle,
        end_angle=end_angle,
        dxfattribs={'layer': layer},
    )
    return msp
def draw_arc_slot(doc, center, radius, start_angle, end_angle, layer):
    """
    Draws an arc in the DXF document.
    
    Parameters:
    doc (ezdxf.document): The DXF document to draw on.
    center (tuple): The (x, y) coordinates of the arc's center.
    radius (float): The radius of the arc.
    start_angle (float): The starting angle of the arc in degrees.
    end_angle (float): The ending angle of the arc in degrees.
    layer (str): The layer to draw the arc on.
    """
    msp = doc.modelspace()
    msp.add_arc(
        center=center,
        radius=radius,
        start_angle=start_angle,
        end_angle=end_angle,
        dxfattribs={'layer': layer}
    )
def draw_slot(doc, x, y, size, direction="가로", option=None, layer='0'):
    """
    Draws a slot (장공) with specified parameters in a DXF document.
    
    Parameters:
    doc (ezdxf.document): The DXF document to draw on.
    x (float): The x-coordinate of the slot's center.
    y (float): The y-coordinate of the slot's center.
    size (str): The size of the slot in the format 'WxH' (e.g., '8x16').
    direction (str): The direction of the slot, either '가로' (horizontal) or '세로' (vertical). Default is '가로'.
    option (str): Optional feature for the slot. If 'cross', draws center lines extending beyond the slot. Default is None.
    layer (str): The layer to draw the slot on. Default is '0'.
    """    
    msp = doc.modelspace()

    # size를 분리하여 폭과 높이 계산
    width, height = map(float, size.lower().split('x'))

    print(f"width : {width}, height : {height}")

    if direction == "가로":
        slot_length = height - width
        slot_width = width
        radius = slot_width / 2
    else:  # 세로
        slot_length = width
        slot_width = height - width
        radius = slot_length / 2

    

    # 중심점을 기준으로 시작점과 끝점 계산
    if direction == "가로":
        start_point = (x - slot_length / 2, y)
        end_point = (x + slot_length / 2, y)
    else:  # 세로
        start_point = (x, y - slot_length / 2)
        end_point = (x, y + slot_length / 2)

    # 직선 부분 그리기
    if direction == "가로":
        msp.add_line((start_point[0], start_point[1] + radius), (end_point[0], end_point[1] + radius), dxfattribs={'layer': layer})
        msp.add_line((start_point[0], start_point[1] - radius), (end_point[0], end_point[1] - radius), dxfattribs={'layer': layer})
    else:  # 세로
        msp.add_line((start_point[0] + radius, start_point[1]), (end_point[0] + radius, end_point[1]), dxfattribs={'layer': layer})
        msp.add_line((start_point[0] - radius, start_point[1]), (end_point[0] - radius, end_point[1]), dxfattribs={'layer': layer})

    # 양 끝의 반원 그리기
    if direction == "가로":
        draw_arc_slot(doc, start_point, radius, 90, 270, layer)  # 반원 방향 수정
        draw_arc_slot(doc, end_point, radius, 270, 90, layer)  # 반원 방향 수정
    else:  # 세로
        draw_arc_slot(doc, start_point, radius, 180, 360, layer)  # 반원 방향 수정
        draw_arc_slot(doc, end_point, radius, 0, 180, layer)  # 반원 방향 수정

    # 옵션이 "cross"인 경우 중심선을 추가로 그리기
    if option == "cross":
        if direction == "가로":
            msp.add_line((x - slot_length / 2 - 8, y), (x + slot_length / 2 + 8, y), dxfattribs={'layer': 'CL'})
            msp.add_line((x, y - slot_width / 2 - 4), (x, y + slot_width / 2 + 4), dxfattribs={'layer': 'CL'})
        else:  # 세로
            msp.add_line((x - slot_width / 2 - 4, y), (x + slot_width / 2 + 4, y), dxfattribs={'layer': 'CL'})
            msp.add_line((x, y - slot_length / 2 - 8), (x, y + slot_length / 2 + 8), dxfattribs={'layer': 'CL'})

    return msp
# 중심선 적색표기 색상 1 적용 적색
def circle_cross(doc, center_x, center_y, radius, layer='0', color='7'):
    """
    DXF 문서에 원을 그리는 함수
    :param doc: ezdxf 문서 객체
    :param center_x: 원의 중심 x 좌표
    :param center_y: 원의 중심 y 좌표
    :param radius: 원의 반지름
    :param layer: 원을 추가할 레이어 이름 (기본값은 '0')
    지름으로 수정 /2 적용
    """
    draw_circle(doc, center_x, center_y, radius, layer=layer, color=color)
    # 적색 십자선 그려주기    
    line(doc, center_x - radius/2 - 5, center_y, center_x +  radius/2 + 5, center_y, layer="CEN" )
    line(doc, center_x , center_y - radius/2 - 5, center_x , center_y +  radius/2 + 5, layer="CEN" )
    return circle_cross

def extract_abs(a, b):
    return abs(a - b)    

# 도면틀 삽입
def insert_block(x, y, block_name, layer=None):    
    scale = 1
    insert_point = (x, y, scale)

    if layer is None :
        # 블록 삽입하는 방법           
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': scale,
            'yscale': scale,
            'rotation': 0
        })       
    else:
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': scale,
            'yscale': scale,
            'rotation': 0,
            'layer' : layer
        })               

# 도면틀 삽입
def insert_frame( x, y, scale , sep , text):    
    if(sep =="drawings_frame"):
        block_name = "drawings_frame"
        insert_point = (x, y, scale)

        # 블록 삽입하는 방법           
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': scale,
            'yscale': scale,
            'rotation': 0
        })

        textstr = f"{text}"       
        draw_Text(doc, x + 2220*scale, y + 220*scale , 23*scale , str(textstr), '0')
        draw_Text(doc, x + (2200+160)*scale, y + (220-88)*scale   , 23*scale , "Light Case Accessory", '0')
        draw_Text(doc, x + (2200-320)*scale, y + (220-88)*scale   , 23*scale , formatted_date , '0')

def envsettings():
    # 하드디스크 고유번호를 가져오는 코드 (시스템에 따라 다를 수 있음)
    disk_id = os.popen('wmic diskdrive get serialnumber').read().strip()
    data = {"DiskID": disk_id}
    with open(license_file_path, 'w', encoding='utf-8') as file:
        json.dump(data, file)

# LED바 길이 계산 (1030, 등도 계산되도록 로직개발)
def CalculateLEDbar(car_length):
    # Subtract 100 from the input value
    result = car_length - 100

    # Adjust to the next lower or equal multiple of 50
    if result % 100 > 50:
        result = (result // 100) * 100 + 100
    else:
        if 0 < result % 100 <= 50:
            # For values just over a multiple of 100, bump up to the next 50
            result = (result // 100) * 100 + 50
        else:
            # Otherwise, round down to the nearest 100
            result = (result // 100) * 100 

    return result

# LED바 길이 계산 031모델 -300
def CalculateLEDbar031(car_length):    
    # CW-300
    result = car_length - 300

    # Adjust to the next lower or equal multiple of 50
    if result % 100 > 50:
        result = (result // 100) * 100 + 100
    else:
        if 0 < result % 100 <= 50:
            # For values just over a multiple of 100, bump up to the next 50
            result = (result // 100) * 100 + 50
        else:
            # Otherwise, round down to the nearest 100
            result = (result // 100) * 100 

    return result

def calculate_light_positions_adjusted_031(dataCD):
    # Base coordinates
    y_coord1 = 125
    y_coord_last = dataCD - 125

    # Check if CD is over 2000 for additional coordinates calculation
    if dataCD > 2000:
        # Middle two coordinates with a gap
        midpoint = dataCD / 2
        y_coord4 = midpoint - 185
        y_coord5 = midpoint + 185

        # Calculating other coordinates with equidistant spacing
        interval = (y_coord4 - y_coord1) / 3
        y_coord2, y_coord3 = y_coord1 + interval, y_coord1 + 2 * interval
        y_coord6, y_coord7 = y_coord5 + interval, y_coord5 + 2 * interval

        # Adjusting coordinates to round off and add truncated decimals
        y_coord2, y_coord3, y_coord6 = adjust_coordinates(y_coord2, y_coord3, y_coord6)
        y_coord7 = round(y_coord7)

        all_y_coords = sorted([y_coord1, y_coord2, y_coord3, y_coord4, y_coord5, y_coord6, y_coord7, y_coord_last])
    else:
        # For CD 2000 or less, calculate additional coordinates
        midpoint = dataCD / 2
        y_coord2 = midpoint - 185
        y_coord3 = midpoint + 185

        # Calculating the additional coordinates
        y_coord4 = (y_coord1 + y_coord2) / 2
        y_coord5 = (y_coord3 + y_coord_last) / 2

        all_y_coords = sorted([y_coord1, y_coord4, y_coord2, y_coord3, y_coord5, y_coord_last])

    return all_y_coords

# Function to adjust coordinates with rounding and adding truncated decimals
def adjust_coordinates(coord1, coord2, coord3):
    adjusted_coord1 = round(coord1)
    decimal_part = coord1 - adjusted_coord1
    adjusted_coord2 = round(coord2 + decimal_part)
    adjusted_coord3 = round(coord3 + (coord2 - adjusted_coord2))

    return adjusted_coord1, adjusted_coord2, adjusted_coord3

def calculate_lc_position_031(inputCW, inputCD):
    # Calculate Y-coordinates for top and bottom holes
    Topholey = inputCD - 40
    Bottomy = 40

    # Initialize the list of x-coordinates for top and bottom holes
    Topholex = []
    Bottomx = []

    # CW에서 양쪽의 여유분 230을 제외한 값을 3등분
    divided_space = (inputCW - 230) // 3
    end_digit = divided_space % 10

    # Determine the number of holes and calculate their positions
    if inputCW > 1400:
        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution
        x1 = 115
        space_available = inputCW - 115 * 2

        if end_digit == 0: 
            # 정확히 나누어 떨어질 때
            middle_space = ((space_available // 3) // 10 * 10)
            side_space = ((space_available - middle_space) // 2 // 10 * 10)
            
        elif end_digit in [3, 4]: 
            # middle_space가 4로 끝나도록 조정
            middle_space = ((space_available // 3) // 10 * 10) + 4
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 3

        elif end_digit in [6, 7]: 
            # middle_space가 6으로 끝나도록 조정
            middle_space = ((space_available // 3) // 10 * 10) + 6
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 7

        else:
            # 기본 분배: 나머지 end_digit의 경우 기본적으로 10의 배수로 처리
            middle_space = (space_available // 3 // 10 * 10)
            side_space = ((space_available - middle_space) // 2 // 10 * 10)

    # 이후 x2 및 x3 계산 등에 사용할 수 있습니다.    

        # Calculating all x-coordinates
        x2 = x1 + side_space
        x3 = x2 + middle_space
        x4 = inputCW - 115

        Topholex = [x1, x2, x3, x4]
        Bottomx = [x1, x2, x3, x4]    
    else:
        # For CW < 1400, there are 3 holes
        x1 = 115
        x2 = (inputCW - 115 * 2) / 2 + 115
        x3 = inputCW - 115
        Topholex = [x1, x2, x3]
        Bottomx = [x1, x2, x3]

    # Create a dictionary to hold the coordinates
    lc_positions = {
        'Topholex': Topholex,
        'Topholey': Topholey,
        'Bottomx': Bottomx,
        'Bottomy': Bottomy,
        'end_digit' : end_digit
    }

    print(f"lc_positions : {lc_positions}")
    return lc_positions

def calculate_lc_vertical_hole_positions_031(inputCW, inputCD):
    # 고정된 x 좌표들 설정
    leftholex = [115, 115, 115]
    rightholex = [inputCW - 115, inputCW - 115, inputCW - 115]

    # y 좌표를 계산하기 위한 로직
    if inputCD <= 1250:
        # CD가 1250 이하일 경우, y1만 존재
        y1 = (inputCD) / 2  # 40*2를 제외하고 중간값을 계산
        vertical_holey = [y1]
    elif inputCD > 1250 and inputCD <2100:
        # CD가 1250을 초과할 경우, y 좌표 3개를 계산
        divided_space = (inputCD - 80) // 3  # 40*2를 제외하고 3등분
        end_digit = divided_space % 10

        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution        
        space_available = inputCD - 40 * 2
        
        if end_digit in [3, 4]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 4
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 3   
            # print (f" 3, 4 선택됨 end_digit : {end_digit}  , side_space : {side_space} , middle_space : {middle_space}")                 
        elif end_digit in [6, 7]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 6
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 7
        else:            
            middle_space = ((space_available // 3) // 10 * 10)             
            side_space = ((space_available - middle_space) // 2 // 10 * 10) 

        # Calculating all x-coordinates
        y1 = side_space + 40    # 40위에서 시작
        y2 = y1 + middle_space

        vertical_holey = [y1, y2]
        # print (f" end_digit : {end_digit}  , side_space : {side_space} , middle_space : {middle_space}")
    # 2100 보다 크면 중간 홀 3개    
    else:         
        divided_space = (inputCD - 80) // 4  # 40*2를 제외하고 4등분
        end_digit = divided_space % 10

        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution        
        space_available = inputCD - 40 * 2
        
        if end_digit in [3, 4]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 4) // 10 * 10) + 4
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 3 // 10 * 10) + 3                    
        elif end_digit in [6, 7]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 4) // 10 * 10) + 6
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 3 // 10 * 10) + 7
        else:            
            middle_space = space_available // 4
            side_space = (space_available - middle_space) // 3

        # Calculating all x-coordinates
        y1 = side_space + 40
        y2 = y1 + middle_space
        y3 = y2 + side_space

        vertical_holey = [y1, y2, y3]

    # 좌표 딕셔너리 반환
    hole_positions = {
        'leftholex': leftholex,
        'rightholex': rightholex,
        'vertical_holey': vertical_holey
    }

    return hole_positions

# 패턴그리는 함수
def draw_patterned_lines(doc, start_x, end_x, y1, y2, layer):
    x = start_x
    gap_count = 0  # 1mm 간격으로 그린 선의 개수를 추적

    while x < end_x:
        line(doc, x, y1, x, y2, layer)

        # 1mm 간격으로 3개의 선을 그린 후 5mm 간격 적용
        gap_count += 1
        if gap_count == 3:
            x += 6
            gap_count = 0  # 간격 카운트 리셋
        else:
            x += 1.5

def ALprofile_draw(doc, x, y, length, direction="side"):   
    width = 42
    if direction=='side':
        # 마구리
        rectangle(doc, x, y, x+width, y - 2, layer="CL")        
        rectangle(doc, x, y - length - 2, x+width, y - length, layer="CL")        

        rectangle(doc, x, y - length - 2, x+width, y-2, layer="CL")        
        rectangle(doc, x + 1, y - 2, x+width-1, y-length, layer="CL")
        rectangle(doc, x+15, y - 50, x+width-15, y-length+50-2, layer="6")
        line(doc, x+15+1.2, y - 50, x+15+1.2, y-length+50-2, layer="hidden")
        line(doc, x+width-15-1.2, y - 50, x+width-15-1.2, y-length+50-2, layer="hidden")
        rectangle(doc, x+15+3, y - 50-2, x+width-15-3, y-length+50-2+2, layer="4")
    if direction=='topbottom':
        # 마구리
        rectangle(doc, x, y, x+2, y - width, layer="CL")        
        rectangle(doc, x + length - 2, y ,  x + length , y - width, layer="CL")        

        rectangle(doc, x , y , x+length, y-width, layer="CL")        
        rectangle(doc, x + 2 , y - 1, x+length-2, y - width + 1, layer="CL")

        rectangle(doc, x + 50, y - 15, x + length - 50, y-width + 15, layer="6")
        line(doc, x + 50, y - 15 - 1.2, x + length - 50,  y - 15 - 1.2, layer="hidden")
        line(doc, x + 50,  y - width + 15 + 1.2 , x + length - 50, y - width + 15 + 1.2 , layer="hidden")        
        rectangle(doc, x+52, y - 18, x+length -52, y-width + 18, layer="4")

#031 홀간격계산
def calculate_hole_positions_separated_lc031(LCD):    
    center = LCD // 2  # 중심 지점
    central_start = center - 185  # 중심 구간 시작
    central_end = center + 185  # 중심 구간 끝

    left_positions = []
    right_positions = []

    # 중심으로부터 왼쪽 계산
    left_current = central_start
    while left_current > 0:
        left_positions.insert(0, left_current)
        left_current -= 220

    # 중심으로부터 오른쪽 계산
    right_current = central_end
    while right_current < LCD:
        right_positions.append(right_current)
        right_current += 220

    # 결과 리스트 합치기
    positions = left_positions + right_positions

    return positions

#031 led bar bracket 홀간격계산
def calculate_ledbar_bottomhole_lc031(ledbar_length):
    center = ledbar_length // 2  # 중심 지점
    central_gap = 350
    interval = 300  # 간격

    left_positions = []
    right_positions = []

    # 중앙 구간 시작 및 종료
    central_start = center - central_gap // 2
    central_end = center + central_gap // 2

    # 중심으로부터 왼쪽 계산
    left_current = central_start
    while left_current > 0:
        left_positions.insert(0, left_current)
        left_current -= interval

    # 중심으로부터 오른쪽 계산
    right_current = central_end
    while right_current < ledbar_length:
        right_positions.append(right_current)
        right_current += interval

    # 결과 리스트 합치기
    positions = left_positions + right_positions

    return positions

def calculate_ledbar_upperhole_lc031(ledbar_length):
    center = ledbar_length // 2  # 중심 지점
    initial_offset = 50  # 최초 간격
    interval = 100  # 이후 간격

    left_positions = []
    right_positions = []

    # 중심으로부터 왼쪽 계산 (초기 간격으로 시작)
    left_current = center - initial_offset
    while left_current > 0:
        left_positions.insert(0, left_current)
        left_current -= interval

    # 중심으로부터 오른쪽 계산 (초기 간격으로 시작)
    right_current = center + initial_offset
    while right_current < ledbar_length:
        right_positions.append(right_current)
        right_current += interval

    # 결과 리스트 합치기
    positions = left_positions + right_positions

    return positions

#008_A 홀간격계산
def calculate_lc_position_008_A(inputCW, inputCD):
    # Calculate Y-coordinates for top and bottom holes
    Topholey = (int)(inputCD - 75)
    Middley = (int)(inputCD/2)
    Bottomy = 75

    firstXgap = 75    
    middle_space = 0

    # Calculating all x-coordinates
    x1 = firstXgap
    x2 = (int)(inputCW/2)
    x3 = (int)(inputCW - firstXgap)

    Topholex = [x1, x2, x3]
    Middlex = [x1, x3]
    Bottomx = [x1, x2, x3]
    
    lc_positions = {
        'Topholex': Topholex,
        'Topholey': Topholey,
        'Middlex': Middlex,
        'Middley': Middley,
        'Bottomx': Bottomx,
        'Bottomy': Bottomy
    }


    return lc_positions


def calculate_lc_position_008_A_CW(inputCW, inputCD):#CW 팝너트 홀
    # Calculate Y-coordinates for top and bottom holes
    Topholey = (int)(inputCD - 13)
    Bottomy = 13

    firstXgap = 13    

    # Calculating all x-coordinates
    x1 = firstXgap
    x2 = (int)(inputCW/2 - firstXgap*2)/3 + firstXgap
    x3 = (int)(inputCW/2 - firstXgap*2)*2/3 + firstXgap
    x4 = (int)(inputCW/2 - firstXgap*2) + firstXgap


    total = x2 + x3 + x4
    # total의 소수 첫째자리(10분위)가 0이 아니면 y3에 0.1을 추가
    if int(total * 10) % 10 != 0:
        x3 += 0.1    

    x2 = (int)(x2*10)/10
    x3 = (int)(x3*10)/10
    x4 = (int)(x4*10)/10
    
    Topholex = [x1, x2, x3, x4]
    Bottomx = [x1, x2, x3, x4]
    
    lc_positions = {
        'Topholex': Topholex,
        'Topholey': Topholey,
        'Bottomx': Bottomx,
        'Bottomy': Bottomy
    }


    return lc_positions


def calculate_lc_position_008_A_CD(inputCW, inputCD):#CD 팝너트 홀
    # Calculate Y-coordinates for top and bottom holes
    Leftx = 13 
    Rightx = (int)(inputCW - 13)

    firstYgap = 13    

    # Calculating all x-coordinates
    y1 = firstYgap
    y2 = int((inputCD/2 - firstYgap*2)/3*10)/10 + firstYgap
    y3 = int((inputCD/2 - firstYgap*2)*2/3*10)/10 + firstYgap
    y4 = int((inputCD/2 - firstYgap)*10)/10
    
    total = y2 + y3 + y4
    # total의 소수 첫째자리(10분위)가 0이 아니면 y3에 0.1을 추가
    if int(total * 10) % 10 != 0:
        y3 += 0.1    

    y2 = int(y2*10)/10
    y3 = int(y3*10)/10
    y4 = int(y4*10)/10
    
    Lefty = [y1, y2, y3, y4]
    Righty = [y1, y2, y3, y4]
    
    lc_positions = {
        'Leftx': Leftx,
        'Lefty': Lefty,
        'Rightx': Rightx,
        'Righty': Righty
    }


    return lc_positions

def calculate_lc_position_008_A_ASSY_horizontal(inputCW, inputCD):
    liney = inputCD / 2

    firstXgap = 37.5    

    x1 = firstXgap
    x2 = (inputCW // 2 - firstXgap * 2) / 3 + firstXgap
    x3 = (inputCW // 2 - firstXgap * 2) * 2 / 3 + firstXgap
    x4 = inputCW / 2 - firstXgap
    
    total = x2 + x3 + x4

    # total의 소수 첫째 자리가 0이 아니면 x3에 0.1을 추가
    if round(total, 1) != total:
        x3 += 0.1    

    # 소수점 첫째 자리까지 유지
    x2 = round(x2, 1)
    x3 = round(x3, 1)
    x4 = round(x4, 1)

    linex = [x1, x2, x3, x4]
    
    lc_positions = {
        'linex': linex,
        'liney': liney,
    }

    return lc_positions

def calculate_lc_position_008_A_ASSY_vertical(inputCW, inputCD):
    linex = inputCW / 2

    firstYgap = 37.5    

    y1 = firstYgap
    base = (inputCD // 2 - firstYgap * 2) / 3  # 정수 연산을 명확히 처리
    
    y2 = base + firstYgap
    y3 = base * 2 + firstYgap
    y4 = inputCD / 2 - firstYgap
    
    total = y2 + y3 + y4

    # 소수 첫째 자리가 0이 아니면 y3에 0.1 추가
    if round(total, 1) != total:
        y3 += 0.1

    # 소수점 첫째 자리까지 유지
    y2, y3, y4 = round(y2, 1), round(y3, 1), round(y4, 1)

    liney = [y1, y2, y3, y4]
    
    lc_positions = {
        'linex': linex,
        'liney': liney,
    }

    return lc_positions

#035 홀간격계산
def calculate_lc_position_035(inputCW, inputCD):
    # Calculate Y-coordinates for top and bottom holes
    Topholey = inputCD - 40
    Bottomy = 40

    # Initialize the list of x-coordinates for top and bottom holes
    Topholex = []
    Bottomx = []

    firstXgap = 110
    side_space = 0
    space_available = 0
    middle_space = 0

    # CW에서 양쪽의 여유분 230을 제외한 값을 3등분
    divided_space = (inputCW - firstXgap*2) // 3
    end_digit = divided_space % 10    

    # Determine the number of holes and calculate their positions
    if inputCW >= 1600:
        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution
        x1 = firstXgap # 라이트케이스 첫홀 위치 25 띄운 치수 제외한 치수
        space_available = inputCW - firstXgap * 2
        
        if end_digit in [3, 4]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 4
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 3                    
        elif end_digit in [6, 7]:
            # The middle space needs to end with 4, therefore itd should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 6
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 7
        else:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) 
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) 

        # Calculating all x-coordinates
        x2 = x1 + side_space
        x3 = x2 + middle_space
        x4 = inputCW - firstXgap

        Topholex = [x1, x2, x3, x4]
        Bottomx = [x1, x2, x3, x4]
    else:
        # For CW < 1400, there are 3 holes
        x1 = firstXgap
        x2 = (inputCW - firstXgap * 2) / 2 + firstXgap
        x3 = inputCW - firstXgap
        Topholex = [x1, x2, x3]
        Bottomx = [x1, x2, x3]

    # Create a dictionary to hold the coordinates
    lc_positions = {
        'Topholex': Topholex,
        'Topholey': Topholey,
        'Bottomx': Bottomx,
        'Bottomy': Bottomy,
        'end_digit' : end_digit
    }

    # print (f" calculate_lc_position_035 함수 end_digit : {end_digit}  , side_space : {side_space} , middle_space : {middle_space}")
    # print (f" {lc_positions}")

    return lc_positions

def calculate_hole_vertical_lc035(LCD):
    # 중심 지점 계산
    center = LCD // 2

    # 남은 거리를 분할하는 내부 함수
    def divide_remaining_distance(remaining_distance):        
        for div in range(3, 5):
            intervals = [remaining_distance // div] * (div - 1)
            intervals.append(remaining_distance - sum(intervals))
            if all(150 <= x <= 500 for x in intervals):
                return intervals
        return []

    # 중심으로부터의 고정 홀 위치
    fixed_holes = [54, center - 15, center + 15, LCD - 54]

    # LCD 값에 따른 추가 홀 위치 계산
    if LCD < 1950:
        # LCD가 1950 미만일 경우 중간 값 추가
        additional_holes = [(54 + (center - 15)) // 2]
    else:
        # LCD가 1950 이상일 경우 3등분한 값 추가
        first_third = 54 + (center - 15 - 54) // 3
        second_third = first_third + (center - 15 - 54) // 3
        additional_holes = [first_third, second_third]

    # 중심 왼쪽 홀 위치 결합 및 정렬
    left_holes = sorted(fixed_holes[:2] + additional_holes)

    # 중심을 기준으로 오른쪽 홀 위치 계산
    right_holes = [LCD - hole for hole in reversed(left_holes)]

    # 전체 홀 위치 결합 및 정렬
    return sorted(set(left_holes + right_holes))

def calculate_lc_vertical_hole_positions_035(inputCW, inputCD):
    # 고정된 x 좌표들 설정
    firstXgap = 110
    leftholex = [firstXgap, firstXgap, firstXgap]
    rightholex = [inputCW - firstXgap, inputCW - firstXgap, inputCW - firstXgap]

    side_space = 0
    space_available = 0

    # print (f"input CD {inputCD/2}")

    # y 좌표를 계산하기 위한 로직
    if inputCD <= 1350:
        # CD가 1250 이하일 경우, y1만 존재
        y1 = inputCD / 2   # 40*2를 제외하고 중간값을 계산
        vertical_holey = [y1]
    elif inputCD > 1350 and inputCD <2100:
        # CD가 1250을 초과할 경우, y 좌표 3개를 계산
        divided_space = (inputCD - 80) // 3  # 40*2를 제외하고 3등분
        end_digit = divided_space % 10

        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution        
        space_available = inputCD - 40 * 2
        
        if end_digit in [3, 4]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 4
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 3   
            # print (f" 3, 4 선택됨 end_digit : {end_digit}  , side_space : {side_space} , middle_space : {middle_space}")                 
        elif end_digit in [6, 7]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 6
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 7
        else:            
            middle_space = ((space_available // 3) // 10 * 10)             
            side_space = ((space_available - middle_space) // 2 // 10 * 10) 

        # Calculating all x-coordinates
        y1 = side_space + 40    # 40위에서 시작
        y2 = y1 + middle_space

        vertical_holey = [y1, y2]
        # print (f" end_digit : {end_digit}  , side_space : {side_space} , middle_space : {middle_space}")
    # 2100 보다 크면 중간 홀 3개    
    else:         
        divided_space = (inputCD - 80) // 4  # 40*2를 제외하고 4등분
        end_digit = divided_space % 10

        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution        
        space_available = inputCD - 40 * 2
        
        if end_digit in [3, 4]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 4) // 10 * 10) + 4
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 3 // 10 * 10) + 3                    
        elif end_digit in [6, 7]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 4) // 10 * 10) + 6
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 3 // 10 * 10) + 7
        else:            
            middle_space = space_available // 4
            side_space = (space_available - middle_space) // 3

        # Calculating all x-coordinates
        y1 = side_space + 40
        y2 = y1 + middle_space
        y3 = y2 + side_space

        vertical_holey = [y1, y2, y3]

    # 좌표 딕셔너리 반환
    hole_positions = {
        'leftholex': leftholex,
        'rightholex': rightholex,
        'vertical_holey': vertical_holey
    }

    return hole_positions

def calculate_cliphole_lc035(distance , length):
    # 중심 지점 계산
    center = length // 2

    # LCD 길이에 따른 홀의 수와 위치 결정
    if length < 1950:
        # LCD가 1950 미만일 때는 3개의 홀
        hole_positions = [distance, center, length - distance]
    else:
        # LCD가 1950 이상일 때는 4개의 홀
        remaining_center = length - 2 * distance  # 양쪽 끝 홀 사이의 거리
        center_hole1 = distance + remaining_center // 4  # 중앙 왼쪽 홀
        center_hole2 = distance + (remaining_center // 4) * 3  # 중앙 오른쪽 홀
        hole_positions = [distance, center_hole1, center_hole2, length - distance]

    return sorted(hole_positions)

def calculate_midplatehole_lc035(length):
    # 전체 길이를 5등분하는 간격 계산
    interval = length / 5

    # 4개의 위치 좌표 계산
    hole_positions = [round(interval * i, 1) for i in range(1, 5)]

    return hole_positions
def calculate_midplate_connecthole_lc035(length):
    # 양 끝 홀 위치를 고정
    first_and_last_hole = 82.6

    # 중간 홀들 간의 거리 계산
    middle_length = length - 2 * first_and_last_hole  # 양쪽 끝 홀을 제외한 길이
    interval = round(middle_length / 7, 1)  # 6등분한 간격, 소수점 첫째 자리까지

    # 홀 위치 계산
    hole_positions = [first_and_last_hole]  # 첫 홀 위치 추가
    current_position = first_and_last_hole

    # 중간 홀 위치 계산 및 추가
    for _ in range(7):
        current_position += interval
        hole_positions.append(round(current_position, 1))

    # 마지막 홀 위치 추가
    hole_positions.append(length - first_and_last_hole)

    return hole_positions

#035 led bar bracket 홀간격계산
def calculate_ledbar_bottomhole_lc035(ledbar_length):
    center = ledbar_length // 2  # 중심 지점
    central_gap = 350
    interval = 300  # 간격

    left_positions = []
    right_positions = []

    # 중앙 구간 시작 및 종료
    central_start = center - central_gap // 2
    central_end = center + central_gap // 2

    # 중심으로부터 왼쪽 계산
    left_current = central_start
    while left_current > 0:
        left_positions.insert(0, left_current)
        left_current -= interval

    # 중심으로부터 오른쪽 계산
    right_current = central_end
    while right_current < ledbar_length:
        right_positions.append(right_current)
        right_current += interval

    # 결과 리스트 합치기
    positions = left_positions + right_positions

    return positions

def calculate_ledbar_upperhole_lc035(ledbar_length):
    center = ledbar_length // 2  # 중심 지점
    initial_offset = 50  # 최초 간격
    interval = 100  # 이후 간격

    left_positions = []
    right_positions = []

    # 중심으로부터 왼쪽 계산 (초기 간격으로 시작)
    left_current = center - initial_offset
    while left_current > 0:
        left_positions.insert(0, left_current)
        left_current -= interval

    # 중심으로부터 오른쪽 계산 (초기 간격으로 시작)
    right_current = center + initial_offset
    while right_current < ledbar_length:
        right_positions.append(right_current)
        right_current += interval

    # 결과 리스트 합치기
    positions = left_positions + right_positions

    return positions



####################################################################################################################################################################################
# 026 천장 자동작도
####################################################################################################################################################################################
def lc026():     
    global args  #수정할때는 global 선언이 필요하다. 단순히 읽기만 할때는 필요없다.    
    global start_time
    global workplace, drawdate, deadlinedate, lctype, secondord, text_style, exit_program, Vcut_rate, Vcut, program_message, formatted_date, text_style_name
    global CW_raw, CD_raw, LC_height, panel_thickness, su, LCframeMaterial, LCplateMaterial
    global input_CD, input_CW,CD,CW,drawdate_str , deadlinedate_str, drawdate_str_short, deadlinedate_str_short    
    global doc, msp , T5_is, text_style, distanceXpos, distanceYpos

    abs_x = 0
    abs_y = 0

    # LC 크기 지정
    LCD = CD-50
    LCW = CW-50

    # 2page 덴크리 다온텍등 등기구 설명 그림
    insert_block(0 , 0 , "lc026_2pageframe")    

    # watt 계산 공식 적용해야 함
    # LED바 기장(단위:m) X 수량 X 15 = 60W 이하
    # LED바 1개당 watt 산출 m단위로 계산.. /1000  026모델은 2개의 led바가 들어감     

    ledsu = 2

    wattCalculate = math.ceil((LCD - 3)/1000 * ledsu * 15)

    if(wattCalculate <= 60):
        watt = 60
    elif(wattCalculate > 60 and wattCalculate < 100):
        watt = 100
    elif(wattCalculate >= 100 and wattCalculate < 150):
        watt = 150
    elif(wattCalculate >= 150 and wattCalculate < 200):
        watt = 200
    elif(wattCalculate >= 200 and wattCalculate < 200):
        watt = 300 

    watt = f"{str(watt)}W"        

    print("L/C 1개당 규격설정 wattCalculate: " + str(watt))      

    print("wattCalculate: " + str(wattCalculate))     

    print("L/C 1개당 규격설정 wattCalculate: " + str(watt))  

    T5_standard = 0
    if T5_is == '있음':
        adjusted_CD = CD - 50  # CD에서 50 뺌

        # 규격 결정
        if adjusted_CD > 1500:
            T5_standard = 1500
        elif adjusted_CD > 1200:
            T5_standard = 1200
        else:
            T5_standard = 900

        print(f"선택된 규격: {T5_standard}")
    else:
        print("T5_is 가 '있음'이 아니므로 규격 선택 안 함")         

    # 도면틀 넣기
    BasicXscale = 2560
    BasicYscale = 1550
    TargetXscale = 800 
    TargetYscale = 300 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale

    frame_scale = 1    

    # 현장명
    textstr = workplace       
    x = abs_x + 80
    y = abs_y + 1000 + 218
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 발주처
    textstr = secondord
    x = abs_x + 80
    y = abs_y + 1000 + 218 - 30
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 타입
    textstr = lctype
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 수량
    textstr = su
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 카사이즈
    textstr = f"W{CW} x D{CD}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 색상 L/C재질이 있는 경우는 그것을 표시하고, 없는 경우는 기본 흑색무광표시
    if LCframeMaterial == 'SPCC 1.6T' : # 기본색상임            
        textstr = f"L / C : 흑색무광"
    else:
        textstr = f"L / C : {LCframeMaterial}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # (등기구업체 : 덴크리)    
    textstr = f"(등기구업체 : 덴크리)"
    x = abs_x + 120
    y = abs_y + 1000 + 90
    draw_Text(doc, x, y , 12, str(textstr), layer='0')    
    # (아답터용량 : 60W )
    textstr = f"(아답터용량 : {watt} )"
    x = abs_x + 120
    y = abs_y + 1000 + 90 - 20
    draw_Text(doc, x, y , 12, str(textstr), layer='0')
    textstr = f"*. 출도일자 : {drawdate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 
    draw_Text(doc, x, y , 8, str(textstr), layer='0')    
    textstr = f"*. 납품일자 : {deadlinedate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 - 15
    draw_Text(doc, x, y , 8, str(textstr), layer='0')

    # 갑지2번째 led바에 대한 내용 기술
    # LED BAR :
    LedBarLength = CalculateLEDbar(CD)

    textstr = f"LED BAR - {LedBarLength}={su*2}(EA)"
    x = abs_x + 711
    y = abs_y + 600
    draw_Text(doc, x, y , 20, str(textstr), layer='0')    
    # SMPS-1(EA)
    textstr = f"SMPS-{su}(EA)"
    x = abs_x + 711 + 60
    y = abs_y + 600 - 35
    draw_Text(doc, x, y , 20, str(textstr), layer='0')    
    # 색상 L/C재질이 있는 경우는 그것을 표시하고, 없는 경우는 기본 흑색무광표시
    if LCframeMaterial == 'SPCC 1.6T' : # 기본색상임            
        # Type
        textstr = f"{lctype}타입"
    else:
        # 대소문자 구분 없이 mr 또는 M/R이 포함되어 있는지 확인
        if "mr" in LCframeMaterial.lower() or "m/r" in LCframeMaterial.lower():
            textstr = f"026 : (밀러타입)"
        else:
            textstr = f"026 : {LCframeMaterial}"

    x = abs_x + 711 + 100
    y = abs_y + 600 - 35 - 85 
    draw_Text(doc, x, y , 20, str(textstr), layer='0')    
    # 현장명
    textstr = f"*.현장명 : {secondord}-{workplace} - {su}(set)"
    x = abs_x + 30
    y = abs_y + 600 
    draw_Text(doc, x, y , 18, str(textstr), layer='0')    
    # (등기구업체 : 덴크리)   
    if T5_standard > 0 : 
        Ledcompany = f"(등기구업체 : 엘엠팩트)"        
        Leditem = "T5"
        Ledwatt = "6500K"
        LedBarLength = T5_standard
        watt = "내장"
    else:
        Ledcompany = f"(등기구업체 : 다온텍)"
        Leditem = "LED BAR"
        Ledwatt = "10000K"
        LedBarLength = CalculateLEDbar(CD)
        watt = f"{str(watt)}W"

    textstr = f"{Ledcompany}"
    x = abs_x + 30 + 200
    y = abs_y + 600 - 50
    draw_Text(doc, x, y , 18, str(textstr), layer='0')    
    # LED BAR-(누드&클립&잭타입)10000K-1400L
    textstr = f"{Leditem} -(누드&클립&잭타입){Ledwatt}-{LedBarLength}L"
    x = abs_x + 30 + 60
    y = abs_y + 600 - 50 - 80
    draw_Text(doc, x, y , 14, str(textstr), layer='0')    
    # (하단) LED BAR-(누드&클립&잭타입)10000K-1400L
    textstr = f"{Leditem} -(누드&클립&잭타입){Ledwatt}-{LedBarLength}L"
    x = abs_x + 30 + 60
    y = abs_y + 600 - 50 - 80 - 360
    draw_Text(doc, x, y , 14, str(textstr), layer='0')    
    
    # car inside box 표기
    xx =  740
    yy =  45
    rectangle(doc, xx,yy,xx+200,yy+100,layer='0')      
    line(doc, xx,yy+50,xx+200,yy+50)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50, yy+50 , 12, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30, yy+25 , 12, str(textstr), layer='0')


    #########################################################################################################
    # LC ASSY 상부 형상
    ##########################################################################################################
    abs_x = abs_x + 2500
    abs_y = abs_y + CD + 450
    if LC_height != 100 :
        insert_block(abs_x , abs_y , "lc026_top_left")    
    else:
        insert_block(abs_x , abs_y , "lc026_top_left_height100")    

    x = abs_x + CW 
    y = abs_y 
    if LC_height != 100 :
        insert_block(x , y , "lc026_top_right")  
    else:
        insert_block(x , y , "lc026_top_right_height100")            

    x1 = abs_x + 350 - 40
    y1 = abs_y + 19.6    
    x2 = x1 + (CW - 350*2 + 80)
    y2 = y1
    line(doc, x1, y1, x2, y2, layer='4')     # 4는 하늘색
    lineto(doc, x2, y1+10, layer='4')     # 4는 하늘색
    lineto(doc, x1, y1+10, layer='4')     # 4는 하늘색
    lineto(doc, x1, y1, layer='4')     # 4는 하늘색
    textstr =  f"LMW(CW-670)"
    dim_string(doc, x1, y2, x2, y2, 45,  textstr, text_height=0.20, text_gap=0.07, direction="up")    

    # 상부선 2mm 간격
    X1 = abs_x - 25
    Y1 = abs_y + LC_height
    X2 = X1 + CW + 50
    Y2 = Y1
    line(doc, X1, Y1, X2, Y2, layer='2')     # 회색
    lineto(doc, X2, Y1+2, layer='2')    
    lineto(doc, X1, Y1+2, layer='2')    
    lineto(doc, X1, Y1, layer='2')     

    # Assy 상부 치수
    x1 = abs_x + 300
    y1 = abs_y + LC_height
    x2 = abs_x + CW - 300
    y2 = y1    
    textstr =  f"CW-600={CW-600}"
    dim_string(doc, x1, y1, x2, y2, 155,  textstr, text_height=0.20, text_gap=0.07, direction="up")
    # Assy 중심 치수
    x1 = abs_x + 350
    y1 = abs_y 
    x2 = abs_x + CW - 350
    y2 = y1    
    textstr =  f"CW-500={CW-500}"
    dim_string(doc, x1, y1, x2, y2, 105,  textstr, text_height=0.20, text_gap=0.07, direction="down")
    # Assy 상단 하부 치수
    x1 = abs_x 
    y1 = abs_y 
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"CW(AA) {CW}"
    dim_string(doc, x1, y1, x2, y2, 160,  textstr, text_height=0.20, text_gap=0.07, direction="down")

    ###########################################################################################################################################
    # 1page Assy Car case
    ###########################################################################################################################################
    insideHeight = 25
    abs_y = abs_y - 450 - CD
    x1 = abs_x
    y1 = abs_y
    x2 = abs_x + CW
    y2 = CD
    rectangle(doc, x1, y1, x2, y2, layer='0')

    # 내부 선 그리기
    x_positions = [(abs_x + insideHeight, abs_x + CW - insideHeight)]
    y_positions = [(abs_y + insideHeight, abs_y + CD - insideHeight)]

    for x_pos in x_positions:
        for y_pos in y_positions:
            # 수평선
            line(doc, x_pos[0], y_pos[0], x_pos[0] + 2, y_pos[0], layer='0')  # 왼쪽
            line(doc, x_pos[1], y_pos[0], x_pos[1] - 2, y_pos[0], layer='0')  # 오른쪽
            line(doc, x_pos[0], y_pos[1], x_pos[0] + 2, y_pos[1], layer='0')  # 왼쪽
            line(doc, x_pos[1], y_pos[1], x_pos[1] - 2, y_pos[1], layer='0')  # 오른쪽

            # 수직선
            line(doc, x_pos[0], y_pos[0], x_pos[0], y_pos[1], layer='0')  # 왼쪽
            line(doc, x_pos[1], y_pos[0], x_pos[1], y_pos[1], layer='0')  # 오른쪽

    # 좌측 fan assy
    insert_block(abs_x + 15 , abs_y  + CD - 70 , "lc026_fan")    
    insert_block(abs_x + insideHeight + 2 , abs_y  + CD - insideHeight , "lc026_rib_top_left")    
    insert_block(abs_x + CW - insideHeight - 2 , abs_y  + CD - insideHeight , "lc026_rib_top_right")    
    insert_block(abs_x + insideHeight + 2 , abs_y  + insideHeight , "lc026_rib_bottom_left")    
    insert_block(abs_x + CW - insideHeight - 2 , abs_y  + insideHeight , "lc026_rib_bottom_right")    
    # adaptor
    insert_block(abs_x + insideHeight + 245 , abs_y  + 50 , "lc026_adptor")    
    # 장공/단공 70
    insert_block(abs_x + 150 , abs_y  + 250 , "circle70")    
    insert_block(abs_x + CW - 150 , abs_y  + 250 , "circle70")    
    insert_block(abs_x + CW - 150 , abs_y  + CD - 250 , "circle70")    
    insert_block(abs_x + 150 , abs_y  + CD - 250 , "circle70")    
    # 장공/단공 20
    insert_block(abs_x + 200 , abs_y  + 200 , "circle20")    
    insert_block(abs_x + CW - 200 , abs_y  + 200 , "circle20")    
    insert_block(abs_x + CW - 200 , abs_y  + CD - 200, "circle20")    
    insert_block(abs_x + 200 , abs_y  + CD - 200 , "circle20")    
    # 장공/단공 11 and 10 장공
    insert_block(abs_x + 100 , abs_y + CD/2  , "circle11plus")    
    insert_block(abs_x + CW - 100 , abs_y + CD/2 , "circle11plus")    
    insert_block(abs_x + CW + 1200 , abs_y + 650 , "lc026assy")    

    # 내부 폴리갈 자리 하늘색    
    # 폴리갈 크기 산출
    polygal_width = LCW - 620
    if CD>=2000 :
        polygal_height = (CD - 29*2)/3
        x1 = abs_x + 335
        y1 = abs_y + 29
        x2 = abs_x + CW - 335
        y2 = abs_y + CD - 29
        y3 = abs_y + 29 + polygal_height*2
        y4 = abs_y + 29 + polygal_height
        rectangle(doc, x1, y1, x2, y2, layer='4')   # 하늘색   
        line(doc, x1, y3 , x2, y3 , layer='4')     # 4는 하늘색
        line(doc, x1, y4 , x2, y4 , layer='4')     # 4는 하늘색
        dim_vertical_left(doc,  abs_x + CW /1.6 , y1, abs_x + CW /1.7 , y2, 150, "JKW", text_height=0.20,  text_gap=0.07)
        dim_vertical_left(doc,  abs_x + CW /1.6 , y1, abs_x + CW /1.7 , y4, 70, "JKW", text_height=0.20,  text_gap=0.07)
        dim_vertical_left(doc,  abs_x + CW /1.6 , y4, abs_x + CW /1.7 , y3, 70, "JKW", text_height=0.20,  text_gap=0.07)        
        dim_vertical_left(doc,  abs_x + CW /1.6 , y3, abs_x + CW /1.7 , y2, 70, "JKW", text_height=0.20,  text_gap=0.07)        
    else:
        polygal_height = LCD/2 - 15    
        x1 = abs_x + 335
        y1 = abs_y + 29
        x2 = abs_x + CW - 335
        y2 = abs_y + CD - 29
        rectangle(doc, x1, y1, x2, y2, layer='4')   # 하늘색   
        line(doc, x1, abs_y + CD/2 , x2, abs_y + CD/2 , layer='4')     # 4는 하늘색
        dim_vertical_left(doc,  abs_x + CW /1.6 , y1, abs_x + CW /1.7 , y2, 150, "JKW", text_height=0.20,  text_gap=0.07)
        dim_vertical_left(doc,  abs_x + CW /1.6 , y1, abs_x + CW /1.7 , (y1+y2)/2, 70, "JKW", text_height=0.20,  text_gap=0.07)
        dim_vertical_left(doc,  abs_x + CW /1.6 , y2, abs_x + CW /1.7 , (y1+y2)/2, 70, "JKW", text_height=0.20,  text_gap=0.07)

    # 인발 파란색(왼쪽)
    side_gap = 250
    top_gap = 27
    x1 = abs_x + side_gap
    y1 = abs_y + top_gap
    x2 = x1 + 100
    y2 = abs_y + CD - top_gap
    rectangle(doc, x1, y1, x2, y2, layer='5')  # 파란색

    # 수직선 간격 배열
    line_gaps = [12.5, 11.5, 2, 11.5, 25, 11.5, 2, 11.5]
    current_x = x1

    for gap in line_gaps:
        current_x += gap
        line(doc, current_x, y1, current_x, y2, layer='55')


    # 인발 파란색 (오른쪽)
    side_gap = CW - 250 - 100
    top_gap = 27
    x1 = abs_x + side_gap
    y1 = abs_y + top_gap
    x2 = x1 + 100
    y2 = abs_y + CD - top_gap
    rectangle(doc, x1, y1, x2, y2, layer='5')  # 파란색

    # 수직선 간격 배열
    line_gaps = [12.5, 11.5, 2, 11.5, 25, 11.5, 2, 11.5]
    current_x = x1

    for gap in line_gaps:
        current_x += gap
        line(doc, current_x, y1, current_x, y2, layer='55')

    # LC 좌우 사각형 (왼쪽)
    side_gap = 85
    top_gap = 25
    x1 = abs_x + side_gap
    y1 = abs_y + top_gap
    x2 = abs_x + side_gap + 25
    y2 = abs_y + CD - top_gap
    rectangle(doc, x1, y1, x2, y2, layer='0')   # 흰색     

    # LC 좌우 사각형 (오른쪽)
    side_gap = CW - 85 - 25
    top_gap = 25
    x1 = abs_x + side_gap
    y1 = abs_y + top_gap
    x2 = abs_x + side_gap + 25
    y2 = abs_y + CD - top_gap
    rectangle(doc, x1, y1, x2, y2, layer='0')   # 흰색    

    # 본천장 및 LC 상단 
    x1 = abs_x 
    y1 = abs_y + CD
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"L/C {CW-50}"
    dim_string(doc, x1 + 25 , y1 - 25, x2 - 25 , y2 - 25, 100,  textstr, text_height=0.20, text_gap=0.07, direction="up")    
    x1 = abs_x 
    y1 = abs_y + CD
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"CAR INSIDE - {CW}"
    dim_string(doc, x1 , y1 , x2  , y2, 140,  textstr, text_height=0.20, text_gap=0.07, direction="up")    

    # 우측 치수선
    x1 = abs_x + CW 
    y1 = abs_y + CD - 25
    x2 = abs_x + CW - 140
    y2 = y1 - 15      
    dim(doc, x1, y2, x1, y1, 265-140, "JKW", direction="right")
    x1 = abs_x + CW - 140 
    y1 = abs_y + CD - 40
    x2 = x1
    y2 = abs_y + CD/2
    dim(doc, x1, y1, x2, y2, 265, "JKW", direction="right")
    x1 = abs_x + CW - 140 
    y1 = abs_y + 40
    x2 = x1
    y2 = abs_y + CD/2
    dim(doc, x1, y1, x2, y2, 265, "JKW", direction="right")
    x1 = abs_x + CW 
    y1 = abs_y + 25
    x2 = abs_x + CW - 140
    y2 = y1 + 15      
    dim(doc, x1, y2, x2, y1, 265-140, "JKW", direction="right")   

    x1 = abs_x + CW - 25
    y1 = abs_y + CD - 25
    x2 = x1
    y2 = abs_y + CD/2
    dim(doc, x1, y1, x2, y2, 225, "JKW", direction="right")   
    x1 = abs_x + CW - 25
    y1 = abs_y + 25
    x2 = x1
    y2 = abs_y + CD/2
    dim(doc, x1, y1, x2, y2, 225, "JKW", direction="right")   
    
    x1 = abs_x + CW - 25
    y1 = abs_y + CD - 25
    x2 = x1
    y2 = abs_y + 25        
    textstr =  f"L/C {CD-panel_thickness*2}"
    # dim_vertical_right_string(doc, x1, y1, x2, y2, 308, textstr , text_height=0.20,  text_gap=0.07)        
    dim(doc,  x2, y2, x1, y1, 308, direction="right", text = textstr)     
    x1 = abs_x + CW
    y1 = abs_y + CD
    x2 = x1
    y2 = abs_y         
    textstr =  f"CAR INSIDE {CD}"    
    dim(doc,  x2, y2, x1, y1, 364,  direction="right", text = textstr)     

    # 도면틀 넣기
    BasicXscale = 2560
    BasicYscale = 1550
    TargetXscale = CW + 2500
    TargetYscale = CD + 1000
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
    # print("스케일 비율 : " + str(frame_scale))   
    frameYpos = abs_y - 450 * frame_scale 
    insert_frame(abs_x  - 500 * frame_scale , frameYpos  , frame_scale, "drawings_frame", workplace)   
    lastXpos = TargetXscale

    # 1page 상단에 car inside 표기
    x = abs_x + CW/4 - 500
    y = abs_y + frame_scale*BasicYscale + 300
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, x, y , 300, str(textstr), layer='0')    

    frameXpos = abs_x + 600

    frameXpos = frameXpos * frame_scale + 1400

    # print("1페이지 이후 간격 더한 위치 frameXpos : " + str(frameXpos))       

    #################################################################################################
    # 2page LC 본 Frame 그리기
    #################################################################################################    
    # 2page LC frame 삽입(레이져 가공)
    rx1 = frameXpos + 1100 
    ry1 = frameYpos + 1300

    if LC_height != 100 :
        insert_block(rx1 , ry1 , "lc026_frame_section")    
    else:
        insert_block(rx1 - 593.2 , ry1 , "lc026_frame_section_height100")    
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 3780 + (CD - 1500)  # 기본 모형이 1450
    TargetYscale = 2304
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
    # print("스케일 비율 : " + str(frame_scale))   

    x1 = frameXpos
    y1 = abs_y - 500 * frame_scale
    insert_frame(frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   
    
    descript_x = rx1 + LCD + 600

    textstr = f"Part Name : Frame"    
    draw_Text(doc, descript_x , y1 + 1445 , 30, str(textstr), layer='0')    
    textstr = f"Mat.Spec : {LCframeMaterial}"
    draw_Text(doc, descript_x, y1 + 1385 , 30, str(textstr), layer='0')    
    if LCframeMaterial == 'SPCC 1.6T' : # 기본색상임                
        textstr = f"Size : {LCD} x 524"    
    else:
        textstr = f"Size : {LCD} x 491.6"            
    draw_Text(doc, descript_x, y1 + 1325 , 30, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, descript_x, y1 + 1265 , 30, str(textstr), layer='0')    

    ry1 = ry1 - 245
    rx1 = frameXpos + 1100

    if LCframeMaterial == 'SPCC 1.6T' : # 기본색상임                
        x1, y1 = rx1 + 2, ry1
        x2, y2 = rx1 + LCD - 2, y1    
        x3, y3 = x2, y2 + 18.5
        x4, y4 = x3, y3 + 48.6
        x5, y5 = x4, y4 + 12.5
        x6, y6 = x5 + 2, y5
        x7, y7 = x6, y6 + 6.1
        x8, y8 = x7, y7 + 246
        x9, y9 = x8, y8 + 48
        x10, y10 = x9, y9 + 115.3
        x11, y11 = x10, y10 + 29
        x12, y12 = rx1, y11
        x13, y13 = x12, y12 - 29
        x14, y14 = x13, y13 - 115.3
        x15, y15 = x14, y14 - 48
        x16, y16 = x15, y15 - 246
        x17, y17 = x16, y16 - 6.1
        x18, y18 = x17 + 2, y17
        x19, y19 = x18, y18 - 12.5  
        x20, y20 = x19, y19 - 48.6
        x21, y21 = x1, y1

    else: # 1.2T인 경우
        x1, y1 = rx1 + 2, ry1
        x2, y2 = rx1 + LCD - 2, y1    
        x3, y3 = x2, y2 + 19.4
        x4, y4 = x3, y3 + 50.4
        x5, y5 = x4, y4 + 15.25
        x6, y6 = x5 + 2, y5
        x7, y7 = x6, y6 + 4.85
        x8, y8 = x7, y7 + 247.8
        x9, y9 = x8, y8 + 48
        x10, y10 = x9, y9 + 76.9
        x11, y11 = x10, y10 + 29
        x12, y12 = rx1, y11
        x13, y13 = x12, y12 - 29
        x14, y14 = x13, y13 - 76.9
        x15, y15 = x14, y14 - 48
        x16, y16 = x15, y15 - 247.8
        x17, y17 = x16, y16 - 4.85
        x18, y18 = x17 + 2, y17
        x19, y19 = x18, y18 - 15.25
        x20, y20 = x19, y19 - 50.4
        x21, y21 = x1, y1        

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 21
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y

    #절곡라인        
    line(doc, x20, y20, x3, y3,  layer='hidden')    # 절곡선  
    line(doc, x16, y16, x7, y7,  layer='hidden')    # 절곡선  
    line(doc, x15, y15, x8, y8,  layer='hidden')    # 절곡선  
    line(doc, x14, y14, x9, y9,  layer='hidden')    # 절곡선  
    line(doc, x13, y13, x10, y10,  layer='hidden')    # 절곡선  

    # 마킹선 추가
    insert_block(x15 , y15-0.7 , "lc026_markline_yellow20")  
    insert_block(x8 , y8-0.7 , "lc026_markline_yellow20_reverse")  

    #반대 절곡라인
    line(doc, x19, y19, x4, y4,  layer='2')    # 절곡선  
    # line(doc, x1, y1, x2, y2, layer='레이져')   
    # draw_circle(doc, x1 + 248, y1 + 208.7 , 70 , layer='레이져', color='5')   
    circle70x1 =  rx1 + 250
    circle70y1 =  ry1 + 213.8
    circle_cross(doc, circle70x1 , circle70y1 , 70 , layer='레이져', color='2')    
    circle70x2 =  rx1 +  (LCD-250) 
    circle70y2 =  circle70y1
    circle_cross(doc, circle70x2 , circle70y2 , 70 , layer='레이져', color='2')        
    dim_diameter(doc, (circle70x1, circle70y1), 70, angle=225, dimstyle="JKW", override=None) # 70은 지름 기록
    circle15x1 =  rx1 + 15
    circle15y1 =  ry1 + 30.65
    circle_cross(doc, circle15x1 , circle15y1 , 15 , layer='레이져', color='2')    
    circle15x2 =  rx1 +  LCD - 15
    circle15y2 =  circle15y1
    circle_cross(doc, circle15x2 , circle15y2 , 15 , layer='레이져', color='2')        
    dim_diameter(doc, (circle15x1, circle15y1), 15, angle=225, dimstyle="JKW", override=None) # 70은 지름 기록
    dim_diameter(doc, (circle15x2, circle15y2), 15, angle=225, dimstyle="JKW", override=None) # 70은 지름 기록
    circle20x1 =  rx1 + 100
    circle20y1 =  ry1 + 431.6
    circle_cross(doc, circle20x1 , circle20y1 , 20 , layer='레이져', color='2')    
    circle20x2 =  rx1 +  LCD - 100
    circle20y2 =  circle20y1
    circle_cross(doc, circle20x2 , circle20y2 , 20 , layer='레이져', color='2')        
    dim_diameter(doc, (circle20x1, circle20y1), 20, angle=225, dimstyle="JKW", override=None) 
    dim_diameter(doc, (circle20x2, circle20y2), 20, angle=315, dimstyle="JKW", override=None) 

    # 중앙 11파이 단공
    circle11x1 =  rx1 + LCD/2
    circle11y1 =  ry1 + 476.6
    circle_cross(doc, circle11x1 , circle11y1 , 11 , layer='레이져', color='2')            
    dim_diameter(doc, (circle11x1, circle11y1), 11, angle=225, dimstyle="JKW", override=None) 

    # 하부 치수선
    dim(doc, x17, y17, x6, y6, 262, text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, x17, y17, x1, y1, 196, text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, x2, y2, x6, y6, 116, text_height=0.22, text_gap=0.07, direction="down")

    # 상부 치수선
    dim(doc, x12, y12, circle70x1 , circle70y1 , 98, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, circle70x1 , circle70y1 , circle70x2 , circle70y2 ,  412, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  circle70x2 , circle70y2, x11, y11 , 412, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  x12, y12, circle11x1, circle11y1, 176, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  circle11x1, circle11y1, x11, y11, 190, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  x12, y12, circle20x1, circle20y1, 240, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  circle20x1, circle20y1, circle20x2, circle20y2, 325, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  circle20x2, circle20y2, x11, y11, 325, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  x12, y12, x11, y11, 308, text_height=0.22, text_gap=0.07, direction="up")

    # 좌측 치수
    dim(doc, x12, y12, circle20x1, circle20y1,  60 , direction="left")
    dim(doc, circle15x1, circle15y1,  x1, y1, 107 , direction="left")
    dim(doc, x17, y17,  x1, y1, 147 , direction="left")
    dim(doc, x12, y12, x17, y17,   147 , direction="left")

    #70파이 치수선 좌측
    dim(doc, circle70x2, circle70y2, circle70x2 - 150, y1, 150 , direction="left")
    dim(doc, circle70x2, circle70y2, circle70x2 - 150, y12, 150 , direction="left")

    # 우측 치수선
    dim(doc, x11, y10, x10, y11,  200 , direction="right", option="reverse")
    dim(doc, x10, y10, x9, y9,  200  , direction="right")
    dim(doc,  x9, y9, x8, y8,  270  , direction="right")
    dim(doc,  x8, y8, x7, y7,  200  , direction="right")
    dim(doc,  x7, y7, x4, y4, 270  , direction="right")
    dim(doc,  x4, y4, x3, y3, 200  , direction="right")
    dim(doc,  x3, y3, x2, y2, 130  , direction="right")
    dim(doc,  x11, y11, x2, y2, 342  , direction="right")

    #################################################################################################
    # 3page Frame cover
    #################################################################################################  
    # 2page LC frame 삽입
    rx1 = abs_x + math.ceil(math.ceil(frame_scale * BasicXscale) / 100) * 100 + 1500 + TargetXscale + 300
    ry1 = frameYpos + 400
    if LCframeMaterial == 'SPCC 1.6T' : # 기본색상임      
        insert_block(rx1 , ry1 , "lc026_framecover_laser")        
        insert_block(rx1 , ry1 , "lc026_framecover_dim")        
    else:
        insert_block(rx1 , ry1 , "lc026_framecover_laser_height100")        
        insert_block(rx1 , ry1 , "lc026_framecover_dim_height100")                

    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 1217
    TargetYscale = 670
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
    # print("3page 스케일 비율 : " + str(frame_scale))  
    insert_frame(  rx1 - 330 , frameYpos , frame_scale, "drawings_frame", workplace)           
    if LCframeMaterial == 'SPCC 1.6T' : # 기본색상임    
        textstr = f" : {su*2} EA"    
        # draw_Text(doc, rx1+780 , ry1 + 35 , 10, str(textstr), layer='0')    
        # draw_Text(doc, rx1+780 , ry1 + 35 - 25 , 10, str(textstr), layer='0')    
    else:    
        descript_x = rx1 + 600
        textstr = f"Part Name : Frame Cover"    
        draw_Text(doc, descript_x , ry1 + 120 , 12, str(textstr), layer='0')    
        textstr = f"Mat.Spec : {LCframeMaterial}"
        draw_Text(doc, descript_x, ry1 + 90 , 12, str(textstr), layer='0')    
        textstr = f"Size : 350 x 145.6"            
        draw_Text(doc, descript_x, ry1 + 60 , 12, str(textstr), layer='0')        
        textstr = f"Quantity : 본도동일 {su*2} EA"    
        draw_Text(doc, descript_x, ry1 + 30 , 12, str(textstr), layer='0')    
        textstr = f"           본도반대 {su*2} EA"    
        draw_Text(doc, descript_x, ry1  , 12, str(textstr), layer='0')    

    #############################################################################################
    # 4page polygal
    #############################################################################################
    abs_x = rx1 - 330 + 1670
    abs_y = frameYpos # 첫 페이지의 도면 하단 기준

    # 폴리갈 크기 산출
    polygal_width = LCW - 620
    if CD>=2000 :
        polygal_height = (CD - 29*2-30)/3     
        ploygal_su = 3
    else:
        polygal_height = LCD/2 - 15    
        ploygal_su = 2
    
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2100 + (polygal_width - 930)
    TargetYscale = 1280 + (polygal_height - 710)
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
    # print("4page 스케일 비율 : " + str(frame_scale))  
    insert_frame(  abs_x , frameYpos , frame_scale, "drawings_frame", workplace)           
    textstr = f" : {su*2} EA"    
    # draw_Text(doc, rx1+780 , ry1 + 35 , 10, str(textstr), layer='0')    
    # draw_Text(doc, rx1+780 , ry1 + 35 - 25 , 10, str(textstr), layer='0')    

    # 폴리갈 전개도 기준점 위치
    rx1 = abs_x + 220
    ry1 = abs_y + 340
    # 폴리갈 쫄대 하단 17높이 표기
    insideHeight = 17    
    x1 = rx1 
    y1 = ry1
    x2 = rx1 + polygal_width 
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')      
    # 폴리갈 쫄대 상부 17높이 표기
    insideHeight = 17    
    x1 = rx1 
    y1 = ry1 + polygal_height + 10 - 17
    x2 = rx1 + polygal_width 
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')    

    insideHeight = polygal_height 
    x1 = rx1 
    y1 = ry1 + 5
    x2 = rx1 + polygal_width 
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='6')            
    # 폴리갈 해치    
    hatch = msp.add_hatch(dxfattribs={'layer': '55'})  # 5는 파란색 코드입니다.

    # "CORK" 패턴으로 해치 패턴 설정
    hatch.set_pattern_fill("CORK", scale=5.0, angle=90, color='5')  # 여기서 scale은 패턴의 크기를 조정합니다. 색상지정은 여기서 해야 된다.

    # 경계선 추가 (여기서는 사각형을 예로 듭니다)
    hatch.paths.add_polyline_path(
        [(x1, y1), (x2, y1), (x2, y2), (x1, y2)], 
        is_closed=True
    )

    textstr =  f"폴리갈 몰딩 절단 칫수 - {polygal_width}"
    dim(doc, x1, y2+5, x2, y2+5, 140, text_height=0.22, direction="up", text= textstr) 
    textstr =  f"폴리갈  - {polygal_width}"
    dim(doc, x1, y2+5, x2, y2+5, 65, text_height=0.22, direction="up", text= textstr) 

    # 우측 치수선    
    dim(doc, x2, y2, x2, y2+5, 80, direction ="right")
    dim(doc, x2, y1, x2, y1-5, 80, direction ="right")   
    textstr =  f"폴리갈  - {math.ceil(polygal_height)}"    
    dim(doc, x2, y1, x2, y2, 80,  direction="right" , text = textstr)     
    
    textstr =  f"폴리갈 몰딩 외경 - {math.ceil(polygal_height+10)}"
    dim(doc, x2, y1-5, x2, y2+5, 140,  direction="right" , text = textstr)     

    # 우측 단면도 
    insideHeight = polygal_height   
    rx1 = x2 + 380
    ry1 = ry1 + 5
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + 6
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='2')     

    insert_block(x1 , y1 , "polygal_rib_bottom")  
    insert_block(x1 , y2 , "polygal_rib")  

    # 단면도 치수선
    dim(doc, x2, y2, x2, y2 + 5, 36, direction = "right")
    dim(doc, x1, y1-5, x1, y2 + 5, 140, direction = "left")
    dim(doc,  x1, y2, x2, y2, 72, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  x1-1.2, y2+5, x2+1.2, y2+5, 112, text_height=0.22, text_gap=0.07, direction="up")
    
    # description
    insideHeight = polygal_height   
    rx1 = x2 + 100
    ry1 = ry1 + LCD/4 + 100
    x1 = rx1 
    y1 = ry1 

    textstr = f"Part Name : 중판"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : 폴리갈 6T"    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : {polygal_width} x {polygal_height}"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*ploygal_su} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')    

    # car inside box 표기
    xx =  x2 + 100
    yy =  y1 - 480
    rectangle(doc, xx,yy,xx+200*2,yy+100*2,layer='0')      
    line(doc, xx,yy+50*2,xx+200*2,yy+50*2)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50*2, yy+50*2 , 24, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30*2, yy+25*2 , 24, str(textstr), layer='0')     

    #############################################################################################
    # 5page Frame 용 AL Profile
    #############################################################################################
    abs_x = rx1 - 330 + 1670

    # 폴리갈 크기 산출
    polygal_width = LCW - 620
    polygal_height = LCD/2 - 8
    
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2100 + (LCD - 1450)
    TargetYscale = 1280 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
    # print("5page 스케일 비율 : " + str(frame_scale))  
    insert_frame(  abs_x , frameYpos , frame_scale, "drawings_frame", workplace)           

    # 그리드 14개 1.5 간격 12.5 시작 2.50 시작 후 1.5 간격 13개 확장
    # 프로파일 전개도 기준점 위치
    rx1 = abs_x + 180
    ry1 = abs_y + 800
    insideHeight = 100   
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + LCD - 4
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')  
    line(doc, x1, y1+12.5, x2, y1+12.5,  layer='0')    # 0 선  
    line(doc, x1, y2-12.5, x2, y2-12.5,  layer='0')    # 0 선      
    # 회색선 다량생성    
    line(doc, x1, y1+12.5+2.50, x2, y1+12.5+2.50,  layer='2')    # 0 선  
    line(doc, x1, y1+37.5, x2, y1+37.5,  layer='2')    # 0 선  
    line(doc, x1, y1+12.5+2.50+50, x2, y1+12.5+2.50+50,  layer='2')    # 0 선      
    for i in range(1, 13 + 1):
        line(doc, x1, y1+12.5+2.50+i*1.5, x2, y1+12.5+2.50+i*1.5,  layer='2')    # 0 선  
        line(doc, x1, y1+50+12.5+2.50+i*1.5, x2, y1+50+12.5+2.50+i*1.5,  layer='2')    # 0 선  

    rx1 = abs_x + 180
    ry1 = abs_y + 800 + 400
    insideHeight = 100   
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + LCD - 4
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')  
    line(doc, x1, y1+12.5, x2, y1+12.5,  layer='0')    # 0 선  
    line(doc, x1, y2-12.5, x2, y2-12.5,  layer='0')    # 0 선  
    # 회색선 다량생성    
    line(doc, x1, y1+12.5+2.50, x2, y1+12.5+2.50,  layer='2')    # 0 선  
    line(doc, x1, y1+37.5, x2, y1+37.5,  layer='2')    # 0 선  
    line(doc, x1, y1+12.5+2.50+50, x2, y1+12.5+2.50+50,  layer='2')    # 0 선      
    for i in range(1, 13 + 1):
        line(doc, x1, y1+12.5+2.50+i*1.5, x2, y1+12.5+2.50+i*1.5,  layer='2')    # 0 선  
        line(doc, x1, y1+50+12.5+2.50+i*1.5, x2, y1+50+12.5+2.50+i*1.5,  layer='2')    # 0 선  

    # 선홍색 LED BAR - 10000K - 1400L (누드&클립&잭타입)
    # 100 단위일때 기본 23mm 떨어짐이지만, 비규격일때는 계산 해주는 로직이 필요하다.    

    if CD-100 == LedBarLength:
        LedSpace = 23
    else:
        LedSpace = 23 - ( LedBarLength-(CD-100) ) / 2

    rx1 = abs_x + 180 + LedSpace
    ry1 = abs_y + 800 + 50 - 2.5
    insideHeight = 5   
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + LCD - 4 - LedSpace * 2
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='6')  

    rx1 = abs_x + 180 + LedSpace
    ry1 = abs_y + 800 + 400 + 50 - 2.5
    insideHeight = 5 
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + LCD -4 - LedSpace *2
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='6')  
    insert_block(x2  , y2 - 2.5 , "lc026_profile_wire")  # smps    

   # 단공 3.2파이 6개
    x_positions = [180 + 173, 180 + (LCD-4)/2, 180 + LCD-4 - 173]
    y_positions = [800 + 50, 800 + 400 + 50]

    for y in y_positions:
        for x in x_positions:
            rx1 = abs_x + x
            ry1 = abs_y + y
            circle_cross(doc, rx1, ry1, 3.2, layer='0', color='4')

    # 블럭삽입 단면도, SMPS 위치
    insert_block(rx1 + 400 , ry1 , "lc026_profile_section")  

    # 치수선
    rx1 = abs_x + 180  
    ry1 = abs_y + 800 + 500
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + LCD - 4
    y2 = ry1        
    dim(doc,  rx1, ry1, x2, y2, 76, text_height=0.22, text_gap=0.07, direction="up")
    rx1 = abs_x + 180 + LedSpace
    ry1 = abs_y + 800 + 400 + 50 - 2.5
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + LCD - 4 - LedSpace*2
    y2 = ry1            
    textstr =  f"{Leditem} - {Ledwatt} - {LedBarLength}L (누드&클립&잭타입)"
    dim_string(doc, rx1, ry1, x2, y2, 140,  textstr, text_height=0.20, text_gap=0.07, direction="down")    

    # description
    insideHeight = polygal_height   
    rx1 = abs_x + 700
    ry1 = abs_y + 600
    x1 = rx1
    y1 = ry1

    textstr = f"Part Name : Frame 용 AL Profile"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : AL"    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : 100 x 34.6 x {LCD-4} "    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')    

    # car inside box 표기
    xx =  x1 + 500
    yy =  y1-180
    rectangle(doc, xx,yy,xx+200*2,yy+100*2,layer='0')      
    line(doc, xx,yy+50*2,xx+200*2,yy+50*2)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50*2, yy+50*2 , 24, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30*2, yy+25*2 , 24, str(textstr), layer='0')    


    # 레이져 도면틀 생성 (공통) 레이져에 배열할때 편리한 기능
    # 1219*1950 철판크기 샘플
    rectangle(doc, 10000, 5000,10000+3400 , 5000+ 2340, layer='0')  
    rectangle(doc, 10000+850, 5000+130,10000+850+1950 , 5000+130+1219, layer='레이져')  
    dim(doc, 10000+850, 5000+130+1219, 10000+850+1950, 5000+130+1219, 100, text_height=0.22, text_gap=0.07, direction="up")
    dim_vertical_left(doc,  10000+850, 5000+130, 10000+850 , 5000+130+1219, 150, "JKW", text_height=0.22,  text_gap=0.07)  

    textstr = f"{secondord} - {workplace} - {drawdate_str_short} 출도, {deadlinedate_str_short} 납기"    
    draw_Text(doc, 10000+300 , 5000+ 2340 - 200 ,  80, str(textstr), layer='0')        
    textstr = f"SPCC Tx1219x      =  장"        
    draw_Text(doc, 10000+1000 , 5000+ 2340 - 200 - 200 ,  90, str(textstr), layer='0')        
    
####################################################################################################################################################################################
# N20 천장 자동작도
####################################################################################################################################################################################
def lcN20():    
    global args  #수정할때는 global 선언이 필요하다. 단순히 읽기만 할때는 필요없다.    
    global start_time
    global workplace, drawdate, deadlinedate, lctype, secondord, text_style, exit_program, Vcut_rate, Vcut, program_message, formatted_date, text_style_name
    global CW_raw, CD_raw, panel_thickness, su, LCframeMaterial, LCplateMaterial
    global input_CD, input_CW,CD,CW,drawdate_str , deadlinedate_str, drawdate_str_short, deadlinedate_str_short
    global doc, msp , T5_is, text_style, distanceXpos, distanceYpos
    # 2page 덴크리 다온텍등 등기구 설명 그림
    insert_block(0 , 0 , "lcN20_2page_frame")        
    abs_x = 0
    abs_y = 0
    # LC 크기 지정
    LCD = CD-50
    LCW = CW-50

    # watt 계산 공식 적용해야 함
    # LED바 기장(단위:m) X 수량 X 15 = 60W 이하
    # LED바 1개당 watt 산출 m단위로 계산.. /1000

    wattCalculate = math.ceil((LCD - 3)/1000 * 2 * 15)

    # print("wattCalculate: " + str(wattCalculate))  

    if(wattCalculate <= 60):
        watt = 60
    elif(wattCalculate > 60 and wattCalculate < 100):
        watt = 100
    elif(wattCalculate >= 100 and wattCalculate < 150):
        watt = 150
    elif(wattCalculate >= 150 and wattCalculate < 200):
        watt = 200
    elif(wattCalculate >= 200 and wattCalculate < 200):
        watt = 300
                
    # print("L/C 1개당 규격설정 wattCalculate: " + str(watt))      

    # T5 규격에 해당되는지 체크하는 부분
    T5_standard = 0
    if T5_is == '있음':
        adjusted_CD = CD - 50  # CD에서 50 뺌

        # 규격 결정
        if adjusted_CD > 1500:
            T5_standard = 1500
        elif adjusted_CD > 1200:
            T5_standard = 1200
        else:
            T5_standard = 900

        # print(f"선택된 규격: {T5_standard}")
    else:
        # print("T5_is 가 '있음'이 아니므로 규격 선택 안 함")       
        pass

    if T5_standard > 0 : 
        Ledcompany = f"(등기구업체 : 엘엠팩트)"        
        Leditem = "T5"
        Ledwatt = "6500K"
        LedBarLength = T5_standard
        watt = "내장"
    else:
        Ledcompany = f"(등기구업체 : 다온텍)"
        Leditem = "LED BAR"
        Ledwatt = "10000K"
        LedBarLength = CalculateLEDbar(CD)
        watt = f"{str(watt)}W"

    ledsu = 4

    if(CD>2000):
        ledsu = 4

    # 도면틀 넣기
    BasicXscale = 2560
    BasicYscale = 1550
    TargetXscale = 800 
    TargetYscale = 300 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale

    frame_scale = 1    

    # 현장명
    textstr = workplace       
    x = abs_x + 80
    y = abs_y + 1000 + 218
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 발주처
    textstr = secondord
    x = abs_x + 80
    y = abs_y + 1000 + 218 - 30
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 타입
    textstr = lctype
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 수량
    textstr = su
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 카사이즈
    textstr = f"W{CW} x D{CD}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 색상
    textstr = f"L / C : 흑색무광"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')

    x = abs_x + 120
    y = abs_y + 1000 + 90
    draw_Text(doc, x, y , 12, str(Ledcompany), layer='0')    
    # (아답터용량 : 60W )
    textstr = f"(아답터용량 : {watt} )"
    x = abs_x + 120
    y = abs_y + 1000 + 90 - 20
    draw_Text(doc, x, y , 12, str(textstr), layer='0')
    textstr = f"*. 출도일자 : {drawdate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 
    draw_Text(doc, x, y , 8, str(textstr), layer='0')    
    textstr = f"*. 납품일자 : {deadlinedate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 - 15
    draw_Text(doc, x, y , 8, str(textstr), layer='0')

    # led바에 대한 내용 기술
    # LED BAR or T5:       

    textstr = f"{Leditem}-{LedBarLength}L={ledsu*su}(EA)"
    x = abs_x + 720 + 60
    y = abs_y + 650
    draw_Text(doc, x, y , 16, str(textstr), layer='0')        

    textstr = f"SMPS({watt})={su*2}(EA)"
    x = abs_x + 720 + 60
    y = abs_y + 650 - 35
    draw_Text(doc, x, y , 16, str(textstr), layer='0')    
    textstr = f"({watt})"    
    draw_Text(doc, 210, 380 , 20, str(textstr), layer='0')            
    draw_Text(doc, 720, 380 , 20, str(textstr), layer='0')    

    # Type
    textstr = f"{lctype}타입"
    x = abs_x + 711 + 100
    y = abs_y + 600 - 20
    draw_Text(doc, x, y , 20, str(textstr), layer='0')    
    # 현장명
    textstr = f"*.현장명 : {secondord}-{workplace} - {su}(set)"
    x = abs_x + 30
    y = abs_y + 650 
    draw_Text(doc, x, y , 20, str(textstr), layer='0')    
    # (등기구업체 : 덴크리)    
    x = abs_x + 30 + 200
    y = abs_y + 600 
    draw_Text(doc, x, y , 20, str(Ledcompany), layer='0')    
    # LED BAR-(누드&클립&잭타입)10000K-1400L
    # draw_Text_direction(doc, x, y, size, text, layer=None, rotation = 90)    
    textstr1 = f"{Leditem}-(누드&클립&잭타입)"
    textstr2 = f"{Ledwatt}-{LedBarLength}L"

    x = abs_x + 235
    y = abs_y + 70    

    # 문자열 회전해서 보여주기
    draw_Text_direction(doc, x, y , 14, str(textstr1), layer='0', rotation = 90)    
    draw_Text_direction(doc, x+30, y+80 , 14, str(textstr2), layer='0', rotation = 90)    

    draw_Text_direction(doc, x+128, y , 14, str(textstr1), layer='0', rotation = 90)    
    draw_Text_direction(doc, x+30+128, y+80 , 14, str(textstr2), layer='0', rotation = 90)    

    draw_Text_direction(doc, x+128*2, y , 14, str(textstr1), layer='0', rotation = 90)    
    draw_Text_direction(doc, x+30+128*2, y+80 , 14, str(textstr2), layer='0', rotation = 90)    

    draw_Text_direction(doc, x+128*3+110, y , 14, str(textstr1), layer='0', rotation = 90)    
    draw_Text_direction(doc, x+30+128*3+110, y+80 , 14, str(textstr2), layer='0', rotation = 90)    
   
    # car inside box 표기
    xx =  790
    yy =  45
    rectangle(doc, xx,yy,xx+200,yy+100,layer='0')      
    line(doc, xx,yy+50,xx+200,yy+50)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50, yy+50 , 12, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30, yy+25 , 12, str(textstr), layer='0')

    # Assy Car case    
    #####################################################################################################################################
    #1page 조립도 만들기
    ##################################################################################################################################### 
    # LC ASSY 상부 형상
    frameXpos = abs_x + 1700
    abs_x = frameXpos + 1700    
    abs_y = abs_y + CD + 500
    insert_block(abs_x , abs_y , "lcN20_top_left")    

    x = abs_x + CW 
    y = abs_y 
    insert_block(x , y , "lcN20_top_right")  

    polygal_width = CW - 342.5*2

    x1 = abs_x + 342.5
    y1 = abs_y + 60
    x2 = x1 + (CW - 342.5*2)
    y2 = y1

    # 상부에 위치한 단면도 폴리갈 표현하기
    line(doc, x1, y1, x1, y1+14, layer='CL') # 적색    
    lineto(doc, x1+1.2, y1+14, layer='CL')   
    lineto(doc, x1+1.2, y1+ 1.2, layer='CL')   
    lineto(doc, x2-1.2, y1+ 1.2, layer='CL')   
    lineto(doc, x2-1.2, y1+14 , layer='CL')   
    lineto(doc, x2, y1+14 , layer='CL')   
    lineto(doc, x2, y1 , layer='CL')   
    lineto(doc, x1, y1 , layer='CL')   
    textstr =  f"폴리갈 및 중판 폭 = {math.ceil(polygal_width)}"
    dim(doc, x1, y2+14, x2, y2+14, 172,  text_height=0.32,  direction="up", text=textstr)      

    # 상부 폴리갈 위 
    line(doc, x1+1.2, y1+6, x2-1.2 , y1+6, layer='2') # 회색

    # 상부선 2.3mm 간격
    X1 = abs_x - 25
    Y1 = abs_y + 150
    X2 = X1 + CW + 50
    Y2 = Y1
    line(doc, X1, Y1, X2, Y2, layer='2')     # 회색
    lineto(doc, X2, Y1+2.3, layer='2')    
    lineto(doc, X1, Y1+2.3, layer='2')    
    lineto(doc, X1, Y1, layer='2')     

    # Assy 상부 치수
    x1 = abs_x + 297.5
    y1 = abs_y + 150 
    x2 = abs_x + CW - 297.5
    y2 = y1        
    dim(doc, x1, y1, x2, y2, 211,  text_height=0.30, text_gap=0.07, direction="up")
    # Assy 중심 치수
    x1 = abs_x + 371
    y1 = abs_y 
    x2 = abs_x + CW - 371
    y2 = y1    
    dim(doc, x1, y1, x2, y2, 116, text_height=0.30, text_gap=0.07, direction="down")
    x1 = abs_x + 25
    y1 = abs_y + 36
    x2 = abs_x + CW - 25
    y2 = y1        
    textstr =  f"Light Case Wide (W) = {CW-50}"    
    dim(doc, x1, y1, x2, y2, 300,  text_height=0.25, direction="down", text=textstr)   
    # Assy 상단 하부 치수
    x1 = abs_x 
    y1 = abs_y 
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"Car Inside (CW) = {CW}"
    dim(doc, x1, y1, x2, y2, 350, text_height=0.25, direction="down", text=textstr)   

    insideHeight = 25
    abs_y = abs_y - CD - 800    
    x1 = abs_x 
    y1 = abs_y 
    x2 = abs_x + CW
    y2 = abs_y + CD    
    rectangle(doc, x1, y1, x2, y2, layer='0')    
    x1 = abs_x + insideHeight
    y1 = abs_y + insideHeight
    x2 = abs_x + CW - insideHeight
    y2 = abs_y + CD - insideHeight
    line(doc, x1, y1, x1+2, y1, layer='0')    
    line(doc, x2, y1, x2-2, y1, layer='0')    
    line(doc, x1, y2, x1+2, y2, layer='0')    
    line(doc, x2, y2, x2-2, y2, layer='0')    
    line(doc, x1, y1, x1,   y2, layer='0')    
    line(doc, x2, y1, x2,   y2, layer='0')        
    
     # 좌측 fan assy
    insert_block(abs_x + 15 , abs_y  + CD - 53 , "lcN20_fan")     # 상단에서 90 떨어진 지점이 53 + 37
    insert_block(abs_x + insideHeight  , abs_y  + CD - insideHeight , "lcN20_rib_top_left")    
    insert_block(abs_x + CW - insideHeight  , abs_y  + CD - insideHeight , "lcN20_rib_top_right")    
    insert_block(abs_x + insideHeight  , abs_y  + insideHeight , "lcN20_rib_bottom_left")    
    insert_block(abs_x + CW - insideHeight , abs_y  + insideHeight , "lcN20_rib_bottom_right")    

    # 중간에 단공 2개
    circle_cross(doc, abs_x + 112.5 , abs_y + CD/2 , 11 , layer='2', color='2')       
    circle_cross(doc, abs_x + CW - 112.5 , abs_y + CD/2 , 11 , layer='2', color='2')     
         

    # 내부 폴리갈 자리 하늘색    
    x1 = abs_x + 342
    y1 = abs_y + 28
    x2 = abs_x + CW - 342
    y2 = abs_y + CD - 28
    rectangle(doc, x1, y1, x2, y2, layer='2')   # 회색    
        
    # 인발 하늘색(왼쪽) LedBarLength 길이에 따라 가변됨    
    top_gap = (CD - LedBarLength)/2
    side_gaps = [46.91, 334.09, CW - 334.09 - 15, CW - 46.91 - 15]

    for side_gap in side_gaps:
        x1 = abs_x + side_gap
        y1 = abs_y + top_gap
        x2 = x1 + 15
        y2 = abs_y + CD - top_gap
        rectangle(doc, x1, y1, x2, y2, layer='4')  # 하늘색
    
    # 선홍색 50mm 폭 형상 4개
    arrayx = [26, 320, CW - 320 - 50, CW - 26 - 50]
    arrayx_detail = [4.8, 5.9, 30, 31, 44.5, 46, 48.5]

    for i, x in enumerate(arrayx):
        x1 = abs_x + x
        y1 = abs_y + 26.2
        x2 = x1 + 50
        y2 = abs_y + CD - 26.2
        rectangle(doc, x1, y1, x2, y2, layer='6')
        
        # 인덱스가 0 또는 2인 경우
        if i in [0, 2]:
            for xx in arrayx_detail:
                xx1 = x1 + xx            
                xx2 = xx1            
                line(doc, xx1, y1, xx2, y2, layer='hidden6')
        
        # 인덱스가 1 또는 3인 경우, 세부 사항을 역순으로 적용
        elif i in [1, 3]:
            for xx in reversed(arrayx_detail):
                xx1 = x1 + 50 - xx  # x1에서 50을 뺀 위치에서 시작하여 xx를 빼줍니다.
                xx2 = xx1
                line(doc, xx1, y1, xx2, y2, layer='hidden6')

    # 폴리갈 크기 산출
    polygal_width = LCW - 635

    if CD>=2000 :            
        polygal_height = math.floor((CD - 28*2 )/3 )

        # 폴리갈 쫄대좌표 상부1개, 중부2개, 하부1개
        arrY = [polygal_height,polygal_height,polygal_height*2,polygal_height]    

        prey = abs_y + 28
        for i, y in enumerate(arrY):  # i는 인덱스, y는 실제 배열 꼭 기억!
            x1 = abs_x + 342.5
            y1 = prey
            x2 = x1 + CW - 342.5*2
            y2 = abs_y + 28 + y
            prey = y2 
            if i== 0 :
                rectangle(doc, x1, y1, x2, y1+17, layer='2')
            elif i == 3:
                rectangle(doc, x1, abs_y - 28 , x2, abs_y - 28 - 17, layer='2')
            else:
                rectangle(doc, x1, y1-17, x2, y2, layer='2')
                rectangle(doc, x1, y1+17, x2, y2, layer='2')        
    
        # 폴리갈 해치    
        hatch = msp.add_hatch(dxfattribs={'layer': 'CL'})  # 5는 파란색 코드입니다.

        # "CORK" 패턴으로 해치 패턴 설정 및 90도 회전
        hatch.set_pattern_fill("CORK", scale=5.0, angle=90.0, color='5')  # 여기서 scale은 패턴 크기, angle은 패턴 회전 각도
        # 하부 폴리갈
        x1 = abs_x + 342.5
        y1 = abs_y + 28 + 5
        x2 = abs_x + CW - 342.5
        y2 = y1 + polygal_height - 10
        # 경계선 추가 (여기서는 사각형을 예로 듭니다)
        hatch.paths.add_polyline_path(
            [(x1, y1), (x2, y1), (x2, y2), (x1, y2)], 
            is_closed=True
        )
        rectangle(doc, x1, y1, x2, y2, layer='CL')       
        dim(doc,  x2, y2+5, x2, y1-5,  310+315,  text_height=0.35,  direction="right")

        # 중간 폴리갈
        x1 = abs_x + 342.5
        y1 = abs_y + 28 + polygal_height + 5
        x2 = abs_x + CW - 342.5
        y2 = y1 + polygal_height - 10
        # 경계선 추가 (여기서는 사각형을 예로 듭니다)
        hatch.paths.add_polyline_path(
            [(x1, y1), (x2, y1), (x2, y2), (x1, y2)], 
            is_closed=True
        )
        rectangle(doc, x1, y1, x2, y2, layer='CL')  
        dim(doc,  x2, y2+5, x2, y1-5,  310+315,  text_height=0.35,  direction="right")
        # 상부 폴리갈
        x1 = abs_x + 342.5
        y1 = abs_y + 28 + polygal_height*2 + 5
        x2 = abs_x + CW - 342.5
        y2 = y1 + polygal_height - 10
        # 경계선 추가 (여기서는 사각형을 예로 듭니다)
        hatch.paths.add_polyline_path(
            [(x1, y1), (x2, y1), (x2, y2), (x1, y2)], 
            is_closed=True
        )
        rectangle(doc, x1, y1, x2, y2, layer='CL')  
        dim(doc,  x2, y2+5, x2, y1-5,  310+315,  text_height=0.35,  direction="right")       

        # 상부 치수선
        # Assy 본천장 및 LC 상단 
        x1 = abs_x + 342.5
        y1 = abs_y + CD - 28
        x2 = x1 + (CW - 342.5*2)
        y2 = y1 
        textstr =  f"폴리갈 및 중판 폭 = {math.ceil(polygal_width)}"
        dim(doc, x1, y2, x2, y2, 130,  text_height=0.32, direction="up", text=textstr)   

    else:
        polygal_height = LCD/2 - 13

        # 폴리갈 쫄대좌표 상부1개, 중부2개, 하부1개
        arrY = [28, CD/2, CD/2-17, CD-28-17]    

        for i, y in enumerate(arrY):  # i는 인덱스, y는 실제 배열 꼭 기억!
            x1 = abs_x + 342.5
            y1 = abs_y + y
            x2 = x1 + CW - 342.5*2
            y2 = y1 + 17
            rectangle(doc, x1, y1, x2, y2, layer='2')
    
        # 폴리갈 해치    
        hatch = msp.add_hatch(dxfattribs={'layer': 'CL'})  # 5는 파란색 코드입니다.

        # "CORK" 패턴으로 해치 패턴 설정 및 90도 회전
        hatch.set_pattern_fill("CORK", scale=5.0, angle=90.0, color='5')  # 여기서 scale은 패턴 크기, angle은 패턴 회전 각도
        # 하부 절반 폴리갈
        x1 = abs_x + 342.5
        y1 = abs_y + 33
        x2 = abs_x + CW - 342.5
        y2 = abs_y + CD/2 - 5
        # 경계선 추가 (여기서는 사각형을 예로 듭니다)
        hatch.paths.add_polyline_path(
            [(x1, y1), (x2, y1), (x2, y2), (x1, y2)], 
            is_closed=True
        )
        rectangle(doc, x1, y1, x2, y2, layer='CL')  

        # 상부 절반 폴리갈    
        x1 = abs_x + 342.5
        y1 = abs_y + CD/2 + 5
        x2 = abs_x + CW - 342.5
        y2 = abs_y + CD - 33
        # 경계선 추가 (여기서는 사각형을 예로 듭니다)
        hatch.paths.add_polyline_path(
            [(x1, y1), (x2, y1), (x2, y2), (x1, y2)], 
            is_closed=True
        )
        rectangle(doc, x1, y1, x2, y2, layer='CL')  

        # 상부 치수선
        # Assy 본천장 및 LC 상단 
        x1 = abs_x + 342.5
        y1 = abs_y + CD - 28
        x2 = x1 + (CW - 342.5*2)
        y2 = y1 
        textstr =  f"폴리갈 및 중판 폭 = {math.ceil(polygal_width)}"
        dim(doc, x1, y2, x2, y2, 130,  text_height=0.32, direction="up", text=textstr)  

        x1 = abs_x + CW - 342.5
        y1 = abs_y + CD - 28
        x2 = x1
        y2 = abs_y + CD/2        
        textstr =  f"중판 {math.ceil(CD/2 - 28)}"    
        dim(doc,  x2, y2, x1, y1,  310+315,  text_height=0.35,  direction="right" , text=textstr)   

        x1 = abs_x + CW - 342.5
        y1 = abs_y + CD/2
        x2 = x1
        y2 = abs_y + 28
        textstr =  f"중판 {math.ceil(CD/2 - 28)}"    
        dim(doc,  x2, y2, x1, y1,  310+315,  text_height=0.35,  direction="right" , text=textstr)   

        
    # LC 본체 세로선 6개 (좌3, 우3)
    arrX = [100, 125, 310, CW - 310, CW-125, CW-100]

    for i, x in enumerate(arrX):  # i는 인덱스, y는 실제 배열 꼭 기억!
        x1 = abs_x + x
        y1 = abs_y + 25
        x2 = x1
        y2 = abs_y + CD - 25
        line(doc, x1, y1, x2, y2, layer='2')
       
    x1 = abs_x 
    y1 = abs_y + CD
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"L/C {CW-50}"
    dim(doc, x1 + 25 , y1 - 25, x2 - 25 , y2 - 25, 210,  text_height=0.32, direction="up", text=textstr)   

    x1 = abs_x 
    y1 = abs_y + CD
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"CAR INSIDE(CW) - {CW}"
    dim(doc, x1 , y1 , x2  , y2, 290, text_height=0.32,  direction="up", text=textstr)   
    # 25 치수 표현
    dim(doc,  x1+25, y2, x1, y1, 185,  text_height=0.30, text_gap=0.07, direction="up")
    dim(doc, x2-25, y1, x2, y2, 185,  text_height=0.30, text_gap=0.07, direction="up")

    # 하부 치수선
    # Assy 본천장 및 LC 상단 
    x1 = abs_x + 25
    y1 = abs_y + 25
    x2 = x1 + 172.5
    y2 = y1 + 240     
    dim(doc, x1, y1, x2, y2, 136,  text_height=0.32, text_gap=0.07, direction="down")    
    x1 = x2
    y1 = y2
    x2 = abs_x + (CW - 25) - 172.5
    y2 = y1     
    dim(doc, x1, y1, x2, y2, 136 + 240,  text_height=0.32, text_gap=0.07, direction="down")    
    x1 = x2
    y1 = y2
    x2 = abs_x + (CW - 25)
    y2 = abs_y + 25
    dim(doc, x1, y1, x2, y2, 136 + 240,  text_height=0.32, text_gap=0.07, direction="down")    

    x1 = abs_x 
    y1 = abs_y 
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"Light Case Wide (W) = {CW-50}"
    dim(doc, x1 + 25 , y1 + 25, x2 - 25 , y2 + 25, 250,  text_height=0.32, direction="down", text=textstr)   

    x1 = abs_x + 46.91
    y1 = abs_y + CD - top_gap
    x2 = x1
    y2 = abs_y + top_gap
    textstr =  f"{Leditem}-{LedBarLength}L"    
    dim(doc,   x1, y1, x2, y2,   240,  text_height=0.35,  direction="left", text=textstr)   
    x1 = abs_x 
    y1 = abs_y + CD 
    x2 = x1 + 112.5
    y2 = abs_y + CD - 45
    dim(doc, x2, y2, x1, y1,  390 + 60 , text_height=0.40, direction="left")
    x1 = abs_x + 112.5
    y1 = abs_y + CD - 45
    x2 = x1
    y2 = abs_y + CD/2
    dim(doc, x1, y1, x2, y2, 390 + 60, text_height=0.40, direction="left")  
    x1 = x2
    y1 = abs_y + CD/2
    x2 = x1
    y2 = abs_y + 45
    dim(doc, x1, y1, x2, y2, 390 + 60, text_height=0.40, direction="left")    
    x1 = x2
    y1 = y2
    x2 = abs_x
    y2 = abs_y 
    dim(doc, x1, y1, x2, y2, 390 + 60, text_height=0.40, direction="left")    
    x1 = abs_x 
    y1 = abs_y + CD
    x2 = x1
    y2 = abs_y         
    textstr =  f"CAR INSIDE (CD) = {CD}"    
    dim(doc,  x1, y1, x2, y2,   500, text_height=0.35 , direction="left", text=textstr)   

    # 우측 치수선(290) 치수선 3개
    x1 = abs_x + CW - 25
    y1 = abs_y + CD - 25
    x2 = abs_x + CW - 197.5
    y2 = y1 - 290      
    dim(doc, x1, y1, x2, y2, 205, text_height=0.40, direction="right")  
    x1 = x2
    y1 = y2
    x2 = abs_x + CW - 112.5
    y2 = abs_y + 315
    dim(doc, x1, y1, x2, y2, 205 + 172.5, text_height=0.40, direction="right")  
    x1 = x2
    y1 = abs_y + 315
    x2 = abs_x + CW - 25
    y2 = abs_y + 25
    dim(doc, x1, y1, x2, y2, 265+27.5, text_height=0.40, direction="right")  

    # 판넬떨어짐 25 표현
    x1 = abs_x + CW 
    y1 = abs_y + CD
    x2 = abs_x + CW - 342.5
    y2 = y1-25
    dim(doc,   x2, y2, x1, y1,  310+315, text_height=0.35 , direction="right")
    x1 = abs_x + CW 
    y1 = abs_y 
    x2 = abs_x + CW - 342.5
    y2 = y1+25
    dim(doc,   x2, y2, x1, y1,  310+315, text_height=0.35 , direction="right")
  
    x1 = abs_x + CW - 25
    y1 = abs_y + CD - 25
    x2 = x1
    y2 = abs_y + 25        
    textstr =  f"L/C {CD-panel_thickness*2}"
    # dim_vertical_right_string(doc, x1, y1, x2, y2, 308, textstr , text_height=0.20,  text_gap=0.07)        
    dim(doc,   x2, y2, x1, y1,  410,  text_height=0.35, direction="right", text=textstr)   
    x1 = abs_x + CW
    y1 = abs_y + CD
    x2 = x1
    y2 = abs_y         
    textstr =  f"CAR INSIDE {CD}"    
    dim(doc,  x2, y2, x1, y1,  490,  text_height=0.35, direction="right" , text=textstr)   

    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1715
    TargetXscale = 5260 
    TargetYscale = 3330 + (CD - 1500)
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
    # print("스케일 비율 : " + str(frame_scale))        
    frameYpos = abs_y - 450 * frame_scale     
    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)       
    # 1page 상단에 car inside 표기
    x = frameXpos + CW/3
    y = abs_y + frame_scale*BasicYscale + 50
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, x, y , 200*frame_scale, str(textstr), layer='0')    

    frameXpos = frameXpos + TargetXscale + 1400

    # print("1페이지 이후 간격 더한 위치 frameXpos : " + str(frameXpos))   

    #################################################################################################
    # 2page frame 전개도
    #################################################################################################    
    # 2page LC frame 삽입(레이져 가공)
    rx1 = frameXpos  + 2350 + (CD-1500)
    ry1 = frameYpos + 1050
    insert_block(rx1 , ry1 , "lcN20_frame_section")    
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2730 + (CD - 1500)  # 기본 모형이 1500
    TargetYscale = 1650 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
    # print("1page 스케일 비율 : " + str(frame_scale))         
    insert_frame(  frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

    rx1 = frameXpos  + 350
    ry1 = ry1 - 245
    
    thickness = 1.2
    vcut = thickness / 2

    x1 = rx1 
    y1 = ry1
    x2 = x1 + 100
    y2 = y1    
    x3 = x2
    y3 = y2 + 25 - thickness
    x4 = x3
    y4 = y3 + 122.4 - thickness
    x5 = x4 + LCD - 200
    y5 = y4 
    x6 = x5 
    y6 = y5 - 122.4 + thickness 
    x7 = x6 
    y7 = y6 - 25 + thickness
    x8 = x7 + 100 
    y8 = y7
    x9 = x8 
    y9 = y8 + 25 - thickness
    x10 = x9 
    y10 = y9  + 62.8 - thickness
    x11 = x10 - 1.5 
    y11 = y10
    x12 = x11 
    y12 = y11 + 86
    x13 = x12 + 1.5
    y13 = y12 
    x14 = x13 
    y14 = y13 + vcut
    x15 = x14 
    y15 = y14 + 224 - thickness
    x16 = x15 
    y16 = y15 + vcut
    x17 = x16 - 1.5
    y17 = y16 
    x18 = x17 
    y18 = y17 + 62.6 - thickness
    x19 = x18
    y19 = y18 + 62.2 - thickness 
    x20 = x19
    y20 = y19 + vcut
    x21 = x20 + 1.5
    y21 = y20

    x22 = x21 
    y22 = y21 + 21.8 - vcut*3
    x23 = x22
    y23 = y22 + 100.3 - thickness
    x24 = x23 
    y24 = y23 + 25 - thickness/3  # 각도 90도가 아니어서 보정계산함
    x25 = x24 - LCD
    y25 = y24 
    x26 = x25 
    y26 = y25 - 25 + thickness/3 
    x27 = x26 
    y27 = y26 - 100.3 + thickness
    x28 = x27 
    y28 = y27 - 21.8 + vcut*3
    x29 = x28 + 1.5 
    y29 = y28 
    x30 = x29
    y30 = y29 - vcut

    x31 = x30
    y31 = y30 - 62.2 + thickness 
    x32 = x31 
    y32 = y31 - 62.6 + thickness
    x33 = x32 - 1.5
    y33 = y32 
    x34 = x33 
    y34 = y33 - vcut
    x35 = x34 
    y35 = y34 - 224 + thickness
    x36 = x35 
    y36 = y35 - vcut
    x37 = x36 + 1.5
    y37 = y36 
    x38 = x37 
    y38 = y37 - 86
    x39 = x38 - 1.5
    y39 = y38  
    x40 = x39
    y40 = y39 - 62.8 + thickness  

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 40
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
    # 4각 2개소     
    rectangle(doc, x1+50, y34 + (y31-y34)/2 - 15 , x1 + LCD/2 - 25, y34 + (y31-y34)/2 - 15 + 30, layer='레이져')  
    rectangle(doc,  x1 + LCD/2 + 25 , y34 + (y31-y34)/2 - 15 , x8 - 50, y34 + (y31-y34)/2 - 15 + 30, layer='레이져')  
      
    #절곡라인        
    line(doc, x40, y40, x3, y3,  layer='hidden')    # 절곡선  
    line(doc, x6, y6, x9, y9,  layer='hidden')    # 절곡선  
    line(doc, x35, y35, x14, y14,  layer='hidden')    # 절곡선  
    line(doc, x34, y34, x15, y15,  layer='hidden')    # 절곡선  
    line(doc, x30, y30, x19, y19,  layer='hidden')    # 절곡선  
    line(doc, x27, y27, x22, y22,  layer='hidden')    # 절곡선  
    line(doc, x26, y26, x23, y23,  layer='hidden')    # 절곡선  
    #반대 절곡라인
    line(doc, x31, y31, x18, y18,  layer='2')    # 절곡선  

    # 3.2파이 단공 하나 중앙
    circle3dot2x1 =  x1 + LCD/2
    circle3dot2y1 =  y34 + (y31-y34)/2
    circle_cross(doc, circle3dot2x1 , circle3dot2y1 , 3.2 , layer='레이져', color='0')    
    
    circle70x1 =  x1 + 290
    circle70y1 =  y35 + 112 - vcut
    circle_cross(doc, circle70x1 , circle70y1 , 70 , layer='레이져', color='0')    
    circle70x2 =  x8 - 290
    circle70y2 =  circle70y1
    circle_cross(doc, circle70x2 , circle70y2 , 70 , layer='레이져', color='0')        
    dim_diameter(doc, (circle70x1, circle70y1), 70, angle=225, dimstyle="JKW", override=None) # 70은 지름 기록
    
    # 20 파이 4개 (방향은 좌측에서 우측 위에서 아래 방향으로 )
    circle20x1 =  x1 + 20
    circle20y1 =  y34 + 24 - vcut
    circle_cross(doc, circle20x1 , circle20y1 , 20 , layer='레이져', color='0')    
    circle20x2 =  x8 - 20
    circle20y2 =  circle20y1
    circle_cross(doc, circle20x2 , circle20y2 , 20 , layer='레이져', color='0')        
    circle20x3 =  x1 + 20
    circle20y3 =  y35 - 24 + vcut
    circle_cross(doc, circle20x3 , circle20y3 , 20 , layer='레이져', color='0')    
    circle20x4 =  x8 - 20
    circle20y4 =  circle20y3
    circle_cross(doc, circle20x4 , circle20y4 , 20 , layer='레이져', color='0')        
    dim_diameter(doc, (circle20x1, circle20y1), 20, angle=225, dimstyle="JKW", override=None) 
    dim_diameter(doc, (circle20x2, circle20y2), 20, angle=315, dimstyle="JKW", override=None) 
    dim_diameter(doc, (circle20x3, circle20y3), 20, angle=315, dimstyle="JKW", override=None) 
    dim_diameter(doc, (circle20x4, circle20y4), 20, angle=315, dimstyle="JKW", override=None) 
    # 11파이 단공 5개
    circle11x1 =  x1 + 20
    circle11y1 =  y25 - 12.5
    circle11x2 =  circle3dot2x1
    circle11y2 =  circle11y1
    circle11x3 =  x8 - 20
    circle11y3 =  circle11y1
    circle11x4 =  x1 + 20
    circle11y4 =  y1 + 12.5
    circle11x5 =  x8 - 20
    circle11y5 =  circle11y4
    circle_cross(doc, circle11x1 , circle11y1 , 11 , layer='레이져', color='2')            
    circle_cross(doc, circle11x2 , circle11y2 , 11 , layer='레이져', color='2')            
    circle_cross(doc, circle11x3 , circle11y3 , 11 , layer='레이져', color='2')            
    circle_cross(doc, circle11x4 , circle11y4 , 11 , layer='레이져', color='2')            
    circle_cross(doc, circle11x5 , circle11y5 , 11 , layer='레이져', color='2')            
    dim_diameter(doc, (circle11x1, circle11y1), 11, angle=225, dimstyle="JKW", override=None) 
    dim_diameter(doc, (circle11x2, circle11y2), 11, angle=225, dimstyle="JKW", override=None) 
    dim_diameter(doc, (circle11x3, circle11y3), 11, angle=225, dimstyle="JKW", override=None) 
    dim_diameter(doc, (circle11x4, circle11y4), 11, angle=225, dimstyle="JKW", override=None) 
    dim_diameter(doc, (circle11x5, circle11y5), 11, angle=225, dimstyle="JKW", override=None) 

    # 홀간격 표시
    dim(doc, circle11x2, circle11y2,  circle11x2, circle11y2+12.5, 120 , direction="left")
    # 50 3개 표시
    dim(doc, x28 , y28, x28 + 50 , y34 + (y31-y34)/2 + 15  , 60 , text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, circle3dot2x1-25 , y17+ (y18-y17)/2 + 15, circle3dot2x1-25+50 , y17+ (y18-y17)/2 + 15 , 60 + 77.6 , text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  x21-50 , y17+ (y18-y17)/2 + 15 , x21 , y21 , 60 + 77.6 , text_height=0.22, text_gap=0.07, direction="up")

    # 1.5 간격
    dim(doc, x32, y32,  x33, y33, 50 , text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, x16, y16,  x17, y17, 50 , text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, x36, y36,  x37, y37, 50 , text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, x12, y12,  x13, y13, 50 , text_height=0.22, text_gap=0.07, direction="up")

    # 홀에 대한 치수선
    dim_vertical_left(doc, x5, y5, x7, y7, 80 ,"JKW", text_height=0.20,  text_gap=0.07)
    dim_vertical_left(doc, x5, y5, x5, y14, 80 ,"JKW", text_height=0.20,  text_gap=0.07)
    dim_vertical_left(doc, x7, y7, circle70x2, circle70y2, 450 ,"JKW", text_height=0.20,  text_gap=0.07)
    dim_vertical_right(doc, circle11x4, circle11y4, circle11x4, circle11y4 - 12.5, 120 ,"JKW", text_height=0.20,  text_gap=0.07)

    # 상부 치수선
    dim(doc, circle11x1 , circle11y1 , x25, y25,  116+12.5, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, circle11x1 , circle11y1, circle11x2 , circle11y2 , 116 + 12.5, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, circle11x2 , circle11y2, circle11x3 , circle11y3,   116 + 12.5, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, circle11x3 , circle11y3, x24, y24,  116 + 12.5, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, x25, y25, x24, y24,  116 + 60 , text_height=0.22, text_gap=0.07, direction="up")

    # 하부 치수선
    dim(doc, circle11x4, circle11y4, x1, y1, 80, text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, circle11x4, circle11y4,circle11x5, circle11y5, 80, text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, circle11x5, circle11y5, x8, y8, 80, text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, x1, y1, x2, y2, 80 + 50, text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, x2, y2, x7, y7, 80 + 50, text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, x7, y7, x8, y8, 80 + 50, text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, x1, y1, circle70x1, circle70y1, 80+150, text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, circle70x1, circle70y1, circle70x2, circle70y2, 80+150 + 283.5, text_height=0.22, text_gap=0.07, direction="down")
    dim(doc, circle70x2, circle70y2, x8, y8, 80+150 + 283.5, text_height=0.22, text_gap=0.07, direction="down")    

    # 좌측 치수
    dim(doc, x25, y25, x31 + 50 - 1.5, y31 - 16,  115 , direction = "left")
    dim(doc, x31 + 50 - 1.5, y31 - 16, x31 + 50 - 1.5, y31 - 16 - 30, 115 + 50 , direction = "left")
    dim(doc, x31 + 50 - 1.5, y31 - 16 - 30, x1 , y1 , 115 + 50 , direction = "left")

    dim(doc, x25, y25 , x28, y28 , 115 + 80 , direction = "left")
    dim(doc, x28, y28 , x33, y33 , 115 + 80 , direction = "left")
    dim(doc, x33, y33 , x36, y36 , 115 + 80 , direction = "left")
    dim(doc, x36, y36 , x39, y39 , 115 + 80 , direction = "left")
    dim(doc, x39, y39 , x1 , y1  , 115 + 80 , direction = "left")

    # 우측 치수
    dim(doc, x23, y23 ,x24, y24,   100  , direction = "right")
    dim(doc, x23, y23, x22, y22 ,  100  , direction = "right")
    dim(doc, x19, y19, x22, y22 ,  145 , direction = "right")
    dim(doc, x19, y19, x18, y18 ,  145 , direction = "right")
    dim(doc, x18, y18, x15, y15 ,  145 , direction = "right")  
    dim(doc, x15, y15, x14, y14 ,  145 , direction = "right")
    dim(doc, x14, y14, x9, y9 ,  145   , direction = "right")
    dim(doc, x9, y9, x8, y8 ,  145     , direction = "right")
    dim(doc, x24, y24, x8, y8 ,145+ 80 , direction = "right")
   
    # description    
    x1 = frameXpos + 1900 + (CD-1500)
    y1 = frameYpos + 605
    textstr = f"Part Name : Light Case Frame"    
    draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : {LCframeMaterial}"    
    draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
    textstr = f"Size : {LCD} x {math.floor((y24-y8)*10)/10}mm"    
    draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')    

    frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 3page
    #################################################################################################  
    # LC frame 삽입
        # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 1680
    TargetYscale = 1024 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
    # print("3page 스케일 비율 : " + str(frame_scale))  
    insert_frame(  frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)      

    rx1 = frameXpos + 510 + 330 - 668.53
    ry1 = frameYpos + 400 - 236 + 526.86
    insert_block(rx1 , ry1 , "lcN20_framecover_laser_left")         # 좌측 laser선
    insert_block(rx1+991.2 , ry1 , "lcN20_framecover_laser_right")        # 우측 laser선
    insert_block(rx1 , ry1 , "lcN20_framecover_dim")                # 치수선 

    # description     (좌측)    
    x1 = frameXpos + 215
    y1 = frameYpos + 415
    textstr = f"Part Name : Cover(좌)"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : {LCframeMaterial}"    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : 346 x 172.6mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')             

    # description     (우측)
    x1 = frameXpos + 1100
    y1 = frameYpos + 415
    textstr = f"Part Name : Cover(우)"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : {LCframeMaterial}"    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : 346 x 172.6mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')            

    frameXpos = frameXpos + TargetXscale + 400
    
    #################################################################################################
    # 4page bracket
    #################################################################################################  
    # 2page LC frame 삽입
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 420
    TargetYscale = 256 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale        
    insert_frame(  frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)      

    rx1 = frameXpos + 160 
    ry1 = frameYpos + 168
    insert_block(rx1 , ry1 , "lcN20_bracket_laser")       
    insert_block(rx1 , ry1 , "lcN20_bracket_dim")       

    # description 
    x1 = frameXpos + 250
    y1 = frameYpos + 200
    textstr = f"Part Name : 중판 이탈방지 B/K"    
    draw_Text(doc, x1 , y1 , 7 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : EGI 1.2T"    
    draw_Text(doc, x1 , y1 -15 , 7, str(textstr), layer='0')    
    textstr = f"Size : 40.2 x 30mm"    
    draw_Text(doc, x1 , y1 -30  , 7, str(textstr), layer='0')        
    textstr = f"Quantity : {su*4} EA"    
    draw_Text(doc, x1 , y1 - 45 , 7, str(textstr), layer='0')             

    frameXpos = frameXpos + TargetXscale + 400

    #############################################################################################
    # 5page 폴리갈
    #############################################################################################
    abs_x = frameXpos
    abs_y = frameYpos # 첫 페이지의 도면 하단 기준

    # 폴리갈 크기 산출
    polygal_width = LCW - 635
    if CD>=2000 :
        polygal_height = math.floor( (CD - 28*2-30)/3 )
        ploygal_su = 3
    else:
        polygal_height = LCD/2 - 13  
        ploygal_su = 2
    
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2100 + (polygal_width - 930)
    TargetYscale = 1280 + (polygal_height - 710)
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
    # print("5page 스케일 비율 : " + str(frame_scale))  
    insert_frame(  frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)           

    # 폴리갈 전개도 기준점 위치
    rx1 = abs_x + 220
    ry1 = abs_y + 340
    # 폴리갈 쫄대 하단 17높이 표기
    insideHeight = 17    
    x1 = rx1 
    y1 = ry1
    x2 = rx1 + polygal_width 
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')      
    # 폴리갈 쫄대 상부 17높이 표기
    insideHeight = 17    
    x1 = rx1 
    y1 = ry1 + polygal_height + 10 - 17
    x2 = rx1 + polygal_width 
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')    

    insideHeight = polygal_height 
    x1 = rx1 
    y1 = ry1 + 5
    x2 = rx1 + polygal_width 
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='6')            
    # 폴리갈 해치    
    hatch = msp.add_hatch(dxfattribs={'layer': '55'})  # 5는 파란색 코드입니다.

    # "CORK" 패턴으로 해치 패턴 설정
    hatch.set_pattern_fill("CORK", scale=10.0, angle= 90.0 , color='5')  # 여기서 scale은 패턴의 크기를 조정합니다. 색상지정은 여기서 해야 된다.

    # 경계선 추가 (여기서는 사각형을 예로 듭니다)
    hatch.paths.add_polyline_path(
        [(x1, y1), (x2, y1), (x2, y2), (x1, y2)], 
        is_closed=True
    )

    textstr =  f"폴리갈 몰딩 절단 칫수 - {polygal_width}"
    dim(doc, x1, y2+5, x2, y2+5, 140,   direction="up", text=textstr)
    textstr =  f"폴리갈  - {polygal_width}"
    dim(doc, x1, y2+5, x2, y2+5, 65,  direction="up", text=textstr)

    # 우측 치수선    
    dim(doc, x2, y2 , x2, y2+5,  80, direction = "right")
    dim(doc, x2, y1, x2, y1-5, 80, direction = "right")
    textstr =  f"폴리갈  - {math.ceil(polygal_height)}"    
    dim(doc, x2, y1, x2, y2, 80,   direction="right", text=textstr)
    
    textstr =  f"폴리갈 몰딩 외경 - {math.ceil(polygal_height+10)}"
    dim(doc, x2, y1-5, x2, y2+5, 140,  direction="right", text=textstr)

    # 우측 단면도 
    insideHeight = polygal_height   
    rx1 = x2 + 380
    ry1 = ry1 + 5
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + 6
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='2')     

    insert_block(x1 , y1 , "polygal_rib_bottom")  
    insert_block(x1 , y2 , "polygal_rib")  

    # 단면도 치수선
    dim(doc, x2, y2, x2, y2 + 5, 36, direction = "right")
    dim(doc, x1, y1-5, x1, y2 + 5, 140, direction = "left")
    dim(doc,  x1, y2, x2, y2, 72, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  x1-1.2, y2+5, x2+1.2, y2+5, 112, text_height=0.22, text_gap=0.07, direction="up")
    
    # description
    insideHeight = polygal_height   
    rx1 = x2 + 100
    ry1 = ry1 + LCD/4 + 100
    x1 = rx1 
    y1 = ry1 

    textstr = f"Part Name : 중판"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : 폴리갈 6T"    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : {math.floor(polygal_width)} x {math.floor(polygal_height)}"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*ploygal_su} EA"   
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')    

    # car inside box 표기
    xx =  x2 + 100
    yy =  y1 - 480
    rectangle(doc, xx,yy,xx+200*2,yy+100*2,layer='0')      
    line(doc, xx,yy+50*2,xx+200*2,yy+50*2)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50*2, yy+50*2 , 24, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30*2, yy+25*2 , 24, str(textstr), layer='0')        

    frameXpos = frameXpos + TargetXscale + 400

    #############################################################################################
    # 6 page
    #############################################################################################
    abs_x = frameXpos 

    # 등기구 크기 산출
    light_width = LCD - 3
    
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2100 + (LCD - 1450)
    TargetYscale = 1280 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale        
    insert_frame(  abs_x , frameYpos , frame_scale, "drawings_frame", workplace)           
    
    # 등기구 
    rx1 = abs_x + 180
    ry1 = abs_y + 600
    insideHeight = 60   
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + light_width
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')      
    for i in range(1, 29 + 1):
        line(doc, x1, y1+i*2, x2, y1+i*2,  layer='6')

    # 단면도 삽입
    insert_block( x2 + 147  , y1  , "lcN20_lightcover")         

    rx1 = abs_x + 180
    ry1 = abs_y + 700 + 450
    insideHeight = 84   
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + light_width
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')  
    line(doc, x1, y1+59, x2, y1+59,  layer='0')    # 0 선  

    # 단면도 삽입
    insert_block( x2 + 147  , y1 + 59 , "lcN20_lightframe") 


    # light frame 치수선
    rx1 = abs_x + 180  
    ry1 = abs_y + 700 + 500
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + light_width
    y2 = ry1        
    dim(doc,  rx1, ry1, x2, y2, 76, text_height=0.22, text_gap=0.07, direction="up")
    # light cover 치수선    
    ry1 = abs_y + 600 + 60
    y1 = ry1     
    y2 = ry1        
    dim(doc,  rx1, ry1, x2, y2, 76, text_height=0.22, text_gap=0.07, direction="up")

    # frame description
    insideHeight = polygal_height   
    rx1 = abs_x + 700
    ry1 = abs_y + 1040
    x1 = rx1 
    y1 = ry1 

    textstr = f"Part Name : 내, 외측 등기구 후레임"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : AL"    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : 84 x 58.8 x {light_width} "    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*4} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')    

    # 등커버 description
    insideHeight = polygal_height   
    rx1 = abs_x + 700
    ry1 = abs_y + 500
    x1 = rx1 
    y1 = ry1 

    textstr = f"Part Name : 등커버"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : PC"    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : 60 x 59 x {light_width} "    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*4} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')    

    # car inside box 표기
    xx =  x1 + 500
    yy =  y1-220
    rectangle(doc, xx,yy,xx+200*2,yy+100*2,layer='0')      
    line(doc, xx,yy+50*2,xx+200*2,yy+50*2)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50*2, yy+50*2 , 24, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30*2, yy+25*2 , 24, str(textstr), layer='0')        

    ####################################################################################################################################################
    # 레이져 도면틀 생성 (공통) 레이져에 배열할때 편리한 기능
    # 1219*1950 철판크기 샘플
    rectangle(doc, 10000, 5000,10000+3400 , 5000+ 2340, layer='0')  
    rectangle(doc, 10000+850, 5000+130,10000+850+1950 , 5000+130+1219, layer='레이져')  
    dim(doc, 10000+850, 5000+130+1219, 10000+850+1950, 5000+130+1219, 100, text_height=0.22, text_gap=0.07, direction="up")
    dim_vertical_left(doc,  10000+850, 5000+130, 10000+850 , 5000+130+1219, 150, "JKW", text_height=0.22,  text_gap=0.07)  

    textstr = f"{secondord} - {workplace} - {drawdate_str_short} 출도, {deadlinedate_str_short} 납기"    
    draw_Text(doc, 10000+300 , 5000+ 2340 - 200 ,  80, str(textstr), layer='0')        
    textstr = f"SPCC  Tx1219x1950=1장"        
    draw_Text(doc, 10000+1000 , 5000+ 2340 - 200 - 200 ,  90, str(textstr), layer='0')        
   
####################################################################################################################################################################################
# 032 천장 자동작도 (프레임홀수정함 2024/11/01)
####################################################################################################################################################################################
def lc032():   
    global args  #수정할때는 global 선언이 필요하다. 단순히 읽기만 할때는 필요없다.    
    global start_time
    global workplace, drawdate, deadlinedate, lctype, secondord, text_style, exit_program, Vcut_rate, Vcut, program_message, formatted_date, text_style_name
    global CW_raw, CD_raw, panel_thickness, su, LCframeMaterial, LCplateMaterial
    global input_CD, input_CW,CD,CW,drawdate_str , deadlinedate_str, drawdate_str_short, deadlinedate_str_short
    global doc, msp , T5_is, text_style, distanceXpos, distanceYpos

    # 2page 덴크리 다온텍등 등기구 설명 그림
    insert_block(0 , 0 , "lc032_2pageframe")        
    abs_x = 0
    abs_y = 0
    # LC 크기 지정 car inside에서 차감
    LCD = CD-50
    LCW = CW-50

    # watt 계산 공식 적용해야 함
    # LED바 기장(단위:m) X 수량 X 15 = 60W 이하
    # LED바 1개당 watt 산출 m단위로 계산.. /1000

    wattCalculate = math.ceil((LCD - 3)/1000 * 2 * 15)

    # print("wattCalculate: " + str(wattCalculate))  

    if(wattCalculate <= 60):
        watt = 60
    elif(wattCalculate > 60 and wattCalculate < 100):
        watt = 100
    elif(wattCalculate >= 100 and wattCalculate < 150):
        watt = 150
    elif(wattCalculate >= 150 and wattCalculate < 200):
        watt = 200
    elif(wattCalculate >= 200 and wattCalculate < 200):
        watt = 300
                
    # print("L/C 1개당 규격설정 wattCalculate: " + str(watt))  

    T5_standard = 0
    if T5_is == '있음':
        adjusted_CD = CD - 50  # CD에서 50 뺌

        # 규격 결정
        if adjusted_CD > 1500:
            T5_standard = 1500
        elif adjusted_CD > 1200:
            T5_standard = 1200
        else:
            T5_standard = 900

        # print(f"선택된 규격: {T5_standard}")
    else:
        # print("T5_is 가 '있음'이 아니므로 규격 선택 안 함")    
        pass

    # 도면틀 넣기
    BasicXscale = 2560
    BasicYscale = 1550
    TargetXscale = 800 
    TargetYscale = 300 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale

    frame_scale = 1    
    ################################################################################################################################################
    # 갑지 1
    ################################################################################################################################################
    # 현장명
    textstr = workplace       
    x = abs_x + 80
    y = abs_y + 1000 + 218
    draw_Text(doc, x, y , 6, str(textstr), layer='0')
    # 발주처
    textstr = secondord
    x = abs_x + 80
    y = abs_y + 1000 + 218 - 30
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 타입
    textstr = lctype
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 수량
    textstr = su
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 카사이즈
    textstr = f"W{CW} x D{CD}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 색상 LC
    textstr = f"L / C : {LCframeMaterial}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')    
    textstr = f"중판 : {LCplateMaterial}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18 - 12
    draw_Text(doc, x, y , 7, str(textstr), layer='0')

    # (등기구업체 : 덴크리)   
    if T5_standard > 0 : 
        Ledcompany = f"(등기구업체 : 엘엠팩트)"        
        Leditem = "T5"
        Ledwatt = "6500K"
        LedBarLength = T5_standard
        watt = "내장"
    else:
        Ledcompany = f"(등기구업체 : 다온텍)"
        Leditem = "LED BAR"
        Ledwatt = "10000K"
        LedBarLength = CalculateLEDbar(CD)
        watt = f"{str(watt)}W"

    x = abs_x + 80
    y = abs_y + 1000 + 90
    draw_Text(doc, x, y , 12, str(Ledcompany), layer='0')     
    # (아답터용량 : 60W )
    textstr = f"(아답터용량 : SMPS 슬림 2-{watt} )"
    x = abs_x + 80
    y = abs_y + 1000 + 90 - 20
    draw_Text(doc, x, y , 12, str(textstr), layer='0')    
    textstr = f"*. 출도일자 : {drawdate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 
    draw_Text(doc, x, y , 8, str(textstr), layer='0')    
    textstr = f"*. 납품일자 : {deadlinedate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 - 15
    draw_Text(doc, x, y , 8, str(textstr), layer='0')

    ################################################################################################################################################
    # 갑지 2
    ################################################################################################################################################
    # led바에 대한 내용 기술
    # LED BAR :

    textstr = f"{Leditem} - {LedBarLength}L = {su*4}(EA)"
    x = abs_x + 720 + 10
    y = abs_y + 650 + 32
    draw_Text(doc, x, y , 14, str(textstr), layer='0')        

    textstr = f"SMPS({watt})={su*2}(EA)"
    x = abs_x + 720 + 10
    y = abs_y + 650 - 35 + 32
    draw_Text(doc, x, y , 14, str(textstr), layer='0')    

    textstr = f"AL PROFILE TUBE - {LCD-3} = {su*2}(EA)" 
    x = abs_x + 720 + 10
    y = abs_y + 650 - 70 + 32
    draw_Text(doc, x, y , 13 , str(textstr), layer='0')    

    textstr = f"SMPS"        
    draw_Text(doc, 120, 480+45 , 20, str(textstr), layer='0')            
    draw_Text(doc, 845, 480+45 , 20, str(textstr), layer='0')    

    textstr = f"({watt})"        
    draw_Text(doc, 120, 480 , 20, str(textstr), layer='0')            
    draw_Text(doc, 845, 480 , 20, str(textstr), layer='0')    

    # Type
    textstr = f"{lctype}타입"
    x = abs_x + 711 - 100
    y = abs_y + 690
    draw_Text(doc, x, y , 20, str(textstr), layer='0')    
    # 현장명
    textstr = f"*.현장명 : {secondord}-{workplace} - {su}(set)"
    x = abs_x + 30
    y = abs_y + 650 
    draw_Text(doc, x, y , 17, str(textstr), layer='0')    
    # (등기구업체 : )    
    x = abs_x + 30 + 200
    y = abs_y + 600 
    draw_Text(doc, x, y , 20, str(Ledcompany), layer='0')    
    
    textstr1 = f"{Leditem}-(누드&클립&잭타입)"
    textstr2 = f"{Ledwatt}-{LedBarLength}L"

    x = abs_x + 235 - 107
    y = abs_y + 90    
    draw_Text_direction(doc, x, y , 14, str(textstr1), layer='0', rotation = 90)    
    draw_Text_direction(doc, x + 30, y+80 , 14, str(textstr2), layer='0', rotation = 90)    

    draw_Text_direction(doc, x+200, y , 14, str(textstr1), layer='0', rotation = 90)    
    draw_Text_direction(doc, x+200 + 30, y+80 , 14, str(textstr2), layer='0', rotation = 90)    

    draw_Text_direction(doc, x+520, y , 14, str(textstr1), layer='0', rotation = 90)    
    draw_Text_direction(doc, x+520 + 30, y+80 , 14, str(textstr2), layer='0', rotation = 90)    

    draw_Text_direction(doc, x+710, y , 14, str(textstr1), layer='0', rotation = 90)    
    draw_Text_direction(doc, x+710 + 30, y+80 , 14, str(textstr2), layer='0', rotation = 90)    
   
    textstr = f"AL PROFILE TUBE - {LCD-3}"    
    x = abs_x + 235 - 107 + 86
    y = abs_y + 90      
    draw_Text_direction(doc, x, y , 14, str(textstr), layer='0', rotation = 90)        
    draw_Text_direction(doc, x + 550, y , 14, str(textstr), layer='0', rotation = 90)    

    # car inside box 표기
    xx =  410
    yy =  45
    rectangle(doc, xx,yy,xx+200,yy+100,layer='0')      
    line(doc, xx,yy+50,xx+200,yy+50)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50, yy+50 , 12, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30, yy+25 , 12, str(textstr), layer='0')
  
    # Assy Car case    
    #####################################################################################################################################
    #1page
    ##################################################################################################################################### 
    # LC ASSY 상부 형상
    frameXpos = abs_x + 1700
    abs_x = frameXpos + 1700    
    abs_y = abs_y + CD + 400
    insert_block(abs_x , abs_y , "lc032_top_left")    

    x = abs_x + CW 
    y = abs_y 
    insert_block(x , y , "lc032_top_right")  

    first_space = 320
    
    # 폴리갈 or 중판 크기 산출
    polygal_width = CW - first_space * 2
    polygal_height = LCD - 6

    # print("polygal_height: " + str(polygal_height))  

    x1 = abs_x + 320
    y1 = abs_y + 100
    x2 = x1 + polygal_width
    y2 = y1
    line(doc, x1, y1, x1, y1+14, layer='4') # 적색    
    lineto(doc, x1+1.2, y1+14, layer='4')   
    lineto(doc, x1+1.2, y1+ 1.2, layer='4')   
    lineto(doc, x2-1.2, y1+ 1.2, layer='4')   
    lineto(doc, x2-1.2, y1+14 , layer='4')   
    lineto(doc, x2, y1+14 , layer='4')   
    lineto(doc, x2, y1 , layer='4')   
    lineto(doc, x1, y1 , layer='4')   
    textstr =  f"중판 폭 = {math.ceil(polygal_width)}"
    dim(doc, x1, y2+14, x2, y2+14, 172, text_height=0.32,  direction="up" , text=textstr)    

    # 녹색 실물 그리기 12t
    rectangle(doc, x1 + 20 , y1 + 1.2 , x2 - 20 , y1 + 13.2 , layer='1')  
    line(doc, x1 + 20 , y1 + 2.4 , x2 - 20 , y1 + 2.4 , layer='1') # 실선
    line(doc, x1 + 20 , y1 + 12 , x2 - 20 , y1 + 12 , layer='hidden1') # 히든선 녹색

    # 상부선 2.3mm 간격
    X1 = abs_x - 25
    Y1 = abs_y + 150
    X2 = X1 + CW + 50
    Y2 = Y1
    line(doc, X1, Y1, X2, Y2, layer='2')     # 회색
    lineto(doc, X2, Y1+2.3, layer='2')    
    lineto(doc, X1, Y1+2.3, layer='2')    
    lineto(doc, X1, Y1, layer='2')     

    # Assy 중심 치수
    x1 = abs_x + 345
    y1 = abs_y + 50
    x2 = abs_x + CW - 345
    y2 = y1    
    dim(doc,  x2, y2,  x1, y1,153.5 + 2.50 , text_height=0.30, text_gap=0.07, direction="down")
    x1 = abs_x + 25
    y1 = abs_y + 36
    x2 = abs_x + CW - 25
    y2 = y1        
    textstr =  f"Light Case Wide (W) = {CW-50}"    
    dim(doc, x1, y1, x2, y2, 260,  text_height=0.30, direction="down" , text=textstr)    
    # 25 간격
    dim(doc, x1, y1, abs_x, abs_y, 260, text_height=0.30, text_gap=0.07, direction="down")
    dim(doc, x2, y1, x2+25, abs_y, 260, text_height=0.30, text_gap=0.07, direction="down")
    
    # Assy 상단 하부 치수
    x1 = abs_x 
    y1 = abs_y 
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"Car Inside (CW) = {CW}"
    dim(doc, x1, y1, x2, y2, 310, text_height=0.30, direction="down" , text=textstr)    

    ###################################
    # 1page Assy 본판 작도
    ###################################
    # 기본 윤곽 car inside
    insideHeight = 25
    abs_y = abs_y - CD - 800    
    x1 = abs_x 
    y1 = abs_y 
    x2 = abs_x + CW
    y2 = abs_y + CD    
    rectangle(doc, x1, y1, x2, y2, layer='0')    
    x1 = abs_x + insideHeight
    y1 = abs_y + insideHeight
    x2 = abs_x + CW - insideHeight
    y2 = abs_y + CD - insideHeight
    rectangle(doc, x1, y1, x2, y2, layer='0')    

    line(doc, x1+1.2, y1, x1+1.2, y2, layer='2')    
    line(doc, x2-1.2, y1, x2-1.2, y2, layer='2')    
    line(doc, x1+1.2, y1+1.2, x2-1.2, y1+1.2, layer='2')    
    line(doc, x1+1.2, y2-1.2, x2-1.2, y2-1.2, layer='2')      
    
     # 4각 모서리의 사각모형 30x220 형상
    rectangle(doc, x1+100, y1, x1+100+220, y1+30, layer='0')  
    rectangle(doc, x1+100, y2-30, x1+100+220, y2, layer='0')  
    rectangle(doc, x2-100-220, y1, x2-100, y1+30, layer='0')  
    rectangle(doc, x2-100-220, y2-30, x2-100, y2, layer='0')  

    # LC 홀 가공
    arrX = [ 145, 145 + 150, CW - (145 + 150), CW - 145,  145, 145 + 150, CW - (145 + 150), CW - 145 ]

    for i, x in enumerate(arrX):  # i는 인덱스, y는 실제 배열 꼭 기억!        
        x = abs_x + x
        if (i > 3 ):
            # i가 짝수일 때 실행할 코드
            y = abs_y + 40
        else:
            # i가 홀수일 때 실행할 코드        
            y = abs_y + CD - 40

        circle_cross(doc, x , y , 11 , layer='0')     

    # 상부 절곡
    rectangle(doc, abs_x + 99.2 , y1 , abs_x + 99.2 + 25.8, y2 , layer='0')    
    rectangle(doc, abs_x + CW - 99.2 - 25.8 , y1 , abs_x + CW - 99.2 , y2, layer='0')    

    # 중간에 단공 2개
    # circle_cross(doc, abs_x + 112.5 , abs_y + CD/2 , 11 , layer='2', color='2')       
    # circle_cross(doc, abs_x + CW - 112.5 , abs_y + CD/2 , 11 , layer='2', color='2')       

    # 내부 중판 자리 하늘색    
    x1 = abs_x + 320
    y1 = abs_y + 28
    x2 = abs_x + CW - 320
    y2 = abs_y + CD/2
    rectangle(doc, x1, y1, x2, y2, layer='4')   # 하늘색 4번
    x1 = abs_x + 320
    y1 = y2
    x2 = abs_x + CW - 320
    y2 = abs_y + CD - 28    
    rectangle(doc, x1, y1, x2, y2, layer='4')   # 하늘색 4번

    # 상부 프레임 표현 녹색선 (좌측)
    x1 = abs_x + 285
    y1 = abs_y + 26.5
    x2 = x1 + 58.8
    y2 = abs_y + CD - 26.5    
    rectangle(doc, x1, y1, x2, y2, layer='1')   # 녹색 1번    
    line(doc, x1+10, y1, x1+10, y2, layer='1') 
    line(doc, x1+29, y1, x1+29, y2, layer='1') 
    line(doc, x1+30, y1, x1+30, y2, layer='1') 

    # 상부 프레임 표현 녹색선 (우측)
    x1 = abs_x + CW - 285
    y1 = abs_y + 26.5
    x2 = x1 - 58.8
    y2 = abs_y + CD - 26.5    
    rectangle(doc, x1, y1, x2, y2, layer='1')   # 녹색 1번    
    line(doc, x1-10, y1, x1-10, y2, layer='1') 
    line(doc, x1-29, y1, x1-29, y2, layer='1') 
    line(doc, x1-30, y1, x1-30, y2, layer='1') 

    # assy smps 2가지 가져오기
    # smps가 안들어가는 T5는 제외
    if T5_standard  < 1 :    
        x = abs_x + 93 
        y = abs_y + 130
        insert_block(x , y , "lc032_assy_smps1")      
        x = abs_x + CW - 169.2
        y = abs_y + 130
        insert_block(x , y , "lc032_assy_smps2")      

    # assy bracket 4개 가져오기 위치는 공식으로 계산
    # 중판을 계산해서 그 크기를 내경기준 + 53해서 간격을 띄운다 하단기준 
    resize_midplate_width = polygal_height/2 + 53 + 50 
    first_point = 400

    x1 = abs_x + 295 
    y1 = abs_y + first_point
    insert_block(x1 , y1 , "lc032_assy_bracket_left")          
    y2 = abs_y + first_point + resize_midplate_width
    insert_block(x1 , y2 , "lc032_assy_bracket_left")      
    x2 = abs_x + CW - 295     
    insert_block(x2 , y1 , "lc032_assy_bracket_right")              
    insert_block(x2 , y2 , "lc032_assy_bracket_right")     

    dim(doc, x1, y1+50, x1, y2, 80,  text_height=0.30, text_gap=0.07, direction="left")    

    if CD-100 == LedBarLength:
        LedSpace = 23
    else:
        LedSpace = 23 - ( LedBarLength-(CD-100) ) / 2

    # AL Tube  외곽 30 width 선
    x1 = abs_x + 25
    y1 = abs_y + 25 + 2
    x2 = x1 + 30 
    y2 = abs_y + CD - 27
    rectangle(doc, x1, y1, x2, y2, layer='CL')   # 적색            

    if T5_standard  < 1 :    
        rectangle(doc, x1+9, y1+LedSpace, x2-9, y2-LedSpace, layer='6')   # 선홍색 
        rectangle(doc, x1+12, y1+LedSpace+2, x2-12, y2-LedSpace-2, layer='4')   # 하늘색               
        # rectangle(doc, x1+13.5, y1+30, x2-13.5, y2-30, layer='2')   # 회색               

        # 직사각형의 크기와 간격
        rectangle_height = 6    
        spacing_y = 7

        # 직사각형 그리는 범위
        inner_x1 = x1 + 13.5
        inner_y1 = y1 + LedSpace +2
        inner_x2 = x2 - 13.5
        inner_y2 = y2 - LedSpace -2

        # 가로 방향 직사각형의 위치
        rect_x1 = inner_x1
        rect_x2 = inner_x2

        # 세로 방향으로 직사각형 그리기 (노란색 LED형상)
        current_y = inner_y1
        while current_y + rectangle_height <= inner_y2:
            rect_y1 = current_y
            rect_y2 = rect_y1 + rectangle_height

            # 직사각형 그리기
            rectangle(doc, rect_x1, rect_y1, rect_x2, rect_y2, layer='3')

            # 다음 직사각형의 시작 y 좌표 업데이트
            current_y += rectangle_height + spacing_y

    # 오른쪽 AL Tube  외곽 30 width 선
    x1 = abs_x + CW - 25 - 30
    y1 = abs_y + 27
    x2 = x1 + 30 
    y2 = abs_y + CD - 27        

    rectangle(doc, x1, y1, x2, y2, layer='CL')   # 적색            
    if T5_standard  < 1 :           
        rectangle(doc, x1+9, y1+LedSpace, x2-9, y2-LedSpace, layer='6')   # 선홍색 
        rectangle(doc, x1+12, y1+LedSpace+2, x2-12, y2-LedSpace-2, layer='4')   # 하늘색               

        # 직사각형의 크기와 간격
        rectangle_height = 6    
        spacing_y = 7

        # 직사각형 그리는 범위
        inner_x1 = x1 + 13.5
        inner_y1 = y1 + LedSpace +2
        inner_x2 = x2 - 13.5
        inner_y2 = y2 - LedSpace -2

        # 가로 방향 직사각형의 위치
        rect_x1 = inner_x1
        rect_x2 = inner_x2

        # 세로 방향으로 직사각형 그리기 (노란색 LED형상)
        current_y = inner_y1
        while current_y + rectangle_height <= inner_y2:
            rect_y1 = current_y
            rect_y2 = rect_y1 + rectangle_height

            # 직사각형 그리기
            rectangle(doc, rect_x1, rect_y1, rect_x2, rect_y2, layer='3')

            # 다음 직사각형의 시작 y 좌표 업데이트
            current_y += rectangle_height + spacing_y

    # 상부 치수선
    # Assy 본천장 및 LC 상단 
    x1 = abs_x + 320
    y1 = abs_y + CD - 28
    x2 = x1 + (CW - 320 * 2)
    y2 = y1 
    textstr =  f" 중판({LCplateMaterial})  = {math.ceil(polygal_width)}"
    dim(doc,  x2, y2, x1, y2, 120, text_height=0.30, direction="down" , text=textstr)    
    
    x1 = abs_x + 320 + (CW - 320 * 2)
    y1 = abs_y + CD - 28
    x2 = x2
    y2 = y1 - (CD - 28*2)/2 
    textstr =  f" 중판({LCplateMaterial})  = {math.ceil(polygal_height/2)}"
    dim(doc,  x1, y1, x2, y2,    100, text_height=0.23, direction="left" , text=textstr)    
    y1 = y2
    y2 = abs_y + 28    
    dim(doc,  x1, y1, x2, y2,   100, text_height=0.23,   direction="left" , text=textstr)    

    x1 = abs_x 
    y1 = abs_y + CD
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"L/C SIZE =  {CW-50}"
    dim(doc, x1 + 25 , y1 - 25, x2 - 25 , y2 - 25, 210, text_height=0.32, direction="up" , text=textstr)    

    x1 = abs_x 
    y1 = abs_y + CD
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"CAR INSIDE(CW) - {CW}"
    dim(doc, x1 , y1 , x2  , y2, 290, text_height=0.32, direction="up" , text=textstr)    

    # 25 치수 표현
    dim(doc,  x1+25, y2,  x1, y1, 80,  text_height=0.30, text_gap=0.07, direction="up")
    dim(doc, x1+25, y2, x1+25+120, y2, 80,  text_height=0.30, text_gap=0.07, direction="up")
    dim(doc, x1+25+120, y2, x1+25+120+150, y2, 80,  text_height=0.30, text_gap=0.07, direction="up")
    dim(doc, x1+25+120+150, y2, x2-25-120-150 , y2, 80,  text_height=0.30, text_gap=0.07, direction="up")
    dim(doc, x2-25-120-150, y2, x2-25-120, y2, 80,  text_height=0.30, text_gap=0.07, direction="up")
    dim(doc, x2-25-120, y2, x2-25, y2, 80,  text_height=0.30, text_gap=0.07, direction="up")
    dim(doc, x2-25, y1, x2, y2, 80,  text_height=0.30, text_gap=0.07, direction="up")

    # 좌측 치수선
    if Leditem == 'T5':
        textstr =  f"{Leditem} (안쪽 배치 녹색선) -{LedBarLength}L"    
        calculate_gap = (CD - 50 - LedBarLength)/2
        x1 = abs_x + 285
        y1 = abs_y + CD - calculate_gap
        x2 = x1
        y2 = abs_y + calculate_gap        
        dim_string(doc,  x2, y2,  x1, y1,  300 + 220, textstr , text_height=0.35,  direction="left")     
    else:
        textstr =  f"{Leditem}-{LedBarLength}L"     
        x1 = abs_x + 46.91
        y1 = abs_y + CD - 50
        x2 = x1
        y2 = abs_y + 50
        dim_string(doc,  x2, y2,  x1, y1,  300, textstr , text_height=0.35,  direction="left")     
    
    x1 = abs_x + 46.91
    y1 = abs_y + CD - 26.5
    x2 = x1
    y2 = abs_y + 26.5
    textstr =  f"내측 AL TUBE 길이 = {str(math.floor(CD-26.5*2))}L"    
    dim(doc,  x1, y1,  x2, y2,   400, text_height=0.32,  direction="left" , text=textstr)    
  
    # 우측 치수선  
    x1 = abs_x + CW - 25
    y1 = abs_y + CD - 25
    x2 = x1 - 120
    y2 = y1 - 15
    dim(doc, x1, y1, x2, y2, 220 - 95, text_height=0.30,  direction="right")        

    x1 = abs_x + CW - 145
    y1 = abs_y + CD - 40
    x2 = x1
    y2 = abs_y + 40    
    dim(doc, x1, y1, x2, y2, 330, text_height=0.30,  direction="right")

    x1 = abs_x + CW - 120 - 25
    y1 = abs_y + 40
    x2 = abs_x + CW - 25
    y2 = y1 - 15
    dim(doc, x1, y1, x2, y2, 220 + 25 , text_height=0.30,  direction="right")

    # 25 간격 치수선 표기
    x1 = abs_x + CW 
    y1 = abs_y + CD 
    x2 = x1 - 25
    y2 = y1 - 25
    dim(doc, x1, y1, x2, y2, 250-25, text_height=0.30,  direction="right")

    x1 = abs_x + CW -25
    y1 = abs_y + 25
    x2 = abs_x + CW
    y2 = abs_y
    dim(doc, x1, y1, x2, y2, 250, text_height=0.30,  direction="right")

    x1 = abs_x + CW - 25
    y1 = abs_y + CD - 25
    x2 = x1
    y2 = abs_y + 25        
    textstr =  f"LIGHT CASE SIZE = {CD-panel_thickness*2}"    
    dim(doc,   x2, y2, x1, y1,   320, text_height=0.35, direction="right" , text=textstr)    
    x1 = abs_x + CW
    y1 = abs_y + CD
    x2 = x1
    y2 = abs_y         
    textstr =  f"CAR INSIDE {CD}"    
    dim(doc,  x2, y2, x1, y1,  410, text_height=0.35, direction="right" , text=textstr)    

    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1715
    TargetXscale = 5000 + (CW - 1600)
    TargetYscale = 2950 + (CD - 1500)
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale        
    frameYpos = abs_y - 350 * frame_scale     
    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)       
    # 1page 상단에 car inside 표기
    x = frameXpos + CW/3
    y = abs_y + frame_scale*BasicYscale + 50
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, x, y , 200*frame_scale, str(textstr), layer='0')    

    frameXpos = frameXpos + TargetXscale + 1400

    #################################################################################################
    # 2page 프레임 전개도 실전 레이져 가공물 본체만들기
    #################################################################################################    
    # 2page LC frame 삽입(레이져 가공)
    rx1 = frameXpos  + 340 
    ry1 = frameYpos + 1960
    insert_block(rx1 , ry1 , "lc032_frame_section")    
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 3780 + (CD - 1500)  # 기본 모형이 1450
    TargetYscale = 2304 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
      
    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

    rx1 = math.ceil(frameXpos  + 1035)
    ry1 = ry1 - 230
    
    thickness = 1.2
    vcut = thickness / 2

    x1 = rx1 
    y1 = ry1
    x2 = x1 + LCD
    y2 = y1    
    x3 = x2
    y3 = y2 + 25.8 - vcut # 45도 적용
    x4 = x3
    y4 = y3 + 106.07 - thickness
    x5 = x4 
    y5 = y4 + 34.5 - vcut 
    x6 = x5 - 1.5
    y6 = y5 
    x7 = x6 
    y7 = y6 + 20 - vcut
    x8 = x7 
    y8 = y7 + 32 - thickness
    x9 = x8 
    y9 = y8 + 20.5 - vcut
    x10 = x9 + 1.5
    y10 = y9 
    x11 = x10 
    y11 = y10 + vcut
    x12 = x11 
    y12 = y11 + 30 - thickness
    x13 = x12 
    y13 = y12 + 15 - vcut
    x14 = x13 
    y14 = y13 + 40.81 - vcut

    x15 = x14 
    y15 = y14 + 174 - vcut
    x16 = x15 - 1.2
    y16 = y15 
    x17 = x16 
    y17 = y16 + 60 - vcut
    x18 = x17 - LCD + thickness*2
    y18 = y17 
    x19 = x18
    y19 = y18 - 60 + vcut
    x20 = x19 - 1.2
    y20 = y19 
    x21 = x20 
    y21 = y20  - 174 + vcut

    x22 = x21 
    y22 = y21 - 40.81 + vcut
    x23 = x22
    y23 = y22 - 15 + vcut
    x24 = x23 
    y24 = y23 - 30 + thickness
    x25 = x24 
    y25 = y24 - vcut
    x26 = x25 + 1.5
    y26 = y25 
    x27 = x26 
    y27 = y26 - 20.5 + vcut
    x28 = x27 
    y28 = y27 - 32 + thickness
    x29 = x28 
    y29 = y28 - 20 + vcut
    x30 = x29 - 1.5
    y30 = y29 
    x31 = x30
    y31 = y30 - 34.5 + vcut 
    x32 = x31 
    y32 = y31 - 106.07 + thickness

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 32
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
      
    #절곡라인        
    line(doc, x32, y32, x3, y3,  layer='hidden')    # 절곡선    
    line(doc, x31, y31, x4, y4,  layer='hidden')    # 절곡선    
    line(doc, x28, y28, x7, y7,  layer='hidden')    # 절곡선    
    line(doc, x24, y24, x11, y11,  layer='hidden')    # 절곡선    
    line(doc, x23, y23, x12, y12,  layer='hidden')    # 절곡선    
    line(doc, x19, y19, x16, y16,  layer='hidden')    # 절곡선    

    #반대 절곡라인
    line(doc, x27, y27, x8, y8,  layer='2')    # 절곡선  
    line(doc, x22, y22, x13, y13,  layer='2')    # 절곡선  
    line(doc, x21, y21, x14, y14,  layer='2')    # 절곡선  

    # 상단 좌우 끝위치 장공
    insert_block(x18 , y18  , "lc032_15x25slot_left")  # 치수선 포함된 것 표현
    insert_block(x18+ 20 - 1.2 , y18 - 34.4  , "lc032_15x25slot_left_laser", layer="레이져") # 중심점 이동
    insert_block(x17 , y17  , "lc032_15x25slot_right", layer="레이져")

    # 4각 라운드 2개소 위치
    secondLineY =  y28 + 16 -vcut 
    insert_block(x28+38.5 ,secondLineY , "lc032_100x200_round", layer="6")     # 중심선 위치 중요
    insert_block(x28+38.5 ,secondLineY , "lc032_100x200_round_laser", layer="레이져")
    insert_block(x7-38.5 ,secondLineY  , "lc032_100x200_round", layer="6")
    insert_block(x7-38.5 ,secondLineY  , "lc032_100x200_round_laser", layer="레이져")
    
    # 3.2파이 두번째 줄 단공 4개 -> 3.2x10 장공으로 수정 (2025/01/24)
    circle3dot2x1 =  x1 + 100
    circle3dot2y1 =  secondLineY   
    circle3dot2x2 =  x1 + 100 + 400
    circle3dot2y2 =  secondLineY   
    circle3dot2x3 =  x2 - 100 - 400
    circle3dot2y3 =  secondLineY  
    circle3dot2x4 =  x2 - 100
    circle3dot2y4 =  secondLineY  
    # circle_cross(doc, circle3dot2x1 , circle3dot2y1 , 3.2 , layer='레이져', color='0')     
    # circle_cross(doc, circle3dot2x2 , circle3dot2y2 , 3.2 , layer='레이져', color='0')     
    # circle_cross(doc, circle3dot2x3 , circle3dot2y3 , 3.2 , layer='레이져', color='0')     
    # circle_cross(doc, circle3dot2x4 , circle3dot2y4 , 3.2 , layer='레이져', color='0')     
    draw_slot(doc, circle3dot2x1 , circle3dot2y1 , "3.2x10", direction="세로", option='cross', layer='레이져')    
    draw_slot(doc, circle3dot2x2 , circle3dot2y2 , "3.2x10", direction="세로", option='cross', layer='레이져')    
    draw_slot(doc, circle3dot2x3 , circle3dot2y3 , "3.2x10", direction="세로", option='cross', layer='레이져')    
    draw_slot(doc, circle3dot2x4 , circle3dot2y4 , "3.2x10", direction="세로", option='cross', layer='레이져')        
    dim_leader_line(doc, circle3dot2x1 , circle3dot2y1, circle3dot2x1+30 , circle3dot2y1+50, "3.2x10",  layer='JKW', style='0', text_height=22)
    dim_leader_line(doc, circle3dot2x2 , circle3dot2y2, circle3dot2x2+30 , circle3dot2y2+50, "3.2x10",  layer='JKW', style='0', text_height=22)
    dim_leader_line(doc, circle3dot2x3 , circle3dot2y3, circle3dot2x3+30 , circle3dot2y3+50, "3.2x10",  layer='JKW', style='0', text_height=22)
    dim_leader_line(doc, circle3dot2x4 , circle3dot2y4, circle3dot2x4+30 , circle3dot2y4+50, "3.2x10",  layer='JKW', style='0', text_height=22)
    # dim_diameter(doc, (circle3dot2x1 , circle3dot2y1), 3.2, angle=225, dimstyle="JKW", override=None)
    # dim_diameter(doc, (circle3dot2x2 , circle3dot2y2), 3.2, angle=225, dimstyle="JKW", override=None)
    # dim_diameter(doc, (circle3dot2x3 , circle3dot2y3), 3.2, angle=225, dimstyle="JKW", override=None)
    # dim_diameter(doc, (circle3dot2x4 , circle3dot2y4), 3.2, angle=225, dimstyle="JKW", override=None)

    # 3.2파이 첫번째 줄 단공 2개 -> 3개 (상단)
    circle3dot2x5 =  x18 + 350
    circle3dot2y5 =  y18 - 29.7 - 7
    circle3dot2x6 =  x17 - 350
    circle3dot2y6 =  y18 - 29.7 - 7
    circle3dot2x7 =  (x17 + x18)/2 
    circle3dot2y7 =  y18 - 29.7 - 7    
    draw_slot(doc, circle3dot2x5 , circle3dot2y5 , "3.2x10", direction="세로", option='cross', layer='레이져')            
    draw_slot(doc, circle3dot2x6 , circle3dot2y6 , "3.2x10", direction="세로", option='cross', layer='레이져')            
    draw_slot(doc, circle3dot2x7 , circle3dot2y7 , "3.2x10", direction="세로", option='cross', layer='레이져')            
    dim_leader_line(doc, circle3dot2x5 , circle3dot2y5, circle3dot2x5 +30, circle3dot2y5+50, "3.2x10",  layer='JKW', style='0', text_height=22)    
    dim_leader_line(doc, circle3dot2x6 , circle3dot2y6, circle3dot2x6 +30, circle3dot2y6+50, "3.2x10",  layer='JKW', style='0', text_height=22)    
    dim_leader_line(doc, circle3dot2x7 , circle3dot2y7, circle3dot2x7 +30, circle3dot2y7+50, "3.2x10",  layer='JKW', style='0', text_height=22)    
    # dim_diameter(doc, (circle3dot2x5 , circle3dot2y5), 3.2, angle=225, dimstyle="JKW", override=None)  
    # dim_diameter(doc, (circle3dot2x6 , circle3dot2y6), 3.2, angle=225, dimstyle="JKW", override=None)  
    # dim_diameter(doc, (circle3dot2x7 , circle3dot2y7), 3.2, angle=225, dimstyle="JKW", override=None)  
    # 15파이 두번째 줄 단공 3개
    circle15x1 =  x1 + 225
    circle15y1 =  secondLineY   
    circle15x2 =  x1 + LCD/2
    circle15y2 =  secondLineY   
    circle15x3 =  x2 - 225
    circle15y3 =  secondLineY   
    circle_cross(doc, circle15x1 , circle15y1, 15 , layer='레이져', color='0')     
    circle_cross(doc, circle15x2 , circle15y2 , 15 , layer='레이져', color='0')     
    circle_cross(doc, circle15x3 , circle15y3 , 15 , layer='레이져', color='0')    
    dim_diameter(doc, (circle15x1 , circle15y1), 15,  angle=225, dimstyle="JKW", override=None) 
    dim_diameter(doc, (circle15x2 , circle15y2), 15,  angle=225, dimstyle="JKW", override=None) 
    dim_diameter(doc, (circle15x3 , circle15y3), 15,  angle=225, dimstyle="JKW", override=None) 
    
    textstr = f"클립홀이 같아야 함"    
    draw_Text(doc, x14 - 500 , y21 + 70 , 25, str(textstr), layer='0')    
  
    # 상부 치수선 (3개 단공 3.2파이연속석)
    dim(doc, x18, y18, circle3dot2x5 , circle3dot2y5 , 50 , text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, circle3dot2x5 , circle3dot2y5 , circle3dot2x7 , circle3dot2y7 , 50 + 29.7 + 7 , text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, circle3dot2x7 , circle3dot2y7 , circle3dot2x6 , circle3dot2y6 , 50 + 29.7 + 7, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, circle3dot2x6 , circle3dot2y6 , x17 , y17 , 50 + 29.7 + 7 , text_height=0.22, text_gap=0.07, direction="up")

    # 1.2 표시
    dim(doc,  x18,y20, x20, y18 , 100 + extract_abs(y20,y18) , text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, x18, y18, x17 , y17 , 100  , text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, x17, y17, x15 , y15 , 100 , text_height=0.22, text_gap=0.07, direction="up")

    dim(doc, x20, y20, x15 , y15 , 150 + extract_abs(y17,y15) , text_height=0.22, text_gap=0.07, direction="up")

    # 하부 치수선
    dim(doc, x1, y1, circle3dot2x1, circle3dot2y1, 46 , text_height=0.22, text_gap=0.07, direction="down")
    dimto(doc, circle3dot2x2, circle3dot2y2, 46 + extract_abs(y1,circle3dot2y1) )
    dimto(doc, circle3dot2x3, circle3dot2y3, 46 + extract_abs(y1,circle3dot2y1) )
    dimto(doc, circle3dot2x4, circle3dot2y4, 46 + extract_abs(y1,circle3dot2y1) )
    dimto(doc, x2, y2 , 46 + extract_abs(y1,circle3dot2y1) )

    dim(doc, x1, y1, circle15x1 , circle15y1 , 146  , text_height=0.22, text_gap=0.07, direction="down")
    dimto(doc,circle15x2 , circle15y2 , 146 + extract_abs(y1,circle3dot2y1) )
    dimto(doc,circle15x3 , circle15y3 , 146 + extract_abs(y1,circle3dot2y1) )
    dimto(doc, x2, y2 , 146 + extract_abs(y1,circle3dot2y1) )

    # 10, 60 좌우 치수선
    dim(doc, x1+10 , circle15y1 , x1, y1 ,  90 + extract_abs(y1,circle3dot2y1)  , text_height=0.22, text_gap=0.07, direction="down", option="reverse")
    dimto(doc, x1+10+60 , circle15y2 , 110 )    
    dim(doc,  x2-10-60 , circle15y2 , x2-10 , circle15y2, 110  , text_height=0.22, text_gap=0.07, direction="down")
    dimto(doc, x2 , y2 , 110 + extract_abs(y1, circle3dot2y1)  )    
    dim(doc, x1, y1, x2 , y2 , 200  , text_height=0.22, text_gap=0.07, direction="down")

    # 4각 라운드
    dim(doc, x1+70, secondLineY + 10, x1+70 , secondLineY - 10 , 80  , text_height=0.22, text_gap=0.07, direction="right")
    dim(doc, x1+70, secondLineY - 10, x1+70 , y1 , 120  , text_height=0.22, text_gap=0.07, direction="right")
    # 1.5 간격
    dim(doc, x25, y25, x26 , y26 , 100  , text_height=0.22, text_gap=0.07, direction="up")
    dim(doc, x9, y9, x10 , y10 , 100  , text_height=0.22, text_gap=0.07, direction="up")

    # 좌측 치수
    dim(doc, x20 , y20 , x18 , y18 ,   80,  direction="left", option="reverse")
    dimto(doc, x25 , y25 , 80 )    
    dimto(doc, x30 , y30 , 80 )    
    dimto(doc, x1, y1 , 132 )
        
    # 우측 치수
    dim(doc,  x2 , y2 , x17, y17 , 460  , text_height=0.22, direction="right")
    dim(doc, x15 , y15 , x17, y17 ,  380  , text_height=0.22, direction="right" , option="reverse")
    dimto(doc, x14 , y14 , 380 )    
    dimto(doc, x13 , y13 , 380 )    
    dimto(doc, x12 , y12 , 330 )    
    dimto(doc, x11 , y11 , 280 )    
    dimto(doc, x8 , y8 , 230 )    
    dimto(doc, x7 , y7 , 180 )    
    dimto(doc, x4 , y4 , 130 )    
    dimto(doc, x3 , y3 , 90 )    
    dimto(doc, x2 , y2 , 70 )    
   
    # main Frame description    
    x1 = frameXpos + 3100 + (CD-1500)
    y1 = frameYpos + 2105
    textstr = f"Part Name : Light Case Frame"    
    draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : {LCframeMaterial}"    
    draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
    textstr = f"Size : {LCD} x {str(math.ceil((y17-y2)*10)/10)}mm"    
    draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')    

    # LC 부속자재들 3가지 위치하기
    rx1 = frameXpos  + 782 
    ry1 = frameYpos + 1100
    insert_block(rx1 , ry1 , "lc032_frame_ribset_laser1", layer="레이져")        
    #  insert_block(rx1 , ry1 , "lc032_frame_ribset_laser2", layer="레이져")        # 제거요청 SMPS 
    insert_block(rx1 , ry1 , "lc032_frame_ribset_laser3", layer="레이져")        
    insert_block(rx1 , ry1 , "lc032_frame_ribset")   # SMPS 제거용으로 수정함       

    # 중판 이탈방지 B/K description    
    x1 = frameXpos + 540 
    y1 = frameYpos + 680
    textstr = f"Part Name : 중판 이탈방지 B/K"    
    draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : EGI 1.2T"    
    draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
    textstr = f"Size : 57.6 X 50 "    
    draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
    textstr = f"Quantity : {su*4} EA"    
    draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')   

    # SMPS 커버 description    
    # x1 = frameXpos + 1300 
    # y1 = frameYpos + 820
    # textstr = f"Part Name : SMPS 커버"    
    # draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
    # textstr = f"Mat.Spec : EGI 1.2T"    
    # draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
    # textstr = f"Size : 220 X 184.4  "    
    # draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
    # textstr = f"Quantity : {su*2} EA"    
    # draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')   

    #  LIGHT CASE 커버 Mat.  description    
    x1 = frameXpos + 2300 
    y1 = frameYpos + 720
    textstr = f"Part Name :  LIGHT CASE 커버Mat.  "    
    draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : {LCframeMaterial}"    
    draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
    textstr = f"Size : 320.8 X 178.8   "    
    draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
    textstr = f"Quantity : 본도동일 {su*2} EA"    
    draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')   
    textstr = f"               본도반대 {su*2} EA"    
    draw_Text(doc, x1 , y1 - 230 , 30, str(textstr), layer='0')   

    frameXpos = frameXpos + TargetXscale + 400

    #############################################################################################
    # 3page PROFILE LED BAR 등 3가지
    #############################################################################################
    abs_x = frameXpos 

    # 등기구 크기 산출
    light_width = LCD - 4
    
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2520 + (LCD - 1450) 
    TargetYscale = 1280 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale   

    frame_scale = math.ceil(frame_scale*10) / 10           
    insert_frame(  abs_x , frameYpos , frame_scale, "drawings_frame", workplace)           
    
    # 등기구 확산커버포함 3개의 rectangle
    rx1 = frameXpos + 240
    ry1 = frameYpos + 1500 *frame_scale 
    insideHeight = 30   
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + light_width
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')          
    insideHeight = 12   
    x3 = x1 + LedSpace
    y3 = y1 + 9 
    x4 = x2 - LedSpace
    y4 = y3 + insideHeight    
    rectangle(doc, x3, y3, x4, y4, layer='6')      
    insideHeight = 6   
    x5 = x3 + 2
    y5 = y3 + 3 
    x6 = x4 - 2
    y6 = y5 + insideHeight    
    rectangle(doc, x5, y5, x6, y6, layer='4')      

    verticalLinex1 = x1 + 225 - 2
    verticalLiney1 = y2 + 10
    verticalLinex2 = x1 + (LCD-4)/2
    verticalLiney2 = y2 + 10
    verticalLinex3 = x2 - 225 + 2
    verticalLiney3 = y2 + 10
    line(doc, verticalLinex1, verticalLiney1, verticalLinex1, verticalLiney1-60, layer="CL")
    line(doc, verticalLinex2, verticalLiney2, verticalLinex2, verticalLiney2-60, layer="CL")
    line(doc, verticalLinex3, verticalLiney3, verticalLinex3, verticalLiney3-60, layer="CL")

    # 하부 치수선
    if Leditem == 'T5':        
        calculate_gap = (CD - 50 - LedBarLength)/2 - 25 + 23  
    else:
        calculate_gap = 23

    dim(doc,  x1 + calculate_gap, y1, x1, y1,  70 , text_height=0.22, text_gap=0.07, direction="down")
    textstr =  f"{Leditem} - {Ledwatt} (누드&클립&잭타입) =  {LedBarLength}L"
    dim(doc, x1 + calculate_gap, y1+9,  x2-calculate_gap, y1+9, 79,   direction="down" , text=textstr)    
    dim(doc,  x2-calculate_gap, y1, x2, y1, 70 , text_height=0.22, text_gap=0.07, direction="down")

    # 상부 치수선
    dim(doc,  x1, y1,verticalLinex1, verticalLiney1, 50 , text_height=0.22, text_gap=0.07, direction="up")
    textstr =  f"{Leditem} 고정용 클립 센터거리 -  {math.ceil(LCD/2 - 225)}L"
    dim(doc, verticalLinex1, verticalLiney1, verticalLinex2, verticalLiney2, 65 - 30 , text_height=0.16,  direction="up" , text=textstr)    
    dim(doc, verticalLinex2, verticalLiney2, verticalLinex3, verticalLiney3, 65 - 30, text_height=0.16, direction="up" , text=textstr)    
    dim(doc,  verticalLinex3, verticalLiney3,x2, y2, 65 - 30 , text_height=0.22, text_gap=0.07, direction="up")
    textstr =  f"{light_width} - (AL TUBE 길이)"
    dim(doc, x1, y1, x2, y2, 70 + 56, direction="up"  , text=textstr)    

    # AL profile 단면도 삽입
    insert_block( x2 + 134  , y1 -9  , "lc032_light_section_alprofile")         

    # AL profile  description    
    x1 = x2 + 320
    y1 = y2 + 67
    textstr = f"Part Name :  PROFILE LED BAR  "    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"               (확산커버포함) "    
    draw_Text(doc, x1 , y1 - 35 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : AL "    
    draw_Text(doc, x1 , y1 - 70 , 20, str(textstr), layer='0')    
    textstr = f"Size : SHP 3020 -  {light_width}mm"    
    draw_Text(doc, x1 , y1 - 105 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 140 , 20, str(textstr), layer='0')   

    #########################################################
    # 외측 등기구 후레임  2번째 줄
    #########################################################    
    rx1 = frameXpos + 165
    ry1 =  frameYpos + 1000 * frame_scale
    insideHeight = 84   
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + light_width + 1
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')  
    line(doc, x1, y1+59, x2, y1+59,  layer='0')    # 0 선  

    # 단면도 삽입
    insert_block( x2 + 147  , y1 , "lc032_lightframe") 

    # 상부 치수선    
    dim(doc,  rx1, y2, x2, y2, 80, text_height=0.22, text_gap=0.07, direction="up")

    # light frame description
    x1 = x2 + 350
    y1 = y1 + 150
    textstr = f"Part Name : 내, 외측 등기구 후레임"    
    draw_Text(doc, x1 , y1 , 22 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : AL"    
    draw_Text(doc, x1 , y1 -40 , 22, str(textstr), layer='0')    
    textstr = f"Size : 84 x 58.8 x {light_width+1} "    
    draw_Text(doc, x1 , y1 - 80 , 22, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 22, str(textstr), layer='0')    

    #########################################################
    # 등커버 3번째 줄
    #########################################################    
    rx1 = frameXpos + 165
    ry1 = frameYpos +  700 * frame_scale 
    insideHeight = 60   
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + light_width + 1
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')  
    for i in range(1, 29 + 1):
        line(doc, x1, y1+i*2, x2, y1+i*2,  layer='6')
    
    # 단면도 삽입
    insert_block( x2 + 147  , y1 , "lc032_lightframe_cover") 

    # 상부 치수선    
    dim(doc,  rx1, y2, x2, y2, 80, text_height=0.22, text_gap=0.07, direction="up")

    # 등커버 description
    x1 = x2 + 350
    y1 = y1 + 100
    textstr = f"Part Name : 등커버"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : PC"    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : 60 x 59 x {light_width+1} "    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')    

    # car inside box 표기
    xx =  x2 + 100
    yy =  y1 - 400
    rectangle(doc, xx,yy,xx+200*2,yy+100*2,layer='0')      
    line(doc, xx,yy+50*2,xx+200*2,yy+50*2)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50*2, yy+50*2 , 24, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30*2, yy+25*2 , 24, str(textstr), layer='0')         

    frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 4page 중판 1 
    #################################################################################################        
         
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2730 + (CD - 1500)/2 + (CW-1600) 
    TargetYscale = 1665 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    

    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

    rx1 = math.ceil(frameXpos  + 710)
    ry1 = frameYpos + 630
    
    thickness = 1.2
    vcut = thickness / 2

    # 중판 전개도 크기 정의
    midplate_width = CW - 320 * 2
    midplate_su = 2

    if(CD>=2000):
        midplate_height = (CD - 28 * 2)/3
        midplate_su = 3
    else:
        midplate_height = (CD - 28 * 2)/2

    topwing1 = 13.8
    topwing2 = 15
    btwing1 = 13.8
    btwing2 = 15
    leftwing1 = 15
    leftwing2 = 16.7
    rightwing1 = 13.8
    rightwing2 = 15

    x1 = rx1 
    y1 = ry1
    x2 = x1 + midplate_width
    y2 = y1    
    x3 = x2
    y3 = y2 + 15 - thickness
    x4 = x3 
    y4 = y3 + 15.5 - vcut
    x5 = x4 
    y5 = y4 + 2.2 - vcut
    x6 = x5 + 27.6 -vcut*2
    y6 = y5 
    x7 = x6 
    y7 = y6 + midplate_height - 18.2
    x8 = x7 - 16.5 + vcut * 2
    y8 = y7 
    x9 = x8 
    y9 = y8 + 15 - vcut * 2
    x10 = x9 - 11.1
    y10 = y9 
    x11 = x10 
    y11 = y10 + 1.6
    x12 = x11 
    y12 = y11 + 15 - vcut * 2
    x13 = x12 
    y13 = y12 + 15 - vcut * 3
    x14 = x13 - midplate_width
    y14 = y13 
    x15 = x14 
    y15 = y14 - 15 + vcut * 3
    x16 = x15
    y16 = y15 - 15 + vcut * 2
    x17 = x16 
    y17 = y16 - 1.6
    x18 = x17 - 11.1
    y18 = y17 
    x19 = x18
    y19 = y18 - 15 + vcut * 2
    x20 = x19 - 16.5 + vcut * 2
    y20 = y19 
    x21 = x20 
    y21 = y20 - midplate_height + 18.2
    x22 = x21 + 27.6 - vcut * 2
    y22 = y21
    x23 = x22
    y23 = y22 - 2.2 + vcut
    x24 = x23 
    y24 = y23 - 15.5 + vcut

    # 세로 절곡점을 찾음
    x25 = x21 + 13.8 -vcut 
    y25 = y21 
    x26 = x23 + vcut
    y26 = y23 
    x27 = x4 - vcut 
    y27 = y4
    x28 = x6 - 13.8 + vcut 
    y28 = y6
    x29 = x7 - 13.8 + vcut 
    y29 = y7
    x30 = x11 - vcut
    y30 = y11 
    x31 = x16 + vcut
    y31 = y16
    x32 = x20  + 13.8 -vcut 
    y32 = y20

    # 전개도 사이즈 저장 #1 중판
    midplatesizex1 = x6 - x21
    midplatesizey1 = y13 - y2

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 24
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
      
    #절곡라인  밴딩방향      
    line(doc, x15, y15, x12, y12,  layer='hidden')   # 절곡선    
    line(doc, x16, y16, x11, y11,  layer='hidden')   # 절곡선    
    line(doc, x23, y23, x4, y4,  layer='hidden')     # 절곡선    

    # 세로절곡선 표현
    line(doc, x32, y32, x25, y25,  layer='hidden')    # 절곡선    
    line(doc, x31, y31, x26, y26,  layer='hidden')    # 절곡선    
    line(doc, x30, y30, x27, y27,  layer='hidden')    # 절곡선    
    line(doc, x29, y29, x28, y28,  layer='hidden')    # 절곡선        

    #반대 절곡라인 반대방향
    line(doc, x24, y24, x3, y3,  layer='2')    # 절곡선     
   
    # 상부 치수선
    dim(doc, x18, y18,  x20, y20 , 120 , direction="up")
    dim(doc, x18, y18,  x14, y14 , 80, direction="up")
    dim(doc,  x13, y13 , x14, y14 ,110 ,direction="up")
    dim(doc, x9, y9 , x13, y13 ,  50 ,direction="up")
    dim(doc, x9, y9 , x7, y7 , 120 ,direction="up")

    # 좌측 치수선
    dim(doc, x14, y14, x18, y18 , 80 , direction="left", option="reverse")
    dimto(doc,  x20, y20 , 125 )    
    dimto(doc,  x21, y21 , 140 )    
    dimto(doc,  x1, y1 ,140  )    

    # 하부 치수선
    dim(doc, x1, y1 ,x21, y21,  120 , direction="down", option="reverse")
    dimto(doc,  x2, y2 , 120 )  
    dimto(doc,  x6, y6 , 120 )  

    dim(doc, x25, y25 ,x21, y21,  250 , direction="down", option="reverse")
    dimto(doc,  x26, y26 , 220 )  
    dimto(doc,  x27, y27 , 220 )      
    dim(doc, x28, y28, x27, y27 ,  220 , direction="down") 
    dim(doc, x28, y28, x6, y6 ,  250 , direction="down") # 치수선 위치 반대로 나오게

    dim(doc, x21, y21, x6, y6 , 290 , direction="down")

    # 우측 치수선
    dim(doc, x12, y12 , x13, y13,  80 , direction="right", option="reverse")
    dimto(doc,  x11, y11 , 140 )  
    dimto(doc,  x4, y4 , 200 )  
    dimto(doc,  x3, y3 , 140 )  
    dimto(doc,  x2, y2 , 100 )  
    dim(doc, x13, y13, x2, y2 , 250 ,  direction="right")    

    # vcut 전개도에 4개소 표시
    dim_leader_line(doc, x31, y31 - extract_abs(y32,y26)/5 , x31 + 100 ,y31 - extract_abs(y32,y26)/5 - 100, "Vcut")    
    dim_leader_line(doc, x30, y30 - extract_abs(y30,y28)/5 , x30 - 100 ,y30 - extract_abs(y30,y28)/5 - 100, "Vcut")    
    dim_leader_line(doc, x26  + extract_abs(x26,x27)/1.3  , y26, x26  + extract_abs(x26,x27)/1.3   + 100 ,y26 + 100, "Vcut")    
    dim_leader_line(doc, x26  + extract_abs(x26,x27)/1.3  , y11, x26  + extract_abs(x26,x27)/1.3   + 100 ,y11 - 100, "Vcut")    

    # 좌측 단면도 그리기
    # 왼쪽 leftwing1 leftwing2
    # 오른쪽 날개 rightwing1 rightwing2
    # 본판 mainplate = mp

    leftwing1 = 15
    leftwing2 = 16.7
    mp = midplate_height
    rightwing1 = 13.8
    rightwing2 = 15

    startx = x1 - 503
    starty = y1 - leftwing1 + btwing1 + btwing2

    x1 = startx
    y1 = starty
    x2 = x1 + thickness    
    y2 = y1    
    x3 = x2  
    y3 = y2 + leftwing1 - thickness
    x4 = x3 + leftwing2 - thickness
    y4 = y3 
    x5 = x4 
    y5 = y4 + midplate_height
    x6 = x5 - rightwing2
    y6 = y5 
    x7 = x6 
    y7 = y6 - rightwing1
    x8 = x7 + thickness
    y8 = y7 
    x9 = x8 
    y9 = y8 + rightwing1 - thickness
    x10 = x9 + rightwing2 - thickness *2
    y10 = y9 
    x11 = x10 
    y11 = y10 - midplate_height + thickness *2
    x12 = x11 - leftwing2 + thickness
    y12 = y11

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 12
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="0")     

    #vcut 원 그리기
    draw_circle(doc, x4, y4, 13, layer='0', color='6') 
    draw_circle(doc, x5, y5, 13, layer='0', color='6') 
    draw_circle(doc, x6, y6, 13, layer='0', color='6') 

    # 치수선
    dim(doc, x1,y1, x4, y4, 110, direction="down")
    dim(doc, x12,y12, x1, y1, 80, direction="left")
    dim(doc, x4,y4, x5, y5, 110, direction="right")
    dim(doc, x6,y6, x5, y5, 100, direction="up")
    dim(doc, x6,y6, x7, y7, 80, direction="left")

    dim_leader_line(doc, x4, y4 , x4 + 130 , y4-100, "Vcut 3개소")      
   
    ###############################################################################
    # 상부 단면도 그리기
    ###############################################################################
    leftwing1 = 13.8
    leftwing2 = 15
    mp = midplate_width
    rightwing1 = 13.8
    rightwing2 = 15

    startx = rx1 + leftwing1
    starty = ry1 + 340 + midplate_height + leftwing2*3

    x1 = startx
    y1 = starty
    x2 = x1 - leftwing1
    y2 = y1    
    x3 = x2  
    y3 = y2 - leftwing2
    x4 = x3 + mp
    y4 = y3 
    x5 = x4 
    y5 = y4 + rightwing2
    x6 = x5 - rightwing1
    y6 = y5 
    x7 = x6 
    y7 = y6 - thickness
    x8 = x7 + rightwing1 - thickness
    y8 = y7 
    x9 = x8 
    y9 = y8 - rightwing2 + thickness*2
    x10 = x9 - mp + thickness *2
    y10 = y9 
    x11 = x10 
    y11 = y10 + leftwing2 - thickness *2
    x12 = x11 + leftwing1 - thickness
    y12 = y11

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 12
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="0")     

    #vcut 원 그리기 4개소
    draw_circle(doc, x2, y2, 13, layer='0', color='6') 
    draw_circle(doc, x3, y3, 13, layer='0', color='6') 
    draw_circle(doc, x4, y4, 13, layer='0', color='6') 
    draw_circle(doc, x5, y5, 13, layer='0', color='6') 

    # 치수선
    dim(doc, x2,y2, x1, y1, 80, direction="up")
    dim(doc, x2,y2, x3, y3, 80, direction="left")
    dim(doc, x5,y5, x4, y4, 80, direction="right")
    dim(doc, x6,y6, x5, y5, 80, direction="up")
    dim(doc, x3,y3, x4, y4, 80, direction="down")
   
    dim_leader_line(doc, x4, y4 , x4 + 50 , y4 - 120, "Vcut 4개소")     

    # 중판 #1 description    
    x1 = frameXpos + 1980 + (midplate_width-960)
    y1 = frameYpos + 1125 +  (midplate_height-722)
    textstr = f"Part Name : 중판#1"    
    draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : {LCplateMaterial}"    
    draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
    textstr = f"Size : {str(math.ceil(midplatesizex1*100)/100)} x {str(math.ceil(midplatesizey1*100)/100)}mm"    
    draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
    textstr = f"Quantity : {su} EA"    
    draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')        

    frameXpos = frameXpos + TargetXscale + 400


    #################################################################################################
    # 5page 중판 #2  중판1보다 4포인트 전개도 추가됨
    #################################################################################################        
         
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2730 + (CD - 1500)/2  + (CW-1600) 
    TargetYscale = 1665 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
       
    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

    rx1 = frameXpos  + 710
    ry1 = frameYpos + 630
    
    thickness = 1.2
    vcut = thickness / 2

    # 중판 전개도 크기 정의
    midplate_width = CW - 320 * 2
    midplate_su = 2

    if(CD>=2000):
        midplate_height = (CD - 28 * 2)/3
        midplate_su = 3
    else:
        midplate_height = (CD - 28 * 2)/2

    topwing1 = 13.8
    topwing2 = 15
    btwing1 = 13.8
    btwing2 = 15
    leftwing1 = 13.8
    leftwing2 = 15
    rightwing1 = 13.8
    rightwing2 = 15

    x1 = rx1 
    y1 = ry1
    x2 = x1 + midplate_width
    y2 = y1    
    x3 = x2
    y3 = y2 + btwing1 - vcut 
    x4 = x3 
    y4 = y3 + btwing2 - vcut * 2
    x5 = x4 
    y5 = y4 + 2.2 - vcut
    x6 = x5 + btwing2 - (thickness*3 + vcut/2) 
    y6 = y5 
    x7 = x6 
    y7 = y6 + btwing1 
    x8 = x7 + rightwing1 + (vcut*3 - 0.3)
    y8 = y7 
    x9 = x8 
    y9 = y8 + midplate_height - (topwing1 + btwing1 + thickness*3 + vcut/2 + 0.5)
    x10 = x9 - rightwing1 - (vcut*3- 0.3)
    y10 = y9 
    x11 = x10 
    y11 = y10 + topwing1 
    x12 = x11 - topwing2 + (thickness*3 + vcut/2) 
    y12 = y11
    x13 = x12 
    y13 = y12 + 2.2 - vcut
    x14 = x13 
    y14 = y13 + topwing2  - vcut * 2
    x15 = x14 
    y15 = y14 + topwing1 - vcut
    x16 = x15 - midplate_width
    y16 = y15 
    x17 = x16 
    y17 = y16 - topwing1  + vcut
    x18 = x17
    y18 = y17 - topwing2 + vcut * 2
    x19 = x18
    y19 = y18 - 2.2 + vcut
    x20 = x19 - topwing2 + (thickness*3 + vcut/2) 
    y20 = y19 
    x21 = x20 
    y21 = y20 - topwing1 
    x22 = x21 - leftwing1 - (vcut*3 - 0.3)
    y22 = y21
    x23 = x22
    y23 = y22 - midplate_height + (topwing1 + btwing1 + thickness*3 + vcut/2 + 0.5)
    x24 = x23 + leftwing1 + (vcut*3 - 0.3)
    y24 = y23
    x25 = x24 
    y25 = y24 - btwing1
    x26 = x25 + btwing2 - (thickness*3 + vcut/2) 
    y26 = y25 
    x27 = x26 
    y27 = y26 - 2.2 + vcut
    x28 = x27 
    y28 = y27 - btwing2 + vcut * 2

    # 세로 절곡점 찾기
    x29 = x23 + leftwing1 - vcut
    y29 = y23
    x30 = x27 + vcut
    y30 = y27 
    x31 = x4 - vcut
    y31 = y4
    x32 = x8 - rightwing1 + vcut
    y32 = y8

    x33 = x9 - rightwing1 + vcut
    y33 = y9
    x34 = x13 - vcut
    y34 = y13
    x35 = x18 + vcut
    y35 = y18
    x36 = x22 + leftwing1 - vcut
    y36 = y22

    # 전개도 사이즈 저장 #1 중판
    midplatesizex2 = x8 - x23 
    midplatesizey2 = y15 - y2

    # 중판1보다 4점 추가됨 주의
    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 28
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
      
    #절곡라인  밴딩 가로 방향      
    line(doc, x28, y28, x3, y3,  layer='hidden')   # 절곡선      
    line(doc, x27, y27, x4, y4,  layer='hidden')   # 절곡선      
    line(doc, x18, y18, x13, y13,  layer='hidden')   # 절곡선          
    line(doc, x17, y17, x14, y14,  layer='hidden')   # 절곡선      

    # 세로절곡선 표현
    line(doc, x36, y36, x29, y29,  layer='hidden')    # 절곡선        
    line(doc, x35, y35, x30, y30,  layer='hidden')    # 절곡선        
    line(doc, x34, y34, x31, y31,  layer='hidden')    # 절곡선        
    line(doc, x33, y33, x32, y32,  layer='hidden')    # 절곡선        

    #반대 절곡라인 반대방향
    # line(doc, x24, y24, x3, y3,  layer='2')    # 절곡선     
   
    # 상부 치수선
    dim(doc, x20, y20, x22, y22 , 120 , direction="up", option="reverse" )
    dimto(doc,  x16, y16 , 85 )
    dimto(doc,  x15, y15 , 140 )    
    dim(doc,  x11, y11 , x15, y15 , 70 , direction="up" )
    dim(doc,  x11, y11 , x9, y9 , 100 , direction="up" )

    # 좌측 치수선
    dim(doc, x20, y20 , x16, y16,  60 , direction="left", option="reverse" )    
    dimto(doc,  x22, y22 , 110 )    
    dimto(doc,  x23, y23 , 160 )    
    dimto(doc,  x25, y25 , 130 )    
    dimto(doc,  x1, y1 , 60 )    

    # 하부 치수선
    dim(doc, x23, y23, x29, y29 , 230 , direction="down")
    dimto(doc,  x30, y30 , 150 )  
    dimto(doc,  x31, y31 , 180 )  
    dimto(doc,  x32, y32 , 150 )      
    dimto(doc,  x8, y8 , 230 )  
    dim(doc, x23, y23, x8, y8 , 280 , direction="down")

    # 우측 치수선
    dim(doc, x15, y15, x14, y14 , 80 , direction="right")
    dimto(doc,  x13, y13 , 120 )      
    dimto(doc,  x4, y4 , 190 )  
    dimto(doc,  x3, y3 , 150 )  
    dimto(doc,  x2, y2 , 110 )  
    dim(doc, x15, y15, x2, y2 , 250 , direction="right")

    # vcut 전개도에 4개소 표시
    dim_leader_line(doc, x35, y35 - extract_abs(y35,y30)/5 , x35 + 100 ,y35 - extract_abs(y35,y30)/5 - 100, "Vcut")    
    dim_leader_line(doc, x13, y13 - extract_abs(y18,y13)/3 , x13 - 100 ,y13 - extract_abs(y18,y13)/3 - 100, "Vcut")        
    dim_leader_line(doc, x30  + extract_abs(x30,x31)/1.3  , y30, x30  + extract_abs(x30,x31)/1.3   + 100 , y28 + 100, "Vcut")    
    dim_leader_line(doc, x28  + extract_abs(x28,x29)/1.3  , y13, x28  + extract_abs(x28,x29)/1.3   + 100 , y13 - 100, "Vcut")    

    # 좌측 단면도 그리기
    # 왼쪽 leftwing1 leftwing2
    # 오른쪽 날개 rightwing1 rightwing2
    # 본판 mainplate = mp 
    # 'ㄷ'자 형상 

    leftwing1 = 13.8
    leftwing2 = 15
    mp = midplate_height
    rightwing1 = 13.8
    rightwing2 = 15

    startx = x1 - 503 - leftwing2
    starty = y1 + leftwing1 + btwing1 + btwing2

    x1 = startx
    y1 = starty
    x2 = x1 - thickness    
    y2 = y1    
    x3 = x2  
    y3 = y2 - leftwing1 
    x4 = x3 + leftwing2 
    y4 = y3 
    x5 = x4 
    y5 = y4 + midplate_height
    x6 = x5 - rightwing2
    y6 = y5 
    x7 = x6 
    y7 = y6 - rightwing1
    x8 = x7 + thickness
    y8 = y7 
    x9 = x8 
    y9 = y8 + rightwing1 - thickness
    x10 = x9 + rightwing2 - thickness *2
    y10 = y9 
    x11 = x10 
    y11 = y10 - midplate_height + thickness *2
    x12 = x11 - leftwing2 + thickness*2
    y12 = y11


    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 12
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="0")     

    #vcut 원 그리기
    draw_circle(doc, x3, y3, 13, layer='0', color='6') 
    draw_circle(doc, x4, y4, 13, layer='0', color='6') 
    draw_circle(doc, x5, y5, 13, layer='0', color='6') 
    draw_circle(doc, x6, y6, 13, layer='0', color='6') 

    # 치수선
    dim(doc, x3,y3, x4, y4, 110, direction="down")
    dim(doc, x2,y2, x3, y3, 80, direction="left")
    dim(doc, x4,y4, x5, y5, 110, direction="right")
    dim(doc, x6,y6, x5, y5, 100, direction="up")
    dim(doc, x6,y6, x7, y7, 80, direction="left")

    dim_leader_line(doc, x4, y4 , x4 + 100 , y4-100, "Vcut 4개소")
   
    ###############################################################################
    # 상부 단면도 그리기
    ###############################################################################
    leftwing1 = 13.8
    leftwing2 = 15
    mp = midplate_width
    rightwing1 = 13.8
    rightwing2 = 15

    startx = rx1 + leftwing1
    starty = ry1 + 340 + midplate_height + leftwing2*3

    x1 = startx
    y1 = starty
    x2 = x1 - leftwing1
    y2 = y1    
    x3 = x2  
    y3 = y2 - leftwing2
    x4 = x3 + mp
    y4 = y3 
    x5 = x4 
    y5 = y4 + rightwing2
    x6 = x5 - rightwing1
    y6 = y5 
    x7 = x6 
    y7 = y6 - thickness
    x8 = x7 + rightwing1 - thickness
    y8 = y7 
    x9 = x8 
    y9 = y8 - rightwing2 + thickness*2
    x10 = x9 - mp + thickness *2
    y10 = y9 
    x11 = x10 
    y11 = y10 + leftwing2 - thickness *2
    x12 = x11 + leftwing1 - thickness
    y12 = y11

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 12
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="0")     

    #vcut 원 그리기 4개소
    draw_circle(doc, x2, y2, 13, layer='0', color='6') 
    draw_circle(doc, x3, y3, 13, layer='0', color='6') 
    draw_circle(doc, x4, y4, 13, layer='0', color='6') 
    draw_circle(doc, x5, y5, 13, layer='0', color='6') 

    # 치수선
    dim(doc, x2,y2, x1, y1, 80, direction="up")
    dim(doc, x2,y2, x3, y3, 80, direction="left")
    dim(doc, x5,y5, x4, y4, 80, direction="right")
    dim(doc, x6,y6, x5, y5, 80, direction="up")
    dim(doc, x3,y3, x4, y4, 80, direction="down")
   
    dim_leader_line(doc, x4, y4 , x4 + 50 , y4 - 100, "Vcut 4개소")

    # 중판 #1 description    
    x1 = frameXpos + 1980 + (midplate_width-960)
    y1 = frameYpos + 1125 +  (midplate_height-722)
    textstr = f"Part Name : 중판#2"    
    draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : {LCplateMaterial}"    
    draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
    textstr = f"Size : {str(math.ceil(midplatesizex2*100)/100)} x {str(math.ceil(midplatesizey2*100)/100)}mm"    
    draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
    textstr = f"Quantity : {su} EA"    
    draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')        

    frameXpos = frameXpos + TargetXscale + 1400

    #################################################################################################
    # 6page 중판 보강 'ㄷ'자 형태
    #################################################################################################        
         
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 1520 + (CW - 1600)  # 기본 모형이 1450
    TargetYscale = 921 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    

    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

    rx1 = math.ceil(frameXpos  + 104)
    ry1 = frameYpos + 630
    
    thickness = 1.2
    vcut = thickness / 2

    # 중판 전개도 크기 정의
    midplate_width = CW - 320 * 2
    midplate_su = 2

    if(CD>=2000):
        midplate_height = (CD - 28 * 2)/3
        midplate_su = 3
    else:
        midplate_height = (CD - 28 * 2)/2

    topwing1 = 13.8
    topwing2 = 15
    btwing1 = 13.8
    btwing2 = 15
    leftwing1 = 13.8
    leftwing2 = 15
    rightwing1 = 13.8
    rightwing2 = 15

    # 중판 width - 40 적용 날개 간섭등 공차 차감
    rib_width = midplate_width - 40

    x1 = rx1 
    y1 = ry1
    x2 = x1 + rib_width
    y2 = y1    
    x3 = x2
    y3 = y2 + 12 - thickness
    x4 = x3
    y4 = y3 + 30 - thickness * 2
    x5 = x4 
    y5 = y4 + 12 - thickness
    x6 = x5 - rib_width
    y6 = y5 
    x7 = x6 
    y7 = y6 - 12 + thickness
    x8 = x7 
    y8 = y7 - 30 + thickness * 2

    # 전개도 사이즈 저장 #1 중판
    midplatesizex = x2 - x1
    midplatesizey = y5 - y1

    # 중판1보다 4점 추가됨 주의
    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 8
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
      
    #절곡라인  밴딩 가로 방향      
    line(doc, x8, y8, x3, y3,  layer='hidden')   # 절곡선      
    line(doc, x7, y4, x4, y4,  layer='hidden')   # 절곡선      
   
    # 상부 치수선
    dim(doc, x6, y6, x5, y5 , 100 , direction="up")

    # 우측 치수선
    dim(doc, x5, y5, x4, y4 , 40 , text_height=0.12, direction="right")
    dimto(doc,  x3, y3 , 150 )      
    dimto(doc,  x2, y2 , 80 )  
    dim(doc, x5, y5, x2, y2 , 210 , direction="right")
    
    # 우측 단면도 그리기
    # 왼쪽 leftwing1 leftwing2
    # 오른쪽 날개 rightwing1 rightwing2
    # 본판 mainplate = mp 
    # 'ㄷ'자 형상 

    leftwing = 12
    mp = 30
    rightwing = 12

    startx = x1 + rib_width + 350
    starty = y1 + (leftwing + rightwing)

    x1 = startx
    y1 = starty
    x2 = x1
    y2 = y1 - thickness    
    x3 = x2 - rightwing 
    y3 = y2 
    x4 = x3 
    y4 = y3 + mp
    x5 = x4 + leftwing
    y5 = y4 
    x6 = x5 
    y6 = y5 - thickness
    x7 = x6 - leftwing + thickness
    y7 = y6 
    x8 = x7 
    y8 = y7 - mp + thickness*2

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 8
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="0")     

    # 단면도 치수선
    dim(doc, x3,y3, x2, y2, 100,  text_height=0.12, direction="down")
    dim(doc, x4,y4, x3, y3, 50,  text_height=0.12, direction="left")    
    dim(doc, x4,y4, x5, y5, 60, text_height=0.12,  direction="up")    

    # 중판 보강 description    
    x1 = frameXpos + 490 + (midplate_width-960) * frame_scale
    y1 = frameYpos + 450 +  (midplate_height-722) * frame_scale
    textstr = f"Part Name : (중판 보강)"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : EGI 1.2T"
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')        
    textstr = f"Size : {math.floor(midplatesizex)} x {math.floor(midplatesizey*10)/10}mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*4} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')        

    frameXpos = frameXpos + TargetXscale + 1400

    ####################################################################################################################################################
    # 레이져 도면틀 생성 (공통) 레이져에 배열할때 편리한 기능
    # 1219*1950 철판크기 샘플
    x1 = 10000+850
    y1 = 5000+130
    x2 = 10000+850+2438
    y2 = y1 +1219
    rectangle(doc, 10000, 5000,10000+8400 , 5000+ 2340 , layer='0')  
    rectangle(doc,x1 , y1 ,x1 + 2438 , y2 , layer='레이져')  
    dim(doc, x1 , y2 , x2 , y2,  100, direction="up")
    rectangle(doc, x1+2600 , y1 , x2+2600 , y2 , layer='레이져')  
    dim(doc, x1+2600, y2, x2+2600, y2, 100, direction="up")
    dim(doc,  x1 , y1, x1 ,y2 , 150, direction="left")

    textstr = f"{secondord} - {workplace} - {drawdate_str_short} 출도, {deadlinedate_str_short} 납기"    
    draw_Text(doc, 10000+300 , 5000+ 2340 - 200 ,  80, str(textstr), layer='0')        
    textstr = f"{LCframeMaterial}"
    draw_Text(doc, 10000+1000 , 5000+ 2340 - 200 - 200 ,  90, str(textstr), layer='0')        
    textstr = f"{LCplateMaterial}"
    draw_Text(doc, 10000+1000+2600 , 5000+ 2340 - 200 - 200 ,  90, str(textstr), layer='0')        
    textstr = f"EGI 1.2T"        
    draw_Text(doc, 10000+1000+5200 , 5000+ 2340 - 200 - 200 ,  90, str(textstr), layer='0')        
   
####################################################################################################################################################################################
# 031 천장 자동작도
####################################################################################################################################################################################
def lc031():   
    global args  #수정할때는 global 선언이 필요하다. 단순히 읽기만 할때는 필요없다.    
    global start_time
    global workplace, drawdate, deadlinedate, lctype, secondord, text_style, exit_program, Vcut_rate, Vcut, program_message, formatted_date, text_style_name
    global CW_raw, CD_raw, panel_thickness, su, LCframeMaterial, LCplateMaterial
    global input_CD, input_CW,CD,CW,drawdate_str , deadlinedate_str, drawdate_str_short, deadlinedate_str_short
    global doc, msp , T5_is, text_style, distanceXpos, distanceYpos

    # 2page 덴크리 다온텍등 등기구 설명 그림
    insert_block(0 , 0 , "lc031_2pageframe")        
    abs_x = 0
    abs_y = 0
    # LC 크기 지정 car inside에서 차감
    LCD = CD-50
    LCW = CW-60  # 라이트 케이스 크기를 정한다. 보통은 50

    # watt 계산 공식 적용해야 함
    # LED바 기장(단위:m) X 수량 X 15 = 60W 이하
    # LED바 1개당 watt 산출 m단위로 계산.. /1000  031모델은 3개의 led바가 들어감     

    ledsu = 3

    if(CD>2000):
        ledsu = 4

    wattCalculate = math.ceil((LCD - 3)/1000 * ledsu * 15)

    # print("wattCalculate: " + str(wattCalculate))  

    if(wattCalculate <= 60):
        watt = 60
    elif(wattCalculate > 60 and wattCalculate < 100):
        watt = 100
    elif(wattCalculate >= 100 and wattCalculate < 150):
        watt = 150
    elif(wattCalculate >= 150 and wattCalculate < 200):
        watt = 200
    elif(wattCalculate >= 200 and wattCalculate < 200):
        watt = 300
                
    # print("L/C 1개당 규격설정 wattCalculate: " + str(watt))  

    T5_standard = 0
    if T5_is == '있음':
        adjusted_CD = CD - 50  # CD에서 50 뺌

        # 규격 결정
        if adjusted_CD > 1500:
            T5_standard = 1500
        elif adjusted_CD > 1200:
            T5_standard = 1200
        else:
            T5_standard = 900

        # print(f"선택된 규격: {T5_standard}")
    else:
        # print("T5_is 가 '있음'이 아니므로 규격 선택 안 함")    
        pass

    # 도면틀 넣기
    BasicXscale = 2560
    BasicYscale = 1550
    TargetXscale = 800 
    TargetYscale = 300 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale

    frame_scale = 1    
    ################################################################################################################################################
    # 갑지 1
    ################################################################################################################################################
    # 현장명
    textstr = workplace       
    x = abs_x + 80
    y = abs_y + 1000 + 218
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 발주처
    textstr = secondord
    x = abs_x + 80
    y = abs_y + 1000 + 218 - 30
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 타입
    textstr = lctype
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 수량
    textstr = su
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 카사이즈
    textstr = f"W{CW} x D{CD}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 색상 LC
    LCcolor = '백색무광'
    if LCframeMaterial=='SPCC 1.2T' :
        LCcolor = '백색무광'
    textstr = f"L / C : {LCcolor}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')       
    LCplateMaterial = 'SPCC 1.2T' 
    textstr = f"중판 : {LCplateMaterial}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18 - 12
    draw_Text(doc, x, y , 7, str(textstr), layer='0')

    # (등기구업체 : 덴크리)   
    if T5_standard > 0 : 
        Ledcompany = f"(등기구업체 : 엘엠팩트)"        
        Leditem = "T5"
        Ledwatt = "6500K"
        LedBarLength = T5_standard
        watt = "내장"
    else:
        Ledcompany = f"(등기구업체 : 다온텍)"
        Leditem = "LED BAR"
        Ledwatt = "10000K"
        LedBarLength = CalculateLEDbar031(CW)
        watt = f"{str(watt)}W"

    x = abs_x + 80
    y = abs_y + 1000 + 90
    draw_Text(doc, x, y , 12, str(Ledcompany), layer='0')     
    # (아답터용량 : 60W ) 031모델은 슬림 아님
    textstr = f"(아답터용량 : SMPS 2-{watt} )"
    x = abs_x + 80
    y = abs_y + 1000 + 90 - 20
    draw_Text(doc, x, y , 12, str(textstr), layer='0')
    textstr = f"*. 출도일자 : {drawdate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 
    draw_Text(doc, x, y , 8, str(textstr), layer='0')    
    textstr = f"*. 납품일자 : {deadlinedate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 - 15
    draw_Text(doc, x, y , 8, str(textstr), layer='0')

    ################################################################################################################################################
    # 갑지 2
    ################################################################################################################################################
    # led바에 대한 내용 기술
    # LED BAR :

    textstr = f"{Leditem} - {LedBarLength}L={su*ledsu*2}(EA)"
    x = abs_x + 720 + 60
    y = abs_y + 650 + 32
    draw_Text(doc, x, y , 12, str(textstr), layer='0')        

    textstr = f"SMPS({watt})={su*2}(EA)"
    x = abs_x + 720 + 60
    y = abs_y + 650 - 35 + 32
    draw_Text(doc, x, y , 12, str(textstr), layer='0')    

    textstr = f"({watt})"            
    draw_Text(doc, 506, 310 , 18, str(textstr), layer='0')    

    # Type
    textstr = f"{lctype}타입"
    x = abs_x + 711 + 100
    y = abs_y + 630
    draw_Text(doc, x, y , 18, str(textstr), layer='0')    
    # 현장명
    textstr = f"*.현장명 : {secondord}-{workplace} - {su}(set)"
    x = abs_x + 30
    y = abs_y + 650 
    draw_Text(doc, x, y , 20, str(textstr), layer='0')    
    # (등기구업체 : )    
    x = abs_x + 30 + 200
    y = abs_y + 600 
    draw_Text(doc, x, y , 20, str(Ledcompany), layer='0')    
    
    textstr1 = f"{Leditem}-(누드&클립&잭타입)"
    textstr2 = f"{Ledwatt}-{LedBarLength}L"

    x = abs_x + 250
    y = 480    
    draw_Text_direction(doc, x, y , 14, str(textstr1), layer='0', rotation = 0)    
    draw_Text_direction(doc, x + 50, y - 30 , 14, str(textstr2), layer='0', rotation = 0)    

    # car inside box 표기
    xx =  790
    yy =  45
    rectangle(doc, xx,yy,xx+200,yy+100,layer='0')      
    line(doc, xx,yy+50,xx+200,yy+50)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50, yy+50 , 12, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30, yy+25 , 12, str(textstr), layer='0')

    frameXpos = 1037

    # Assy Car case    
    #####################################################################################################################################
    # 1page 조립도 Assy
    ##################################################################################################################################### 
    # LC ASSY 상부 형상    
    abs_x = frameXpos + 1700    
    frameXpos = abs_x - 500
    abs_y = abs_y + CD + 500
    insert_block(abs_x , abs_y , "lc031_top_left")    

    x = abs_x + CW 
    y = abs_y 
    insert_block(x , y , "lc031_top_right")  

    first_space = 135

    # 폴리갈 or 중판 크기 산출
    polygal_width = CW - first_space * 2
    polygal_height = LCD - 6

    # print("polygal_height: " + str(polygal_height))  

    textstr =  f"중판 폭 = {math.ceil(polygal_width)}"
    dim(doc, abs_x + first_space , y + 13, abs_x + CW - first_space , y + 13, 45 ,  text_height=0.26, direction="up", text= textstr)    

    # 중판 녹색 실물 그리기 6t
    rectangle(doc, abs_x + 135 , abs_y + 7 , abs_x + CW - 135  ,abs_y + 7 + 6 , layer='1')  

    # 상부선 2.0mm 간격
    x1 = abs_x - 25
    y1 = abs_y + 150
    x2 = x1 + CW + 50
    y2 = y1
    rectangle(doc, x1, y1, x2, y2+2, layer='0')   
 
    # 1.2T 상부 히든선 추가
    x1 = abs_x + 130
    y1 = abs_y + 150-1.2
    x2 = x1 + CW - 130
    y2 = y1  
    line(doc, x1, y1, x2, y2, layer='hidden') 

    # megenta LED바 브라켓 및 led바 표현
    rectangle(doc, abs_x + 200 , y1 - 25 , abs_x + CW - 200 , y1-20  , layer='6')  
    insert_block(abs_x + 360  , abs_y + 128.8 , "lc031_ledbar_bracket_assy")    
    insert_block(abs_x + CW - 360 - 40  , abs_y + 128.8 , "lc031_ledbar_bracket_assy")    
    
    # Assy 상부 의장면 OPen size
    x1 = abs_x + 155
    y1 = abs_y + 7
    x2 = abs_x + CW - 155
    y2 = y1    
    textstr =  f"의장면 Open Size = {CW-155*2}"      
    dim(doc, x1, y1, x2, y2, 90.5, text_height=0.30, text_gap=0.07, direction="down" , text = textstr)    

    x1 = abs_x + 30
    y1 = abs_y + 36
    x2 = abs_x + CW - 30
    y2 = y1        
    textstr =  f"Light Case Wide (W) = {CW-60}"    
    dim(doc, x1, y1, x2, y2, 260,  text_height=0.30, text_gap=0.07, direction="down",  text = textstr)

    # # 25 간격
    # dim(doc, x1, y1, abs_x, abs_y, 260, text_height=0.30, text_gap=0.07, direction="down")
    # dim(doc, x2, y1, x2+25, abs_y, 260, text_height=0.30, text_gap=0.07, direction="down")
    
    # Assy 상단 하부 치수
    x1 = abs_x 
    y1 = abs_y 
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"Car Inside (CW) = {CW}"
    dim(doc, x1, y1, x2, y2, 310,  text_height=0.30, text_gap=0.07, direction="down", text=textstr)

    ###################################
    # 1page Assy 본판 작도
    ###################################
    # 기본 윤곽 car inside
    width_insideGap = 30
    height_insideGap = 25
    abs_y = abs_y - CD - 800    
    x1 = abs_x 
    y1 = abs_y 
    x2 = abs_x + CW
    y2 = abs_y + CD    
    rectangle(doc, x1, y1, x2, y2, layer='hidden')    
    x1 = abs_x + width_insideGap
    y1 = abs_y + height_insideGap
    x2 = abs_x + CW - width_insideGap
    y2 = abs_y + CD - height_insideGap
    rectangle(doc, x1, y1, x2, y2, layer='0')    

    # 상하 30 철판 표현
    x1 = abs_x + 132
    y1 = abs_y + height_insideGap
    x2 = abs_x + CW - 132
    y2 = y1+30
    rectangle(doc, x1, y1, x2, y2, layer='0')    
    x1 = abs_x + 132
    y1 = abs_y + CD - height_insideGap
    x2 = abs_x + CW - 132
    y2 = y1-30
    rectangle(doc, x1, y1, x2, y2, layer='0')    

    # 양쪽 테두리 30 흰색선
    x1 = abs_x + 100
    y1 = abs_y + 25
    x2 = x1 + 30
    y2 = abs_y + CD - 25
    rectangle(doc, x1, y1, x2, y2, layer='0') 
    x1 = abs_x + CW - 100 - 30
    y1 = abs_y + 25
    x2 = x1 + 30
    y2 = abs_y + CD - 25
    rectangle(doc, x1, y1, x2, y2, layer='0') 

    # 내부 중판 자리 하늘색    
    x1 = abs_x + 360
    y1 = abs_y + 30
    x2 = x1 + 40
    y2 = abs_y + CD - 30
    rectangle(doc, x1, y1, x2, y2, layer='4')   # 하늘색 4번
    Rx1 = abs_x + CW - 360 - 40
    Ry1 = abs_y + 30
    Rx2 = Rx1 + 40
    Ry2 = abs_y + CD - 30
    rectangle(doc, Rx1, Ry1, Rx2, Ry2, layer='4')   # 하늘색 4번

    # 중간 LED Bracket 치수선
    dim(doc, abs_x + 130 , abs_y + CD - 340  , x1 , abs_y + CD - 340 ,   100, text_height=0.30, text_gap=0.07, direction="up")            
    dimcontinue(doc,  x2 , abs_y + CD - 340 )          
    dimcontinue(doc,  Rx1 , abs_y + CD - 340 )          
    dimcontinue(doc,  Rx2 , abs_y + CD - 340 )          
    dimcontinue(doc,  abs_x + CW - 130 , abs_y + CD - 340 )          

    # LED바 그리기 CD2000, 2000이상 구분해서 하기
    y_coords = calculate_light_positions_adjusted_031(CD)
    for i, coord in enumerate(y_coords):
        rectangle(doc, abs_x+150 , abs_y + CD - coord - 5, abs_x+CW-150, abs_y + CD - coord + 5, layer='6')  # led바 표시 magenta
        circle_cross(doc, abs_x + 380,  abs_y + CD - coord , 3.2 , layer='4') # 3.2 pie 
        circle_cross(doc, abs_x + CW - 380,  abs_y + CD - coord , 3.2 , layer='4') # 3.2 pie 
        if(i==3):
            textstr = f"{Leditem} - {Ledwatt}-{LedBarLength}L"            
            dim_string(doc,  abs_x+150 ,  abs_y + CD - coord  ,  abs_x+CW-150,  abs_y + CD - coord ,  100,  textstr, text_height=0.32, text_gap=0.07, direction="up")                         
            # print(f"i 번호 {i} textstr : {textstr}")

    # 좌측치수선
        if i == 0:
            dim(doc, abs_x + 360 , abs_y + CD - 30  ,  abs_x + 150  ,abs_y + CD - coord,  480, text_height=0.30, text_gap=0.07, direction="left")
        else:
            dimcontinue(doc,  abs_x + 150 , abs_y + CD - coord )         

    # 좌측치수선
    #마지막 치수선 추가            
    dimcontinue(doc,  abs_x + 360 , abs_y + 30 )      
    dim(doc, abs_x + 360 , abs_y + CD - 30  , abs_x + 360 , abs_y + 30,   540, text_height=0.30, text_gap=0.07, direction="left")    

    # LC 홀 가공 (기존 본천장 로직 적용해서 만들것)
    inputCW = CW
    inputCD = CD

    lc_positions = calculate_lc_position_031(inputCW, inputCD)

    # Topholex와 Topholey 좌표 추출
    Topholex = lc_positions['Topholex']
    Topholey = lc_positions['Topholey']

    # Bottomx와 Bottomy 좌표 추출
    Bottomx = lc_positions['Bottomx']
    Bottomy = lc_positions['Bottomy']    

    # 필요한 경우 다른 홀들에 대해서도 circle_cross 함수를 호출할 수 있음
    # 예를 들어, 모든 상단 홀에 대해 반복
    for x in Topholex:
        circle_cross(doc, abs_x + x, abs_y + Topholey, 11, layer='레이져')    
        # print(f"topholex x좌표: {x} y좌표 {abs_y + Topholey} abs_y좌표 : {abs_y}  Topholey : {Topholey}")     

    for x in Bottomx:
        circle_cross(doc, abs_x + x, abs_y + Bottomy, 11, layer='레이져')            

    # 세로 홀 위치 좌표 계산
    hole_positions = calculate_lc_vertical_hole_positions_031(inputCW, inputCD)

    # 각 홀의 좌표를 추출
    leftholex = hole_positions['leftholex']
    rightholex = hole_positions['rightholex']
    vertical_holey = hole_positions['vertical_holey']

    # 리스트 길이 확인
    num_holes = min(len(leftholex), len(rightholex), len(vertical_holey))

    # circle_cross 함수 호출 시 리스트 길이를 확인하여 오류 방지
    for i in range(num_holes):
        circle_cross(doc, abs_x + leftholex[i], abs_y + vertical_holey[i], 11, layer='레이져')
        circle_cross(doc, abs_x + rightholex[i], abs_y + vertical_holey[i], 11, layer='레이져')
        # print(f"leftholex x좌표: {leftholex[i]} abs_x좌표 {abs_y }  vertical_holey : {vertical_holey[i]}")    

    # assy smps 2가지 가져오기
    # smps가 안들어가는 T5는 제외
    if T5_standard  < 1 :    
        x = abs_x + 32 
        y = abs_y + 125
        insert_block(x , y , "lc031_assy_smps")      
        x = abs_x + CW - 98 - 32
        y = abs_y + 125
        insert_block(x , y , "lc031_assy_smps")     

    if CD-100 == LedBarLength:
        LedSpace = 23
    else:
        LedSpace = 23 - ( LedBarLength-(CD-100) ) / 2

    # 상부 치수선
    x1 = abs_x 
    y1 = abs_y + CD
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"L/C SIZE =  {CW-60}"
    dim(doc, x1 + 30 , y1 - 25, x2 - 30 , y2 - 25, 210,  text_height=0.32, text_gap=0.07, direction="up",  text = textstr)   

    x1 = abs_x 
    y1 = abs_y + CD
    x2 = abs_x + CW
    y2 = y1    
    textstr =  f"CAR INSIDE(CW) = {CW}"
    dim(doc, x1 , y1 , x2  , y2, 290,  text_height=0.32, text_gap=0.07, direction="up", text = textstr)    

    # 30 치수 표현
    dim(doc,  x1+30, y2,  x1, y1, 185,  text_height=0.30, text_gap=0.07, direction="up")    
    dim(doc, x2-30, y1, x2, y2, 185,  text_height=0.30, text_gap=0.07, direction="up")
  
    # 우측 치수선  
    x1 = abs_x + CW - 30
    y1 = abs_y + CD - 25
    x2 = abs_x + CW - 115
    y2 = y1 - 15
    dim(doc, x1, y1, x2, y2, 220 - 95, text_height=0.30,  direction="right")        

    x1 = abs_x + CW - 115
    y1 = abs_y + 40
    x2 = abs_x + CW - 30
    y2 = y1 - 15
    dim(doc, x1, y1, x2, y2, 220 + 25 , text_height=0.30,  direction="right")

    # 25 간격 치수선 표기
    x1 = abs_x + CW 
    y1 = abs_y + CD 
    x2 = x1 - 30
    y2 = y1 - 25
    dim(doc, x1, y1, x2, y2, 250-25, text_height=0.30,  direction="right")

    x1 = abs_x + CW - 30
    y1 = abs_y + 25
    x2 = abs_x + CW
    y2 = abs_y
    dim(doc, x1, y1, x2, y2, 250, text_height=0.30,  direction="right")

    # 상단, 하단 홀의 오른쪽 끝 y 좌표 추출
    right_end_topholey = Topholey
    right_end_bottomy = Bottomy

    # 세로 방향 홀의 y 좌표 추출
    vertical_holes_rightholey = vertical_holey

    # 모든 오른쪽 끝 홀의 y 좌표를 하나의 리스트로 결합
    all_right_end_holes_y_coords = [right_end_topholey] + vertical_holes_rightholey + [right_end_bottomy]

    # 모든 오른쪽 끝 홀의 y 좌표를 오름차순으로 정렬
    sorted_y_coords = sorted(all_right_end_holes_y_coords)

    # 정렬된 y 좌표 출력
    last_ypos = 40
    for i, y_coord in enumerate(sorted_y_coords):        
        # print(f"세로 홀의 y 좌표: {y_coord}")
        if i>0:
            dim(doc, abs_x + CW -115 , abs_y + last_ypos, abs_x + CW -115 , abs_y +  y_coord, 274, text_height=0.30,  direction="right")
            last_ypos = y_coord

    x1 = abs_x + CW - 25
    y1 = abs_y + CD - 25
    x2 = x1
    y2 = abs_y + 25        
    textstr =  f"LIGHT CASE SIZE = {CD-panel_thickness*2}"
    # dim_vertical_right_string(doc, x1, y1, x2, y2, 308, textstr , text_height=0.20,  text_gap=0.07)        
    dim(doc,   x2, y2, x1, y1,   285, text_height=0.32, direction="right", text = textstr)    
    x1 = abs_x + CW
    y1 = abs_y + CD
    x2 = x1
    y2 = abs_y         
    textstr =  f"CAR INSIDE {CD}"    
    dim(doc,  x2, y2, x1, y1,  350, text_height=0.32, direction="right", text = textstr)     

    #########################################################################
    # 우측단면도 표기 (블럭 삽입)
    #########################################################################
    x = abs_x + CW + 450
    y = abs_y + CD
    insert_block(x , y , "lc031_assy_side_section_top")         
    x1 = x
    y1 = abs_y 
    insert_block(x1 , y1 , "lc031_assy_side_section_bottom")         
    # if(CD <= 2000):
    x2 = x + 128
    y2 = abs_y + CD/2 
    insert_block(x2 , y2 , "lc031_assy_side_section_midplate")         
    rectangle(doc, x-2.3, y + 30 , x, y1 - 30, layer="0")

    # 폴리갈 or 중판 크기 산출
    first_space = 46
    polygal_width = (CD - 47.1 - 235.95 - 48.95)/2
    polygal_height = LCD - 6    

    print (f"단면도 polygal_width : {polygal_width}")

    # CD 2000 이하 폴리갈 표현 1번
    x3 = x + 120.8
    y3 = y - 47.1
    x4 = x3 + 6
    y4 = y3 - polygal_width 
    rectangle(doc, x3, y3 , x4, y4, layer="4")    
    x5 = x + 120.8
    y5 = y1 + 48.95
    x6 = x5 + 6
    y6 = y5 + polygal_width 
    rectangle(doc, x5, y5 , x6, y6, layer="4")    
    dim(doc, x + 119.6 , y-42.1, x + 119.6 , y-42.1 - polygal_width - 10, 66, text_height=0.30,  direction="left")
    dim(doc, x + 119.6 , abs_y + 43.95, x + 119.6 , abs_y + 43.95 + polygal_width + 10, 66, text_height=0.30,  direction="left")

    x7 = x + 135
    y7 = y 
    x8 = x7 
    y8 = y7 - 25    
    dim(doc, x7, y7, x8, y8, 66,  direction="right")
    x9 = x + 135
    y9 = y1 
    x10 = x9 
    y10 = y9 + 25        
    dim(doc, x9, y9, x10, y10, 66,  direction="right")
    
    textstr = f"Light Case Wide (D) = {LCD}"
    dim(doc, x8, y8, x10 , y10,  120, text_height=0.3, direction="right", text=textstr)
    
    textstr = f"Car Inside (CD) = {CD}"
    dim(doc, x9, y9, x7 , y7,  230, text_height=0.3, direction="right", text=textstr)

    ####################################################################################
    # LC031 모델은 새로운 중판모양을 우측 화면에 표시한다. 
    # 기존의 공식을 이용해서 그린다.
    ####################################################################################
    # 1page Assy 본판 작도 코드 재사용
    ###################################

    Right_abs_x = abs_x + CW + 1300
    Right_abs_y = abs_y 
    
    # 기본 윤곽 car inside
    width_insideGap = 30
    height_insideGap = 25
   
    x1 = Right_abs_x 
    y1 = Right_abs_y 
    x2 = Right_abs_x + CW
    y2 = Right_abs_y + CD    
    rectangle(doc, x1, y1, x2, y2, layer='hidden')    
    x1 = Right_abs_x + width_insideGap
    y1 = Right_abs_y + height_insideGap
    x2 = Right_abs_x + CW - width_insideGap
    y2 = Right_abs_y + CD - height_insideGap
    rectangle(doc, x1, y1, x2, y2, layer='0')    

    # 알루미늄 형상 화면 표시
    # 첫 번째 구간
    start_x1 = Right_abs_x + 30
    end_x1 = Right_abs_x + 145
    y1_1 = Right_abs_y + 26.2
    y2_1 = Right_abs_y + CD - 26.2
    draw_patterned_lines(doc, start_x1, end_x1, y1_1, y2_1, '2')

    # 두 번째 구간
    start_x2 = Right_abs_x + CW - 145
    end_x2 = Right_abs_x + CW - 30
    y1_2 = Right_abs_y + 26.2
    y2_2 = Right_abs_y + CD - 26.2
    draw_patterned_lines(doc, start_x2, end_x2, y1_2, y2_2, '2')            

    # 내부 중판 자리 하늘색    
    x1 = Right_abs_x + 135
    y1 = Right_abs_y + 46
    x2 = Right_abs_x + CW - 135
    y2 = y1 + polygal_width + 10
    rectangle(doc, x1, y1, x2, y2, layer='4')   # 하늘색 4번

    # 폴리갈 해치    
    hatch = msp.add_hatch(dxfattribs={'layer': '5'})  # 5는 파란색, CL 적색 코드입니다.

    # "CORK" 패턴으로 해치 패턴 설정 및 90도 회전
    hatch.set_pattern_fill("CORK", scale=5.0, angle=0.0, color='5')  # 여기서 scale은 패턴 크기, angle은 패턴 회전 각도

    # 경계선 추가 
    hatch.paths.add_polyline_path(
        [(x1, y1), (x2, y1), (x2, y2), (x1, y2)], 
        is_closed=True
    )

    x1 = Right_abs_x + 135
    y1 = Right_abs_y + CD - 46
    x2 = Right_abs_x + CW - 135
    y2 = y1 - polygal_width - 10
    rectangle(doc, x1, y1, x2, y2, layer='4')   # 하늘색 4번


    # 폴리갈 해치    
    hatch = msp.add_hatch(dxfattribs={'layer': '5'})  # 5는 파란색, CL 적색 코드입니다.

    # "CORK" 패턴으로 해치 패턴 설정 및 90도 회전
    hatch.set_pattern_fill("CORK", scale=5.0, angle=0.0, color='5')  # 여기서 scale은 패턴 크기, angle은 패턴 회전 각도

    # 경계선 추가 
    hatch.paths.add_polyline_path(
        [(x1, y1), (x2, y1), (x2, y2), (x1, y2)], 
        is_closed=True
    )    

    rectangle(doc, Right_abs_x + 30 , Right_abs_y + 25, Right_abs_x + CW - 30, Right_abs_y + 25 + 1.2, layer='0')   
    rectangle(doc, Right_abs_x + 30 , Right_abs_y + CD - 25, Right_abs_x + CW - 30, Right_abs_y + CD - 25 - 1.2, layer='0')   
    # 30 + 115선 두개
    rectangle(doc, Right_abs_x + 30 , Right_abs_y + 25, Right_abs_x + 30 + 115, Right_abs_y + CD - 25 - 1.2, layer='0')  
    rectangle(doc, Right_abs_x + CW - 30 - 115 , Right_abs_y + 25,Right_abs_x + CW - 30, Right_abs_y + CD - 25 - 1.2, layer='0')        

    #####################################################
    # 상부 치수선
    #####################################################
    x1 = Right_abs_x 
    y1 = Right_abs_y + CD
    x2 = Right_abs_x + CW
    y2 = y1    
    textstr =  f"L/C SIZE =  {CW-60}"
    dim(doc, x1 + 30 , y1 - 25, x2 - 30 , y2 - 25, 210,  text_height=0.32, text_gap=0.07, direction="up",  text = textstr)   

    x1 = Right_abs_x 
    y1 = Right_abs_y + CD
    x2 = Right_abs_x + CW
    y2 = y1    
    textstr =  f"CAR INSIDE(CW) = {CW}"
    dim(doc, x1 , y1 , x2  , y2, 290,  text_height=0.32, text_gap=0.07, direction="up", text = textstr)    

    # 30 치수 표현
    dim(doc,  x1+30, y2,  x1, y1, 185,  text_height=0.30, text_gap=0.07, direction="up")    
    dim(doc, x2-30, y1, x2, y2, 185,  text_height=0.30, text_gap=0.07, direction="up")

    # 상부 3번째 줄 표기
    x1 = Right_abs_x + 145
    y1 = Right_abs_y + CD - 26.2
    x2 = Right_abs_x + CW - 145
    y2 = y1    
    textstr =  f"의장면 Open Size = {CW-290}"
    dim(doc, x1 , y1 , x2  , y2, 130,  text_height=0.28, text_gap=0.07, direction="up", text = textstr)        

    # 상부 좌우 115 치수 표현
    dim(doc,  x1-115, y2,  x1, y1, 130,  text_height=0.28, text_gap=0.07, direction="up")    
    dim(doc,  x2, y1,  x2+115, y2, 130,  text_height=0.28, text_gap=0.07, direction="up")

    #####################################################
    # 하부 치수선
    #####################################################
    distance = 135
    x1 = Right_abs_x + distance
    y1 = Right_abs_y + 26.2
    x2 = Right_abs_x + CW - distance
    y2 = y1    
    textstr =  f"중판 Size = {CW-distance*2}"
    dim(doc, x1 , y1 , x2  , y2, 130,  text_height=0.28, text_gap=0.07, direction="down", text = textstr)        

    # 하부 좌우 105 치수 표현
    dim(doc,  x1-105, y2,  x1, y1, 130,  text_height=0.28, text_gap=0.07, direction="down")    
    dim(doc,  x2, y1,  x2+105, y2, 130,  text_height=0.28, text_gap=0.07, direction="down")

    #####################################################
    # 좌측 치수선
    #####################################################
    x1 = Right_abs_x + 135
    y1 = Right_abs_y + CD - 46
    x2 = x1
    y2 = y1 - polygal_width - 10   
    textstr =  f"폴리갈 외경 = {math.ceil(polygal_width+10)}"
    dim(doc, x1 , y1 , x2  , y2, 180 ,  text_height=0.25, text_gap=0.07, direction="left", text = textstr)            
    y3 = y2
    y4 = y3 - 220
    textstr =  f"220"
    dim(doc, x1 , y3 , x2  , y4, 180 ,  text_height=0.25, text_gap=0.07, direction="left", text = textstr)    
    y5 = y4
    y6 = y5 - polygal_width - 10   
    textstr =  f"폴리갈 외경 = {math.ceil(polygal_width+10)}"
    dim(doc, x1 , y5 , x2  , y6, 180 ,  text_height=0.25, text_gap=0.07, direction="left", text = textstr)            
  
    #####################################################
    # 우측 치수선
    #####################################################
    # 25 간격 치수선 표기
    x1 = Right_abs_x + CW 
    y1 = Right_abs_y + CD 
    x2 = x1 - 30
    y2 = y1 - 25
    dim(doc, x1, y1, x2, y2, 200-25, text_height=0.30,  direction="right")

    x1 = Right_abs_x + CW - 30
    y1 = Right_abs_y + 25
    x2 = Right_abs_x + CW
    y2 = Right_abs_y
    dim(doc, x1, y1, x2, y2, 200, text_height=0.30,  direction="right")

    x1 = Right_abs_x + CW - 25
    y1 = Right_abs_y + CD - 25
    x2 = x1
    y2 = Right_abs_y + 25        
    textstr =  f"L/C SIZE(D) = {CD-panel_thickness*2}"
    # dim_vertical_right_string(doc, x1, y1, x2, y2, 308, textstr , text_height=0.20,  text_gap=0.07)        
    dim(doc,   x2, y2, x1, y1,   235, text_height=0.32, direction="right", text = textstr)    
    x1 = Right_abs_x + CW
    y1 = Right_abs_y + CD
    x2 = x1
    y2 = Right_abs_y         
    textstr =  f"CAR INSIDE(CD) {CD}"    
    dim(doc,  x2, y2, x1, y1,  300, text_height=0.32, direction="right", text = textstr)   

    #####################################################
    # 블럭 삽입 - 이탈방지 브라켓관련 블럭
    #####################################################
    # STS M/R 화면에 지정
    textstr = f"STS M/R"       
    x = Right_abs_x + CW/2 - len(textstr)*60/2
    y = Right_abs_y + CD/2 - 15
    draw_Text(doc, x, y , 60, str(textstr), layer='0')    
    # 구름마크 문구
    x = Right_abs_x 
    y = Right_abs_y + CD
    insert_block(x , y , "lc031_assy_side_section_memo")              
    # 좌측 B/K 표시
    x = Right_abs_x + 145
    y = Right_abs_y + 46 + (polygal_width+10)/2
    insert_block(x , y , "lc031_assy_side_section_bracket_left")              
    y = Right_abs_y + 46 + (polygal_width+10)/2 + polygal_width + 220
    insert_block(x , y , "lc031_assy_side_section_bracket_left")              
    # 우측 B/K 표시
    x = Right_abs_x + CW - 145
    y = Right_abs_y + 46 + (polygal_width+10)/2
    insert_block(x , y , "lc031_assy_side_section_bracket_right")              
    y = Right_abs_y + 46 + (polygal_width+10)/2 + polygal_width + 220
    insert_block(x , y , "lc031_assy_side_section_bracket_right")              

    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1715
    if CW > 1600:
        TargetXscale = 5400 + (CW - 1600)*1.7
    else:
        TargetXscale = 5400 + (CW - 1600)

    TargetYscale = 3300 + (CD - 1500) if CD<1500 else 3300 + (CD - 1500)*1.7     

    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale            
    frameYpos = abs_y - 450 * frame_scale     
    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)       
    # 1page 상단에 car inside 표기
    x = frameXpos + CW/3
    y = abs_y + frame_scale*BasicYscale + 50
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, x, y , 200*frame_scale, str(textstr), layer='0')    

    frameXpos = frameXpos + TargetXscale + 1500 if CD<1600 else frameXpos + TargetXscale + 1500 + (CD - 1600)*1.3

    #################################################################################################
    # 2page 프레임 본체만들기 (레이져가공)
    #################################################################################################        
    rx1 = frameXpos  + 410  + LCD + 355
    ry1 = frameYpos + 1600
    insert_block(rx1 , ry1 , "lc031_frame_section_side")    
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765

    TargetXscale = 3000  if CD < 1500 else 3000 + (CD - 1500) * 1.2 
    TargetYscale = 1800 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    

    # print("2page 스케일 비율 : " + str(frame_scale))           
    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

    rx1 = frameXpos  + 400
       
    thickness = 1.2
    vcut = thickness / 2

    x1 = rx1 
    y1 = ry1
    x2 = x1 + LCD - 3
    y2 = y1    
    x3 = x2
    y3 = y2 + 12 - vcut 
    x4 = x3
    y4 = y3 + 30 - thickness
    x5 = x4 
    y5 = y4 + vcut 
    x6 = x5 + 1.5
    y6 = y5 
    x7 = x6 
    y7 = y6 + 20 - vcut*3
    x8 = x7 
    y8 = y7 + 88
    x9 = x8 
    y9 = y8 + 29.6
    x10 = x9 - LCD
    y10 = y9 
    x11 = x10 
    y11 = y10 - 29.6
    x12 = x11 
    y12 = y11 - 88
    x13 = x12 
    y13 = y12 - 20 + vcut*3
    x14 = x13 + 1.5
    y14 = y13 
    x15 = x14 
    y15 = y14 - vcut 
    x16 = x15 
    y16 = y15 - 30 + thickness

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 16
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
      
    #절곡라인  (히든선 : 정방향)      
    line(doc, x16, y16, x3, y3, layer='hidden')    # 절곡선      
    line(doc, x15, y15, x4, y4, layer='hidden')    # 절곡선      
    line(doc, x12, y12, x7, y7, layer='hidden')    # 절곡선      
    line(doc, x11, y11, x8, y8, layer='hidden')    # 절곡선      

    # 상부 치수소
    dim(doc,  x10+15, y11+14.6 , x10 , y10 , 118, direction="up", option="reverse")            
   
    # 리스트 길이 확인
    if CD > 2000:
        num_holes = 4
    else:
        num_holes = 3

    for i in range(num_holes + 1):
        if i == 0 or i == num_holes:
            if i == 0:
                circle_cross(doc, x10 + 15, y11 + 14.6, 11, layer='레이져')
            else:
                circle_cross(doc, x10 + LCD - 15, y11 + 14.6, 11, layer='레이져')
                dimcontinue(doc, x10 + LCD - 15, y11 + 14.6)
        else:
            if i-1 < len(vertical_holey):
                circle_cross(doc, x10 + 15 + vertical_holey[i-1] - 40, y11 + 14.6, 11, layer='레이져')
                dimcontinue(doc, x10 + 15 + vertical_holey[i-1] - 40, y11 + 14.6)
                print(f"vertical Y : {vertical_holey[i-1]-40}")
            else:
                print(f"Index {i-1} is out of range for vertical_holey list.")
        if i == 1:
            baseX = x10 + 15 + vertical_holey[i-1] - 40 if i-1 < len(vertical_holey) else x10 + 15
            baseY = y10 - 15


    dimcontinue(doc,  x9, y9)
    # 상부 전체치수
    dim(doc,  x10, y10 , x9 , y9 , 165,  direction="up")         
    # 11파이 단공 라이트케이스 홀 
    dim(doc,  baseX, y10 , baseX , baseY , 100,  direction="left")  
    # 11파이 지시선
    dim_leader_line(doc, baseX, baseY , baseX + 100 ,baseY + 50, f"{num_holes} -%%C11 Hole (M8 Pop Nut)")   
    # vcut
    dim_leader_line(doc, x1 + 200, y3 , x1 + 200 + 50, y1 - 60 , f"V-Cut Line")   
    dim_leader_line(doc, x1 + 400, y4 , x1 + 400 + 70, y1 - 40 , f"V-Cut Line")   

    # 좌측치수선
    dim(doc,  x10, y10 , x13 , y13 , 123,  direction="left")         
    dim(doc,  x13 , y13 , x1 , y1 , 200,  direction="left")         
    
    # 하부 치수
    dim(doc,  x13, y13 , x1 , y1 , 100,  direction="down")         
    dim(doc,  x6 , y6 , x2, y2 ,  100,  direction="down")         

    # 오른쪽 치수
    dim(doc,  x3, y3 , x2 , y2 , 80,  direction="right")         
    dim(doc,  x4, y4 , x7 , y7 , 80,  direction="right")         
    dim(doc,  x8, y8 , x9 , y9 , 80,  direction="right")         
    dim(doc,  x4, y4 , x3 , y3 , 150,  direction="right")         
    dim(doc,  x7, y7 , x8 , y8 , 150,  direction="right")         
    dim(doc,  x2, y2 , x9 , y9 , 195,  direction="right")        

    # main Frame description    
    x1 = frameXpos + 950 + (CD-1500)
    y1 = frameYpos + 1400
    textstr = f"Part Name : Light Case (좌, 우)"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : SPCC 1.2t "    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : {LCD} x {str(math.floor((y10-y2)*10)/10)}mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')         

    ############################################################
    # 031 Front Frame  두번째 모양
    ############################################################   
    rx1 = frameXpos  + 400  + LCW + 355    
    ry1 = frameYpos + 700    
    insert_block(rx1 , ry1 , "lc031_frame_section_front")        

    rx1 = frameXpos  + 400
    # 알루미늄과 만나는 라운형태의 모양 가져오기
    basex = rx1 - 55.07
    y = ry1 + 201.7
    insert_block(basex , y , "lc031_frame_roundshape_front_right_update2", layer='레이져')       
    x = rx1 + (CW - 311) + 55.09 
    insert_block(x , y , "lc031_frame_roundshape_front_left_update2", layer='레이져')              
    thickness = 1.2
    vcut = thickness / 2

    x1 = rx1 
    y1 = ry1
    x2 = x1 + CW - 311
    y2 = y1    
    x3 = x2
    y3 = y2 + 14 - vcut 
    x4 = x3
    y4 = y3 + 12 - thickness
    x5 = x4 
    y5 = y4 + 7 - thickness
    x6 = x5 
    y6 = y5 + 23.45 - vcut
    x7 = x6 + 49
    y7 = y6 + 0.95
    x8 = x7 + 6.09
    y8 = y7 + 149.1 - thickness
    x9 = x8 - 31.58
    y9 = y8 
    x10 = x9 
    y10 = y9 + 0.3
    x11 = x10 
    y11 = y10 + 30 - thickness
    x12 = x11 - LCW - 0.69 + (102.345*2)  # 0.69는 오차값임
    y12 = y11 

    x13 = x12 
    y13 = y12 - 30 + thickness
    x14 = x13 
    y14 = y13 - 0.3
    x15 = x14 - 31.58
    y15 = y14
    x16 = x15 + 6.09
    y16 = y15 - 149.1 + thickness
    x17 = x16 + 49
    y17 = y16 - 0.95
    x18 = x17
    y18 = y17 - 23.45 + vcut
    x19 = x18
    y19 = y18 - 7 + thickness
    x20 = x19
    y20 = y19 - 12 + thickness
    x21 = x1
    y21 = y1

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 6
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    # line(doc, prev_x, prev_y, x1, y1, layer="레이져")      

    prev_x, prev_y = x8, y8  
    lastNum = 15
    for i in range(8, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    
    prev_x, prev_y = x17, y17  
    lastNum = 21
    for i in range(17, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y    
      
    #절곡라인  (히든선 : 정방향)      
    line(doc, x14, y14, x9, y9,  layer='hidden')    # 절곡선    누락으로 추가
    line(doc, x19, y19, x4, y4,  layer='hidden')    # 절곡선      
    line(doc, x18, y18, x5, y5,  layer='hidden')    # 절곡선      
    line(doc, x16, y16, x7, y7,  layer='hidden')    # 절곡선      
    #절곡라인  (실선 : 역방향)      
    line(doc, x20, y20, x3, y3,  layer='2') 

    # vcut 4개소 문구 삽입
    x = rx1 + 180
    # 알루미늄과 만나는 라운형태의 모양 가져오기    
    y = ry1 + 110
    insert_block(x , y , "lc031_frame_roundshape_vcut")       
    
    # 상부 치수
    dim(doc,  x12 ,  y11 - 15 , x15,  y15 ,  95, direction="up", option="reverse")            
   
    # 리스트 길이 확인
    if CW >= 1600:
        num_holes = 2
    else:
        num_holes = 1

    for i in range(num_holes + 1 ):
        if i != 0 :
            circle_cross(doc, x12 + Topholex[i]-115-17 , y11 - 15, 11, layer='레이져')       
            dimcontinue(doc, x12 + Topholex[i]-115-17,  y11 - 15)
            print (f"Front Case x 위치 : {Topholex[i]-115-17}")
        if i == 1:
            baseX = x12 + Topholex[i]-115-17
            baseY =  y11 - 15

    dimcontinue(doc,  x11, y11)
    dimcontinue(doc,  x8, y8)
    # 상부 3두번째 줄    
    dim(doc,  x12 , y12 , x12-102.37, y12-100,  140,  direction="up", option='reverse')         
    dimcontinue(doc,  x11, y11)
    dimcontinue(doc,  x11+102.37, y11- 100)    
    # 상부 3번째 줄
    dim(doc,  x12-102.37 , y12-100 , x11+102.37 , y11-100 , 300,  direction="up")         

    # 11파이 단공 라이트케이스 홀 
    dim(doc,  baseX, y11 , baseX , baseY , 100,  direction="left")  
    # 11파이 지시선
    dim_leader_line(doc, baseX, baseY , baseX + 100 ,baseY + 50, f"{num_holes} -%%C11 Hole (M8 Pop Nut)")   

    # 좌측치수선
    dim(doc,  x12 , y12 , x15, y15 ,  200,  direction="left")         
    dim(doc,  x15 , y15 , x17 , y17 , 170,  direction="left")         
    dim(doc,  x1 , y1 , x17 , y17 , 210,  direction="left")            
    
    # 하부 치수
    dim(doc,  x1, y1 , x1-125.85 , y1+130 , 100,  direction="down", option="reverse")         
    dimcontinue(doc,  x2 , y2)
    dimcontinue(doc,  x2+125.85 , y2+130)

    # 오른쪽 치수
    dim(doc,  x10 , y10 ,  x11, y11 , 180,  direction="right", option='reverse')         
    dimcontinue(doc,  x7 , y7)
    dimcontinue(doc,  x5 , y5)
    dim(doc,  x5, y5 , x4 , y4 , 260,  direction="right")     
    dim(doc,  x4, y4 , x3 , y3 , 300,  direction="right")              
    dim(doc,  x3, y3 , x2 , y2 , 320,  direction="right")     
    # 오른쪽 전체치수
    dim(doc,  x11, y11 , x2 , y2 , 380,  direction="right")       

    # main Frame description    
    x1 = frameXpos + 950 + (CW-1600)
    y1 = frameYpos + 520
    textstr = f"Part Name : Light Case (앞, 뒤)"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : SPCC 1.2t "    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : {LCW+0.69} x {str(math.floor((y11-y2)*10)/10)}mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')    

    frameXpos = frameXpos + TargetXscale + 1500 if CD<1600 else frameXpos + TargetXscale + 1500 + (CD - 1600)*1.3

    #################################################################################################
    # 3page 중판 이탈방지 B/K , LED BAR 고정 B/K, LED 선 고정 B/K
    #################################################################################################        
    rx1 = frameXpos  + 434
    ry1 = frameYpos + 1473
    insert_block(rx1 , ry1 , "lc031_rib_zshape_laser")    
    insert_block(rx1 , ry1 , "lc031_rib_zshape")    

    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 3000 + (CW - 1600) if CD < 1500 else 3000 + (CW - 1600) + (CD - 1600) * 1.1
    TargetYscale = 1800 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    

    # print("3page 스케일 비율 : " + str(frame_scale))           
    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

    x1 = frameXpos + 190
    y1 = frameYpos + 1250
    textstr = f"Part Name : 중판 이탈방지 B/K"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : EGI 1.2T"    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : 40.2 x 30mm"    
    draw_Text(doc, x1 , y1 -80  , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*4} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')    

    ######################################################
    # LED BAR 고정 B/K
    ######################################################

    rx1 = frameXpos  + 1000
    ry1 = frameYpos + 1418

    thickness = 1.5
    vcut = thickness / 2

    x1 = rx1 
    y1 = ry1
    x2 = x1 + LCD - 10 - 26*2
    y2 = y1    
    x3 = x2
    y3 = y2 + 1.2
    x4 = x3 + 26
    y4 = y3 
    x5 = x4 
    y5 = y4 + 18.8 - thickness
    x6 = x5 
    y6 = y5 + 40 - thickness *2
    x7 = x6 
    y7 = y6 + 18.8 - thickness
    x8 = x7 - 26
    y8 = y7 
    x9 = x8 
    y9 = y8 + 1.2
    x10 = x9 - LCD + 10 + 26*2
    y10 = y9 
    x11 = x10 
    y11 = y10 - 1.2
    x12 = x11 - 26
    y12 = y11 
    x13 = x12 
    y13 = y12 - 18.8 + thickness
    x14 = x13 
    y14 = y13 - 40 + thickness *2
    x15 = x14 
    y15 = y14 - 18.8 + thickness
    x16 = x15 + 26
    y16 = y15 

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 16
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
      
    #절곡라인  (히든선 : 정방향)      
    line(doc, x14, y14, x5, y5,  layer='hidden')    # 절곡선      
    line(doc, x13, y13, x6, y6,  layer='hidden')    # 절곡선      

    #  상부 치수 옆으로 이어지는 치수선 95 ~
    positions = calculate_hole_positions_separated_lc031(LCD - 10)
    # print(" calculate_hole_positions_separated_lc031 X좌표:", positions)    
    dim(doc,  x12 + positions[0], y14+18.5 , x12 , y12 , 110, direction="up", option="reverse")            
   
    # 리스트 길이 확인

    for i in range(len(positions) ):
        circle_cross(doc,  x12 + positions[i] , y14+18.5 , 3.2, layer='레이져')   
        if i != 0:            
            dimcontinue(doc,   x12 + positions[i] , y14+18.5 )
        if i == 1 :
            baseX = x12 + positions[i]
            baseY = y14+18.5
        if i == 2 :
            baseX1 = x12 + positions[i]
            baseY1 = y14+18.5

    dimcontinue(doc,  x7, y7)
    # 상부 전체치수
    dim(doc,  x12, y12 , x7 , y7 , 155,  direction="up")         
    # 3.2파이 단공 라이트케이스 홀 
    dim(doc,  baseX, y13 , baseX , baseY , 50,  direction="left")  
    dim(doc,  baseX , baseY , baseX, y14 , 80,  direction="right")  
    # 3.2파이 지시선
    dim_leader_line(doc, baseX1, baseY1 , baseX1 + 100 ,baseY1 + 50, f"{len(positions)} -%%C3.2 Hole")   

    # 좌측치수선
    dim(doc,  x12 , y12 , x13 + positions[0] , y14+18.5 ,  100,  direction="left")         
    dim(doc,  x1, y1 , x13 + positions[0] , y14+18.5 ,  100 ,  direction="left")         
    
    # 하부 치수
    dim(doc,  x1, y1 , x15 , y15 , 100,  direction="down", option='reverse')         
    dimcontinue(doc,  x2 , y2)
    dim(doc,  x2, y2 , x4 , y4 , 100,  direction="down")

    # 오른쪽 치수
    dim(doc,   x6 , y6 ,  x7, y7 ,80,  direction="right")         
    dim(doc,  x6, y6 , x5 , y5 , 90,  direction="right")         
    dim(doc,  x5, y5 , x4 , y4 , 140,  direction="right")         
    dim(doc,  x7, y7 , x4 , y4 , 200,  direction="right")        

    # LED BAR 고정 B/K 단면도 삽입
    xx = frameXpos  + 2740 + (CD-1500)
    yy = frameYpos + 1418
    insert_block(xx , yy , "lc031_section_rib_ledbar_bracket")       

    # main Frame description    
    x1 = frameXpos + 1350 + (CD-1500)
    y1 = frameYpos + 1230
    textstr = f"Part Name : LED BAR 고정 B/K"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : SPCC 1.6T "    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : {LCD-10} x {str(math.floor((y7-y4)*10)/10)}mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')         

    ######################################################
    # LED 선 고정 B/K
    ######################################################

    rx1 = frameXpos  + 1000
    ry1 = frameYpos + 1418

    thickness = 1.5
    vcut = thickness / 2

    x1 = rx1 
    y1 = ry1
    x2 = x1 + LCD - 10 - 26*2
    y2 = y1    
    x3 = x2
    y3 = y2 + 1.2
    x4 = x3 + 26
    y4 = y3 
    x5 = x4 
    y5 = y4 + 18.8 - thickness
    x6 = x5 
    y6 = y5 + 40 - thickness *2
    x7 = x6 
    y7 = y6 + 18.8 - thickness
    x8 = x7 - 26
    y8 = y7 
    x9 = x8 
    y9 = y8 + 1.2
    x10 = x9 - LCD + 10 + 26*2
    y10 = y9 
    x11 = x10 
    y11 = y10 - 1.2
    x12 = x11 - 26
    y12 = y11 
    x13 = x12 
    y13 = y12 - 18.8 + thickness
    x14 = x13 
    y14 = y13 - 40 + thickness *2
    x15 = x14 
    y15 = y14 - 18.8 + thickness
    x16 = x15 + 26
    y16 = y15 

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 16
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
      
    #절곡라인  (히든선 : 정방향)      
    line(doc, x14, y14, x5, y5,  layer='hidden')    # 절곡선      
    line(doc, x13, y13, x6, y6,  layer='hidden')    # 절곡선      

    #  상부 치수 옆으로 이어지는 치수선 95 ~
    positions = calculate_hole_positions_separated_lc031(LCD - 10)
    # print(" calculate_hole_positions_separated_lc031 X좌표:", positions)    
    dim(doc,  x12 + positions[0], y14+18.5 , x12 , y12 , 110, direction="up", option="reverse")            
   
    # 리스트 길이 확인

    for i in range(len(positions) ):
        if i != 0:
            circle_cross(doc,  x12 + positions[i] , y14+18.5 , 3.2, layer='레이져')   
            dimcontinue(doc,   x12 + positions[i] , y14+18.5 )
        if i == 1 :
            baseX = x12 + positions[i]
            baseY = y14+18.5
        if i == 2 :
            baseX1 = x12 + positions[i]
            baseY1 = y14+18.5

    dimcontinue(doc,  x7, y7)
    # 상부 전체치수
    dim(doc,  x12, y12 , x7 , y7 , 155,  direction="up")         
    # 3.2파이 단공 라이트케이스 홀 
    dim(doc,  baseX, y13 , baseX , baseY , 50,  direction="left")  
    dim(doc,  baseX , baseY , baseX, y14 , 80,  direction="right")  
    # 3.2파이 지시선
    dim_leader_line(doc, baseX1, baseY1 , baseX1 + 100 ,baseY1 + 50, f"{len(positions)} -%%C3.2 Hole")   

    # 좌측치수선
    dim(doc,  x12 , y12 , x13 + positions[0] , y14+18.5 ,  100,  direction="left")         
    dim(doc,  x1, y1 , x13 + positions[0] , y14+18.5 ,  100 ,  direction="left")         
    
    # 하부 치수
    dim(doc,  x1, y1 , x15 , y15 , 100,  direction="down", option='reverse')         
    dimcontinue(doc,  x2 , y2)
    dim(doc,  x2, y2 , x4 , y4 , 100,  direction="down")

    # 오른쪽 치수
    dim(doc,   x6 , y6 ,  x7, y7 ,80,  direction="right")         
    dim(doc,  x6, y6 , x5 , y5 , 90,  direction="right")         
    dim(doc,  x5, y5 , x4 , y4 , 140,  direction="right")         
    dim(doc,  x7, y7 , x4 , y4 , 200,  direction="right")        

    # LED BAR 고정 B/K 단면도 삽입
    xx = frameXpos  + 2740 + (CD-1500)
    yy = frameYpos + 1418
    insert_block(xx , yy , "lc031_section_rib_ledbar_bracket")       

    # main Frame description    
    x1 = frameXpos + 1350 + (CD-1500)
    y1 = frameYpos + 1230
    textstr = f"Part Name : LED BAR 고정 B/K"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : SPCC 1.6T "    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : {LCD-10} x {str(math.floor((y7-y4)*10)/10)}mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')         

    ######################################################
    # LED 선 고정 B/K
    ######################################################

    rx1 = frameXpos  + 400
    ry1 = frameYpos + 560

    thickness = 1.2
    vcut = thickness / 2

    x1 = rx1 
    y1 = ry1
    x2 = x1 + CW - 200
    y2 = y1    
    x3 = x2
    y3 = y2 + 20 - thickness
    x4 = x3
    y4 = y3 + 20 - thickness
    x5 = x4 -  CW + 200
    y5 = y4 
    x6 = x5 
    y6 = y5 - 18.8 + thickness

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 6
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
      
    #절곡라인  (히든선 : 정방향)      
    line(doc, x6, y6, x3, y3,  layer='hidden')    # 절곡선      

    #  225 300 350 300 255 간격계산
    positions = calculate_ledbar_bottomhole_lc031(CW-200)    
    # print(" calculate_hole_positions_bottom_lc031 X좌표:", positions)    
    dim(doc,  x1 + positions[0], y1+10 , x1 , y1 , 150, direction="down", option="reverse")            
   
    for i in range(len(positions) ):
        circle_cross(doc,  x1 + positions[i] , y1+10 , 5 , layer='레이져')   
        if i != 0:
            dimcontinue(doc,   x1 + positions[i] , y1+10 )
        if i == 0 :
            baseX = x1 + positions[i]
            baseY = y1+10

    dimcontinue(doc,  x2, y2)

    # 위의 100간격 단공 여러개 표현
    positions_upper = calculate_ledbar_upperhole_lc031(CW-200)        
    # print(" calculate_hole_positions_positions_upper_lc031 X좌표:", positions_upper)    
    dim(doc,  x1 + positions_upper[0], y5-10 , x1 , y5 , 100, direction="down", option="reverse")    
    for i in range(len(positions_upper) ):
        circle_cross(doc,  x1 + positions_upper[i] , y5-10 , 5 , layer='레이져')   
        if i != 0:
            dimcontinue(doc,   x1 + positions_upper[i] , y5-10  )
        if i == 0 :
            baseXupper = x1 + positions_upper[i]
            baseYupper = y5-10 

    dimcontinue(doc,  x2, y2)
    # 상부 전체치수
    dim(doc,  x5, y5 , x4 , y4 , 120,  direction="up")         

    # 5파이 단공
    dim(doc,  x5, y5 , baseXupper , baseYupper , 100,  direction="left")  
    dim(doc,  x1, y1 , baseX , baseY , 100,  direction="left")  

    positions_sum = positions + positions_upper
    # 5파이 지시선
    dim_leader_line(doc, baseX, baseY , baseX + 100 ,baseY + 50, f"{len(positions_sum)} -%%C5 Hole")   
    
    # 오른쪽 치수
    dim(doc,   x4 , y4 ,  x3, y3 ,80,  direction="right")         
    dim(doc,  x2, y2 , x3 , y3 , 50,  direction="right")         
    dim(doc,  x4, y4 , x2 , y2 , 150,  direction="right")        

    # LED BAR wire 고정 B/K 단면도 삽입
    xx = frameXpos  + 2051 + (CW-1600)
    yy = frameYpos + 581
    insert_block(xx , yy , "lc031_section_rib_ledbar_bracket_wire")       

    # main Frame description    
    x1 = frameXpos + 2250 + (CW-1600)
    y1 = frameYpos + 680
    textstr = f"Part Name : LED 선 고정 B/K"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : SPCC 1.2T "    
    draw_Text(doc, x1 , y1 - 40 , 20, str(textstr), layer='0')    
    textstr = f"Size : {CW-200} x {str(math.floor((y4-y2)*10)/10)}mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*1} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')         

    frameXpos = frameXpos + TargetXscale + 1500 if CD<1600 else frameXpos + TargetXscale + 1500 + (CD - 1600)*1.3

    #################################################################################################
    # 4page Light Case Frame AL 알루미늄
    #################################################################################################        
    rx1 = frameXpos  + 340 + LCD + 150
    ry1 = frameYpos + 950
    insert_block(rx1 , ry1 , "lc031_AL_frame_section")    

    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2500 + (CD - 1500) + (CW - 1600) * 1.3 
    TargetYscale = 1530 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    

    # print("3page 스케일 비율 : " + str(frame_scale))           
    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

    rx1 = frameXpos  + 340  

    thickness = 1.2
    vcut = thickness / 2

    x1 = rx1 
    y1 = ry1
    x2 = x1 + LCD - 3
    y2 = y1    
    x3 = x2
    y3 = y2 + 150
    x4 = x3
    y4 = y3 + 30
    x5 = x4 - LCD + 3
    y5 = y4 
    x6 = x5 
    y6 = y5 - 30
    x7 = x6 
    y7 = y6 - 150

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 7
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="0")          
    
    line(doc, x6, y6, x3, y3,  layer='0') 

    # 상부 전체치수
    dim(doc,  x5, y5 , x4 , y4 , 150,  direction="up")            

    # main Frame description    
    x1 = frameXpos + 850 + (CD-1500)
    y1 = frameYpos + 720
    textstr = f"Part Name : Light Case Frame"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : AL "    
    draw_Text(doc, x1 , y1 - 40 , 20, str(textstr), layer='0')    
    textstr = f"Size : 125.35 x 92 x {LCD-3}mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')       

    # car inside box 표기
    xx =  x1 + 800
    yy =  y1 - 180
    rectangle(doc, xx,yy,xx+200*2,yy+100*2,layer='0')      
    line(doc, xx,yy+50*2,xx+200*2,yy+50*2)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50*2, yy+50*2 , 24, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30*2, yy+25*2 , 24, str(textstr), layer='0')          

    frameXpos = frameXpos + TargetXscale + 1500 if CD<1600 else frameXpos + TargetXscale + 1500 + (CD - 1600)*1.3

    #############################################################################################
    # 5page 중판
    #############################################################################################
    abs_x = frameXpos
    abs_y = frameYpos # 첫 페이지의 도면 하단 기준

    # 폴리갈 크기 산출
    polygal_width = CW - 270
    polygal_height = (CD - 46*2 - 240)/2
    
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2200 if(CW<1600) else 2200 + (CW - 1600) * 1.1
    TargetYscale = 1700 if(CD<1600) else 1700 + (CD - 1600) * 0.7
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    
    # print("5page 스케일 비율 : " + str(frame_scale))  
    insert_frame(  frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)           

    # 폴리갈 전개도 기준점 위치
    rx1 = frameXpos + 220
    ry1 = frameYpos + 690

    insideHeight = 17    
    x1 = rx1 
    y1 = ry1
    x2 = rx1 + polygal_width 
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')      
    insideHeight = 17    
    x1 = rx1 
    y1 = ry1 + polygal_height - 7
    x2 = rx1 + polygal_width 
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='0')    

    insideHeight = polygal_height   
    x1 = rx1 
    y1 = ry1 + 5
    x2 = rx1 + polygal_width 
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='6')            
    # 폴리갈 해치    
    hatch = msp.add_hatch(dxfattribs={'layer': '55'})  # 5는 파란색 코드입니다.

    # "CORK" 패턴으로 해치 패턴 설정
    hatch.set_pattern_fill("CORK", scale=10.0, angle= 0 , color='5')  # 여기서 scale은 패턴의 크기를 조정합니다. 색상지정은 여기서 해야 된다.

    # 경계선 추가 (여기서는 사각형을 예로 듭니다)
    hatch.paths.add_polyline_path(
        [(x1, y1), (x2, y1), (x2, y2), (x1, y2)], 
        is_closed=True
    )

    textstr =  f"폴리갈 몰딩 절단 칫수 - {polygal_width}"
    dim_string(doc, x1, y2+5, x2, y2+5, 140,  textstr,  text_gap=0.07, direction="up")    # 글자키움
    textstr =  f"폴리갈  - {polygal_width}"
    dim_string(doc, x1, y2+5, x2, y2+5, 65,  textstr, text_gap=0.07, direction="up")    

    # 우측 치수선    
    dim(doc, x2, y2 , x2, y2+5,  80, direction = "right")
    dim(doc, x2, y1, x2, y1-5, 80, direction = "right")
    textstr =  f"폴리갈  - {math.ceil(polygal_height)}"    
    dim(doc, x2, y1, x2, y2, 80,  direction="right", text= textstr)      # aligned로 나오지 않을때 역방향 개발    
    
    textstr =  f"폴리갈 몰딩 외경 - {math.ceil(polygal_height+10)}"
    dim(doc, x2, y1-5, x2, y2+5, 140,  direction="right", text= textstr)     # aligned로 나오지 않을때 역방향 개발    

    # 우측 단면도 
    insideHeight = polygal_height   
    rx1 = x2 + 380
    ry1 = ry1 + 5
    x1 = rx1 
    y1 = ry1 
    x2 = rx1 + 6
    y2 = y1 + insideHeight    
    rectangle(doc, x1, y1, x2, y2, layer='2')     

    insert_block(x1 , y1 , "polygal_rib_bottom")  
    insert_block(x1 , y2 , "polygal_rib")  

    # 단면도 치수선
    dim(doc, x2, y2, x2, y2 + 5, 36, direction = "right")
    dim(doc, x1, y1-5, x1, y2 + 5, 140, direction = "left")
    dim(doc,  x1, y2, x2, y2, 72, text_height=0.22, text_gap=0.07, direction="up")
    dim(doc,  x1-1.2, y2+5, x2+1.2, y2+5, 112, text_height=0.22, text_gap=0.07, direction="up")
    
    # description
    insideHeight = polygal_height   
    rx1 = frameXpos + 800
    ry1 = frameYpos + 550
    x1 = rx1 
    y1 = ry1 

    textstr = f"Part Name : 중판"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : 폴리갈 6T"    
    draw_Text(doc, x1 , y1 -40 , 20, str(textstr), layer='0')    
    textstr = f"Size : {math.floor(polygal_width)} x {math.floor(polygal_height)}"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')    

    # car inside box 표기
    xx =  x1 + 800
    yy =  y1 - 200
    rectangle(doc, xx,yy,xx+200*2,yy+100*2,layer='0')      
    line(doc, xx,yy+50*2,xx+200*2,yy+50*2)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50*2, yy+50*2 , 24, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30*2, yy+25*2 , 24, str(textstr), layer='0')        

    frameXpos = frameXpos + TargetXscale + 1500 if CD<1600 else frameXpos + TargetXscale + 1500 + (CD - 1600)*1.3

    #################################################################################################
    # 6page 중판 보강, 센터 미러판
    #################################################################################################        
         
    # 도면틀 넣기
    BasicXscale = 2813
    BasicYscale = 1765
    TargetXscale = 2940 + (CW - 1600)
    TargetYscale = 1792 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale    

    # print("6page 스케일 비율 : " + str(frame_scale))           
    insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

    rx1 = frameXpos  + 735
    ry1 = frameYpos + 1670
    
    thickness = 1.2
    vcut = thickness / 2

    # 중판 전개도 크기 정의
    midplate_width = CW - 270
    midplate_su = 2

    topwing = 15
    midbridge = 30
    bottomwing = 15
    
    # 중판 width - 40 적용 날개 간섭등 공차 차감

    x1 = rx1 
    y1 = ry1
    x2 = x1 + midplate_width
    y2 = y1    
    x3 = x2
    y3 = y2 + bottomwing - thickness
    x4 = x3
    y4 = y3 + midbridge - thickness * 2
    x5 = x4 
    y5 = y4 + topwing - thickness
    x6 = x5 - midplate_width
    y6 = y5 
    x7 = x6 
    y7 = y6 - topwing + thickness
    x8 = x7 
    y8 = y7 - midbridge + thickness * 2

    # 전개도 사이즈 저장 #1 중판
    midplatesizex = x2 - x1
    midplatesizey = y5 - y1

    # 중판1보다 4점 추가됨 주의
    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 8
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
      
    #절곡라인  밴딩 가로 방향      
    line(doc, x8, y8, x3, y3,  layer='hidden')   # 절곡선      
    line(doc, x7, y4, x4, y4,  layer='hidden')   # 절곡선      
   
    # 상부 치수선
    dim(doc, x6, y6, x5, y5 , 100 , direction="up")

    # 우측 치수선
    dim(doc, x5, y5, x4, y4 , 40 , text_height=0.12, direction="right")
    dimto(doc,  x3, y3 , 150 )      
    dimto(doc,  x2, y2 , 80 )  
    dim(doc, x5, y5, x2, y2 , 210 , direction="right")
    
    # 우측 단면도 그리기    
    # 'ㄷ'자 형상 

    startx = x1 + midplate_width + 350
    starty = y1 + (topwing + bottomwing)

    x1 = startx
    y1 = starty
    x2 = x1
    y2 = y1 - thickness    
    x3 = x2 - bottomwing 
    y3 = y2 
    x4 = x3 
    y4 = y3 + midbridge
    x5 = x4 + topwing
    y5 = y4 
    x6 = x5 
    y6 = y5 - thickness
    x7 = x6 - topwing + thickness
    y7 = y6 
    x8 = x7 
    y8 = y7 - midbridge + thickness*2

    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 8
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="0")     

    # 단면도 치수선
    dim(doc, x3,y3, x2, y2, 100,  text_height=0.12, direction="down")
    dim(doc, x4,y4, x3, y3, 50,  text_height=0.12, direction="left")    
    dim(doc, x4,y4, x5, y5, 60, text_height=0.12,  direction="up")    

    # 중판 보강 description    
    x1 = frameXpos + 1040 + (midplate_width-1330)
    y1 = frameYpos + 1500 
    textstr = f"Part Name : 중판 보강(샤링 품목)"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : EGI 1.2T"
    draw_Text(doc, x1 , y1 - 40 , 20, str(textstr), layer='0')        
    textstr = f"Size : {math.floor(midplatesizex)} x {math.floor(midplatesizey*10)/10}mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su*2} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')       

    #####################################
    # 6page 센터 미러판
    #####################################

    rx1 = frameXpos  + 450
    ry1 = frameYpos + 500
    
    thickness = 1.2
    vcut = thickness / 2

    # 중판 전개도 크기 정의
    midplate_width = CW - 312
    midplate_su = 2
    
    midbridge = 250
    
    x1 = rx1 
    y1 = ry1
    x2 = x1 + midplate_width
    y2 = y1    
    x3 = x2
    y3 = y2 + midbridge
    x4 = x3 - midplate_width
    y4 = y3 

    # 전개도 사이즈 저장 #1 중판
    midplatesizex = x2 - x1
    midplatesizey = y4 - y1

    # 중판1보다 4점 추가됨 주의
    prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
    lastNum = 4
    for i in range(1, lastNum + 1):
        curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
        line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
        prev_x, prev_y = curr_x, curr_y
    line(doc, prev_x, prev_y, x1, y1, layer="레이져")            
 
    # 상부 치수선
    dim(doc, x4, y4, x3, y3 , 100 , direction="up")

    # 우측 치수선
    dim(doc, x3, y3, x2, y2 , 80 , direction="right")

    rx1 = x2 + 250
    ry1 = y1
    insert_block(rx1 , ry1 , "lc031_midplate_section")      

    # 센터 미러판 description    
    x1 = frameXpos + 840 + (midplate_width-1330)
    y1 = frameYpos + 600 
    textstr = f"Part Name : 센터 미러판"    
    draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
    textstr = f"Mat.Spec : STS M/R 1.2T"
    draw_Text(doc, x1 , y1 - 40 , 20, str(textstr), layer='0')        
    textstr = f"Size : {math.floor(midplatesizex)} x {math.floor(midplatesizey)}mm"    
    draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
    textstr = f"Quantity : {su} EA"    
    draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')             

    frameXpos = frameXpos + TargetXscale + 1500 if CD<1600 else frameXpos + TargetXscale + 1500 + (CD - 1600)*1.3

    ####################################################################################################################################################
    # 레이져 도면틀 생성 (공통) 레이져에 배열할때 편리한 기능
    # 1219*1950 철판크기 샘플
    x1 = 10000+850
    y1 = 5000+130
    x2 = 10000+850+2438
    y2 = y1 +1219    
    x3 = 10000+850+2438+850+2438+ 1850  
    y3 = y1 +1219
    rectangle(doc, 10000, 5000,10000+13400 , 5000+ 2340 , layer='0')  
    rectangle(doc,x1 , y1 ,x1 + 2438 , y2 , layer='레이져')  
    dim(doc, x1 , y2 , x2 , y2,  100, direction="up")
    # rectangle(doc, x1+2600 , y1 , x2+2600 , y2 , layer='레이져')  
    # dim(doc, x1+2600, y2, x2+2600, y2, 100, direction="up")
    # dim(doc,  x1 , y1, x1 ,y2 , 150, direction="left")
    rectangle(doc, x3 , y1 , x3+2600 , y2 , layer='레이져')  
    dim(doc, x3, y3, x3+2600, y2, 100, direction="up")
    dim(doc,  x3 , y1, x3 ,y2 , 150, direction="left")

    textstr = f"{secondord} - {workplace} - {drawdate_str_short} 출도, {deadlinedate_str_short} 납기"    
    draw_Text(doc, 10000+300 , 5000+ 2340 - 200 ,  120, str(textstr), layer='0')        

    y = 5000+ 2340 - 650
    textstr = f"SPCC 1.2tX1219X = 장"
    draw_Text(doc, 10000+1000 ,y ,  90, str(textstr), layer='0')        
    textstr = f"SPCC 1.6T"        
    draw_Text(doc, 10000+1000+2600 , y ,  90, str(textstr), layer='0')            
    textstr = f"EGI 1.2T"        
    draw_Text(doc, 10000+1000+5200 ,y ,  90, str(textstr), layer='0')        
    textstr = f"{LCplateMaterial}"
    draw_Text(doc, 10000+1000+5200+2600 , y ,  90, str(textstr), layer='0')            
   
####################################################################################################################################################################################
# lc008_A 천장 자동작도
####################################################################################################################################################################################
def lc008_A():   
    global args  #수정할때는 global 선언이 필요하다. 단순히 읽기만 할때는 필요없다.    
    global start_time
    global workplace, drawdate, deadlinedate, lctype, secondord, text_style, exit_program, Vcut_rate, Vcut, program_message, formatted_date, text_style_name
    global CW_raw, CD_raw, panel_thickness, su, LCframeMaterial, LCplateMaterial
    global input_CD, input_CW,CD,CW,drawdate_str , deadlinedate_str, drawdate_str_short, deadlinedate_str_short
    global doc, msp , T5_is, text_style, distanceXpos, distanceYpos

    # 2page 덴크리 다온텍등 등기구 설명 그림
    insert_block(0 , 0 , "lc008_A_2pageframe")        
    abs_x = 0
    abs_y = 0
    # LC 크기 지정 car inside에서 차감
    LCD = CD-50
    LCW = CW-50  # 라이트 케이스 크기를 정한다. 보통은 50

    # watt 계산 공식 적용해야 함
    # LED바 기장(단위:m) X 수량 X 15 = 60W 이하
    # LED바 1개당 watt 산출 m단위로 계산.. /1000  031모델은 3개의 led바가 들어감     
        
    ledsu = 4    

    if(CD>2000):
        ledsu = 4

    wattCalculate = math.ceil((LCD - 3)/1000 * ledsu * 15)

    print("wattCalculate: " + str(wattCalculate))  

    if(wattCalculate <= 60):
        watt = 60
    elif(wattCalculate > 60 and wattCalculate < 100):
        watt = 100
    elif(wattCalculate >= 100 and wattCalculate < 150):
        watt = 150
    elif(wattCalculate >= 150 and wattCalculate < 200):
        watt = 200
    elif(wattCalculate >= 200 and wattCalculate < 200):
        watt = 300
                
    # print("L/C 1개당 규격설정 wattCalculate: " + str(watt))  

    T5_standard = 0
    if T5_is == '있음':
        adjusted_CD = CD - 50  # CD에서 50 뺌

        # 규격 결정
        if adjusted_CD > 1500:
            T5_standard = 1500
        elif adjusted_CD > 1200:
            T5_standard = 1200
        else:
            T5_standard = 900

        # print(f"선택된 규격: {T5_standard}")
    # else:
    #     print("T5_is 가 '있음'이 아니므로 규격 선택 안 함")    

    # 도면틀 넣기
    BasicXscale = 2560
    BasicYscale = 1550
    TargetXscale = 800 
    TargetYscale = 300 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale

    frame_scale = 1    
    ########################################################
    # 갑지 1
    ########################################################
    # 현장명
    textstr = workplace       
    x = abs_x + 80
    y = abs_y + 1000 + 218
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 발주처
    textstr = secondord
    x = abs_x + 80
    y = abs_y + 1000 + 218 - 30
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 타입
    textstr = lctype
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 수량
    textstr = su
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 카사이즈
    textstr = f"W{CW} x D{CD}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 색상 LC
    textstr = f"L / C : 흑색무광"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')   

    textstr = f"중판 : STS MR 1.2t"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18 - 12
    draw_Text(doc, x, y , 7, str(textstr), layer='0')

    # 등기구 업체
    LedBarLength = 0
    if T5_standard > 0 : 
        Ledcompany = f"(등기구업체 : 엘엠팩트)"        
        Leditem = "T5"
        Ledwatt = "6500K"
        LedBarLength_CD = T5_standard
        LedBarLength_CW = T5_standard
        watt = "내장"
    else:
        Ledcompany = f"(등기구업체 : 다온텍)"
        Leditem = "LED BAR"
        Ledwatt = "10000K"
        LedBarLength_CD = CalculateLEDbar(CD-70)
        LedBarLength_CW = CalculateLEDbar(CW-170)
        # print (f"035 LedBarLength_CD : {LedBarLength_CD}")
        # print (f"035 LedBarLength_CW : {LedBarLength_CW}")
        watt = f"{str(watt)}W"

    x = abs_x + 80
    y = abs_y + 1000 + 120
    draw_Text(doc, x, y , 8, str(Ledcompany), layer='0')     
    textstr = f"(할로겐 전구색 3000K (스텐링)일체형-다온텍)"
    x = abs_x + 80
    y = abs_y + 1000 + 120 - 15
    draw_Text(doc, x, y , 8, str(textstr), layer='0')
    
    altube_CW = CW-200
    altube_CD = CD-250

    textstr = f"*. 출도일자 : {drawdate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 
    draw_Text(doc, x, y , 8, str(textstr), layer='0')    
    textstr = f"*. 납품일자 : {deadlinedate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 - 15
    draw_Text(doc, x, y , 8, str(textstr), layer='0')

    ################################################################################################################################################
    # 갑지 2
    ################################################################################################################################################
    # 현장명
    textstr = f"* 현장명 : {secondord}-{workplace} - {su}(set)"
    x = abs_x + 30
    y = abs_y + 650 
    draw_Text(doc, x, y , 20, str(textstr), layer='0')        
    x = abs_x + 30 + 200
    y = abs_y + 600 
    draw_Text(doc, x, y , 20, str(Ledcompany), layer='0')    
    
    textstr = f"{watt}"
    x = abs_x + 835
    y = 395 + 110
    draw_Text_direction(doc, x + 50, y - 30 , 20, str(textstr), layer='0', rotation = 0)    

    textstr = f"LED BAR - 3000K -{altube_CW}L" #상
    x = abs_x + 410
    y = 550
    draw_Text(doc, x, y , 12, str(textstr), layer='0')
    
    textstr = f"(누드형 클립형 잭타입)" #상
    x = abs_x + 410 +10
    y = 530
    draw_Text(doc, x, y , 12, str(textstr), layer='0')

    textstr = f"LED BAR - 3000K -{altube_CW}L" #하
    x = abs_x + 410
    y = 544-460
    draw_Text(doc, x, y , 12, str(textstr), layer='0')
    
    textstr = f"(누드형 클립형 잭타입)" #하
    x = abs_x + 410 +10
    y = 524-460
    draw_Text(doc, x, y , 12, str(textstr), layer='0')

    x = abs_x + 240
    y = 230    
    textstr = f"LED BAR - 3000K -{altube_CD}L" #좌
    draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)        

    x = abs_x + 260
    y = 240    
    textstr = f"(누드형 클립형 잭타입)" #좌
    draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)        
       
    x = abs_x + 720
    y = 220    
    textstr = f"LED BAR - 3000K -{altube_CD}L" #우
    draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)        

    x = abs_x + 740
    y = 230    
    textstr = f"(누드형 클립형 잭타입)" #우
    draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)   
                   
    frameXpos = 1437

    # Assy Car case    
    #####################################################################################################################################
    # 1page 조립도 Assy
    ##################################################################################################################################### 
    # LC ASSY 상부 형상    
    
    ###################################
    # 1page Assy 상부 단면도 작도
    ###################################
    program = 1
    if program:
        abs_x = frameXpos + 700    
        abs_y = abs_y + CD + 500
        insert_block(abs_x, abs_y, "lc008_A_top_left")    

        x = abs_x + CW 
        y = abs_y 
        insert_block(x, y, "lc008_A_top_right")  
        
        x = abs_x + CW/2 
        y = abs_y 
        insert_block(x, y, "lc008_A_center")  
        
        x = abs_x + (CW-100)/4 + 50
        y = abs_y 
        insert_block(x, y, "lc008_A_Halogen")
        
        x = abs_x + CW-((CW-100)/4 + 50)
        insert_block(x, y, "lc008_A_Halogen")
        
        first_space = 135

        # 흰색선
        line(doc, abs_x + 50, abs_y, abs_x + CW - 100 + 50, abs_y, layer='0')
        line(doc, abs_x + 50 + 1.2, abs_y + 1.2, abs_x + CW - 100 + 50 -2.4, abs_y + 1.2, layer='0')
        line(doc, abs_x + 100, abs_y + 30, abs_x + CW - 100, abs_y + 30, layer='0')

        # 상부선 2.3mm 간격
        x1 = abs_x - 25
        y1 = abs_y + 100
        x2 = x1 + CW + 50
        y2 = y1
        rectangle(doc, x1, y1, x2, y2+2.3, layer='0')   
    
        x1 = abs_x
        y1 = abs_y + 100
        x2 = abs_x + CW
        y2 = y1    
        textstr =  f"Car Inside - {CW}"      
        dim(doc, x1, y1, x2, y2, 197, text_height=0.30, text_gap=0.07, direction="up" , text = textstr)    

        # 단면도 하부 치수
        x1 = abs_x + 50
        y1 = abs_y 
        x2 = abs_x + CW - 50
        y2 = y1    
        dim(doc, x1, y1, x2, y2, 240,  text_height=0.30, text_gap=0.07, direction="down")
        
        # 단면도 하부 치수/2
        x1 = abs_x + 50
        y1 = abs_y 
        x2 = abs_x + CW/2
        y2 = y1    
        dim(doc, x1, y1, x2, y2, 120,  text_height=0.50, direction="down")
        
        x1 = abs_x + CW/2
        y1 = abs_y 
        x2 = abs_x + CW -50
        y2 = y1    
        dim(doc, x1, y1, x2, y2, 120,  text_height=0.50, direction="down")

    ###################################
    # 1page Assy 본판 작도
    ###################################
    program = 1
    if program:    
        # 기본 윤곽 car inside
        width_insideGap = 60
        height_insideGap = 60
        abs_y = abs_y - CD - 800    
        x1 = abs_x 
        y1 = abs_y 
        x2 = abs_x + CW
        y2 = abs_y + CD    
        rectangle(doc, x1, y1, x2, y2, layer='2')    

        # 상하 30 철판 표현
        ribGap = 30
        # x방향 CW와 LC 떨어진 gap
        Xgap = 60
        x1 = abs_x + 90
        y1 = abs_y + height_insideGap
        x2 = abs_x + CW - 90
        y2 = y1+ribGap
        rectangle(doc, x1, y1, x2, y2, layer='0')  #상      
        y1 = abs_y + CD - height_insideGap    
        y2 = y1-ribGap
        rectangle(doc, x1, y1, x2, y2, layer='0')  #하

        # 양쪽 테두리 30 흰색선
        x1 = abs_x + 60
        y1 = abs_y + height_insideGap        
        x2 = x1 + 30
        y2 = abs_y + CD - height_insideGap
        rectangle(doc, x1, y1, x2, y2, layer='0') 
        x1 = abs_x + CW - 60 - ribGap
        x2 = x1 + ribGap
        rectangle(doc, x1, y1, x2, y2, layer='0') 
        
        # 상하좌우1.6T hidden 선
        line(doc, abs_x + 60 + 30 -1.6, abs_y + 60, abs_x + 60 + 30 -1.6, abs_y + CD - 60, layer="Hidden") #좌
        line(doc, abs_x + CW - 60 - 30 + 1.6, abs_y + 60, abs_x + CW - 60- 30+ 1.6, abs_y + CD - 60, layer="Hidden") #우
        line(doc, abs_x + 90, abs_y + CD - 90 + 1.6, abs_x + CW - 90, abs_y + CD - 90 + 1.6, layer="Hidden") #상
        line(doc, abs_x + 90, abs_y + 90 - 1.6, abs_x + CW - 90, abs_y + 90 - 1.6, layer="Hidden") #하
        
        #내부 가로폭치수
        dim(doc, abs_x + 90 - 1.6, abs_y + CD - 90 + 1.6, abs_x + CW - 90 + 1.6, abs_y + CD - 90 + 1.6, 120, text_height=0.50, direction="down")

        # LC 홀 가공 (기존 본천장 로직 적용해서 만들것)
        inputCW = CW
        inputCD = CD

        lc_positions = calculate_lc_position_008_A(inputCW, inputCD)

    # lc_positions = {
    #     'Topholex': Topholex,
    #     'Topholey': Topholey,
    #     'Middlex': Middlex,
    #     'Middley': Middley,
    #     'Bottomx': Bottomx,
    #     'Bottomy': Bottomy
    # }        
        # Topholex와 Topholey 좌표 추출
        Topholex = lc_positions['Topholex']
        Topholey = lc_positions['Topholey']

        # middle 좌표 추출
        Middlex = lc_positions['Middlex']
        Middley = lc_positions['Middley']    

        # Bottomx와 Bottomy 좌표 추출
        Bottomx = lc_positions['Bottomx']
        Bottomy = lc_positions['Bottomy']    

        # 필요한 경우 다른 홀들에 대해서도 circle_cross 함수를 호출할 수 있음
        # 예를 들어, 모든 상단 홀에 대해 반복
        for x in Topholex:
            circle_cross(doc, abs_x + x, abs_y + Topholey, 11, layer='0')    
            # print(f"topholex x좌표: {x} y좌표 {abs_y + Topholey} abs_y좌표 : {abs_y}  Topholey : {Topholey}")     

        for x in Middlex:
            circle_cross(doc, abs_x + x, abs_y + Middley, 11, layer='0')            

        for x in Bottomx:
            circle_cross(doc, abs_x + x, abs_y + Bottomy, 11, layer='0')            

        # 상부 치수선
        x1 = abs_x 
        y1 = abs_y + CD
        x2 = abs_x + CW
        y2 = y1    
        textstr =  f"CAR INSIDE(CW) = {CW}"
        dim(doc, x1, y1, x2, y2, 300, text_height=0.30, direction="up", text = textstr)    

        x1 = abs_x 
        y1 = abs_y + CD
        x2 = abs_x + CW
        y2 = y1    
        dim(doc, x1 + Xgap, y1 - Xgap, x2 - Xgap, y2 - Xgap, 290, direction="up") 
        
        dim(doc, x1 + Xgap, y2 - Xgap, x1, y2, 210, direction="up", option="reverse")    
        dimcontinue(doc, x1 + CW/2, y2 - Xgap)    
        dimcontinue(doc, x2 - Xgap, y2 - Xgap)    
        dimcontinue(doc, x2, y2)
        
        dim(doc, x1 + 75, y1 - 75, x1, y1, 140, direction="up", option="reverse")    
        dimcontinue(doc, x1 + CW/2, y2 - 75)    
        dimcontinue(doc, x2 - 75, y2 - 75)    
        dimcontinue(doc, x2, y2)
        
        # 우측 60 간격 치수선 표기
        x1 = abs_x + CW 
        y1 = abs_y + CD 
        x2 = x1 - Xgap
        y2 = y1 - 60
        dim(doc, x1, y1, x2, y2, 115, text_height=0.30,  direction="right")

        x1 = abs_x + CW - Xgap
        y1 = abs_y + 60 
        x2 = abs_x + CW
        y2 = abs_y
        dim(doc, x1, y1, x2, y2, 200, text_height=0.30,  direction="right")

        # 상단, 하단 홀의 오른쪽 끝 y 좌표 추출
        right_end_topholey = Topholey
        right_end_bottomy = Bottomy

        # 세로 방향 홀의 y 좌표 추출

        vertical_holey = [Middley]
        vertical_holes_rightholey = vertical_holey

        # 모든 오른쪽 끝 홀의 y 좌표를 하나의 리스트로 결합
        all_right_end_holes_y_coords = [right_end_topholey] + vertical_holes_rightholey + [right_end_bottomy]

        # 모든 오른쪽 끝 홀의 y 좌표를 오름차순으로 정렬
        sorted_y_coords = sorted(all_right_end_holes_y_coords)

        # 정렬된 y 좌표 출력
        # 좌측 치수
        last_ypos = 75
        dim(doc, abs_x, abs_y, abs_x + 60, abs_y + last_ypos, 120, text_height=0.30, direction="left") 
        for i, y_coord in enumerate(sorted_y_coords):        
            # print(f"세로 홀의 y 좌표: {y_coord}")
            if i>0:
                dimcontinue(doc, abs_x + 60, abs_y + y_coord)
                last_ypos = y_coord
        dimcontinue(doc, abs_x, abs_y + CD, distance = 120, option='reverse')             
        
        dim(doc, abs_x + 60, abs_y + 60, abs_x + 60 , abs_y + CD - 60, 340, text_height=0.30,  direction="left") 
        dim(doc, abs_x, abs_y, abs_x + 60, abs_y + 60, 195, text_height=0.30,  direction="left") 
        dimcontinue(doc, abs_x + 60, abs_y + CD/2) 
        dimcontinue(doc, abs_x + 60, abs_y + CD - 60) 
        dimcontinue(doc, abs_x, abs_y + CD) 
        
        x1 = abs_x + CW - Xgap
        y1 = abs_y + CD - 60
        x2 = x1
        y2 = abs_y + 60        
        textstr =  f"LIGHT CASE = {CD-60*2}"
        dim(doc, x2, y2, x1, y1, 200, text_height=0.32, direction="right", text = textstr)    
        x1 = abs_x + CW
        y1 = abs_y + CD
        x2 = x1
        y2 = abs_y         
        textstr =  f"CAR INSIDE(CD) = {CD}"    
        dim(doc,  x2, y2, x1, y1,  210, text_height=0.32, direction="right", text = textstr)     
    #########################################################################
    # 1page 우측단면도 표기 (블럭 삽입)
    #########################################################################
    program = 1
    if program:
        x = abs_x + CW + 450
        y = abs_y + CD
        insert_block(x, y, "lc008_A_assy_side_section_top")         
        x1 = x
        y1 = abs_y 
        insert_block(x1, y1, "lc008_A_assy_side_section_bottom")         
        x2 = x + 70
        y2 = abs_y + CD/2 
        insert_block(x2, y2, "lc008_A_assy_side_section_midbolt")      

        rectangle(doc, x - 2.3, y + 25, x, y1 - 25, layer="0")

        # 흰색선 삽입
        line(doc, x + 100, y - 50, x + 100, y1 + 50, layer="0")
        line(doc, x + 100 - 1.2, y - 50 - 1.2,  x + 100 - 1.2, y1 + 50 + 1.2, layer="0")
        line(doc, x + 70,  y - 90,  x + 70, y1 + 90, layer="0")
        
        dim(doc, x + 100 , y - 50,  x + 100, y1 + CD/2,  300, text_height=0.3, direction="right")
        dim(doc, x + 100 , y1 + 50,  x + 100, y1 + CD/2,  300, text_height=0.3, direction="right")

        textstr = f"Car Inside (CD) = {CD}"
        dim(doc, x, y, x1 , y1,  100, text_height=0.3, direction="left", text=textstr)
  
    #################################
    # 1page 우측 의장면 Assy 본판
    #################################    
    program = 1
    if program:    
        # 기본 윤곽 car inside
        y = abs_y
        
        x1 = x + 600
        x2 = x1 + CW
        y1 = y
        y2 = y1 + CD
        textstr =  f"CAR INSIDE(CW) = {CW}"    
        dim(doc,  x1, y2, x2, y2,  150, text_height=0.30, direction="up", text = textstr)                 
        textstr =  f"CAR INSIDE(CD) = {CD}"    
        dim(doc,  x2, y1, x2, y2,  160, text_height=0.30, direction="right", text = textstr)                 
        
        width_insideGap = 50
        height_insideGap = 50
        y1 = abs_y
        x2 = x1 + CW
        y2 = abs_y + CD    
        rectangle(doc, x1, y1, x2, y2, layer='2')
        
        x1 = x1 + width_insideGap
        y1 = y1 + height_insideGap
        x2 = x1 + CW - width_insideGap*2
        y2 = y1 + CD - height_insideGap*2
        rectangle(doc, x1, y1, x2, y2, layer='0')
                        
        ICW = CW - width_insideGap*2 #의장면 가로폭
        ICD = CD - height_insideGap*2 #의장면 세로폭
        
        #센터 세로선 3줄
        line(doc, x1 + ICW/2, y1, x1 + ICW/2, y2, layer="Hidden")
        line(doc, x1 + ICW/2-1.2, y1, x1 + ICW/2-1.2, y2, layer="Hidden")
        line(doc, x1 + ICW/2+1.2, y1, x1 + ICW/2+1.2, y2, layer="Hidden")
        
        #센터 가로줄
        line(doc, x1,  y1 + ICD/2, x2, y1 + ICD/2, layer="Hidden") 
        line(doc, x1 + 1.2,  y1 + ICD/2+1.2, x1 + ICW/2-1.2, y1 + ICD/2+1.2, layer="Hidden") #센터 위
        line(doc, x2-1.2,  y1 + ICD/2+1.2, x1 + ICW/2+1.2, y1 + ICD/2+1.2, layer="Hidden")
        line(doc, x1 + 1.2,  y1 + ICD/2-1.2, x1 + ICW/2-1.2, y1 + ICD/2-1.2, layer="Hidden") #센터 아래
        line(doc, x2-1.2,  y1 + ICD/2-1.2, x1 + ICW/2+1.2, y1 + ICD/2-1.2, layer="Hidden")
        
        #테두리
        rectangle(doc, x1 + 26, y2 - 25, x1 + ICW/2-1.2, y2 - 1.2, layer='Hidden')#상좌
        rectangle(doc, x2 - 26, y2 - 25, x1 + ICW/2+1.2, y2 - 1.2, layer='Hidden')#상우
        rectangle(doc, x1 + 26, y1 + 25, x1 + ICW/2-1.2, y1 + 1.2, layer='Hidden')#하좌
        rectangle(doc, x2 - 26, y1 + 25, x1 + ICW/2+1.2, y1 + 1.2, layer='Hidden')#하우
        
        rectangle(doc, x1,  y1 , x1 + 25, y2, layer='Hidden')#좌
        rectangle(doc, x2 ,  y1, x2 - 25, y2, layer='Hidden')#우
        line(doc, x1 + 1.2,  y1, x1 + 1.2, y2, layer="Hidden")#좌
        line(doc, x2 - 1.2,  y1, x2 - 1.2, y2, layer="Hidden")#우
        
        insert_block(x1 + ICW/4, y1 + ICD/4, "lc008_A_Halogen_S")    
        insert_block(x1 + ICW/4, y2 - ICD/4, "lc008_A_Halogen_S")    
        insert_block(x2 - ICW/4, y1 + ICD/4, "lc008_A_Halogen_S")    
        insert_block(x2 - ICW/4, y2 - ICD/4, "lc008_A_Halogen_S")    
        
        # LC 홀 가공 (기존 본천장 로직 적용해서 만들것)
        
        ICW = CW - width_insideGap*2 #의장면 가로폭
        ICD = CD - height_insideGap*2 #의장면 세로폭
        
        inputCW = ICW
        inputCD = ICD

        lc_positions = calculate_lc_position_008_A_CW(inputCW, inputCD)

        print(f" 1page inputCW :{ inputCW}  inputCD : {inputCD}lc_positions : {lc_positions} ")
    # lc_positions = {
    #     'Topholex': Topholex,
    #     'Topholey': Topholey,
    #     'Bottomx': Bottomx,
    #     'Bottomy': Bottomy
    # }        

        # Topholex와 Topholey 좌표 추출
        Topholex = lc_positions['Topholex']
        Topholey = lc_positions['Topholey']

        # Bottomx와 Bottomy 좌표 추출
        Bottomx = lc_positions['Bottomx']
        Bottomy = lc_positions['Bottomy']    
        
        temp_x = x1
        temp_y = y1
        
        for x in Topholex:
            circle_cross(doc, temp_x + x, temp_y + Topholey, 9, layer='0')    
            circle_cross(doc, temp_x + x, temp_y + Bottomy, 9, layer='0')          
            
        temp_x = x1 + ICW/2
        temp_y = y1
        
        for x in Topholex:
            circle_cross(doc, temp_x + x, temp_y + Topholey, 9, layer='0')    
            circle_cross(doc, temp_x + x, temp_y + Bottomy, 9, layer='0')            

        lc_positions = calculate_lc_position_008_A_CD(inputCW, inputCD)

        # 'Leftx': Leftx,
        # 'Lefty': Lefty,
        # 'Rightx': Rightx,
        # 'Righty': Righty
        
        Leftx = lc_positions['Leftx']
        Lefty = lc_positions['Lefty']

        Rightx = lc_positions['Rightx']
        Righty = lc_positions['Righty']    
        
        temp_x = x1
        temp_y = y1 + ICD/2

        # print(f"temp_x: {temp_x} temp_y: {temp_y} ")        
        
        for y in Lefty:
            circle_cross(doc, temp_x + Leftx , temp_y + y, 9, layer='0')    
            circle_cross(doc, temp_x + Rightx , temp_y + y, 9, layer='0')            
        
        
        temp_x = x1
        temp_y = y1

        # print(f"temp_x: {temp_x} temp_y: {temp_y} ")        
        
        for y in Lefty:
            circle_cross(doc, temp_x + Leftx , temp_y + y, 9, layer='0')    
            circle_cross(doc, temp_x + Rightx , temp_y + y, 9, layer='0')            
            
        
        #의장면 조립홀 간격별 선 그리기
        lc_positions = calculate_lc_position_008_A_ASSY_horizontal(inputCW, inputCD)
        
        linex = lc_positions['linex']
        liney = lc_positions['liney']
        
        temp_x = x1
        temp_y = y1

        for x in linex:
            insert_block(temp_x + x , temp_y + liney , "lc008_A_ASSY_line")    
            
        temp_x = x1 + ICW/2
        temp_y = y1
        
        for x in linex:
            insert_block(temp_x + x , temp_y + liney , "lc008_A_ASSY_line")        
        
        lc_positions = calculate_lc_position_008_A_ASSY_vertical(inputCW, inputCD)
        
        linex = lc_positions['linex']
        liney = lc_positions['liney']
        
        temp_x = x1
        temp_y = y1

        for y in liney:
            insert_block(temp_x + linex , temp_y + y , "lc008_A_ASSY_line_")    
            
        temp_x = x1
        temp_y = y1 + ICD/2

        for y in liney:
            insert_block(temp_x + linex , temp_y + y , "lc008_A_ASSY_line_")    
        
        # LED BAR 폴리선 좌표(굵은선)
        points = [(x1 + 50, y1 + 60), (x2 - 50, y1 + 60)]
        msp.add_lwpolyline(points, dxfattribs={"const_width": 10, "color": 8})  # 색상 지정 8번 회색
        
        points = [(x1 + 50, y2 - 60), (x2 - 50, y2 - 60)]
        msp.add_lwpolyline(points, dxfattribs={"const_width": 10, "color": 8})  # 색상 지정 8번 회색
        
        points = [(x1 + 60, y1 + 75), (x1 + 60, y2 - 75)]
        msp.add_lwpolyline(points, dxfattribs={"const_width": 10, "color": 8})  # 색상 지정 8번 회색
        
        points = [(x2 - 60, y1 + 75), (x2 - 60, y2 - 75)]
        msp.add_lwpolyline(points, dxfattribs={"const_width": 10, "color": 8})  # 색상 지정 8번 회색
        
        x1 = x1 - width_insideGap
        y1 = y1 - height_insideGap
        x2 = x1 + CW
        y2 = y1 + CD
        
        textstr = f"LED BAR = {CW - 200}"
        dim(doc, x1 + 100, y2 - 110, x2 - 100, y2 - 110, 200, text_height=0.30, direction="up", text = textstr)                 
        textstr = f"LED BAR = {CD - 250}"
        dim(doc, x2 - 100, y1 + 125, x2 - 100, y2 - 125, 200, text_height=0.30, direction="right", text = textstr)                 
        
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1715
        TargetXscale = 5500 + (CW - 1600) * 1.6 + (CD - 1500) * 1.5
        TargetYscale = 3400 + (CD - 1500)
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("스케일 비율 : " + str(frame_scale))        
        frameYpos = abs_y - 450 * frame_scale     
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)      
         
        # 1page 상단에 car inside 표기
        x = frameXpos + CW/3
        y = abs_y + frame_scale*BasicYscale + 50
        textstr = f"W{CW} x D{CD}"    
        draw_Text(doc, x, y , 200*frame_scale, str(textstr), layer='0')    

        frameXpos = frameXpos + TargetXscale + 400

    # Assy Car case    
    #####################################################################################################################################
    # 2page 조립도 Assy 2번째
    ##################################################################################################################################### 
    # LC ASSY 상부 형상        
    
    ###################################
    # 2page Assy 상부 단면도 작도
    ###################################
    program = 1
    if program:
            
        abs_x = frameXpos + 200        
        abs_y = abs_y + CD + 800 
        insert_block(abs_x + 1500, abs_y, "lc008_A_top_left_2")

        x = abs_x + CW + 1500
        y = abs_y 
        insert_block(x, y, "lc008_A_top_right")
        
        x = abs_x + CW/2 + 1500
        y = abs_y 
        insert_block(x, y, "lc008_A_center")
        
        x = abs_x + (CW - 100)/4 + 50 + 1500
        y = abs_y 
        insert_block(x, y, "lc008_A_Halogen")
        
        x = abs_x + CW - ((CW - 100)/4 + 50) + 1500
        insert_block(x, y, "lc008_A_Halogen")
        
        temp_x = abs_x + 1500 
        temp_y = abs_y
        # 흰색선
        line(doc, temp_x + 50, temp_y, temp_x + CW - 50, temp_y, layer='0')
        line(doc, temp_x + 50 + 1.2, temp_y + 1.2, temp_x + CW - 50 - 2.4, temp_y + 1.2, layer='0')
        line(doc, temp_x + 75, temp_y + 30, temp_x + CW - 75, temp_y + 30, layer='0')

        # 상부선 2.3mm 간격
        x1 = temp_x - 25
        y1 = temp_y + 100
        x2 = x1 + CW + 50
        y2 = y1
        rectangle(doc, x1, y1, x2, y2 + 2.3, layer='0')   
    
        x1 = temp_x
        y1 = temp_y + 100
        x2 = temp_x + CW
        y2 = y1    
        textstr =  f"Car Inside (CW) = {CW}"      
        dim(doc, x1, y1, x2, y2, 197, text_height=0.30, text_gap=0.07, direction="up" , text = textstr)    

        # 단면도 하부 치수
        x1 = temp_x + 50
        y1 = temp_y 
        x2 = temp_x + CW - 50
        y2 = y1    
        dim(doc, x1, y1, x2, y2, 240,  text_height=0.30, text_gap=0.07, direction="down")
        
        # 단면도 하부 치수/2
        x1 = temp_x + 50
        y1 = temp_y 
        x2 = temp_x + CW/2
        y2 = y1    
        dim(doc, x1, y1, x2, y2, 120,  text_height=0.50, direction="down")
        
        x1 = temp_x + CW/2
        y1 = temp_y 
        x2 = temp_x + CW -50
        y2 = y1    
        dim(doc, x1, y1, x2, y2, 120,  text_height=0.50, direction="down")

    ###################################
    # 2page Assy
    ###################################
    program = 1
    if program:    
        # 기본 윤곽 car inside
        width_insideGap = 85
        height_insideGap = 85
        abs_x = abs_x + 500    
        abs_y = abs_y - CD - 800    
        x1 = abs_x + 550
        y1 = abs_y 
        x2 = abs_x + CW
        y2 = abs_y + CD    

    #########################################################################
    # 2page 좌측단면도 표기 (블럭 삽입)
    #########################################################################
    program = 1
    if program:
        x1 = abs_x + 100
        y1 = abs_y
        x2 = abs_x + 100
        y2 = abs_y + CD
        insert_block(x1, y2, "lc008_A_assy_side_section_top")         
        insert_block(x1, y1, "lc008_A_assy_side_section_bottom")         
        insert_block(x1 + 70, y1 + CD/2, "lc008_A_assy_side_section_midbolt")      

        rectangle(doc, x1, y1 - 25, x1 - 2.3, y2 + 25, layer="0")

        # 흰색선 삽입
        line(doc, x1 + 100, y1 + 50, x1 + 100, y2 - 50, layer="0")
        line(doc, x1 + 100 - 1.2, y1 + 50 + 1.2, x1 + 100 - 1.2, y2 - 50 - 1.2, layer="0")
        line(doc, x1 + 70, y1 + 90, x1 + 70, y2 - 90, layer="0")

        dim(doc, x1 + 100, y1 + 50, x1 + 100, y1 + CD/2, 300, text_height=0.3, direction="right")
        dimcontinue(doc, x1 + 100, y2 - 50)

        textstr = f"Car Inside (CD) = {CD}"
        dim(doc, x1, y1, x2, y2, 200, text_height=0.3, direction="left", text=textstr)

    #########################################################################
    # 2page 우측단면도 표기 (블럭 삽입)
    #########################################################################
    program = 1
    if program:
        x1 = abs_x + CW*2.1
        y1 = abs_y
        x2 = x1
        y2 = abs_y + CD
        insert_block(x1, y2, "lc008_A_assy_side_section_top_R")         
        insert_block(x1, y1, "lc008_A_assy_side_section_bottom_R")         
        insert_block(x1 - 70, y1 + CD/2, "lc008_A_assy_side_section_midbolt_R")      

        rectangle(doc, x1, y1 - 25, x1 + 2.3, y2 + 25, layer="0")

        # 흰색선 삽입
        line(doc, x1 - 100, y1 + 50, x1 - 100, y2 - 50, layer="0")
        line(doc, x1 - 100 + 1.2, y1 + 50 + 1.2, x1 - 100 + 1.2, y2 - 50 - 1.2, layer="0")
        line(doc, x1 - 70, y1 + 90, x1 - 70, y2 - 90, layer="0")

        dim(doc, x1 - 100, y1 + 50, x1 - 100, y1 + CD/2, 300, text_height=0.3, direction="left")
        dimcontinue(doc, x1 - 100, y2 - 50)

        textstr = f"Car Inside (CD) = {CD}"
        dim(doc, x1, y1, x2, y2, 200, text_height=0.3, direction="right", text=textstr)
    
    #################################
    # 2page 의장면 Assy 본판
    #################################    
    program = 1
    if program:    
        # 기본 윤곽 car inside
        x = abs_x
        y = abs_y
        
        x1 = x + 1000
        x2 = x1 + CW
        y1 = y
        y2 = y1 + CD
        
        #상부 치수
        textstr =  f"CAR INSIDE(CW) = {CW}"    
        dim(doc, x1, y2, x2, y2,  250, text_height=0.30, direction="up", text = textstr)                        
        dim(doc, x1 + 50, y2 - 50,  x1, y2,  250, direction="up", option="reverse")      
        dimcontinue(doc, x2 - 50, y2 - 50)
        dimcontinue(doc, x2, y2)
        dim(doc, x1 + 50, y2 - 50,  x1 + CW/2, y2 - 50,  200, direction="up")      
        dimcontinue(doc, x2 - 50, y2 - 50)
        #좌측 치수
        textstr =  f"CAR INSIDE(CD) = {CD}"    
        dim(doc,  x1, y2, x1, y1,  300, text_height=0.30, direction="left", text = textstr)                        
        dim(doc,  x1, y1, x1 + 50, y1 + 50,  200, direction="left")      
        dimcontinue(doc, x1 + 50, y2 - 50)
        dimcontinue(doc, x1, y2)
        dim(doc, x1 + 50, y1 + 50,  x1 + 50, y1 + CD/2,  150, direction="left")      
        dimcontinue(doc, x1 - 50, y2 - 50)  
                           
        width_insideGap = 50
        height_insideGap = 50
        x1 = x + 1000
        y1 = abs_y
        x2 = x1 + CW
        y2 = abs_y + CD    
        rectangle(doc, x1, y1, x2, y2, layer='2')
        
        x1 = x1 + width_insideGap
        y1 = y1 + height_insideGap
        x2 = x1 + CW - width_insideGap*2
        y2 = y1 + CD - height_insideGap*2
        rectangle(doc, x1, y1, x2, y2, layer='0')
        
        ICW = CW - width_insideGap*2 #의장면 가로폭
        ICD = CD - height_insideGap*2 #의장면 세로폭
        
        #센터 세로선 3줄
        line(doc, x1 + ICW/2,  y1,   x1 + ICW/2, y2, layer="Hidden")
        line(doc, x1 + ICW/2-1.2,  y1,   x1 + ICW/2-1.2, y2, layer="Hidden")
        line(doc, x1 + ICW/2+1.2,  y1,   x1 + ICW/2+1.2, y2, layer="Hidden")
        
        #센터 가로줄
        line(doc, x1,  y1 + ICD/2,   x2, y1 + ICD/2, layer="Hidden") 
        line(doc, x1 + 1.2,  y1 + ICD/2+1.2,   x1 + ICW/2-1.2, y1 + ICD/2+1.2, layer="Hidden") #센터 위
        line(doc, x2-1.2,  y1 + ICD/2+1.2,   x1 + ICW/2+1.2, y1 + ICD/2+1.2, layer="Hidden")
        line(doc, x1 + 1.2,  y1 + ICD/2-1.2,   x1 + ICW/2-1.2, y1 + ICD/2-1.2, layer="Hidden") #센터 아래
        line(doc, x2-1.2,  y1 + ICD/2-1.2,   x1 + ICW/2+1.2, y1 + ICD/2-1.2, layer="Hidden")
        
        #테두리
        rectangle(doc, x1 + 26,  y2 - 25,  x1 + ICW/2-1.2, y2 - 1.2, layer='Hidden')#상좌
        rectangle(doc, x2 - 26,  y2 - 25,  x1 + ICW/2+1.2, y2 - 1.2, layer='Hidden')#상우
        rectangle(doc, x1 + 26,  y1 + 25,  x1 + ICW/2-1.2, y1 + 1.2, layer='Hidden')#하좌
        rectangle(doc, x2 - 26,  y1 + 25,  x1 + ICW/2+1.2, y1 + 1.2, layer='Hidden')#하우
        
        rectangle(doc, x1,  y1 ,  x1 + 25, y2, layer='Hidden')#좌
        rectangle(doc, x2 ,  y1,  x2 - 25, y2, layer='Hidden')#우
        line(doc, x1 + 1.2,  y1,  x1 + 1.2, y2, layer="Hidden")#좌
        line(doc, x2 - 1.2,  y1,  x2 - 1.2, y2, layer="Hidden")#우
             
        insert_block(x1 + ICW/4 , y1 + ICD/4 , "lc008_A_Halogen_S")    
        insert_block(x1 + ICW/4 , y2 - ICD/4 , "lc008_A_Halogen_S")    
        insert_block(x2 - ICW/4 , y1 + ICD/4 , "lc008_A_Halogen_S")    
        insert_block(x2 - ICW/4 , y2 - ICD/4 , "lc008_A_Halogen_S")    
        
        # LC 홀 가공 (기존 본천장 로직 적용해서 만들것)      
        ICW = CW - width_insideGap*2 #의장면 가로폭
        ICD = CD - height_insideGap*2 #의장면 세로폭
        
        inputCW = ICW
        inputCD = ICD

        lc_positions = calculate_lc_position_008_A_CW(inputCW, inputCD)

    # lc_positions = {
    #     'Topholex': Topholex,
    #     'Topholey': Topholey,
    #     'Bottomx': Bottomx,
    #     'Bottomy': Bottomy
    # }        

        # Topholex와 Topholey 좌표 추출
        Topholex = lc_positions['Topholex']
        Topholey = lc_positions['Topholey']

        # Bottomx와 Bottomy 좌표 추출
        Bottomx = lc_positions['Bottomx']
        Bottomy = lc_positions['Bottomy']    

        temp_x = x1
        temp_y = y1
        
        for x in Topholex:
            circle_cross(doc, temp_x + x, temp_y + Topholey, 9, layer='0')    
            circle_cross(doc, temp_x + x, temp_y + Bottomy, 9, layer='0')          
            
        temp_x = x1 + ICW/2
        temp_y = y1
        
        for i, x in enumerate(Topholex):
            circle_cross(doc, temp_x + x, temp_y + Topholey, 9, layer='0')    
            if i == 0:
                dim(doc, x1+x , y1 , x1, y1 , 150,  direction="down", option='reverse')    
            else:
                dimcontinue(doc , x1 + x, y1)
        for i, x in enumerate(Topholex):
            circle_cross(doc, temp_x + x, temp_y + Bottomy, 9, layer='0')           
            if i == 0:
                dim(doc, x1 + CW/2 + x -50, y1, x1 + CW/2 + x -50-26, y1 , 150, direction="down", option='reverse')    
            else:
                dimcontinue(doc, x1 + CW/2 + x -50, y1)
        dimcontinue(doc ,x2, y1) 
                
        lc_positions = calculate_lc_position_008_A_CD(inputCW, inputCD)

        # 'Leftx': Leftx,
        # 'Lefty': Lefty,
        # 'Rightx': Rightx,
        # 'Righty': Righty

        Leftx = lc_positions['Leftx']
        Lefty = lc_positions['Lefty']

        Rightx = lc_positions['Rightx']
        Righty = lc_positions['Righty']    
        
        temp_x = x1
        temp_y = y1 + ICD/2

        # print(f"temp_x: {temp_x} temp_y: {temp_y} ")        
        
        for i, y in enumerate(Lefty):
            circle_cross(doc, temp_x + Leftx , temp_y + y, 9, layer='0')    
            if i == 0:
                dim(doc, x2 , y1 + y , x2, y1 , 150,  direction="right", option='reverse')    
            else:
                dimcontinue(doc , x2, y1 + y)

        for i, y in enumerate(Lefty):
            circle_cross(doc, temp_x + Rightx , temp_y + y, 9, layer='0')       
            if i == 0:
                dim(doc, x2 , y1 + y + ICD/2, x2, y1 + ICD/2-13 , 150,  direction="right", option='reverse')    
            else:     
                dimcontinue(doc , x2, y1 + y + ICD/2)
        dimcontinue(doc ,x2, y2 )
        
        temp_x = x1
        temp_y = y1

        # print(f"temp_x: {temp_x} temp_y: {temp_y} ")        
        
        for y in Lefty:
            circle_cross(doc, temp_x + Leftx , temp_y + y, 9, layer='0')    
            circle_cross(doc, temp_x + Rightx , temp_y + y, 9, layer='0')            
        
        #의장면 조립홀 간격별 선 그리기
        #가로
        lc_positions = calculate_lc_position_008_A_ASSY_horizontal(inputCW, inputCD)
        
        linex = lc_positions['linex']
        liney = lc_positions['liney']
        
        temp_x = x1
        temp_y = y1

        for i, x in enumerate(linex):
            insert_block(temp_x + x , temp_y + liney , "lc008_A_ASSY_line")    
            if i == 0:
                dim(doc, x1 + x , y1 + ICD/2 , x1, y1 + ICD/2 , 80,  direction="down", option='reverse')    
            else:
                dimcontinue(doc , x1 + x, y1 + ICD/2)
        
        temp_x = x1 + ICW/2
        temp_y = y1
        
        for i, x in enumerate(linex):
            insert_block(temp_x + x , temp_y + liney , "lc008_A_ASSY_line")    
            if i == 0:
                dim(doc, x1 + ICW/2 + x , y1 + ICD/2 , x1 + ICW/2-37.5, y1 + ICD/2 , 80,  direction="down", option='reverse')    
            else:
                dimcontinue(doc , x1 + ICW/2  + x, y1 + ICD/2)
            
        dim(doc, x2 - 37.5, y1 + ICD/2 , x2 , y1 + ICD/2 , 80,  direction="down")    
        
        #세로
        lc_positions = calculate_lc_position_008_A_ASSY_vertical(inputCW, inputCD)
        
        linex = lc_positions['linex']
        liney = lc_positions['liney']
        
        temp_x = x1
        temp_y = y1
        
        for i, y in enumerate(liney):
            insert_block(temp_x + linex , temp_y + y , "lc008_A_ASSY_line_")    
            if i == 0:
                dim(doc, x1 + ICW/2 , y1 + y , x1 + ICW/2, y1 , 100,  direction="right", option='reverse')    
            else:
                dimcontinue(doc , x1 + ICW/2, y1 + y)
                            
        temp_x = x1
        temp_y = y1 + ICD/2

        for i, y in enumerate(liney):
            insert_block(temp_x + linex , temp_y + y , "lc008_A_ASSY_line_")    
            if i == 0:
                dim(doc, x1 + ICW/2 , y1 + y + ICD/2 , x1 + ICW/2, y1 + ICD/2-37.5, 100,  direction="right", option='reverse')    
            else:
                dimcontinue(doc , x1 + ICW/2, y1 + y + ICD/2)
                
        dim(doc , x1 + ICW/2 , y2 - 37.5, x1 + ICW/2, y2, 100,  direction="right")    
       
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1715
        TargetXscale = 5500 + (CW - 1600)  + (CD - 1500)   # 5500은 CW 1600 크기의 도면전체 크기를 의미함
        TargetYscale = 3400 + (CD - 1500)                            # 3400은 CD 1500 크기의 도면전체 크기를 의미함
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("스케일 비율 : " + str(frame_scale))        
        frameYpos = abs_y - 450 * frame_scale   
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)       

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 3page 프레임 좌우 본체만들기 (레이져가공)
    #################################################################################################        
    program = 1
    if program:
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = BasicXscale + (CD - 1500) 
        TargetYscale = BasicYscale
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        insert_frame(frameXpos, frameYpos, frame_scale, "drawings_frame", workplace)   

        rx1 = frameXpos + LCD + 1000
        ry1 = frameYpos + 1100 * frame_scale
        insert_block(rx1, ry1, "lc008_A_frame_section")    

        rx1 = frameXpos + 500
        
        thickness = 1.5
        vcut = thickness / 2

        x1 = rx1
        y1 = ry1
        x2 = x1 + CD - 120 + 14
        y2 = y1    
        x3 = x2 
        y3 = y2 + 37 - thickness
        x4 = x3 - 37
        y4 = y3
        x5 = x4
        y5 = y4 + 70- thickness*2
        x6 = x5 + 30
        y6 = y5
        x7 = x6 
        y7 = y6 + 30 - thickness
        x8 = x7 - CD + 120
        y8 = y7
        x9 = x8
        y9 = y8 - 30 + thickness
        x10 = x9 + 30
        y10 = y9
        x11 = x10
        y11 = y10 - 70 + thickness*2
        x12 = x11 - 37
        y12 = y11

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 12
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
        
        #절곡라인  (히든선 : 정방향)      
        line(doc, x10, y10, x5, y5,  layer='hidden')     
        line(doc, x11, y11, x4, y4,  layer='hidden')     
        
        #11파이, 20파이 홀
        draw_circle(doc, x8 + 15, y8 - 15, 11, layer="레이져") 
        draw_circle(doc, x8 + 15 + (CD-150)/2, y8 - 15, 11, layer="레이져") 
        draw_circle(doc, x7 - 15, y8 - 15, 11, layer="레이져") 
        
        draw_circle(doc, x8 + 70, y1 + 54, 20, layer="레이져") 
        draw_circle(doc, x7 - 70, y1 + 54, 20, layer="레이져") 
                
        width_insideGap = 50
        height_insideGap = 50
        
        ICW = CW - width_insideGap*2 #의장면 가로폭
        ICD = CD - height_insideGap*2 #의장면 세로폭
        
        inputCW = ICW
        inputCD = ICD

        lc_positions = calculate_lc_position_008_A_CD(inputCW, inputCD)

        Leftx = lc_positions['Leftx']
        Lefty = lc_positions['Lefty']

        Rightx = lc_positions['Rightx']
        Righty = lc_positions['Righty']    
        
        temp_x = x1
        temp_y = y1 + 10

        # print(f"temp_x: {temp_x} temp_y: {temp_y} ")        
        
        for i, y in enumerate(Lefty):
            circle_cross(doc, temp_x + Leftx + y - 16, temp_y, 9, layer='레이져')    
            if i == 0:
                dim(doc, x1 + y - 3, y1 , x1, y1 , 100, text_height=0.22, direction="down", option='reverse', localdimstyle="0.2 JKW")    
            else:
                dimcontinue(doc , x1 + y - 3, y1)
        
        for i, y in enumerate(Lefty):
            circle_cross(doc, temp_x + ICD/2 + Leftx + y - 16, temp_y, 9, layer='레이져')    
            if i == 0:
                dim(doc, x1 + ICD/2 + y - 3, y1 , x1 + ICD/2-16, y1 , 100,  direction="down", option='reverse', localdimstyle="0.2 JKW")      
            else:
                dimcontinue(doc , x1 + ICD/2 + y - 3, y1)
        dimcontinue(doc ,x2, y2 )
        #지시선
        dim_leader_line(doc, x1 + (CD-120)/2+10, y1+10 , x1 + (CD-120)/2 + 50 ,y1+15+40, f"8-%%C9 Hole")       
        
        dim_leader_line(doc, x8 + (CD-120)/2, y8-15, x8 + (CD-120)/2 + 50 ,y8+ 30, f"3-%%C11 Holes (M8 Pop Nut)")       
        dim(doc, x8 + (CD-120)/2, y8 , x8 + (CD-120)/2 , y8-15, 100, direction="left")        
            
        dim_leader_line(doc, x1 + 77, y1 + 54, x1+77+30 ,y1+54+ 30, f"2-%%C20 Holes")       
        
        # 상부 전체치수
        dim(doc, x8+15, y8-15, x8, y8, 150, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x8+15+(CD-150)/2, y7-15)
        dimcontinue(doc, x7-15, y7-15)
        dimcontinue(doc, x7, y7)
        
        dim(doc, x10, y10, x8, y8, 210, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x5, y5)
        dimcontinue(doc, x7, y7)
        
        dim(doc, x8+70 , y8-77, x8, y8, 310, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x7-70, y7-77)
        dimcontinue(doc, x7, y7)
        
        dim(doc, x8 , y8, x12, y12, 280, direction="up", option='reverse', localdimstyle="0.2 JKW") 
        dimcontinue(doc, x7, y7)
        dimcontinue(doc, x3, y3)

        # 좌측 치수선
        dim(doc,  x8, y8 , x9 , y9 , 150,  direction="left", localdimstyle="0.2 JKW")           
        dimcontinue(doc, x12 , y12)
        dimcontinue(doc, x1 , y1, option='reverse')

        # 우측 치수선
        dim(doc,  x6, y6 , x7 , y7 , 160,  direction="right", localdimstyle="0.2 JKW")
        dimcontinue(doc,  x3, y3 )
        dimcontinue(doc,  x2, y2 )
           
        dim(doc,  x7, y7 , x2 , y2 , 220,  direction="right", localdimstyle="0.2 JKW")
        dim(doc, x1 , y1 , x2, y2 ,150 ,  direction="down", localdimstyle="0.2 JKW")
        
        #내부 홀위치 치수선
        dim(doc, x2 , y2 , x2-10, y2+10 ,50 ,  direction="left", localdimstyle="0.2 JKW")
        dim(doc, x1 , y1 , x1+77, y2+54 ,130 ,  direction="right", localdimstyle="0.2 JKW")
                            
        # main Frame description    
        x1 = frameXpos + CD + 250
        y1 = frameYpos + 800 * frame_scale
        textstr = f"Part Name : 하부 Light Case (좌, 우)"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : EGI 1.6t "    
        draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
        textstr = f"Size : {CD-106} x 131mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su*2} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')         

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 4page 프레임 본체만들기 (레이져가공) - 앞뒤판 프레임
    #################################################################################################        
    program = 1
    if program:
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = BasicXscale + (CW - 1500) 
        TargetYscale = BasicYscale
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        insert_frame(frameXpos, frameYpos, frame_scale, "drawings_frame", workplace)   

        rx1 = frameXpos + LCW + 1000
        ry1 = frameYpos + 1100 * frame_scale
        insert_block(rx1, ry1, "lc008_A_frame_section")    

        rx1 = frameXpos  + 500
        
        thickness = 1.5
        vcut = thickness / 2

        x1 = rx1
        y1 = ry1
        x2 = x1 + CW - 176.8
        y2 = y1    
        x3 = x2 
        y3 = y2 + 37 - thickness
        x4 = x3
        y4 = y3 + 70- thickness*2
        x5 = x4
        y5 = y4 + 30 - thickness
        x6 = x5 - CW + 176.8
        y6 = y5
        x7 = x6 
        y7 = y6 - 30 + thickness
        x8 = x7
        y8 = y7 - 70 + thickness*2

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 8
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
        
        #절곡라인  (히든선 : 정방향)      
        line(doc, x7, y7, x4, y4,  layer='hidden')     
        line(doc, x8, y8, x3, y3,  layer='hidden')     
        
        #11파이, 20파이 홀
        draw_circle(doc, x6 + (CW-176.8)/2, y6 - 15, 11, layer="레이져") 

        draw_circle(doc, x1 + 41.6, y1 + 54, 20, layer="레이져") 
        draw_circle(doc, x2 - 41.6, y1 + 54, 20, layer="레이져") 
        
        width_insideGap = 50
        height_insideGap = 50
        
        ICW = CW - width_insideGap*2 #의장면 가로폭
        ICD = CD - height_insideGap*2 #의장면 세로폭
        
        inputCW = ICW
        inputCD = ICD

        lc_positions = calculate_lc_position_008_A_CW(inputCW, inputCD)

        # print(f" 4page inputCW :{ inputCW}  inputCD : {inputCD}lc_positions : {lc_positions} ")
        
    # lc_positions = {
    #     'Bottomx': Bottomx,
    #     'Bottomy': Bottomy
    # }        

        # Bottomx와 Bottomy 좌표 추출
        Bottomx = lc_positions['Bottomx']
        Bottomy = lc_positions['Bottomy']    

        temp_x = x1
        temp_y = y1
        
        for i, x in enumerate(Bottomx[1:]):       
            circle_cross(doc, temp_x + x-38.4, temp_y + 10, 9, layer='레이져')                
            # print(f"x{i} : {x}")
            if i == 0:
                dim(doc, x1+x-38.4 , y1 , x1, y1 , 100,  direction="down", option='reverse', localdimstyle="0.2 JKW")    
            else:
                dimcontinue(doc , x1 + x-38.4, y1)
        
        for i, x in enumerate(Bottomx[:-1]):       
            circle_cross(doc, temp_x + x-38.4 + CW/2-50, temp_y + 10, 9, layer='레이져')    
            # print(f"x{i} : {x}")
            if i == 0:
                dim(doc, x1+x-38.4 + CW/2-50 , y1 , x1+x-38.4 + CW/2-50-26, y1 , 100,  direction="down", option='reverse', localdimstyle="0.2 JKW")    
            else:
                dimcontinue(doc , x1 + x-38.4 + CW/2-50, y1)
        dimcontinue(doc ,x2, y2 )

        # 상부 전체치수
        dim(doc, x6+(CW-176.8)/2, y6-15, x6, y6, 100, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x5, y5)

        dim(doc, x6+41.6, y1+54, x6, y6, 200, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x5-41.6, y1+54)
        dimcontinue(doc, x5, y5)
        
        dim(doc, x6, y6, x5, y5, 170, direction="up", localdimstyle="0.2 JKW")
        
        dim_leader_line(doc, x6 + (CW-176.8)/2, y6-15, x8 + (CW-176.8)/2 + 50 ,y6+ 30, f"%%C11 Holes (M8 Pop Nut)")       
        dim(doc, x6 + (CW-176.8)/2, y6 , x6 + (CW-176.8)/2 , y6-15, 100, direction="left", localdimstyle="0.2 JKW")
            
        dim_leader_line(doc, x1 + 41.6, y1 + 54, x1+41.6+30 ,y1+54+ 100, f"2-%%C20 Holes")       
        dim(doc, x1, y1 , x1+41.6 , y1+54, 100, direction="right", localdimstyle="0.2 JKW")
        
        # 오른쪽 치수
        dim(doc,  x3, y3 , x2 , y2 , 80,  direction="right", localdimstyle="0.2 JKW")
        dim(doc,  x4, y4 , x3 , y3 , 80,  direction="right", localdimstyle="0.2 JKW")
        dim(doc,  x4, y4 , x5 , y5 , 80,  direction="right", localdimstyle="0.2 JKW")

        # 오른쪽 전체치수
        dim(doc,  x5, y5 , x2 , y2 , 150,  direction="right", localdimstyle="0.2 JKW")
        
        # main Frame description    
        x1 = frameXpos + 950 + (CW-1600)
        y1 = frameYpos + 600 * frame_scale
        textstr = f"Part Name : Light Case (앞, 뒤)"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : EGI 1.6t"    
        draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
        textstr = f"Size : {CW - 176.8} x 131mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su*2} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')    

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 5page 라이트 케이스 (좌 뒤, 우 앞)
    #################################################################################################
    program = 1
    if program:         
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = 2730 + (max(CW, CD) - 1600) 
        TargetYscale = 1665 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("4page 스케일 비율 : " + str(frame_scale))           
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

        rx1 = frameXpos  + 710 + 500
        ry1 = frameYpos + 550 * frame_scale
        
        thickness = 1.2
        vcut = thickness / 2

        # 라이트케이스 절곡도 크기 정의
        lc_width = CW/2 - 50
        lc_height = CD/2 - 50 

        topwing1 = 30
        btwing1 = 30
        btwing2 = 25
        leftwing1 = 30
        leftwing2 = 25
        rightwing1 = 30

        x1 = rx1 
        y1 = ry1
        x2 = x1 + lc_width 
        y2 = y1    
        x3 = x2
        y3 = y2 + btwing2 - thickness 
        x4 = x3 
        y4 = y3 + btwing1 - vcut - thickness
        x5 = x4
        y5 = y4 + 0.9
        x6 = x5 + rightwing1 - vcut*2 - 1.5
        y6 = y5
        x7 = x6
        y7 = y6 + 24
        x8 = x7 + 1.5
        y8 = y7
        x9 = x8
        y9 = y8 + lc_height - 24 - 3
        x10 = x9 - rightwing1 + vcut*2
        y10 = y9
        x11 = x10
        y11 = y10 + 0.9
        x12 = x11 
        y12 = y11 + topwing1 - vcut
        x13 = x12 - lc_width
        y13 = y12 
        x14 = x13
        y14 = y13 - topwing1 + vcut
        x15 = x14 
        y15 = y14 - 0.9
        x16 = x15 - leftwing1 + vcut*4
        y16 = y15 
        x17 = x16 - leftwing2 + thickness
        y17 = y16
        x18 = x17
        y18 = y17 - lc_height + 24 + 3
        x19 = x18 + leftwing2 - thickness
        y19 = y18 
        x20 = x19 + 0.3
        y20 = y19
        x21 = x20 
        y21 = y20 - 24 
        x22 = x21 + leftwing1 - vcut*3 - 0.9
        y22 = y21
        x23 = x22
        y23 = y22 - 0.9
        x24 = x23 
        y24 = y23 - btwing1 + vcut*3

        # 전개도 사이즈 저장 #1 lc
        lcsizex2 = x8 - x18
        lcsizey2 = y12 - y2

        Radiusval = 3          
        lastNum = 24

        prev_x, prev_y = eval('x1'), eval('y1')

        def draw_arc(doc, start, end, radius, option=None):
            """ option='reverse'는 시계방향으로 점을 회전할때 이 것을 사용한다. """
            if option == 'reverse' :                
                line(doc, prev_x, prev_y, start[0], start[1], layer='레이져')                
                add_arc_between_points(doc, end, start, radius)
            else:                
                line(doc, prev_x, prev_y, start[0], start[1], layer='레이져')                
                add_arc_between_points(doc, start, end, radius)

        for i in range(2, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')

            if i == 6:
                draw_arc(doc, (curr_x - Radiusval, curr_y), (curr_x, curr_y + Radiusval), Radiusval)
                prev_x, prev_y = curr_x, curr_y + Radiusval  

            elif i == 21:
                draw_arc(doc, (curr_x, curr_y + Radiusval), (curr_x + Radiusval, curr_y), Radiusval)
                prev_x, prev_y = curr_x + Radiusval, curr_y  

            else:
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y
                
        
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
        
        #절곡라인 밴딩 가로 방향      
        line(doc, x24, y24, x3, y3,  layer='hidden')   # 절곡선      
        line(doc, x23 + 0.6, y23, x4 - 0.6, y4,  layer='hidden')   # 절곡선      
        line(doc, x14 + 0.6, y14, x11 - 0.6, y11,  layer='hidden')   # 절곡선      

        # 세로절곡선 표현
        line(doc, x19, y19, x16, y16,  layer='hidden')    # 절곡선        
        line(doc, x23 + 0.6, y23, x14 + 0.6, y14,  layer='hidden')    # 절곡선        
        line(doc, x4 - 0.6, y4, x11 - 0.6, y11,  layer='hidden')    # 절곡선        
        
        #하부 홀
        lc_positions = calculate_lc_position_008_A_CW(inputCW, inputCD)

        Bottomx = lc_positions['Bottomx']
        Bottomy = lc_positions['Bottomy']   
        
        temp_x = x1
        temp_y = y1
        
        for i, x in enumerate(Bottomx):       
            circle_cross(doc, temp_x + x, temp_y + 12, 9, layer='레이져')    
            # print(f"x{i} : {x}")
            if i == 0:
                dim(doc, x1 + x , y1 , x1, y1 , 150,  direction="down", option='reverse', localdimstyle="0.2 JKW")    
            else:
                dimcontinue(doc , x1 + x, y1)
        dimcontinue(doc ,x2, y2 )
        
        dim_leader_line(doc, x1 + 13, y1 + 12, x1 + 13 + 50, y1 + 12 - 50, f"4-%%C9 Holes (M6 Pop Nut)")       
        dim(doc, x2, y2, x2 - 13, y2 + 12, 50, direction="left", localdimstyle="0.2 JKW")
               
        #상부 홀
        lc_positions = calculate_lc_position_008_A_ASSY_horizontal(inputCW, inputCD)
        
        linex = lc_positions['linex']
        liney = lc_positions['liney']
        
        temp_x = x13
        temp_y = y13

        for i, x in enumerate(linex):
            circle_cross(doc, temp_x + x, temp_y - 15, 8, layer='레이져')    
            if i == 0:
                dim(doc, temp_x + x , temp_y , temp_x, temp_y , 70,  direction="up", option='reverse', localdimstyle="0.2 JKW")    
            else:
                dimcontinue(doc , temp_x + x, temp_y)
        dimcontinue(doc ,x12, y12 )
        
        text = "8-%%C8 Holes"
        dim_leader(doc,  x12 - 37.5, y12 - 15, x12 - 37.5 - 50, y12 - 15 - 100, text, direction="rightToleft", distance=13)   # distance 13은 글자높이임       
        dim(doc, x12 - 37.5, y12 - 15, x12, y12, 100, direction="left", localdimstyle="0.2 JKW")
               
        #좌측 홀
        lc_positions = calculate_lc_position_008_A_CD(inputCW, inputCD)

        Leftx = lc_positions['Leftx']
        Lefty = lc_positions['Lefty']
        
        temp_x = x18
        temp_y = y18
        
        for i, y in enumerate(Lefty[1:]):
            circle_cross(doc, temp_x - 1 + Leftx , temp_y + y - 25.5, 9, layer='레이져')    
            if i == 0:
                dim(doc, temp_x, temp_y + y - 25.5, temp_x, temp_y, 120,  direction="left", option='reverse', localdimstyle="0.2 JKW")
            else:
                dimcontinue(doc , temp_x, temp_y + y - 25.5)
        dimcontinue(doc ,x17, y17, option='reverse' )

        dim_leader_line(doc, x17 + 12, y17 - 11.5, x17 + 12 + 50, y17 - 11.5 - 50, f"3-%%C9 Holes (M6 Pop Nut)")       
        dim(doc,  x17 + 12, y17 - 11.5, x17, y17, 170, direction="down", localdimstyle="0.2 JKW")
               
        #우측 홀
        lc_positions = calculate_lc_position_008_A_ASSY_vertical(inputCW, inputCD)
        
        linex = lc_positions['linex']
        liney = lc_positions['liney']
        
        temp_x = x6
        temp_y = y6
        
        for i, y in enumerate(liney):
            circle_cross(doc, temp_x - 15 + 1.5, temp_y + y - 1.5, 8, layer='레이져')    
            if i == 0:
                dim(doc, temp_x - 5 , temp_y + y - 1.5, temp_x - 5, temp_y , 140,  direction="right", option='reverse', localdimstyle="0.2 JKW")
            else:
                dimcontinue(doc , temp_x - 5, temp_y + y - 1.5)
        dimcontinue(doc ,x9, y9 )
        dim(doc,  x9 - 15, y9 - 36, x9, y9, 30, direction="down", localdimstyle="0.2 JKW")
        
        # insert_block(x1 + lc_width/2 , y4 + lc_height/2 - 0.6 , "lc008_A_Halogen_S2")    
        circle_cross(doc, x1 + lc_width/2 , y4 + lc_height/2 - 0.6, 72, layer='레이져')
        text = "%%C72 Hole"
        dim_leader(doc, x1 + lc_width/2 , y4 + lc_height/2 - 0.6, x1 + lc_width/2 - 100 , y4 + lc_height/2 - 0.6 - 100, text, direction="rightToleft", distance=13)   # distance 13은 글자높이임
        
        dim(doc, x23 + 0.6, y23 + lc_height/2, x1 + lc_width/2 , y4 + lc_height/2, 100, direction="up", localdimstyle="0.2 JKW")
        dimcontinue(doc ,x10 - 0.6, y4 + lc_height/2)
        dim(doc, x23 + lc_width/2, y23, x1 + lc_width/2 , y4 + lc_height/2 - 0.6, 100, direction="right", localdimstyle="0.2 JKW")
        dimcontinue(doc ,x14 + lc_width/2, y14)

        # 상부 치수선
        dim(doc, x13, y13, x17, y17, 130, direction="up", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x12, y12)
        dimcontinue(doc, x9, y9)
        dim(doc, x17, y17, x9, y9, 200, direction="up", localdimstyle="0.2 JKW")

        # 좌측 치수선
        dim(doc, x13, y13, x17, y17, 230, direction="left", localdimstyle="0.2 JKW")
        dimcontinue(doc, x18, y18)
        dimcontinue(doc, x21, y21)
        dimcontinue(doc, x1, y1, 230, option="reverse")
        dim(doc, x13, y13, x1, y1, 280, direction="left", localdimstyle="0.2 JKW")

        # 하부 치수선
        dim(doc, x21, y21, x18, y18, 250, direction="down", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x1, y1)
        dimcontinue(doc, x2, y2)
        dimcontinue(doc, x6, y6)
        dimcontinue(doc, x8, y8, 100)

        # 우측 치수선
        dim(doc, x8, y8, x6, y6, 60, direction="right", localdimstyle="0.2 JKW")
        dim(doc, x6, y6, x2, y2, 200, direction="right", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x9, y9)
        dimcontinue(doc, x12, y12)

        #절곡 치수선
        dim(doc, x24, y24, x1, y1, 70, direction="right", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x23, y23)
        dimcontinue(doc, x14, y14)
        dimcontinue(doc, x13, y13)
        
        dim(doc, x19, y19, x18, y18, 80, direction="up", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x22 + 0.6, y22)
        dimcontinue(doc, x4 - 0.6, y4)
        dimcontinue(doc, x8, y8)
        
        # main Frame description    
        x1 = frameXpos + 150
        y1 = ry1 + CD/2 * 1.4
        textstr = f"Part Name : 의장면 Light Case (좌 뒤, 우 앞)"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : STS 201 M/R 1.2t"    
        draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
        textstr = f"Size : {round(lcsizex2, 1)} x {round(lcsizey2, 1)}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su*2} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')    
        
    ###############################################################################
    # 5page 좌측 단면도 그리기
    ###############################################################################
    # 본판 mainplate = mp 
    program = 1
    if program:
        mp = lc_height

        startx = rx1 - 600
        starty = ry1 + 77.6

        x1 = startx
        y1 = starty
        x2 = x1 - thickness    
        y2 = y1    
        x3 = x2  
        y3 = y2 - btwing2 
        x4 = x3 + btwing1 
        y4 = y3 
        x5 = x4 
        y5 = y4 + lc_height
        x6 = x5 - topwing1
        y6 = y5 
        x7 = x6 
        y7 = y6 - thickness
        x8 = x7 + topwing1 - thickness
        y8 = y7 
        x9 = x8
        y9 = y8 - lc_height + thickness *2 
        x10 = x9 - topwing1 + thickness*2
        y10 = y9

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 10
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="0")     

        #vcut 원 그리기
        draw_circle(doc, x4, y4, 13, layer='0', color='6') 
        draw_circle(doc, x5, y5, 13, layer='0', color='6') 

        # 치수선
        dim(doc, x3,y3, x4, y4, 110, direction="down", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x2,y2, x3, y3, 80, direction="left", localdimstyle="0.2 JKW")
        dim(doc, x4,y4, x5, y5, 110, direction="right", localdimstyle="0.2 JKW")
        dim(doc, x6,y6, x5, y5, 100, direction="up", localdimstyle="0.2 JKW")

        dim_leader_line(doc, x4, y4 , x4 + 50 , y4-50, "Vcut 2개소")
        
        insert_block(x6 + topwing1/2, y5, "lc008_A_lefttop")    
        insert_block(x3, y3 + 13, "lc008_A_leftbottom")    
        
        insert_block(x4, y4 + lc_height/2, "lc008_A_Halogen_L")    
        dim(doc, x4, y4, x4, y4 + lc_height/2, 60, direction="right", localdimstyle="0.2 JKW")
        dimcontinue(doc, x5, y5)
   
    ###############################################################################
    # 5page 상부 단면도 그리기
    ###############################################################################
    program = 1
    if program:    
        mp = lc_width

        startx = rx1 + leftwing1
        starty = ry1 + 400 + lc_height + leftwing2*3

        x1 = startx
        y1 = starty
        x2 = x1 - leftwing2
        y2 = y1    
        x3 = x2  
        y3 = y2 - leftwing1
        x4 = x3 + mp
        y4 = y3 
        x5 = x4 
        y5 = y4 + rightwing1
        x6 = x5 - thickness
        y6 = y5
        x7 = x6
        y7 = y6 - rightwing1 + thickness
        x8 = x7 - mp + thickness *2
        y8 = y7 
        x9 = x8 
        y9 = y8 + leftwing1 - thickness *2
        x10 = x9 + leftwing2 - thickness
        y10 = y9

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 10
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="0")     

        #vcut 원 그리기 4개소
        draw_circle(doc, x3, y3, 13, layer='0', color='6') 
        draw_circle(doc, x4, y4, 13, layer='0', color='6') 

        # 치수선
        dim(doc, x2,y2, x1, y1, 80, direction="up", localdimstyle="0.2 JKW")
        dim(doc, x2,y2, x3, y3, 80, direction="left", localdimstyle="0.2 JKW")
        dim(doc, x5,y5, x4, y4, 80, direction="right", localdimstyle="0.2 JKW")
        dim(doc, x3,y3, x4, y4, 130, direction="down", localdimstyle="0.2 JKW")
    
        dim_leader_line(doc, x4, y4 , x4 + 50 , y4 - 100, "Vcut 2개소")
        insert_block(x3 + lc_width/2, y3, "lc008_A_Halogen")    
        dim(doc, x3, y3, x3 + lc_width/2, y3, 80, direction="down", localdimstyle="0.2 JKW")
        dimcontinue(doc, x4, y4)
        
        insert_block(x2 + 13, y2, "lc008_A_topleft")    
        insert_block(x4, y4 + rightwing1/2, "lc008_A_topright")    
        
        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 6page 라이트 케이스 (우 뒤, 좌 앞)
    #################################################################################################
    program = 1
    if program:         
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = 2730 + (max(CW, CD) - 1600) 
        TargetYscale = 1665 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("4page 스케일 비율 : " + str(frame_scale))           
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

        rx1 = frameXpos  + 710 + 400 + 400
        ry1 = frameYpos + 550 * frame_scale
        
        thickness = 1.2
        vcut = thickness / 2

        # 라이트케이스 절곡도 크기 정의
        lc_width = CW/2 - 50
        lc_height = CD/2 - 50 

        topwing1 = 30
        btwing1 = 30
        btwing2 = 25
        leftwing1 = 30
        rightwing1 = 30
        rightwing2 = 25

        x1 = rx1
        y1 = ry1
        x2 = x1 - lc_width 
        y2 = y1    
        x3 = x2
        y3 = y2 + btwing2 - thickness 
        x4 = x3 
        y4 = y3 + btwing1 - vcut - thickness
        x5 = x4
        y5 = y4 + 0.9
        x6 = x5 - rightwing1 + vcut*2 + 1.5
        y6 = y5
        x7 = x6
        y7 = y6 + 24
        x8 = x7 - 1.5
        y8 = y7
        x9 = x8
        y9 = y8 + lc_height - 24 - 3
        x10 = x9 + rightwing1 - vcut*2
        y10 = y9
        x11 = x10
        y11 = y10 + 0.9
        x12 = x11 
        y12 = y11 + topwing1 - vcut
        x13 = x12 + lc_width
        y13 = y12 
        x14 = x13
        y14 = y13 - topwing1 + vcut
        x15 = x14 
        y15 = y14 - 0.9
        x16 = x15 + leftwing1 - vcut*4
        y16 = y15 
        x17 = x16 + leftwing2 - thickness
        y17 = y16
        x18 = x17
        y18 = y17 - lc_height + 24 + 3
        x19 = x18 - leftwing2 + thickness
        y19 = y18 
        x20 = x19 - 0.3
        y20 = y19
        x21 = x20 
        y21 = y20 - 24 
        x22 = x21 - leftwing1 + vcut*3 + 0.9
        y22 = y21
        x23 = x22
        y23 = y22 - 0.9
        x24 = x23 
        y24 = y23 - btwing1 + vcut*3
        
        # 전개도 사이즈 저장 #1 lc
        lcsizex2 = x17 - x9
        lcsizey2 = y13 - y1

        Radiusval = 3          
        lastNum = 24

        prev_x, prev_y = eval('x1'), eval('y1')

        for i in range(2, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')

            if i == 6:                
                draw_arc(doc, (curr_x + Radiusval, curr_y ), (curr_x , curr_y + Radiusval), Radiusval, option='reverse')
                prev_x, prev_y = curr_x , curr_y + Radiusval

            elif i == 21:
                draw_arc(doc,  (curr_x, curr_y +  Radiusval),  (curr_x -  Radiusval, curr_y ), Radiusval, option='reverse')
                prev_x, prev_y = curr_x - Radiusval, curr_y  

            else:
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y
        
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
        
        #절곡라인 밴딩 가로 방향      
        line(doc, x24, y24, x3, y3,  layer='hidden')   # 절곡선      
        line(doc, x23 - 0.6, y23, x4 + 0.6, y4,  layer='hidden')   # 절곡선      
        line(doc, x14 - 0.6, y14, x11 + 0.6, y11,  layer='hidden')   # 절곡선      

        # 세로절곡선 표현
        line(doc, x19, y19, x16, y16,  layer='hidden')    # 절곡선        
        line(doc, x23 - 0.6, y23, x14 - 0.6, y14,  layer='hidden')    # 절곡선        
        line(doc, x4 + 0.6, y4, x11 + 0.6, y11,  layer='hidden')    # 절곡선       
        
        #하부 홀
        lc_positions = calculate_lc_position_008_A_CW(inputCW, inputCD)

        Bottomx = lc_positions['Bottomx']
        Bottomy = lc_positions['Bottomy']   
        
        temp_x = x1
        temp_y = y1
        
        for i, x in enumerate(Bottomx):       
            circle_cross(doc, temp_x - x, temp_y + 12, 9, layer='레이져')    
            # print(f"x{i} : {x}")
            if i == 0:
                dim(doc, x1 - x , y1 , x1, y1 , 150,  direction="down", option='reverse', localdimstyle="0.2 JKW")    
            else:
                dimcontinue(doc , x1 - x, y1)
        dimcontinue(doc ,x2, y2 )
        
        dim_leader_line(doc, x2 + 13, y2 + 12, x2 + 13 + 50, y2 + 12 - 50, f"4-%%C9 Holes (M6 Pop Nut)")
        dim(doc, x2, y2, x2 + 13, y2 + 12, 50, direction="right", localdimstyle="0.2 JKW")
               
        #상부 홀
        lc_positions = calculate_lc_position_008_A_ASSY_horizontal(inputCW, inputCD)
        
        linex = lc_positions['linex']
        liney = lc_positions['liney']
        
        temp_x = x13
        temp_y = y13

        for i, x in enumerate(linex):
            circle_cross(doc, temp_x - x, temp_y - 15, 8, layer='레이져')    
            if i == 0:
                dim(doc, temp_x - x , temp_y , temp_x, temp_y , 90,  direction="up", option='reverse', localdimstyle="0.2 JKW")
            else:
                dimcontinue(doc , temp_x - x, temp_y)
        dimcontinue(doc ,x12, y12 )
        
        dim_leader_line(doc, x12 + 37.5, y12 - 15, x12 + 37.5 + 50, y12 - 15 - 100, f"8-%%C8 Holes")       
        dim(doc, x12 + 37.5, y12 - 15, x12, y12, 50, direction="right", localdimstyle="0.2 JKW")
               
        #우측 홀
        lc_positions = calculate_lc_position_008_A_CD(inputCW, inputCD)

        Leftx = lc_positions['Leftx']
        Lefty = lc_positions['Lefty']
        
        temp_x = x18
        temp_y = y18
        
        for i, y in enumerate(Lefty[1:]):
            circle_cross(doc, temp_x - 25 + Leftx , temp_y + y - 25.5, 9, layer='레이져')    
            if i == 0:
                dim(doc, temp_x, temp_y + y - 25.5, temp_x, temp_y , 140,  direction="right", option='reverse', localdimstyle="0.2 JKW")
            else:
                dimcontinue(doc , temp_x, temp_y + y - 25.5)
        dimcontinue(doc ,x17, y17)

        dim_leader_line(doc, x17 - 12, y17 - 11.5, x17 + 12 + 50, y17 - 11.5 + 150, f"3-%%C9 Holes (M6 Pop Nut)")       
        dim(doc,  x17 - 12, y17 - 11.5, x17, y17, 100, direction="down", localdimstyle="0.2 JKW")
               
        #좌측 홀
        lc_positions = calculate_lc_position_008_A_ASSY_vertical(inputCW, inputCD)
        
        linex = lc_positions['linex']
        liney = lc_positions['liney']
        
        temp_x = x6
        temp_y = y6
        
        for i, y in enumerate(liney):
            circle_cross(doc, temp_x + 12 + 1.5, temp_y + y - 1.5, 8, layer='레이져')    
            if i == 0:
                dim(doc, temp_x + 5 , temp_y + y - 1.5, temp_x + 5, temp_y, 130,  direction="left", option='reverse', localdimstyle="0.2 JKW")
            else:
                dimcontinue(doc , temp_x + 5, temp_y + y - 1.5, 130)
        dimcontinue(doc ,x9, y9, 123.5, option='reverse')
        dim(doc, x9 + 15, y9 - 36, x9, y9, 30, direction="down", localdimstyle="0.2 JKW")

        # insert_block(x1 - lc_width/2 , y4 + lc_height/2 - 0.6 , "lc008_A_Halogen_S2")    
        circle_cross(doc, x1 - lc_width/2 , y4 + lc_height/2 - 0.6, 72, layer='레이져')
        text = "%%C72 Hole"
        dim_leader(doc, x1 - lc_width/2 , y4 + lc_height/2 - 0.6, x1 - lc_width/2 - 100 , y4 + lc_height/2 - 0.6 - 100, text, direction="rightToleft", distance=13)   # distance 13은 글자높이임
        
        dim(doc, x23 - 0.6, y23 + lc_height/2, x1 - lc_width/2 , y4 + lc_height/2, 100, direction="up", localdimstyle="0.2 JKW")
        dimcontinue(doc ,x10 + 0.6, y4 + lc_height/2)
        dim(doc, x23 - lc_width/2, y23, x1 - lc_width/2 , y4 + lc_height/2 - 0.6, 100, direction="right", localdimstyle="0.2 JKW")
        dimcontinue(doc ,x14 - lc_width/2, y14)

        # 상부 치수선
        dim(doc, x13, y13, x17, y17, 150, direction="up", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x12, y12)
        dimcontinue(doc, x9, y9)
        dim(doc, x17, y17, x9, y9, 220, direction="up", localdimstyle="0.2 JKW")

        # 우측 치수선
        dim(doc, x17, y17, x13, y13, 200, direction="right", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x18, y18)
        dimcontinue(doc, x21, y21, 224.1, option="reverse")
        dim(doc, x21, y21, x1, y1, 224.1, direction="right", localdimstyle="0.2 JKW")
        dim(doc, x13, y13, x1, y1, 300, direction="right", localdimstyle="0.2 JKW")

        # 하부 치수선
        dim(doc, x21, y21, x18, y18, 250, direction="down", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x1, y1)
        dimcontinue(doc, x2, y2)
        dimcontinue(doc, x6, y6)
        dimcontinue(doc, x8, y8, 100)

        # 좌측 치수선
        dim(doc, x8, y8, x6, y6, 60, direction="left", option="reverse", localdimstyle="0.2 JKW")
        dim(doc, x2, y2, x6, y6, 200, direction="left", localdimstyle="0.2 JKW")
        dimcontinue(doc, x9, y9)
        dimcontinue(doc, x12, y12, 200, option="reverse")

        #절곡 치수선
        dim(doc, x24, y24, x1, y1, 70, direction="left", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x23, y23, option="reverse")
        dimcontinue(doc, x14, y14)
        dimcontinue(doc, x13, y13)
        
        dim(doc, x19, y19, x18, y18, 80, direction="up", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x22 - 0.6, y22)
        dimcontinue(doc, x4 + 0.6, y4)
        dimcontinue(doc, x8, y8)
        
        # main Frame description    
        x1 = frameXpos + 2000 - 150
        y1 = ry1 + CD/2 * 1.4 + 100
        textstr = f"Part Name : 의장면 Light Case (우 뒤, 좌 앞)"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : STS 201 M/R 1.2t"    
        draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
        textstr = f"Size : {round(lcsizex2, 1)} x {round(lcsizey2, 1)}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su*2} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')    
        
    ###############################################################################
    # 6page 우측 단면도 그리기
    ###############################################################################
    # 본판 mainplate = mp 
    program = 1
    if program:
        mp = lc_height

        startx = x1 - btwing2 + 300
        starty = ry1 + 77.6

        x1 = startx
        y1 = starty
        x2 = x1 + thickness    
        y2 = y1    
        x3 = x2  
        y3 = y2 - btwing2 
        x4 = x3 - btwing1 
        y4 = y3 
        x5 = x4 
        y5 = y4 + lc_height
        x6 = x5 + topwing1
        y6 = y5 
        x7 = x6 
        y7 = y6 - thickness
        x8 = x7 - topwing1 + thickness
        y8 = y7 
        x9 = x8
        y9 = y8 - lc_height + thickness *2 
        x10 = x9 + topwing1 - thickness*2
        y10 = y9

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 10
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="0")     

        #vcut 원 그리기
        draw_circle(doc, x4, y4, 13, layer='0', color='6') 
        draw_circle(doc, x5, y5, 13, layer='0', color='6') 

        # 치수선
        dim(doc, x3, y3, x4, y4, 110, direction="down", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x3, y3, x2, y2, 110, direction="right", localdimstyle="0.2 JKW")
        dim(doc, x4, y4, x5, y5, 110, direction="left", localdimstyle="0.2 JKW")
        dim(doc, x6, y6, x5, y5, 100, direction="up", localdimstyle="0.2 JKW")

        dim_leader_line(doc, x4, y4 , x4 + 50 , y4-50, "Vcut 2개소")
        
        insert_block(x6 - topwing1/2, y5, "lc008_A_righttop")    
        insert_block(x3, y3 + 13, "lc008_A_rightbottom")    
        
        insert_block(x4, y4 + lc_height/2, "lc008_A_Halogen_R")    
        dim(doc, x4, y4, x4, y4 + lc_height/2, 60, direction="left", localdimstyle="0.2 JKW")
        dimcontinue(doc, x5, y5)
   
    ###############################################################################
    # 6page 상부 단면도 그리기
    ###############################################################################
    program = 1
    if program:    
        mp = lc_width

        startx = rx1 - 24.8
        starty = ry1 + 400 + lc_height + leftwing2*3

        x1 = startx
        y1 = starty
        x2 = x1 + leftwing2
        y2 = y1    
        x3 = x2  
        y3 = y2 - leftwing1
        x4 = x3 - mp
        y4 = y3 
        x5 = x4 
        y5 = y4 + rightwing1
        x6 = x5 + thickness
        y6 = y5
        x7 = x6
        y7 = y6 - rightwing1 + thickness
        x8 = x7 + mp - thickness *2
        y8 = y7 
        x9 = x8 
        y9 = y8 + leftwing1 - thickness *2
        x10 = x9 - leftwing2 + thickness
        y10 = y9

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 10
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="0")     

        #vcut 원 그리기 4개소
        draw_circle(doc, x3, y3, 13, layer='0', color='6') 
        draw_circle(doc, x4, y4, 13, layer='0', color='6') 

        # 치수선
        dim(doc, x2,y2, x1, y1, 80, direction="up", localdimstyle="0.2 JKW")
        dim(doc, x2,y2, x3, y3, 80, direction="right", localdimstyle="0.2 JKW")
        dim(doc, x5,y5, x4, y4, 80, direction="left", localdimstyle="0.2 JKW")
        dim(doc, x3,y3, x4, y4, 130, direction="down", localdimstyle="0.2 JKW")
    
        dim_leader_line(doc, x3, y3 ,x3 + 50 ,y4 - 100, "Vcut 2개소")
        insert_block(x3 - lc_width/2, y3, "lc008_A_Halogen")    
        dim(doc, x3, y3, x3 - lc_width/2, y3, 80, direction="down", localdimstyle="0.2 JKW")
        dimcontinue(doc, x4, y4)
        
        insert_block(x2 - 13, y2, "lc008_A_topright2")    
        insert_block(x4, y4 + rightwing1/2, "lc008_A_topleft2")    
        
        frameXpos = frameXpos + TargetXscale + 400
    #############################################################################################
    # 7page  LED BAR B/K
    #############################################################################################
    program = 1
    if program:    
        abs_x = frameXpos 

        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = 1000
        TargetYscale = 100 
        frame_scale = 0.3   

        frame_scale = math.ceil(frame_scale*10) / 10       
        insert_frame(abs_x , frameYpos , frame_scale, "drawings_frame", workplace)          
        
        x1 = frameXpos + 360
        y1 = frameYpos + 1450 *frame_scale 
       
        # LED BAR B/K 단면도 삽입
        insert_block(x1, y1 - 50, "lc008_A_led bar B.K")
        insert_block(x1, y1 - 50, "lc008_A_led bar B.K_laser")

        # LED BAR B/K  description    
        x = x1
        y = y1 - 150
        textstr = f"Part Name :  LED BAR B/K "    
        draw_Text(doc, x , y , 20 , str(textstr), layer='0')        
        textstr = f"Mat.Spec : EGI  1.2t "    
        draw_Text(doc, x , y - 40 , 20, str(textstr), layer='0')    
        textstr = f"Size : 40 x 32.2mm"    
        draw_Text(doc, x , y - 80 , 20, str(textstr), layer='0')        
        textstr = f"Quantity : {su*8} EA"    
        draw_Text(doc, x , y - 120 , 20, str(textstr), layer='0')   

    program = 1
    if program:
        ####################################################################################################################################################
        # 레이져 도면틀 생성 (공통) 레이져에 배열할때 편리한 기능
        # 1219*1950 철판크기 샘플
        x1 = 10000+850
        y1 = 5000+130
        x2 = 10000+850+2438
        y2 = y1 +1219    
        x3 = 10000+850+2438+850+2438+ 1850  
        y3 = y1 +1219
        rectangle(doc, 10000, 5000,10000+13400 , 5000+ 2340 , layer='0')  
        rectangle(doc,x1 , y1 ,x1 + 2438 , y2 , layer='레이져')  
        dim(doc, x1 , y2 , x2 , y2,  100, direction="up")
        # rectangle(doc, x1+2600 , y1 , x2+2600 , y2 , layer='레이져')  
        # dim(doc, x1+2600, y2, x2+2600, y2, 100, direction="up")
        # dim(doc,  x1 , y1, x1 ,y2 , 150, direction="left")
        rectangle(doc, x3 , y1 , x3+2600 , y2 , layer='레이져')  
        dim(doc, x3, y3, x3+2600, y2, 100, direction="up")
        dim(doc,  x3 , y1, x3 ,y2 , 150, direction="left")

        textstr = f"{secondord} - {workplace} - {drawdate_str_short} 출도, {deadlinedate_str_short} 납기"    
        draw_Text(doc, 10000+300 , 5000+ 2340 - 200 ,  120, str(textstr), layer='0')        

        y = 5000+ 2340 - 650
        textstr = f"SPCC 1.2tX1219X = 장"
        draw_Text(doc, 10000+1000 ,y ,  90, str(textstr), layer='0')        
        textstr = f"SPCC 1.6T"        
        draw_Text(doc, 10000+1000+2600 , y ,  90, str(textstr), layer='0')            
        textstr = f"EGI 1.2T"        
        draw_Text(doc, 10000+1000+5200 ,y ,  90, str(textstr), layer='0')        
        textstr = f"{LCplateMaterial}"
        draw_Text(doc, 10000+1000+5200+2600 , y ,  90, str(textstr), layer='0')            
   
####################################################################################################################################################################################
# lc035 천장 자동작도
####################################################################################################################################################################################
def lc035():   
    global args  #수정할때는 global 선언이 필요하다. 단순히 읽기만 할때는 필요없다.    
    global start_time
    global workplace, drawdate, deadlinedate, lctype, secondord, text_style, exit_program, Vcut_rate, Vcut, program_message, formatted_date, text_style_name
    global CW_raw, CD_raw, panel_thickness, su, LCframeMaterial, LCplateMaterial
    global input_CD, input_CW,CD,CW,drawdate_str , deadlinedate_str, drawdate_str_short, deadlinedate_str_short
    global doc, msp , T5_is, text_style, distanceXpos, distanceYpos

    # 2page 덴크리 다온텍등 등기구 설명 그림
    insert_block(0 , 0 , "lc035_2pageframe")        
    abs_x = 0
    abs_y = 0
    # LC 크기 지정 car inside에서 차감
    LCD = CD-50
    LCW = CW-50  # 라이트 케이스 크기를 정한다. 보통은 50
        
    ledsu = 4    

    if(CD>2000):
        ledsu = 4

    wattCalculate = math.ceil((LCD - 3)/1000 * ledsu * 15)

    print("wattCalculate: " + str(wattCalculate))  

    if(wattCalculate <= 60):
        watt = 60
    elif(wattCalculate > 60 and wattCalculate < 100):
        watt = 100
    elif(wattCalculate >= 100 and wattCalculate < 150):
        watt = 150
    elif(wattCalculate >= 150 and wattCalculate < 200):
        watt = 200
    elif(wattCalculate >= 200 and wattCalculate < 200):
        watt = 300
                
    # print("L/C 1개당 규격설정 wattCalculate: " + str(watt))  

    T5_standard = 0
    if T5_is == '있음':
        adjusted_CD = CD - 50  # CD에서 50 뺌

        # 규격 결정
        if adjusted_CD > 1500:
            T5_standard = 1500
        elif adjusted_CD > 1200:
            T5_standard = 1200
        else:
            T5_standard = 900    

    # 도면틀 넣기
    BasicXscale = 2560
    BasicYscale = 1550
    TargetXscale = 800 
    TargetYscale = 300 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale

    frame_scale = 1    
    ########################################################
    # 갑지 1
    ########################################################
    # 현장명
    textstr = workplace       
    x = abs_x + 80
    y = abs_y + 1000 + 218
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 발주처
    textstr = secondord
    x = abs_x + 80
    y = abs_y + 1000 + 218 - 30
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 타입
    textstr = lctype
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 수량
    textstr = su
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 카사이즈
    textstr = f"W{CW} x D{CD}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 색상 LC
    LCcolor = LCframeMaterial
    if LCframeMaterial=='SPCC 1.2T' :
        LCcolor = '백색무광'
    textstr = f"L / C : {LCcolor}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')   

    textstr = f"중판 : {LCplateMaterial}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18 - 12
    draw_Text(doc, x, y , 7, str(textstr), layer='0')

    # 등기구 업체
    LedBarLength = 0
    if T5_standard > 0 : 
        Ledcompany = f"(등기구업체 : 엘엠팩트)"        
        Leditem = "T5"
        Ledwatt = "6500K"
        LedBarLength_CD = T5_standard
        LedBarLength_CW = T5_standard
        watt = "내장"
    else:
        Ledcompany = f"(등기구업체 : 다온텍)"
        Leditem = "LED BAR"
        Ledwatt = "10000K"
        LedBarLength_CD = CalculateLEDbar(CD-70)
        LedBarLength_CW = CalculateLEDbar(CW-170)
        # print (f"035 LedBarLength_CD : {LedBarLength_CD}")
        # print (f"035 LedBarLength_CW : {LedBarLength_CW}")
        watt = f"{str(watt)}W"

    x = abs_x + 80
    y = abs_y + 1000 + 120
    draw_Text(doc, x, y , 8, str(Ledcompany), layer='0')     
    textstr = f"(아답터용량 : 슬림형 SMPS -{watt} - {su}EA)"
    x = abs_x + 80
    y = abs_y + 1000 + 120 - 15
    draw_Text(doc, x, y , 8, str(textstr), layer='0')
    
    altube_CW = CW-68*2
    altube_CD = CD-26*2

    textstr = f"1) AL PROFILE LED BAR TUBE - (확산커버포함)-SHP4045-{altube_CW}L={su*2}(set)"
    x = abs_x + 80
    y = abs_y + 1000 + 140 - 15 - 30
    draw_Text(doc, x, y , 5, str(textstr), layer='0')
    
    textstr = f"(LED BAR-누드&클립&잭타입-10000K-{LedBarLength_CW}L)"
    x = abs_x + 80 + 90
    y = abs_y + 1000 + 140 - 15 - 30 - 10
    draw_Text(doc, x, y , 5, str(textstr), layer='0')
    
    textstr = f"2) AL PROFILE LED BAR TUBE - (확산커버포함)-SHP4045-{altube_CD}L={su*2}(set)"
    x = abs_x + 80
    y = abs_y + 1000 + 140 - 15 - 30 -10 -10
    draw_Text(doc, x, y , 5, str(textstr), layer='0')
    
    textstr = f"(LED BAR-누드&클립&잭타입-10000K-{LedBarLength_CD}L)"
    x = abs_x + 80 + 90
    y = abs_y + 1000 + 140 - 15 - 30 - 10 - 10 - 10 
    draw_Text(doc, x, y , 5, str(textstr), layer='0')

    textstr = f"*. 출도일자 : {drawdate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 
    draw_Text(doc, x, y , 8, str(textstr), layer='0')    
    textstr = f"*. 납품일자 : {deadlinedate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 - 15
    draw_Text(doc, x, y , 8, str(textstr), layer='0')

    ################################################################################################################################################
    # 갑지 2
    ################################################################################################################################################
 
    # 현장명
    textstr = f"* 현장명 : {secondord}-{workplace} - {su}(set)"
    x = abs_x + 30
    y = abs_y + 650 
    draw_Text(doc, x, y , 20, str(textstr), layer='0')        
    x = abs_x + 30 + 200
    y = abs_y + 600 
    draw_Text(doc, x, y , 20, str(Ledcompany), layer='0')    
    
    textstr = f"{watt}"
    x = abs_x + 1300 - 445
    y = 395 + 80
    draw_Text_direction(doc, x + 50, y - 30 , 14, str(textstr), layer='0', rotation = 0)    

    textstr = f"AL PROFILE TUBE (마구리포함) -{altube_CW}L"
    x = abs_x + 290 - 35
    y = 514
    draw_Text(doc, x, y , 12, str(textstr), layer='0')
    
    textstr = f"(LED BAR-누드&클립&잭타입-10000K-{LedBarLength_CW}L)"
    x = abs_x + 290 - 65
    y = 405
    draw_Text(doc, x, y , 12, str(textstr), layer='0')

    textstr = f"AL PROFILE TUBE (마구리포함) -{altube_CW}L"
    x = abs_x + 290 - 35
    y = 164
    draw_Text(doc, x, y , 12, str(textstr), layer='0')
    
    textstr = f"(LED BAR-누드&클립&잭타입-10000K-{LedBarLength_CW}L)"
    x = abs_x + 290 - 65
    y = 60
    draw_Text(doc, x, y , 12, str(textstr), layer='0')

    x = abs_x + 70
    y = 130    
    textstr = f"AL PROFILE TUBE (마구리포함) -{altube_CD}L"
    draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)        

    x = abs_x + 180
    y = 120    
    textstr = f"LED BAR-누드&클립&잭타입-10000K-{LedBarLength_CD}L"
    draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)        
       
    x = abs_x + 700
    y = 130    
    textstr = f"AL PROFILE TUBE (마구리포함) -{altube_CD}L"
    draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)        

    x = abs_x + 609
    y = 120    
    textstr = f"LED BAR-누드&클립&잭타입-10000K-{LedBarLength_CD}L"
    draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)   

    # car inside box 표기
    xx =  790
    yy =  145
    rectangle(doc, xx,yy,xx+200,yy+100,layer='0')      
    line(doc, xx,yy+50,xx+200,yy+50)
    textstr = f"CAR INSIDE"    
    draw_Text(doc, xx+50, yy+50 , 12, str(textstr), layer='0')
    textstr = f"W{CW} x D{CD}"    
    draw_Text(doc, xx+30, yy+25 , 12, str(textstr), layer='0')         
       
    frameXpos = 1037

    # Assy Car case    
    #####################################################################################################################################
    # 1page 조립도 Assy
    ##################################################################################################################################### 
    # LC ASSY 상부 형상    
    program = 1
    if program:
        abs_x = frameXpos + 1700    
        frameXpos = abs_x - 500
        abs_y = abs_y + CD + 500
        insert_block(abs_x , abs_y , "lc035_top_left")    

        x = abs_x + CW 
        y = abs_y 
        insert_block(x , y , "lc035_top_right")  

        first_space = 135

        # 폴리갈 or 중판 크기 산출
        polygal_width = CW - first_space * 2
        polygal_height = LCD - 6

        # print("polygal_height: " + str(polygal_height))  

        # 중판 녹색 실물 그리기 1.2t
        rectangle(doc, abs_x + 69 , abs_y + 19.8 , abs_x + CW - 69  ,abs_y + 19.8 + 34 , layer='1')  
        # offset 1.2T
        rectangle(doc, abs_x + 69 , abs_y + 19.8 , abs_x + CW - 69  ,abs_y + 19.8 + 34 , layer='1', offset=1.2)  
        # 흰색선
        line(doc, abs_x + 67 , abs_y + 45 , abs_x + CW - 67  ,abs_y + 45 , layer='0')

        # 상부선 2.0mm 간격
        x1 = abs_x - 25
        y1 = abs_y + 150
        x2 = x1 + CW + 50
        y2 = y1
        rectangle(doc, x1, y1, x2, y2+2, layer='0')   
    
        # 1.2T 상부 히든선 추가
        x1 = abs_x + 130
        y1 = abs_y + 150-1.2
        x2 = x1 + CW - 130
        y2 = y1  
        line(doc, x1, y1, x2, y2, layer='hidden') 

        x1 = abs_x + 25
        y1 = abs_y + 90
        x2 = abs_x + CW - 25
        y2 = y1    
        textstr =  f"Light Case (W) = {LCW}"      
        dim(doc, x1, y1, x2, y2, 197, text_height=0.30, text_gap=0.07, direction="up" , text = textstr)    
        
        # Assy 상부 의장면 OPen size
        x1 = abs_x + 67
        y1 = abs_y  
        x2 = abs_x + CW - 67
        y2 = y1       
        textstr =  f"{CW - 67*2}"   
        dim(doc, x1, y1, x2, y2, 73.15, text_height=0.30, text_gap=0.07, direction="down",  text = textstr)

        x1 = abs_x + 25
        y1 = abs_y + 36
        x2 = abs_x + CW - 25
        y2 = y1        
        textstr =  f"Light Case Wide (W) = {LCW}"    
        dim(doc, x1, y1, x2, y2, 166,  text_height=0.30, text_gap=0.07, direction="down",  text = textstr)

        # Assy 하부 치수
        x1 = abs_x 
        y1 = abs_y 
        x2 = abs_x + CW
        y2 = y1    
        textstr =  f"Car Inside (CW) = {CW}"
        dim(doc, x1, y1, x2, y2, 240,  text_height=0.30, direction="down", text=textstr)

    ###################################
    # 1page Assy 본판 작도
    ###################################
    program = 1
    if program:    
        # 기본 윤곽 car inside
        width_insideGap = 25
        height_insideGap = 25
        abs_y = abs_y - CD - 800    
        x1 = abs_x 
        y1 = abs_y 
        x2 = abs_x + CW
        y2 = abs_y + CD    
        rectangle(doc, x1, y1, x2, y2, layer='hidden')    
        x1 = abs_x + width_insideGap
        y1 = abs_y + height_insideGap
        x2 = abs_x + CW - width_insideGap
        y2 = abs_y + CD - height_insideGap
        rectangle(doc, x1, y1, x2, y2, layer='0')    

        # 상하 30 철판 표현
        ribGap = 30
        # x방향 CW와 LC 떨어진 gap
        Xgap = 25
        x1 = abs_x + 125
        y1 = abs_y + height_insideGap
        x2 = abs_x + CW - 125
        y2 = y1+ribGap
        rectangle(doc, x1, y1, x2, y2, layer='0')        
        y1 = abs_y + CD - height_insideGap    
        y2 = y1-ribGap
        rectangle(doc, x1, y1, x2, y2, layer='0')    

        # 양쪽 테두리 30 흰색선
        x1 = abs_x + 95
        y1 = abs_y + height_insideGap        
        x2 = x1 + 30
        y2 = abs_y + CD - height_insideGap
        rectangle(doc, x1, y1, x2, y2, layer='0') 
        x1 = abs_x + CW - 95 - ribGap
        x2 = x1 + ribGap
        rectangle(doc, x1, y1, x2, y2, layer='0') 

        # 좌측 fan assy (좌측상단 기준)
        insert_block(abs_x + 58 , abs_y  + CD - 70 , "lc026_fan")         

        # LC 홀 가공 (기존 본천장 로직 적용해서 만들것)
        inputCW = CW
        inputCD = CD

        lc_positions = calculate_lc_position_035(inputCW, inputCD)

        # Topholex와 Topholey 좌표 추출
        Topholex = lc_positions['Topholex']
        Topholey = lc_positions['Topholey']

        # Bottomx와 Bottomy 좌표 추출
        Bottomx = lc_positions['Bottomx']
        Bottomy = lc_positions['Bottomy']    

        # 필요한 경우 다른 홀들에 대해서도 circle_cross 함수를 호출할 수 있음
        # 예를 들어, 모든 상단 홀에 대해 반복
        for x in Topholex:
            circle_cross(doc, abs_x + x, abs_y + Topholey, 11, layer='레이져')    
            # print(f"topholex x좌표: {x} y좌표 {abs_y + Topholey} abs_y좌표 : {abs_y}  Topholey : {Topholey}")     

        for x in Bottomx:
            circle_cross(doc, abs_x + x, abs_y + Bottomy, 11, layer='레이져')            

        # 세로 홀 위치 좌표 계산
        hole_positions = calculate_lc_vertical_hole_positions_035(inputCW, inputCD)

        # 각 홀의 좌표를 추출
        leftholex = hole_positions['leftholex']
        rightholex = hole_positions['rightholex']
        vertical_holey = hole_positions['vertical_holey']

        # 리스트 길이 확인
        num_holes = min(len(leftholex), len(rightholex), len(vertical_holey))

        # circle_cross 함수 호출 시 리스트 길이를 확인하여 오류 방지
        for i in range(num_holes):
            circle_cross(doc, abs_x + leftholex[i], abs_y + vertical_holey[i], 11, layer='레이져')
            circle_cross(doc, abs_x + rightholex[i], abs_y + vertical_holey[i], 11, layer='레이져')
            # print(f"leftholex x좌표: {leftholex[i]} abs_x좌표 {abs_y }  vertical_holey : {vertical_holey[i]}")    

        # assy smps 2가지 가져오기
        # smps가 안들어가는 T5는 제외
        if T5_standard  < 1 :    
            x = abs_x + 32 
            y = abs_y + 125
            insert_block(x , y , "lc035_assy_smps")      
            x = abs_x + CW - 98 - 32
            y = abs_y + 125
            insert_block(x , y , "lc035_assy_smps")     

        if CD-70 == LedBarLength_CD:
            LedSpace_CD = 49
        else:
            LedSpace_CD = 49 - ( LedBarLength_CD-(CD-70) ) / 2

        if CW-100 == LedBarLength_CW:
            LedSpace_CW = 57
        else:
            LedSpace_CW = 57 - ( LedBarLength_CW-(CW-100) ) / 2

        # 상부 치수선
        x1 = abs_x 
        y1 = abs_y + CD
        x2 = abs_x + CW
        y2 = y1    
        textstr =  f"L/C SIZE =  {CW-50}"
        dim(doc, x1 + Xgap , y1 - 25, x2 - Xgap , y2 - 25, 160,  text_height=0.32, text_gap=0.07, direction="up",  text = textstr)   

        x1 = abs_x 
        y1 = abs_y + CD
        x2 = abs_x + CW
        y2 = y1    
        textstr =  f"CAR INSIDE(CW) = {CW}"
        dim(doc, x1 , y1 , x2  , y2, 240,  text_height=0.32, text_gap=0.07, direction="up", text = textstr)    

        # 30 치수 표현
        dim(doc,  x1+Xgap, y2,  x1, y1, 185,  text_height=0.30, text_gap=0.07, direction="up")    
        dim(doc, x2-Xgap, y1, x2, y2, 185,  text_height=0.30, text_gap=0.07, direction="up")
    
        # 우측 치수선  
        x1 = abs_x + CW - Xgap
        y1 = abs_y + CD - 25
        x2 = abs_x + CW - 110
        y2 = y1 - 15
        dim(doc, x1, y1, x2, y2, 150 , text_height=0.30,  direction="right")        

        x1 = abs_x + CW - 110
        y1 = abs_y + 40
        x2 = abs_x + CW - Xgap   
        y2 = y1 - 15
        dim(doc, x1, y1, x2, y2, 150 + 25 , text_height=0.30,  direction="right")

        # 25 간격 치수선 표기
        x1 = abs_x + CW 
        y1 = abs_y + CD 
        x2 = x1 - Xgap
        y2 = y1 - 25
        dim(doc, x1, y1, x2, y2, 80, text_height=0.30,  direction="right")

        x1 = abs_x + CW - Xgap
        y1 = abs_y + 25
        x2 = abs_x + CW
        y2 = abs_y
        dim(doc, x1, y1, x2, y2, 150, text_height=0.30,  direction="right")

        # 상단, 하단 홀의 오른쪽 끝 y 좌표 추출
        right_end_topholey = Topholey
        right_end_bottomy = Bottomy

        # 세로 방향 홀의 y 좌표 추출
        vertical_holes_rightholey = vertical_holey

        # 모든 오른쪽 끝 홀의 y 좌표를 하나의 리스트로 결합
        all_right_end_holes_y_coords = [right_end_topholey] + vertical_holes_rightholey + [right_end_bottomy]

        # 모든 오른쪽 끝 홀의 y 좌표를 오름차순으로 정렬
        sorted_y_coords = sorted(all_right_end_holes_y_coords)

        # 정렬된 y 좌표 출력
        last_ypos = 40
        dim(doc, abs_x  , abs_y , abs_x + 110 , abs_y +  last_ypos, 120, text_height=0.30,  direction="left")
        for i, y_coord in enumerate(sorted_y_coords):        
            # print(f"세로 홀의 y 좌표: {y_coord}")
            if i>0:
                dimcontinue(doc,  abs_x + 110 , abs_y +  y_coord)
                last_ypos = y_coord
        dimcontinue(doc,  abs_x , abs_y + CD , distance = 120, option='reverse' )             

        x1 = abs_x + CW - Xgap
        y1 = abs_y + CD - 25
        x2 = x1
        y2 = abs_y + 25        
        textstr =  f"LIGHT CASE SIZE = {CD-panel_thickness*2}"
        # dim_vertical_right_string(doc, x1, y1, x2, y2, 308, textstr , text_height=0.20,  text_gap=0.07)        
        dim(doc,   x2, y2, x1, y1,   200, text_height=0.32, direction="right", text = textstr)    
        x1 = abs_x + CW
        y1 = abs_y + CD
        x2 = x1
        y2 = abs_y         
        textstr =  f"CAR INSIDE {CD}"    
        dim(doc,  x2, y2, x1, y1,  260, text_height=0.32, direction="right", text = textstr)     

        # 하부 치수선    
        last_xpos = 110
        dim(doc, abs_x  , abs_y , abs_x + last_xpos , abs_y + 40, 110,  direction="down")
        for i, x_coord in enumerate(Topholex):        
            # print(f"가로 홀의 x 좌표: {x_coord}")
            if i>0:
                dim(doc, abs_x + last_xpos , abs_y + 40, abs_x + x_coord , abs_y + 40, 150,  direction="down")
                last_xpos = x_coord    
        dim(doc, abs_x + last_xpos   , abs_y + 40 , abs_x + CW , abs_y, 150,  direction="down")

        # 하부 전체치수
        x1 = abs_x 
        y1 = abs_y 
        x2 = abs_x  + CW
        y2 = abs_y         
        textstr =  f"CAR INSIDE(CW) = {CW}"    
        dim(doc,  x1, y1, x2, y2,  200, text_height=0.30, direction="down", text = textstr)                 

    #########################################################################
    # 우측단면도 표기 (블럭 삽입)
    #########################################################################
    program = 1
    if program:
        x = abs_x + CW + 450
        y = abs_y + CD
        insert_block(x , y , "lc035_assy_side_section_top")         
        x1 = x
        y1 = abs_y 
        insert_block(x1 , y1 , "lc035_assy_side_section_bottom")         
        x2 = x + 113.2
        y2 = abs_y + CD/2 
        insert_block(x2 , y2 , "lc035_assy_side_section_midbolt")      

        rectangle(doc, x-2.3, y + 30 , x, y1 - 30, layer="0")

        # 폴리갈 or 중판 크기 산출
        first_space = 46
        midplate_width = (CD - 47.1 - 235.95 - 48.95)/2
        polygal_height = LCD - 6    

        # print (f"단면도 midplate_width : {midplate_width}")

        # 녹색 중판 표현하기 (상부)
        line(doc, x + 96.2, y - 89,  x + 96.2, y - 89 + 20, layer="1")
        lineto(doc,  x + 96.2 + 34, y - 89 + 20, layer="1")
        lineto(doc,  x + 96.2 + 34, y - CD/2  , layer="1")
        lineto(doc,  x + 96.2 + 34 - 34, y - CD/2 , layer="1")
        lineto(doc,  x + 96.2 + 34 - 34, y - CD/2  - 10 , layer="1")
        lineto(doc,  x + 96.2 + 34 - 34 - 1.2, y - CD/2  - 10 , layer="1")
        lineto(doc,  x + 96.2 + 34 - 34 - 1.2, y - CD/2  - 10 + 11.2 , layer="1")
        lineto(doc,  x + 96.2 + 34 - 34 - 1.2, y - CD/2  - 10 + 11.2 , layer="1")
        lineto(doc,  x + 96.2 + 34 - 34 - 1.2 + 34, y - CD/2 - 10 + 11.2 , layer="1")
        lineto(doc,  x + 96.2 + 34 - 34 - 1.2 + 34, y  - 69 - 1.2 , layer="1")
        lineto(doc,  x + 96.2 + 34 - 34 - 1.2 + 34 - 31.6, y  - 69 - 1.2 , layer="1")
        lineto(doc,  x + 96.2 + 34 - 34 - 1.2 + 34 - 31.6, y  - 69 - 1.2 - 18.8 , layer="1")
        lineto(doc,  x + 96.2 + 34 - 34 - 1.2 + 34 - 31.6 - 1.2, y  - 69 - 1.2 - 18.8 , layer="1")

        # 녹색 중판 표현하기 (하부)
        line(doc, x + 96.2, y1 + 89,  x + 96.2, y1 + 89 - 20, layer="1")
        lineto(doc,  x + 96.2 + 34, y1 + 89 - 20, layer="1")
        lineto(doc,  x + 96.2 + 34, y1 + CD/2, layer="1")
        lineto(doc,  x + 96.2  , y1 + CD/2, layer="1")
        lineto(doc,  x + 96.2  , y1 + CD/2 - 10, layer="1")
        lineto(doc,  x + 96.2 + 1.1 , y1 + CD/2 - 10, layer="1")
        lineto(doc,  x + 96.2 + 1.1 , y1 + CD/2 - 1.2, layer="1")
        lineto(doc,  x + 96.2 + 32.8 , y1  + CD/2 - 1.2, layer="1")
        lineto(doc,  x + 96.2 + 32.8 , y1 + 89 - 20 + 1.2, layer="1")
        lineto(doc,  x + 96.2 + 1.2 , y1 + 89 - 20 + 1.2, layer="1")
        lineto(doc,  x + 96.2 + 1.2, y1 + 89 , layer="1")
        lineto(doc,  x + 96.2 , y1 + 89 , layer="1")

        # 흰색선 삽입
        line(doc, x + 105,  y - 67,  x + 105, y1 + 67, layer="0")

        x7 = x + 135
        y7 = y 
        x8 = x7 
        y8 = y7 - 25        
        x9 = x + 135
        y9 = y1 
        x10 = x9 
        y10 = y9 + 25          
        
        # 25 + 42 중심 치수선
        dim(doc, x10 , y10+42,  x8, y8-42,  120 - 3.86, text_height=0.3, direction="right")

        textstr = f"Light Case Wide (D) = {LCD}"
        dim(doc, x10 , y10,  x8, y8,  200, text_height=0.3, direction="right", text=textstr)
        
        textstr = f"Car Inside (CD) = {CD}"
        dim(doc, x9, y9, x7 , y7,  310, text_height=0.3, direction="right", text=textstr)

    ####################################################################################
    # lc035 모델은 새로운 중판모양을 우측 화면에 표시한다. 
    # 기존의 공식을 이용해서 그린다.
    ####################################################################################
    # 1page Assy 본판 작도 코드 재사용
    #################################
    program = 1
    if program:
        Right_abs_x = abs_x + CW + 1300
        Right_abs_y = abs_y 
        
        # ALprofile_draw
        # 첫 번째 구간
        startx = Right_abs_x + 25
        starty = Right_abs_y + CD - 25    
        ALprofile_draw(doc, startx, starty, altube_CD, direction="side")    
        startx = Right_abs_x + 25 + 42        
        ALprofile_draw(doc, startx, starty, altube_CW, direction="topbottom")
        startx = Right_abs_x + CW - 25 - 42    
        ALprofile_draw(doc, startx, starty, altube_CD, direction="side")
        startx = Right_abs_x + 25 + 42   
        starty = Right_abs_y + 25 + 42
        ALprofile_draw(doc, startx, starty, altube_CW, direction="topbottom")

        # 기본 윤곽 car inside
        width_insideGap = 25
        height_insideGap = 25
    
        x1 = Right_abs_x 
        y1 = Right_abs_y 
        x2 = Right_abs_x + CW
        y2 = Right_abs_y + CD    
        rectangle(doc, x1, y1, x2, y2, layer='hidden')    
        # 중판 녹색선 표현     
        rectangle(doc, x1+69, y1+69, x2-69, y2-69, layer='1')             
        line(doc, x1+69, y1+CD/2, x2-69, y1+CD/2, layer='1')      
        textstr =  f"중판 STS M/R = {math.floor(CW-69*2)}"            
        dim(doc, x1+69, y1+CD/2, x2-69, y1+CD/2, 150,  text_height=0.30,  direction="down",  text = textstr)    
        textstr =  f"중판 STS M/R = {math.floor(CD/2-69)}"               
        dim(doc, x1+CD/2, y2 - 69, x1+CD/2, y1+CD/2,  350, text_height=0.30,  direction="left",  text = textstr)              
        dim(doc, x1+CD/2, y1 + 69, x1+CD/2, y1+CD/2,  350, text_height=0.30,  direction="left",  text = textstr)             
        
        x1 = Right_abs_x + width_insideGap
        y1 = Right_abs_y + height_insideGap
        x2 = Right_abs_x + CW - width_insideGap
        y2 = Right_abs_y + CD - height_insideGap
        rectangle(doc, x1, y1, x2, y2, layer='0')    
        rectangle(doc, x1+42, y1+42, x2-42, y2-42, layer='0')       

        #####################################################
        # 상부 치수선
        #####################################################
        x1 = Right_abs_x + 25 
        y1 = Right_abs_y + CD 
        x2 = Right_abs_x + CW - 25
        y2 = y1    
        
        dim(doc,  x1+42  , y1-25,  x1 , y1-25 ,130,  direction="up", option='reverse')
        dimcontinue(doc, x2-42 , y2 - 25)
        dimcontinue(doc, x2 , y2 - 25)

        #####################################################
        # 하부 치수선
        #####################################################
        x1 = Right_abs_x 
        y1 = Right_abs_y  
        x2 = Right_abs_x + CW
        y2 = y1 + CD   
        textstr =  f"L/C SIZE =  {LCW}"
        dim(doc, x1 + 25 , y1 + 25, x2 - 25 , y1 + 25, 160,  text_height=0.30,  direction="down",  text = textstr)   
        textstr =  f"CAR INSIDE(CW) = {CW}"
        dim(doc, x1 , y1 , x2  , y2, 240,  text_height=0.30, direction="down", text = textstr)    

        # 좌우 25 치수 표현
        dim(doc,  x1+25, y1 + 25,  x1, y1, 135,  text_height=0.30, text_gap=0.07, direction="down")    
        dim(doc, x2-25, y1 + 25, x2, y1, 135,  text_height=0.30, text_gap=0.07, direction="down")

        #####################################################
        # 좌측 치수선
        #####################################################
        x1 = Right_abs_x 
        y1 = Right_abs_y + CD 
        x2 = x1
        y2 = Right_abs_y 
        textstr =  f"L/C SIZE =  {LCW}"    
        dim(doc, x1 , y1 , x1 +25  , y1 - 25, 180 ,  direction="left")
        dim(doc, x1+25 , y1-25 , x2+25  , y2 + 25, 180 ,  text_height=0.25, text_gap=0.07, direction="left", text = textstr)            
        dim(doc, x2 , y2 , x2 + 25  , y2 + 25, 180 ,  direction="left")

        textstr =  f"Car Inside (D) = {math.ceil(CD)}"
        dim(doc, x1 , y1 , x2  , y2, 280 ,  text_height=0.25, text_gap=0.07, direction="left", text = textstr)            
    
        #####################################################
        # 우측 치수선
        #####################################################
        # 42 간격 치수선 표기

        x1 = Right_abs_x + CW  
        y1 = Right_abs_y + CD 
        x2 = x1
        y2 = Right_abs_y 
        dim(doc,  x1-25-42, y1-25-42,   x1-25, y1-25,   235, text_height=0.30, direction="right", option='reverse')
        dimcontinue(doc, x2-25-42, y2+25+42   )
        dimcontinue(doc, x2-25, y2+25   )
        
        x1 = Right_abs_x + CW
        y1 = Right_abs_y + CD
        x2 = x1
        y2 = Right_abs_y         
        textstr =  f"CAR INSIDE(CD) {CD}"    
        dim(doc,  x2, y2, x1, y1,  300, text_height=0.32, direction="right", text = textstr)   

        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1715
        TargetXscale = 5500 + (CW - 1600) * 1.6 + (CD - 1500) * 1.5
        TargetYscale = 3400 + (CD - 1500)
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("스케일 비율 : " + str(frame_scale))        
        frameYpos = abs_y - 450 * frame_scale     
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)       
        # 1page 상단에 car inside 표기
        x = frameXpos + CW/3
        y = abs_y + frame_scale*BasicYscale + 50
        textstr = f"W{CW} x D{CD}"    
        draw_Text(doc, x, y , 200*frame_scale, str(textstr), layer='0')    

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 2page 프레임 본체만들기 (레이져가공)
    #################################################################################################        
    program = 1
    if program:
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = 2940 + (CD - 1500)  # 기본 모형이 1450
        TargetYscale = 1800 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("2page 스케일 비율 : " + str(frame_scale))           
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

        rx1 = frameXpos  + 410  + LCD + 600
        ry1 = frameYpos + 1100 * frame_scale
        insert_block(rx1 , ry1 , "lc035_frame_section_side")    

        rx1 = frameXpos  + 500
        
        thickness = 1.2
        vcut = thickness / 2

        x1 = rx1 + 40.8
        y1 = ry1
        x2 = x1 + LCD - 40.8*2
        y2 = y1    
        x3 = x2 
        y3 = y2 + 23.2 - vcut
        x4 = x3
        y4 = y3 + 10 - vcut*3
        x5 = x4 + 40.8
        y5 = y4 
        x6 = x5 
        y6 = y5 + vcut
        x7 = x6 
        y7 = y6 + 42 - thickness
        x8 = x7 
        y8 = y7 + 45 - thickness
        x9 = x8 
        y9 = y8 + 91.2
        x10 = x9 
        y10 = y9 + 29.6
        x11 = x10 - LCD
        y11 = y10
        x12 = x11 
        y12 = y11 - 29.6
        x13 = x12 
        y13 = y12 - 91.2
        x14 = x13 
        y14 = y13 - 45 + thickness
        x15 = x14 
        y15 = y14 - 42 + thickness
        x16 = x15
        y16 = y15 - vcut
        x17 = x16 + 40.8
        y17 = y16
        x18 = x17
        y18 = y17 - 10 + vcut*3

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 18
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
        
        #절곡라인  (히든선 : 정방향)      
        line(doc, x15, y15, x6, y6,  layer='hidden')     
        line(doc, x14, y14, x7, y7,  layer='hidden')     
        line(doc, x13, y13, x8, y8,  layer='hidden')     
        line(doc, x12, y12, x9, y9,  layer='hidden')     

        #절곡라인  (실선 : 역방향)      
        line(doc, x18, y18, x3, y3,  layer='2')  

        # 양쪽 전선홀 삽입 80x24 round   
        insert_block(rx1+10 , y15+6.9 , "lc035_27x80_roundrectangle")    
        insert_block(rx1+LCD - 10 - 80 , y15+6.9 , "lc035_27x80_roundrectangle")    

        # 상부 치수선 및 LCD방향 라이트케이스 홀 삽입 11파이
        # 프레임 상부 정렬된 x 좌표 출력 (CD방향 단공 배열을 이용함)    
        dim(doc, x11  , y11 , x10 , y10, 125, direction="up")    
        for i, y_coord in enumerate(sorted_y_coords):  
            xval =  y_coord - 25     
            # print(f"CD방향 홀의 i변수 {i}  , y 좌표: {xval}")        
            circle_cross(doc, x11 + xval , y11 - 15 , 11, layer='레이져')
            if i>0 :
                dimcontinue(doc,  x11 + xval , y11 - 15 )
            else:
                dim(doc,  x11+15, y11 - 15 , x11 , y11 , 70, direction="up", option="reverse") 
            if i==1:
                baseX, baseY = x11 + xval , y11 - 15

        dimcontinue(doc,  x10 , y10 )

        # 11파이 지시선
        dim_leader_line(doc, baseX, baseY , baseX + 100 ,baseY + 50, f"{len(sorted_y_coords)} -%%C11 Hole (M8 Pop Nut)")       
        dim(doc,   baseX, baseY , baseX  , y11, 100, direction="left")

        # 하부 단공치수, 배열된 것을 구현함 결합되는 홀임
        lc035_vertical_holes = calculate_hole_vertical_lc035(LCD)
        # print(f"035 lc035_vertical_holes {lc035_vertical_holes}")
        
        for i, y_coord in enumerate(lc035_vertical_holes):  
            xval =  y_coord 
            y = y1 + 10
            # print(f"CD 방향 중판 홀의 i변수 {i}  , x 좌표: {xval}")        
            circle_cross(doc, x15 + xval , y , 10, layer='레이져')
            if i>0 :
                dimcontinue(doc,  x11 + xval , y )
            else:
                dim(doc, x1+13.2 , y1+10,  x1  , y1 , 200, direction="down", option='reverse') 
            if i==3:
                baseX, baseY = x15 + xval , y

        dimcontinue(doc,  x2 , y2 )

        # 10파이 지시선
        dim_leader_line(doc, baseX, baseY , baseX + 50 ,baseY - 30, f"{len(lc035_vertical_holes)} -%%C10 Hole")       
        dim(doc,   baseX, baseY , baseX  , y2, 100, direction="left")        

        # 하부 클립홀 15파이 3개~4개 타공
        lc035_clipholes = calculate_cliphole_lc035(225, LCD)
        # print(f"035 Clip hole {lc035_clipholes}")
        
        for i, y_coord in enumerate(lc035_clipholes):  
            xval =  y_coord 
            y = y15 + 21 - vcut
            # print(f" 중판 clip 홀의 i변수 {i}  , x 좌표: {xval}")        
            circle_cross(doc, x15 + xval , y , 15, layer='레이져')
            if i>0 :
                dimcontinue(doc,  x15 + xval , y )
            else:
                dim(doc, x15+225, y,  x15  , y15 , 150, direction="down", option='reverse') 
            if i==1:
                baseX, baseY = x15 + xval , y

        dimcontinue(doc,  x5 , y5 )
        dim_leader_line(doc, baseX, baseY , baseX + 50 ,baseY + 50, f"{len(lc035_clipholes)} -%%C15 Hole")       
        dim(doc,   baseX, baseY , baseX  , y1, 100, direction="left")

        # print(f"lc035_vertical_holes 개수 {len(lc035_vertical_holes)} ")        

        # 하부 전체치수
        dim(doc, x1 , y1,  x16  , y16 , 280, direction="down", option='reverse') 
        dimcontinue(doc, x2 , y2)
        dimcontinue(doc, x5 , y5)

        # 좌측치수선
        dim(doc,  x11, y11 , x16 , y16 , 150,  direction="left")           
        dimcontinue(doc, x1 , y1, option='reverse')

        # 오른쪽 치수
        dim(doc,  x9, y9 , x10 , y10 , 110,  direction="right")         
        dimcontinue(doc,  x8, y8 )
        dim(doc,  x7, y7 , x8 , y8 , 160,  direction="right")         

        dim(doc,  x6, y6 , x7 , y7 , 50,  direction="right")             
        dim(doc,  x6, y6 , x3 , y3 , 50,  direction="right")        
        dim(doc,  x3, y3 , x2 , y2 , 150,  direction="right")        
        dim(doc,  x10, y10 , x2 , y2 , 220,  direction="right")        

        # main Frame description    
        x1 = frameXpos + 950 + (CD-1500)
        y1 = frameYpos + 600 * frame_scale
        textstr = f"Part Name : Light Case (좌, 우)"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : {LCframeMaterial} "    
        draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
        textstr = f"Size : {LCD} x {str(math.floor((y10-y2)*10)/10)}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su*2} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')         

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 3page 프레임 본체만들기 (레이져가공) - 앞뒤판 프레임
    #################################################################################################        
    program = 1
    if program:
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = 2940 + (CD - 1500)  # 기본 모형이 1450
        TargetYscale = 1800 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("3page 스케일 비율 : " + str(frame_scale))           
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)     

        rx1 = frameXpos  + 400  + LCW + 600
        ry1 = frameYpos + 1100 * frame_scale    
        insert_block(rx1 , ry1 , "lc035_frame_section_front")      

        rx1 = frameXpos  + 500
        
        thickness = 1.2
        vcut = thickness / 2

        LCW = CW - 53

        x1 = rx1 + 62.5
        y1 = ry1
        x2 = x1  + LCW - 62.5*2
        y2 = y1    
        x3 = x2
        y3 = y2 + 23.2 - vcut
        x4 = x3
        y4 = y3 + vcut
        x5 = x4 + 22
        y5 = y4 
        x6 = x5 
        y6 = y5 + 10 - vcut*3
        x7 = x6 
        y7 = y6 + 42 - thickness
        x8 = x7 
        y8 = y7 + vcut
        x9 = x8 + 40.5
        y9 = y8 
        x10 = x9 
        y10 = y9  + 43.11
        x11 = x10 - 69.06
        y11 = y10 + 59.19
        x12 = x11 - 29.94
        y12 = y11
        x13 = x12 
        y13 = y12 + vcut/2
        x14 = x13 
        y14 = y13 + 30 - thickness
        x15 = x14 - LCW + 99*2
        y15 = y14
        x16 = x15 
        y16 = y15 - 30 + thickness
        x17 = x16 
        y17 = y16 - vcut/2
        x18 = x17 - 29.94
        y18 = y17 
        x19 = x18 - 69.06
        y19 = y18 - 59.19
        x20 = x19
        y20 = y19 - 43.11
        x21 = x20 + 40.5
        y21 = y20
        x22 = x21 
        y22 = y21 - vcut
        x23 = x22
        y23 = y22 - 42 + thickness
        x24 = x23 
        y24 = y23 - 10 + vcut*3
        x25 = x24 + 22
        y25 = y24
        x26 = x25
        y26 = y25 - vcut

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 26
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
        
        #절곡라인  (히든선 : 정방향)      
        line(doc, x23, y23, x6, y6,  layer='hidden')     
        line(doc, x22, y22, x7, y7,  layer='hidden')     
        line(doc, x16, y16, x13, y13,  layer='hidden')     

        #절곡라인  (실선 : 역방향)      
        line(doc, x26, y26, x3, y3,  layer='2')  

        # 양쪽 전선홀 삽입 80x24 round   
        insert_block(x23+10 , y23+6.9 , "lc035_27x80_roundrectangle")    
        insert_block(x6 - 10 - 80 , y6+6.9 , "lc035_27x80_roundrectangle")    

        # 상부 치수선 및 LCD방향 라이트케이스 홀 삽입 11파이
        # 프레임 상부 정렬된 x 좌표 출력 (CD방향 단공 배열을 이용함)    
        dim(doc, x19  , y19 , x10 , y10, 300, direction="up")    
        
        modified_Topholex = Topholex[1:-1]
        # print(f"W방향 Topholex {Topholex}")   
        # print(f"W방향 modified_Topholex {modified_Topholex}")   
        for i, y_coord in enumerate(modified_Topholex):  
            xval = y_coord - 15.5 - 110 # 첫홀 간격 빼줌
            y = y15 - 15        
            # print(f"W방향 홀의 i변수 {i}  , y 좌표: {xval}")        
            circle_cross(doc, x15 + xval , y , 11, layer='레이져')        
            if i== 0:
                dim(doc,  x15 + xval  , y  , x15   , y15 , 100, direction="up", option="reverse") 
            else:
                dimcontinue(doc,  x15 + xval , y )        
            
            if i==0:
                baseX, baseY = x15 + xval , y

        dimcontinue(doc,  x14 , y14 )

        # 11파이 지시선
        dim_leader_line(doc, baseX, baseY , baseX + 100 ,baseY + 50, f"{len(modified_Topholex)} -%%C11 Hole (M8 Pop Nut)")       
        dim(doc,   baseX, baseY , baseX  , y15, 100, direction="left")

        # 하부 클립홀 15파이 3개~4개 타공
        lc035_front_clipholes = calculate_cliphole_lc035( 233, CW-134)
        # print(f"035 Clip hole {lc035_front_clipholes}")
        
        for i, y_coord in enumerate(lc035_front_clipholes):  
            xval =  y_coord 
            y = y23 + 21 - vcut
            # print(f" 중판 clip 홀의 i변수 {i}  , x 좌표: {xval}")        
            circle_cross(doc, x23 + xval , y , 15, layer='레이져')
            if i>0 :
                dimcontinue(doc,  x23 + xval , y )
            else:
                dim(doc, x24+233, y,  x23  , y23 , 120, direction="down", option='reverse') 
            if i==1:
                baseX, baseY = x23 + xval , y

        dimcontinue(doc,  x5 , y5 )
        dim_leader_line(doc, baseX, baseY , baseX + 50 ,baseY + 50, f"{len(lc035_front_clipholes)} -%%C15 Hole")       
        dim(doc,   baseX, baseY , baseX  , y1, 50, direction="left")

        # 하부 중판과 연결홀 파이11  4개
        lc035_midplateholes = calculate_midplatehole_lc035(CW-178)  #178 1600일때  1422 실제 중판W 계산원리
        # print(f"035 lc035_midplateholes {lc035_midplateholes}")
        
        for i, coord in enumerate(lc035_midplateholes):  
            xval = x1 + coord 
            y = y1 + 10
            # print(f" 중판 lc035_midplateholes 홀의 i변수 {i}  , x 좌표: {xval}")        
            circle_cross(doc, xval , y , 10, layer='레이져')
            if i>0 :
                dimcontinue(doc,  xval , y )
            else:
                dim(doc, xval, y,  x1 , y1 , 155, direction="down", option='reverse') 
            if i==2:
                baseX, baseY = xval , y

        dimcontinue(doc,  x2 , y2 )
        dim_leader_line(doc, baseX, baseY , baseX + 50 ,baseY - 20, f"{len(lc035_midplateholes)} -%%C10 Hole")       
        dim(doc,   baseX  , y1, baseX, baseY , 50, direction="left")
    
        # 상부 좌측 69... 치수
        dim(doc,  x18  , y18  , x19   , y19 , 150, direction="up")
        # 상부 2번째 행 3개치수선
        dim(doc,  x15  , y15  , x19   , y19 , 150, direction="up", option='reverse')
        dimcontinue(doc,  x14 , y14 )
        dimcontinue(doc,  x10 , y10 )

        # 좌측치수선
        dim(doc, x19 , y19 , x20, y20 , 55,  direction="left")
        dim(doc, x15 , y15 , x18, y18 , 220,  direction="left")
        dimcontinue(doc, x20 , y20)
        dim(doc, x20, y20 ,x24 , y24,  180,  direction="left")    
        dimcontinue(doc, x1 , y1, distance=242.5, option='reverse')

        # 오른쪽 치수
        dim(doc,  x13, y13 , x14 , y14 , 150,  direction="right", option='reverse')           
        dimcontinue(doc,  x7, y7 )
        dim(doc,  x7, y7 , x6 , y6 , 90,  direction="right")         

        dim(doc,  x3, y3 , x6 , y6 , 170,  direction="right")                 
        dim(doc,  x3, y3 , x2 , y2 , 200,  direction="right")          
        # 오른쪽 전체치수
        dim(doc,  x14, y14 , x2 , y2 , 250,  direction="right")          

        # 하부 전체치수
        dim(doc, x1 , y1,  x24  , y24 , 210, direction="down", option='reverse') 
        dimcontinue(doc, x2 , y2)
        dimcontinue(doc, x5 , y5)    
        # 하부 위에서 4행 치수선
        dim(doc, x24 , y24,  x20  , y20 , 300, direction="down", option='reverse') 
        dimcontinue(doc, x5 , y5)
        dimcontinue(doc, x9 , y9)    
        # 하부 위에서 5행 치수선    
        dim(doc, x1, y1, x20  , y20 ,  350, direction="down")
        dim(doc, x2, y2, x9  , y9 ,  350, direction="down")    
        
        # main Frame description    
        x1 = frameXpos + 950 + (CW-1600)
        y1 = frameYpos + 600 * frame_scale
        textstr = f"Part Name : Light Case (앞, 뒤)"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : {LCframeMaterial}"    
        draw_Text(doc, x1 , y1 -60 , 30, str(textstr), layer='0')    
        textstr = f"Size : {LCW+0.69} x {str(math.floor((y14-y2)*10)/10)}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su*2} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')    

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 4page 중판 (앞쪽)  'ㄷ'자 형태에서 날개 한번 떠 꺾인 형태 2R 라운드 있음
    #################################################################################################
    program = 1
    if program:         
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = 2730 + (CW - 1600) 
        TargetYscale = 1665 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("4page 스케일 비율 : " + str(frame_scale))           
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

        rx1 = frameXpos  + 710
        ry1 = frameYpos + 630 * frame_scale
        
        thickness = 1.2
        vcut = thickness / 2

        # 중판 전개도 크기 정의
        midplate_su = 2
        midplate_width = CW - 138
        midplate_height = (CD-69*2)/midplate_su 

        if(CD>=2000):        
            midplate_su = 3

        topwing1 = 10
        topwing2 = 34
        btwing1 = 20
        btwing2 = 34
        leftwing1 = 20
        leftwing2 = 34
        rightwing1 = 20
        rightwing2 = 34

        x1 = rx1 
        y1 = ry1
        x2 = x1 + midplate_width - (leftwing1) - (rightwing1) - 1 
        y2 = y1    
        x3 = x2
        y3 = y2 + btwing1 - vcut 
        x4 = x3 
        y4 = y3 + vcut * 1.5
        x5 = x4 + (rightwing1-1)
        y5 = y4
        x6 = x5
        y6 = y5 + rightwing2 - 2.7 
        x7 = x6 + rightwing1 + rightwing2 - vcut * 1.5
        y7 = y6
        x8 = x7 
        y8 = y7 + vcut
        x9 = x8 
        y9 = y8 + midplate_height - vcut*2
        x10 = x9
        y10 = y9 + vcut 
        x11 = x10 - rightwing1 - rightwing2 + vcut * 1.5
        y11 = y10 
        x12 = x11 
        y12 = y11 + topwing2 - vcut*2 - 1.5
        x13 = x12 - rightwing1 + 1
        y13 = y12 
        x14 = x13 
        y14 = y13 + vcut * 1.5
        x15 = x14 
        y15 = y14 + topwing1 - vcut
        x16 = x15 - midplate_width + leftwing1 + rightwing1 + 1
        y16 = y15 
        x17 = x16 
        y17 = y16 - topwing1  + vcut
        x18 = x17
        y18 = y17 - vcut * 1.5
        x19 = x18 - leftwing1 + 1
        y19 = y18 
        x20 = x19 
        y20 = y19 - topwing2 + vcut*2 + 1.5
        x21 = x20 - rightwing1 - rightwing2 + vcut * 1.5 
        y21 = y20 
        x22 = x21 
        y22 = y21 - vcut
        x23 = x22
        y23 = y22 - midplate_height + vcut*2
        x24 = x23 
        y24 = y23 - vcut
        x25 = x24 + leftwing1 + leftwing2 - vcut * 1.5 
        y25 = y24
        x26 = x25 
        y26 = y25 - leftwing2 + 2.7 
        x27 = x26 + leftwing1 - 1
        y27 = y26 
        x28 = x27 
        y28 = y27 - vcut * 1.5

        # 세로 절곡점 찾기
        x29 = x24 + leftwing1 - vcut
        y29 = y24
        x30 = x25 - vcut*1.5
        y30 = y25 
        x31 = x6 + vcut*1.5
        y31 = y6
        x32 = x7 - rightwing1 + vcut
        y32 = y7

        x33 = x10 - rightwing1 + vcut
        y33 = y10
        x34 = x11 + vcut*1.5
        y34 = y11
        x35 = x20 - vcut*1.5
        y35 = y20
        x36 = x21 + leftwing1 - vcut
        y36 = y21

        # 전개도 사이즈 저장 #1 중판
        midplatesizex2 = x10 - x21 
        midplatesizey2 = y15 - y2

        # 중판1보다 4점 추가됨 주의
        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        # 2R 그리기
        Radiusval = 2          
        lastNum = 28
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            if i==0:
                line(doc, prev_x, prev_y , curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y            
            # if i==5 or i==12 or i==19 or i==26 :
            if i==5 :
                x = curr_x - Radiusval
                y = curr_y + Radiusval            
                line(doc, prev_x, prev_y, x , curr_y , layer='레이져')   # 우상향     
                add_arc_between_points(doc, (x, curr_y), (x + Radiusval, curr_y + Radiusval), Radiusval)
                x = curr_x 
                y = curr_y + Radiusval            
                prev_x, prev_y = x, y
            elif i==12 :
                x = curr_x - Radiusval
                y = curr_y - Radiusval            
                line(doc, prev_x, prev_y, curr_x , y , layer='레이져')   # 좌상향     
                add_arc_between_points(doc, (curr_x , y) , (x , curr_y) , Radiusval)
                x = curr_x -  Radiusval             
                y = curr_y 
                prev_x, prev_y = x, y
            elif i==19 :
                x = curr_x + Radiusval
                y = curr_y - Radiusval            
                line(doc, prev_x, prev_y, x , curr_y , layer='레이져')   # 우하향     
                add_arc_between_points(doc, (x , curr_y) , (curr_x , y) , Radiusval)
                x = curr_x 
                y = curr_y - Radiusval             
                prev_x, prev_y = x, y
            elif i==26 :
                x = curr_x + Radiusval
                y = curr_y + Radiusval            
                line(doc, prev_x, prev_y, curr_x , y ,  layer='레이져')   # 우하향     
                add_arc_between_points(doc,(curr_x , y) , (x , curr_y) , Radiusval)
                x = curr_x + Radiusval             
                y = curr_y 
                prev_x, prev_y = x, y
            else:            
                line(doc, prev_x, prev_y , curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y

        line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
        
        #절곡라인  밴딩 가로 방향      
        line(doc, x28, y28, x3, y3,  layer='hidden')   # 절곡선      
        line(doc, x23, y23, x8, y8,  layer='hidden')   # 절곡선      
        line(doc, x22, y22, x9, y9,  layer='hidden')   # 절곡선      
        line(doc, x17, y17, x14, y14,  layer='hidden')   # 절곡선      

        # 세로절곡선 표현
        line(doc, x36, y36, x29, y29,  layer='hidden')    # 절곡선        
        line(doc, x35, y35, x30, y30,  layer='hidden')    # 절곡선        
        line(doc, x34, y34, x31, y31,  layer='hidden')    # 절곡선        
        line(doc, x33, y33, x32, y32,  layer='hidden')    # 절곡선        

        # 상부 7개 200피치 단공 9파이 
        #178+1을 빼준다. 1600일때  1422 실제 중판W 계산원리 양쪽 0.5씩 빼주는 형태임   
        midplate_connetholes = calculate_midplate_connecthole_lc035(CW - 34.8) # 1600 CW기준 34.8 전개도임     
                
        for i, coord in enumerate(midplate_connetholes):  
            xval = x22 + coord 
            y = y22 + topwing2/2 - vcut            
            # print(f" 중판 상부 i : {i}")    
            if i>0 and i<len(midplate_connetholes)-1 :
                circle_cross(doc, xval , y , 9, layer='레이져')     
                dc(doc,  xval , y )
            elif i==0:
                circle_cross(doc, xval , y , 9, layer='레이져')     
                d(doc, x21 , y21 , xval, y, 150, direction="down")

            if i==3:
                baseX, baseY = xval , y

        dc(doc,  x10 , y10 )

        dim_leader_line(doc, baseX, baseY , baseX + 150 ,baseY -50, f"{len(midplate_connetholes)} -%%C9 Hole")       
        dim(doc,   baseX  , y16,  baseX, baseY , 100, direction="left")
    
        # 상부 중판과 연결홀 200피치 9개 홀 
        #178+1을 빼준다. 1600일때  1422 실제 중판W 계산원리 양쪽 0.5씩 빼주는 형태임   
     
        for i, coord in enumerate(lc035_midplateholes):  
            xval = x1 + coord - 0.5
            y = y1 + 10
            # print(f" 중판 lc035_midplateholes 홀의 i변수 {i}  , x 좌표: {coord}")        
            circle_cross(doc, xval , y , 9, layer='레이져')
            if i>0 :
                dimcontinue(doc,  xval , y )
            else:
                dim(doc, xval, y,  x1 , y1 , 120, direction="down", option='reverse') 
            if i==0:
                baseX, baseY = xval , y

        dimcontinue(doc,  x2 , y2 )
        dim_leader_line(doc, baseX, baseY , baseX - 50 ,baseY +150, f"{len(lc035_midplateholes)} -%%C9 Hole(M6 Pop Nut)")       
        dim(doc,   baseX  , y1,  baseX, baseY , 50, direction="left")
    
        # 좌우 결합되는 단공, 배열된 것을 구현함 결합되는 홀임        
        # print(f"035 lc035_vertical_holes {lc035_vertical_holes}")
        
        half_of_lc035_vertical_holes = lc035_vertical_holes[:len(lc035_vertical_holes)//2]

        for i, coord in enumerate(half_of_lc035_vertical_holes):  
            xx = x24 + 10
            XX_right = x10 -10   # 오른쪽 홀
            y = y21 + 39 - coord # 단공 첫홀이 54가 15로 나와야 하니 + 39해줘야 한다.
            # print(f"CD 방향 중판 홀의 i변수 {i}  , x 좌표: {xval}")        
            circle_cross(doc, xx , y , 9, layer='레이져')
            circle_cross(doc, XX_right , y , 9, layer='레이져')
            if i>0 :
                dc(doc,  xx , y ,distance = 58 )
            else:
                d(doc, x21  , y21 , xx, y,   100, direction="left") 

        dc(doc,  x24 , y24, option='reverse')

        # 상부 치수선
        dim(doc, x36, y36, x21, y21 , 120 , direction="up", option="reverse" )
        dimcontinue(doc, x35, y35)
        dim(doc, x35, y35, x34, y34 , 160 , direction="up")    
        dim(doc, x33, y33, x10, y10 , 120 , direction="up", option="reverse" )
        dimcontinue(doc, x34, y34)    
        # 상부 전체치수
        dim(doc, x21, y21, x10, y10 , 210 , direction="up")

        # 좌측 치수선
        d(doc, x16,  y16, x21, y21 , 150 , direction="left")
        dc(doc, x24, y24)
        dc(doc, x1, y1, option='reverse')
        d(doc, x16, y16 , x1, y1,  290 , direction="left")

        # 하부 치수선
        dim(doc, x1, y1, x26 , y26 , 160 , direction="down" )
        dim(doc, x2, y2, x5 , y5 , 160 , direction="down" )

        dim(doc, x1, y1, x24 , y24 , 240 , direction="down" , option='reverse')
        dimcontinue(doc,  x2, y2)    
        dimcontinue(doc,  x7, y7)      
        dim(doc, x24, y24, x7 , y7 , 350 , direction="down")

        # 우측 치수선
        dim(doc, x14, y14, x15, y15,  160 , direction="right")
        dim(doc,  x9, y9, x14, y14, 160 , direction="right", option='reverse')
        dimcontinue(doc,  x8, y8)
        dimcontinue(doc,  x3, y3)
        dim(doc, x3, y3, x2, y2,  160 , direction="right")   
        dim(doc, x15, y15, x2, y2 , 310 , direction="right")

    # 좌측 단면도 그리기
    # 왼쪽 leftwing1 leftwing2
    # 오른쪽 날개 rightwing1 rightwing2
    # 본판 mainplate = mp 
    # 'ㄷ'자 형상 
    program = 1
    if program:
        mp = midplate_height

        startx = x1 - 503 - btwing2
        starty = y1 + btwing1 + topwing1 + topwing2

        x1 = startx
        y1 = starty
        x2 = x1 - thickness    
        y2 = y1    
        x3 = x2  
        y3 = y2 - btwing1 
        x4 = x3 + btwing2 
        y4 = y3 
        x5 = x4 
        y5 = y4 + midplate_height
        x6 = x5 - topwing2
        y6 = y5 
        x7 = x6 
        y7 = y6 - topwing1
        x8 = x7 + thickness
        y8 = y7 
        x9 = x8 
        y9 = y8 + topwing1 - thickness
        x10 = x9 + topwing2 - thickness *2
        y10 = y9 
        x11 = x10 
        y11 = y10 - midplate_height + thickness *2
        x12 = x11 - topwing2 + thickness*2
        y12 = y11

        insert_block(x1 , y1 , "lc035_popnut_leftbottom")   

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 12
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="0")     

        #vcut 원 그리기
        draw_circle(doc, x3, y3, 13, layer='0', color='6') 
        draw_circle(doc, x4, y4, 13, layer='0', color='6') 
        draw_circle(doc, x5, y5, 13, layer='0', color='6') 
        draw_circle(doc, x6, y6, 13, layer='0', color='6') 

        # 치수선
        d(doc, x3,y3, x4, y4, 110, direction="down", option='reverse')
        d(doc, x2,y2, x3, y3, 80, direction="left")
        dim(doc, x4,y4, x5, y5, 110, direction="right")
        dim(doc, x6,y6, x5, y5, 100, direction="up")
        dim(doc, x6,y6, x7, y7, 80, direction="left")

        dim_leader_line(doc, x4, y4 , x4 + 50 , y4-150, "Vcut 4개소")
   
    ###############################################################################
    # 상부 단면도 그리기
    ###############################################################################
    program = 1
    if program:    
        mp = midplate_width

        startx = rx1 + leftwing1
        starty = ry1 + 340 + midplate_height + leftwing2*3

        x1 = startx
        y1 = starty
        x2 = x1 - leftwing1
        y2 = y1    
        x3 = x2  
        y3 = y2 - leftwing2
        x4 = x3 + mp
        y4 = y3 
        x5 = x4 
        y5 = y4 + rightwing2
        x6 = x5 - rightwing1
        y6 = y5 
        x7 = x6 
        y7 = y6 - thickness
        x8 = x7 + rightwing1 - thickness
        y8 = y7 
        x9 = x8 
        y9 = y8 - rightwing2 + thickness*2
        x10 = x9 - mp + thickness *2
        y10 = y9 
        x11 = x10 
        y11 = y10 + leftwing2 - thickness *2
        x12 = x11 + leftwing1 - thickness
        y12 = y11

        insert_block(x1 , y1 , "lc035_popnut_topleft")       
        insert_block(x6 , y6 , "lc035_popnut_topright")       

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 12
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="0")     

        #vcut 원 그리기 4개소
        draw_circle(doc, x2, y2, 13, layer='0', color='6') 
        draw_circle(doc, x3, y3, 13, layer='0', color='6') 
        draw_circle(doc, x4, y4, 13, layer='0', color='6') 
        draw_circle(doc, x5, y5, 13, layer='0', color='6') 

        # 치수선
        dim(doc, x2,y2, x1, y1, 80, direction="up")
        dim(doc, x2,y2, x3, y3, 80, direction="left")
        dim(doc, x5,y5, x4, y4, 80, direction="right")
        dim(doc, x6,y6, x5, y5, 80, direction="up")
        dim(doc, x3,y3, x4, y4, 80, direction="down")
    
        dim_leader_line(doc, x4, y4 , x4 + 50 , y4 - 100, "Vcut 4개소")

        # 중판 #1 description    
        x1 = frameXpos + 600 + (midplate_width-960)
        y1 = frameYpos + 1100 +  (midplate_height-722)
        textstr = f"Part Name : 중판(앞쪽)"    
        draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : {LCplateMaterial}"    
        draw_Text(doc, x1 , y1 - 40 , 20, str(textstr), layer='0')    
        textstr = f"Size : {str(math.ceil(midplatesizex2*100)/100)} x {str(math.ceil(midplatesizey2*100)/100)}mm"    
        draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
        textstr = f"Quantity : {su} EA"    
        draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')        

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 5page 중판 (뒷쪽) 
    #################################################################################################        
    program = 1
    if program:         
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = 2730 + (CW - 1600)
        TargetYscale = 1665 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("5page 스케일 비율 : " + str(frame_scale))           
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

        rx1 = frameXpos  + 710
        ry1 = frameYpos + 630 * frame_scale
        
        thickness = 1.2
        vcut = thickness / 2

        # 중판 전개도 크기 정의
        midplate_su = 2
        midplate_width = CW - 138
        midplate_height = (CD-69*2)/midplate_su 

        if(CD>=2000):        
            midplate_su = 3

        topwing1 = 11.2
        topwing2 = 35.2
        btwing1 = 20
        btwing2 = 34
        leftwing1 = 20
        leftwing2 = 34
        rightwing1 = 20
        rightwing2 = 34

        x1 = rx1 
        y1 = ry1 + midplate_height + 93.2
        x2 = x1 + midplate_width - (leftwing1) - (rightwing1) - 1 
        y2 = y1    
        x3 = x2
        y3 = y2 - btwing1 + vcut               # + → -
        x4 = x3 
        y4 = y3 - vcut * 1.5                   # + → -
        x5 = x4 + (rightwing1 - 1)
        y5 = y4
        x6 = x5
        y6 = y5 - rightwing2 + 2.7             # + → -
        x7 = x6 + rightwing1 + rightwing2 - vcut * 1.5
        y7 = y6
        x8 = x7 
        y8 = y7 - vcut                         # + → -
        x9 = x8 
        y9 = y8 - midplate_height + vcut * 2  # + → -
        x10 = x9
        y10 = y9 - vcut                        # + → -
        x11 = x10 - rightwing1 - rightwing2 + vcut * 1.5
        y11 = y10 
        x12 = x11 
        y12 = y11 - topwing2 + vcut * 2 + 2    # + → -
        x13 = x12 - rightwing1 + 1 - 6
        y13 = y12 
        x14 = x13 
        y14 = y13 - vcut * 2 - vcut / 3        # + → -
        x15 = x14 
        y15 = y14 - topwing1 + vcut            # + → -
        x16 = x15 - midplate_width + leftwing1 + rightwing1 + 1 + 6 * 2
        y16 = y15 
        x17 = x16 
        y17 = y16 + topwing1 - vcut            # - → +
        x18 = x17
        y18 = y17 + vcut * 2 + vcut / 3        # - → +
        x19 = x18 - leftwing1 + 1 - 6
        y19 = y18 
        x20 = x19 
        y20 = y19 + topwing2 - vcut * 2 - 2    # - → +
        x21 = x20 - rightwing1 - rightwing2 + vcut * 1.5 
        y21 = y20 
        x22 = x21 
        y22 = y21 + vcut                       # - → +
        x23 = x22
        y23 = y22 + midplate_height - vcut * 2 # - → +
        x24 = x23 
        y24 = y23 + vcut                       # - → +
        x25 = x24 + leftwing1 + leftwing2 - vcut * 1.5 
        y25 = y24
        x26 = x25 
        y26 = y25 + leftwing2 - 2.7            # - → +
        x27 = x26 + leftwing1 - 1
        y27 = y26 
        x28 = x27 
        y28 = y27 + vcut * 1.5                 # - → +


        # 세로 절곡점 찾기
        x29 = x24 + leftwing1 - vcut
        y29 = y24
        x30 = x25 - vcut*1.5
        y30 = y25 
        x31 = x6 + vcut*1.5
        y31 = y6
        x32 = x7 - rightwing1 + vcut
        y32 = y7

        x33 = x10 - rightwing1 + vcut
        y33 = y10
        x34 = x11 + vcut*1.5
        y34 = y11
        x35 = x20 - vcut*1.5
        y35 = y20
        x36 = x21 + leftwing1 - vcut
        y36 = y21

        # 전개도 사이즈 저장 #1 중판
        midplatesizex2 = x10 - x21 
        midplatesizey2 = y15 - y2

        # 중판1보다 4점 추가됨 주의
        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        # 2R 그리기
        Radiusval = 2          
        lastNum = 28
        
        def draw_arc(doc, start, end, radius, option=None):
            """ option='reverse'는 시계방향으로 점을 회전할때 이 것을 사용한다. """
            if option == 'reverse' :                
                line(doc, prev_x, prev_y, start[0], start[1], layer='레이져')                
                add_arc_between_points(doc, end, start, radius)
            else:                
                line(doc, prev_x, prev_y, start[0], start[1], layer='레이져')                
                add_arc_between_points(doc, start, end, radius)
                
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            if i==0:
                line(doc, prev_x, prev_y , curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y            
            # if i==5 or i==12 or i==19 or i==26 :
            if i==5 :
                draw_arc(doc, (curr_x - Radiusval, curr_y), (curr_x, curr_y - Radiusval), Radiusval, option='reverse')
                prev_x, prev_y = curr_x , curr_y - Radiusval
            elif i == 12:
                draw_arc(doc, (curr_x, curr_y + Radiusval), (curr_x - Radiusval, curr_y), Radiusval, option='reverse')
                prev_x, prev_y = curr_x - Radiusval, curr_y  
            elif i==19 :
                draw_arc(doc, (curr_x + Radiusval, curr_y), (curr_x, curr_y + Radiusval), Radiusval, option='reverse')
                prev_x, prev_y = curr_x, curr_y + Radiusval
            elif i==26 :
                draw_arc(doc, (curr_x, curr_y - Radiusval), (curr_x + Radiusval, curr_y), Radiusval, option='reverse')
                prev_x, prev_y = curr_x + Radiusval, curr_y  
            else:            
                line(doc, prev_x, prev_y , curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y

        line(doc, prev_x, prev_y, x1, y1, layer="레이져")          
        
        #절곡라인  밴딩 가로 방향      
        line(doc, x28, y28, x3, y3, layer='hidden')
        line(doc, x23, y23, x8, y8, layer='hidden')
        line(doc, x22, y22, x9, y9, layer='hidden')
        # 절곡 실선
        line(doc, x17, y17, x14, y14,  layer='2')   

        # 세로절곡선 표현
        line(doc, x36, y36, x29, y29, layer='hidden')   
        line(doc, x35, y35, x30, y30, layer='hidden')   
        line(doc, x34, y34, x31, y31, layer='hidden')   
        line(doc, x33, y33, x32, y32, layer='hidden')   

        # 상부 7개 200피치 단공 9파이 
        #178+1을 빼준다. 1600일때  1422 실제 중판W 계산원리 양쪽 0.5씩 빼주는 형태임   
        midplate_connetholes = calculate_midplate_connecthole_lc035(CW - 34.8) # 1600 CW기준 34.8 전개도임     
        for i, coord in enumerate(midplate_connetholes):  
            xval = x22 + coord 
            y = y22 - topwing2/2 + vcut            
            if i>0 and i<len(midplate_connetholes)-1 :
                circle_cross(doc, xval, y, 9, layer='레이져')
                dc(doc, xval, y)
            elif i==0:
                circle_cross(doc, xval , y , 9, layer='레이져')
                d(doc, x21 , y21 , xval, y, 100, direction="up")

            if i==3:
                baseX, baseY = xval , y

        dc(doc, x10, y10)

        dim_leader_line(doc, baseX, baseY, baseX + 150 ,baseY -50, f"{len(midplate_connetholes)} -%%C9 Hole")       
        dim(doc, baseX, y16, baseX, baseY, 100, direction="left")
    
        # 상부 중판과 연결홀 200피치 9개 홀 
        #178+1을 빼준다. 1600일때  1422 실제 중판W 계산원리 양쪽 0.5씩 빼주는 형태임    
        for i, coord in enumerate(lc035_midplateholes):  
            xval = x1 + coord - 0.5
            y = y1 - 10
            print(f" 중판 lc035_midplateholes 홀의 i변수 {i}  , x 좌표: {xval}")        
            circle_cross(doc, xval, y, 9, layer='레이져')
            if i>0 :
                dimcontinue(doc, xval, y)
            else:
                dim(doc, xval, y, x1, y1, 100, direction="up", option='reverse') 
            if i==0:
                baseX, baseY = xval, y

        dimcontinue(doc, x2, y2)
        dim_leader_line(doc, baseX, baseY, baseX + 50, baseY - 100, f"{len(lc035_midplateholes)} -%%C9 Hole(M6 Pop Nut)")       
        dim(doc, baseX, y1, baseX, y1 - 10, 50, direction="left")
    
        # 좌우 결합되는 단공, 배열된 것을 구현함 결합되는 홀임        
        # print(f"035 lc035_vertical_holes {lc035_vertical_holes}")       
        half_of_lc035_vertical_holes = lc035_vertical_holes[:len(lc035_vertical_holes)//2]

        for i, coord in enumerate(half_of_lc035_vertical_holes):  
            xx = x24 + 10
            XX_right = x10 -10   # 오른쪽 홀
            y = y21 - 39 + coord # 단공 첫홀이 54가 15로 나와야 하니 + 39해줘야 한다.
            # print(f"CD 방향 중판 홀의 i변수 {i}  , x 좌표: {xval}")        
            circle_cross(doc, xx , y , 9, layer='레이져')
            circle_cross(doc, XX_right , y , 9, layer='레이져')
            if i>0 :
                dc(doc,  xx , y ,distance = 58 )
            else:
                d(doc, x21  , y21 , xx, y,   100, direction="left") 

        dc(doc,  x24 , y24, option='reverse')

        # 상부 치수선
        dim(doc, x36, y36, x21, y21 , 160 , direction="down", option="reverse" )
        dimcontinue(doc, x35, y35)
        dim(doc, x35, y35, x34, y34 , 160 , direction="down")    
        dim(doc, x33, y33, x10, y10 , 160 , direction="down", option="reverse" )
        dimcontinue(doc, x34, y34)    
        # 상부 전체치수
        dim(doc, x21, y21, x10, y10 , 210 , direction="down")

        # 좌측 치수선
        d(doc, x16,  y16, x21, y21 , 150 , direction="left")
        dc(doc, x24, y24)
        dc(doc, x1, y1, option='reverse')
        d(doc, x16, y16 , x1, y1,  290 , direction="left")

        # 하부 치수선
        dim(doc, x1, y1, x26 , y26 , 90 , direction="up" )
        dim(doc, x2, y2, x5 , y5 , 90 , direction="up" )

        dim(doc, x1, y1, x24 , y24 , 130 , direction="up" , option='reverse')
        dimcontinue(doc,  x2, y2)    
        dimcontinue(doc,  x7, y7)      
        dim(doc, x24, y24, x7 , y7 , 230 , direction="up")

        # 우측 치수선
        dim(doc, x14, y14, x15, y15, 150, direction="right")
        dim(doc, x9, y9, x14, y14, 150, direction="right", option='reverse')
        dimcontinue(doc, x8, y8)
        dimcontinue(doc, x3, y3)
        dim(doc, x3, y3, x2, y2,  150 , direction="right")   
        dim(doc, x15, y15, x2, y2 , 310 , direction="right")

    # 좌측 단면도 그리기
    # 왼쪽 leftwing1 leftwing2
    # 오른쪽 날개 rightwing1 rightwing2
    # 본판 mainplate = mp 
    # 'ㄷ'자 형상 
    program = 1
    if program:
        mp = midplate_height

        startx = x1 - 503 - btwing2
        starty = y1 - 71.6

        x1 = startx
        y1 = starty
        x2 = x1 - thickness    
        y2 = y1    
        x3 = x2  
        y3 = y2 + btwing1        # - → +
        x4 = x3 + btwing2 
        y4 = y3 
        x5 = x4 
        y5 = y4 - midplate_height    # + → -
        x6 = x5 - topwing2 + thickness
        y6 = y5 
        x7 = x6 
        y7 = y6 - topwing1 + thickness    # + → -
        x8 = x7 - thickness
        y8 = y7 
        x9 = x8 
        y9 = y8 + topwing1        # - → +
        x10 = x9 + topwing2 - thickness 
        y10 = y9 
        x11 = x10 
        y11 = y10 + midplate_height - thickness *2    # - → +
        x12 = x11 - btwing2 + thickness*2
        y12 = y11

        # 분할선 녹색
        x = (x8+x5)/2
        y = y5
        line(doc, x,y+10, x, y-20, layer="1")
        dim(doc, x,y+10, x5, y5, 50, direction="down")

        insert_block(x1 , y1 + 20, "lc035_popnut_leftbottom")   

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 12
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="0")     

        #vcut 원 그리기
        draw_circle(doc, x3, y3, 13, layer='0', color='6') 
        draw_circle(doc, x4, y4, 13, layer='0', color='6') 
        draw_circle(doc, x5, y5, 13, layer='0', color='6') 
        draw_circle(doc, x6, y6, 13, layer='0', color='6') 

        # 치수선
        d(doc, x3,y3, x4, y4, 110, direction="up", option='reverse')
        d(doc, x2,y2, x3, y3, 80, direction="left")
        dim(doc, x4,y4, x5, y5, 110, direction="right")
        dim(doc, x8,y8, x5, y5, 100, direction="down")
        dim(doc, x8,y8, x9, y9, 80, direction="left")

        dim_leader_line(doc, x4, y4 , x4 + 50 , y4+150, "Vcut 4개소")
   
    ###############################################################################
    # 상부 단면도 그리기
    ###############################################################################
    program = 1
    if program:    
        mp = midplate_width

        startx = rx1 + leftwing1
        starty = ry1 + 340 + midplate_height + leftwing2*3

        x1 = startx - 20.5
        y1 = starty
        x2 = x1 - leftwing1
        y2 = y1    
        x3 = x2  
        y3 = y2 - leftwing2
        x4 = x3 + mp
        y4 = y3 
        x5 = x4 
        y5 = y4 + rightwing2
        x6 = x5 - rightwing1
        y6 = y5 
        x7 = x6 
        y7 = y6 - thickness
        x8 = x7 + rightwing1 - thickness
        y8 = y7 
        x9 = x8 
        y9 = y8 - rightwing2 + thickness*2
        x10 = x9 - mp + thickness *2
        y10 = y9 
        x11 = x10 
        y11 = y10 + leftwing2 - thickness *2
        x12 = x11 + leftwing1 - thickness
        y12 = y11

        insert_block(x1 , y1 , "lc035_popnut_topleft")       
        insert_block(x6 , y6 , "lc035_popnut_topright")       

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 12
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="0")     

        #vcut 원 그리기 4개소
        draw_circle(doc, x2, y2, 13, layer='0', color='6') 
        draw_circle(doc, x3, y3, 13, layer='0', color='6') 
        draw_circle(doc, x4, y4, 13, layer='0', color='6') 
        draw_circle(doc, x5, y5, 13, layer='0', color='6') 

        # 치수선
        dim(doc, x2,y2, x1, y1, 80, direction="up")
        dim(doc, x2,y2, x3, y3, 80, direction="left")
        dim(doc, x5,y5, x4, y4, 80, direction="right")
        dim(doc, x6,y6, x5, y5, 80, direction="up")
        dim(doc, x3,y3, x4, y4, 80, direction="down")
    
        dim_leader_line(doc, x4, y4 , x4 + 50 , y4 - 100, "Vcut 4개소")

        # 중판(뒷쪽)  description    
        x1 = frameXpos + 600 + (midplate_width-960)
        y1 = frameYpos + 1100 +  (midplate_height-722)
        textstr = f"Part Name : 중판(뒤쪽)"    
        draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : {LCplateMaterial}"    
        draw_Text(doc, x1 , y1 - 40 , 20, str(textstr), layer='0')    
        textstr = f"Size : {str(math.ceil(midplatesizex2*100)/100)} x {str(math.ceil(midplatesizey2*100)/100)}mm"    
        draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
        textstr = f"Quantity : {su} EA"    
        draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')        

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 6page 의장면 보강대(샤링가공)
    #################################################################################################        
    program = 1
    if program:         
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = 2310 + (CW - 1600)
        TargetYscale = 1408 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("6page 스케일 비율 : " + str(frame_scale))           
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)   

        rx1 = frameXpos  + 260
        ry1 = frameYpos + 880 * frame_scale
        
        thickness = 1.5
        vcut = thickness / 2

        # 중판 전개도 크기 정의
        make_width = midplate_width - 50
        midplate_su = 2

        topwing = 32
        midbridge = 40
        bottomwing = 32
        
        # 중판 width - 40 적용 날개 간섭등 공차 차감

        x1 = rx1 
        y1 = ry1
        x2 = x1 + make_width
        y2 = y1    
        x3 = x2
        y3 = y2 + bottomwing - thickness
        x4 = x3
        y4 = y3 + midbridge - thickness * 2
        x5 = x4 
        y5 = y4 + topwing - thickness
        x6 = x5 - make_width
        y6 = y5 
        x7 = x6 
        y7 = y6 - topwing + thickness
        x8 = x7 
        y8 = y7 - midbridge + thickness * 2

        # 전개도 사이즈 저장 #1 중판
        midplatesizex = x2 - x1
        midplatesizey = y5 - y1

        # 중판1보다 4점 추가됨 주의
        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 8
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
        
        #절곡라인  밴딩 가로 방향      
        line(doc, x8, y8, x3, y3,  layer='hidden')   # 절곡선      
        line(doc, x7, y4, x4, y4,  layer='hidden')   # 절곡선      
    
        # 상부 치수선
        dim(doc, x6, y6, x5, y5 , 100 , direction="up")

        # 우측 치수선
        d(doc, x5, y5, x4, y4 , 80 , direction="right", option='reverse')
        d(doc, x4, y4,  x3, y3, 130 , direction="right", option='reverse')
        dc(doc, x2, y2 )
        d(doc, x5, y5, x2, y2 , 220 , direction="right")
        
        # 우측 단면도 그리기    
        # 'ㄷ'자 형상 

        startx = x1 + midplate_width + 350
        starty = y1 + (topwing + bottomwing)

        x1 = startx
        y1 = starty
        x2 = x1
        y2 = y1 - thickness    
        x3 = x2 - bottomwing 
        y3 = y2 
        x4 = x3 
        y4 = y3 + midbridge
        x5 = x4 + topwing
        y5 = y4 
        x6 = x5 
        y6 = y5 - thickness
        x7 = x6 - topwing + thickness
        y7 = y6 
        x8 = x7 
        y8 = y7 - midbridge + thickness*2

        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 8
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="0")     

        # 단면도 치수선
        dim(doc, x3,y3, x2, y2, 100,  text_height=0.12, direction="down")
        dim(doc, x4,y4, x3, y3, 50,  text_height=0.12, direction="left")    
        dim(doc, x4,y4, x5, y5, 60, text_height=0.12,  direction="up")    

        # 중판 보강 description    
        x1 = frameXpos + (900 + (CW-1600)) * frame_scale
        y1 = frameYpos + 700 * frame_scale
        textstr = f"Part Name : 의장면 보강대(샤링가공)"    
        draw_Text(doc, x1 , y1 , 20 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : EGI 1.6T"
        draw_Text(doc, x1 , y1 - 40 , 20, str(textstr), layer='0')        
        textstr = f"Size : {math.floor(midplatesizex)} x {math.floor(midplatesizey*10)/10}mm"    
        draw_Text(doc, x1 , y1 - 80 , 20, str(textstr), layer='0')        
        textstr = f"Quantity : {su*midplate_su*2} EA"    
        draw_Text(doc, x1 , y1 - 120 , 20, str(textstr), layer='0')       

        frameXpos = frameXpos + TargetXscale + 400

    #############################################################################################
    # 7page PROFILE LED BAR (좌,우)
    #############################################################################################
    program = 1
    if program:    
        abs_x = frameXpos 

        # 등기구 크기 산출
        light_width = LCD - 2
        
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = 2310 + max(CD - 1500, CW-1600)
        TargetYscale = 1408 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale   

        frame_scale = math.ceil(frame_scale*10) / 10       
        # print("7page 스케일 비율 : " + str(frame_scale))  
        insert_frame(  abs_x , frameYpos , frame_scale, "drawings_frame", workplace)          

        LedSpace_CD = (light_width - LedBarLength_CD)/2
        
        # 등기구 확산커버포함 3개의 rectangle    
        rx1 = frameXpos + 360
        ry1 = frameYpos + 1450 *frame_scale 
        insideHeight = 45   
        x1 = rx1 
        y1 = ry1 
        x2 = rx1 + light_width
        y2 = y1 + insideHeight    
        rectangle(doc, x1, y1, x2, y2, layer='0')          
        insideHeight = 12   
        x3 = x1 + LedSpace_CD
        y3 = y1 + 16.5 
        x4 = x2 - LedSpace_CD
        y4 = y3 + insideHeight    
        rectangle(doc, x3, y3, x4, y4, layer='6')      
        insideHeight = 6   
        x5 = x3 + 2
        y5 = y3 + 3 
        x6 = x4 - 2
        y6 = y5 + insideHeight    
        rectangle(doc, x5, y5, x6, y6, layer='4')    

        # 양쪽에 홀그리기
        draw_circle(doc, x1 + 20, y1+22.5 , 20 , layer='0')
        draw_circle(doc, x2 - 20, y1+22.5 , 20 , layer='0')    

        #  마구리 라인
        line(doc, x1+2, y1, x1+2, y2, layer="0")
        line(doc, x2-2, y1, x2-2, y2, layer="0")

        # 클립홀 위치에 따른 치수선 및 적색선 그리기
        for i, coord in enumerate(lc035_clipholes):  
            xval = x1 + coord -1 
            y = y1 - 3
            # print(f" 중판 clip 홀의 i변수 {i}  , x 좌표: {xval}")        
            if i<len(lc035_clipholes):
                line(doc, xval, y, xval, y2+3, layer="CL")
            if i==1 or i==2:
                textstr =  f"{Leditem} 고정 센터 -  {coord-preval}L"
                dc(doc,  xval , y2+3 ,text=textstr)
            elif i==3:
                dc(doc,  xval , y2+3)
            else:
                d(doc, x1 , y2,  xval  , y2+3 , 80, direction="up" ) 
            preval = coord 

        dc(doc,  x2 , y2+3 )

        # 하부 치수선
        dim(doc,  x1 + LedSpace_CD, y1, x1, y1,  70 , text_height=0.22, text_gap=0.07, direction="down")
        textstr =  f"{Leditem} - {Ledwatt} (누드&클립&잭타입) =  {LedBarLength_CD}L"
        dim(doc, x1 + LedSpace_CD, y1+16.5,  x2-LedSpace_CD, y1+16.5, 79+7.5,   direction="down" , text=textstr)    
        dim(doc,  x2-LedSpace_CD, y1, x2, y1, 70 , text_height=0.22, text_gap=0.07, direction="down")

        textstr =  f"{light_width} - (AL TUBE 길이)"
        dim(doc, x1, y1, x2, y2, 200, direction="up"  , text=textstr)    

        # AL profile 단면도 삽입
        insert_block( x2 + 120  , y1 -9  , "lc035_section_profile")         

        # AL profile  description    
        x = (x1 + x2)/2 - 150
        y = y1 - 150
        textstr = f"Part Name :  PROFILE LED BAR(좌,우)  "    
        draw_Text(doc, x , y , 20 , str(textstr), layer='0')        
        textstr = f"Mat.Spec : AL "    
        draw_Text(doc, x , y - 40 , 20, str(textstr), layer='0')    
        textstr = f"Size : SHP 4045 -  {light_width}L"    
        draw_Text(doc, x , y - 80 , 20, str(textstr), layer='0')        
        textstr = f"Quantity : {su*2} EA"    
        draw_Text(doc, x , y - 120 , 20, str(textstr), layer='0')   

    #############################################################################################
    #  PROFILE LED BAR (앞,뒤)
    #############################################################################################
    program = 1
    if program:    
        abs_x = frameXpos 

        # 등기구 크기 산출
        light_width = CW - 136

        LedSpace_CW = (light_width - LedBarLength_CW)/2
        
        # 등기구 확산커버포함 3개의 rectangle    
        rx1 = frameXpos + 360
        ry1 = frameYpos + 650 *frame_scale 
        insideHeight = 45   
        x1 = rx1 
        y1 = ry1 
        x2 = rx1 + light_width
        y2 = y1 + insideHeight    
        rectangle(doc, x1, y1, x2, y2, layer='0')          
        insideHeight = 12   
        x3 = x1 + LedSpace_CW
        y3 = y1 + 16.5 
        x4 = x2 - LedSpace_CW
        y4 = y3 + insideHeight    
        rectangle(doc, x3, y3, x4, y4, layer='6')      
        insideHeight = 6   
        x5 = x3 + 2
        y5 = y3 + 3 
        x6 = x4 - 2
        y6 = y5 + insideHeight    
        rectangle(doc, x5, y5, x6, y6, layer='4')      

        # 양쪽에 홀그리기
        draw_circle(doc, x1 + 20, y1+22.5 , 20 , layer='0')
        draw_circle(doc, x2 - 20, y1+22.5 , 20 , layer='0')

        # 클립홀 위치에 따른 치수선 및 적색선 그리기
        for i, coord in enumerate(lc035_front_clipholes):  
            xval = x1 + coord -1 
            y = y1 - 3
            # print(f" 중판 clip 홀의 i변수 {i}  , x 좌표: {xval}")        
            if i<len(lc035_clipholes):
                line(doc, xval, y, xval, y2+3, layer="CL")
            if i==1 or i==2:
                textstr =  f"{Leditem} 고정 센터 -  {coord-preval}L"
                dc(doc,  xval , y2+3 ,text=textstr)
            elif i==3:
                dc(doc,  xval , y2+3)
            else:
                d(doc, x1 , y2,  xval  , y2+3 , 80, direction="up" ) 
            preval = coord 

        dc(doc,  x2 , y2+3 )

        #  마구리 라인
        line(doc, x1+2, y1, x1+2, y2, layer="0")
        line(doc, x2-2, y1, x2-2, y2, layer="0")

        # 하부 치수선
        dim(doc,  x1 + LedSpace_CW, y1, x1, y1,  70 , text_height=0.22, text_gap=0.07, direction="down")
        textstr =  f"{Leditem} - {Ledwatt} (누드&클립&잭타입) =  {LedBarLength_CW}L"
        dim(doc, x1 + LedSpace_CW, y1+16.5,  x2-LedSpace_CW, y1+16.5, 79+7.5,   direction="down" , text=textstr)    
        dim(doc,  x2-LedSpace_CW, y1, x2, y1, 70 , text_height=0.22, text_gap=0.07, direction="down")

        # 상부 치수선
        textstr =  f"{light_width} - (AL TUBE 길이)"
        dim(doc, x1, y1, x2, y2, 200, direction="up"  , text=textstr)    

        # AL profile 단면도 삽입
        insert_block( x2 + 120  , y1 -9  , "lc035_section_profile")         

        # AL profile  description    
        x = (x1 + x2)/2 - 150
        y = y1 - 150
        textstr = f"Part Name :  PROFILE LED BAR(앞,뒤)  "    
        draw_Text(doc, x , y , 20 , str(textstr), layer='0')        
        textstr = f"Mat.Spec : AL "    
        draw_Text(doc, x , y - 40 , 20, str(textstr), layer='0')    
        textstr = f"Size : SHP 4045 -  {light_width}L"    
        draw_Text(doc, x , y - 80 , 20, str(textstr), layer='0')        
        textstr = f"Quantity : {su*2} EA"    
        draw_Text(doc, x , y - 120 , 20, str(textstr), layer='0')   

        # car inside box 표기
        xx =  x2 + 100
        yy =  y1 + 280
        rectangle(doc, xx,yy,xx+200*2,yy+100*2,layer='0')      
        line(doc, xx,yy+50*2,xx+200*2,yy+50*2)
        textstr = f"CAR INSIDE"    
        draw_Text(doc, xx+50*2, yy+50*2 , 24, str(textstr), layer='0')
        textstr = f"W{CW} x D{CD}"    
        draw_Text(doc, xx+30*2, yy+25*2 , 24, str(textstr), layer='0')  

        frameXpos = frameXpos + TargetXscale + 400

    program = 1
    if program:
        ####################################################################################################################################################
        # 레이져 도면틀 생성 (공통) 레이져에 배열할때 편리한 기능
        # 1219*1950 철판크기 샘플
        x1 = 10000+850
        y1 = 5000+130
        x2 = 10000+850+2438
        y2 = y1 +1219    
        x3 = 10000+850+2438+850+2438+ 1850  
        y3 = y1 +1219
        rectangle(doc, 10000, 5000,10000+13400 , 5000+ 2340 , layer='0')  
        rectangle(doc,x1 , y1 ,x1 + 2438 , y2 , layer='레이져')  
        dim(doc, x1 , y2 , x2 , y2,  100, direction="up")
        # rectangle(doc, x1+2600 , y1 , x2+2600 , y2 , layer='레이져')  
        # dim(doc, x1+2600, y2, x2+2600, y2, 100, direction="up")
        # dim(doc,  x1 , y1, x1 ,y2 , 150, direction="left")
        rectangle(doc, x3 , y1 , x3+2600 , y2 , layer='레이져')  
        dim(doc, x3, y3, x3+2600, y2, 100, direction="up")
        dim(doc,  x3 , y1, x3 ,y2 , 150, direction="left")

        textstr = f"{secondord} - {workplace} - {drawdate_str_short} 출도, {deadlinedate_str_short} 납기"    
        draw_Text(doc, 10000+300 , 5000+ 2340 - 200 ,  120, str(textstr), layer='0')        

        y = 5000+ 2340 - 650
        textstr = f"SPCC 1.2tX1219X = 장"
        draw_Text(doc, 10000+1000 ,y ,  90, str(textstr), layer='0')        
        textstr = f"SPCC 1.6T"        
        draw_Text(doc, 10000+1000+2600 , y ,  90, str(textstr), layer='0')            
        textstr = f"EGI 1.2T"        
        draw_Text(doc, 10000+1000+5200 ,y ,  90, str(textstr), layer='0')        
        textstr = f"{LCplateMaterial}"
        draw_Text(doc, 10000+1000+5200+2600 , y ,  90, str(textstr), layer='0')            

####################################################################################################################################################################################
# 본천장 자동작도
####################################################################################################################################################################################
def MainCeiling():   
    global args  #수정할때는 global 선언이 필요하다. 단순히 읽기만 할때는 필요없다.    
    global start_time
    global workplace, drawdate, deadlinedate, lctype, secondord, text_style, exit_program, Vcut_rate, Vcut, program_message, formatted_date, text_style_name
    global CW_raw, CD_raw, panel_thickness, su, LCframeMaterial, LCplateMaterial
    global input_CD, input_CW,CD,CW,drawdate_str, deadlinedate_str, drawdate_str_short, deadlinedate_str_short
    global doc, msp, T5_is, text_style, distanceXpos, distanceYpos, OP_raw
    global SplitYpos1, SplitYpos2, SplitYpos3, MDFirstSpace, WGap, MDCenterBK, MDSecondSpace

    # # 2page 덴크리 다온텍등 등기구 설명 그림
    # insert_block(0 , 0 , "lc008_A_2pageframe")        
    abs_x = 0
    abs_y = -5000
    # # LC 크기 지정 car inside에서 차감
    LCD = CD-50
    LCW = CW-50  # 라이트 케이스 크기를 정한다. 보통은 50

    # # watt 계산 공식 적용해야 함
    # # LED바 기장(단위:m) X 수량 X 15 = 60W 이하
    # # LED바 1개당 watt 산출 m단위로 계산.. /1000  031모델은 3개의 led바가 들어감     

    ledsu = 4    

    if(CD>2000):
        ledsu = 4

    wattCalculate = math.ceil((LCD - 3)/1000 * ledsu * 15)

    print("wattCalculate: " + str(wattCalculate))  

    if(wattCalculate <= 60):
        watt = 60
    elif(wattCalculate > 60 and wattCalculate < 100):
        watt = 100
    elif(wattCalculate >= 100 and wattCalculate < 150):
        watt = 150
    elif(wattCalculate >= 150 and wattCalculate < 200):
        watt = 200
    elif(wattCalculate >= 200 and wattCalculate < 200):
        watt = 300
                
    # print("L/C 1개당 규격설정 wattCalculate: " + str(watt))  

    T5_standard = 0
    if T5_is == '있음':
        adjusted_CD = CD - 50  # CD에서 50 뺌

        # 규격 결정
        if adjusted_CD > 1500:
            T5_standard = 1500
        elif adjusted_CD > 1200:
            T5_standard = 1200
        else:
            T5_standard = 900

        print(f"선택된 규격: {T5_standard}")
    else:
        print("T5_is 가 '있음'이 아니므로 규격 선택 안 함")    

    # 도면틀 넣기
    BasicXscale = 2560
    BasicYscale = 1550
    TargetXscale = 800 
    TargetYscale = 300 
    if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
        frame_scale = TargetXscale/BasicXscale
    else:
        frame_scale = TargetYscale/BasicYscale

    frame_scale = 1    
    ########################################################
    # 갑지 1
    ########################################################
    # 현장명
    textstr = workplace       
    x = abs_x + 80
    y = abs_y + 1000 + 218
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 발주처
    textstr = secondord
    x = abs_x + 80
    y = abs_y + 1000 + 218 - 30
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 타입
    textstr = lctype
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 수량
    textstr = su
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 카사이즈
    textstr = f"W{CW} x D{CD}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')
    # 색상 LC
    LCcolor = LCframeMaterial
    if LCframeMaterial=='EGI 1.6t' :
        LCcolor = '흑색무광'
    textstr = f"L / C : {LCcolor}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18
    draw_Text(doc, x, y , 7, str(textstr), layer='0')   

    textstr = f"중판 : {LCplateMaterial}"
    x = abs_x + 260
    y = abs_y + 1000 + 214 + 10 - 18 - 18 - 18 - 12
    draw_Text(doc, x, y , 7, str(textstr), layer='0')

    # 등기구 업체
    LedBarLength = 0
    if T5_standard > 0 : 
        Ledcompany = f"(등기구업체 : 엘엠팩트)"        
        Leditem = "T5"
        Ledwatt = "6500K"
        LedBarLength_CD = T5_standard
        LedBarLength_CW = T5_standard
        watt = "내장"
    else:
        Ledcompany = f"(등기구업체 : 다온텍)"
        Leditem = "LED BAR"
        Ledwatt = "10000K"
        LedBarLength_CD = CalculateLEDbar(CD-70)
        LedBarLength_CW = CalculateLEDbar(CW-170)
        # print (f"035 LedBarLength_CD : {LedBarLength_CD}")
        # print (f"035 LedBarLength_CW : {LedBarLength_CW}")
        watt = f"{str(watt)}W"

    x = abs_x + 80
    y = abs_y + 1000 + 120
    draw_Text(doc, x, y , 8, str(Ledcompany), layer='0')     
    textstr = f"(할로겐 전구색 3000K (스텐링)일체형-다온텍)"
    x = abs_x + 80
    y = abs_y + 1000 + 120 - 15
    draw_Text(doc, x, y , 8, str(textstr), layer='0')
    
    altube_CW = CW-250
    altube_CD = CD-300

    textstr = f"*. 출도일자 : {drawdate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 
    draw_Text(doc, x, y , 8, str(textstr), layer='0')    
    textstr = f"*. 납품일자 : {deadlinedate_str}"
    x = abs_x + 230
    y = abs_y + 1000 + 30 - 15
    draw_Text(doc, x, y , 8, str(textstr), layer='0')

    ################################################################################################################################################
    # 갑지 2
    ################################################################################################################################################
    # 현장명
    # textstr = f"* 현장명 : {secondord}-{workplace} - {su}(set)"
    # x = abs_x + 30
    # y = abs_y + 650 
    # draw_Text(doc, x, y , 20, str(textstr), layer='0')        
    # x = abs_x + 30 + 200
    # y = abs_y + 600 
    # draw_Text(doc, x, y , 20, str(Ledcompany), layer='0')    
    
    # textstr = f"{watt}"
    # x = abs_x + 835
    # y = 395 + 110
    # draw_Text_direction(doc, x + 50, y - 30 , 20, str(textstr), layer='0', rotation = 0)    

    # textstr = f"LED BAR - 3000K -{altube_CW}L" #상
    # x = abs_x + 410
    # y = 550
    # draw_Text(doc, x, y , 12, str(textstr), layer='0')
    
    # textstr = f"(누드형 클립형 잭타입)" #상
    # x = abs_x + 410 +10
    # y = 530
    # draw_Text(doc, x, y , 12, str(textstr), layer='0')

    # textstr = f"LED BAR - 3000K -{altube_CW}L" #하
    # x = abs_x + 410
    # y = 544-460
    # draw_Text(doc, x, y , 12, str(textstr), layer='0')
    
    # textstr = f"(누드형 클립형 잭타입)" #하
    # x = abs_x + 410 +10
    # y = 524-460
    # draw_Text(doc, x, y , 12, str(textstr), layer='0')

    # x = abs_x + 240
    # y = 230    
    # textstr = f"LED BAR - 3000K -{altube_CD}L" #좌
    # draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)        

    # x = abs_x + 260
    # y = 240    
    # textstr = f"(누드형 클립형 잭타입)" #좌
    # draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)        
       
    # x = abs_x + 720
    # y = 220    
    # textstr = f"LED BAR - 3000K -{altube_CD}L" #우
    # draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)        

    # x = abs_x + 740
    # y = 230    
    # textstr = f"(누드형 클립형 잭타입)" #우
    # draw_Text_direction(doc, x , y , 11, str(textstr), layer='0', rotation = 90)   
                   
    frameXpos = 1437

    #####################################################################################################################################
    # 1page Assy 본판 작도
    #####################################################################################################################################     
    program = 1    
    if program:                
        abs_x = frameXpos + 700    
        abs_y = abs_y + CD + 500
        # 기본 윤곽 car inside
        MCW = CW + 50
        MCD = CD + 50

        def common_cal(CD, CW, ceilingdivision, ELYpos, sidePanel):
            global SplitYpos1, SplitYpos2, SplitYpos3

            print(f"CD, CW, ceilingdivision, ELYpos, sidePanel : {CD}, {CW}, {ceilingdivision}, {ELYpos}, {sidePanel}")         
            if CD <= 1200:
                SplitYpos1 = CD
            elif 1200 < CD <= 1400 and ceilingdivision == 2:
                SplitYpos1 = math.floor((MCD - 500) / 100) * 100
                if SplitYpos1 > 1100:
                    SplitYpos1 = 1100
                SplitYpos2 = MCD - SplitYpos1
            elif (CD <= 2400 and ceilingdivision == 2) or (CD > 2400 and ceilingdivision == 2):
                SplitYpos1 = math.floor((MCD - 500) / 100) * 100
                if SplitYpos1 > 1100:
                    SplitYpos1 = 1100
                SplitYpos2 = MCD - SplitYpos1
                sumOB = 0
                changeSplitYpos1 = 0  

                if sidePanel is None:
                    sidePanel = []

                for currentVar in sidePanel:
                    sumOB += currentVar
                    if currentVar and sumOB != CD:
                        if SplitYpos1 - 20 <= sumOB <= SplitYpos1 + 20:
                            changeSplitYpos1 = SplitYpos1 - 20

                if changeSplitYpos1 > 0:
                    SplitYpos1 = changeSplitYpos1
                    SplitYpos2 = MCD - SplitYpos1
            elif CD > 2400 and ceilingdivision == 3:
                SplitYpos1 = 1000
                SplitYpos2 = math.floor((MCD - 1000 - 500) / 100) * 100
                SplitYpos3 = MCD - SplitYpos1 - SplitYpos2

            MDFirstSpace = math.floor(MCD - ELYpos - 65)

            if ceilingdivision == 1:
                MDFirstSpace = math.floor((SplitYpos1 - ELYpos - 130) / 4)
            elif ceilingdivision == 2:
                if 1800 <= CD <= 2400:
                    MDFirstSpace = math.floor((SplitYpos1 + SplitYpos2 - ELYpos - 130) / 5)
                else:
                    MDFirstSpace = math.floor(SplitYpos2 - ELYpos - 130)
            elif ceilingdivision == 3:
                MDFirstSpace = math.floor(SplitYpos3 - ELYpos - 130)
            
            MDSecondSpace = math.floor(MCD - SplitYpos2 - 65 - MDFirstSpace - 130 - 30)

            if MDSecondSpace > 400:
                MDCenterBK = (MCD - (SplitYpos2 + 65 + MDFirstSpace + 130 + 30))/2
            else:
                MDCenterBK = 0

            Tlength = CW - 62
            sampleGap = 60.5
            WGap = (Tlength - 60 - sampleGap * 2) / 3

            if WGap != math.floor(WGap):
                sampleGap = 61
                WGap = (Tlength - 60 - sampleGap * 2) / 3

            if WGap != math.floor(WGap):
                sampleGap = 60
                WGap = (Tlength - 60 - sampleGap * 2) / 3

            print(f"SplitYpos1:{SplitYpos1}")
            print(f"SplitYpos2:{SplitYpos2}")
            print(f"SplitYpos3:{SplitYpos3}")
            print(f"MDFirstSpace:{MDFirstSpace}")
            print(f"MDSecondSpace:{MDSecondSpace}")
            print(f"MDCenterBK:{MDCenterBK}")
                
            return SplitYpos1, SplitYpos2, SplitYpos3, MDFirstSpace, WGap, MDCenterBK, MDSecondSpace
        
        SplitYpos1, SplitYpos2, SplitYpos3, MDFirstSpace, WGap, MDCenterBK, MDSecondSpace = common_cal(CD, CW, ceilingdivision, ELYpos, sidePanel)
        
        Xgap = 30
        Ygap = 30
        abs_y = abs_y - CD - 800    
        x1 = abs_x 
        y1 = abs_y 
        x2 = abs_x + MCW
        y2 = abs_y + MCD
        rectangle(doc, x1, y1, x2, y1 + (MCD - SplitYpos1), layer='0')    
        rectangle(doc, x1, y1 + (MCD - SplitYpos1), x2, y2, layer='0')    
        line(doc, x1, y1 + Ygap, x2, y1 + Ygap, layer="0")
        line(doc, x1, y1 + Ygap - 2, x2, y1 + Ygap - 2, layer="0")
        line(doc, x1, y2 - Ygap, x2, y2 - Ygap, layer="0")
        line(doc, x1, y2 - Ygap + 2, x2, y2 - Ygap + 2, layer="0")
        line(doc, x1 + Xgap, y1 + Ygap, x1 + Xgap, y2 - Ygap, layer="0")
        line(doc, x1 + Xgap - 2, y1 + Ygap, x1 + Xgap - 2, y2 - Ygap, layer="0")
        line(doc, x2 - Xgap, y1 + Ygap, x2 - Xgap, y2 - Ygap, layer="0")
        line(doc, x2 - Xgap + 2, y1 + Ygap, x2 - Xgap + 2, y2 - Ygap, layer="0")

        #테두리 장공
        sum = 0
        for x in frontPanel[:-1] :
            sum = sum + x
            insert_block(x1 + sum, y1 + slotposition, "8x20 slot_")
            dim(doc, x1 + sum, y1, x1 + sum - x, y1, 70, direction="down", localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + sum, y1)
        dimcontinue(doc, x2, y1 )

        sum = 0
        for x in rearPanel[:-1] :
            sum = sum + x
            insert_block(x1 + sum, y2 - slotposition, "8x20 slot_")
            dim(doc, x1 + sum, y2, x1 + sum - x, y2, 70, direction="up", localdimstyle="0.2 JKW")    
            dimcontinue(doc , x1 + sum, y2)
        dimcontinue(doc ,x2, y2 )

        sum = 0
        for y in sidePanel[:-1] :
            sum = sum + y
            insert_block(x1 + slotposition, y1 + sum, "8x20 slot")
            
        sum = 0
        for y in sidePanel[:-1] :
            sum = sum + y
            insert_block(x2 - slotposition, y1 + sum, "8x20 slot")
            dim(doc, x2, y1 + sum, x2, y1 + sum - y, 150, direction="right", localdimstyle="0.2 JKW")    
            dimcontinue(doc, x2, y1 + sum)
        dimcontinue(doc, x2, y2 )

        inputCW = CW
        inputCD = CD
        
        if lctype == "035":
            #LC 조립홀 가로
            lc_positions = calculate_lc_position_035(inputCW, inputCD)

            Topholex = lc_positions['Topholex']
            Topholey = lc_positions['Topholey']

            Bottomx = lc_positions['Bottomx']
            Bottomy = lc_positions['Bottomy']    

            for i, x in enumerate(Topholex):
                circle_cross(doc, x1 + 25 + x, y2 - 65, 13, layer='0')
                if i == 0:
                    dim(doc, x1 + 25 + x, y2 - 60, x1, y2, 180, direction="up", option='reverse', localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + x + 25, y2 - 60)
            dimcontinue(doc, x2, y2) 
            
                #LC 조립홀 가로(header)
            if headertype == "이노에스템":
                for i, x in enumerate(Bottomx):
                    insert_block(x1 + 25 + x, y1 + 65, "Ceiling_LChole_ino")
                    if i == 0:
                        dim(doc, x1 + 25 + x, y1 + 50, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
                    else:
                        dimcontinue(doc, x1 + x + 25, y1 + 50)
                dimcontinue(doc, x2, y1) 
            elif headertype == "에이스코리아" or headertype == "세화정공":
                for i, x in enumerate(Bottomx):
                    insert_block(x1 + 25 + x, y1 + 65, "Ceiling_LChole_sehwa")
                    if i == 0:
                        dim(doc, x1 + 25 + x, y1 + 50, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
                    else:
                        dimcontinue(doc, x1 + x + 25, y1 + 50)
                dimcontinue(doc, x2, y1) 
            else:
                for i, x in enumerate(Bottomx):
                    circle_cross(doc, x1 + 25 + x, y1 + 65, 13, layer='0')
                    if i == 0:
                        dim(doc, x1 + 25 + x, y1 + 50, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
                    else:
                        dimcontinue(doc, x1 + x + 25, y1 + 50)
                dimcontinue(doc, x2, y1) 
            #LC 조립홀 세로
            hole_positions = calculate_lc_vertical_hole_positions_035(inputCW, inputCD)

            leftholex = hole_positions['leftholex']
            rightholex = hole_positions['rightholex']
            vertical_holey = hole_positions['vertical_holey']

            num_holes = min(len(leftholex), len(rightholex), len(vertical_holey))# 리스트 길이 확인
            # circle_cross 함수 호출 시 리스트 길이를 확인하여 오류 방지
            for i in range(num_holes):
                circle_cross(doc, x1 + 25 + leftholex[i], y1 + 25 + vertical_holey[i], 13, layer='0')        
                circle_cross(doc, x1 + 25 + rightholex[i], y1 + 25 + vertical_holey[i], 13, layer='0')
                if i == 0:
                    dim(doc, x2 - 135, y1 + 65, x1 + 25 + rightholex[i], y1 + 25 + vertical_holey[i], 335, direction="right", localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + 25 + rightholex[i] , y1 + 25 + vertical_holey[i])      
            dimcontinue(doc, x2 - 125, y2 - 65)
            dim(doc, x2, y1, x2 - 125, y1 + 65, 200, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x2, y2, x2 - 125, y2 - 65, 200, direction="right", localdimstyle="0.2 JKW")
            
        elif lctype == "008_A":
            lc_positions = calculate_lc_position_008_A(inputCW, inputCD)

            # Topholex와 Topholey 좌표 추출
            Topholex = lc_positions['Topholex']
            Topholey = lc_positions['Topholey']

            # middle 좌표 추출
            Middlex = lc_positions['Middlex']
            Middley = lc_positions['Middley']    

            # Bottomx와 Bottomy 좌표 추출
            Bottomx = lc_positions['Bottomx']
            Bottomy = lc_positions['Bottomy']    

            # 필요한 경우 다른 홀들에 대해서도 circle_cross 함수를 호출할 수 있음
            # 예를 들어, 모든 상단 홀에 대해 반복
            for i, x in enumerate(Topholex):
                circle_cross(doc, abs_x + 25 + x, abs_y + 25 + Topholey, 11, layer='0')
                if i == 0:
                    dim(doc, x1 + 25 + x, y2 - 90, x1, y2, 210, direction="up", option='reverse', localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + x + 25, y2 - 90)
            dimcontinue(doc, x2, y2) 

            for x in Middlex:
                circle_cross(doc, abs_x + 25 + x, abs_y + 25 + Middley, 11, layer='0')            

            for x in Bottomx:
                circle_cross(doc, abs_x + 25 + x, abs_y + 25 + Bottomy, 11, layer='0')     
                       
            dim(doc, x2 - 90, y1 + 100, x2, y1, 290, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x2 - 90, y1 + MCD/2) 
            dimcontinue(doc, x2 - 90, y2 - 100) 
            dimcontinue(doc, x2, y2) 

        elif lctype == "031":
            lc_positions = calculate_lc_position_031(inputCW, inputCD)

            Topholex = lc_positions['Topholex']
            Topholey = lc_positions['Topholey']

            Bottomx = lc_positions['Bottomx']
            Bottomy = lc_positions['Bottomy']    
            
            for i, x in enumerate(Topholex):
                circle_cross(doc, abs_x + 25 + x, abs_y + 25 + Topholey, 13, layer='0')
                if i == 0:
                    dim(doc, x1 + 25 + x, y2 - 60, x1, y2, 180, direction="up", option='reverse', localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + x + 25, y2 - 60)
            dimcontinue(doc, x2, y2) 

            if headertype == "이노에스템":
                for i, x in enumerate(Bottomx):
                    insert_block(x1 + 25 + x, y1 + 65, "Ceiling_LChole_ino")
                    if i == 0:
                        dim(doc, x1 + 25 + x, y1 + 50, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
                    else:
                        dimcontinue(doc, x1 + x + 25, y1 + 50)
                dimcontinue(doc, x2, y1) 
            elif headertype == "에이스코리아" or headertype == "세화정공":
                for i, x in enumerate(Bottomx):
                    insert_block(x1 + 25 + x, y1 + 65, "Ceiling_LChole_sehwa")
                    if i == 0:
                        dim(doc, x1 + 25 + x, y1 + 50, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
                    else:
                        dimcontinue(doc, x1 + x + 25, y1 + 50)
                dimcontinue(doc, x2, y1) 
            else:
                for i, x in enumerate(Bottomx):
                    circle_cross(doc, x1 + 25 + x, y1 + 65, 13, layer='0')
                    if i == 0:
                        dim(doc, x1 + 25 + x, y1 + 50, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
                    else:
                        dimcontinue(doc, x1 + x + 25, y1 + 50)
                dimcontinue(doc, x2, y1) 
                
            # 세로 홀 위치 좌표 계산
            hole_positions = calculate_lc_vertical_hole_positions_031(inputCW, inputCD)

            # 각 홀의 좌표를 추출
            leftholex = hole_positions['leftholex']
            rightholex = hole_positions['rightholex']
            vertical_holey = hole_positions['vertical_holey']

            # 리스트 길이 확인
            num_holes = min(len(leftholex), len(rightholex), len(vertical_holey))

            # circle_cross 함수 호출 시 리스트 길이를 확인하여 오류 방지
            for i in range(num_holes):
                circle_cross(doc, abs_x + 25 + leftholex[i], abs_y + 25 + vertical_holey[i], 13, layer='0')
                circle_cross(doc, abs_x + 25 + rightholex[i], abs_y + 25 + vertical_holey[i], 13, layer='0')
                if i == 0:
                    dim(doc, x2 - 135, y1 + 65, x1 + 25 + rightholex[i], y1 + 25 + vertical_holey[i], 335, direction="right", localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + 25 + rightholex[i] , y1 + 25 + vertical_holey[i], 340)      
            dimcontinue(doc, x2 - 125, y2 - 65, 340)
            dim(doc, x2, y1, x2 - 125, y1 + 65, 200, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x2, y2, x2 - 125, y2 - 65, 200, direction="right", localdimstyle="0.2 JKW")
            
        elif lctype == "N20": 
            if headertype == "이노에스템":
                insert_block(x1 + 137.5, y1 + 70, "Ceiling_LChole_N20")
                insert_block(x1 + 137.5 + 185, y1 + 70, "Ceiling_LChole_N20")
                insert_block(x2 - 137.5, y1 + 70, "Ceiling_LChole_N20")
                insert_block(x2 - 137.5 - 185, y1 + 70, "Ceiling_LChole_N20")
            elif headertype == "세화정공"or headertype == "에이스코리아":
                insert_block(x1 + 137.5, y1 + 70, "Ceiling_LChole_N20_acekorea")
                insert_block(x1 + 137.5 + 185, y1 + 70, "Ceiling_LChole_N20_acekorea")
                insert_block(x2 - 137.5, y1 + 70, "Ceiling_LChole_N20_acekorea")
                insert_block(x2 - 137.5 - 185, y1 + 70, "Ceiling_LChole_N20_acekorea")
            else:  
                circle_cross(doc, x1 + 137.5, y1 + 70, 13, layer='0')
                circle_cross(doc, x1 + 137.5 + 185, y1 + 70, 13, layer='0')
                circle_cross(doc, x2 - 137.5, y1 + 70, 13, layer='0')
                circle_cross(doc, x2 - 137.5 - 185, y1 + 70, 13, layer='0')
                
            circle_cross(doc, x1 + 137.5, y2 - 70, 13, layer='0')
            circle_cross(doc, x1 + 137.5 + 185, y2 - 70, 13, layer='0')
            circle_cross(doc, x2 - 137.5, y2 - 70, 13, layer='0')
            circle_cross(doc, x2 - 137.5 - 185, y2 - 70, 13, layer='0')
            
            dim(doc, x1 + 137.5, y2 - 60, x1, y2, 180, direction="up", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 137.5 + 185, y2 - 60)
            dimcontinue(doc, x2 - (137.5 + 185), y2 - 60)
            dimcontinue(doc, x2 - 137.5, y2 - 60)
            dimcontinue(doc, x2, y2)
            dim(doc, x1 + 137.5, y1 + 60, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 137.5 + 185, y1 + 60)
            dimcontinue(doc, x2 - (137.5 + 185), y1 + 60)
            dimcontinue(doc, x2 - 137.5, y1 + 60)
            dimcontinue(doc, x2, y1)
            dim(doc, x2 - 125, y1 + 70, x2, y1, 330, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x2 - 125, y2 - 70)
            dimcontinue(doc, x2, y2)

        elif lctype == "032":
            if headertype == "이노에스템":
                insert_block(x1 + 170, y1 + 65, "Ceiling_LChole_ino")
                insert_block(x1 + 170 + 150, y1 + 65, "Ceiling_LChole_ino")
                insert_block(x2 - 170, y1 + 65, "Ceiling_LChole_ino")
                insert_block(x2 - 170 - 150, y1 + 65, "Ceiling_LChole_ino")
            elif headertype == "에이스코리아" or headertype == "세화정공":
                insert_block(x1 + 170, y1 + 65, "Ceiling_LChole_sehwa")
                insert_block(x1 + 170 + 150, y1 + 65, "Ceiling_LChole_sehwa")
                insert_block(x2 - 170, y1 + 65, "Ceiling_LChole_sehwa")
                insert_block(x2 - 170 - 150, y1 + 65, "Ceiling_LChole_sehwa")
            else:  
                circle_cross(doc, x1 + 170, y1 + 65, 13, layer='0')
                circle_cross(doc, x1 + 170 + 150, y1 + 65, 13, layer='0')
                circle_cross(doc, x2 - 170, y1 + 65, 13, layer='0')
                circle_cross(doc, x2 - 170 - 150, y1 + 65, 13, layer='0')

            circle_cross(doc, x1 + 170, y2 - 65, 13, layer='0')
            circle_cross(doc, x1 + 170 + 150, y2 - 65, 13, layer='0')
            circle_cross(doc, x2 - 170, y2 - 65, 13, layer='0')
            circle_cross(doc, x2 - 170 - 150, y2 - 65, 13, layer='0')
            
            dim(doc, x1 + 170, y2 - 60, x1, y2, 180, direction="up", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 170 + 150, y2 - 60)
            dimcontinue(doc, x2 - (170 + 150), y2 - 60)
            dimcontinue(doc, x2 - 170, y2 - 60)
            dimcontinue(doc, x2, y2)
            dim(doc, x1 + 170, y1 + 60, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 170 + 150, y1 + 60)
            dimcontinue(doc, x2 - (170 + 150), y1 + 60)
            dimcontinue(doc, x2 - 170, y1 + 60)
            dimcontinue(doc, x2, y1)
            dim(doc, x2 - 160, y1 + 65, x2, y1, 360, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x2 - 160, y2 - 65)
            dimcontinue(doc, x2, y2)

        elif lctype == "026":
            if headertype == "이노에스템":
                insert_block(x1 + 165, y1 + 65, "Ceiling_LChole_ino")
                insert_block(x1 + 165 + 160, y1 + 65, "Ceiling_LChole_ino")
                insert_block(x2 - 165, y1 + 65, "Ceiling_LChole_ino")
                insert_block(x2 - 165 - 160, y1 + 65, "Ceiling_LChole_ino")
            elif headertype == "에이스코리아" or headertype == "세화정공":
                insert_block(x1 + 165, y1 + 65, "Ceiling_LChole_sehwa")
                insert_block(x1 + 165 + 160, y1 + 65, "Ceiling_LChole_sehwa")
                insert_block(x2 - 165, y1 + 65, "Ceiling_LChole_sehwa")
                insert_block(x2 - 165 - 160, y1 + 65, "Ceiling_LChole_sehwa")
            else:
                circle_cross(doc, x1 + 165, y1 + 65, 13, layer='0')
                circle_cross(doc, x1 + 165 + 160, y1 + 65, 13, layer='0')
                circle_cross(doc, x2 - 165, y1 + 65, 13, layer='0')
                circle_cross(doc, x2 - 165 - 160, y1 + 65, 13, layer='0')
                
            circle_cross(doc, x1 + 165, y2 - 65, 13, layer='0')
            circle_cross(doc, x1 + 165 + 160, y2 - 65, 13, layer='0')
            circle_cross(doc, x2 - 165, y2 - 65, 13, layer='0')
            circle_cross(doc, x2 - 165 - 160, y2 - 65, 13, layer='0')
            circle_cross(doc, x1 + 165, y1 + MCD/2, 13, layer='0')
            circle_cross(doc, x2 - 165, y1 + MCD/2, 13, layer='0')
            
            dim(doc, x1 + 165, y2 - 60, x1, y2, 180, direction="up", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 165 + 160, y2 - 60)
            dimcontinue(doc, x2 - (165 + 160), y2 - 60)
            dimcontinue(doc, x2 - 165, y2 - 60)
            dimcontinue(doc, x2, y2)
            dim(doc, x1 + 165, y1 + 60, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 165 + 160, y1 + 60)
            dimcontinue(doc, x2 - (165 + 160), y1 + 60)
            dimcontinue(doc, x2 - 165, y1 + 60)
            dimcontinue(doc, x2, y1)
            dim(doc, x2 - 160, y1 + 65, x2, y1, 360, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x2 - 160, y1 + MCD/2)
            dimcontinue(doc, x2 - 160, y2 - 65)
            dimcontinue(doc, x2, y2)

        #하부 Header
        if headertype == "이노에스템":
            rectangle(doc, x1 + 31, y1 + 32.5, x2 - 31, y1 + 32.5 + 55, layer='4')
            rectangle(doc, x1 + 76, y1 + 32.5, x2 - 76, y1 + 32.5 + 15, layer='4')
            rectangle(doc, x1 + 76, y1 + 32.5 + 40, x2 - 76, y1 + 32.5 + 15 + 40, layer='4')
            rectangle(doc, x1 + 31, y1 + 107.5, x2 - 31, y1 + 107.5 + 55, layer='4')
            rectangle(doc, x1 + 76, y1 + 107.5, x2 - 76, y1 + 107.5 + 15, layer='4')
            rectangle(doc, x1 + 76, y1 + 107.5 + 40, x2 - 76, y1 + 107.5 + 55, layer='4')
        elif headertype == "정일산업":
            rectangle(doc, x1 + 31, y1 + 30, x2 - 31, y1 + 30 + 75, layer='4')
            rectangle(doc, x1 + 76, y1 + 30, x2 - 76, y1 + 30 + 15, layer='4')
            rectangle(doc, x1 + 76, y1 + 30 + 60, x2 - 76, y1 + 30 + 15 + 60, layer='4')
            rectangle(doc, x1 + 31, y1 + 105, x2 - 31, y1 + 105 + 75, layer='4')
            rectangle(doc, x1 + 76, y1 + 105, x2 - 76, y1 + 105 + 15, layer='4')
            rectangle(doc, x1 + 76, y1 + 105 + 60, x2 - 76, y1 + 105 + 15 + 60, layer='4')
        elif headertype == "에이스코리아":
            rectangle(doc, x1 + 31, y1 + 30, x2 - 31, y1 + 30 + 55, layer='4')
            rectangle(doc, x1 + 76, y1 + 30, x2 - 76, y1 + 30 + 15, layer='4')
            rectangle(doc, x1 + 76, y1 + 30 + 40, x2 - 76, y1 + 30 + 15 + 40, layer='4')
            rectangle(doc, x1 + 31, y1 + 85, x2 - 31, y1 + 85 + 55, layer='4')
            rectangle(doc, x1 + 76, y1 + 85, x2 - 76, y1 + 85 + 15, layer='4')
            rectangle(doc, x1 + 76, y1 + 85 + 40, x2 - 76, y1 + 85 + 15 + 40, layer='4')
        elif headertype == "세화정공":
            rectangle(doc, x1 + 31, y1 + 35, x2 - 31, y1 + 35 + 100, layer='4')
            line(doc, x1 + 31, y1 + 35 + 13, x2 - 31, y1 + 35 + 13, layer="4")
            line(doc, x1 + 31, y1 + 35 + 50, x2 - 31, y1 + 35 + 50, layer="4")
            line(doc, x1 + 31, y1 + 35 + 87, x2 - 31, y1 + 35 + 87, layer="4")
            line(doc, x1 + 31, y1 + 35 + 2.3, x2 - 31, y1 + 35 + 2.3, layer="Hidden")
            line(doc, x1 + 31, y1 + 35 + 2.3 + 45.4, x2 - 31, y1 + 35 + 2.3 + 45.4, layer="Hidden")
            line(doc, x1 + 31, y1 + 35 + 2.3 + 50, x2 - 31, y1 + 35 + 2.3 + 50, layer="Hidden")
            line(doc, x1 + 31, y1 + 35 + 2.3 + 95.4, x2 - 31, y1 + 35 + 2.3 + 95.4, layer="Hidden")
                
            temp_x1 = x1 + 31
            temp_y1 = y1 + 72
            temp_x2 = temp_x1 + CW/2 - 181
            temp_y2 = temp_y1
            temp_x3 = temp_x2
            temp_y3 = temp_y2 + 10.7
            temp_x4 = temp_x3 + 50
            temp_y4 = temp_y3
            temp_x5 = temp_x4
            temp_y5 = temp_y4 - 10.7
            temp_x6 = temp_x5 + 200
            temp_y6 = temp_y5
            temp_x7 = temp_x6
            temp_y7 = temp_y6 + 10.7
            temp_x8 = temp_x7 + 50
            temp_y8 = temp_y7
            temp_x9 = temp_x8
            temp_y9 = temp_y8 - 10.7
            temp_x10 = temp_x9 + CW/2 - 131
            temp_y10 = temp_y9
            temp_x11 = temp_x10
            temp_y11 = temp_y10 + 26
            temp_x12 = temp_x11 - CW/2 + 131
            temp_y12 = temp_y11
            temp_x13 = temp_x12
            temp_y13 = temp_y12 - 10.7
            temp_x14 = temp_x13 - 50
            temp_y14 = temp_y13
            temp_x15 = temp_x14
            temp_y15 = temp_y14 + 10.7
            temp_x16 = temp_x15 - 200
            temp_y16 = temp_y15
            temp_x17 = temp_x16
            temp_y17 = temp_y16 - 10.7
            temp_x18 = temp_x17 - 50
            temp_y18 = temp_y17
            temp_x19 = temp_x18
            temp_y19 = temp_y18 + 10.7
            temp_x20 = temp_x19 - CW/2 + 181
            temp_y20 = temp_y19

            prev_x, prev_y = temp_x1, temp_y1  # 첫 번째 점으로 초기화
            lastNum = 20
            for i in range(1, lastNum + 1):
                curr_x, curr_y = eval(f'temp_x{i}'), eval(f'temp_y{i}')
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="4")
                prev_x, prev_y = curr_x, curr_y
            line(doc, prev_x, prev_y, temp_x1, temp_y1, layer="4")

        #하부 가로 긴 브라켓
        insert_block(x1 + 31, y1 + ELYpos, "Ceiling_Bracket_horizontal_holeL")
        insert_block(x2 - 31, y1 + ELYpos, "Ceiling_Bracket_horizontal_holeR")
        rectangle(doc, x1 + 31, y1 + ELYpos + 47, x2 - 31, y1 + ELYpos + 65, layer='2')#상
        rectangle(doc, x1 + 31, y1 + ELYpos - 65, x2 - 31, y1 + ELYpos - 47, layer='2')#하
        line(doc, x1 + 31, y1 + ELYpos - 48.6, x2 - 31, y1 + ELYpos - 48.6, layer="Hidden")
        line(doc, x1 + 31, y1 + ELYpos + 48.6, x2 - 31, y1 + ELYpos + 48.6, layer="Hidden")
        insert_block(x1 + 31 + (CW - 12)/3, y1 + ELYpos, "Ceiling_Bracket_horizontal_hole4")
        insert_block(x2 - 31 - (CW - 12)/3, y1 + ELYpos, "Ceiling_Bracket_horizontal_hole4")
        #하부 가로 긴 브라켓 길이 저장
        mcsizex1 = x2 - 31 - (x1 + 31)
        
        #좌 세로 브라켓
        rectangle(doc, x1 + (MCW/3 - 130), y1 + ELYpos + 66 , x1 + MCW/3, y2 - 30.5, layer='2')
        rectangle(doc, x1 + (MCW/3 - 130) + 18, y1 + ELYpos + 66, x1 + MCW/3 - 18, y2 - 30.5, layer='2')
        rectangle(doc, x1 + (MCW/3 - 130) + 16, y1 + ELYpos + 66, x1 + MCW/3 - 16, y2 - 30.5, layer='Hidden')
        insert_block(x1 + (MCW/3 - 65), y1 + ELYpos + 96, "Ceiling_Bracket_vertical_hole2")
        insert_block(x1 + (MCW/3 - 65), y2 - 60.5, "Ceiling_Bracket_vertical_hole2")
        insert_block(x1 + (MCW/3 - 65), y1 + ELYpos + 66 + (MCD - (ELYpos + 96.5))/3, "Ceiling_Bracket_vertical_hole4")
        insert_block(x1 + (MCW/3 - 65), y1 + ELYpos + 66 + (MCD - (ELYpos + 96.5))*2/3, "Ceiling_Bracket_vertical_hole4")
        #좌우 세로 긴 브라켓 길이 저장
        mcsizex3 = y2 - 30.5 - (y1 + ELYpos + 66)
        #우 세로 브라켓
        rectangle(doc, x1 + MCW*2/3, y1 + ELYpos + 66, x1 + (MCW*2/3 + 130), y2 - 30.5, layer='2')
        rectangle(doc, x1 + MCW*2/3 + 18, y1 + ELYpos + 66, x1 + (MCW*2/3 + 130) - 18, y2 - 30.5, layer='2')
        rectangle(doc, x1 + MCW*2/3 + 16, y1 + ELYpos + 66, x1 + (MCW*2/3 + 130) - 16, y2 - 30.5, layer='Hidden')
        insert_block(x1 + (MCW*2/3 + 65), y1 + ELYpos + 96, "Ceiling_Bracket_vertical_hole2")
        insert_block(x1 + (MCW*2/3 + 65), y2 - 60.5, "Ceiling_Bracket_vertical_hole2")
        insert_block(x1 + (MCW*2/3 + 65), y1 + ELYpos + 66 + (MCD - (ELYpos + 96.5))/3, "Ceiling_Bracket_vertical_hole4")
        insert_block(x1 + (MCW*2/3 + 65), y1 + ELYpos + 66 + (MCD - (ELYpos + 96.5))*2/3, "Ceiling_Bracket_vertical_hole4")
        
        #중간 가로 브라켓 상
        rectangle(doc, x1 + MCW/3 + 0.5, y2 - 30 - MDFirstSpace - 130, x1 + MCW*2/3 - 0.5, y2 - 30 - MDFirstSpace, layer='2')
        rectangle(doc, x1 + MCW/3 + 0.5, y2 - 30 - MDFirstSpace - 130 + 18, x1 + MCW*2/3 - 0.5, y2 - 30 - MDFirstSpace - 130 + 112, layer='2')
        rectangle(doc, x1 + MCW/3 + 0.5, y2 - 30 - MDFirstSpace - 130 + 16, x1 + MCW*2/3 - 0.5, y2 - 30 - MDFirstSpace - 130 + 114, layer='Hidden')
        insert_block(x1 + MCW/3 + 30.5, y2 - 30 - MDFirstSpace - 65, "Ceiling_Bracket_horizontal_hole2")
        insert_block(x1 + MCW*2/3 - 30.5, y2 - 30 - MDFirstSpace - 65, "Ceiling_Bracket_horizontal_hole2")
        insert_block(x1 + MCW/2, y2 - 30 - MDFirstSpace - 65, "Ceiling_Bracket_horizontal_hole4")
        
        #중간 가로 브라켓 중 (MDSecondSpace가 400이상일 때만 그림)
        if MDSecondSpace >= 400:
            rectangle(doc, x1 + MCW/3 + 0.5, y2 - 30 - MDFirstSpace - 130 - MDCenterBK - 65, x1 + MCW*2/3 - 0.5, y2 - 30 - MDFirstSpace - 130 - MDCenterBK + 65, layer='2')
            rectangle(doc, x1 + MCW/3 + 0.5, y2 - 30 - MDFirstSpace - 130 + 18 - MDCenterBK - 65, x1 + MCW*2/3 - 0.5, y2 - 30 - MDFirstSpace - 130 + 112 - MDCenterBK - 65, layer='2')
            rectangle(doc, x1 + MCW/3 + 0.5, y2 - 30 - MDFirstSpace - 130 + 16 - MDCenterBK - 65, x1 + MCW*2/3 - 0.5, y2 - 30 - MDFirstSpace - 130 + 114 - MDCenterBK - 65, layer='Hidden')
            insert_block(x1 + MCW/3 + 30.5, y2 - 30 - MDFirstSpace - 65 - MDCenterBK - 65, "Ceiling_Bracket_horizontal_hole2")
            insert_block(x1 + MCW*2/3 - 30.5, y2 - 30 - MDFirstSpace - 65 - MDCenterBK - 65, "Ceiling_Bracket_horizontal_hole2")
            insert_block(x1 + MCW/2, y2 - 30 - MDFirstSpace - 65 - MDCenterBK - 65, "Ceiling_Bracket_horizontal_hole4")
            
        #중간 가로 브라켓 하
        rectangle(doc, x1 + MCW/3 + 0.5, y1 + (MCD - SplitYpos1) - 65, x1 + MCW*2/3 - 0.5, y1 + (MCD - SplitYpos1) + 65, layer='2')
        rectangle(doc, x1 + MCW/3 + 0.5, y1 + (MCD - SplitYpos1) - 65 + 18, x1 + MCW*2/3 - 0.5, y1 + (MCD - SplitYpos1) + 65 - 18, layer='2')
        rectangle(doc, x1 + MCW/3 + 0.5, y1 + (MCD - SplitYpos1) - 65 + 16, x1 + MCW*2/3 - 0.5, y1 + (MCD - SplitYpos1) + 65 - 16, layer='Hidden')
        insert_block(x1 + MCW/3 + 30.5, y1 + (MCD - SplitYpos1), "Ceiling_Bracket_horizontal_hole2")
        insert_block(x1 + MCW*2/3 - 30.5, y1 + (MCD - SplitYpos1), "Ceiling_Bracket_horizontal_hole2")
        insert_block(x1 + MCW/2, y1 + (MCD - SplitYpos1), "Ceiling_Bracket_horizontal_hole4")
        
        #좌측 가로 브라켓
        rectangle(doc, x1 + 30.5, y1 + (MCD - SplitYpos1) - 15, x1 + MCW/3 - 130 - 0.5, y1 + (MCD - SplitYpos1) + 15, layer='2')
        line(doc, x1 + 30.5, y1 + (MCD - SplitYpos1) + 13, x1 + MCW/3 - 130 - 0.5, y1 + (MCD - SplitYpos1) + 13, layer="2")
        rectangle(doc, x1 + 30.5, y2 - 700 - 15, x1 + MCW/3 - 130 - 0.5, y2 - 700 + 15, layer='2')
        line(doc, x1 + 30.5, y2 - 700 + 13, x1 + MCW/3 - 130 - 0.5, y2 - 700 + 13, layer="2")
        #좌우 가로 브라켓 길이 저장
        mcsizex2 = x1 + MCW/3 - 130 - 0.5 - (x1 + 30.5)
        #우측 가로 브라켓
        rectangle(doc, x2 - 30.5, y1 + (MCD - SplitYpos1) - 15, x2 - (MCW/3 - 130 - 0.5), y1 + (MCD - SplitYpos1) + 15, layer='2')
        line(doc, x2 - 30.5, y1 + (MCD - SplitYpos1) + 13, x2 - (MCW/3 - 130 - 0.5), y1 + (MCD - SplitYpos1) + 13, layer="2")
        rectangle(doc, x2 - 30.5, y2 - 700 - 15, x2 - (MCW/3 - 130 - 0.5), y2 - 700 + 15, layer='2')
        line(doc, x2 - 30.5, y2 - 700 + 13, x2 - (MCW/3 - 130 - 0.5), y2 - 700 + 13, layer="2")
        
        #팬, 서스테이너
        insert_block(x1 + 75, y2 - 140, "Ceiling_Fan")
        line(doc, x1 - 90, y1 + 25 - 140, x1 + 10, y1 + 25 - 140, layer="0")#140 선
        insert_block(x1 + MCW/2 - AS_raw/2, y1 + 25 - 140 + EE_raw, "Ceiling_Sustainer")
        insert_block(x1 + MCW/2 - AS_raw/2 + AS_raw, y1 + 25 - 140 + EE_raw, "Ceiling_Sustainer_R")
        
        #상부 치수선  
        dim(doc, x1, y2, x2, y2, 250, direction="up", localdimstyle="0.2 JKW")#상부 전체
        dim(doc, x1 + Xgap, y2 - Ygap, x1, y2, 210,  text_height=0.32, text_gap=0.07, direction="up", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 + MCW/3 - 130, y2 - Ygap)
        dimcontinue(doc, x1 + MCW/3, y2 - Ygap)
        dimcontinue(doc, x1 + MCW*2/3, y2 - Ygap)
        dimcontinue(doc, x1 + MCW*2/3 + 130, y2 - Ygap)
        dimcontinue(doc, x2 - Xgap, y2 - Ygap)
        dimcontinue(doc, x2, y2)
        #좌측 치수선
        dim(doc, x1, y2, x1 + 70, y2 - 140, 100, direction="left", localdimstyle="0.2 JKW")#팬
        dimcontinue(doc, x1 + 70, y2 - 520)
        dim(doc, x1 + 75, y2 - 530, x1, y2 - 500, 100, direction="down", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 + 165, y2 - 530)
        dim(doc, x1, y2, x1, y1 + (MCD - SplitYpos1), 150, direction="left", localdimstyle="0.2 JKW")#분할치수
        dimcontinue(doc, x1, y1)
        dim(doc, x1, y1 + 25, x1, y1 + 25 - 140, 50, direction="left", localdimstyle="0.2 JKW")#140 선
        textstr =  f"EE = {EE_raw}"
        dim(doc, x1 - 90, y1 + 25 - 140 + EE_raw, x1 - 90, y1 + 25 - 140, 200, direction="left", localdimstyle="0.2 JKW", text=textstr)#EE
        dim(doc, x1, y1, x1 + 33, y1 + ELYpos, 100, direction="left", localdimstyle="0.2 JKW")#header
        #하부 치수선
        textstr =  f"AS = {AS_raw}"
        dim(doc, x1 + MCW/2 - AS_raw/2, y1 + 585, x1 + MCW/2 - AS_raw/2 + AS_raw, y1 + 585, 790, direction="down", localdimstyle="0.2 JKW", text=textstr)#EE
        #우측 치수선
        dim(doc, x2, y2, x2, y1, 260, direction="right", localdimstyle="0.2 JKW")#전체
        #내부 치수선
        dim(doc, x1 + 225, y1 + ELYpos + 30, x1, y1 + ELYpos + 30, 100, direction="up", option="reverse", localdimstyle="0.2 JKW")#header 위 가로 브라켓
        dimcontinue(doc, x2 - 225, y1 + ELYpos + 30)
        dimcontinue(doc, x2, y1 + ELYpos + 30)
        
        dim(doc, x1 + MCW/3, y1 + ELYpos + 65.5, x1 + MCW/3, y2 - 30, 100, direction="right", localdimstyle="0.2 JKW")
        dim(doc, x1 + MCW/3, y1 + ELYpos + 65.5, x1 + MCW/3, y1 + (MCD - SplitYpos1) - 65, 50, direction="right", localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 + MCW/3, y1 + (MCD - SplitYpos1) - 65 + 130)
        dimcontinue(doc, x1 + MCW/3, y2 - 30 - MDFirstSpace - 130 - MDCenterBK - 65)
        dimcontinue(doc, x1 + MCW/3, y2 - 30 - MDFirstSpace - 130 - MDCenterBK - 65 + 130)
        dimcontinue(doc, x1 + MCW/3, y2 - 30 - MDFirstSpace - 130)
        dimcontinue(doc, x1 + MCW/3, y2 - 30 - MDFirstSpace)
        dimcontinue(doc, x1 + MCW/3, y2 - 30)
    
    #########################################################################
    # 1page 우측단면도 표기 (블럭 삽입)
    #########################################################################
    program = 1
    if program:
        x1 = x1 + MCW + 850
        insert_block(x1, y2, "Ceiling_Top_section")
        insert_block(x1 - 30, y1 + 25 - 140 + EE_raw, "Ceiling_Mid_section")
        
        if headertype == "이노에스템":
            insert_block(x1, y1, "Ceiling_Bottom_section_inoestem")
        if headertype == "정일산업":
            insert_block(x1, y1, "Ceiling_Bottom_section_joungil")
        if headertype == "에이스코리아":
            insert_block(x1, y1, "Ceiling_Bottom_section_acekorea")
        if headertype == "세화정공":
            insert_block(x1, y1, "Ceiling_Bottom_section_sehwa")

        rectangle(doc, x1, y1, x1 + 2.3, y2, layer="0")
        line(doc, x1 + 2.3, y1 + 50, x1 + 2.3, y2 - 50, layer="6")
        line(doc, x1 + 2.3 + 15, y1 + 50, x1 + 2.3 + 15, y2 - 50, layer="6")
        line(doc, x1 + 2.3 + 115, y1 + 50, x1 + 2.3 + 115, y2 - 50, layer="6")
        line(doc, x1 + 2.3 + 150, y1 + 50, x1 + 2.3 + 150, y2 - 50, layer="6")
        line(doc, x1 - 100, y1 + 31, x1 - 100, y1 + 25 - 140 + EE_raw - 200, layer="2")
        line(doc, x1 - 100, y2 - 31, x1 - 100, y1 + 25 - 140 + EE_raw + 200, layer="2")
        
        dim(doc, x1, y1, x1 - 30, y1 + 25 - 140 + EE_raw, 200, direction="left", localdimstyle="0.2 JKW")
        dim(doc, x1, y1, x1 - 30, y1 + 30, 260, direction="left", localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 - 100, y1 + 25 - 140 + EE_raw - 200)
        dimcontinue(doc, x1 - 100, y1 + 25 - 140 + EE_raw + 200)
        dimcontinue(doc, x1 - 100, y2 - 30)
        dimcontinue(doc, x1 - 100, y2, option="reverse")
        dim(doc, x1, y1, x1 - 30, y1 + 25 - 140 + EE_raw - 150, 320, direction="left", localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 - 30, y1 + 25 - 140 + EE_raw + 150)
        dimcontinue(doc, x1, y2)
        dim(doc, x1, y1, x1, y2, 380, direction="left", localdimstyle="0.2 JKW")

        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1715
        TargetXscale = 3800 + (CW - 1600) * 1.6 + (CD - 1500) * 1.5
        TargetYscale = 1700 + (CD - 1500)
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        # print("스케일 비율 : " + str(frame_scale))        
        frameYpos = abs_y - 500 * frame_scale     
        insert_frame(frameXpos, frameYpos, frame_scale, "drawings_frame", workplace)      
         
        frameXpos = frameXpos + TargetXscale + 400

    #####################################################################################################################################
    # 2page 천장 본판
    ##################################################################################################################################### 
    program = 1
    if program:    
        # 기본 윤곽 car inside
        x1 = frameXpos + 400
        y1 = abs_y 
        x2 = x1 + MCW
        y2 = y1 + MCD
        rectangle(doc, x1, y1, x2, y2, layer='레이져')
        line(doc, x1, y1 + (MCD - SplitYpos1), x2, y1 + (MCD - SplitYpos1), layer="레이져")

        #테두리 장공
        sum = 0
        for x in frontPanel[:-1] :
            sum = sum + x
            insert_block(x1 + sum, y1 + slotposition, "8x20 slot_Laser")
            dim(doc, x1 + sum, y1, x1 + sum - x, y1, 70, direction="down", localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + sum, y1)
        dimcontinue(doc, x2, y1)

        sum = 0
        for x in rearPanel[:-1] :
            sum = sum + x
            insert_block(x1 + sum, y2 - slotposition, "8x20 slot_Laser")
            dim(doc, x1 + sum, y2, x1 + sum - x, y2, 70, direction="up", localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + sum, y2)
        dimcontinue(doc, x2, y2)

        sum = 0
        for y in sidePanel[:-1] :
            sum = sum + y
            insert_block(x1 + slotposition, y1 + sum, "8x20 slotLaser")
            
        sum = 0
        for y in sidePanel[:-1] :
            sum = sum + y
            insert_block(x2 - slotposition, y1 + sum, "8x20 slotLaser")
            dim(doc, x2, y1 + sum, x2, y1 + sum - y, 150, direction="right", localdimstyle="0.2 JKW")    
            dimcontinue(doc, x2, y1 + sum)
        dimcontinue(doc, x2, y2)

        inputCW = CW
        inputCD = CD
        
        if lctype == "035":
            #LC 조립홀 가로
            lc_positions = calculate_lc_position_035(inputCW, inputCD)

            Topholex = lc_positions['Topholex']
            Topholey = lc_positions['Topholey']

            Bottomx = lc_positions['Bottomx']
            Bottomy = lc_positions['Bottomy']    

            for i, x in enumerate(Topholex):
                circle_cross(doc, x1 + 25 + x, y2 - 65, 13, layer='레이져')
                if i == 0:
                    dim(doc, x1 + 25 + x, y2 - 60, x1, y2, 180, direction="up", option='reverse', localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + x + 25, y2 - 60)
            dimcontinue(doc ,x2, y2) 

            for i, x in enumerate(Bottomx):
                circle_cross(doc, x1 + 25 + x, y1 + 65, 13, layer='레이져')
                if i == 0:
                    dim(doc, x1 + 25 + x, y1 + 55, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + x + 25, y1 + 50)
            dimcontinue(doc ,x2, y1) 
            
            #LC 조립홀 세로
            hole_positions = calculate_lc_vertical_hole_positions_035(inputCW, inputCD)

            leftholex = hole_positions['leftholex']
            rightholex = hole_positions['rightholex']
            vertical_holey = hole_positions['vertical_holey']

            num_holes = min(len(leftholex), len(rightholex), len(vertical_holey))# 리스트 길이 확인

            # circle_cross 함수 호출 시 리스트 길이를 확인하여 오류 방지
            for i in range(num_holes):
                circle_cross(doc, x1 + 25 + leftholex[i], y1 + 25 + vertical_holey[i], 13, layer='레이져')        
                circle_cross(doc, x1 + 25 + rightholex[i], y1 + 25 + vertical_holey[i], 13, layer='레이져')
                if i == 0:
                    dim(doc, x2 - 135, y1 + 65, x1 + 25 + rightholex[i], y1 + 25 + vertical_holey[i], 360, direction="right", localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + 25 + rightholex[i], y1 + 25 + vertical_holey[i])      
            dimcontinue(doc, x2 - 125, y2 - 65)
            dim(doc, x2, y1, x2 - 125, y1 + 65, 225, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x2, y2, x2 - 125, y2 - 65, 225, direction="right", localdimstyle="0.2 JKW")
            
        elif lctype == "008_A":
            lc_positions = calculate_lc_position_008_A(inputCW, inputCD)

            # Topholex와 Topholey 좌표 추출
            Topholex = lc_positions['Topholex']
            Topholey = lc_positions['Topholey']

            # middle 좌표 추출
            Middlex = lc_positions['Middlex']
            Middley = lc_positions['Middley']    

            # Bottomx와 Bottomy 좌표 추출
            Bottomx = lc_positions['Bottomx']
            Bottomy = lc_positions['Bottomy']    

            # 필요한 경우 다른 홀들에 대해서도 circle_cross 함수를 호출할 수 있음
            # 예를 들어, 모든 상단 홀에 대해 반복
            for i, x in enumerate(Topholex):
                circle_cross(doc, x1 + 25 + x, y1 + 25 + Topholey, 11, layer='레이져')
                if i == 0:
                    dim(doc, x1 + 25 + x, y2 - 90, x1, y2, 210, direction="up", option='reverse', localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + x + 25, y2 - 90)
            dimcontinue(doc, x2, y2) 

            for x in Middlex:
                circle_cross(doc, x1 + 25 + x, y1 + 25 + Middley, 11, layer='레이져')            

            for x in Bottomx:
                circle_cross(doc, x1 + 25 + x, y1 + 25 + Bottomy, 11, layer='레이져')     
                       
            dim(doc, x2 - 90, y1 + 100, x2, y1, 290, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x2 - 90, y1 + MCD/2) 
            dimcontinue(doc, x2 - 90, y2 - 100) 
            dimcontinue(doc, x2, y2) 
            
        elif lctype == "031":
            lc_positions = calculate_lc_position_031(inputCW, inputCD)

            Topholex = lc_positions['Topholex']
            Topholey = lc_positions['Topholey']

            Bottomx = lc_positions['Bottomx']
            Bottomy = lc_positions['Bottomy']    
            
            for i, x in enumerate(Topholex):
                circle_cross(doc, x1 + 25 + x, y1 + 25 + Topholey, 13, layer='레이져')
                if i == 0:
                    dim(doc, x1 + 25 + x, y2 - 60, x1, y2, 180, direction="up", option='reverse', localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + x + 25, y2 - 60)
            dimcontinue(doc, x2, y2) 

            for i, x in enumerate(Bottomx):
                circle_cross(doc, x1 + 25 + x, y1 + 65, 13, layer='레이져')
                if i == 0:
                    dim(doc, x1 + 25 + x, y1 + 50, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + x + 25, y1 + 50)
            dimcontinue(doc, x2, y1) 
                
            # 세로 홀 위치 좌표 계산
            hole_positions = calculate_lc_vertical_hole_positions_031(inputCW, inputCD)

            # 각 홀의 좌표를 추출
            leftholex = hole_positions['leftholex']
            rightholex = hole_positions['rightholex']
            vertical_holey = hole_positions['vertical_holey']

            # 리스트 길이 확인
            num_holes = min(len(leftholex), len(rightholex), len(vertical_holey))

            # circle_cross 함수 호출 시 리스트 길이를 확인하여 오류 방지
            for i in range(num_holes):
                circle_cross(doc, x1 + 25 + leftholex[i], y1 + 25 + vertical_holey[i], 13, layer='레이져')
                circle_cross(doc, x1 + 25 + rightholex[i], y1 + 25 + vertical_holey[i], 13, layer='레이져')
                if i == 0:
                    dim(doc, x2 - 135, y1 + 65, x1 + 25 + rightholex[i], y1 + 25 + vertical_holey[i], 335, direction="right", localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + 25 + rightholex[i] , y1 + 25 + vertical_holey[i], 340)      
            dimcontinue(doc, x2 - 125, y2 - 65, 340)
            dim(doc, x2, y1, x2 - 125, y1 + 65, 200, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x2, y2, x2 - 125, y2 - 65, 200, direction="right", localdimstyle="0.2 JKW")
            
        elif lctype == "N20": 
            circle_cross(doc, x1 + 137.5, y1 + 70, 13, layer='레이져')
            circle_cross(doc, x1 + 137.5 + 185, y1 + 70, 13, layer='레이져')
            circle_cross(doc, x2 - 137.5, y1 + 70, 13, layer='레이져')
            circle_cross(doc, x2 - 137.5 - 185, y1 + 70, 13, layer='레이져')
            circle_cross(doc, x1 + 137.5, y2 - 70, 13, layer='레이져')
            circle_cross(doc, x1 + 137.5 + 185, y2 - 70, 13, layer='레이져')
            circle_cross(doc, x2 - 137.5, y2 - 70, 13, layer='레이져')
            circle_cross(doc, x2 - 137.5 - 185, y2 - 70, 13, layer='레이져')
            dim(doc, x1 + 137.5, y2 - 60, x1, y2, 180, direction="up", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 137.5 + 185, y2 - 60)
            dimcontinue(doc, x2 - (137.5 + 185), y2 - 60)
            dimcontinue(doc, x2 - 137.5, y2 - 60)
            dimcontinue(doc, x2, y2)
            dim(doc, x1 + 137.5, y1 + 60, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 137.5 + 185, y1 + 60)
            dimcontinue(doc, x2 - (137.5 + 185), y1 + 60)
            dimcontinue(doc, x2 - 137.5, y1 + 60)
            dimcontinue(doc, x2, y1)
            dim(doc, x2 - 125, y1 + 70, x2, y1, 330, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x2 - 125, y2 - 70)
            dimcontinue(doc, x2, y2)

        elif lctype == "032":
            circle_cross(doc, x1 + 170, y1 + 65, 13, layer='레이져')
            circle_cross(doc, x1 + 170 + 150, y1 + 65, 13, layer='레이져')
            circle_cross(doc, x2 - 170, y1 + 65, 13, layer='레이져')
            circle_cross(doc, x2 - 170 - 150, y1 + 65, 13, layer='레이져')
            circle_cross(doc, x1 + 170, y2 - 65, 13, layer='레이져')
            circle_cross(doc, x1 + 170 + 150, y2 - 65, 13, layer='레이져')
            circle_cross(doc, x2 - 170, y2 - 65, 13, layer='레이져')
            circle_cross(doc, x2 - 170 - 150, y2 - 65, 13, layer='레이져')
            dim(doc, x1 + 170, y2 - 60, x1, y2, 180, direction="up", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 170 + 150, y2 - 60)
            dimcontinue(doc, x2 - (170 + 150), y2 - 60)
            dimcontinue(doc, x2 - 170, y2 - 60)
            dimcontinue(doc, x2, y2)
            dim(doc, x1 + 170, y1 + 60, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 170 + 150, y1 + 60)
            dimcontinue(doc, x2 - (170 + 150), y1 + 60)
            dimcontinue(doc, x2 - 170, y1 + 60)
            dimcontinue(doc, x2, y1)
            dim(doc, x2 - 160, y1 + 65, x2, y1, 360, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x2 - 160, y2 - 65)
            dimcontinue(doc, x2, y2)

        elif lctype == "026":
            circle_cross(doc, x1 + 165, y1 + 65, 13, layer='레이져')
            circle_cross(doc, x1 + 165 + 160, y1 + 65, 13, layer='레이져')
            circle_cross(doc, x2 - 165, y1 + 65, 13, layer='레이져')
            circle_cross(doc, x2 - 165 - 160, y1 + 65, 13, layer='레이져')
            
            circle_cross(doc, x1 + 165, y2 - 65, 13, layer='레이져')
            circle_cross(doc, x1 + 165 + 160, y2 - 65, 13, layer='레이져')
            circle_cross(doc, x2 - 165, y2 - 65, 13, layer='레이져')
            circle_cross(doc, x2 - 165 - 160, y2 - 65, 13, layer='레이져')
            circle_cross(doc, x1 + 165, y1 + MCD/2, 13, layer='레이져')
            circle_cross(doc, x2 - 165, y1 + MCD/2, 13, layer='레이져')
            
            dim(doc, x1 + 165, y2 - 60, x1, y2, 180, direction="up", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 165 + 160, y2 - 60)
            dimcontinue(doc, x2 - (165 + 160), y2 - 60)
            dimcontinue(doc, x2 - 165, y2 - 60)
            dimcontinue(doc, x2, y2)
            dim(doc, x1 + 165, y1 + 60, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
            dimcontinue(doc, x1 + 165 + 160, y1 + 60)
            dimcontinue(doc, x2 - (165 + 160), y1 + 60)
            dimcontinue(doc, x2 - 165, y1 + 60)
            dimcontinue(doc, x2, y1)
            dim(doc, x2 - 160, y1 + 65, x2, y1, 360, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x2 - 160, y1 + MCD/2)
            dimcontinue(doc, x2 - 160, y2 - 65)
            dimcontinue(doc, x2, y2)
            
        circle_cross(doc, x1 + 225, y1 + ELYpos, 20, layer='레이져')
        circle_cross(doc, x2 - 225, y1 + ELYpos, 20, layer='레이져')
        circle_cross(doc, x1 + 40, y1 + ELYpos, 8, layer='레이져')
        text = "2-%%C20 Hole(전선인입홀)"
        dim_leader(doc, x1 + 225, y1 + ELYpos, x1 + 225 + 50, y1 + ELYpos + 50, text, direction="leftToright", distance=13)   # distance 13은 글자높이임       
        text = "%%C8"
        dim_leader(doc, x1 + 40, y1 + ELYpos, x1 + 40 + 50, y1 + ELYpos + 50, text, direction="leftToright", distance=13)   # distance 13은 글자높이임       
        text = "%%C8x20 Slot"
        dim_leader(doc, x1 + 80, y1 + slotposition, x1 + 80 + 50, y1 + slotposition + 100, text, direction="leftToright", distance=12)
      
        insert_block(x1 + 75, y2 - 140, "Ceiling_Fan_laser")
        dim(doc, x1, y2, x1 + 43, y2 - 147, 50, direction="left", localdimstyle="0.2 JKW")#팬
        dimcontinue(doc, x1 + 43, y2 - 513)
        dim(doc, x1, y2, x1 + 70, y2 - 140, 100, direction="left", localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 + 70, y2 - 520)
        dim(doc, x1 + 75, y2 - 530, x1, y2 - 500, 100, direction="down", option="reverse", localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 + 165, y2 - 530)
        dim(doc, x1, y2 - 147, x1 + 43, y2 - 147, 50, direction="down", localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 + 93, y2 - 147)
        text = "4-%%C9.2"
        dim_leader(doc, x1 + 165, y2 - 520, x1 + 165 + 50, y2 - 520 + 50, text, direction="leftToright", distance=13)   # distance 13은 글자높이임       
                
        dim(doc, x1 + 225, y1 + 267, x1, y1 + 225, 100, direction="up", option="reverse", localdimstyle="0.2 JKW")#헤더 위 가로 브라켓
        dimcontinue(doc, x2 - 225, y1 + 267)
        dimcontinue(doc, x2, y1 + 267)
        dim(doc, x1, y1, x1 + 30, y1 + 252.5, 100, direction="left", localdimstyle="0.2 JKW")
        dim(doc, x1 + MCW/2 - 30, y1, x1 + MCW/2 - 30, y1 + slotposition, 20, direction="left", option="reverse", localdimstyle="0.2 JKW")#테두리장공폭
        dim(doc, x1 + MCW/2 - 30, y2, x1 + MCW/2 - 30, y2 - slotposition, 20, direction="left", option="reverse", localdimstyle="0.2 JKW")
        dim(doc, x1, y1 + 118, x1 + slotposition, y1 + 118, 30, direction="up", localdimstyle="0.2 JKW")
        dim(doc, x2, y1 + 118, x2 - slotposition, y1 + 118, 30, direction="up", localdimstyle="0.2 JKW")
        dim(doc, x1, y2, x1, y1 + (MCD - SplitYpos1), 150, direction="left", localdimstyle="0.2 JKW")#분할치수
        dimcontinue(doc, x1, y1)
        dim(doc, x2, y1, x2, y2, 300, direction="right", localdimstyle="0.2 JKW")#전체
        dim(doc, x1, y1, x2, y1, 210, direction="down", localdimstyle="0.2 JKW")
        
        # main Frame description    
        x1 = x1 + CW + 500
        y1 = y1 + 1300
        textstr = f"Part Name : Ceiling Plate"
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
        textstr = f"Mat.Spec : SPCC 2.3T"
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
        textstr = f"Size : {MCW} x 1000mm"
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
        textstr = f"Quantity : {su} EA"
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')

        y1 = y1 - 1000
        textstr = f"Part Name : Ceiling Plate"
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
        textstr = f"Mat.Spec : SPCC 2.3T"
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
        textstr = f"Size : {MCW} x {MCD - 1000}mm"
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
        textstr = f"Quantity : {su} EA"
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')
        
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1715
        TargetXscale = 3800 + (CW - 1600) * 1.6 + (CD - 1500) * 1.5 
        TargetYscale = 1700 + (CD - 1500) 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale

        # print("스케일 비율 : " + str(frame_scale))        
        frameYpos = abs_y - 500 * frame_scale   
        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)       

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 3page front, rear, side B/K
    #################################################################################################        
    program = 1
    if program:
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = BasicXscale + (max(CW,CD) - 1500) 
        TargetYscale = BasicYscale
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        insert_frame(frameXpos, frameYpos, frame_scale, "drawings_frame", workplace)   

    ###################################################
    # 3page front B/K (발보호판)
    ###################################################
        rx1 = frameXpos + 200
        ry1 = frameYpos + 1100 * frame_scale + 580
        insert_block(rx1, ry1 + 26, "Ceiling_Bracket_Section_R")

        def footOP(OP_raw):
            """엑셀에서 가져온 OP_raw 값에 따라 footOP 값을 반환하는 함수"""
            if OP_raw == 700:
                return 895
            elif OP_raw == 800:
                return 1055
            elif OP_raw == 900:
                return 1215
            elif OP_raw == 1000:
                return 1330
            elif OP_raw == 1100:
                return 1425
            else:
                return None  # 정의되지 않은 값은 None 반환
        
        footOP_value = footOP(OP_raw)
        
        rx1 = frameXpos + 400
        
        thickness = 2
        vcut = thickness / 2

        x1 = rx1
        y1 = ry1
        x2 = x1 + MCW
        y2 = y1
        x3 = x2
        y3 = y2 + 30 - thickness
        x4 = x3
        y4 = y3 + 130 - thickness
        x5 = x4 - (MCW - footOP_value)/2
        y5 = y4
        x6 = x5 - footOP_value
        y6 = y5
        x7 = x1
        y7 = y6
        x8 = x7
        y8 = y7 - 130 + thickness

        Radiusval = 20          
        lastNum = 8
        
        prev_x, prev_y = eval('x1'), eval('y1')

        def draw_arc(doc, start, end, radius, option=None):
            """ option='reverse'는 시계방향으로 점을 회전할때 이 것을 사용한다. """
            if option == 'reverse' :                
                line(doc, prev_x, prev_y, start[0], start[1], layer='레이져')                
                add_arc_between_points(doc, end, start, radius)
            else:                
                line(doc, prev_x, prev_y, start[0], start[1], layer='레이져')                
                add_arc_between_points(doc, start, end, radius)

        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')

            # print (f" x{i} : prev_x : {prev_x} , curr_x : {curr_x}")

            if i == 4:
                draw_arc(doc, (curr_x, curr_y - Radiusval), (curr_x - Radiusval, curr_y), Radiusval)
                prev_x, prev_y = curr_x - Radiusval, curr_y              
                
                # (x5, y5) → (x6, y6) 구간은 건너뛰기
            elif i == 6:
                prev_x, prev_y = curr_x, curr_y

            elif i == 7:
                draw_arc(doc, (curr_x+Radiusval , curr_y ), (curr_x , curr_y - Radiusval), Radiusval)
                prev_x, prev_y = curr_x , curr_y-Radiusval               

            else:
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y
                
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")
        
        block_name = f"Ceiling_foot_OP{OP_raw}"
        block_name_laser = f"Ceiling_foot_OP{OP_raw}_laser"
        
        insert_block(x5, y5, block_name)
        insert_block(x5, y5, block_name_laser)
        
        
        #절곡라인  (히든선 : 정방향)      
        line(doc, x8, y8, x3, y3, layer='hidden')     
        #front 장공
        sum = 0
        for x in frontPanel[:-1] :
            sum = sum + x
            insert_block(x1 + sum, y1 + slotposition, "8x20 slot_Laser")
            dim(doc, x1 + sum, y1, x1 + sum - x, y1, 70, direction="down", localdimstyle="0.2 JKW")    
            dimcontinue(doc , x1 + sum, y2)
        dimcontinue(doc, x2, y2)

        circle_cross(doc, x1 + 150, y1 + 30 - thickness, 30, layer='레이져')
        circle_cross(doc, x2 - 150, y1 + 30 - thickness, 30, layer='레이져')

        # 상부 전체치수
        dim(doc, x4, y4, x7, y7, 140, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x4, y4, x4 - (MCW - footOP_value)/2 - 20, y4, 80, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x7, y7, x7 + (MCW - footOP_value)/2 + 20, y7, 80, direction="up", option='reverse', localdimstyle="0.2 JKW")
        
        # 우측 치수선
        dim(doc, x3, y3, x2, y2, 100, direction="right", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x4, y4)
        dim(doc, x2, y2, x4, y4, 150, direction="right", localdimstyle="0.2 JKW")
        
        text = "8-%%C30 Hole"
        dim_leader(doc, x1 + 150, y1 + 28, x1 + 150 + 50, y1 + 28 + 50, text, direction="leftToright", distance=17)       
        text = "%%C8x20 Slot"
        dim_leader(doc, x2 - 80, y1 + slotposition, x2 - 80 - 100, y1 + slotposition + 100, text, direction="rightToleft", distance=12)
        text = "2-R20"
        dim_leader(doc, x4 - 10, y4 - 10, x4 + 50, y4 + 50, text, direction="leftToright", distance=17)
                            
        # main Frame description    
        x1 = x1 + max(MCW,MCD) + 200
        y1 = y1 + 100
        textstr = f"Part Name : Ceiling Front B/K"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : SPCC 2.3t"    
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')    
        textstr = f"Size : 156 x {MCW}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')         

    ###################################################
    # 3page rear B/K
    ###################################################
        rx1 = frameXpos + 200
        ry1 = frameYpos + 1100 * frame_scale
        insert_block(rx1, ry1 + 26, "Ceiling_Bracket_Section")    

        rx1 = frameXpos + 400
        
        thickness = 2
        vcut = thickness / 2

        x1 = rx1
        y1 = ry1
        x2 = x1 + MCW
        y2 = y1    
        x3 = x2 
        y3 = y2 + 30 - thickness
        x4 = x3
        y4 = y3 + 130 - thickness
        x5 = x4 - MCW
        y5 = y4
        x6 = x5
        y6 = y5 - 130 + thickness

        Radiusval = 20          
        lastNum = 6

        prev_x, prev_y = eval('x1'), eval('y1')
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')

            if i == 4:
                draw_arc(doc, (curr_x, curr_y - Radiusval), (curr_x - Radiusval, curr_y), Radiusval)
                prev_x, prev_y = curr_x - Radiusval, curr_y  

            elif i == 5:
                draw_arc(doc, (curr_x + Radiusval, curr_y), (curr_x, curr_y - Radiusval), Radiusval)
                prev_x, prev_y = curr_x, curr_y - Radiusval

            else:
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y
                
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")      

        #절곡라인  (히든선 : 정방향)      
        line(doc, x6, y6, x3, y3, layer='hidden')     
        #rear장공
        sum = 0
        for x in rearPanel[:-1] :
            sum = sum + x
            insert_block(x1 + sum, y1 + slotposition, "8x20 slot_Laser")
            dim(doc, x1 + sum, y1, x1 + sum - x, y1, 70, direction="down", localdimstyle="0.2 JKW")    
            dimcontinue(doc , x1 + sum, y2)
        dimcontinue(doc, x2, y2)

        circle_cross(doc, x1 + 150, y1 + 30 - thickness, 30, layer='레이져')
        circle_cross(doc, x2 - 150, y1 + 30 - thickness, 30, layer='레이져')

        # 상부 전체치수
        dim(doc, x4, y4, x5, y5, 150, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x4, y4, x4 - 150, y4 - 128, 100, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x5, y5, x5 + 150, y4 - 128, 100, direction="up", option='reverse', localdimstyle="0.2 JKW")
        # 우측 치수선
        dim(doc, x3, y3, x2, y2, 100, direction="right", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x4, y4)
        dim(doc, x2, y2, x4, y4, 150, direction="right", localdimstyle="0.2 JKW")
        
        text = "2-%%C30 Hole"
        dim_leader(doc, x1 + 150, y1 + 28, x1 + 150 + 50, y1 + 28 + 50, text, direction="leftToright", distance=17)       
        text = "%%C8x20 Slot"
        dim_leader(doc, x2 - 80, y1 + slotposition, x2 - 80 - 100, y1 + slotposition + 100, text, direction="rightToleft", distance=12)
        text = "2-R20"
        dim_leader(doc, x4 - 10, y4 - 10, x4 + 50, y4 + 50, text, direction="leftToright", distance=17)
                            
        # main Frame description    
        x1 = x1 + max(MCW,MCD) + 200
        y1 = y1 + 100
        textstr = f"Part Name : Ceiling Rear B/K"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : SPCC 2.3t"    
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')    
        textstr = f"Size : 156 x {MCW}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')         
    ###################################################
    # 3page side B/K
    ###################################################
        rx1 = frameXpos + 200
        ry1 = frameYpos + 1100 * frame_scale - 580
        insert_block(rx1, ry1 + 26, "Ceiling_Bracket_Section")    

        rx1 = frameXpos + 400
        
        thickness = 2
        vcut = thickness / 2

        x1 = rx1
        y1 = ry1
        x2 = x1 + CD - 12
        y2 = y1
        x3 = x2
        y3 = y2 + 30 - thickness
        x4 = x3
        y4 = y3 + 130 - thickness
        x5 = x4 - MCD + (25 - 140 + EE_raw + 200) + 31
        y5 = y4
        x6 = x5
        y6 = y5 - 100
        x7 = x6 - 400
        y7 = y6
        x8 = x7
        y8 = y7 + 100
        x9 = x4 - CD + 12
        y9 = y8
        x10 = x9
        y10 = y9 - 130 + thickness
        
        Radiusval = 20          
        lastNum = 10

        prev_x, prev_y = eval('x1'), eval('y1')
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')

            if i == 5:
                draw_arc(doc, (curr_x + Radiusval, curr_y), (curr_x, curr_y - Radiusval), Radiusval)
                prev_x, prev_y = curr_x, curr_y - Radiusval

            elif i == 8:
                draw_arc(doc, (curr_x, curr_y - Radiusval), (curr_x - Radiusval, curr_y), Radiusval)
                prev_x, prev_y = curr_x - Radiusval, curr_y
                
            elif i == 9:
                draw_arc(doc, (curr_x + Radiusval, curr_y), (curr_x, curr_y - Radiusval), Radiusval)
                prev_x, prev_y = curr_x, curr_y - Radiusval
                
            else:
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y
                
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")      

        #절곡라인  (히든선 : 정방향)      
        line(doc, x10, y10, x3, y3, layer='hidden')     
        #side장공
        sum = 0
        for x in sidePanel[:-1] :
            sum = sum + x
            insert_block(x1 + sum, y1 + slotposition, "8x20 slot_Laser")
            dim(doc, x1 + sum, y1, x1 + sum - x, y1, 70, direction="down", localdimstyle="0.2 JKW")    
            dimcontinue(doc , x1 + sum, y2)
        dimcontinue(doc, x2, y2)

        circle_cross(doc, x1 + 150, y1 + 30 - thickness, 30, layer='레이져')
        circle_cross(doc, x2 - 150, y1 + 30 - thickness, 30, layer='레이져')

        # 상부 전체치수
        dim(doc, x4, y4, x9, y9, 150, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x4, y4, x4 - 150, y4 - 128, 50, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x9, y9, x9 + 150, y4 - 128, 50, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x9, y9, x8, y8, 100, direction="up", localdimstyle="0.2 JKW")
        dimcontinue(doc, x5, y5)
        dimcontinue(doc, x4, y4)
        # 우측 치수선
        dim(doc, x3, y3, x2, y2, 100, direction="right", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x4, y4)
        dim(doc, x2, y2, x4, y4, 150, direction="right", localdimstyle="0.2 JKW")
        dim(doc, x7, y7, x8, y8, 50, direction="right", localdimstyle="0.2 JKW")
        
        text = "2-%%C30 Hole"
        dim_leader(doc, x1 + 150, y1 + 28, x1 + 150 + 50, y1 + 28 + 50, text, direction="leftToright", distance=17)       
        text = "%%C8x20 Slot"
        dim_leader(doc, x2 - 80, y1 + slotposition, x2 - 80 - 100, y1 + slotposition + 100, text, direction="rightToleft", distance=12)
        text = "3-R20"
        dim_leader(doc, x9 + 10, y9 - 10, x9 - 50, y4 + 50, text, direction="rightToleft", distance=17)
                            
        # main Frame description    
        x1 = x1 + max(MCW,MCD) + 200
        y1 = y1 + 100
        textstr = f"Part Name : Ceiling Side B/K"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : SPCC 2.3t"    
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')    
        textstr = f"Size : 156 x {CD-12}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : 본도 동일 - {su} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')    
        textstr = f"             본도 반대 - {su} EA"    
        draw_Text(doc, x1 , y1 - 240 , 30, str(textstr), layer='0')    

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 4page Ceiling  Bracket
    #################################################################################################
    program = 1
    if program:
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = BasicXscale + (CW - 1500) 
        TargetYscale = BasicYscale
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    
        insert_frame(frameXpos, frameYpos, frame_scale, "drawings_frame", workplace)   

    ###################################################
    # 4page header B/K
    ###################################################
        rx1 = frameXpos + 150
        ry1 = frameYpos + 1100 * frame_scale + 500
        
        if headertype == "세화정공":
            insert_block(rx1, ry1 + 67.5, "Ceiling_header_section_sehwa")    
            
            thickness = 2
            vcut = thickness / 2

            x1 = rx1 + 300
            y1 = ry1
            x2 = x1 + CW/2 - 181
            y2 = y1
            x3 = x2
            y3 = y2 + 13 - thickness
            x4 = x3 + 50
            y4 = y3
            x5 = x4
            y5 = y4 - 13 + thickness
            x6 = x5 + 200
            y6 = y5
            x7 = x6
            y7 = y6 + 13 - thickness
            x8 = x7 + 50
            y8 = y7
            x9 = x8
            y9 = y8 - 13 + thickness
            x10 = x9 + CW/2 - 131
            y10 = y9
            x11 = x10
            y11 = y10 + 13 - thickness
            x12 = x11
            y12 = y11 + 40 - thickness*2
            x13 = x12
            y13 = y12 + 50 - thickness*2
            x14 = x13
            y14 = y13 + 40 - thickness*2
            x15 = x14
            y15 = y14 + 13 - thickness
            x16 = x15 - (CW/2 - 131)
            y16 = y15
            x17 = x16
            y17 = y16- 13 + thickness
            x18 = x17 - 50
            y18 = y17
            x19 = x18
            y19 = y18 + 13 - thickness 
            x20 = x19 - 200
            y20 = y19
            x21 = x20
            y21 = y20 - 13 + thickness
            
            x22 = x21 - 50 
            y22 = y21
            x23 = x22
            y23 = y22 + 13 - thickness
            x24 = x23 - (CW/2 - 181)
            y24 = y23
            x25 = x24
            y25 = y24 - 13 + thickness 
            x26 = x25
            y26 = y25 - 40 + thickness*2
            x27 = x26
            y27 = y26 - 50 + thickness*2
            x28 = x27
            y28 = y27 - 40 + thickness*2

            prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
            lastNum = 28
            for i in range(1, lastNum + 1):
                curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y
            line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
            
            #절곡라인  (히든선 : 정방향)      
            line(doc, x11, y11, x28, y28, layer='hidden')
            line(doc, x12, y12, x27, y27, layer='hidden')     
            line(doc, x13, y13, x26, y26, layer='hidden')     
            line(doc, x14, y14, x25, y25, layer='hidden')     
            
            # 상부 전체치수
            dim(doc, x15, y15, x24, y24, 150, direction="up", localdimstyle="0.2 JKW")
            dim(doc, x16, y16, x15, y15, 100, direction="up", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x19, y19)
            dimcontinue(doc, x20, y20)
            dimcontinue(doc, x23, y23)
            # 우측 치수
            dim(doc, x15, y15, x10, y10, 260, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x11, y11, x10, y10, 80, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x12, y12, x11, y11, 130, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x13, y13, x12, y12, 180, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x13, y13, x14, y14, 130, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x14, y14, x15, y15, 80, direction="right", localdimstyle="0.2 JKW")
            
            if lctype == "035":
                lc_positions = calculate_lc_position_035(inputCW, inputCD)
                for i, x in enumerate(Bottomx):
                    insert_block(x1 - 6 + x, y1 + 70, "Ceiling_LChole_B.K_sehwa")
                    if i == 0:
                        dim(doc, x1 - 6 + x, y1 + 55, x1, y1, 190, direction="down", option='reverse', localdimstyle="0.2 JKW")    
                    else:
                        dimcontinue(doc, x1 - 6 + x , y1 + 50)
                dimcontinue(doc ,x10, y1) 
                # # 하부 치수
                # dim(doc, x1 + 134, y1 + 47, x1, y1, 100, direction="down", option='reverse', localdimstyle="0.2 JKW")
                # dimcontinue(doc, x1 + 134 + 160, y1 + 47)
                # dimcontinue(doc, x10 - 134 - 160, y1 + 47)
                # dimcontinue(doc, x10 - 134, y1 + 47)
                # dimcontinue(doc, x10, y1)

                # circle_cross(doc, x1 + 134, y1 + 70, 35, layer='레이져')
                # circle_cross(doc, x1 + 134 + 160, y1 + 70, 35, layer='레이져')
                # circle_cross(doc, x10 - 134, y1 + 70, 35, layer='레이져')
                # circle_cross(doc, x10 - 134 - 160, y1 + 70, 35, layer='레이져')
            
            # main Frame description    
            x1 = x1 + CW + 300
            y1 = y1 + 100
            textstr = f"Part Name : Ceiling Bracket"
            draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
            textstr = f"Mat.Spec : SPCC 2.3t"
            draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
            textstr = f"Size : 140 x {CW - 12}mm"
            draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
            textstr = f"Quantity : {su*2} EA"
            draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')

        elif headertype == "정일산업":

            insert_block(rx1, ry1 + 56.5, "Ceiling_header_section_joungil")    
            
            thickness = 2
            vcut = thickness / 2

            x1 = rx1 + 300
            y1 = ry1
            x2 = x1 + 50
            y2 = y1
            x3 = x2
            y3 = y2 - 20 + thickness
            x4 = x3 + CW - 112
            y4 = y3
            x5 = x4
            y5 = y4 + 20 - thickness
            x6 = x5 + 50
            y6 = y5
            x7 = x6 
            y7 = y6 + 45 - thickness*2
            x8 = x7
            y8 = y7 + 50 - thickness*2
            x9 = x8
            y9 = y8 + 45 - thickness*2
            x10 = x9 - 50
            y10 = y9
            x11 = x10
            y11 = y10 + 20 - thickness
            x12 = x11 - CW + 112
            y12 = y11
            x13 = x12
            y13 = y12 - 20 + thickness
            x14 = x13 - 50
            y14 = y13
            x15 = x14
            y15 = y14 - 45 + thickness*2
            x16 = x15
            y16 = y15 - 50 + thickness*2

            prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
            lastNum = 16
            for i in range(1, lastNum + 1):
                curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y
            line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
            
            #절곡라인  (히든선 : 정방향)      
            line(doc, x2, y2, x5, y5, layer='hidden')     
            line(doc, x16, y16, x7, y7, layer='hidden')     
            line(doc, x15, y15, x8, y8, layer='hidden')     
            line(doc, x13, y13, x10, y10, layer='hidden')     
            
            # 상부 전체치수
            dim(doc, x9, y9, x14, y14, 150, direction="up", localdimstyle="0.2 JKW")
            dim(doc, x12, y12, x14, y14, 80, direction="up", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x11, y11)
            dimcontinue(doc, x9, y9)
            # 우측 치수
            dim(doc, x11, y11, x4, y4, 260, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x9, y9, x11, y11, 80, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x8, y8, 160)
            dimcontinue(doc, x7, y7)
            dimcontinue(doc, x6, y6, 160)
            dimcontinue(doc, x4, y4)
            # 좌측 치수
            dim(doc, x12, y12, x14, y14, 160, direction="left", localdimstyle="0.2 JKW")
            dimcontinue(doc, x1, y1)
            dimcontinue(doc, x3, y3, 160, option='reverse')
            dim(doc, x1, y1, x1 + 125, y1 + 70, 50, direction="left", localdimstyle="0.2 JKW")
            dimcontinue(doc, x12, y12)
            # 하부 치수
            dim(doc, x1 + 139, y1 + 70, x1, y1, 160, direction="down", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x1 + 139 + 150, y1 + 70)
            dimcontinue(doc, x6 - 139 - 150, y1 + 70)
            dimcontinue(doc, x6 - 139, y1 + 70)
            dimcontinue(doc, x6, y1)
    
            circle_cross(doc, x1 + 139, y1 + 70, 35, layer='레이져')
            circle_cross(doc, x1 + 139 + 150, y1 + 70, 35, layer='레이져')
            circle_cross(doc, x6 - 139, y1 + 70, 35, layer='레이져')
            circle_cross(doc, x6 - 139 - 150, y1 + 70, 35, layer='레이져')
            
            text = "4-%%C35"
            dim_leader(doc, x1 + 139, y1 + 70, x1 + 139 + 50, y1 + 70 + 50, text, direction="leftToright", distance=13)   # distance 13은 글자높이임       

            # main Frame description    
            x1 = x1 + CW + 300
            y1 = y1 + 100
            textstr = f"Part Name : Ceiling Bracket"
            draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
            textstr = f"Mat.Spec : SPCC 2.3t"
            draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
            textstr = f"Size : 189 x {CW - 12}mm"
            draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
            textstr = f"Quantity : {su*2} EA"
            draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')

        elif headertype == "에이스코리아" or headertype == "이노에스템":

            insert_block(rx1, ry1 + 56.5, "Ceiling_header_section")    
            
            thickness = 2
            vcut = thickness / 2

            x1 = rx1 + 300
            y1 = ry1
            x2 = x1 + 50
            y2 = y1
            x3 = x2
            y3 = y2 - 15 + thickness
            x4 = x3 + CW - 112
            y4 = y3
            x5 = x4
            y5 = y4 + 15 - thickness
            x6 = x5 + 50
            y6 = y5
            x7 = x6 
            y7 = y6 + 35 - thickness*2
            x8 = x7
            y8 = y7 + 55 - thickness*2
            x9 = x8
            y9 = y8 + 35 - thickness*2
            x10 = x9 - 50
            y10 = y9
            x11 = x10
            y11 = y10 + 15 - thickness
            x12 = x11 - CW + 112
            y12 = y11
            x13 = x12
            y13 = y12 - 15 + thickness
            x14 = x13 - 50
            y14 = y13
            x15 = x14
            y15 = y14 - 35 + thickness*2
            x16 = x15
            y16 = y15 - 55 + thickness*2

            prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
            lastNum = 16
            for i in range(1, lastNum + 1):
                curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y
            line(doc, prev_x, prev_y, x1, y1, layer="레이져")      
            
            #절곡라인  (히든선 : 정방향)      
            line(doc, x2, y2, x5, y5, layer='hidden')     
            line(doc, x16, y16, x7, y7, layer='hidden')     
            line(doc, x15, y15, x8, y8, layer='hidden')     
            line(doc, x13, y13, x10, y10, layer='hidden')     
            
            # 상부 전체치수
            dim(doc, x9, y9, x14, y14, 150, direction="up", localdimstyle="0.2 JKW")
            dim(doc, x12, y12, x14, y14, 100, direction="up", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x11, y11)
            dimcontinue(doc, x9, y9)
            # 우측 치수
            dim(doc, x11, y11, x4, y4, 260, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x9, y9, x11, y11, 80, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x8, y8, 160)
            dimcontinue(doc, x7, y7)
            dimcontinue(doc, x6, y6, 160)
            dimcontinue(doc, x4, y4)
            # 좌측 치수
            dim(doc, x12, y12, x14, y14, 160, direction="left", localdimstyle="0.2 JKW")
            dimcontinue(doc, x1, y1)
            dimcontinue(doc, x3, y3, 160, option='reverse')
            
            # main Frame description    
            x1 = x1 + CW + 300
            y1 = y1 + 100
            textstr = f"Part Name : Ceiling Bracket"
            draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
            textstr = f"Mat.Spec : SPCC 2.3t"
            draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
            textstr = f"Size : 139 x {CW - 12}mm"
            draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
            textstr = f"Quantity : {su} EA"
            draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')

    ###################################################
    # 4page header B/K - 2
    ###################################################
        rx1 = frameXpos + 150
        ry1 = frameYpos + 1100 * frame_scale
        
        if headertype == "에이스코리아" or headertype == "이노에스템":

            insert_block(rx1, ry1, "Ceiling_header_section")    
            
            thickness = 2
            vcut = thickness / 2

            x1 = rx1 + 300
            y1 = ry1
            x2 = x1 + 50
            y2 = y1
            x3 = x2
            y3 = y2 - 15 + thickness
            x4 = x3 + CW - 112
            y4 = y3
            x5 = x4
            y5 = y4 + 15 - thickness
            x6 = x5 + 50
            y6 = y5
            x7 = x6 
            y7 = y6 + 35 - thickness*2
            x8 = x7
            y8 = y7 + 55 - thickness*2
            x9 = x8
            y9 = y8 + 35 - thickness*2
            x10 = x9 - 50
            y10 = y9
            x11 = x10
            y11 = y10 + 15 - thickness
            x12 = x11 - CW + 112
            y12 = y11
            x13 = x12
            y13 = y12 - 15 + thickness
            x14 = x13 - 50
            y14 = y13
            x15 = x14
            y15 = y14 - 35 + thickness*2
            x16 = x15
            y16 = y15 - 55 + thickness*2

            prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
            lastNum = 16
            for i in range(1, lastNum + 1):
                curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')

                # (x11, y11) → (x12, y12) 구간은 건너뛰기
                if i == 12:
                    prev_x, prev_y = curr_x, curr_y
                
                else:
                    line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                    prev_x, prev_y = curr_x, curr_y
                
            line(doc, prev_x, prev_y, x1, y1, layer="레이져")  
            
            lc_positions = calculate_lc_position_035(inputCW, inputCD)

            Topholex = lc_positions['Topholex']
            Topholey = lc_positions['Topholey']

            Bottomx = lc_positions['Bottomx']
            Bottomy = lc_positions['Bottomy']    
            
            for i, x in enumerate(Bottomx):
                insert_block(x1 - 6 + x, y1 + 66.5, "Ceiling_LChole_B.K") 
                if i == 0:
                    dim(doc, x1 - 6 + x, y1 + 30, x1, y1, 100, direction="down", option='reverse', localdimstyle="0.2 JKW")    
                else:
                    dimcontinue(doc, x1 + x - 6, y1 + 50)
            dimcontinue(doc, x6, y6)    
        
            for i, x in enumerate(Bottomx):
                if i != 0:
                    # 나머지 선: (x + 17.5, y12) -> (다음 x - 17.5, y12)
                    line(doc, x12 - 56 + Bottomx[i-1] + 17.5, y12, x12 - 56 + x - 17.5, y12, layer="레이져")
            line(doc, x12, y12, x12 + 36.5, y12, layer="레이져")
            line(doc, x11, y11, x11 - 36.5, y12, layer="레이져")
            
            #절곡라인  (히든선 : 정방향)      
            line(doc, x2, y2, x5, y5, layer='hidden')
            line(doc, x16, y16, x7, y7, layer='hidden')     
            line(doc, x15, y15, x8, y8, layer='hidden')     
            line(doc, x13, y13, x10, y10, layer='hidden')     
            
            # 상부 전체치수
            dim(doc, x9, y9, x14, y14, 150, direction="up", localdimstyle="0.2 JKW")
            dim(doc, x12, y12, x14, y14, 100, direction="up", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x11, y11)
            dimcontinue(doc, x9, y9)
            # 우측 치수
            dim(doc, x11, y11, x4, y4, 260, direction="right", localdimstyle="0.2 JKW")
            dim(doc, x9, y9, x11, y11, 80, direction="right", option='reverse', localdimstyle="0.2 JKW")
            dimcontinue(doc, x8, y8, 160)
            dimcontinue(doc, x7, y7)
            dimcontinue(doc, x6, y6, 160)
            dimcontinue(doc, x4, y4)
            # 좌측 치수
            dim(doc, x12, y12, x14, y14, 210, direction="left", localdimstyle="0.2 JKW")
            dimcontinue(doc, x1, y1)
            dimcontinue(doc, x3, y3, 210, option='reverse')
            dim(doc, x3, y3, x3 + 54, y3 + 79.5, 160, direction="left", localdimstyle="0.2 JKW")
            dim(doc, x12, y12, x3 + 54, y3 + 79.5, 160, direction="left", localdimstyle="0.2 JKW")

            # main Frame description    
            x1 = x1 + CW + 300
            y1 = y1 + 100
            textstr = f"Part Name : Ceiling Bracket"
            draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
            textstr = f"Mat.Spec : SPCC 2.3t"
            draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
            textstr = f"Size : 139 x {CW - 12}mm"
            draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
            textstr = f"Quantity : {su} EA"
            draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')

    ###################################################
    # 4page 좌우 가로 브라켓
    ###################################################
        rx1 = frameXpos + 150
        ry1 = frameYpos + 1100 * frame_scale -500
        insert_block(rx1, ry1, "Ceiling_Bracket_30x30_section")    
        
        x1 = rx1 + 150
        y1 = ry1
        x2 = x1 + mcsizex2
        y2 = y1
        x3 = x2
        y3 = y2 + 28
        x4 = x3
        y4 = y3 + 28
        x5 = x1
        y5 = y4
        x6 = x5
        y6 = y5 - 28
        
        prev_x, prev_y = x1, y1  # 첫 번째 점으로 초기화
        lastNum = 6
        for i in range(1, lastNum + 1):
            curr_x, curr_y = eval(f'x{i}'), eval(f'y{i}')
            line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
            prev_x, prev_y = curr_x, curr_y
        line(doc, prev_x, prev_y, x1, y1, layer="레이져")  
        
        #절곡라인  (히든선 : 정방향)      
        line(doc, x3, y3, x6, y6, layer='hidden')
        # 우측 치수
        dim(doc, x2, y2, x4, y4, 120, direction="right", localdimstyle="0.2 JKW")
        dim(doc, x3, y3, x2, y2, 60, direction="right", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x4, y4, 60)
        # 하부 치수
        dim(doc, x2, y2, x1, y1, 120, direction="down", localdimstyle="0.2 JKW")
        dim(doc, x1, y1, x1 + mcsizex2/2, y1, 60, direction="down", localdimstyle="0.2 JKW")
        dimcontinue(doc, x2, y2, 60)
        
        circle_cross(doc, x1 + mcsizex2/2, y3, 30, layer='레이져')
        text = "%%C30"
        dim_leader(doc, x1 + mcsizex2/2, y3, x1 + mcsizex2/2 + 50, y3 + 70, text, direction="leftToright", distance=13)   # distance 13은 글자높이임       
        
        # main Frame description    
        x1 = x1 + mcsizex2 + 200
        y1 = y1 + 50
        textstr = f"Part Name : Ceiling Bracket"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
        textstr = f"Mat.Spec : SPCC 2.3t"    
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
        textstr = f"Size : 56 x {mcsizex2:.1f}mm"
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
        textstr = f"Quantity : {su} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')

    ###################################################
    # 4page 서스테이너 브라켓
    ###################################################
        rx1 = x1 + mcsizex2 + 500
        ry1 = ry1 - 50
        insert_block(rx1, ry1, "Ceiling_Bracket_sustainer")    
        insert_block(rx1, ry1, "Ceiling_Bracket_sustainer_L")    
        
        # main Frame description    
        x1 = x1 + mcsizex2 + 1100
        textstr = f"Part Name : Ceiling Bracket"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : SPCC 2.3t"    
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')    
        textstr = f"Size : 96 x 389mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su*2} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')    

        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 5page Ceiling  Bracket
    #################################################################################################
    program = 1
    if program:         
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = BasicXscale + CW - 2000
        TargetYscale = 1665 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)
        
    ###################################################
    # 5page 하부 가로 긴 브라켓
    ###################################################
        rx1 = frameXpos + 200
        ry1 = frameYpos + 1100 * frame_scale + 400
        
        x1 = rx1 + 300
        y1 = ry1
        x2 = x1 + mcsizex1
        y2 = y1
        x3 = x2
        y3 = y2 + 181.2
        x4 = x1
        y4 = y3
        
        line(doc, x1, y1, x2, y2, layer='레이져')
        line(doc, x3, y3, x4, y4, layer='레이져')
        
        insert_block(x1, y1, "Ceiling_Bracket_horizontal_L")
        insert_block(x1, y1, "Ceiling_Bracket_horizontal_L_")
        insert_block(x2, y2, "Ceiling_Bracket_horizontal_R")
        insert_block(x2, y2, "Ceiling_Bracket_horizontal_R_")
        insert_block(x1 + (CW - 12)/3, y1 + 90.6, "Ceiling_Bracket_horizontal_hole4_laser")
        insert_block(x1 + (CW - 12)*2/3, y1 + 90.6, "Ceiling_Bracket_horizontal_hole4_laser")
        
        circle_cross(doc, x1 + 30, y1 + 9, 3.2, layer='레이져')
        circle_cross(doc, x1 + 30, y4 - 9, 3.2, layer='레이져')
        circle_cross(doc, x2 - 30, y1 + 9, 3.2, layer='레이져')
        circle_cross(doc, x2 - 30, y4 - 9, 3.2, layer='레이져')
        text = "12-%%C3.2"
        dim_leader(doc, x1 + 30, y1 + 9, x1 + 30 + 50, y1 + 9 - 50, text, direction="leftToright", distance=13)   # distance 13은 글자높이임       
        dim(doc, x2 - 30, y1 + 9, x2, y1, 50, direction="left", localdimstyle="0.2 JKW")

        rectangle(doc, x1 - 31 + (MCW/3 - 110), y1 + 26.5, x1 - 31 + (MCW/3 - 20), y1 + 43.5, layer='레이져')
        rectangle(doc, x1 - 31 + (MCW/3 - 110), y4 - 26.5, x1 - 31 + (MCW/3 - 20), y4 - 43.5, layer='레이져')
        rectangle(doc, x1 - 31 + (MCW*2/3 + 110), y1 + 26.5, x1 - 31 + (MCW*2/3 + 20), y1 + 43.5, layer='레이져')
        rectangle(doc, x1 - 31 + (MCW*2/3 + 110), y4 - 26.5, x1 - 31 + (MCW*2/3 + 20), y4 - 43.5, layer='레이져')

        #절곡라인  (히든선 : 정방향)      
        line(doc, x1, y1 + 16.5, x2, y1 + 16.5, layer='2')     
        line(doc, x1, y4 - 16.5, x2, y4 - 16.5, layer='2')     
        line(doc, x1, y1 + 43.5, x2, y1 + 43.5, layer='hidden')     
        line(doc, x1, y4 - 43.5, x2, y4 - 43.5, layer='hidden')     
        
        # 상부 전체치수
        dim(doc, x4 + 194, y4 - 65, x4, y4, 110, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x2 - 194, y4 - 65)
        dimcontinue(doc, x3, y3)
        dim(doc, x4, y4, x4 + 30, y4, 100, direction="up", localdimstyle="0.2 JKW")
        dimcontinue(doc, x4 + (CW - 12)/3 - 30, y4)
        dimcontinue(doc, x4 + (CW - 12)/3 + 30, y4)
        dimcontinue(doc, x4 + (CW - 12)*2/3 - 30, y4)
        dimcontinue(doc, x4 + (CW - 12)*2/3 + 30, y4)
        dimcontinue(doc, x3 - 30, y4)
        dimcontinue(doc, x3, y4)
        # 하부 전체치수
        dim(doc, x2, y2, x1, y1, 130, direction="down", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x1 - 31 + (MCW/3 - 110), y1 + 26.5, x1, y1, 100, direction="down", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 - 31 + (MCW/3 - 110) + 90, y1 + 26.5)
        dimcontinue(doc, x1 - 31 + (MCW*2/3 + 20), y1 + 26.5)
        dimcontinue(doc, x1 - 31 + (MCW*2/3 + 20) + 90, y1 + 26.5)
        dimcontinue(doc, x2, y2)
        
        # main Frame description    
        x1 = x1 + CW - 500
        y1 = y1 - 200
        textstr = f"Part Name : Ceiling Bracket"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : SPCC 1.6T"    
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')    
        textstr = f"Size : 181.2 x {CW - 12}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')    
        
    ###################################################
    # 5page cover plate
    ###################################################
        rx1 = frameXpos + 500
        ry1 = frameYpos + 1100 * frame_scale - 400
        
        coverplatex1 = (CW - 12)/3 - 62 + 58
        
        x1 = rx1
        y1 = ry1
        x2 = x1 + coverplatex1
        y2 = y1
        x3 = x2
        y3 = y2 + 130
        x4 = x1
        y4 = y3
        
        insert_block(rx1, ry1, "Ceiling_Coverplate")
        insert_block(rx1, ry1, "Ceiling_Coverplate_laser")
        insert_block(x2, y2, "Ceiling_CoverplateR")
        insert_block(x2, y2, "Ceiling_Coverplate_laserR")
        
        line(doc, x1, y1, x2, y2, layer='레이져')
        line(doc, x3, y3, x4, y4, layer='레이져')
        
        # 상부 치수
        dim(doc, x1 + 28, y4, x2 - 28, y4, 80, direction="up", localdimstyle="0.2 JKW")
        dim(doc, x1, y4, x2, y3, 140, direction="up", localdimstyle="0.2 JKW")
        
        # main Frame description    
        x1 = x1 + coverplatex1 + 400
        y1 = y1 + 100
        textstr = f"Part Name : Cover Plate"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
        textstr = f"Mat.Spec : SPCC 1.6t"    
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
        textstr = f"Size : 130 x {coverplatex1:.1f}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
        textstr = f"Quantity : {su*3} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')
        
        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 6page Ceiling  Bracket
    #################################################################################################
    program = 1
    if program:         
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = BasicXscale + CW - 2000
        TargetYscale = 1665 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)
        
    ###################################################
    # 6page 좌우 세로 긴 브라켓
    ###################################################
        rx1 = frameXpos + 200
        ry1 = frameYpos + 1100 * frame_scale + 400
        
        x1 = rx1 + 300
        y1 = ry1
        x2 = x1 + mcsizex3
        y2 = y1
        x3 = x2
        y3 = y2 + 181.2
        x4 = x1
        y4 = y3
        
        line(doc, x1, y1, x2, y2, layer='레이져')
        line(doc, x3, y3, x2, y2, layer='레이져')
        line(doc, x3, y3, x4, y4, layer='레이져')
        line(doc, x1, y1, x4, y4, layer='레이져')
        
        insert_block(x1, y1, "Ceiling_Bracket_horizontal_L2_")
        insert_block(x2, y2, "Ceiling_Bracket_horizontal_R_2")
        insert_block(x1 + (MCD - (ELYpos + 96.5))/3, y1 + 90.6, "Ceiling_Bracket_horizontal_hole4_laser")
        insert_block(x1 + (MCD - (ELYpos + 96.5))*2/3, y1 + 90.6, "Ceiling_Bracket_horizontal_hole4_laser")
        
        circle_cross(doc, x1 + 30, y1 + 9, 3.2, layer='레이져')
        circle_cross(doc, x1 + 30, y4 - 9, 3.2, layer='레이져')
        circle_cross(doc, x2 - 30, y1 + 9, 3.2, layer='레이져')
        circle_cross(doc, x2 - 30, y4 - 9, 3.2, layer='레이져')
        text = "12-%%C3.2"
        dim_leader(doc, x1 + 30, y1 + 9, x1 + 30 + 50, y1 + 9 + 50, text, direction="leftToright", distance=13)   # distance 13은 글자높이임       

        rectangle(doc, x1 + MDFirstSpace + 20, y1 + 26.5, x1 + MDFirstSpace + 110, y1 + 43.5, layer='레이져')#좌
        rectangle(doc, x1 + MDFirstSpace + 20, y4 - 26.5, x1 + MDFirstSpace + 110, y4 - 43.5, layer='레이져')
        rectangle(doc, x1 + mcsizex3/2 - 45, y1 + 26.5, x1 + mcsizex3/2 + 45, y1 + 43.5, layer='레이져')#중
        rectangle(doc, x1 + mcsizex3/2 - 45, y4 - 26.5, x1 + mcsizex3/2 + 45, y4 - 43.5, layer='레이져')
        rectangle(doc, x2 - MDFirstSpace - 20, y1 + 26.5, x2 - MDFirstSpace - 110, y1 + 43.5, layer='레이져')#우
        rectangle(doc, x2 - MDFirstSpace - 20, y4 - 26.5, x2 - MDFirstSpace - 110, y4 - 43.5, layer='레이져')

        #절곡라인  (히든선 : 정방향)      
        line(doc, x1, y1 + 16.5, x2, y1 + 16.5, layer='2')     
        line(doc, x1, y4 - 16.5, x2, y4 - 16.5, layer='2')     
        line(doc, x1, y1 + 43.5, x2, y1 + 43.5, layer='hidden')     
        line(doc, x1, y4 - 43.5, x2, y4 - 43.5, layer='hidden')     
        
        # 상부 전체치수
        dim(doc, x4, y4, x4 + 30, y4, 50, direction="up", localdimstyle="0.2 JKW")
        dimcontinue(doc, x4 + (MCD - (ELYpos + 96.5))/3 - 30, y4)
        dimcontinue(doc, x4 + (MCD - (ELYpos + 96.5))/3 + 30, y4)
        dimcontinue(doc, x4 + (MCD - (ELYpos + 96.5))*2/3 - 30, y4)
        dimcontinue(doc, x4 + (MCD - (ELYpos + 96.5))*2/3 + 30, y4)
        dimcontinue(doc, x3 - 30, y4)
        dimcontinue(doc, x3, y4)
        # 하부 전체치수
        dim(doc, x2, y2, x1, y1, 130, direction="down", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x1 + MDFirstSpace + 20, y1 + 26.5, x1, y1, 100, direction="down", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 + MDFirstSpace + 20 + 90, y1 + 26.5)
        dimcontinue(doc, x1 + mcsizex3/2 - 45, y1 + 26.5)
        dimcontinue(doc, x1 + mcsizex3/2 + 45, y1 + 26.5)
        dimcontinue(doc, x2 - MDFirstSpace - 20 - 90, y1 + 26.5)
        dimcontinue(doc, x2 - MDFirstSpace - 20, y1 + 26.5)
        dimcontinue(doc, x2, y2)
        
        # main Frame description    
        x1 = x1 + CW - 500
        y1 = y1 - 200
        textstr = f"Part Name : Ceiling Bracket"
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
        textstr = f"Mat.Spec : SPCC 1.6T"
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
        textstr = f"Size : 181.2 x {mcsizex3:.1f}mm"
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
        textstr = f"Quantity : {su} EA"
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')
        
    ###################################################
    # 6page cover plate
    ###################################################
        rx1 = frameXpos + 500
        ry1 = frameYpos + 1100 * frame_scale - 400
        
        coverplatex2 = (MCD - (ELYpos + 96.5))/3 - 62 + 58
        
        x1 = rx1
        y1 = ry1
        x2 = x1 + coverplatex2
        y2 = y1
        x3 = x2
        y3 = y2 + 130
        x4 = x1
        y4 = y3
        
        insert_block(rx1, ry1, "Ceiling_Coverplate")
        insert_block(rx1, ry1, "Ceiling_Coverplate_laser")
        insert_block(x2, y2, "Ceiling_CoverplateR")
        insert_block(x2, y2, "Ceiling_Coverplate_laserR")
        
        line(doc, x1, y1, x2, y2, layer='레이져')
        line(doc, x3, y3, x4, y4, layer='레이져')
        
        # 상부 치수
        dim(doc, x1 + 28, y4, x2 - 28, y4, 80, direction="up", localdimstyle="0.2 JKW")
        dim(doc, x1, y4, x2, y3, 140, direction="up", localdimstyle="0.2 JKW")
        
        # main Frame description    
        x1 = x1 + coverplatex2 + 400
        y1 = y1 + 100
        textstr = f"Part Name : Cover Plate"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
        textstr = f"Mat.Spec : SPCC 1.6t"    
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
        textstr = f"Size : 130 x {coverplatex2:.1f}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
        textstr = f"Quantity : {su*3} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')
        
        frameXpos = frameXpos + TargetXscale + 400

    #################################################################################################
    # 7page Ceiling  Bracket
    #################################################################################################
    program = 1
    if program:         
        # 도면틀 넣기
        BasicXscale = 2813
        BasicYscale = 1765
        TargetXscale = BasicXscale + CW - 2000
        TargetYscale = 1665 
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale    

        insert_frame( frameXpos , frameYpos , frame_scale, "drawings_frame", workplace)
        
    ###################################################
    # 7page 중간 가로 브라켓
    ###################################################
        rx1 = frameXpos + 200
        ry1 = frameYpos + 1100 * frame_scale + 400
        
        x1 = rx1 + 300
        y1 = ry1
        x2 = x1 + MCW/3 - 1
        y2 = y1
        x3 = x2
        y3 = y2 + 181.2
        x4 = x1
        y4 = y3
        
        line(doc, x1, y1, x2, y2, layer='레이져')
        line(doc, x3, y3, x2, y2, layer='레이져')
        line(doc, x3, y3, x4, y4, layer='레이져')
        line(doc, x1, y1, x4, y4, layer='레이져')
        
        insert_block(x1, y1, "Ceiling_Bracket_horizontal_L2_")
        insert_block(x2, y2, "Ceiling_Bracket_horizontal_R_2")
        insert_block(x1 + (MCW/3 - 1)/2, y1 + 90.6, "Ceiling_Bracket_horizontal_hole4_laser")

        circle_cross(doc, x1 + 30, y1 + 9, 3.2, layer='레이져')
        circle_cross(doc, x1 + 30, y4 - 9, 3.2, layer='레이져')
        circle_cross(doc, x2 - 30, y1 + 9, 3.2, layer='레이져')
        circle_cross(doc, x2 - 30, y4 - 9, 3.2, layer='레이져')

        text = "8-%%C3.2"
        dim_leader(doc, x1 + 30, y1 + 9, x1 + 30 + 50, y1 + 9 + 50, text, direction="leftToright", distance=13)
        
        rectangle(doc, x1 + (MCW/3 - 1)/2 - 45, y1 + 26.5, x1 + (MCW/3 - 1)/2 + 45, y1 + 43.5, layer='레이져')
        rectangle(doc, x1 + (MCW/3 - 1)/2 - 45, y4 - 26.5, x1 + (MCW/3 - 1)/2 + 45, y4 - 43.5, layer='레이져')
        #절곡라인  (히든선 : 정방향)      
        line(doc, x1, y1 + 16.5, x2, y1 + 16.5, layer='2')     
        line(doc, x1, y4 - 16.5, x2, y4 - 16.5, layer='2')     
        line(doc, x1, y1 + 43.5, x2, y1 + 43.5, layer='hidden')     
        line(doc, x1, y4 - 43.5, x2, y4 - 43.5, layer='hidden')     
        
        # 상부 전체치수
        dim(doc, x3, y3, x4, y4, 150, direction="up", option='reverse', localdimstyle="0.2 JKW")
        dim(doc, x4, y4, x4 + 30, y4, 100, direction="up", localdimstyle="0.2 JKW")
        dimcontinue(doc, x4 + (MCW/3 - 1)/2 - 30, y4)
        dimcontinue(doc, x4 + (MCW/3 - 1)/2 + 30, y4)
        dimcontinue(doc, x3 - 30, y4)
        dimcontinue(doc, x3, y4)
        # 하부 전체치수
        dim(doc, x1 + (MCW/3 - 1)/2 - 45, y1 + 26.5, x1, y1, 100, direction="down", option='reverse', localdimstyle="0.2 JKW")
        dimcontinue(doc, x1 + (MCW/3 - 1)/2 + 45, y1 + 26.5)
        dimcontinue(doc, x2, y2)
        
        # main Frame description    
        x1 = x1 + CW - 500
        y1 = y1 + 100
        textstr = f"Part Name : Ceiling Bracket"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')    
        textstr = f"Mat.Spec : SPCC 1.6T"    
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')    
        textstr = f"Size : 181.2 x {(MCW/3 - 1):.1f}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')        
        textstr = f"Quantity : {su*3} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')    
        
    ###################################################
    # 7page cover plate
    ###################################################
        rx1 = frameXpos + 500
        ry1 = frameYpos + 1100 * frame_scale - 400
        
        coverplatex3 = (MCW/3 - 1)/2 - 62 + 58
        
        x1 = rx1
        y1 = ry1
        x2 = x1 + coverplatex3
        y2 = y1
        x3 = x2
        y3 = y2 + 130
        x4 = x1
        y4 = y3
        
        insert_block(rx1, ry1, "Ceiling_Coverplate")
        insert_block(rx1, ry1, "Ceiling_Coverplate_laser")
        insert_block(x2, y2, "Ceiling_CoverplateR")
        insert_block(x2, y2, "Ceiling_Coverplate_laserR")
        
        line(doc, x1, y1, x2, y2, layer='레이져')
        line(doc, x3, y3, x4, y4, layer='레이져')
        
        # 상부 치수
        dim(doc, x1 + 28, y4, x2 - 28, y4, 80, direction="up", localdimstyle="0.2 JKW")
        dim(doc, x1, y4, x2, y3, 140, direction="up", localdimstyle="0.2 JKW")
        
        # main Frame description    
        x1 = x1 + coverplatex1 + 400
        y1 = y1 + 100
        textstr = f"Part Name : Cover Plate"    
        draw_Text(doc, x1 , y1 , 30 , str(textstr), layer='0')
        textstr = f"Mat.Spec : SPCC 1.6t"    
        draw_Text(doc, x1 , y1 - 60 , 30, str(textstr), layer='0')
        textstr = f"Size : 130 x {coverplatex3:.1f}mm"    
        draw_Text(doc, x1 , y1 - 120 , 30, str(textstr), layer='0')
        textstr = f"Quantity : {su*6} EA"    
        draw_Text(doc, x1 , y1 - 180 , 30, str(textstr), layer='0')
        
        frameXpos = frameXpos + TargetXscale + 400     
   

# @Gooey(program_name='Jamb cladding 자동작도 프로그램 ver01', tabbed_groups=True, navigation='Tabbed')
@Gooey(encoding='utf-8', program_name='미래기업 천장 & LIGHT CASE 자동작도', tabbed_groups=True, navigation='Tabbed', show_success_modal=False,  default_size=(800, 600))
def main():
    global args  #수정할때는 global 선언이 필요하다. 단순히 읽기만 할때는 필요없다.    
    global start_time
    global workplace, drawdate, deadlinedate, lctype, secondord, text_style, exit_program, Vcut_rate, Vcut, program_message, formatted_date, text_style_name
    global CW_raw, CD_raw, panel_thickness, LC_height, su, LCframeMaterial, LCplateMaterial
    global input_CD, input_CW,CD,CW,drawdate_str , deadlinedate_str, drawdate_str_short, deadlinedate_str_short
    global doc, msp , T5_is, text_style
    global frontPanel, rearPanel, sidePanel, slotposition , AS_raw, EE_raw, ELYpos, OP_raw, ceilingdivision
    global headertype

    # 현재 날짜와 시간을 가져옵니다.
    current_datetime = datetime.now()

    # 2023-12-10 원하는 형식으로 날짜를 문자열로 변환합니다.
    formatted_date = current_datetime.strftime('%Y-%m-%d')

    # 현재 날짜와 시간을 '년월일_시분초' 형식으로 가져옴
    # current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    current_time = datetime.now().strftime("%H%M%S")

    # 찾은 .xlsm 파일 목록 순회
    for file_path in xlsm_files:

        # 엑셀 파일 열기
        # file_path = xlsm_files[0]
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        sheet_name = '발주(수정)'  # 원하는 시트명으로 변경 기본정보가 있는 시트
        sheet = workbook[sheet_name]

        doc = ezdxf.readfile(os.path.join(application_path, '', 'block.dxf'))
        msp = doc.modelspace()        

        # 5, 6, 7, 8, 9행의 2열 값을 직접 가져오기 엑셀시트의 행열에 대한 개념이 있어야 한다.
        workplace = sheet.cell(row=3, column=4).value
        drawdate = sheet.cell(row=3, column=16).value
        deadlinedate = sheet.cell(row=3, column=19).value # 납기일 추가
        lctype = sheet.cell(row=8, column=22).value
        secondord = sheet.cell(row=2, column=4).value
        CW_raw = sheet.cell(row=16, column=1).value
        CD_raw = sheet.cell(row=16, column=7).value
        panel_thickness = sheet.cell(row=7, column=10).value
        LC_height = sheet.cell(row=7, column=25).value
        su = sheet.cell(row=6, column=10).value 
        LCframeMaterial = sheet.cell(row=11, column=16).value
        LCplateMaterial = sheet.cell(row=11, column=22).value
        input_CD = sheet.cell(row=12, column=18).value
        input_CW = sheet.cell(row=12, column=22).value
        T5_is = sheet.cell(row=12, column=25).value ####확인
        entrytype = sheet.cell(row=2, column=23).value
        doortype = sheet.cell(row=6, column=4).value
        capacity = sheet.cell(row=6, column=7).value
        OP_raw = sheet.cell(row=6, column=13).value
        EE_raw = sheet.cell(row=6, column=16).value
        AS_raw = sheet.cell(row=6, column=19).value
        Sehwa_Cutfromcenter = sheet.cell(row=6, column=22).value
        baseMaterial = sheet.cell(row=7, column=4).value
        LHtype = sheet.cell(row=7, column=13).value
        emergencytype = sheet.cell(row=7, column=16).value
        emergencyRow = sheet.cell(row=7, column=19).value
        emergencyCol = sheet.cell(row=7, column=22).value
        slotposition = sheet.cell(row=8, column=4).value
        sillspace = sheet.cell(row=8, column=10).value
        headertype = sheet.cell(row=8, column=16).value
        fan_su = sheet.cell(row=12, column=1).value
        fantop = sheet.cell(row=12, column=4).value
        fanside = sheet.cell(row=12, column=7).value
        ceilingdivision = sheet.cell(row=16, column=10).value
        ELYpos = sheet.cell(row=16, column=12).value
        
        # 19행, 4열부터 22개의 값 가져와서 리스트로 저장
        frontPanel = [
            sheet.cell(row=19, column=col).value 
            for col in range(4, 26)  # 4열부터 25열까지 (총 22개)
            if sheet.cell(row=19, column=col).value not in (0, None)
        ]
        rearPanel = [
            sheet.cell(row=20, column=col).value 
            for col in range(4, 26)  # 4열부터 25열까지 (총 22개)
            if sheet.cell(row=20, column=col).value not in (0, None)
        ]
        sidePanel = [
            sheet.cell(row=21, column=col).value 
            for col in range(4, 26)  # 4열부터 25열까지 (총 22개)
            if sheet.cell(row=21, column=col).value not in (0, None)
        ]


        # LCframeMaterial이 프레임 재질 공백이나 None일 경우 'SPCC 1.2T'를 할당
        if not LCframeMaterial or LCframeMaterial.isspace():
            if lctype == '026':
                LCframeMaterial = 'SPCC 1.6T'
            else:
                LCframeMaterial = 'SPCC 1.2T'        

        # LCplateMaterial이 중판 공백이나 None일 경우 'SPCC 1.2T'를 할당
        if not LCplateMaterial or LCplateMaterial.isspace():
            LCplateMaterial = 'SPCC 1.2T'        

        # LC_height가 None이거나 공백 문자열이면 150을 할당
        if LC_height is None or (isinstance(LC_height, str) and LC_height.isspace()):
            LC_height = 150

        # 작도일자 추출
        drawdate_str = "{0}년 {1:02d}월 {2:02d}일".format(drawdate.year, drawdate.month, drawdate.day)
        drawdate_str_short = "{0:02d}/{1:02d}".format(drawdate.month, drawdate.day)

        try:
            deadlinedate_str = "{0}년 {1:02d}월 {2:02d}일".format(deadlinedate.year, deadlinedate.month, deadlinedate.day)
            deadlinedate_str_short = "{0:02d}/{1:02d}".format(deadlinedate.month, deadlinedate.day)
        except AttributeError:
            print("deadlinedate is not a valid datetime object")

        if input_CD not in [None, '', 0] and input_CW not in [None, '', 0]:
            # input_CD와 input_CW가 모두 공백이나 0이 아닌 경우 실행할 코드 강제입력에 대한 처리
            CW = input_CD
            CD = input_CW
        else:    
            CW = CW_raw - panel_thickness * 2
            CD = CD_raw - panel_thickness * 2

        # 파일 이름에 사용할 수 없는 문자 정의
        invalid_chars = '<>:"/\\|?*'
        # 정규식을 사용하여 유효하지 않은 문자 제거
        cleaned_file_name = re.sub(f'[{re.escape(invalid_chars)}]', '', f"{workplace}_{lctype}_CW{str(CW)}xCD{str(CD)}_{current_time}")
        # 결과 파일 이름
        file_name = f"{cleaned_file_name}.dxf"

        exit_program = False

        # 1.2T vcut 기준자료
        Bending_rate = 1.2
        Vcut = True
        Vcut_rate = Bending_rate/2

        program_message = \
            '''
        프로그램 실행결과입니다.
        -------------------------------------
        {0}
        -------------------------------------
        이용해 주셔서 감사합니다.
        '''    

        args = parse_arguments()    
        # 숫자로 변수를 선언해 준다.
        
        # 프로그램 시작 시간 기록
        start_time = time.time()

        # 항상 본천장 작도함
        MainCeiling() 

        if not args.config:
            if args.opt1:
                if(lctype=='026'):
                    lc026()            
                elif(lctype=='N20'):
                    lcN20()            
                elif(lctype=='032'):
                    lc032()            
                elif(lctype=='031'):
                    lc031()            
                elif(lctype=='035'):
                    lc035()            
                elif(lctype=='008_A'):
                    lc008_A()                

        # display_message()    
        end_time = time.time()

        # 실행 시간 계산
        execution_time = end_time - start_time

        # 시간 형식으로 변환
        hours, remainder = divmod(execution_time, 3600)
        minutes, seconds = divmod(remainder, 60)

        # 결과를 포맷팅하는 부분 수정
        parts = []
        if hours:
            parts.append(f"{int(hours)}h")
        if minutes:
            parts.append(f"{int(minutes)}m")
        parts.append(f"{int(seconds) + 3 }초")  # 초를 '초'로 표시

        formatted_execution_time = " ".join(parts)

        log_login()
        print(f"실행 시간: {formatted_execution_time}")

        # 생성된 파일 이름으로 DXF 파일 저장
        # file_name =f"c:\\python\\{file_name}" 
        file_name =f"c:\\python\\mirae_ceiling_ver01\\{file_name}" 
        doc.saveas(file_name)
        print(f"파일이 '{file_name}'로 저장 완료!")              

if __name__ == '__main__':
    main()
    sys.exit()