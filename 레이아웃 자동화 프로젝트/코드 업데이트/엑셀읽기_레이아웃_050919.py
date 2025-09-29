## 레이아웃 개발 시작(2025/08/28)


import math
import ezdxf
from ezdxf.enums import TextEntityAlignment
from datetime import datetime
import openpyxl
import os
import glob
import time
import os
import sys
import io
from datetime import datetime

import json
from gooey import Gooey, GooeyParser
import warnings
import re
import requests
        
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 경고 메시지 필터링
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")

# 전역 변수 초기화
saved_DimXpos = 0
saved_DimYpos = 0
saved_Xpos = 0
saved_Ypos = 0
saved_direction = "up"
saved_text_height = 0.22
saved_text_gap = 0.05
dimdistance = 0
dim_horizontalbase = 0
dim_verticalbase = 0
distanceXpos = 0
distanceYpos = 0

start_time = 0

workplace = ""
drawdate = None
deadlinedate = None
lctype = ""
secondord = ""
CW_raw = 0
CD_raw = 0
panel_thickness = 0
carpanel_thickness = 0
LC_height = 0
su = 0
LCframeMaterial = ""
LCplateMaterial = ""
input_CD = 0
input_CW = 0
CD = 0
CW = 0
drawdate_str = ""
deadlinedate_str = ""
T5_is =""
frontPanel, rearPanel, sidePanel = None, None, None
slotposition = 0
EE_raw = 0
AS_raw = 0
ceilingdivision = 0
SplitYpos1, SplitYpos2, SplitYpos3 = 0, 0, 0
ELYpos = 0
footOP = 0
headertype, emergencytype = None, None
emergencyRow, emergencyCol = 0, 0

CW, CD, CH, HWW, HWD, HH = 0,0,0,0,0,0
wall_thickness = 0
OP = 0
CWW = 0
CHW = 0

# 폴더 내의 모든 .xlsm 파일을 검색
application_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
excel_saved_file = os.path.join(application_path, 'layout_excel')
xlsm_files = glob.glob(os.path.join(excel_saved_file, '*.xlsx'))
jamb_ini = os.path.join(application_path, 'data', 'jamb.json')
license_file_path = os.path.join(application_path, 'data', 'hdsettings.json') # 하드디스크 고유번호 인식

# DXF 파일 로드
doc = ezdxf.readfile(os.path.join(application_path, '', 'layout.dxf'))
msp = doc.modelspace()

# TEXTSTYLE 정의
text_style_name = 'JKW'  # 원하는 텍스트 스타일 이름
if text_style_name not in doc.styles:
    text_style = doc.styles.new(
        name=text_style_name,
        dxfattribs={
            'font': 'Arial.ttf', # TrueType 글꼴 파일명            
        }
    )
else:
    text_style = doc.styles.get(text_style_name)    

# 찾은 .xlsm 파일 목록 출력
# for xlsm_file in xlsm_files:
#     print(xlsm_file)

def log_login():
    # PHP 파일의 URL 서버에서는 아이피를 저장한다. 업체 아이피를 기록한다.
    url = f"http://8440.co.kr/autopanel/savelog.php?company=미래기업&content={workplace}_{lctype}_{su}EA"

    # HTTP 요청 보내기
    response = requests.get(url)
    
    # 요청이 성공했는지 확인
    if response.status_code == 200:
        # print("logged successfully.")
        print(response.json())
    else:
        print("Failed to log login time.")
        print(response.text)   
        exit(1)

def parse_arguments_settings():
    parser = GooeyParser()
    settings = parser.add_argument_group('설정')
    settings.add_argument('--config', action='store_true', default=True, help='라이센스')
    settings.add_argument('--password', widget='PasswordField', gooey_options={'visible': True, 'show_label': True}, help='비밀번호')
    return parser.parse_args()

def parse_arguments():
    parser = GooeyParser()

    group1 = parser.add_argument_group('라이트케이스')
    # group1.add_argument('--opt1', action='store_true', help='기본')
    # group1.add_argument('--opt2', action='store_true', help='도장홀')
    # group1.add_argument('--opt3', action='store_true', help='쪽쟘 상판끝 라운드(추영덕소장)')    
    group1.add_argument('--opt1', action='store_true', default=True, help='기본')

    settings = parser.add_argument_group('설정')
    settings.add_argument('--config', action='store_true', help='라이센스')
    settings.add_argument('--password', widget='PasswordField', gooey_options={'visible': True, 'show_label': True}, help='비밀번호')

    return parser.parse_args()
def display_message():
    message = program_message.format('\n'.join(sys.argv[1:])).split('\n')
    delay = 1.5 / len(message)

    for line in message:
        print(line)
        time.sleep(delay)
# 환경설정 가져오기(하드공유 번호)
def load_env_settings():
    try:
        with open(license_file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return data.get("DiskID")
    except FileNotFoundError:
        return None
def get_current_disk_id():
    return os.popen('wmic diskdrive get serialnumber').read().strip()
# None이면 0을 리턴하는 함수
def validate_or_default(value):
    if value is None:
        return 0
    return value
def find_intersection(start1, end1, start2, end2):
    # 교차점 계산 (두 직선이 수직 및 수평으로 만나는 경우에 대해서만)
    if start1[0] == end1[0]:
        return (start1[0], start2[1])
    else:
        return (start2[0], start1[1])
def calculate_fillet_point(center, point, radius):
    # 필렛 접점 계산
    dx = point[0] - center[0]
    dy = point[1] - center[1]
    if abs(dx) > abs(dy):
        return (center[0] + radius * (1 if dx > 0 else -1), center[1])
    else:
        return (center[0], center[1] + radius * (1 if dy > 0 else -1))
def calculate_angle(center, point):
    """함수 설명:
    math.atan2(y, x): 점 (x, y)에 대한 아크탄젠트(atan2) 값을 반환합니다.
    이는 (0, 0)을 기준으로 (x, y)까지의 반시계 방향으로 회전한 각도를 라디안 단위로 반환합니다.
    math.degrees(): atan2의 결과(라디안)를 **각도(도)**로 변환합니다.
    (point[1] - center[1], point[0] - center[0]): center에서 point까지의 상대적인 y, x 차이를 구합니다."
    """
    # 각도 계산
    return math.degrees(math.atan2(point[1] - center[1], point[0] - center[0]))
def add_90_degree_fillet(doc, start1, end1, start2, end2, radius):
    msp = doc.modelspace()

    # 교차점 찾기
    intersection_point = find_intersection(start1, end1, start2, end2)

    # 필렛 접점 계산
    point1 = calculate_fillet_point(intersection_point, end1, radius)
    point2 = calculate_fillet_point(intersection_point, end2, radius)

    # 각도 계산
    start_angle = calculate_angle(intersection_point, point1)
    end_angle = calculate_angle(intersection_point, point2)

    # 필렛 원호 그리기
    msp.add_arc(
        center=intersection_point,
        radius=radius,
        start_angle=start_angle,
        end_angle=end_angle,
        dxfattribs={'layer': '레이져'},
    )    
    return msp
def calculate_midpoint(point1, point2):
    return ((point1[0] + point2[0]) / 2, (point1[1] + point2[1]) / 2)
def calculate_circle_center(midpoint, radius, point1, point2):
    dx = point2[0] - point1[0]
    dy = point2[1] - point1[1]
    dist = math.sqrt(dx**2 + dy**2)
    factor = math.sqrt(radius**2 - (dist / 2)**2) / dist
    return (midpoint[0] - factor * dy, midpoint[1] + factor * dx)
def add_arc_between_points(doc, point1, point2, radius):
    msp = doc.modelspace()
    
    # 원의 중심 계산 (반지름이 마이너스일때는 point1, point2 변경)
    if radius < 0 :
        # 중점 계산
        midpoint = calculate_midpoint(point1, point2)
        center = calculate_circle_center(midpoint, radius * - 1, point2, point1)
    else:
        midpoint = calculate_midpoint(point1, point2)
        center = calculate_circle_center(midpoint, radius, point1, point2)

    # 각도 계산
    start_angle = calculate_angle(center, point1)
    end_angle = calculate_angle(center, point2)

    # 아크 그리기
    msp.add_arc(
        center=center,
        radius=radius,
        start_angle=start_angle,
        end_angle=end_angle,
        dxfattribs={'layer': '레이져'},
    )    
    return msp
def dim_leader_line(doc, start_x, start_y, end_x, end_y, text, layer='JKW', style='0', text_height=22):
    msp = doc.modelspace()

    text_style_name = 'JKW'
        
    # override 설정
    override_settings = {
        'dimasz': 15
    }    

    # 지시선 추가
    leader = msp.add_leader(
        vertices=[(start_x, start_y), (end_x, end_y)], # 시작점과 끝점
        dxfattribs={
            'dimstyle': layer,
            'layer': layer
        },
        override=override_settings
    )

    # 텍스트 추가 (선택적)
    if text:
        msp.add_mtext(text, dxfattribs={
            'insert': (end_x + 10, end_y + 5), # 텍스트 위치 조정
            'layer': style,
            'char_height': text_height,
            'style': text_style_name,
            'attachment_point': 1  # 텍스트 정렬 방식 설정
        })

    return leader
def dim_leader(doc, start_x, start_y, end_x, end_y, text, text_height=20, direction=None, option=None, distance=None):
    """  distance=13 # 글자크기에 대한 정의임  """
    global saved_DimXpos, saved_DimYpos, saved_text_height, saved_text_gap, saved_direction, dimdistance, dim_horizontalbase, dim_verticalbase, distanceXpos, distanceYpos 
    
    msp = doc.modelspace()
    layer = '0'        
    text_style_name = 'JKW'            
    # text_style_name = selected_dimstyle
    # override 설정
    override_settings = {
        'dimasz': 15
    }

    # 텍스트 위치 조정 및 꺾이는 지점 설정
    if option is None:    
        text_offset_x = 20
        text_offset_y = 20
    else:
        text_offset_x = 0
        text_offset_y = 0        
    
    if direction == 'leftToright':
        mid_x = end_x - text_offset_x
        mid_y = end_y  # 텍스트 앞에서 꺾임
        text_position = (end_x, end_y)
    elif direction == 'rightToleft':
        mid_x = end_x + text_offset_x
        mid_y = end_y  # 텍스트 앞에서 꺾임
        if distance is None:
            text_position = (end_x - len(text) * 13, end_y)
        else:
            text_position = (end_x - len(text) * distance, end_y)
    else:
        mid_x = (start_x + end_x) / 2
        mid_y = (start_y + end_y) / 2
        text_position = (end_x + text_offset_x, end_y + text_offset_y)

    # 지시선 추가
    leader = msp.add_leader(
        vertices=[(start_x, start_y), (mid_x, mid_y-text_height/2), (end_x, end_y-text_height/2)], # 시작점, 중간점(문자 앞에서 꺾임), 끝점
        dxfattribs={
            'dimstyle': text_style_name,
            'layer': layer,
            'color': 3  # 녹색 (AutoCAD 색상 인덱스에서 3번은 녹색)
        },
        override=override_settings
    )

    if option is None:
        # 텍스트 추가 (선택적)
        if text:
            msp.add_mtext(text, dxfattribs={
                'insert': text_position,
                'layer': layer,
                'char_height': text_height,
                'style': text_style_name,
                'attachment_point': 1, # 텍스트 정렬 방식 설정
                'color': 0  # 노란색 (AutoCAD 색상 인덱스에서 2번은 노란색)
            })

    return leader
def line(doc, x1, y1, x2, y2, layer=None):
    global saved_Xpos, saved_Ypos  # 전역 변수로 사용할 것임을 명시
    
    # 모델 공간 가져오기
    msp = doc.modelspace()
    
    # 선 추가
    start_point = (x1, y1)
    end_point = (x2, y2)
    if layer:
        # 절곡선 22 layer는 ltscale을 조정한다
        if(layer=="22"):
            msp.add_line(start=start_point, end=end_point, dxfattribs={'layer': layer, 'ltscale' : 30})
        else:
            msp.add_line(start=start_point, end=end_point, dxfattribs={'layer': layer })
    else:
        msp.add_line(start=start_point, end=end_point)

    # 다음 선분의 시작점을 업데이트
    saved_Xpos = x2
    saved_Ypos = y2        
def lineto(doc, x, y, layer=None):
    global saved_Xpos, saved_Ypos  # 전역 변수로 사용할 것임을 명시
    
    # 현재 위치를 시작점으로 설정
    start_x = saved_Xpos
    start_y = saved_Ypos

    # 끝점 좌표 계산
    end_x = x
    end_y = y

    # 모델 공간 (2D)을 가져옴
    msp = doc.modelspace()

    # 선 추가
    start_point = (start_x, start_y)
    end_point = (end_x, end_y)
    if layer:
        msp.add_line(start=start_point, end=end_point, dxfattribs={'layer': layer})
    else:
        msp.add_line(start=start_point, end=end_point)

    # 다음 선분의 시작점을 업데이트
    saved_Xpos = end_x
    saved_Ypos = end_y
def lineclose(doc, start_index, end_index, layer='레이져'):    

    firstX, firstY = globals()[f'X{start_index}'], globals()[f'Y{start_index}']
    prev_x, prev_y = globals()[f'X{start_index}'], globals()[f'Y{start_index}']

    # start_index+1부터 end_index까지 반복합니다.
    for i in range(start_index + 1, end_index + 1):
        # 현재 인덱스의 좌표를 가져옵니다.
        curr_x, curr_y = globals()[f'X{i}'], globals()[f'Y{i}']

        # 이전 좌표에서 현재 좌표까지 선을 그립니다.
        line(doc, prev_x, prev_y, curr_x, curr_y, layer)

        # 이전 좌표 업데이트
        prev_x, prev_y = curr_x, curr_y

        print(f"prev_x {prev_x}" )
        print(f"prev_y {prev_y}" )
    
    # 마지막으로 첫번째 점과 연결
    line(doc, prev_x, prev_y, firstX, firstY, layer)        
def rectangle(doc, x1, y1, dx, dy, layer=None, offset=None):
    if offset is not None:
        # 네 개의 선분으로 직사각형 그리기 offset 추가
        line(doc, x1+offset, y1+offset, dx-offset, y1+offset, layer=layer)   
        lineto(doc, dx - offset, dy - offset, layer=layer)  
        lineto(doc, x1 + offset, dy - offset, layer=layer)  
        lineto(doc, x1 + offset, y1 + offset, layer=layer)  
    else:        
        # 네 개의 선분으로 직사각형 그리기
        line(doc, x1, y1, dx, y1, layer=layer)   
        line(doc, dx, y1, dx, dy, layer=layer)   
        line(doc, dx, dy, x1, dy, layer=layer)   
        line(doc, x1, dy, x1, y1, layer=layer)   
def add_angular_dim_2l(drawing, base, line1, line2, location=None, text=None, text_rotation=None, dimstyle='0', override=None, dxfattribs=None):
    # Create a new dimension line
    dim = drawing.dimstyle.add(
        'EZ_ANGULAR_2L',
        dxfattribs={
            'dimstyle': dimstyle,
            'dimtad': 1, # Place text above dimension line
            'dimtih': False, # Align text horizontally to dimension line
            'dimtoh': False, # Align text outside horizontal
            'dimdec': 1
        }
    )
    
    # If override is provided, update the dimension style
    if override:
        dim.update(override)
    
    # Create angular dimension
    angular_dim = drawing.add_angular_dim_2l(
        base=base,
        line1=line1,
        line2=line2,
        location=location,
        text=text,
        text_rotation=text_rotation,
        dimstyle=dim.name,
        override=override,
        dxfattribs=dxfattribs,
    )
    
    # Render the dimension
    angular_dim.render()
    return angular_dim
def dim_angular(doc, x1, y1, x2, y2, x3, y3, x4, y4, distance=80, direction="left", dimstyle="JKW"):
    msp = doc.modelspace()

    # 두 선분의 좌표
    p1 = (x1, y1)
    p2 = (x2, y2)
    p3 = (x3, y3)
    p4 = (x4, y4)

    # 치수선의 위치 계산
    base_x = (p1[0] + p2[0] + p3[0] + p4[0]) / 4
    base_y = (p1[1] + p2[1] + p3[1] + p4[1]) / 4

    if direction=="left":
        base = (base_x - distance, base_y)
        # 각도 치수선 추가
        dimension = msp.add_angular_dim_2l(
            base=base, # 치수선 위치
            line1=(p1, p2), # 첫 번째 선의 시작점과 끝점
            line2=(p3, p4), # 두 번째 선의 시작점과 끝점
            dimstyle=dimstyle, # 치수 스타일        
            override={'dimtxt': 0.27, 'dimgap': 0.02, 'dimscl' : 1, 'dimlfac' : 1, 'dimclrt': 7, 'dimdsep': 46, 'dimdec': 0 } # 선색 백색 소수점 . 표기
        )

    elif direction=="up":
        base = (base_x, base_y + distance)          
        # "dimtad": 1, # 0=center; 1=above; 4=below;  
        # 각도 치수선 추가
        dimension = msp.add_angular_dim_2l(
            base=base, # 치수선 위치
            line1=(p1, p2), # 첫 번째 선의 시작점과 끝점
            line2=(p3, p4), # 두 번째 선의 시작점과 끝점
            dimstyle=dimstyle, # 치수 스타일        
            override={'dimtxt': 0.27, 'dimgap': 0.02, 'dimscl' : 1, 'dimlfac' : 1, 'dimclrt': 7, 'dimdsep': 46, 'dimdec': 0 } # 선색 백색 소수점 . 표기
        )        
    elif direction=="down":
        base = (base_x, base_y - distance)                    
    else:
        base = (base_x + distance, base_y)    
        # 각도 치수선 추가
        dimension = msp.add_angular_dim_2l(
            base=base, # 치수선 위치
            line1=(p1, p2), # 첫 번째 선의 시작점과 끝점
            line2=(p3, p4), # 두 번째 선의 시작점과 끝점
            dimstyle=dimstyle, # 치수 스타일        
            override={'dimtxt': 0.27, 'dimgap': 0.02, 'dimscl' : 1, 'dimlfac' : 1, 'dimclrt': 7, 'dimdsep': 46, 'dimdec': 0 } # 선색 백색 소수점 . 표기
        )        

    # 치수선의 기하학적 형태 생성
    dimension.render()
    return dimension
def dim_diameter(doc, center, diameter, angle, dimstyle="JKW", override=None):
    msp = doc.modelspace()
    
    # 기본 지름 치수선 추가
    dimension = msp.add_diameter_dim(
        center=center, # 원의 중심점
        radius=diameter/2, # 반지름
        angle=angle, # 치수선 각도
        dimstyle=dimstyle, # 치수 스타일
        override={"dimtoh": 1}    # 추가 스타일 설정 (옵션) 지시선이 한번 꺾여서 글자각도가 표준형으로 나오는 옵션
    )
    
    # 치수선의 기하학적 형태 생성
    dimension.render()
def dim_string(doc, x1, y1, x2, y2, dis, textstr, text_height=0.22, text_gap=0.05, direction="up"):
    global saved_DimXpos, saved_DimYpos, saved_text_height, saved_text_gap, saved_direction, dimdistance, dim_horizontalbase, dim_verticalbase

    msp = doc.modelspace()
    dim_style = 'JKW'
    layer = "JKW"

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점이 있는지 확인
    if dimension_value % 1 != 0:
        dimdec = 1  # 소수점이 있는 경우 소수점 첫째 자리까지 표시
    else:
        dimdec = 0  # 소수점이 없는 경우 소수점 표시 없음

    # override 설정
    override_settings = {
        'dimtxt': text_height,
        'dimgap': text_gap,
        'dimscl': 1,
        'dimlfac': 1,
        'dimclrt': 7,
        'dimdsep': 46,
        'dimdec': dimdec,
        'dimtix': 1, # 치수선 내부에 텍스트를 표시 (필요에 따라)
        'dimtad': 1 # 치수선 상단에 텍스트를 표시 (필요에 따라)               
    }

    # 방향에 따른 치수선 추가
    if direction == "up":
        dimension = msp.add_linear_dim(
            base=(x1, y1 + dis),
            dimstyle=dim_style,
            p1=(x1, y1),
            p2=(x2, y2),
            dxfattribs={'layer': layer},
            text = textstr,
            override=override_settings
        )
    elif direction == "down":
        dimension = msp.add_linear_dim(
            dimstyle=dim_style,
            base=(x1, y1 - dis),
            p1=(x1, y1),
            p2=(x2, y2),
            dxfattribs={'layer': layer},
            text = textstr,         
            override=override_settings
        )
    elif direction == "left":        
        dimension =  msp.add_linear_dim(
            dimstyle=dim_style,
            base=(x1 - dis, y1),
            angle=90,
            p1=(x1, y1),
            p2=(x2, y2),
            dxfattribs={'layer': layer},
            text = textstr, 
            override=override_settings
        )     
    elif direction == "right":        
        dimension =  msp.add_linear_dim(
            dimstyle=dim_style,
            base=(x1 + dis, y1),
            p1=(x1, y1),
            p2=(x2, y2),
            dxfattribs={'layer': layer},
            text = textstr, 
            angle=270,         
            override=override_settings
        )
    else:
        raise ValueError("Invalid direction. Use 'up', 'down', or 'aligned'.")

    dimension.render()
    return dimension

def d(doc, x1, y1, x2, y2, dis, text_height=0.22, text_gap=0.05, direction="up", option=None, startoption=None, text=None) :
    global saved_DimXpos, saved_DimYpos, saved_text_height, saved_text_gap, saved_direction, dimdistance, dim_horizontalbase, dim_verticalbase, distanceXpos, distanceYpos

    # Option 처리
    if option == 'reverse':
        x1, x2 = x2, x1
        y1, y2 = y2, y1     
        saved_DimXpos, saved_DimYpos = x2, y2
    else:
        saved_DimXpos, saved_DimYpos = x2, y2

    dimdistance = dis
    saved_text_height = text_height
    saved_text_gap = text_gap
    saved_direction = direction

    # 연속선 구현을 위한 구문
    if startoption is None:        
        if direction == "left":             
            distance = min(x1, x2) - dis            
            dim_horizontalbase = distance
        elif direction == "right":   
            distance = max(x1, x2) + dis                       
            dim_horizontalbase = distance 
        elif direction == "up":   
            distance = max(y1, y2) + dis            
            dim_verticalbase = distance
        elif direction == "down":     
            distance = min(y1, y2) - dis                                  
            dim_horizontalbase = distance
    else:
        if direction == "left":             
            distance = distanceXpos
        elif direction == "right":                        
            distance = distanceXpos
        elif direction == "up":            
            distance = distanceYpos
        elif direction == "down":     
            distance = distanceYpos
     

    # flip을 선언하면 치수선의 시작과 끝을 바꾼다. 저장된 좌표는 지장없다.
    # 치수선의 시작점 끝점에 따라 치수선이 나오는 것을 만들기 위함이다. 
    # 연속치수선때는 좌표가 바뀌면 안되기때문에 고려한 부분이다.

    msp = doc.modelspace()
    dim_style = 'JKW'
    layer = "JKW"

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점 확인
    dimdec = 1 if dimension_value % 1 != 0 else 0

    # override 설정
    override_settings = {      
        'dimtxt': text_height, 
        'dimgap': text_gap if text_gap is not None else 0.05, # 여기에서 dimgap에 기본값 설정
        'dimscl': 1,
        'dimlfac': 1,
        'dimclrt': 7,
        'dimdsep': 46,
        'dimdec': dimdec,
        'dimtix': 1, 
        'dimtad': 1  
    }

    # 방향에 따른 치수선 추가
    if direction == "up":
        add_dim_args = {
            'base': ((x1+x2)/2, distance),
            'dimstyle': dim_style,
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        # 절대값으로 거리를 저장
        if startoption is None:
            distanceYpos = max(y1, y2) + dis           
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "down":
        add_dim_args = {
            'dimstyle': dim_style,
            'base': ((x1+x2)/2, distance),
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        # 절대값으로 거리를 저장
        if startoption is None:
            distanceYpos = min(y1, y2) - dis         
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "left":
        if option == 'reverse':   
            add_dim_args = {
                'dimstyle': dim_style,
                'base': (distance, (y1+y2)/2),
                'angle': 90,
                'p1': (x1, y1),
                'p2': (x2, y2),
                'override': override_settings
            }
            # 절대값으로 거리를 저장
            if startoption is None:
                distanceXpos = min(x1, x2) - dis 
        else:
            add_dim_args = {
                'dimstyle': dim_style,
                'base': (distance, (y1+y2)/2),
                'angle': 90,
                'p1': (x2, y2),
                'p2': (x1, y1),
                'override': override_settings
            }       
            # 절대값으로 거리를 저장
            if startoption is None:
                distanceXpos = min(x1, x2) - dis 
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "right":
        add_dim_args = {
            'dimstyle': dim_style,
            'base': (distance, (y1+y2)/2),
            'angle': 90,
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        # 절대값으로 거리를 저장
        if startoption is None:
            distanceXpos = max(x1, x2) + dis         
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "aligned":
        add_dim_args = {
            'dimstyle': dim_style,
            'points': [(x1, y1), (x2, y2)],
            'distance': dis,
            'dxfattribs': {'layer': layer},
            'override': override_settings
        }
        if text is not None :
            add_dim_args['text'] = text
        return msp.add_multi_point_linear_dim(**add_dim_args)

    else:
        raise ValueError("Invalid direction. Use 'up', 'down', 'left', 'right', or 'aligned'.")

def dc(doc, x, y, distance=None, option=None, text=None) :    
    global saved_DimXpos, saved_DimYpos, saved_text_height, saved_text_gap, saved_direction, dimdistance, dim_horizontalbase, dim_verticalbase, distanceXpos, distanceYpos

    x1 = saved_DimXpos
    y1 = saved_DimYpos
    x2 = x
    y2 = y        

    if distance is not None :
        dimdistance = distance    

    # reverse 옵션 처리
    if option == 'reverse':
        x1, x2 = x2, x1
        y1, y2 = y2, y1                

    d(doc, x1, y1, x2, y2, dimdistance, text_height=saved_text_height, text_gap=saved_text_gap, direction=saved_direction, startoption='continue', text=text)

def dim(doc, x1, y1, x2, y2, dis, text_height=0.22, text_gap=0.05, direction="up", option=None, startoption=None, text=None, localdimstyle=None):
    global saved_DimXpos, saved_DimYpos, saved_text_height, saved_text_gap, saved_direction, dimdistance, dim_horizontalbase, dim_verticalbase, dim_style

    # Option 처리
    if option == 'reverse':
        saved_DimXpos, saved_DimYpos = x1, y1
    else:
        saved_DimXpos, saved_DimYpos = x2, y2

    dimdistance = dis
    saved_text_height = text_height
    saved_text_gap = text_gap
    saved_direction = direction

    # 연속선 구현을 위한 구문
    if startoption is None:
        if direction == "left":             
            dim_horizontalbase = dis - (x1 - x2)
        elif direction == "right":                        
            dim_horizontalbase = dis 
        elif direction == "up":            
            dim_verticalbase = dis
        elif direction == "down":                        
            dim_verticalbase = dis   

    # flip을 선언하면 치수선의 시작과 끝을 바꾼다. 저장된 좌표는 지장없다.
    # 치수선의 시작점 끝점에 따라 치수선이 나오는 것을 만들기 위함이다. 
    # 연속치수선때는 좌표가 바뀌면 안되기때문에 고려한 부분이다.

    msp = doc.modelspace()
    if localdimstyle is None :
        dim_style = 'JKW'
        layer = "JKW"
    elif localdimstyle == '0.2 JKW':
        dim_style = '0.2 JKW'
    elif localdimstyle == '0.5 JKW':
        dim_style = '0.5 JKW'        

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점 확인
    dimdec = 1 if dimension_value % 1 != 0 else 0

    # override 설정
    override_settings = {      
        'dimtxt': text_height, 
        'dimgap': text_gap if text_gap is not None else 0.05, # 여기에서 dimgap에 기본값 설정
        'dimscl': 1,
        'dimlfac': 1,
        'dimclrt': 7,
        'dimdsep': 46,
        'dimdec': dimdec,
        'dimtix': 1, 
        'dimtad': 1  
    }

    # 방향에 따른 치수선 추가
    if direction == "up":
        add_dim_args = {
            'base': (x1, y1 + dis),
            'dimstyle': dim_style,
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "down":
        add_dim_args = {
            'dimstyle': dim_style,
            'base': (x1, y1 - dis),
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "left":
        if option == 'reverse':   
            add_dim_args = {
                'dimstyle': dim_style,
                'base': (x2 - dis, y2),
                'angle': 90,
                'p1': (x1, y1),
                'p2': (x2, y2),
                'override': override_settings
            }
        else:
            add_dim_args = {
                'dimstyle': dim_style,
                'base': (x1 - dis, y1),
                'angle': 90,
                'p1': (x2, y2),
                'p2': (x1, y1),
                'override': override_settings
            }        
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "right":
        add_dim_args = {
            'dimstyle': dim_style,
            'base': (x1 + dis, y1),
            'angle': 90,
            'p1': (x1, y1),
            'p2': (x2, y2),
            'override': override_settings
        }
        if text is not None :
            add_dim_args['text'] = text
            msp.add_linear_dim(**add_dim_args).render()
            return
        else:
            return msp.add_linear_dim(**add_dim_args)

    elif direction == "aligned":
        add_dim_args = {
            'dimstyle': dim_style,
            'points': [(x1, y1), (x2, y2)],
            'distance': dis,
            'dxfattribs': {'layer': layer},
            'override': override_settings
        }
        if text is not None :
            add_dim_args['text'] = text
        return msp.add_multi_point_linear_dim(**add_dim_args)

    else:
        raise ValueError("Invalid direction. Use 'up', 'down', 'left', 'right', or 'aligned'.")

def dimcontinue(doc, x, y, distance=None, option=None) :    
    global saved_DimXpos
    global saved_DimYpos
    global saved_text_height
    global saved_text_gap
    global saved_direction
    global dimdistance
    global dim_horizontalbase
    global dim_verticalbase

    x1 = saved_DimXpos
    y1 = saved_DimYpos
    x2 = x
    y2 = y        



# 방향에 대한 정의를 하고 연속적으로 차이를 감지해서 처리하는 것이다.
    if saved_direction=="left" :
        dimdistance = dim_horizontalbase
        # 재계산해야 함.        
        dim_horizontalbase = dimdistance - (x1 - x2)        
    if saved_direction=="right" :
        dimdistance = dim_horizontalbase 
        # 재계산해야 함.
        dim_horizontalbase = dimdistance - (x2 - x1)
    if saved_direction=="up" :
        dimdistance = dim_verticalbase
        # 재계산해야 함.
        dim_verticalbase = dimdistance - (y2 - y1)
    if saved_direction=="down" :
        dimdistance = dim_verticalbase 
        # 재계산해야 함.
        dim_verticalbase = dimdistance - (y1 - y2)

    if distance is not None :
        dimdistance = distance    

    # reverse 옵션 처리
    if option == 'reverse':
        x1, x2 = x2, x1
        y1, y2 = y2, y1                

    dim(doc, x1, y1, x2, y2, dimdistance, text_height=saved_text_height, text_gap=saved_text_gap, direction=saved_direction, startoption='continue', localdimstyle = dim_style)

def dimto(doc, x2, y2, dis, text_height=0.20, text_gap=None, option=None) :    
    global saved_DimXpos
    global saved_DimYpos
    global saved_text_height
    global saved_text_gap
    global saved_direction

    if text_gap is not None :
        text_gap = saved_text_gap
        saved_text_gap = text_gap

    if text_height is not None :
        text_height = saved_text_height
        saved_text_height = text_height

    dim(doc, saved_DimXpos, saved_DimYpos, x2, y2, dis, text_height=text_height, text_gap=text_gap, direction=saved_direction, option=option)

def create_vertical_dim(doc, x1, y1, x2, y2, dis, angle, layer=None, text_height=0.22, text_gap=0.05):
    msp = doc.modelspace()
    dim_style = layer  # 치수 스타일 이름
    points = [(x1, y1), (x2, y2)]

    if angle==None :
        angle = 270

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점이 있는지 확인
    if dimension_value % 1 != 0:
        dimdec = 1  # 소수점이 있는 경우 소수점 첫째 자리까지 표시
    else:
        dimdec = 1  # 소수점이 없는 경우 소수점 표시 없음    

    return msp.add_multi_point_linear_dim(
        base=(x1 + dis if angle == 270 else x1 - dis, y1), #40은 보정
        points = points,
        angle = angle,
        dimstyle = dim_style,
        discard = True,
        dxfattribs = {'layer': layer},
        # 치수 문자 위치 조정 (0: 치수선 위, 1: 치수선 옆) 'dimtmove': 1 
        # override={'dimtxt': text_height, 'dimscl': 1, 'dimlfac': 1, 'dimclrt': 7, 'dimdec': dimdec, 'dimdsep': 46, 'dimtmove': 3  }
        override = {'dimtxt': text_height, 'dimgap': text_gap, 'dimscl': 1, 'dimlfac': 1, 'dimclrt': 7, 'dimdec': dimdec, 'dimdsep': 46, 'dimtmove': 3  }
    )

def dim_vertical_right(doc, x1, y1, x2, y2, dis, layer="JKW", text_height=0.22, text_gap=0.07, angle=None):
    return create_vertical_dim(doc, x1, y1, x2, y2, dis, angle, layer, text_height, text_gap)

def dim_vertical_left(doc, x1, y1, x2, y2, dis, layer=None, text_height=0.22, text_gap=0.07):
    return create_vertical_dim(doc, x1, y1, x2, y2, dis, 90, layer, text_height, text_gap)
  
def create_vertical_dim_string(doc, x1, y1, x2, y2, dis, angle, textstr, text_height=0.22, text_gap=0.07):    
    msp = doc.modelspace()
    dim_style = 'JKW'
    layer = "JKW"
    points = [(x1, y1), (x2, y2)]

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점이 있는지 확인
    if dimension_value % 1 != 0:
        dimdec = 1  # 소수점이 있는 경우 소수점 첫째 자리까지 표시
    else:
        dimdec = 1  # 소수점이 없는 경우 소수점 표시 없음    

    return msp.add_multi_point_linear_dim(
        base=(x1 + dis if angle == 270 else x1 - dis, y1), #40은 보정
        points=points,
        angle=angle,
        dimstyle=dim_style,
        discard=True,
        dxfattribs={'layer': layer},        
        # 치수 문자 위치 조정 (0: 치수선 위, 1: 치수선 옆) 'dimtmove': 1 
        override={'dimtxt': text_height, 'dimgap': text_gap, 'dimscl': 1, 'dimlfac': 1, 'dimclrt': 7, 'dimdec': dimdec, 'dimdsep': 46, 'dimtmove': 3, 'text' : textstr      }
    )

def dim_vertical_right_string(doc, x1, y1, x2, y2, dis, textstr, text_height=0.22, text_gap=0.07):
    return create_vertical_dim_string(doc, x1, y1, x2, y2, dis, 270, textstr, text_height, text_gap)

def dim_vertical_left_string(doc, x1, y1, x2, y2, dis, textstr, text_height=0.22, text_gap=0.07):
    return create_vertical_dim_string(doc, x1, y1, x2, y2, dis, 90, textstr, text_height, text_gap)

def draw_Text_direction(doc, x, y, size, text, layer=None, rotation = 90):
    # 모델 공간 (2D)을 가져옴
    msp = doc.modelspace()

    # MText 객체 생성
    mtext = msp.add_mtext(
        text, # 텍스트 내용
        dxfattribs={
            'layer': layer, # 레이어 지정
            'style': text_style_name, # 텍스트 스타일 지정
            'char_height': size, # 문자 높이 (크기) 지정
        }
    )

    # MText 위치와 회전 설정
    mtext.set_location(insert=(x, y), attachment_point=1, rotation=rotation)

    return mtext

def draw_Text(doc, x, y, size, text, layer=None):
    # 모델 공간 (2D)을 가져옴
    msp = doc.modelspace()

    # 텍스트 추가 및 생성된 Text 객체 가져오기
    text_entity = msp.add_text(
        text, # 텍스트 내용
        dxfattribs={
            'layer': layer, # 레이어 지정
            'style': text_style_name, # 텍스트 스타일 지정
            'height': size, # 텍스트 높이 (크기) 지정
        }
    )

    # Text 객체의 위치 설정
    # text_entity.set_placement((x, y), align=TextEntityAlignment.CENTER)  # 텍스트 위치 및 정렬 설정
    # Text 객체의 위치 설정 (가로 및 세로 중앙 정렬)
    text_entity.set_placement((x, y), align=TextEntityAlignment.MIDDLE_LEFT)

def draw_circle(doc, center_x, center_y, radius, layer='0', color='7'):
    """
    DXF 문서에 원을 그리는 함수
    :param doc: ezdxf 문서 객체
    :param center_x: 원의 중심 x 좌표
    :param center_y: 원의 중심 y 좌표
    :param radius: 원의 반지름
    :param layer: 원을 추가할 레이어 이름 (기본값은 '0')
    지름으로 수정 /2 적용
    """
    msp = doc.modelspace()
    circle = msp.add_circle(center=(center_x, center_y), radius=radius/2, dxfattribs={'layer': layer, 'color' : color})
    return circle
def draw_arc_Global(doc, x1, y1, x2, y2, radius, direction, layer='레이져'): # radius는 반지름을 넣는다. 지름이 아님
    msp = doc.modelspace()
    
    # 점들과 반지름을 이용하여 중점과 원의 중심 계산
    midpoint = calculate_midpoint((x1, y1), (x2, y2))
    center = calculate_circle_center(midpoint, radius, (x1, y1), (x2, y2))

    # 시작 각도와 끝 각도 계산
    start_angle = calculate_angle(center, (x1, y1))
    end_angle = calculate_angle(center, (x2, y2))

    # 방향에 따라 각도 조정
    if direction == 'up' or direction == 'down':
        if start_angle > end_angle:
            start_angle, end_angle = end_angle, start_angle
        if direction == 'down':
            start_angle += 180
            end_angle += 180
    elif direction == 'left' or direction == 'right':
        if start_angle > end_angle:
            start_angle, end_angle = end_angle, start_angle
        if direction == 'right':
            start_angle += 180
            end_angle += 180

    # 아크 그리기
    msp.add_arc(
        center=center,
        radius=radius,
        start_angle=start_angle,
        end_angle=end_angle,
        dxfattribs={'layer': layer},
    )
    return msp
def draw_arc_slot(doc, center, radius, start_angle, end_angle, layer):
    """
    Draws an arc in the DXF document.
    
    Parameters:
    doc (ezdxf.document): The DXF document to draw on.
    center (tuple): The (x, y) coordinates of the arc's center.
    radius (float): The radius of the arc.
    start_angle (float): The starting angle of the arc in degrees.
    end_angle (float): The ending angle of the arc in degrees.
    layer (str): The layer to draw the arc on.
    """
    msp = doc.modelspace()
    msp.add_arc(
        center=center,
        radius=radius,
        start_angle=start_angle,
        end_angle=end_angle,
        dxfattribs={'layer': layer}
    )
def draw_slot(doc, x, y, size, direction="가로", option=None, layer='0'):
    """
    Draws a slot (장공) with specified parameters in a DXF document.
    
    Parameters:
    doc (ezdxf.document): The DXF document to draw on.
    x (float): The x-coordinate of the slot's center.
    y (float): The y-coordinate of the slot's center.
    size (str): The size of the slot in the format 'WxH' (e.g., '8x16').
    direction (str): The direction of the slot, either '가로' (horizontal) or '세로' (vertical). Default is '가로'.
    option (str): Optional feature for the slot. If 'cross', draws center lines extending beyond the slot. Default is None.
    layer (str): The layer to draw the slot on. Default is '0'.
    """    
    msp = doc.modelspace()

    # size를 분리하여 폭과 높이 계산
    width, height = map(float, size.lower().split('x'))

    print(f"width : {width}, height : {height}")

    if direction == "가로":
        slot_length = height - width
        slot_width = width
        radius = slot_width / 2
    else:  # 세로
        slot_length = width
        slot_width = height - width
        radius = slot_length / 2

    

    # 중심점을 기준으로 시작점과 끝점 계산
    if direction == "가로":
        start_point = (x - slot_length / 2, y)
        end_point = (x + slot_length / 2, y)
    else:  # 세로
        start_point = (x, y - slot_length / 2)
        end_point = (x, y + slot_length / 2)

    # 직선 부분 그리기
    if direction == "가로":
        msp.add_line((start_point[0], start_point[1] + radius), (end_point[0], end_point[1] + radius), dxfattribs={'layer': layer})
        msp.add_line((start_point[0], start_point[1] - radius), (end_point[0], end_point[1] - radius), dxfattribs={'layer': layer})
    else:  # 세로
        msp.add_line((start_point[0] + radius, start_point[1]), (end_point[0] + radius, end_point[1]), dxfattribs={'layer': layer})
        msp.add_line((start_point[0] - radius, start_point[1]), (end_point[0] - radius, end_point[1]), dxfattribs={'layer': layer})

    # 양 끝의 반원 그리기
    if direction == "가로":
        draw_arc_slot(doc, start_point, radius, 90, 270, layer)  # 반원 방향 수정
        draw_arc_slot(doc, end_point, radius, 270, 90, layer)  # 반원 방향 수정
    else:  # 세로
        draw_arc_slot(doc, start_point, radius, 180, 360, layer)  # 반원 방향 수정
        draw_arc_slot(doc, end_point, radius, 0, 180, layer)  # 반원 방향 수정

    # 옵션이 "cross"인 경우 중심선을 추가로 그리기
    if option == "cross":
        if direction == "가로":
            msp.add_line((x - slot_length / 2 - 8, y), (x + slot_length / 2 + 8, y), dxfattribs={'layer': 'CL'})
            msp.add_line((x, y - slot_width / 2 - 4), (x, y + slot_width / 2 + 4), dxfattribs={'layer': 'CL'})
        else:  # 세로
            msp.add_line((x - slot_width / 2 - 4, y), (x + slot_width / 2 + 4, y), dxfattribs={'layer': 'CL'})
            msp.add_line((x, y - slot_length / 2 - 8), (x, y + slot_length / 2 + 8), dxfattribs={'layer': 'CL'})

    return msp
# 중심선 적색표기 색상 1 적용 적색
def circle_cross(doc, center_x, center_y, radius, layer='0', color='7'):
    """
    DXF 문서에 원을 그리는 함수
    :param doc: ezdxf 문서 객체
    :param center_x: 원의 중심 x 좌표
    :param center_y: 원의 중심 y 좌표
    :param radius: 원의 반지름
    :param layer: 원을 추가할 레이어 이름 (기본값은 '0')
    지름으로 수정 /2 적용
    """
    draw_circle(doc, center_x, center_y, radius, layer=layer, color=color)
    # 적색 십자선 그려주기    
    line(doc, center_x - radius/2 - 5, center_y, center_x +  radius/2 + 5, center_y, layer="CEN" )
    line(doc, center_x, center_y - radius/2 - 5, center_x, center_y +  radius/2 + 5, layer="CEN" )
    return circle_cross

def extract_abs(a, b):
    return abs(a - b)    

# 도면틀 삽입
def insert_block(x, y, block_name, layer=None):    
    global msp  # 전역 msp 사용 명시
    scale = 1
    insert_point = (x, y, scale)

    if layer is None :
        # 블록 삽입하는 방법           
        msp.add_blockref(block_name, insert_point, dxfattribs={
            'xscale': scale,
            'yscale': scale,
            'rotation': 0
        })       
    else:
        msp.add_blockref(block_name, insert_point, dxfattribs={
            'xscale': scale,
            'yscale': scale,
            'rotation': 0,
            'layer' : layer
        })               

# 도면틀 삽입
def insert_frame(x, y, scale, sep, text):    
    if(sep =="drawings_frame"):
        block_name = "drawings_frame"
        insert_point = (x, y, scale)

        # 블록 삽입하는 방법           
        msp.add_blockref(block_name, insert_point, dxfattribs={
            'xscale': scale,
            'yscale': scale,
            'rotation': 0
        })

        textstr = f"{text}"       
        draw_Text(doc, x + 2220*scale, y + 220*scale, 23*scale, str(textstr), '0')
        draw_Text(doc, x + (2200+160)*scale, y + (220-88)*scale, 23*scale, "Light Case Accessory", '0')
        draw_Text(doc, x + (2200-320)*scale, y + (220-88)*scale, 23*scale, formatted_date, '0')

def envsettings():
    # 하드디스크 고유번호를 가져오는 코드 (시스템에 따라 다를 수 있음)
    disk_id = os.popen('wmic diskdrive get serialnumber').read().strip()
    data = {"DiskID": disk_id}
    with open(license_file_path, 'w', encoding='utf-8') as file:
        json.dump(data, file)

# LED바 길이 계산 (1030, 등도 계산되도록 로직개발)
def CalculateLEDbar(car_length):
    # Subtract 100 from the input value
    result = car_length - 100

    # Adjust to the next lower or equal multiple of 50
    if result % 100 > 50:
        result = (result // 100) * 100 + 100
    else:
        if 0 < result % 100 <= 50:
            # For values just over a multiple of 100, bump up to the next 50
            result = (result // 100) * 100 + 50
        else:
            # Otherwise, round down to the nearest 100
            result = (result // 100) * 100 

    return result

# LED바 길이 계산 031모델 -300
def CalculateLEDbar031(car_length):    
    # CW-300
    result = car_length - 300

    # Adjust to the next lower or equal multiple of 50
    if result % 100 > 50:
        result = (result // 100) * 100 + 100
    else:
        if 0 < result % 100 <= 50:
            # For values just over a multiple of 100, bump up to the next 50
            result = (result // 100) * 100 + 50
        else:
            # Otherwise, round down to the nearest 100
            result = (result // 100) * 100 

    return result

def calculate_light_positions_adjusted_031(dataCD):
    # Base coordinates
    y_coord1 = 125
    y_coord_last = dataCD - 125

    # Check if CD is over 2000 for additional coordinates calculation
    if dataCD > 2000:
        # Middle two coordinates with a gap
        midpoint = dataCD / 2
        y_coord4 = midpoint - 185
        y_coord5 = midpoint + 185

        # Calculating other coordinates with equidistant spacing
        interval = (y_coord4 - y_coord1) / 3
        y_coord2, y_coord3 = y_coord1 + interval, y_coord1 + 2 * interval
        y_coord6, y_coord7 = y_coord5 + interval, y_coord5 + 2 * interval

        # Adjusting coordinates to round off and add truncated decimals
        y_coord2, y_coord3, y_coord6 = adjust_coordinates(y_coord2, y_coord3, y_coord6)
        y_coord7 = round(y_coord7)

        all_y_coords = sorted([y_coord1, y_coord2, y_coord3, y_coord4, y_coord5, y_coord6, y_coord7, y_coord_last])
    else:
        # For CD 2000 or less, calculate additional coordinates
        midpoint = dataCD / 2
        y_coord2 = midpoint - 185
        y_coord3 = midpoint + 185

        # Calculating the additional coordinates
        y_coord4 = (y_coord1 + y_coord2) / 2
        y_coord5 = (y_coord3 + y_coord_last) / 2

        all_y_coords = sorted([y_coord1, y_coord4, y_coord2, y_coord3, y_coord5, y_coord_last])

    return all_y_coords

# Function to adjust coordinates with rounding and adding truncated decimals
def adjust_coordinates(coord1, coord2, coord3):
    adjusted_coord1 = round(coord1)
    decimal_part = coord1 - adjusted_coord1
    adjusted_coord2 = round(coord2 + decimal_part)
    adjusted_coord3 = round(coord3 + (coord2 - adjusted_coord2))

    return adjusted_coord1, adjusted_coord2, adjusted_coord3

def calculate_lc_position_031(inputCW, inputCD):
    # Calculate Y-coordinates for top and bottom holes
    Topholey = inputCD - 40
    Bottomy = 40

    # Initialize the list of x-coordinates for top and bottom holes
    Topholex = []
    Bottomx = []

    # CW에서 양쪽의 여유분 230을 제외한 값을 3등분
    divided_space = (inputCW - 230) // 3
    end_digit = divided_space % 10

    # Determine the number of holes and calculate their positions
    if inputCW > 1400:
        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution
        x1 = 115
        space_available = inputCW - 115 * 2

        if end_digit == 0: 
            # 정확히 나누어 떨어질 때
            middle_space = ((space_available // 3) // 10 * 10)
            side_space = ((space_available - middle_space) // 2 // 10 * 10)
            
        elif end_digit in [3, 4]: 
            # middle_space가 4로 끝나도록 조정
            middle_space = ((space_available // 3) // 10 * 10) + 4
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 3

        elif end_digit in [6, 7]: 
            # middle_space가 6으로 끝나도록 조정
            middle_space = ((space_available // 3) // 10 * 10) + 6
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 7

        else:
            # 기본 분배: 나머지 end_digit의 경우 기본적으로 10의 배수로 처리
            middle_space = (space_available // 3 // 10 * 10)
            side_space = ((space_available - middle_space) // 2 // 10 * 10)

    # 이후 x2 및 x3 계산 등에 사용할 수 있습니다.    

        # Calculating all x-coordinates
        x2 = x1 + side_space
        x3 = x2 + middle_space
        x4 = inputCW - 115

        Topholex = [x1, x2, x3, x4]
        Bottomx = [x1, x2, x3, x4]    
    else:
        # For CW < 1400, there are 3 holes
        x1 = 115
        x2 = (inputCW - 115 * 2) / 2 + 115
        x3 = inputCW - 115
        Topholex = [x1, x2, x3]
        Bottomx = [x1, x2, x3]

    # Create a dictionary to hold the coordinates
    lc_positions = {
        'Topholex': Topholex,
        'Topholey': Topholey,
        'Bottomx': Bottomx,
        'Bottomy': Bottomy,
        'end_digit' : end_digit
    }

    # print(f"lc_positions : {lc_positions}")
    return lc_positions

def calculate_lc_vertical_hole_positions_031(inputCW, inputCD):
    # 고정된 x 좌표들 설정
    leftholex = [115, 115, 115]
    rightholex = [inputCW - 115, inputCW - 115, inputCW - 115]

    # y 좌표를 계산하기 위한 로직
    if inputCD <= 1250:
        # CD가 1250 이하일 경우, y1만 존재
        y1 = (inputCD) / 2  # 40*2를 제외하고 중간값을 계산
        vertical_holey = [y1]
    elif inputCD > 1250 and inputCD <2100:
        # CD가 1250을 초과할 경우, y 좌표 3개를 계산
        divided_space = (inputCD - 80) // 3  # 40*2를 제외하고 3등분
        end_digit = divided_space % 10

        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution        
        space_available = inputCD - 40 * 2
        
        if end_digit in [3, 4]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 4
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 3   
            # print (f" 3, 4 선택됨 end_digit : {end_digit}, side_space : {side_space}, middle_space : {middle_space}")                 
        elif end_digit in [6, 7]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 6
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 7
        else:            
            middle_space = ((space_available // 3) // 10 * 10)             
            side_space = ((space_available - middle_space) // 2 // 10 * 10) 

        # Calculating all x-coordinates
        y1 = side_space + 40    # 40위에서 시작
        y2 = y1 + middle_space

        vertical_holey = [y1, y2]
        # print (f" end_digit : {end_digit}, side_space : {side_space}, middle_space : {middle_space}")
    # 2100 보다 크면 중간 홀 3개    
    else:         
        divided_space = (inputCD - 80) // 4  # 40*2를 제외하고 4등분
        end_digit = divided_space % 10

        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution        
        space_available = inputCD - 40 * 2
        
        if end_digit in [3, 4]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 4) // 10 * 10) + 4
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 3 // 10 * 10) + 3                    
        elif end_digit in [6, 7]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 4) // 10 * 10) + 6
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 3 // 10 * 10) + 7
        else:            
            middle_space = space_available // 4
            side_space = (space_available - middle_space) // 3

        # Calculating all x-coordinates
        y1 = side_space + 40
        y2 = y1 + middle_space
        y3 = y2 + side_space

        vertical_holey = [y1, y2, y3]

    # 좌표 딕셔너리 반환
    hole_positions = {
        'leftholex': leftholex,
        'rightholex': rightholex,
        'vertical_holey': vertical_holey
    }

    return hole_positions

# 패턴그리는 함수
def draw_patterned_lines(doc, start_x, end_x, y1, y2, layer):
    x = start_x
    gap_count = 0  # 1mm 간격으로 그린 선의 개수를 추적

    while x < end_x:
        line(doc, x, y1, x, y2, layer)

        # 1mm 간격으로 3개의 선을 그린 후 5mm 간격 적용
        gap_count += 1
        if gap_count == 3:
            x += 6
            gap_count = 0  # 간격 카운트 리셋
        else:
            x += 1.5

def ALprofile_draw(doc, x, y, length, direction="side"):   
    width = 42
    if direction=='side':
        # 마구리
        rectangle(doc, x, y, x+width, y - 2, layer="CL")        
        rectangle(doc, x, y - length - 2, x+width, y - length, layer="CL")        

        rectangle(doc, x, y - length - 2, x+width, y-2, layer="CL")        
        rectangle(doc, x + 1, y - 2, x+width-1, y-length, layer="CL")
        rectangle(doc, x+15, y - 50, x+width-15, y-length+50-2, layer="6")
        line(doc, x+15+1.2, y - 50, x+15+1.2, y-length+50-2, layer="hidden")
        line(doc, x+width-15-1.2, y - 50, x+width-15-1.2, y-length+50-2, layer="hidden")
        rectangle(doc, x+15+3, y - 50-2, x+width-15-3, y-length+50-2+2, layer="4")
    if direction=='topbottom':
        # 마구리
        rectangle(doc, x, y, x+2, y - width, layer="CL")        
        rectangle(doc, x + length - 2, y, x + length, y - width, layer="CL")        

        rectangle(doc, x, y, x+length, y-width, layer="CL")        
        rectangle(doc, x + 2, y - 1, x+length-2, y - width + 1, layer="CL")

        rectangle(doc, x + 50, y - 15, x + length - 50, y-width + 15, layer="6")
        line(doc, x + 50, y - 15 - 1.2, x + length - 50, y - 15 - 1.2, layer="hidden")
        line(doc, x + 50, y - width + 15 + 1.2, x + length - 50, y - width + 15 + 1.2, layer="hidden")        
        rectangle(doc, x+52, y - 18, x+length -52, y-width + 18, layer="4")

#031 홀간격계산
def calculate_hole_positions_separated_lc031(LCD):    
    center = LCD // 2  # 중심 지점
    central_start = center - 185  # 중심 구간 시작
    central_end = center + 185  # 중심 구간 끝

    left_positions = []
    right_positions = []

    # 중심으로부터 왼쪽 계산
    left_current = central_start
    while left_current > 0:
        left_positions.insert(0, left_current)
        left_current -= 220

    # 중심으로부터 오른쪽 계산
    right_current = central_end
    while right_current < LCD:
        right_positions.append(right_current)
        right_current += 220

    # 결과 리스트 합치기
    positions = left_positions + right_positions

    return positions

#031 led bar bracket 홀간격계산
def calculate_ledbar_bottomhole_lc031(ledbar_length):
    center = ledbar_length // 2  # 중심 지점
    central_gap = 350
    interval = 300  # 간격

    left_positions = []
    right_positions = []

    # 중앙 구간 시작 및 종료
    central_start = center - central_gap // 2
    central_end = center + central_gap // 2

    # 중심으로부터 왼쪽 계산
    left_current = central_start
    while left_current > 0:
        left_positions.insert(0, left_current)
        left_current -= interval

    # 중심으로부터 오른쪽 계산
    right_current = central_end
    while right_current < ledbar_length:
        right_positions.append(right_current)
        right_current += interval

    # 결과 리스트 합치기
    positions = left_positions + right_positions

    return positions

def calculate_ledbar_upperhole_lc031(ledbar_length):
    center = ledbar_length // 2  # 중심 지점
    initial_offset = 50  # 최초 간격
    interval = 100  # 이후 간격

    left_positions = []
    right_positions = []

    # 중심으로부터 왼쪽 계산 (초기 간격으로 시작)
    left_current = center - initial_offset
    while left_current > 0:
        left_positions.insert(0, left_current)
        left_current -= interval

    # 중심으로부터 오른쪽 계산 (초기 간격으로 시작)
    right_current = center + initial_offset
    while right_current < ledbar_length:
        right_positions.append(right_current)
        right_current += interval

    # 결과 리스트 합치기
    positions = left_positions + right_positions

    return positions

#008_A 홀간격계산
def calculate_lc_position_008_A(inputCW, inputCD):
    # Calculate Y-coordinates for top and bottom holes
    Topholey = (int)(inputCD - 75)
    Middley = (int)(inputCD/2)
    Bottomy = 75

    firstXgap = 75    
    middle_space = 0

    # Calculating all x-coordinates
    x1 = firstXgap
    x2 = (int)(inputCW/2)
    x3 = (int)(inputCW - firstXgap)

    Topholex = [x1, x2, x3]
    Middlex = [x1, x3]
    Bottomx = [x1, x2, x3]
    
    lc_positions = {
        'Topholex': Topholex,
        'Topholey': Topholey,
        'Middlex': Middlex,
        'Middley': Middley,
        'Bottomx': Bottomx,
        'Bottomy': Bottomy
    }


    return lc_positions


def calculate_lc_position_008_A_CW(inputCW, inputCD):#CW 팝너트 홀
    # Calculate Y-coordinates for top and bottom holes
    Topholey = (int)(inputCD - 13)
    Bottomy = 13

    firstXgap = 13    

    # Calculating all x-coordinates
    x1 = firstXgap
    x2 = (int)(inputCW/2 - firstXgap*2)/3 + firstXgap
    x3 = (int)(inputCW/2 - firstXgap*2)*2/3 + firstXgap
    x4 = (int)(inputCW/2 - firstXgap*2) + firstXgap


    total = x2 + x3 + x4
    # total의 소수 첫째자리(10분위)가 0이 아니면 y3에 0.1을 추가
    if int(total * 10) % 10 != 0:
        x3 += 0.1    

    x2 = (int)(x2*10)/10
    x3 = (int)(x3*10)/10
    x4 = (int)(x4*10)/10
    
    Topholex = [x1, x2, x3, x4]
    Bottomx = [x1, x2, x3, x4]
    
    lc_positions = {
        'Topholex': Topholex,
        'Topholey': Topholey,
        'Bottomx': Bottomx,
        'Bottomy': Bottomy
    }


    return lc_positions


def calculate_lc_position_008_A_CD(inputCW, inputCD):#CD 팝너트 홀
    # Calculate Y-coordinates for top and bottom holes
    Leftx = 13 
    Rightx = (int)(inputCW - 13)

    firstYgap = 13    

    # Calculating all x-coordinates
    y1 = firstYgap
    y2 = int((inputCD/2 - firstYgap*2)/3*10)/10 + firstYgap
    y3 = int((inputCD/2 - firstYgap*2)*2/3*10)/10 + firstYgap
    y4 = int((inputCD/2 - firstYgap)*10)/10
    
    total = y2 + y3 + y4
    # total의 소수 첫째자리(10분위)가 0이 아니면 y3에 0.1을 추가
    if int(total * 10) % 10 != 0:
        y3 += 0.1    

    y2 = int(y2*10)/10
    y3 = int(y3*10)/10
    y4 = int(y4*10)/10
    
    Lefty = [y1, y2, y3, y4]
    Righty = [y1, y2, y3, y4]
    
    lc_positions = {
        'Leftx': Leftx,
        'Lefty': Lefty,
        'Rightx': Rightx,
        'Righty': Righty
    }


    return lc_positions

def calculate_lc_position_008_A_ASSY_horizontal(inputCW, inputCD):
    liney = inputCD / 2

    firstXgap = 37.5    

    x1 = firstXgap
    x2 = (inputCW // 2 - firstXgap * 2) / 3 + firstXgap
    x3 = (inputCW // 2 - firstXgap * 2) * 2 / 3 + firstXgap
    x4 = inputCW / 2 - firstXgap
    
    total = x2 + x3 + x4

    # total의 소수 첫째 자리가 0이 아니면 x3에 0.1을 추가
    if round(total, 1) != total:
        x3 += 0.1    

    # 소수점 첫째 자리까지 유지
    x2 = round(x2, 1)
    x3 = round(x3, 1)
    x4 = round(x4, 1)

    linex = [x1, x2, x3, x4]
    
    lc_positions = {
        'linex': linex,
        'liney': liney,
    }

    return lc_positions

def calculate_lc_position_008_A_ASSY_vertical(inputCW, inputCD):
    linex = inputCW / 2

    firstYgap = 37.5    

    y1 = firstYgap
    base = (inputCD // 2 - firstYgap * 2) / 3  # 정수 연산을 명확히 처리
    
    y2 = base + firstYgap
    y3 = base * 2 + firstYgap
    y4 = inputCD / 2 - firstYgap
    
    total = y2 + y3 + y4

    # 소수 첫째 자리가 0이 아니면 y3에 0.1 추가
    if round(total, 1) != total:
        y3 += 0.1

    # 소수점 첫째 자리까지 유지
    y2, y3, y4 = round(y2, 1), round(y3, 1), round(y4, 1)

    liney = [y1, y2, y3, y4]
    
    lc_positions = {
        'linex': linex,
        'liney': liney,
    }

    return lc_positions

#035 홀간격계산
def calculate_lc_position_035(inputCW, inputCD):
    # Calculate Y-coordinates for top and bottom holes
    Topholey = inputCD - 40
    Bottomy = 40

    # Initialize the list of x-coordinates for top and bottom holes
    Topholex = []
    Bottomx = []

    firstXgap = 110
    side_space = 0
    space_available = 0
    middle_space = 0

    # CW에서 양쪽의 여유분 230을 제외한 값을 3등분
    divided_space = (inputCW - firstXgap*2) // 3
    end_digit = divided_space % 10    

    # Determine the number of holes and calculate their positions
    if inputCW >= 1600:
        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution
        x1 = firstXgap # 라이트케이스 첫홀 위치 25 띄운 치수 제외한 치수
        space_available = inputCW - firstXgap * 2
        
        if end_digit in [3, 4]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 4
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 3                    
        elif end_digit in [6, 7]:
            # The middle space needs to end with 4, therefore itd should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 6
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 7
        else:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) 
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) 

        # Calculating all x-coordinates
        x2 = x1 + side_space
        x3 = x2 + middle_space
        x4 = inputCW - firstXgap

        Topholex = [x1, x2, x3, x4]
        Bottomx = [x1, x2, x3, x4]
    else:
        # For CW < 1400, there are 3 holes
        x1 = firstXgap
        x2 = (inputCW - firstXgap * 2) / 2 + firstXgap
        x3 = inputCW - firstXgap
        Topholex = [x1, x2, x3]
        Bottomx = [x1, x2, x3]

    # Create a dictionary to hold the coordinates
    lc_positions = {
        'Topholex': Topholex,
        'Topholey': Topholey,
        'Bottomx': Bottomx,
        'Bottomy': Bottomy,
        'end_digit' : end_digit
    }

    # print (f" calculate_lc_position_035 함수 end_digit : {end_digit}, side_space : {side_space}, middle_space : {middle_space}")
    # print (f" {lc_positions}")

    return lc_positions

def calculate_hole_vertical_lc035(LCD):
    # 중심 지점 계산
    center = LCD // 2

    # 남은 거리를 분할하는 내부 함수
    def divide_remaining_distance(remaining_distance):        
        for div in range(3, 5):
            intervals = [remaining_distance // div] * (div - 1)
            intervals.append(remaining_distance - sum(intervals))
            if all(150 <= x <= 500 for x in intervals):
                return intervals
        return []

    # 중심으로부터의 고정 홀 위치
    fixed_holes = [54, center - 15, center + 15, LCD - 54]

    # LCD 값에 따른 추가 홀 위치 계산
    if LCD < 1950:
        # LCD가 1950 미만일 경우 중간 값 추가
        additional_holes = [(54 + (center - 15)) // 2]
    else:
        # LCD가 1950 이상일 경우 3등분한 값 추가
        first_third = 54 + (center - 15 - 54) // 3
        second_third = first_third + (center - 15 - 54) // 3
        additional_holes = [first_third, second_third]

    # 중심 왼쪽 홀 위치 결합 및 정렬
    left_holes = sorted(fixed_holes[:2] + additional_holes)

    # 중심을 기준으로 오른쪽 홀 위치 계산
    right_holes = [LCD - hole for hole in reversed(left_holes)]

    # 전체 홀 위치 결합 및 정렬
    return sorted(set(left_holes + right_holes))

def calculate_lc_vertical_hole_positions_035(inputCW, inputCD):
    # 고정된 x 좌표들 설정
    firstXgap = 110
    leftholex = [firstXgap, firstXgap, firstXgap]
    rightholex = [inputCW - firstXgap, inputCW - firstXgap, inputCW - firstXgap]

    side_space = 0
    space_available = 0

    # print (f"input CD {inputCD/2}")

    # y 좌표를 계산하기 위한 로직
    if inputCD <= 1350:
        # CD가 1250 이하일 경우, y1만 존재
        y1 = inputCD / 2   # 40*2를 제외하고 중간값을 계산
        vertical_holey = [y1]
    elif inputCD > 1350 and inputCD <2100:
        # CD가 1250을 초과할 경우, y 좌표 3개를 계산
        divided_space = (inputCD - 80) // 3  # 40*2를 제외하고 3등분
        end_digit = divided_space % 10

        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution        
        space_available = inputCD - 40 * 2
        
        if end_digit in [3, 4]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 4
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 3   
            # print (f" 3, 4 선택됨 end_digit : {end_digit}, side_space : {side_space}, middle_space : {middle_space}")                 
        elif end_digit in [6, 7]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 3) // 10 * 10) + 6
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 2 // 10 * 10) + 7
        else:            
            middle_space = ((space_available // 3) // 10 * 10)             
            side_space = ((space_available - middle_space) // 2 // 10 * 10) 

        # Calculating all x-coordinates
        y1 = side_space + 40    # 40위에서 시작
        y2 = y1 + middle_space

        vertical_holey = [y1, y2]
        # print (f" end_digit : {end_digit}, side_space : {side_space}, middle_space : {middle_space}")
    # 2100 보다 크면 중간 홀 3개    
    else:         
        divided_space = (inputCD - 80) // 4  # 40*2를 제외하고 4등분
        end_digit = divided_space % 10

        # For CW >= 1400, there are 4 holes
        # Calculate space between holes with special distribution        
        space_available = inputCD - 40 * 2
        
        if end_digit in [3, 4]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 4) // 10 * 10) + 4
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 3 // 10 * 10) + 3                    
        elif end_digit in [6, 7]:
            # The middle space needs to end with 4, therefore it should be the nearest multiple of 10 plus 4
            middle_space = ((space_available // 4) // 10 * 10) + 6
            # The side spaces will take the remaining space divided by 2 and adjusted to end with 3
            side_space = ((space_available - middle_space) // 3 // 10 * 10) + 7
        else:            
            middle_space = space_available // 4
            side_space = (space_available - middle_space) // 3

        # Calculating all x-coordinates
        y1 = side_space + 40
        y2 = y1 + middle_space
        y3 = y2 + side_space

        vertical_holey = [y1, y2, y3]

    # 좌표 딕셔너리 반환
    hole_positions = {
        'leftholex': leftholex,
        'rightholex': rightholex,
        'vertical_holey': vertical_holey
    }

    return hole_positions

def calculate_cliphole_lc035(distance, length):
    # 중심 지점 계산
    center = length // 2

    # LCD 길이에 따른 홀의 수와 위치 결정
    if length < 1950:
        # LCD가 1950 미만일 때는 3개의 홀
        hole_positions = [distance, center, length - distance]
    else:
        # LCD가 1950 이상일 때는 4개의 홀
        remaining_center = length - 2 * distance  # 양쪽 끝 홀 사이의 거리
        center_hole1 = distance + remaining_center // 4  # 중앙 왼쪽 홀
        center_hole2 = distance + (remaining_center // 4) * 3  # 중앙 오른쪽 홀
        hole_positions = [distance, center_hole1, center_hole2, length - distance]

    return sorted(hole_positions)

def calculate_midplatehole_lc035(length):
    # 전체 길이를 5등분하는 간격 계산
    interval = length / 5

    # 4개의 위치 좌표 계산
    hole_positions = [round(interval * i, 1) for i in range(1, 5)]

    return hole_positions
def calculate_midplate_connecthole_lc035(length):
    # 양 끝 홀 위치를 고정
    first_and_last_hole = 82.6

    # 중간 홀들 간의 거리 계산
    middle_length = length - 2 * first_and_last_hole  # 양쪽 끝 홀을 제외한 길이
    interval = round(middle_length / 7, 1)  # 6등분한 간격, 소수점 첫째 자리까지

    # 홀 위치 계산
    hole_positions = [first_and_last_hole]  # 첫 홀 위치 추가
    current_position = first_and_last_hole

    # 중간 홀 위치 계산 및 추가
    for _ in range(7):
        current_position += interval
        hole_positions.append(round(current_position, 1))

    # 마지막 홀 위치 추가
    hole_positions.append(length - first_and_last_hole)

    return hole_positions

#035 led bar bracket 홀간격계산
def calculate_ledbar_bottomhole_lc035(ledbar_length):
    center = ledbar_length // 2  # 중심 지점
    central_gap = 350
    interval = 300  # 간격

    left_positions = []
    right_positions = []

    # 중앙 구간 시작 및 종료
    central_start = center - central_gap // 2
    central_end = center + central_gap // 2

    # 중심으로부터 왼쪽 계산
    left_current = central_start
    while left_current > 0:
        left_positions.insert(0, left_current)
        left_current -= interval

    # 중심으로부터 오른쪽 계산
    right_current = central_end
    while right_current < ledbar_length:
        right_positions.append(right_current)
        right_current += interval

    # 결과 리스트 합치기
    positions = left_positions + right_positions

    return positions

def calculate_ledbar_upperhole_lc035(ledbar_length):
    center = ledbar_length // 2  # 중심 지점
    initial_offset = 50  # 최초 간격
    interval = 100  # 이후 간격

    left_positions = []
    right_positions = []

    # 중심으로부터 왼쪽 계산 (초기 간격으로 시작)
    left_current = center - initial_offset
    while left_current > 0:
        left_positions.insert(0, left_current)
        left_current -= interval

    # 중심으로부터 오른쪽 계산 (초기 간격으로 시작)
    right_current = center + initial_offset
    while right_current < ledbar_length:
        right_positions.append(right_current)
        right_current += interval

    # 결과 리스트 합치기
    positions = left_positions + right_positions

    return positions

def layout(doc):
    global HWW, HWD, wall_thickness, OP, CW, CD, CWW, CHW, left_clearance, rear_clearance, carpanel_thickness  # 전역 변수 사용을 위해 global 선언
    
    abs_x = 0
    abs_y = 0

    rx1 = abs_x + 5000
    ry1 = abs_y + 2000
    x1, y1 = rx1, ry1
    
    # HWW, HWD가 정의되지 않았을 수 있으므로 체크
    if HWW is None or HWD is None:
        print("HWW 또는 HWD가 정의되지 않았습니다.")
        return
    
    x2, y2 = x1 + HWW, y1 + HWD

    # OP + 좌우 150씩 추가해서 앞부분 자르기
    cut_width = OP + 300  # OP + 좌우 150씩 (150 * 2 = 300)
    cut_x1 = x1 + (HWW - cut_width) / 2  # 중앙 정렬
    cut_x2 = cut_x1 + cut_width

    # 외부 사각형 좌표 계산 (벽 두께만큼 바깥쪽으로)
    outer_x1 = x1 - wall_thickness
    outer_y1 = y1 - wall_thickness
    outer_x2 = x2 + wall_thickness
    outer_y2 = y2 + wall_thickness
    
    # 외부 테두리 (호이스트웨이 외벽) - 앞부분 자른 형태
    # 상단 선 (앞부분 자름)
    line(doc, outer_x1, outer_y1, cut_x1, outer_y1, layer='0')  # 왼쪽 부분
    line(doc, cut_x2, outer_y1, outer_x2, outer_y1, layer='0')  # 오른쪽 부분
    # 우측 선
    line(doc, outer_x2, outer_y1, outer_x2, outer_y2, layer='0')
    # 하단 선
    line(doc, outer_x2, outer_y2, outer_x1, outer_y2, layer='0')
    # 좌측 선
    line(doc, outer_x1, outer_y2, outer_x1, outer_y1, layer='0')
    
    # 자른 부분의 좌우 연결선 (바깥선과 안쪽선 연결)
    line(doc, cut_x1, outer_y1, cut_x1, y1, layer='0')  # 왼쪽 연결선
    line(doc, cut_x2, outer_y1, cut_x2, y1, layer='0')  # 오른쪽 연결선
    
    # 내부 사각형 그리기 (호이스트웨이 내부 공간) - 앞부분 자른 형태
    # 상단 선 (앞부분 자름)
    line(doc, x1, y1, cut_x1, y1, layer='0')  # 왼쪽 부분
    line(doc, cut_x2, y1, x2, y1, layer='0')  # 오른쪽 부분
    # 우측 선
    line(doc, x2, y1, x2, y2, layer='0')
    # 하단 선
    line(doc, x2, y2, x1, y2, layer='0')
    # 좌측 선
    line(doc, x1, y2, x1, y1, layer='0')
    
    # 엘리베이터 카 사각형 그리기 (CW, CD는 내부 사이즈)
    # 카를 Rear clearance와 LEFT clearance 값을 이용해서 배치
    
    # 카 외곽선 좌표 (내부 사이즈 + 패널 두께)
    car_outer_width = CW + (carpanel_thickness * 2)
    car_outer_depth = CD + (carpanel_thickness * 2)
    
    car_x1 = x1 + left_clearance
    car_y2 = y2 - rear_clearance  # 뒤쪽(y2)에서 rear_clearance만큼 앞으로
    car_x2 = car_x1 + car_outer_width  # 외곽선 크기
    car_y1 = car_y2 - car_outer_depth  # 외곽선 크기
    
    # 카 내부 사각형 좌표 (실제 CW, CD 크기)
    inner_car_x1 = car_x1 + carpanel_thickness
    inner_car_y1 = car_y1 + carpanel_thickness
    inner_car_x2 = inner_car_x1 + CW  # 내부는 정확히 CW 크기
    inner_car_y2 = inner_car_y1 + CD  # 내부는 정확히 CD 크기
    
    print(f"carpanel_thickness 값: {carpanel_thickness}")
    print(f"카 외곽선: ({car_x1}, {car_y1}) ~ ({car_x2}, {car_y2})")
    print(f"카 내부선: ({inner_car_x1}, {inner_car_y1}) ~ ({inner_car_x2}, {inner_car_y2})")
    
    # 엘리베이터 카 외곽선 그리기 (CW + 패널두께*2, CD + 패널두께*2)
    # line(doc, car_x1, car_y1, car_x2, car_y1, layer='0')
    line(doc, car_x2, car_y1 - 55, car_x2, car_y2, layer='0')
    line(doc, car_x2, car_y2, car_x1, car_y2, layer='0')
    line(doc, car_x1, car_y2, car_x1, car_y1 - 55, layer='0')
    
    # 엘리베이터 카 내부 사각형 그리기 (실제 CW, CD 크기)
    line(doc, inner_car_x1, inner_car_y1, inner_car_x1 + (CW - OP)/2, inner_car_y1, layer='0')
    line(doc, inner_car_x2 - (CW - OP)/2, inner_car_y1, inner_car_x2, inner_car_y1, layer='0')
    line(doc, inner_car_x2, inner_car_y1, inner_car_x2, inner_car_y2, layer='0')
    line(doc, inner_car_x2, inner_car_y2, inner_car_x1, inner_car_y2, layer='0')
    line(doc, inner_car_x1, inner_car_y2, inner_car_x1, inner_car_y1, layer='0')
    
    line(doc, inner_car_x1 + (CW - OP)/2, inner_car_y1 - 80, inner_car_x1 + (CW - OP)/2, inner_car_y1, layer='0')
    line(doc, inner_car_x2 - (CW - OP)/2, inner_car_y1 - 80, inner_car_x2 - (CW - OP)/2, inner_car_y1, layer='0')
    
    
    # 치수선 그리기
    dim_offset = 50  # 치수선과 도형 사이의 거리
    
    # HWW 치수선 (호이스트웨이 폭) - 상단
    dim(doc, x1, y1, x2, y1, dim_offset + 360, direction="down", text=f"HWW {HWW}")
    
    # HWD 치수선 (호이스트웨이 깊이) - 우측
    dim(doc, x1, y1, x1, y2, dim_offset + 300, direction="left", option='reverse', text=f"HWD {HWD}")
    
    # CW 치수선 (카 내부 폭) - 카 내부선 상단
    dim(doc, inner_car_x1, inner_car_y2, inner_car_x2, inner_car_y2, dim_offset + 330, direction="up", text=f"CW {CW}")
    
    # CD 치수선 (카 내부 깊이) - 카 내부선 우측
    dim(doc, inner_car_x1, inner_car_y1, inner_car_x1, inner_car_y2, dim_offset + 600, direction="left", option='reverse', text=f"CD {CD}")
    
    # OP 치수선 (개구부) - 카의 실제 위치에 맞게 수정
    op_dim_x1 = inner_car_x1 + (CW - OP)/2  # 카 내부에서 OP 시작점
    op_dim_x2 = inner_car_x1 + (CW + OP)/2  # 카 내부에서 OP 끝점
    dim(doc, op_dim_x1, y1, op_dim_x2, y1, dim_offset + 300, direction="down", text=f"OP {OP}")
    
    # 플랫폼 카도어 추가 (길이: OP*2+100, 높이: 60)
    platform_car_door_width = OP * 2 + 100
    platform_car_door_height = 60
    platform_car_door_x1 = inner_car_x1 + (CW - platform_car_door_width) / 2  # 카 중앙에 정렬
    platform_car_door_y1 = inner_car_y1 - platform_car_door_height - 80  # 카 앞쪽으로 80mm 간격
    platform_car_door_x2 = platform_car_door_x1 + platform_car_door_width
    platform_car_door_y2 = platform_car_door_y1 + platform_car_door_height
    
    # 플랫폼 카도어 그리기
    line(doc, platform_car_door_x1, platform_car_door_y1, platform_car_door_x2, platform_car_door_y1, layer='0')
    line(doc, platform_car_door_x2, platform_car_door_y1, platform_car_door_x2, platform_car_door_y2, layer='0')
    line(doc, platform_car_door_x2, platform_car_door_y2, platform_car_door_x1, platform_car_door_y2, layer='0')
    line(doc, platform_car_door_x1, platform_car_door_y2, platform_car_door_x1, platform_car_door_y1, layer='0')
    
    # 플랫폼 홀도어 추가 (플랫폼 카도어 앞에 30mm 간격)
    platform_hall_door_width = OP * 2 - 10  # 플랫폼 카도어와 같은 길이
    platform_hall_door_height = 60  # 플랫폼 카도어와 같은 높이
    platform_hall_door_x1 = inner_car_x1 + (CW - platform_hall_door_width) / 2  # 카 중앙에 정렬
    platform_hall_door_y1 = platform_car_door_y1 - platform_hall_door_height - 30  # 플랫폼 카도어 앞에 30mm 간격
    platform_hall_door_x2 = platform_hall_door_x1 + platform_hall_door_width
    platform_hall_door_y2 = platform_hall_door_y1 + platform_hall_door_height
    
    # 플랫폼 홀도어 그리기
    line(doc, platform_hall_door_x1, platform_hall_door_y1, platform_hall_door_x2, platform_hall_door_y1, layer='0')
    line(doc, platform_hall_door_x2, platform_hall_door_y1, platform_hall_door_x2, platform_hall_door_y2, layer='0')
    line(doc, platform_hall_door_x2, platform_hall_door_y2, platform_hall_door_x1, platform_hall_door_y2, layer='0')
    line(doc, platform_hall_door_x1, platform_hall_door_y2, platform_hall_door_x1, platform_hall_door_y1, layer='0')
    
    insert_block(inner_car_x1 + (CW - OP)/2, inner_car_y1 - 230, "jamb_left")
    insert_block(inner_car_x2 - (CW - OP)/2, inner_car_y1 - 230, "jamb_right")
    
    
    # CW 치수선 (카 내부 폭) - 카 내부선 상단
    dim(doc, inner_car_x1, inner_car_y2, inner_car_x2, inner_car_y2, dim_offset + 330, direction="up", text=f"CW {CW}")
    
    # CD 치수선 (카 내부 깊이) - 카 내부선 우측
    dim(doc, inner_car_x1, inner_car_y1, inner_car_x1, inner_car_y2, dim_offset + 600, direction="left", option='reverse', text=f"CD {CD}")

    HWW = int(HWW)
    HWD = int(HWD)
    wall_thickness = int(wall_thickness)
    OP = int(OP)
    CW = int(CW)
    CD = int(CD)

    print(f"x1: {0} y1: {0}")
    print(f"HWW: {HWW} HWD: {HWD}")
    print(f"벽 두께: {wall_thickness} mm")
    print(f"OP: {OP} mm (앞부분 자름: {OP + 300} mm)")
    print(f"내부 공간 크기: {HWW + wall_thickness * 2} x {HWD + wall_thickness * 2}")
    print("CW:", CW, "CD:", CD)
    print("카 위치:", "(", car_x1, ",", car_y1, ") ~ (", car_x2, ",", car_y2, ")")
    x2 = x1 + HWW
    y2 = y1 + HWD
    print("x2:", x2, "y2:", y2)

@Gooey(encoding='utf-8', program_name='미래기업 레이아웃 자동작도', tabbed_groups=True, navigation='Tabbed', show_success_modal=False, default_size=(800, 600))
def main():
    global args  #수정할때는 global 선언이 필요하다. 단순히 읽기만 할때는 필요없다.    
    global start_time
    global workplace, drawdate, deadlinedate, lctype, secondord, text_style, exit_program, Vcut_rate, Vcut, program_message, formatted_date, text_style_name
    global CW, CD, CH, HWW, HWD, HH, wall_thickness, OP, CWW, CHW, rear_clearance, left_clearance, carpanel_thickness
    
    # 전역 변수 초기화
    workplace = ""
    drawdate = None
    deadlinedate = None
    lctype = ""
    secondord = ""
    exit_program = False
    Vcut_rate = 0
    Vcut = 0
    program_message = ""
    formatted_date = ""
    text_style_name = 'JKW'

    # 현재 날짜와 시간을 가져옵니다.
    current_datetime = datetime.now()

    # 2023-12-10 원하는 형식으로 날짜를 문자열로 변환합니다.
    formatted_date = current_datetime.strftime('%Y-%m-%d')

    # 현재 날짜와 시간을 '년월일_시분초' 형식으로 가져옴
    # current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    current_time = datetime.now().strftime("%H%M%S")

    # 찾은 .xlsm 파일 목록 순회
    for file_path in xlsm_files:

        # 엑셀 파일 열기
        # file_path = xlsm_files[0]
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        sheet_name = 'Sheet1'  # 원하는 시트명으로 변경 기본정보가 있는 시트
        sheet = workbook[sheet_name]

        # DXF 파일 다시 로드 (전역 변수 덮어쓰기)
        global doc, msp
        doc = ezdxf.readfile(os.path.join(application_path, '', 'layout.dxf'))
        msp = doc.modelspace()
        
        # 엑셀에서 A열에서 텍스트를 찾아 해당 행의 C열 값을 가져오는 함수
        def find_value_in_column_A(sheet, search_text):
            """A열에서 특정 텍스트를 찾아 해당 행의 C열 값을 반환"""
            for row in range(1, sheet.max_row + 1):
                cell_a = sheet.cell(row=row, column=1)  # A열
                if cell_a.value and str(cell_a.value).strip().upper() == search_text.upper():
                    # 같은 행의 C열 값 반환
                    cell_c = sheet.cell(row=row, column=3)
                    value = cell_c.value
                    if value is not None:
                        try:
                            return float(value)
                        except (ValueError, TypeError):
                            return value
                    return None
            return None
        
        # A열에서 텍스트를 찾아 C열 값 가져오기
        CW = find_value_in_column_A(sheet, "CW")
        CD = find_value_in_column_A(sheet, "CD")
        CH = find_value_in_column_A(sheet, "CH")
        HWW = find_value_in_column_A(sheet, "HWW")
        HWD = find_value_in_column_A(sheet, "HWD")
        HH = find_value_in_column_A(sheet, "HH")
        wall_thickness = find_value_in_column_A(sheet, "Wall thickness")
        CWW = find_value_in_column_A(sheet, "CWW")
        CHW = find_value_in_column_A(sheet, "CHW")
        OP = find_value_in_column_A(sheet, "OP")
        carpanel_thickness = find_value_in_column_A(sheet, "carpanel_thickness")
        
        # clearance 값들을 정확한 텍스트로 찾기
        rear_clearance = find_value_in_column_A(sheet, "rear_clearance")
        left_clearance = find_value_in_column_A(sheet, "left_clearance")
        
        # None 값 체크 및 기본값 설정
        if CW is None:
            CW = 0
        if CD is None:
            CD = 0
        if CH is None:
            CH = 0
        if HWW is None:
            HWW = 0
        if HWD is None:
            HWD = 0
        if HH is None:
            HH = 0
        if wall_thickness is None:
            wall_thickness = 200  # 기본값 200mm
        if OP is None:
            OP = 0  # 기본값 0
        if CWW is None:
            CWW = 0  # 기본값 0
        if CHW is None:
            CHW = 0  # 기본값 0
        if carpanel_thickness is None:
            carpanel_thickness = 25  # 기본값 25mm
        if rear_clearance is None:
            rear_clearance = 150  # 기본값 150mm
        if left_clearance is None:
            left_clearance = 100  # 기본값 100mm
            
        # 최종 값 확인
        print(f"최종 carpanel_thickness: {carpanel_thickness}")
    

        # A열, B열의 1행부터 마지막행까지 데이터를 배열로 가져오기
        max_row = sheet.max_row
        
                
        def draw_table(doc, start_x, start_y, col_widths, row_height, data, text_height=200, layer="0"):
            msp = doc.modelspace()
            
            rows = len(data)       # 행 개수
            cols = len(col_widths) # 열 개수

            # ---- 텍스트 ----
            for r in range(rows):
                row_data = data[r]
                y = start_y - r * row_height
                
                for c in range(cols):
                    x = start_x + sum(col_widths[:c])
                    value = row_data[c] if c < len(row_data) else ""
                    
                    # 텍스트 추가
                    msp.add_mtext(
                        str(value),
                        dxfattribs={
                            "char_height": text_height,
                            "style": "JKW",
                            "layer": layer,
                        }
                    ).set_location((x + 20, y - row_height + 20))

            # ---- 가로줄 ----
            for r in range(rows + 1):  # 행 개수 + 마지막 줄
                y = start_y - r * row_height
                msp.add_line(
                    (start_x, y),
                    (start_x + sum(col_widths), y),
                    dxfattribs={"layer": layer}
                )

            # ---- 세로줄 ----
            x = start_x
            msp.add_line((x, start_y), (x, start_y - rows * row_height), dxfattribs={"layer": layer})
            for c in range(cols):
                x += col_widths[c]
                msp.add_line((x, start_y), (x, start_y - rows * row_height), dxfattribs={"layer": layer})


        # 엑셀에서 A, B열 데이터 읽기
        data = []
        for row in range(1, max_row + 1):
            a_val = sheet.cell(row=row, column=2).value
            b_val = sheet.cell(row=row, column=3).value
            data.append([a_val, b_val])

        # 표 그리기
        draw_table(
            doc,
            start_x=100, start_y=5000,
            col_widths=[1300, 1300],  # A, B열 폭
            row_height=25,          # 행 높이
            data=data,
            text_height=12,
            layer="0"
        )

        doc.saveas("excel_table.dxf")

        # layout 함수 호출 - 선 그리기
        layout(doc)

        # 작도일자 추출
        # drawdate가 None이 아닌 경우에만 처리
        if drawdate is not None:
            drawdate_str = "{0}년 {1:02d}월 {2:02d}일".format(drawdate.year, drawdate.month, drawdate.day)
            drawdate_str_short = "{0:02d}/{1:02d}".format(drawdate.month, drawdate.day)
        else:
            drawdate_str = ""
            drawdate_str_short = ""
            print("drawdate is not a valid datetime object")

        # deadlinedate가 None이 아닌 경우에만 처리
        if deadlinedate is not None:
            deadlinedate_str = "{0}년 {1:02d}월 {2:02d}일".format(deadlinedate.year, deadlinedate.month, deadlinedate.day)
            deadlinedate_str_short = "{0:02d}/{1:02d}".format(deadlinedate.month, deadlinedate.day)
        else:
            deadlinedate_str = ""
            deadlinedate_str_short = ""
            print("deadlinedate is not a valid datetime object")


        # 파일 이름에 사용할 수 없는 문자 정의
        invalid_chars = '<>:"/\\|?*'
        # 정규식을 사용하여 유효하지 않은 문자 제거
        cleaned_file_name = re.sub(f'[{re.escape(invalid_chars)}]', '', f"{workplace}_{lctype}_CW{str(CW)}xCD{str(CD)}_{current_time}")
        # 결과 파일 이름
        file_name = f"{cleaned_file_name}.dxf"

        exit_program = False


        program_message = \
            '''
        프로그램 실행결과입니다.
        -------------------------------------
        {0}
        -------------------------------------
        이용해 주셔서 감사합니다.
        '''    
        
        args = parse_arguments()    
        # 숫자로 변수를 선언해 준다.
        
        # 프로그램 시작 시간 기록
        start_time = time.time()

        # display_message()    
        end_time = time.time()

        # 실행 시간 계산
        execution_time = end_time - start_time

        # 시간 형식으로 변환
        hours, remainder = divmod(execution_time, 3600)
        minutes, seconds = divmod(remainder, 60)

        # 결과를 포맷팅하는 부분 수정
        parts = []
        if hours:
            parts.append(f"{int(hours)}h")
        if minutes:
            parts.append(f"{int(minutes)}m")
        parts.append(f"{int(seconds) + 3 }초")  # 초를 '초'로 표시

        formatted_execution_time = " ".join(parts)

        log_login()
        print(f"실행 시간: {formatted_execution_time}")

        # 생성된 파일 이름으로 DXF 파일 저장
        # file_name =f"c:\\python\\{file_name}" 
        file_name =f"c:\\python\\mirae_layout\\{file_name}" 
        doc.saveas(file_name)
        print(f"파일이 '{file_name}'로 저장 완료!")              

if __name__ == '__main__':
    main()
    sys.exit()