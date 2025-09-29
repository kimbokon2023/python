## 새로운 로직개발 헤밍구조에 대한 통합버전 개발 쪽쟘, 멍텅구리 통합
## 실전 사용하면서 수정사항 반영 B1, B2 크기 조정 pyinstaller 실행 후 반영되는 값 적용
## 2024년 6월 3일 직각쟘에 대한 각도 표시 요청(이미래과장)
## 2024/07/09 성우빌딩 SO 와이드 좌측 상하 각도 이상함
## 2024/08/10 와이드쟘 마이너스 각도 -1도까지 비교하고 만들기
## 10월 22일 김진섭 현장 코드 수정

import math
import ezdxf
from ezdxf.enums import TextEntityAlignment
from datetime import datetime
import openpyxl
import os
import glob
import time
import os
import sys
import io
from datetime import datetime

import json
from gooey import Gooey, GooeyParser
import warnings
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')    

# 경고 메시지 필터링
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")

# 현재 날짜와 시간을 가져옵니다.
current_datetime = datetime.now()

# 2023-12-10 원하는 형식으로 날짜를 문자열로 변환합니다.
formatted_date = current_datetime.strftime('%Y-%m-%d')

# # UserWarning 경고 무시 설정
# warnings.filterwarnings("ignore", category=UserWarning)
# # 모든 경고 무시
# warnings.filterwarnings("ignore")

# 전역 변수로 선언
saved_Xpos = 0
saved_Ypos = 0
start_time = 0
HPI_count = 0

# fromNum = 1
# toNum = 30

# for i in range(fromNum, toNum + 1):
#     exec(f"X{i} = 0")
#     exec(f"Y{i} = 0")

# 현재 날짜와 시간을 '년월일_시분초' 형식으로 가져옴
current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

# 폴더 내의 모든 .xlsm 파일을 검색
application_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
excel_saved_file = os.path.join(application_path, 'excel')
xlsm_files = glob.glob(os.path.join(excel_saved_file, '*.xlsm'))
jamb_ini = os.path.join(application_path, 'data', 'jamb.json')
license_file_path = os.path.join(application_path, 'data', 'settings.json')

# 찾은 .xlsm 파일 목록 출력
for xlsm_file in xlsm_files:
    print(xlsm_file)

# 엑셀 파일 열기
file_path = xlsm_files[0]
workbook = openpyxl.load_workbook(file_path, data_only=True)

sheet_name = '정보'  # 원하는 시트명으로 변경
sheet = workbook[sheet_name]

# 5, 6, 7, 8, 9행의 2열 값을 직접 가져오기
workplace = sheet.cell(row=1, column=2).value
HPI_type = sheet.cell(row=5, column=2).value
worker = sheet.cell(row=6, column=2).value
widejamb_material = sheet.cell(row=7, column=2).value
normal_material = sheet.cell(row=8, column=2).value
narrow_material = sheet.cell(row=9, column=2).value
# print("와이드 잼브 재료: " +  widejamb_material)
# print("노멀 재료: " +  normal_material)
# print("나로우 재료: " +  narrow_material)

# 파일 이름에 사용할 수 없는 문자 정의
invalid_chars = '<>:"/\\|?*'
# 정규식을 사용하여 유효하지 않은 문자 제거
cleaned_file_name = re.sub(f'[{re.escape(invalid_chars)}]', '', f"{workplace}_{current_time}")
# 결과 파일 이름
file_name = f"{cleaned_file_name}.dxf"

# 엑셀 파일 닫기
workbook.close()

# 엑셀 파일 열기
file_path = xlsm_files[0]
workbook = openpyxl.load_workbook(file_path, data_only=True)

# 시트 선택 (시트명을 지정)
sheet_name = '쪽쟘제작'  # 원하는 시트명으로 변경
sheet = workbook[sheet_name]

# 2차원 배열을 저장할 리스트 생성
data_2d = []

# 새로운 DXF 파일 생성
# 새로운 DXF 문서 생성
# doc = ezdxf.new()

# DXF 파일 로드
doc = ezdxf.readfile(os.path.join(application_path, 'dimstyle', 'dimstyle4.dxf'))

# DIMSTYLE 정의 (새로운 DIMSTYLE 생성 또는 기존 DIMSTYLE 수정)
if '미래' not in doc.dimstyles:
    dimstyle = doc.dimstyles.new('DimStyle')
else:
    dimstyle = doc.dimstyles.get('미래')

# TEXTSTYLE 정의
text_style_name = 'JKW'  # 원하는 텍스트 스타일 이름
if text_style_name not in doc.styles:
    text_style = doc.styles.new(
        name=text_style_name,
        dxfattribs={
            'font': 'Arial.ttf',  # TrueType 글꼴 파일명            
        }
    )
else:
    text_style = doc.styles.get(text_style_name)     

# 새로운 레이어 생성
doc.layers.new(name='bending', dxfattribs={'color': 3})
# 선형축척 1로 설정

# 레이어 이름 설정
layer_name = '22'

# 레이어가 존재하는지 확인하고, 없으면 새로 생성
if layer_name not in doc.layers:
    # 여기서 레이어의 추가적인 속성을 정의할 수 있습니다.
    bendingline_layer = doc.layers.new(name=layer_name, dxfattribs={'color': 8})
else:
    bendingline_layer = doc.layers.get(layer_name)

# 레이어의 색상을 8번으로 고정
bendingline_layer.dxf.color = 8
bendingline_layer.dxf.linetype = 'DASHED2'

msp = doc.modelspace()

# dim 스타일 정의
dim_style = doc.styles.new(
    name='MIRAE'  # 스타일 이름    
)

exit_program = False

# 1.2T vcut 기준자료
Bending_rate = 1.2
Vcut = True
Vcut_rate = Bending_rate

program_message = \
    '''
프로그램 실행결과입니다.
-------------------------------------
{0}
-------------------------------------
이용해 주셔서 감사합니다.
'''

def save_file(workplace):
    # 현재 시간 가져오기
    current_time = datetime.now().strftime("%Y%m%d%H%M%S")

    # 파일 이름에 사용할 수 없는 문자 정의
    invalid_chars = '<>:"/\\|?*'
    # 정규식을 사용하여 유효하지 않은 문자 제거
    cleaned_file_name = re.sub(f'[{re.escape(invalid_chars)}]', '', f"{workplace}_{current_time}")

    # 결과 파일이 저장될 디렉토리
    output_directory = "c:/python/jamb_ver10/jamb작업완료"

    # 디렉토리가 존재하지 않으면 생성
    os.makedirs(output_directory, exist_ok=True)

    # 결과 파일 이름
    file_name = f"{cleaned_file_name}.dxf"
    # 전체 파일 경로 생성
    full_file_path = os.path.join(output_directory, file_name)

    # 파일 경로 반환
    return full_file_path    

def display_message():
    message = program_message.format('\n'.join(sys.argv[1:])).split('\n')
    delay = 1.5 / len(message)

    for line in message:
        print(line)
        time.sleep(delay)

# 환경설정 가져오기(하드공유 번호)
def load_env_settings():
    try:
        with open(license_file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return data.get("DiskID")
    except FileNotFoundError:
        return None

def get_current_disk_id():
    return os.popen('wmic diskdrive get serialnumber').read().strip()

# None이면 0을 리턴하는 함수
def validate_or_default(value):
    if value is None:
        return 0
    return value

def find_intersection(start1, end1, start2, end2):
    # 교차점 계산 (두 직선이 수직 및 수평으로 만나는 경우에 대해서만)
    if start1[0] == end1[0]:
        return (start1[0], start2[1])
    else:
        return (start2[0], start1[1])

def calculate_fillet_point(center, point, radius):
    # 필렛 접점 계산
    dx = point[0] - center[0]
    dy = point[1] - center[1]
    if abs(dx) > abs(dy):
        return (center[0] + radius * (1 if dx > 0 else -1), center[1])
    else:
        return (center[0], center[1] + radius * (1 if dy > 0 else -1))

def calculate_angle(center, point):
    # 각도 계산
    return math.degrees(math.atan2(point[1] - center[1], point[0] - center[0]))

def add_90_degree_fillet(doc, start1, end1, start2, end2, radius):
    msp = doc.modelspace()

    # 교차점 찾기
    intersection_point = find_intersection(start1, end1, start2, end2)

    # 필렛 접점 계산
    point1 = calculate_fillet_point(intersection_point, end1, radius)
    point2 = calculate_fillet_point(intersection_point, end2, radius)

    # 각도 계산
    start_angle = calculate_angle(intersection_point, point1)
    end_angle = calculate_angle(intersection_point, point2)

    # 필렛 원호 그리기
    msp.add_arc(
        center=intersection_point,
        radius=radius,
        start_angle=start_angle,
        end_angle=end_angle,
        dxfattribs={'layer': '레이져'},
    )    
    return msp

def dim_leader_line(doc, start_x, start_y, end_x, end_y, text, layer=None, style='0', text_height=22):
    msp = doc.modelspace()

    # 지서선 추가
    leader = msp.add_leader(
        vertices=[(start_x, start_y), (end_x, end_y)],  # 시작점과 끝점
        dxfattribs={
            'dimstyle': layer,
            'layer': layer
        }
    )

    # 텍스트 추가 (선택적)
    if text:
        msp.add_mtext(text, dxfattribs={
            'insert': (end_x + 10, end_y + 5),  # 텍스트 위치 조정
            'layer': style,
            'char_height': text_height,
            'style': text_style_name,
            'attachment_point': 1  # 텍스트 정렬 방식 설정
        })

    return leader

def line(doc, x1, y1, x2, y2, layer=None):
    global saved_Xpos, saved_Ypos  # 전역 변수로 사용할 것임을 명시
    
    # 선 추가
    start_point = (x1, y1)
    end_point = (x2, y2)
    if layer:
        # 절곡선 22 layer는 ltscale을 조정한다
        if(layer=="22"):
            msp.add_line(start=start_point, end=end_point, dxfattribs={'layer': layer, 'ltscale' : 30})
        else:
            msp.add_line(start=start_point, end=end_point, dxfattribs={'layer': layer})
    else:
        msp.add_line(start=start_point, end=end_point)

    # 다음 선분의 시작점을 업데이트
    saved_Xpos = x2
    saved_Ypos = y2        

def lineto(doc, x, y, layer=None):
    global saved_Xpos, saved_Ypos  # 전역 변수로 사용할 것임을 명시
    
    # 현재 위치를 시작점으로 설정
    start_x = saved_Xpos
    start_y = saved_Ypos

    # 끝점 좌표 계산
    end_x = x
    end_y = y

    # 모델 공간 (2D)을 가져옴
    msp = doc.modelspace()

    # 선 추가
    start_point = (start_x, start_y)
    end_point = (end_x, end_y)
    if layer:
        msp.add_line(start=start_point, end=end_point, dxfattribs={'layer': layer})
    else:
        msp.add_line(start=start_point, end=end_point)

    # 다음 선분의 시작점을 업데이트
    saved_Xpos = end_x
    saved_Ypos = end_y

def lineclose(doc, start_index, end_index, layer='레이져'):    

    firstX, firstY = globals()[f'X{start_index}'], globals()[f'Y{start_index}']
    prev_x, prev_y = globals()[f'X{start_index}'], globals()[f'Y{start_index}']

    # start_index+1부터 end_index까지 반복합니다.
    for i in range(start_index + 1, end_index + 1):
        # 현재 인덱스의 좌표를 가져옵니다.
        curr_x, curr_y = globals()[f'X{i}'], globals()[f'Y{i}']

        # 이전 좌표에서 현재 좌표까지 선을 그립니다.
        line(doc, prev_x, prev_y, curr_x, curr_y, layer)

        # 이전 좌표 업데이트
        prev_x, prev_y = curr_x, curr_y

        print(f"prev_x {prev_x}" )
        print(f"prev_y {prev_y}" )
    
    # 마지막으로 첫번째 점과 연결
    line(doc, prev_x, prev_y, firstX , firstY, layer)        

def rectangle(doc, x1, y1, dx, dy, layer=None):
    # 네 개의 선분으로 직사각형 그리기
    line(doc, x1, y1, dx, y1, layer=layer)   
    line(doc, dx, y1, dx, dy, layer=layer)   
    line(doc, dx, dy, x1, dy, layer=layer)   
    line(doc, x1, dy, x1, y1, layer=layer)   

def add_angular_dim_2l(drawing, base, line1, line2, location=None, text=None, text_rotation=None, dimstyle='0', override=None, dxfattribs=None):
    # Create a new dimension line
    dim = drawing.dimstyle.add(
        'EZ_ANGULAR_2L',
        dxfattribs={
            'dimstyle': dimstyle,
            'dimtad': 1,  # Place text above dimension line
            'dimtih': False,  # Align text horizontally to dimension line
            'dimtoh': False,  # Align text outside horizontal
            'dimdec': 1
        }
    )
    
    # If override is provided, update the dimension style
    if override:
        dim.update(override)
    
    # Create angular dimension
    angular_dim = drawing.add_angular_dim_2l(
        base=base,
        line1=line1,
        line2=line2,
        location=location,
        text=text,
        text_rotation=text_rotation,
        dimstyle=dim.name,
        override=override,
        dxfattribs=dxfattribs,
    )
    
    # Render the dimension
    angular_dim.render()
    return angular_dim

def dim_angular(doc, x1, y1, x2, y2, x3, y3, x4, y4, distance=80, direction="left", dimstyle="mirae"):
    msp = doc.modelspace()

    # 두 선분의 좌표
    p1 = (x1, y1)
    p2 = (x2, y2)
    p3 = (x3, y3)
    p4 = (x4, y4)

    # 치수선의 위치 계산
    base_x = (p1[0] + p2[0] + p3[0] + p4[0]) / 4
    base_y = (p1[1] + p2[1] + p3[1] + p4[1]) / 4

    if direction == "left":
        base = (base_x - distance, base_y)
    elif direction == "up":
        base = (base_x, base_y + distance)
    elif direction == "down":
        base = (base_x, base_y - distance)
    else:
        base = (base_x + distance, base_y)

    # 벡터가 동일한지 확인
    if p1 == p2 or p3 == p4:
        print("동일한 점들로는 각도 치수를 생성할 수 없습니다.")
        return None

    # 각도 치수선 추가
    dimension = msp.add_angular_dim_2l(
        base=base,  # 치수선 위치
        line1=(p1, p2),  # 첫 번째 선의 시작점과 끝점
        line2=(p3, p4),  # 두 번째 선의 시작점과 끝점
        dimstyle=dimstyle,  # 치수 스타일
        override={'dimtxt': 0.27, 'dimgap': 0.02, 'dimscl': 1, 'dimlfac': 1, 'dimclrt': 7, 'dimdsep': 46, 'dimdec': 0}  # 선색 백색 소수점 . 표기
    )

    # 치수선의 기하학적 형태 생성
    dimension.render()
    return dimension


def dim_diameter(doc, center, radius, angle, dimstyle="JKW", override=None):
    msp = doc.modelspace()
    
    # 기본 지름 치수선 추가
    dimension = msp.add_diameter_dim(
        center=center,  # 원의 중심점
        radius=radius,  # 반지름
        angle=angle,    # 치수선 각도
        dimstyle=dimstyle,  # 치수 스타일
        override=override   # 추가 스타일 설정 (옵션)
    )
    
    # 치수선의 기하학적 형태 생성
    dimension.render()
    return dimension


def dim_linear(doc, x1, y1, x2, y2, textstr, dis, direction="up", layer=None, text_height=0.22,  text_gap=0.07):
    msp = doc.modelspace()
    dim_style = '미래'

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점이 있는지 확인
    if dimension_value % 1 != 0:
        dimdec = 1  # 소수점이 있는 경우 소수점 첫째 자리까지 표시
    else:
        dimdec = 1  # 소수점이 없는 경우 소수점 표시 없음

    # override 설정
    override_settings = {
        'dimtxt': text_height,
        'dimgap': text_gap,
        'dimscl': 1,
        'dimlfac': 1,
        'dimclrt': 7,
        'dimdsep': 46,
        'dimdec': dimdec
    }

    # 방향에 따른 치수선 추가
    if direction == "up":
        dimension = msp.add_linear_dim(
            base=(x1, y1 + dis),
            dimstyle=dim_style,
            p1=(x1, y1),
            p2=(x2, y2),
            dxfattribs={'layer': layer},
            override=override_settings
        )
    elif direction == "down":
        dimension = msp.add_linear_dim(
            dimstyle=dim_style,
            base=(x1, y1 - dis),
            p1=(x1, y1),
            p2=(x2, y2),
            dxfattribs={'layer': layer},
            override=override_settings
        )
    elif direction == "aligned":
        dimension = msp.add_aligned_dim(
            dimstyle=dim_style,
            p1=(x1, y1),
            p2=(x2, y2),
            distance=dis,
            dxfattribs={'layer': layer},
            override=override_settings
        )
    else:
        raise ValueError("Invalid direction. Use 'up', 'down', or 'aligned'.")

    dimension.render()
    return dimension

def create_vertical_dim(doc, x1, y1, x2, y2, dis, angle, layer='0', text_height=0.22, text_gap=0.07):
    msp = doc.modelspace()
    dim_style = "미래"  # 치수 스타일 이름
    points = [(x1, y1), (x2, y2)]

    # 치수값 계산
    dimension_value = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    # 소수점이 있는지 확인
    if dimension_value % 1 != 0:
        dimdec = 1  # 소수점이 있는 경우 소수점 첫째 자리까지 표시
    else:
        dimdec = 0  # 소수점이 없는 경우 소수점 표시 없음    

    return msp.add_multi_point_linear_dim(
        base=(x1 + dis if angle == 270 else x1 - dis , y1),  #40은 보정
        points=points,
        angle=angle,
        dimstyle=dim_style,
        discard=True,
        dxfattribs={'layer': layer},
        # 치수 문자 위치 조정 (0: 치수선 위, 1: 치수선 옆) 'dimtmove': 1 
        override={'dimtxt': text_height, 'dimgap': text_gap, 'dimscl': 1, 'dimlfac': 1, 'dimclrt': 7, 'dimdec': dimdec, 'dimdsep': 46, 'dimtmove': 0, 'dimtix': 1 }
    )

def dim_vertical_right(doc, x1, y1, x2, y2, dis, layer=None, text_height=0.22,  text_gap=0.07):
    return create_vertical_dim(doc, x1, y1, x2, y2, dis, 270, layer, text_height, text_gap)

def dim_vertical_left(doc, x1, y1, x2, y2, dis, layer=None, text_height=0.22,  text_gap=0.07):
    return create_vertical_dim(doc, x1, y1, x2, y2, dis, 90, layer, text_height, text_gap)

def draw_Text(doc, x, y, size, text, layer=None):
    # 모델 공간 (2D)을 가져옴
    msp = doc.modelspace()

    # 텍스트 추가 및 생성된 Text 객체 가져오기
    text_entity = msp.add_text(
        text,  # 텍스트 내용
        dxfattribs={
            'layer': layer,  # 레이어 지정
            'style': text_style_name,  # 텍스트 스타일 지정
            'height': size,  # 텍스트 높이 (크기) 지정
        }
    )

    # Text 객체의 위치 설정
    # text_entity.set_placement((x, y), align=TextEntityAlignment.CENTER)  # 텍스트 위치 및 정렬 설정
    # Text 객체의 위치 설정 (가로 및 세로 중앙 정렬)
    text_entity.set_placement((x, y), align=TextEntityAlignment.MIDDLE_LEFT)

def drawcircle(doc, center_x, center_y, radius, layer='0', color='7'):
    """
    DXF 문서에 원을 그리는 함수
    :param doc: ezdxf 문서 객체
    :param center_x: 원의 중심 x 좌표
    :param center_y: 원의 중심 y 좌표
    :param radius: 원의 반지름
    :param layer: 원을 추가할 레이어 이름 (기본값은 '0')
    """
    msp = doc.modelspace()
    circle = msp.add_circle(center=(center_x, center_y), radius=radius, dxfattribs={'layer': layer, 'color' : color})
    return circle

def extract_abs(a, b):
    return abs(a - b)    

#######################################################################################################################################################################################
#  HPI 없는 경우 문구 출력
#######################################################################################################################################################################################
def draw_NoHPI(centerBaseX, centerBaseY):
    Textstr = "HPI 무"       
    draw_Text(doc, centerBaseX - len(Textstr)*42/2, centerBaseY + 60, 42 , str(Textstr), '레이져')

#######################################################################################################################################################################################
#  HIP 작도
#######################################################################################################################################################################################
def draw_HPI(centerBaseX, centerBaseY, R, HPI_height):

    global HPI_count

    HPI_count += 1

    if(HPI_type=="TKEK HD S20D-OR"):
        X1 = centerBaseX - 130
        Y1 = centerBaseY + HPI_height + 40 - Vcut_rate
        X2 = centerBaseX + 130
        Y2 = Y1
        line(doc, X1, Y1, X2, Y2, layer='레이져')       
        X3 = X2
        Y3 = Y1 - 80
        lineto(doc, X3, Y3, layer='레이져')          
        X4 = X1
        Y4 = Y3
        lineto(doc, X4, Y4, layer='레이져')              
        lineto(doc, X1, Y1, layer='레이져')              
        # 대각선
        line(doc, X1, Y1, X3, Y3, layer='CL')       
        line(doc, X4, Y4, X2, Y2, layer='CL')       
        
        #3.4파이 홀 4개
        drawcircle(doc, X1 - 35, Y1 - 17,  1.7 , layer='레이져') # 도장홀 3.4파이
        drawcircle(doc, X2 + 35, Y1 - 17,  1.7 , layer='레이져') # 도장홀 3.4파이
        drawcircle(doc, X4 - 35, Y4 + 17,  1.7 , layer='레이져') # 도장홀 3.4파이
        drawcircle(doc, X3 + 35, Y3 + 17,  1.7 , layer='레이져') # 도장홀 3.4파이

        # HPI 상부 치수선 3개
        dim_linear(doc,  centerBaseX-R/2, Y1, X1, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X1, Y1, X2, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X2, Y1, centerBaseX+R/2, Y1, "", 35,  direction="up", layer='dim')
        # 3.4파이 4개
        dim_leader_line(doc, X1 - 35, Y1 - 17, X1 , Y1 + 110, "4-3.4%%c", layer="dim", style='0')                            

        # 좌측 홀간격 치수선
        dim_vertical_left(doc, X1 - 35, Y1 - 17,  X4 - 35, Y4 + 17,  60, 'dim', text_height=0.22, text_gap=0.07)
        dim_vertical_left(doc, X1 , Y1,  X4, Y4 ,  190, 'dim', text_height=0.22, text_gap=0.07)

        # 하부 홀간격 치수선
        dim_linear(doc,  X4 - 35, Y4 + 17, X3 + 35, Y3 + 17, "", 67,  direction="down", layer='dim')

        # 우측 HPI height 치수선
        dim_vertical_right(doc, X3 , centerBaseY + HPI_height - Vcut_rate,  X3 , centerBaseY ,  140, 'dim', text_height=0.22, text_gap=0.07)

    if(HPI_type=="TKEK HD S20D-OR(손상민 5파이타공)"):
        X1 = centerBaseX - 130
        Y1 = centerBaseY + HPI_height + 40 - Vcut_rate
        X2 = centerBaseX + 130
        Y2 = Y1
        line(doc, X1, Y1, X2, Y2, layer='레이져')       
        X3 = X2
        Y3 = Y1 - 80
        lineto(doc, X3, Y3, layer='레이져')          
        X4 = X1
        Y4 = Y3
        lineto(doc, X4, Y4, layer='레이져')              
        lineto(doc, X1, Y1, layer='레이져')              
        # 대각선
        line(doc, X1, Y1, X3, Y3, layer='CL')       
        line(doc, X4, Y4, X2, Y2, layer='CL')       
        
        #3.4파이 홀 4개
        drawcircle(doc, X1 - 35, Y1 - 17,  2.5 , layer='레이져') # 도장홀 3.4파이
        drawcircle(doc, X2 + 35, Y1 - 17,  2.5 , layer='레이져') # 도장홀 3.4파이
        drawcircle(doc, X4 - 35, Y4 + 17,  2.5 , layer='레이져') # 도장홀 3.4파이
        drawcircle(doc, X3 + 35, Y3 + 17,  2.5 , layer='레이져') # 도장홀 3.4파이

        # HPI 상부 치수선 3개
        dim_linear(doc,  centerBaseX-R/2, Y1, X1, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X1, Y1, X2, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X2, Y1, centerBaseX+R/2, Y1, "", 35,  direction="up", layer='dim')
        # 3.4파이 4개
        dim_leader_line(doc, X1 - 35, Y1 - 17, X1 , Y1 + 110, "4-5%%c", layer="dim", style='0')                                    

        # 좌측 홀간격 치수선
        dim_vertical_left(doc, X1 - 35, Y1 - 17,  X4 - 35, Y4 + 17,  60, 'dim', text_height=0.22, text_gap=0.07)
        dim_vertical_left(doc, X1 , Y1,  X4, Y4 ,  190, 'dim', text_height=0.22, text_gap=0.07)

        # 하부 홀간격 치수선
        dim_linear(doc,  X4 - 35, Y4 + 17, X3 + 35, Y3 + 17, "", 67,  direction="down", layer='dim')

        # 우측 HPI height 치수선
        dim_vertical_right(doc, X3 , centerBaseY + HPI_height - Vcut_rate,  X3 , centerBaseY ,  140, 'dim', text_height=0.22, text_gap=0.07)

    if(HPI_type=="OTIS HIX-A162" or HPI_type== "OTIS HIX-A164/165"):
        X1 = centerBaseX - 105
        Y1 = centerBaseY + HPI_height + 38 - Vcut_rate
        X2 = centerBaseX + 105
        Y2 = Y1
        line(doc, X1, Y1, X2, Y2, layer='레이져')       
        X3 = X2
        Y3 = Y1 - 76
        lineto(doc, X3, Y3, layer='레이져')          
        X4 = X1
        Y4 = Y3
        lineto(doc, X4, Y4, layer='레이져')              
        lineto(doc, X1, Y1, layer='레이져')              
        # 대각선
        line(doc, X1, Y1, X3, Y3, layer='CL')       
        line(doc, X4, Y4, X2, Y2, layer='CL')       

        # HPI 상부 치수선 3개
        dim_linear(doc,  centerBaseX-R/2, Y1, X1, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X1, Y1, X2, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X2, Y1, centerBaseX+R/2, Y1, "", 35,  direction="up", layer='dim')
      
        # 우측 HPI height 치수선
        dim_vertical_right(doc, X3 , centerBaseY + HPI_height - Vcut_rate,  X3 , centerBaseY ,  140, 'dim', text_height=0.22, text_gap=0.07)

    if(HPI_type=="OTIS HIX-A201 돌출형"):
        X1 = centerBaseX - 150
        Y1 = centerBaseY + HPI_height + 36 - Vcut_rate
        X2 = centerBaseX + 150
        Y2 = Y1
        line(doc, X1, Y1, X2, Y2, layer='레이져')       
        X3 = X2
        Y3 = Y1 - 72
        lineto(doc, X3, Y3, layer='레이져')          
        X4 = X1
        Y4 = Y3
        lineto(doc, X4, Y4, layer='레이져')              
        lineto(doc, X1, Y1, layer='레이져')              
        # 대각선
        line(doc, X1, Y1, X3, Y3, layer='CL')       
        line(doc, X4, Y4, X2, Y2, layer='CL')       
        
        #3.3파이 홀 3개
        drawcircle(doc, X1 - 17.5, Y1 - 1,  1.65 , layer='레이져') 
        drawcircle(doc, X4 - 17.5, Y4 + 1,  1.65 , layer='레이져') 
        drawcircle(doc, X3 + 17.5, Y3 + 36 ,  1.65 , layer='레이져') 

        # HPI 상부 치수선 3개
        dim_linear(doc,  centerBaseX-R/2, Y1, X1, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X1, Y1, X2, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X2, Y1, centerBaseX+R/2, Y1, "", 35,  direction="up", layer='dim')
        # 3.4파이 4개
        dim_leader_line(doc, X1 - 17.5, Y1 - 1, X1 , Y1 + 110, "3-3.3%%c", layer="dim", style='0')                            

        # 좌측 홀간격 치수선
        dim_vertical_left(doc, X1 - 17.5, Y1 - 1,  X4 - 17.5, Y4 + 1,  60, 'dim', text_height=0.22, text_gap=0.07)
        dim_vertical_left(doc, X1 , Y1,  X4, Y4 ,  190, 'dim', text_height=0.22, text_gap=0.07)

        # 하부 홀간격 치수선
        dim_linear(doc,  X4 - 17.5, Y4 + 1, X3 + 17.5, Y3 + 36, "", 67,  direction="down", layer='dim')

        # 우측 HPI height 치수선
        dim_vertical_right(doc, X3 , centerBaseY + HPI_height - Vcut_rate,  X3 , centerBaseY ,  140, 'dim', text_height=0.22, text_gap=0.07)

    if(HPI_type=="OTIS HIL-A206 돌출형"):
        X1 = centerBaseX - 165
        Y1 = centerBaseY + HPI_height + 30 - Vcut_rate
        X2 = centerBaseX + 165
        Y2 = Y1
        line(doc, X1, Y1, X2, Y2, layer='레이져')       
        X3 = X2
        Y3 = Y1 - 60
        lineto(doc, X3, Y3, layer='레이져')          
        X4 = X1
        Y4 = Y3
        lineto(doc, X4, Y4, layer='레이져')              
        lineto(doc, X1, Y1, layer='레이져')              
        # 대각선
        line(doc, X1, Y1, X3, Y3, layer='CL')       
        line(doc, X4, Y4, X2, Y2, layer='CL')       

        # HPI 상부 치수선 3개
        dim_linear(doc,  centerBaseX-R/2, Y1, X1, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X1, Y1, X2, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X2, Y1, centerBaseX+R/2, Y1, "", 35,  direction="up", layer='dim')
      
        # 우측 HPI height 치수선
        dim_vertical_right(doc, X3 , centerBaseY + HPI_height - Vcut_rate,  X3 , centerBaseY ,  140, 'dim', text_height=0.22, text_gap=0.07)

        #3.4파이 홀 4개
        drawcircle(doc, X1 - 22, Y1  ,  1.65 , layer='레이져') # 도장홀 3.4파이
        drawcircle(doc, X2 + 2, Y2 + 6.8,  1.65 , layer='레이져') # 도장홀 3.4파이
        drawcircle(doc, X4 - 22,  Y4 ,  1.65 , layer='레이져') # 도장홀 3.4파이
        drawcircle(doc, X3 + 2, Y3 - 6.8,  1.65 , layer='레이져') # 도장홀 3.4파이
        # 3.4파이 4개
        dim_leader_line(doc,  X1-22 , Y1,  X1 , Y1 + 110, "4-%%c3.3", layer="dim", style='0')                                    

        # 좌측 홀간격 치수선        
        dim_vertical_left(doc, X1-22 , Y1,  X4 - 22, Y4 ,  120, 'dim', text_height=0.22, text_gap=0.07)

        # 하부 홀간격 중심 치수선
        dim_linear(doc, X4 - 22, Y4 , centerBaseX , Y4  , "", 67,  direction="down", layer='dim')        
        dim_linear(doc,  centerBaseX , Y4 ,  X3 + 2, Y3 - 6.8, "", 67,  direction="down", layer='dim')     

    if(HPI_type=="OTIS HIL-A207 돌출형"):
        # 편심형 주의
        X1 = centerBaseX - 130
        Y1 = centerBaseY + HPI_height + 30 - Vcut_rate
        X2 = centerBaseX + 160
        Y2 = Y1
        line(doc, X1, Y1, X2, Y2, layer='레이져')       
        X3 = X2
        Y3 = Y1 - 60
        lineto(doc, X3, Y3, layer='레이져')          
        X4 = X1
        Y4 = Y3
        lineto(doc, X4, Y4, layer='레이져')              
        lineto(doc, X1, Y1, layer='레이져')              
        # 대각선
        line(doc, X1, Y1, X3, Y3, layer='CL')       
        line(doc, X4, Y4, X2, Y2, layer='CL')       

        #3.3파이 홀 5개 (3개에서 5개로 늘어남)
        HoleX1 = X1 - 86
        HoleY1 = Y1 + 10
        drawcircle(doc, HoleX1 , HoleY1 ,  1.65 , layer='레이져')
        # 새로 생긴 홀
        HoleX2 = X1 - 86 + 9.5
        HoleY2 = HoleY1 - 5
        drawcircle(doc, HoleX2 , HoleY2 ,  1.65 , layer='레이져')         
        HoleX3 = X4 - 86
        HoleY3 = Y4 - 10
        drawcircle(doc, HoleX3 , HoleY3 ,  1.65, layer='레이져') 
        # 새로 생긴 홀 
        HoleX4 = X4 - 86 + 9.5
        HoleY4 = Y4 - 10 + 5      
        drawcircle(doc,  HoleX4 , HoleY4 ,  1.65 , layer='레이져') 
        HoleX5 = X3 + 43.5
        HoleY5 = Y3 + 30              
        drawcircle(doc, HoleX5 , HoleY5 ,  1.65 , layer='레이져') 

        # HPI 상부 치수선 3개
        dim_linear(doc,  centerBaseX-R/2, Y1, X1, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X1, Y1, X2, Y1, "", 35,  direction="up", layer='dim')
        dim_linear(doc,  X2, Y1, centerBaseX+R/2, Y1, "", 35,  direction="up", layer='dim')

        # HPI 하부 치수선 2개
        dim_linear(doc,   HoleX3, HoleY3, centerBaseX, centerBaseY, "", 35 + 70,  direction="down", layer='dim')
        dim_linear(doc,   centerBaseX, centerBaseY, HoleX5, HoleY5, "", 35 + 10.6  ,  direction="down", layer='dim')
        
        # 3.3파이 지시선
        dim_leader_line(doc, X1 - 86, Y1 + 10, X1-50 , Y1 + 70, "3-3.3%%c", layer="dim", style='0')                            

        # 좌측 홀간격 치수선
        dim_vertical_left(doc, X1 - 86, Y1 + 10,  X4 - 86, Y4 - 10,  60, 'dim', text_height=0.22, text_gap=0.07)
        dim_vertical_left(doc, X1 , Y1,  X4, Y4 ,  60, 'dim', text_height=0.22, text_gap=0.07)

        # 상부 홀간격 치수선
        dim_linear(doc, HoleX1 , HoleY1 ,  HoleX5 , HoleY5 ,   "", 100,  direction="up", layer='dim')

        # 우측 HPI height 치수선
        dim_vertical_right(doc, X3 , centerBaseY + HPI_height - Vcut_rate,  X3 , centerBaseY ,  140, 'dim', text_height=0.22,  text_gap=0.07) 

############################################################################################################################################################################################
# HPI Bracket Block 삽입
############################################################################################################################################################################################
def draw_HPI_bracket():
    if(HPI_type == "OTIS HIX-A162"):    
        block_name = "HIX_A162_laser"
        insert_point = (-2000, 2000)

        # 블록 삽입하는 방법           
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': 1,
            'yscale': 1,
            'rotation': 0,                        
            'layer': '레이져'  # 레이어 지정
        })        
        block_name = "HIX_A162_dim"
        insert_point = (-2000, 2000)

        # 블록 삽입하는 방법           
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': 1,
            'yscale': 1,
            'rotation': 0
        })        
        Textstr = f"{HPI_count} EA "       
        draw_Text(doc, -1400, 1739, 42 , str(Textstr), '0')
    if(HPI_type == "OTIS HIX-A164/165"):    
        block_name = "HIX_A165_laser"
        insert_point = (-2000, 2000)

        # 블록 삽입하는 방법           
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': 1,
            'yscale': 1,
            'rotation': 0,                        
            'layer': '레이져'  # 레이어 지정
        })        
        block_name = "HIX_A165_dim"
        insert_point = (-2000, 2000)

        # 블록 삽입하는 방법           
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': 1,
            'yscale': 1,
            'rotation': 0
        })        
        Textstr = f"{HPI_count} EA "       
        draw_Text(doc, -1400+ 300 , 1739 + 950, 42 , str(Textstr), '0')

#############################
# 도면틀 삽입
#############################        
def insert_frame( x, y, scale , sep , text):    
    if(sep =="top"):
        block_name = "transom_frame"
        insert_point = (x, y, scale)

        # 블록 삽입하는 방법           
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': scale,
            'yscale': scale,
            'rotation': 0
        })        
        Textstr = f"{text}"       
        draw_Text(doc, x + 2200*scale, y + 220*scale , 25 , str(Textstr), '0')
        draw_Text(doc, x + (2200+200)*scale, y + (220-88)*scale   , 25 , "트랜섬", '0')
        draw_Text(doc, x + (2200-320)*scale, y + (220-88)*scale   , 25 , formatted_date , '0')

    # 멍텅구리 상판   
    if(sep =="top_normal"):
        block_name = "transom_frame"
        insert_point = (x, y, scale)

        # 블록 삽입하는 방법           
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': scale,
            'yscale': scale,
            'rotation': 0
        })        
        Textstr = f"{text}"       
        draw_Text(doc, x + 2200*scale, y + 230*scale , 17 , str(Textstr), '0')
        draw_Text(doc, x + (2200+200)*scale, y + (220-88)*scale   , 20 , "트랜섬", '0')
        draw_Text(doc, x + (2200-340)*scale, y + (220-88)*scale   , 20 , formatted_date , '0')
    if(sep =="side_normal"):
        block_name = "side_frame"
        insert_point = (x, y, scale)

        # 블록 삽입하는 방법           
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': scale,
            'yscale': scale,
            'rotation': 0
        })        
        Textstr = f"{text}"       
        draw_Text(doc, x + (2200+1430)*scale, y + (220+130)*scale , 25 , str(Textstr), '0')
        draw_Text(doc, x + (2200+150+1550)*scale, y + (220-92+100)*scale   , 30 , "SIDE JAMB", '0')
        draw_Text(doc, x + (2200-320+1100)*scale, y + (220-92+100)*scale   , 30 , formatted_date , '0')
    if(sep =="side"):
        block_name = "side_frame"
        insert_point = (x, y, scale)

        # 블록 삽입하는 방법           
        msp.add_blockref(block_name , insert_point, dxfattribs={
            'xscale': scale,
            'yscale': scale,
            'rotation': 0
        })        
        Textstr = f"{text}"       
        draw_Text(doc, x + (2200+1430)*scale, y + (220+130)*scale , 30 , str(Textstr), '0')
        draw_Text(doc, x + (2200+200+1550)*scale, y + (220-88+100)*scale   , 30 , "SIDE JAMB", '0')
        draw_Text(doc, x + (2200-320+1100)*scale, y + (220-88+100)*scale   , 30 , formatted_date , '0')

##########################################################################################################################################
# 손상민소장 UFrame 그리기
##########################################################################################################################################
def draw_Uframe(Abs_Xpos, Abs_Ypos, U,G ,R, C1, C2, Floor_mc, Floor_des):   
    Size1 = round(U + G - Bending_rate,1)
    Size2 = math.floor((R + C1 + C2 + 50)/10)*10
    TXpos = Abs_Xpos - 5000
    X1 = TXpos 
    Y1 = Abs_Ypos 
    X2 = X1 + Size2
    Y2 = Y1
    line(doc, X1, Y1, X2, Y2, layer='레이져')       
    X3 = X2
    Y3 = Y2 - U + Vcut_rate
    lineto(doc, X3, Y3, layer='레이져')          
    X4 = X3
    Y4 = Y3 - G + Vcut_rate
    lineto(doc, X4, Y4, layer='레이져')          
    X5 = X1
    Y5 = Y4
    lineto(doc, X5, Y5, layer='레이져')          
    X6 = X1
    Y6 = Y2 - U + Vcut_rate
    lineto(doc, X6, Y6, layer='레이져')          
    lineto(doc,  X1, Y1, layer='레이져')

    # 절곡선
    line(doc, X6, Y6, X3, Y3,  layer='22')   

    # 문구 설정
    if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
        Pre_text = ""
    else:
        Pre_text = str(Floor_mc) + "-"
    # Textstr 생성
    Textstr = f"({Pre_text}{str(Floor_des)})"       
    sizedes = f"{Size2} x {str(Size1)}"       

    draw_Text(doc, X2-500, Y2 - 40, 30, str(Textstr), '레이져')

    dim_linear(doc,  X1, Y1, X2, Y2,  "",  100 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      
    dim_vertical_right(doc, X2, Y2,  X3, Y3,    50, 'dim', text_height=0.22, text_gap=0.07)  
    dim_vertical_right(doc, X4, Y4,  X3, Y3,    100, 'dim', text_height=0.22, text_gap=0.07)  
    dim_vertical_right(doc, X2, Y2,  X4, Y4,    150, 'dim', text_height=0.22, text_gap=0.07)  

    X1 = TXpos + Size2 + 400 
    Y1 = Abs_Ypos  - U
    X2 = X1
    Y2 = Y1 + U 
    X3 = X1 + G
    Y3 = Y1
    
    line(doc, X1, Y1, X2, Y2, layer='0')                                    
    line(doc, X1, Y1, X3, Y3, layer='0')                                    
    dim_vertical_left(doc,  X1, Y1, X2, Y2, 60, 'dim' ,text_height=0.22,  text_gap=0.07)                             
    dim_linear(doc,  X3, Y3, X2, Y2,  "",  80 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
    drawcircle(doc, X1, Y1 , 5 , layer='0', color='6') 
    dim_leader_line(doc, X1, Y1 , X1 - 100 , Y1-50, "V-cut", layer="dim", style='0')      
    
    # 문구
    Textstr = f"Part Name : 상판 쫄대({Pre_text}{str(Floor_des)})"    
    draw_Text(doc, X2+100, Y2 + 40, 22, str(Textstr), '0')
    Textstr = f"Mat.Spec : {widejamb_material}"    
    draw_Text(doc, X2+100, Y2 , 22, str(Textstr), '0')
    Textstr = f"Size : {sizedes}"    
    draw_Text(doc, X2+100, Y2 - 40 , 22, str(Textstr), '0')
    Textstr = f"Quantity : 1 EA"    
    draw_Text(doc, X2+100, Y2 - 80, 22, str(Textstr), '0')

##########################################################################################################################################
# 3UP 전체(상판,기둥)
##########################################################################################################################################
def execute_3up():    
    # 시트 선택 (시트명을 지정)
    sheet_name = '3UP제작'  # 원하는 시트명으로 변경
    sheet = workbook[sheet_name]

    # 2차원 배열을 저장할 리스트 생성
    data_2d = []

    # 1행부터 20행까지의 값을 2차원 배열에 저장
    for row_num in range(7, sheet.max_row + 1):  # 7행부터 20행까지 반복
        cell_value = sheet.cell(row=row_num, column=2).value  # 2열의 값 가져오기
        if cell_value is not None and cell_value != 'F':  # 값이 None이 아닌 경우에만 배열에 추가
            row_values =  [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, sheet.max_column + 1)]        
            data_2d.append(row_values)

    # 엑셀 파일 닫기
    workbook.close()

    if not data_2d:
        print("해당 데이터가 없습니다.")
        return  # 함수를 빠져나감    

    ########################################################################################################################################################################
    # 3UP 상판 전개도
    ########################################################################################################################################################################

    Abs_Xpos = 83000
    Abs_Ypos = 0

    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        OP = row[2]
        R = row[3]
        JD1 = row[4]
        JD2 = row[5]
        JB1_up = row[6]
        JB1_mid = row[7]
        JB1_down = row[8]
        JB2_up = row[9]
        JB2_mid = row[10]
        JB2_down = row[11]
        jb1_gap = row[12]
        jb2_gap = row[13]
        H1 = row[14]
        H2 = row[15]
        C1 = row[16]        
        C2 = row[17]
        A1 = row[18]
        A2 = row[19]
        LH1 = row[20]
        LH2_top = row[21]
        LH2_mid = row[22]
        LH2_bottom = row[23]
        RH1 = row[24]
        RH2_top = row[25]        
        RH2_mid = row[26]        
        RH2_bottom = row[27]        
        U = row[28]        
        G = row[29]        
        UW = row[30]        
        SW = row[31]        
        K1_up = row[32]        
        K1_mid = row[33]        
        K1_down = row[34]        
        K2_up = row[35]        
        K2_mid = row[36]        
        K2_down = row[37]                
        Upper_size = row[38]
        Left_side_size = row[39]
        Right_side_size = row[40]
        E1 = row[41]
        E2 = row[42]
        Angle = row[46]
        Top_pop1 = round(row[47] + 2, 1) #소수점 첫째자리까지만 나오게
        Top_pop2 = round(row[48] + 2, 1)         
        Kadapt =  row[58]   # K값 J값에 적용여부 넣기 1이면 넣기
        G_Setvalue = row[59]  # 손상민 소장 G값 강제로 적용 작지 52번째

        # 데이터 검증 및 기본값 설정        
        K1_up = validate_or_default(K1_up)
        K1_mid = validate_or_default(K1_mid)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_mid = validate_or_default(K2_mid)
        K2_down = validate_or_default(K2_down)
        U = validate_or_default(U)
        E1 = validate_or_default(E1)
        E2 = validate_or_default(E2)
        G = validate_or_default(G)        
        G_Setvalue = validate_or_default(G_Setvalue)        
        OP = validate_or_default(OP)
        Kadapt = validate_or_default(Kadapt)

        if(worker=="손상민" and G_Setvalue > 0) :
            # 상부 돌출
            Top_pop1 = G_Setvalue
            Top_pop2 = G_Setvalue
       
        if(Kadapt==1):
            K1_up = 0            
            K2_up = 0            
            K1_down = 0            
            K2_down = 0            
             
        if(Top_pop1<4):
            Top_pop1 = 4
        if(Top_pop2<4):
            Top_pop2 = 4                

        if(OP>0):             
            if(G > 0):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U + G - Vcut_rate*3
                X2 = Abs_Xpos + H1 + H2 + R 
                Y2 = Y1      
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Abs_Ypos + U - Vcut_rate * 2           
                line(doc, X2, Y2, X3, Y3, layer='레이져')          
            if(G < 1):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U  - Vcut_rate
                X2 = Abs_Xpos + H1 + H2 + R 
                Y2 = Y1      
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Y2

            X4 = X3
            Y4 = Abs_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')
            X5 = X4
            Y5 = Abs_Ypos - Top_pop2 + Vcut_rate        
            lineto(doc,  X5, Y5, layer='레이져')    
            X6 = X5 - H2
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='레이져')

            if(K1_up>0 or K2_up>0):
                X7 = Abs_Xpos + H1 + (R-OP)/2 + OP                
                Y7 = Abs_Ypos - JD2 + Vcut_rate
                lineto(doc,  X7, Y7, layer='레이져')
            else:
                X7 = Abs_Xpos + H1 + (R-OP)/2 + OP    
                Y7 = Abs_Ypos - JD2 + Vcut_rate * 2
                lineto(doc,  X7, Y7, layer='레이져')  
                
            if(K1_up>0 or K2_up>0 ):
                X8 = X7
                Y8 = Abs_Ypos - JD2 - K2_up + Vcut_rate*2
                lineto(doc,  X8, Y8, layer='레이져') 
                X9 = X8 
                Y9 = Abs_Ypos - JD2 - K2_up - UW + Vcut_rate*3     
                lineto(doc,  X9, Y9, layer='레이져')   
                X10 = Abs_Xpos + H1 + (R-OP)/2 
                Y10 = Abs_Ypos - JD1 - K1_up - UW + Vcut_rate*3
                lineto(doc,  X10, Y10, layer='레이져')      
                X11 = X10 
                Y11 =  Abs_Ypos - JD1 - K1_up + Vcut_rate*2
                lineto(doc,  X11, Y11, layer='레이져')                                      
            else:
                X8 = X7
                Y8 = Y7
                X9 = X8
                Y9 = Abs_Ypos - JD2 - UW + Vcut_rate*3            
                lineto(doc,  X9, Y9, layer='레이져')   
                X10 = Abs_Xpos + H1 + (R-OP)/2 
                Y10 = Abs_Ypos - JD1 - UW + Vcut_rate*3
                lineto(doc,  X10, Y10, layer='레이져')      
                X11 = X10                 
                Y11 =  Abs_Ypos - JD1 + Vcut_rate*2
                lineto(doc,  X11, Y11, layer='레이져')                                

            X12 = X11
            if(K1_up>0 or K2_up>0):
                Y12 = Abs_Ypos - JD1 + Vcut_rate 
            else:  
                Y12 = Abs_Ypos - JD1 + Vcut_rate * 2
            
            lineto(doc,  X12, Y12, layer='레이져')                   
            X13 = Abs_Xpos + H1
            Y13 = Abs_Ypos - Top_pop2 + Vcut_rate 
            lineto(doc,  X13, Y13, layer='레이져')      
            X14 = Abs_Xpos
            Y14 = Y13
            lineto(doc,  X14, Y14, layer='레이져')      
            X15 = X14 
            Y15 = Abs_Ypos
            lineto(doc,  X15, Y15, layer='레이져')     
            if(G > 0):         
                X16 = X15 
                Y16 = Abs_Ypos + U - Vcut_rate * 2   
                lineto(doc,  X16, Y16, layer='레이져')      
                lineto(doc,  X1, Y1, layer='레이져')                  
            if(G < 1):         
                X16 = X15 
                Y16 = Abs_Ypos + U - Vcut_rate 
                lineto(doc,  X1, Y1, layer='레이져')                  

            # if args.opt17:
            #     drawcircle(doc, X10 + 12.5, Y10 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이
            #     drawcircle(doc, X9 - 12.5, Y9 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이

            # 상판 절곡선
            if(G>0):
                line(doc, X16, Y16, X3, Y3,  layer='22')        
            line(doc, X15, Y15, X4, Y4,  layer='22')
            line(doc, X11, Y11, X8, Y8,  layer='22')
    
            # 상판 우측 절곡치수선
            if(G>0):        
                dim_vertical_right(doc, X2, Y2, X3, Y3,  100, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X4, Y4, X3, Y3,  50, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X4, Y4, X8, Y8,  50, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc,  X9, Y9, X8, Y8, extract_abs(X8,X5) + 50, 'dim', text_height=0.22, text_gap=0.07)   # 뒷날개
            dim_vertical_right(doc, X2, Y2,  X9, Y9,  extract_abs(X9,X2) + 100, 'dim', text_height=0.22, text_gap=0.07)             

            # 상판 좌측 절곡치수선
            if(JD1 != JD2 or K1_up != K2_up):
                if(G>0):                    
                    dim_vertical_left(doc, X1, Y1, X16, Y16,  extract_abs(X1,X16) + 100, 'dim', text_height=0.22, text_gap=0.07)  
                dim_vertical_left(doc, X15, Y15, X16, Y16,  extract_abs(X15,X16) + 50, 'dim', text_height=0.22, text_gap=0.07)                              
                dim_vertical_left(doc, X11, Y11, X15, Y15,  extract_abs(X11,X15) + 50 , 'dim', text_height=0.22, text_gap=0.07)              
                dim_vertical_left(doc, X10, Y10, X11, Y11,  H1 + (R-OP)/2 + extract_abs(X10,X11) + 50 , 'dim', text_height=0.22, text_gap=0.07)              
                dim_vertical_left(doc, X10, Y10, X1, Y1,  extract_abs(X10,X1) + 200, 'dim', text_height=0.22, text_gap=0.07)                          
            else:    
                dim_vertical_left(doc, X10, Y10, X11, Y11,   H1 + (R-OP)/2  + 10, 'dim', text_height=0.22,  text_gap=0.07) 
                dim_vertical_left(doc, X10, Y10, X15, Y15,   H1 + (R-OP)/2  + 100 , 'dim', text_height=0.22,  text_gap=0.07) 
                dim_vertical_left(doc, X10, Y10, X16, Y16,   H1 + (R-OP)/2  + 150, 'dim', text_height=0.22,  text_gap=0.07) 

            # 상판 하부 치수선 3개
            if(Angle>1):
                dim_linear(doc,  X9, Y9, X6, Y6, "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X10, Y10, X9, Y9,  "", 100 + extract_abs(JD1+K1_up, JD2+K2_up),  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            if(Angle>1):
                dim_linear(doc,  X13, Y13, X10, Y10,  "",  extract_abs(Y6,Y9) + 100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            # 상판 돌출부위 3.4 치수선
            dim_vertical_right(doc, X15, Y15, X13, Y13, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_left(doc, X4, Y4, X6, Y6, 100, 'dim', text_height=0.22,  text_gap=0.07)  

            # 기둥과의 접선 부위 aligned 표현 각도표현
            dim_linear(doc, X13, Y13, X11, Y11,   "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) # 좌측
            if(Angle>1):
                dim_angular(doc, X12, Y12, X13, Y13, X11, Y11, X12, Y12 + 10,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다        
            dim_linear(doc, X8, Y8, X6, Y6,  "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) # 좌측
            if(Angle>1):            
                dim_angular(doc, X8, Y8  , X7, Y7 + 10,  X7, Y7,  X6, Y6,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다

            # 상판 K1 K2 값 부분 치수선 
            if(K1_up > 3 or K2_up > 3 ):
                dim_vertical_right(doc, X11, Y11, X12, Y12, 80, 'dim', text_height=0.22,  text_gap=0.07)  
                dim_vertical_left(doc, X8, Y8, X7, Y7, 80, 'dim', text_height=0.22,  text_gap=0.07)  

            # 전개도 상부 치수선
            dim_linear(doc,  X1, Y1, X2, Y2, "", 120,  direction="up", layer='dim')
            dim_linear(doc,  X13, Y13, X6, Y6,  "", extract_abs(Y1,Y13) + 50,  direction="up", layer='dim')
            dim_linear(doc,  X1, Y1, X13, Y13, "",  50,  direction="up", layer='dim')
            dim_linear(doc,  X6, Y6, X2, Y2,  "",  extract_abs(Y2,Y6) + 50,  direction="up", layer='dim')

            # 상부 Vcut 치수 표기
            if(U>0 and G>0):   
               dim_vertical_left(doc, X1+H1+OP/2, Y1, X1+H1+OP/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  

            # 상판 좌우가 다를때 Vcut 상부에 치수선 표기
            # if((JD1 != JD2 or K1_up != K2_up) and G>0):
            #     dim_vertical_left(doc, X1 + R/2, Y1, X1 + R/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            # 문구 설정
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"        
            Textstr = f"({Pre_text}{str(Floor_des)})"       

            X2 = Abs_Xpos + (H1 + R + H2)/2 - len(Textstr)*50/2
            Y2 = Abs_Ypos - JD1/1.5
            draw_Text(doc, X2, Y2, 50, str(Textstr), '레이져')

            # 문구            
            sizedes = Upper_size                  
            
            Textstr = f"Part Name : 상판({Pre_text}{str(Floor_des)})"    
            draw_Text(doc, X9+150, Y9 + 50 - 220, 28, str(Textstr), '0')
            Textstr = f"Mat.Spec : {normal_material}"    
            draw_Text(doc, X9+150, Y9 - 220, 28, str(Textstr), '0')
            Textstr = f"Size : {sizedes}"    
            draw_Text(doc, X9+150, Y9 - 50 - 220 , 28, str(Textstr), '0')
            Textstr = f"Quantity : 1 EA"    
            draw_Text(doc, X9+150, Y9 - 100 - 220 , 28, str(Textstr), '0')                        

            # 도면틀 넣기
            BasicXscale = 2560
            BasicYscale = 1550
            TargetXscale = 800 + R + U * 2 + 200
            TargetYscale = 300 + JD1 + G + U + UW
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            insert_frame(X1 - (U + 580) ,Y1 - ( 1000 + JD1 + K1_up + G + U + UW) , frame_scale, "top_normal", workplace)        

            ###########################################################################################################################################################################
            # 3UP 상판 좌측 단면도    (좌우가 다른 모양일때 단면도를 그려준다)
            ###########################################################################################################################################################################
            if(JD1 != JD2 or K1_up != K2_up):
                Section_Xpos = Abs_Xpos - H1 - 400           

                if(G > 0):
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos - G
                    X2 = Section_Xpos
                    Y2 = Abs_Ypos  

                    line(doc, X1, Y1, X2, Y2, layer='0')                    
                    dim_vertical_left(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)     
                    X3 = Section_Xpos + U
                    Y3 = Abs_Ypos  
                    lineto(doc,  X3, Y3, layer='0')    
                    dim_linear(doc,  X2, Y2, X3, Y3, "", 50,  direction="up", layer='dim')                
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                    dim_leader_line(doc, X3, Y3 , X2-130 , Y2-100, "V-cut 3개", layer="dim", style='0')                
                if(G < 1):
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos 
                    X2 = Section_Xpos
                    Y2 = Abs_Ypos                  
                    X3 = Section_Xpos + U
                    Y3 = Abs_Ypos   
                    line(doc, X2, Y2, X3, Y3, layer='0')                  
                    dim_linear(doc,  X2, Y2, X3, Y3, "", 50,  direction="up", layer='dim')    
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                    dim_leader_line(doc, X3, Y3 , X2-130 , Y2-100, "V-cut 2개", layer="dim", style='0')                                
                
                X4 = X3 
                Y4 = Y3 - JD1
                lineto(doc,  X4, Y4, layer='0')                   
                if(K1_up > 0):
                    X5 = X4 
                    Y5 = Y4 - K1_up
                    lineto(doc,  X5, Y5, layer='0')   
                    dim_vertical_right(doc,  X3, Y3, X4, Y4, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X4, Y4, X5, Y5, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X3, Y3, X5, Y5, 110,'dim' ,text_height=0.22,  text_gap=0.07)                                     
                else:
                    dim_vertical_right(doc,  X3, Y3, X4, Y4, 110,'dim' ,text_height=0.22,  text_gap=0.07)     
                    X5 = X4 
                    Y5 = Y4

                drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')         
                X6 = X5 - UW
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')   
                dim_linear(doc,  X6, Y6, X5, Y5,  "", 80,  direction="down", layer='dim')            
                
                if(K1_up > 0):                 
                    # K값 부분 라인 그리기
                    line(doc, X4, Y4, X4 - UW, Y4, layer='22')     
                
                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JD1 != JD2 or K1_up != K2_up ):
                    rectangle(doc, X4+50, Y5+(JD1+K1_up)/2 - 75 , X4 + 50 + 100, Y5+(JD1+K1_up)/2 + 75, layer='0')               
                    rectangle(doc, X4+285,  Y5+(JD1+K1_up)/2 - 75 , X4+285+100, Y5+(JD1+K1_up)/2 + 75 , layer='0')                                                         
                
            ################################################################################################################################################################
            # 3UP 상판 우측 단면도 
            ################################################################################################################################################################
            Section_Xpos = Abs_Xpos + R + H1 + H2 + 400

            if(G>0):
                X1 = Section_Xpos 
                Y1 = Abs_Ypos - G
                X2 = Section_Xpos
                Y2 = Abs_Ypos  

                line(doc, X1, Y1, X2, Y2, layer='0')                    
                dim_vertical_right(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)     
                X3 = Section_Xpos - U
                Y3 = Abs_Ypos  
                lineto(doc,  X3, Y3, layer='0')    
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim')
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                dim_leader_line(doc, X3, Y3 , X3+50 , Y2-100, "V-cut 3개", layer="dim", style='0')                

            if(G<1):
                X1 = Section_Xpos 
                Y1 = Abs_Ypos
                X2 = Section_Xpos
                Y2 = Abs_Ypos  
                X3 = Section_Xpos - U
                Y3 = Abs_Ypos  
                line(doc, X1, Y1, X3, Y3, layer='0')                            
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim')    
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                dim_leader_line(doc, X3, Y3 , X3 + 50 , Y2-100, "V-cut 2개", layer="dim", style='0')                            
                
            X4 = X3 
            Y4 = Y3 - JD2
            lineto(doc,  X4, Y4, layer='0')   
            if(K1_up > 0):
                X5 = X4 
                Y5 = Y4 - K2_up
                lineto(doc,  X5, Y5, layer='0')   
                dim_vertical_left(doc,  X3, Y3, X4, Y4, 60,'dim' ,text_height=0.22,  text_gap=0.07)                     
                dim_vertical_left(doc,  X4, Y4, X5, Y5, 60,'dim' ,text_height=0.22,  text_gap=0.07)     
                dim_vertical_left(doc,  X3, Y3, X5, Y5, 110,'dim' ,text_height=0.22,  text_gap=0.07)                                     
            else:
                dim_vertical_left(doc,  X3, Y3, X4, Y4, 110,'dim' ,text_height=0.22,  text_gap=0.07)                     
                X5 = X4 
                Y5 = Y4

            X6 = X5 + UW
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='0')   
            dim_linear(doc,  X5, Y5, X6, Y6,   "", 80,  direction="down", layer='dim')
            
            drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
            
            if(K2_up > 0):                 
                # K값 부분 라인 그리기
                line(doc, X4, Y4, X4 + UW, Y4, layer='22')      

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JD1 != JD2 or K1_up != K2_up ):
                rectangle(doc, X4-50, Y5+(JD2+K2_up)/2 - 75 , X4 - 50 - 100, Y5+(JD2+K2_up)/2 + 75, layer='0')               
                rectangle(doc, X4-285,  Y5+(JD2+K2_up)/2 - 75 , X4-285-100,  Y5+(JD2+K2_up)/2 + 75 , layer='0')                                                         
            
        # 다음 도면 간격 계산
        Abs_Ypos -= 4000

    Abs_Xpos = 83000 + 5000
    Abs_Ypos = 500

    ############################################################################################################################################################################
    # 3UP 좌기둥 전개도
    ############################################################################################################################################################################

    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        OP = row[2]
        R = row[3]
        JD1 = row[4]
        JD2 = row[5]
        JB1_up = row[6]
        JB1_mid = row[7]
        JB1_down = row[8]
        JB2_up = row[9]
        JB2_mid = row[10]
        JB2_down = row[11]
        jb1_gap = row[12]
        jb2_gap = row[13]
        H1 = row[14]
        H2 = row[15]
        C1 = row[16]        
        C2 = row[17]
        A1 = row[18]
        A2 = row[19]
        LH1 = row[20]
        LH2_top = row[21]
        LH2_mid = row[22]
        LH2_bottom = row[23]
        RH1 = row[24]
        RH2_top = row[25]        
        RH2_mid = row[26]        
        RH2_bottom = row[27]        
        U = row[28]        
        G = row[29]        
        UW = row[30]        
        SW = row[31]        
        K1_up = row[32]        
        K1_mid = row[33]        
        K1_down = row[34]        
        K2_up = row[35]        
        K2_mid = row[36]        
        K2_down = row[37]                
        Upper_size = row[38]
        Left_side_size = row[39]
        Right_side_size = row[40]
        E1 = row[41]
        E2 = row[42]
        Angle = row[46]
        Top_pop1 = round(row[47] + 2, 1) #소수점 첫째자리까지만 나오게
        Top_pop2 = round(row[48] + 2, 1)         
        Kadapt =  row[58]   # K값 J값에 적용여부 넣기 1이면 넣기
        G_Setvalue = row[59]  # 손상민 소장 G값 강제로 적용 작지 52번째

        if(Kadapt==1):
            K1_up = 0
            K1_mid = 0
            K1_down = 0
            K2_up = 0
            K2_mid = 0
            K2_down = 0   

        OP = validate_or_default(OP)

        if(OP>0):          
            X1 = Abs_Xpos 
            Y1 = Abs_Ypos + JB1_up + K1_up +  SW - Vcut_rate * 5
            X2 = Abs_Xpos + LH2_top - 2
            Y2 = Y1
            X3 = X2            
            Y3 = Y2 - SW + Vcut_rate * 2             
            X4 = X3 + 2
            Y4 = Y3            
            X5 = X4
            Y5 = Y4 + (JB1_mid - JB1_up)            
            X6 = X5 
            Y6 = Y5 + SW - Vcut_rate * 2
            X7 = Abs_Xpos + LH2_top + LH2_mid - 2
            Y7 = Abs_Ypos + JB1_mid + SW - Vcut_rate * 5            
            X8 = X7 
            Y8 = Y7 - SW + Vcut_rate * 2             
            X9 = X8 + 2
            Y9 = Y8

            X10 = X9
            Y10 = Y9 + (JB1_down - JB1_mid)            
            X11 = X10 
            Y11 = Y10 + SW - Vcut_rate * 2          
            X12 = Abs_Xpos + LH2_top + LH2_mid + LH2_bottom 
            Y12 = Abs_Ypos + JB1_down + SW - Vcut_rate * 5               
            X13 = X12
            Y13 = Y12 - SW + Vcut_rate * 2  
            X14 = Abs_Xpos + LH1 
            Y14 = Abs_Ypos

            if(A1>0):       
                X15 = X14
                Y15 = Abs_Ypos - C1 + Vcut_rate * 2                
                X16 = X15
                Y16 = Y15 - A1 + Vcut_rate                 
                X17 = Abs_Xpos             
                Y17 = Abs_Ypos - C1 - A1 + Vcut_rate * 3                
                X18 = Abs_Xpos             
                Y18 = Abs_Ypos - C1 + Vcut_rate * 2                

            if(A1<1):       
                X15 = X14
                Y15 = Abs_Ypos - C1 + Vcut_rate                
                X16 = X15
                Y16 = Y15                
                X17 = Abs_Xpos             
                Y17 = Y16                
                X18 = Abs_Xpos             
                Y18 = Abs_Ypos - C1 + Vcut_rate * 1                

            X19 = Abs_Xpos 
            Y19 = Abs_Ypos

            X20 = Abs_Xpos 
            Y20 = Abs_Ypos + JB1_up + K1_up  - Vcut_rate * 3 

            prev_x, prev_y = X1, Y1  # 첫 번째 점으로 초기화
            for i in range(1, 20 + 1):
                curr_x, curr_y = eval(f'X{i}'), eval(f'Y{i}')
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y
                # print(prev_x)
            # 마지막으로 첫번째 점과 연결
            line(doc, X1, Y1, prev_x, prev_y, layer="레이져")

            # 좌기둥 절곡선 layer 22 
            line(doc, X20, Y20, X3, Y3,  layer='22')
            line(doc, X5, Y5, X8, Y8,  layer='22')
            line(doc, X10, Y10, X13, Y13,  layer='22')
            line(doc, X19, Y19, X14, Y14,  layer='22')
            if(A1>0):
                line(doc, X18, Y18, X15, Y15,  layer='22')         

            # 뒷날개쪽 치수선 3단
            dim_linear(doc,  X1, Y1, X12, Y12, "", 160 + extract_abs(Y1,Y12),  direction="up", layer='dim')                           
            dim_linear(doc,  X1, Y1, X6, Y6, "", 60  + extract_abs(Y1,Y12),  direction="up", layer='dim')                           
            dim_linear(doc,  X6, Y6, X11, Y11, "", 60  + extract_abs(Y1,Y6),  direction="up", layer='dim')                           
            dim_linear(doc,  X11, Y11, X12, Y12, "", 60,  direction="up", layer='dim')                    

            # 도어두께 표시 상부
            dim_vertical_right(doc, X4,  Y4, X5, Y5, 100, 'dim', text_height=0.22, text_gap=0.01)                         
            dim_vertical_right(doc, X9,  Y9, X10, Y10, 100, 'dim', text_height=0.22, text_gap=0.01)                         

            # 기둥 우측 절곡치수선
            # 뒷날개
            dim_vertical_right(doc, X12,  Y12, X13, Y13, 60  + extract_abs(X13,X14) , 'dim', text_height=0.22, text_gap=0.01)                 
            # JB
            dim_vertical_right(doc, X13,  Y13, X14, Y14, 60  + extract_abs(X13,X14)  , 'dim', text_height=0.22, text_gap=0.01)                              
            # C1
            dim_vertical_right(doc, X14,  Y14, X15, Y15, 60 + extract_abs(X13,X14)  , 'dim', text_height=0.22, text_gap=0.01)                                          
            if(A1>0):
                dim_vertical_right(doc, X15,  Y15, X16, Y16,  120 + extract_abs(X13,X14) , 'dim', text_height=0.22, text_gap=0.01)  
                # V컷 치수표현
                dim_vertical_left(doc, X14,  Y14, X16, Y16,   120 + extract_abs(X13,X14) , 'dim', text_height=0.22, text_gap=0.01)  
            dim_vertical_right(doc, X12,  Y12, X16, Y16,   extract_abs(X13,X14)  + 200 , 'dim', text_height=0.22, text_gap=0.01)      

            # 기둥 좌측 절곡치수선  
            # 뒷날개
            dim_vertical_left(doc, X1, Y1, X20, Y20,  80, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X19, Y19, X20, Y20, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_left(doc, X19, Y19, X18, Y18,  80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_left(doc, X17, Y17, X18, Y18,  150 , 'dim', text_height=0.22, text_gap=0.01)
            dim_vertical_left(doc, X1, Y1, X17,  Y17,  230 , 'dim', text_height=0.22, text_gap=0.01)

            # 기둥 하부 치수선 2    
            dim_linear(doc,   X17, Y17, X16, Y16, "", 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   

            # LH1, LH2 치수크기 차이로 인해 설치
            if (LH1 > LH2_top+LH2_mid+LH2_bottom) :
                dim_linear(doc,  X12, Y12, X14, Y14, "",  180 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)       
            if (LH1 < LH2_top+LH2_mid+LH2_bottom) :
                dim_linear(doc,  X16, Y16, X13, Y13, "", 180 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"       

            Xpos = Abs_Xpos + 200  
            Ypos = Abs_Ypos + JB1_up / 3.5
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-좌"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')        

            Xpos = Abs_Xpos + LH1/2 - 200
            Ypos = Abs_Ypos + JB1_up / 3.5 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-좌"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/2 - 200
            Ypos = Abs_Ypos + JB1_up / 3.5 + 30 - 40
            Textstr = f"Mat.Spec : {normal_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/1.2 -200
            Ypos = Abs_Ypos + JB1_up / 3.5 + 30
            Textstr = f"Size : {Left_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/1.2 - 200
            Ypos = Abs_Ypos + JB1_up / 3.5 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')       

            # 기둥 좌우 도면틀 넣기
            BasicXscale = 4487
            BasicYscale = 2770
            # 우측단면도가 있을 경우 감안해야 함. 
            TargetXscale = 600 + max(LH1, LH2_top + LH2_mid +LH2_bottom) + 600

            TargetYscale = 300*4 + (JB2_down + C1 + A1 + K1_up + SW) * 2
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            # 우측 단면도가 있는 경우 도면틀 기본점
            insert_frame(X1 - 950 , Abs_Ypos - (JB2_up+C2+A2+K2_up + SW + 1300 + C1 + A1 ) , frame_scale, "side_normal", workplace)                

            ##############################################################################################################################################################
            # 3UP 좌기둥 좌측 단면도    
            ##############################################################################################################################################################
            Section_Xpos = Abs_Xpos - 600
            Section_Ypos = Abs_Ypos

            X1 = Section_Xpos - C1 
            Y1 = Section_Ypos + A1
            X2 = X1
            Y2 = Y1 - A1
            X3 = X2 + C1
            Y3 = Y2
            X4 = X3
            Y4 = Y3 + JB1_up + K1_up
            X5 = X4
            Y5 = Y4 + (JB1_mid - JB1_up)  
            X6 = X5
            Y6 = Y5 + (JB1_down - JB1_mid)  
            X7 = X6 - SW
            Y7 = Y6
            X8 = X5 - SW
            Y8 = Y5
            X9 = X4 - SW
            Y9 = Y4

            prev_x, prev_y = X1, Y1  # 첫 번째 점으로 초기화
            for i in range(1, 7 + 1):
                curr_x, curr_y = eval(f'X{i}'), eval(f'Y{i}')
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
                prev_x, prev_y = curr_x, curr_y

            line(doc, X5, Y5, X8, Y8, layer='0')  
            line(doc, X4, Y4, X9, Y9, layer='0')  
           
            if(A1>0):
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6')             
                dim_vertical_left(doc, X1, Y1, X2, Y2, 50, 'dim', text_height=0.22,  text_gap=0.07)                 
            if(C1>0):                
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6') 
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 80,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)                 

            if(A1>0 and C1>0):                
                dim_leader_line(doc, X3, Y3 , X3+50, Y3-70, "V-cut 2개소", layer="dim", style='0')
            else:
                if(C1>0):
                    dim_leader_line(doc, X3, Y3 , X3+50, Y3-70, "V-cut", layer="dim", style='0')

            # JB
            dim_vertical_right(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X5, Y5, X4, Y4, 80, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X6, Y6, X5, Y5, 80, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X3, Y3, X5, Y5, 160, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X3, Y3, X6, Y6, 240, 'dim', text_height=0.22, text_gap=0.07)  

            # 뒷날개
            if(SW>0):
               dim_linear(doc, X7, Y7,  X6, Y6, "", 100,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)    

######################################################################################################################################################
# 3UP 우기둥 전개도
######################################################################################################################################################

            Abs_Ypos -= 2000 
            R_Ypos = Abs_Ypos + 1550 - (C1 + C2 + A1 + A2)             

            if(OP>0):          
                if(A2>0):  
                    X1 = Abs_Xpos 
                    Y1 = R_Ypos + C2 + A2 - Vcut_rate * 3

                    X2 = Abs_Xpos + RH1
                    Y2 = Y1
                    
                    X3 = X2            
                    Y3 = Y2 - A2 + Vcut_rate

                    X4 = X3
                    Y4 = Y3 - C2 + Vcut_rate*2               

                    X5 = Abs_Xpos + RH2_top + RH2_mid + RH2_bottom 
                    Y5 = Y4 - JB2_down + Vcut_rate * 3   

                    X6 = X5 
                    Y6 = Y5 - SW + Vcut_rate * 2
                    
                    X7 = Abs_Xpos + RH2_top + + RH2_mid
                    Y7 = Y6
                    
                    X8 = X7 
                    Y8 = Y7 + SW - Vcut_rate * 2 
                    
                    X9 = X8
                    Y9 = Y8 + (JB2_down - JB2_mid)

                    X10 = X9 - 2
                    Y10 = Y9
                                
                    X11 = X10
                    Y11 = Y10 - SW + Vcut_rate * 2
                    
                    X12 = Abs_Xpos + RH2_top          
                    Y12 = R_Ypos - JB2_mid - SW + Vcut_rate * 5
                
                    X13 = X12            
                    Y13 = Y12 + SW - Vcut_rate * 2                              
                    
                    X14 = X13
                    Y14 = Y13 + (JB2_mid - JB2_up)

                    X15 = X14 - 2
                    Y15 = Y14
                                
                    X16 = X15
                    Y16 = Y15 - SW + Vcut_rate * 2
                                
                    X17 = Abs_Xpos  
                    Y17 = Y16
                    
                    X18 = X17            
                    Y18 = Y17 + SW - Vcut_rate * 2                              

                    X19 = Abs_Xpos 
                    Y19 = R_Ypos                   

                    X20 = Abs_Xpos 
                    Y20 = Y19 + C2 - Vcut_rate * 2        

                if(A2 < 1):  
                    X1 = Abs_Xpos 
                    Y1 = R_Ypos + C2 - Vcut_rate 

                    X2 = Abs_Xpos + RH1
                    Y2 = Y1
                    
                    X3 = X2            
                    Y3 = Y2

                    X4 = X3
                    Y4 = Y3 - C2 + Vcut_rate               

                    X5 = Abs_Xpos + RH2_top + RH2_mid + RH2_bottom 
                    Y5 = Y4 - JB2_down + Vcut_rate * 3   

                    X6 = X5 
                    Y6 = Y5 - SW + Vcut_rate * 2
                    
                    X7 = Abs_Xpos + RH2_top + + RH2_mid
                    Y7 = Y6
                    
                    X8 = X7 
                    Y8 = Y7 + SW - Vcut_rate * 2 
                    
                    X9 = X8
                    Y9 = Y8 + (JB2_down - JB2_mid)

                    X10 = X9 - 2
                    Y10 = Y9
                                
                    X11 = X10
                    Y11 = Y10 - SW + Vcut_rate * 2
                    
                    X12 = Abs_Xpos + RH2_top          
                    Y12 = R_Ypos - JB2_mid - SW + Vcut_rate * 5
                
                    X13 = X12            
                    Y13 = Y12 + SW - Vcut_rate * 2                              
                    
                    X14 = X13
                    Y14 = Y13 + (JB2_mid - JB2_up)

                    X15 = X14 - 2
                    Y15 = Y14
                                
                    X16 = X15
                    Y16 = Y15 - SW + Vcut_rate * 2
                                
                    X17 = Abs_Xpos  
                    Y17 = Y16
                    
                    X18 = X17            
                    Y18 = Y17 + SW - Vcut_rate * 2                              

                    X19 = Abs_Xpos 
                    Y19 = R_Ypos                   

                    X20 = Abs_Xpos 
                    Y20 = Y19 + C2 - Vcut_rate         

                prev_x, prev_y = X1, Y1  # 첫 번째 점으로 초기화
                for i in range(1, 20 + 1):
                    curr_x, curr_y = eval(f'X{i}'), eval(f'Y{i}')
                    line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                    prev_x, prev_y = curr_x, curr_y
                    # print(prev_x)
                # 마지막으로 첫번째 점과 연결
                line(doc, X1, Y1, prev_x, prev_y, layer="레이져")
                
                # 우기둥 절곡선 layer 22 
                line(doc, X18, Y18, X15, Y15,  layer='22')
                line(doc, X8, Y8, X5, Y5,  layer='22')
                line(doc, X13, Y13, X10, Y10,  layer='22')                
                line(doc, X19, Y19, X4, Y4,  layer='22')                
                if(A2>0):
                    line(doc, X20, Y20, X3, Y3,  layer='22')         

                # 뒷날개쪽 치수선
                dim_linear(doc,  X17, Y17, X6, Y6, "", 180 + extract_abs(Y17,Y6),  direction="down", layer='dim')                           
                dim_linear(doc,  X17, Y17, X12, Y12, "", 60  + extract_abs(Y17,Y6),  direction="down", layer='dim')                           
                dim_linear(doc,  X12, Y12, X7, Y7, "", 60  + extract_abs(Y12,Y6),  direction="down", layer='dim')                           
                dim_linear(doc,  X7, Y7, X6, Y6, "", 60,  direction="down", layer='dim')                    

                # 도어두께 표시 상부
                dim_vertical_right(doc, X9,  Y9, X8, Y8, 100, 'dim', text_height=0.22, text_gap=0.01)                         
                dim_vertical_right(doc, X13,  Y13, X14, Y14, 100, 'dim', text_height=0.22, text_gap=0.01)                         

                # 기둥 우측 절곡치수선 C
                dim_vertical_right(doc, X3,  Y3, X4, Y4, 60  + extract_abs(X4,X5) , 'dim', text_height=0.22, text_gap=0.01)                 
                dim_vertical_right(doc, X4,  Y4, X5, Y5, 60  + extract_abs(X4,X5) , 'dim', text_height=0.22, text_gap=0.01)                 
                # JB                
                dim_vertical_right(doc, X5,  Y5, X6, Y6, 60 + extract_abs(X4,X5) , 'dim', text_height=0.22, text_gap=0.01)                                          
                if(A1>0):
                    dim_vertical_right(doc, X2,  Y2, X3, Y3,  120 + extract_abs(X4,X5), 'dim', text_height=0.22, text_gap=0.01)  
                dim_vertical_right(doc, X2,  Y2, X6, Y6,   extract_abs(X4,X5) + 200 , 'dim', text_height=0.22, text_gap=0.01)      

                # 좌측 절곡치수선  
                if(A1>0):
                    dim_vertical_left(doc, X1, Y1, X20, Y20,  150 , 'dim', text_height=0.22, text_gap=0.01)
                # C    
                dim_vertical_left(doc, X20, Y20, X19,  Y19,  80 , 'dim', text_height=0.22, text_gap=0.01)                                        
                # JB    
                dim_vertical_left(doc, X19, Y19, X18, Y18, 80 , 'dim', text_height=0.22, text_gap=0.01)
                # 뒷날개                      
                dim_vertical_left(doc, X18, Y18, X17, Y17,  80 , 'dim', text_height=0.22, text_gap=0.01)  
                # 좌측 전체 치수선
                dim_vertical_left(doc, X1, Y1, X17, Y17,  220 , 'dim', text_height=0.22, text_gap=0.01)  

                # 기둥 상부 치수선 2    
                dim_linear(doc,   X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)   

                # RH1, RH2 치수크기 차이로 인해 설치
                if (RH1 > RH2_top+RH2_mid+RH2_bottom) :
                    dim_linear(doc,  X6, Y6,  X4, Y4, "",  180  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)       
                if (RH1 < RH2_top+RH2_mid+RH2_bottom) :
                    dim_linear(doc,   X2, Y2, X5, Y5, "", 180   ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      

                # 전개도 문구        
                if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                    Pre_text = ""
                else:
                    Pre_text = str(Floor_mc) + "-"       

                Xpos = Abs_Xpos + 200  
                Ypos = R_Ypos - JB2_up / 3.5
                Textstr = f"코팅상"       
                draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
                Textstr = f"({Pre_text}{str(Floor_des)})-우"         
                draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')        

                Xpos = Abs_Xpos + RH1/2 - 200
                Ypos = R_Ypos - JB2_up / 3.5 + 30
                Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-우"       
                draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
                Xpos = Abs_Xpos + RH1/2 - 200
                Ypos = R_Ypos - JB2_up / 3.5 + 30 - 40
                Textstr = f"Mat.Spec : {normal_material} "       
                draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
                Xpos = Abs_Xpos + RH1/1.2 -200
                Ypos = R_Ypos - JB2_up / 3.5 + 30
                Textstr = f"Size : {Right_side_size} "       
                draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
                Xpos = Abs_Xpos + RH1/1.2 - 200
                Ypos = R_Ypos - JB2_up / 3.5 + 30 - 40
                Textstr = f"Quantity : 1 EA "       
                draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')       

##############################################################################################################################################################
# 3UP 우기둥 좌측 단면도    
##############################################################################################################################################################
                Section_Xpos = Abs_Xpos - 600
                Section_Ypos = R_Ypos

                X1 = Section_Xpos - C2 
                Y1 = Section_Ypos - A2
                X2 = X1
                Y2 = Y1 + A2
                X3 = X2 + C2
                Y3 = Y2
                X4 = X3
                Y4 = Y3 - JB2_up - K2_up
                X5 = X4
                Y5 = Y4 - (JB2_mid - JB2_up)
                X6 = X5
                Y6 = Y5 - (JB2_down - JB2_mid)
                X7 = X6 - SW
                Y7 = Y6
                X8 = X5 - SW
                Y8 = Y5
                X9 = X4 - SW
                Y9 = Y4
                
                prev_x, prev_y = X1, Y1  # 첫 번째 점으로 초기화
                for i in range(1, 7 + 1):
                    curr_x, curr_y = eval(f'X{i}'), eval(f'Y{i}')
                    line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
                    prev_x, prev_y = curr_x, curr_y

                line(doc, X5, Y5, X8, Y8, layer='0')  
                line(doc, X4, Y4, X9, Y9, layer='0')  
            
                if(A2>0):
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6')             
                    dim_vertical_left(doc, X1, Y1, X2, Y2, 50, 'dim', text_height=0.22,  text_gap=0.07)                 
                if(C2>0):                
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6') 
                    dim_linear(doc,   X2, Y2, X3, Y3,  "", 80,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)                 

                if(A2>0 and C2>0):                
                    dim_leader_line(doc, X3, Y3 , X3+50, Y3+70, "V-cut 2개소", layer="dim", style='0')
                else:
                    if(C2>0):
                        dim_leader_line(doc, X3, Y3 , X3+50, Y3+70, "V-cut", layer="dim", style='0')
                # JB
                dim_vertical_right(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                dim_vertical_right(doc, X5, Y5, X4, Y4, 80, 'dim', text_height=0.22, text_gap=0.07)  
                dim_vertical_right(doc, X6, Y6, X5, Y5, 80, 'dim', text_height=0.22, text_gap=0.07)  
                dim_vertical_right(doc, X3, Y3, X5, Y5, 160, 'dim', text_height=0.22, text_gap=0.07)  
                dim_vertical_right(doc, X3, Y3, X6, Y6, 240, 'dim', text_height=0.22, text_gap=0.07)  
                # 뒷날개
                if(SW>0):
                    dim_linear(doc,   X7, Y7,  X6, Y6, "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)    

            Abs_Ypos -= 2000 

##########################################################################################################################################
# 2UP 전체(상판,기둥)
##########################################################################################################################################
def execute_2up():      

    # 시트 선택 (시트명을 지정)
    sheet_name = '2UP제작'  # 원하는 시트명으로 변경
    sheet = workbook[sheet_name]

    # 2차원 배열을 저장할 리스트 생성
    data_2d = []

    # 1행부터 20행까지의 값을 2차원 배열에 저장
    for row_num in range(7, sheet.max_row + 1):  # 7행부터 20행까지 반복
        cell_value = sheet.cell(row=row_num, column=2).value  # 2열의 값 가져오기
        if cell_value is not None and cell_value != 'F':  # 값이 None이 아닌 경우에만 배열에 추가
            row_values =  [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, sheet.max_column + 1)]        
            data_2d.append(row_values)

    # 엑셀 파일 닫기
    workbook.close()

    if not data_2d:
        print("해당 데이터가 없습니다.")
        return  # 함수를 빠져나감    

    ########################################################################################################################################################################
    # 2UP 상판 전개도
    ########################################################################################################################################################################

    Abs_Xpos = 73000
    Abs_Ypos = 0

    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        OP = row[2]
        R = row[3]
        JD1 = row[4]
        JD2 = row[5]
        JB1_up = row[6]
        JB1_down = row[7]
        JB2_up = row[8]
        JB2_down = row[9]
        jb1_gap = row[10]
        jb2_gap = row[11]
        H1 = row[12]
        H2 = row[13]
        C1 = row[14]        
        C2 = row[15]
        A1 = row[16]
        A2 = row[17]
        LH1 = row[18]
        LH2_top = row[19]
        LH2_bottom = row[20]
        RH1 = row[21]
        RH2_top = row[22]        
        RH2_bottom = row[23]        
        U = row[24]        
        G = row[25]        
        UW = row[26]        
        SW = row[27]        
        K1_up = row[28]        
        K1_down = row[29]        
        K2_up = row[30]        
        K2_down = row[31]                
        Upper_size = row[32]
        Left_side_size = row[33]
        Right_side_size = row[34]                
        Angle = row[40]
        Top_pop1 = round(row[41] + 2, 1) #소수점 첫째자리까지만 나오게
        Top_pop2 = round(row[42] + 2, 1)         
        Kadapt =  row[52]   # K값 J값에 적용여부 넣기 1이면 넣기
        G_Setvalue = row[53]  # 손상민 소장 G값 강제로 적용 작지 52번째

        # 데이터 검증 및 기본값 설정        
        K1_up = validate_or_default(K1_up)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_down = validate_or_default(K2_down)
        U = validate_or_default(U)
        G = validate_or_default(G)        
        G_Setvalue = validate_or_default(G_Setvalue)        
        OP = validate_or_default(OP)
        Kadapt = validate_or_default(Kadapt)

        # print("작업소장 : ")
        # print(worker)
        # print("G_setvalue : ")
        # print(G_Setvalue)        
        if(worker=="손상민" and G_Setvalue > 0) :
            # 상부 돌출
            Top_pop1 = G_Setvalue
            Top_pop2 = G_Setvalue
            # print("G_setvalue : ")
            # print(G_Setvalue)
       
        if(Kadapt==1):
            K1_up = 0            
            K2_up = 0            
            K1_down = 0            
            K2_down = 0            
             
        if(Top_pop1<4):
            Top_pop1 = 4
        if(Top_pop2<4):
            Top_pop2 = 4                

        if(OP>0):             
            if(G > 0):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U + G - Vcut_rate*3
                X2 = Abs_Xpos + H1 + H2 + R 
                Y2 = Y1      
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Abs_Ypos + U - Vcut_rate * 2           
                line(doc, X2, Y2, X3, Y3, layer='레이져')          
            if(G < 1):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U  - Vcut_rate
                X2 = Abs_Xpos + H1 + H2 + R 
                Y2 = Y1      
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Y2

            X4 = X3
            Y4 = Abs_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')
            X5 = X4
            Y5 = Abs_Ypos - Top_pop2 + Vcut_rate        
            lineto(doc,  X5, Y5, layer='레이져')    
            X6 = X5 - H2
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='레이져')

            if(K1_up>0 or K2_up>0):
                X7 = Abs_Xpos + H1 + (R-OP)/2 + OP                
                Y7 = Abs_Ypos - JD2 + Vcut_rate
                lineto(doc,  X7, Y7, layer='레이져')
            else:
                X7 = Abs_Xpos + H1 + (R-OP)/2 + OP    
                Y7 = Abs_Ypos - JD2 + Vcut_rate * 2
                lineto(doc,  X7, Y7, layer='레이져')  
                
            if(K1_up>0 or K2_up>0 ):
                X8 = X7
                Y8 = Abs_Ypos - JD2 - K2_up + Vcut_rate*2
                lineto(doc,  X8, Y8, layer='레이져') 
                X9 = X8 
                Y9 = Abs_Ypos - JD2 - K2_up - UW + Vcut_rate*3     
                lineto(doc,  X9, Y9, layer='레이져')   
                X10 = Abs_Xpos + H1 + (R-OP)/2 
                Y10 = Abs_Ypos - JD1 - K1_up - UW + Vcut_rate*3
                lineto(doc,  X10, Y10, layer='레이져')      
                X11 = X10 
                Y11 =  Abs_Ypos - JD1 - K1_up + Vcut_rate*2
                lineto(doc,  X11, Y11, layer='레이져')                                      
            else:
                X8 = X7
                Y8 = Y7
                X9 = X8
                Y9 = Abs_Ypos - JD2 - UW + Vcut_rate*3            
                lineto(doc,  X9, Y9, layer='레이져')   
                X10 = Abs_Xpos + H1 + (R-OP)/2 
                Y10 = Abs_Ypos - JD1 - UW + Vcut_rate*3
                lineto(doc,  X10, Y10, layer='레이져')      
                X11 = X10                 
                Y11 =  Abs_Ypos - JD1 + Vcut_rate*2
                lineto(doc,  X11, Y11, layer='레이져')                                

            X12 = X11
            if(K1_up>0 or K2_up>0):
                Y12 = Abs_Ypos - JD1 + Vcut_rate 
            else:  
                Y12 = Abs_Ypos - JD1 + Vcut_rate * 2
            
            lineto(doc,  X12, Y12, layer='레이져')                   
            X13 = Abs_Xpos + H1
            Y13 = Abs_Ypos - Top_pop2 + Vcut_rate 
            lineto(doc,  X13, Y13, layer='레이져')      
            X14 = Abs_Xpos
            Y14 = Y13
            lineto(doc,  X14, Y14, layer='레이져')      
            X15 = X14 
            Y15 = Abs_Ypos
            lineto(doc,  X15, Y15, layer='레이져')     
            if(G > 0):         
                X16 = X15 
                Y16 = Abs_Ypos + U - Vcut_rate * 2   
                lineto(doc,  X16, Y16, layer='레이져')      
                lineto(doc,  X1, Y1, layer='레이져')                  
            if(G < 1):         
                X16 = X15 
                Y16 = Abs_Ypos + U - Vcut_rate 
                lineto(doc,  X1, Y1, layer='레이져')                  

            # if args.opt17:
            #     drawcircle(doc, X10 + 12.5, Y10 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이
            #     drawcircle(doc, X9 - 12.5, Y9 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이

            # 상판 절곡선
            if(G>0):
                line(doc, X16, Y16, X3, Y3,  layer='22')        
            line(doc, X15, Y15, X4, Y4,  layer='22')
            line(doc, X11, Y11, X8, Y8,  layer='22')
    
            # 상판 우측 절곡치수선
            if(G>0):        
                dim_vertical_right(doc, X2, Y2, X3, Y3,  100, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X4, Y4, X3, Y3,  50, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X4, Y4, X8, Y8,  50, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc,  X9, Y9, X8, Y8, extract_abs(X8,X5) + 50, 'dim', text_height=0.22, text_gap=0.07)   # 뒷날개
            dim_vertical_right(doc, X2, Y2,  X9, Y9,  extract_abs(X9,X2) + 100, 'dim', text_height=0.22, text_gap=0.07)             

            # 상판 좌측 절곡치수선
            if(JD1 != JD2 or K1_up != K2_up):
                if(G>0):                    
                    dim_vertical_left(doc, X1, Y1, X16, Y16,  extract_abs(X1,X16) + 100, 'dim', text_height=0.22, text_gap=0.07)  
                dim_vertical_left(doc, X15, Y15, X16, Y16,  extract_abs(X15,X16) + 50, 'dim', text_height=0.22, text_gap=0.07)                              
                dim_vertical_left(doc, X11, Y11, X15, Y15,  extract_abs(X11,X15) + 50 , 'dim', text_height=0.22, text_gap=0.07)              
                dim_vertical_left(doc, X10, Y10, X11, Y11,  H1 + (R-OP)/2 + extract_abs(X10,X11) + 50 , 'dim', text_height=0.22, text_gap=0.07)              
                dim_vertical_left(doc, X10, Y10, X1, Y1,  extract_abs(X10,X1) + 200, 'dim', text_height=0.22, text_gap=0.07)                          
            else:    
                dim_vertical_left(doc, X10, Y10, X11, Y11,   H1 + (R-OP)/2  + 10, 'dim', text_height=0.22,  text_gap=0.07) 
                dim_vertical_left(doc, X10, Y10, X15, Y15,   H1 + (R-OP)/2  + 100 , 'dim', text_height=0.22,  text_gap=0.07) 
                dim_vertical_left(doc, X10, Y10, X16, Y16,   H1 + (R-OP)/2  + 150, 'dim', text_height=0.22,  text_gap=0.07) 

            # 상판 하부 치수선 3개
            if(Angle>1):
                dim_linear(doc,  X9, Y9, X6, Y6, "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X10, Y10, X9, Y9,  "", 100 + extract_abs(JD1+K1_up, JD2+K2_up),  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            if(Angle>1):
                dim_linear(doc,  X13, Y13, X10, Y10,  "",  extract_abs(Y6,Y9) + 100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            # 상판 돌출부위 3.4 치수선
            dim_vertical_right(doc, X15, Y15, X13, Y13, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_left(doc, X4, Y4, X6, Y6, 100, 'dim', text_height=0.22,  text_gap=0.07)  

            # 기둥과의 접선 부위 aligned 표현 각도표현
            dim_linear(doc, X13, Y13, X11, Y11,   "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) # 좌측
            if(Angle>1):
                dim_angular(doc, X12, Y12, X13, Y13, X11, Y11, X12, Y12 + 10,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다        
            dim_linear(doc, X8, Y8, X6, Y6,  "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) # 좌측
            if(Angle>1):            
                dim_angular(doc, X8, Y8  , X7, Y7 + 10,  X7, Y7,  X6, Y6,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다

            # 상판 K1 K2 값 부분 치수선 
            if(K1_up > 3 or K2_up > 3 ):
                dim_vertical_right(doc, X11, Y11, X12, Y12, 80, 'dim', text_height=0.22,  text_gap=0.07)  
                dim_vertical_left(doc, X8, Y8, X7, Y7, 80, 'dim', text_height=0.22,  text_gap=0.07)  

            # 전개도 상부 치수선
            dim_linear(doc,  X1, Y1, X2, Y2, "", 120,  direction="up", layer='dim')
            dim_linear(doc,  X13, Y13, X6, Y6,  "", extract_abs(Y1,Y13) + 50,  direction="up", layer='dim')
            dim_linear(doc,  X1, Y1, X13, Y13, "",  50,  direction="up", layer='dim')
            dim_linear(doc,  X6, Y6, X2, Y2,  "",  extract_abs(Y2,Y6) + 50,  direction="up", layer='dim')

            # 상부 Vcut 치수 표기
            if(U>0 and G>0):   
               dim_vertical_left(doc, X1+H1+OP/2, Y1, X1+H1+OP/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  

            # 상판 좌우가 다를때 Vcut 상부에 치수선 표기
            # if((JD1 != JD2 or K1_up != K2_up) and G>0):
            #     dim_vertical_left(doc, X1 + R/2, Y1, X1 + R/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            # 문구 설정
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"        
            Textstr = f"({Pre_text}{str(Floor_des)})"       

            X2 = Abs_Xpos + (H1 + R + H2)/2 - len(Textstr)*50/2
            Y2 = Abs_Ypos - JD1/1.5
            draw_Text(doc, X2, Y2, 50, str(Textstr), '레이져')

            # 문구            
            sizedes = Upper_size                  
            
            Textstr = f"Part Name : 상판({Pre_text}{str(Floor_des)})"    
            draw_Text(doc, X9+150, Y9 + 50 - 220, 28, str(Textstr), '0')
            Textstr = f"Mat.Spec : {normal_material}"    
            draw_Text(doc, X9+150, Y9 - 220, 28, str(Textstr), '0')
            Textstr = f"Size : {sizedes}"    
            draw_Text(doc, X9+150, Y9 - 50 - 220 , 28, str(Textstr), '0')
            Textstr = f"Quantity : 1 EA"    
            draw_Text(doc, X9+150, Y9 - 100 - 220 , 28, str(Textstr), '0')                        

            # 도면틀 넣기
            BasicXscale = 2560
            BasicYscale = 1550
            TargetXscale = 800 + R + U * 2 + 200
            TargetYscale = 300 + JD1 + G + U + UW
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            insert_frame(X1 - (U + 580) ,Y1- ( 900 + JD1 + K1_up + G + U + UW) , frame_scale, "top_normal", workplace)        

            ###########################################################################################################################################################################
            # 2UP 상판 좌측 단면도    (좌우가 다른 모양일때 단면도를 그려준다)
            ###########################################################################################################################################################################
            if(JD1 != JD2 or K1_up != K2_up):
                Section_Xpos = Abs_Xpos - H1 - 400           

                if(G > 0):
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos - G
                    X2 = Section_Xpos
                    Y2 = Abs_Ypos  

                    line(doc, X1, Y1, X2, Y2, layer='0')                    
                    dim_vertical_left(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)     
                    X3 = Section_Xpos + U
                    Y3 = Abs_Ypos  
                    lineto(doc,  X3, Y3, layer='0')    
                    dim_linear(doc,  X2, Y2, X3, Y3, "", 50,  direction="up", layer='dim')                
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                    dim_leader_line(doc, X3, Y3 , X2-130 , Y2-100, "V-cut 3개", layer="dim", style='0')                
                if(G < 1):
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos 
                    X2 = Section_Xpos
                    Y2 = Abs_Ypos                  
                    X3 = Section_Xpos + U
                    Y3 = Abs_Ypos   
                    line(doc, X2, Y2, X3, Y3, layer='0')                  
                    dim_linear(doc,  X2, Y2, X3, Y3, "", 50,  direction="up", layer='dim')    
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                    dim_leader_line(doc, X3, Y3 , X2-130 , Y2-100, "V-cut 2개", layer="dim", style='0')                                
                
                X4 = X3 
                Y4 = Y3 - JD1
                lineto(doc,  X4, Y4, layer='0')                   
                if(K1_up > 0):
                    X5 = X4 
                    Y5 = Y4 - K1_up
                    lineto(doc,  X5, Y5, layer='0')   
                    dim_vertical_right(doc,  X3, Y3, X4, Y4, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X4, Y4, X5, Y5, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X3, Y3, X5, Y5, 110,'dim' ,text_height=0.22,  text_gap=0.07)                                     
                else:
                    dim_vertical_right(doc,  X3, Y3, X4, Y4, 110,'dim' ,text_height=0.22,  text_gap=0.07)     
                    X5 = X4 
                    Y5 = Y4

                drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')         
                X6 = X5 - UW
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')   
                dim_linear(doc,  X6, Y6, X5, Y5,  "", 80,  direction="down", layer='dim')            
                
                if(K1_up > 0):                 
                    # K값 부분 라인 그리기
                    line(doc, X4, Y4, X4 - UW, Y4, layer='22')     
                
                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JD1 != JD2 or K1_up != K2_up ):
                    rectangle(doc, X4+50, Y5+(JD1+K1_up)/2 - 75 , X4 + 50 + 100, Y5+(JD1+K1_up)/2 + 75, layer='0')               
                    rectangle(doc, X4+285,  Y5+(JD1+K1_up)/2 - 75 , X4+285+100, Y5+(JD1+K1_up)/2 + 75 , layer='0')                                                         
                
            ################################################################################################################################################################
            # 2UP 상판 우측 단면도 
            ################################################################################################################################################################
            Section_Xpos = Abs_Xpos + R + H1 + H2 + 400

            if(G>0):
                X1 = Section_Xpos 
                Y1 = Abs_Ypos - G
                X2 = Section_Xpos
                Y2 = Abs_Ypos  

                line(doc, X1, Y1, X2, Y2, layer='0')                    
                dim_vertical_right(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)     
                X3 = Section_Xpos - U
                Y3 = Abs_Ypos  
                lineto(doc,  X3, Y3, layer='0')    
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim')
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                dim_leader_line(doc, X3, Y3 , X3+50 , Y2-100, "V-cut 3개", layer="dim", style='0')                

            if(G<1):
                X1 = Section_Xpos 
                Y1 = Abs_Ypos
                X2 = Section_Xpos
                Y2 = Abs_Ypos  
                X3 = Section_Xpos - U
                Y3 = Abs_Ypos  
                line(doc, X1, Y1, X3, Y3, layer='0')                            
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim')    
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                dim_leader_line(doc, X3, Y3 , X3 + 50 , Y2-100, "V-cut 2개", layer="dim", style='0')                            
                
            X4 = X3 
            Y4 = Y3 - JD2
            lineto(doc,  X4, Y4, layer='0')   
            if(K1_up > 0):
                X5 = X4 
                Y5 = Y4 - K2_up
                lineto(doc,  X5, Y5, layer='0')   
                dim_vertical_left(doc,  X3, Y3, X4, Y4, 60,'dim' ,text_height=0.22,  text_gap=0.07)                     
                dim_vertical_left(doc,  X4, Y4, X5, Y5, 60,'dim' ,text_height=0.22,  text_gap=0.07)     
                dim_vertical_left(doc,  X3, Y3, X5, Y5, 110,'dim' ,text_height=0.22,  text_gap=0.07)                                     
            else:
                dim_vertical_left(doc,  X3, Y3, X4, Y4, 110,'dim' ,text_height=0.22,  text_gap=0.07)                     
                X5 = X4 
                Y5 = Y4

            X6 = X5 + UW
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='0')   
            dim_linear(doc,  X5, Y5, X6, Y6,   "", 80,  direction="down", layer='dim')
            
            drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
            
            if(K2_up > 0):                 
                # K값 부분 라인 그리기
                line(doc, X4, Y4, X4 + UW, Y4, layer='22')      

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JD1 != JD2 or K1_up != K2_up ):
                rectangle(doc, X4-50, Y5+(JD2+K2_up)/2 - 75 , X4 - 50 - 100, Y5+(JD2+K2_up)/2 + 75, layer='0')               
                rectangle(doc, X4-285,  Y5+(JD2+K2_up)/2 - 75 , X4-285-100,  Y5+(JD2+K2_up)/2 + 75 , layer='0')                                                         
            
            # 다음 도면 간격 계산
            Abs_Ypos -= 4000

    Abs_Xpos = 73000 + 5000
    Abs_Ypos = 500

    ############################################################################################################################################################################
    # 2UP 좌기둥 전개도
    ############################################################################################################################################################################

    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        OP = row[2]
        R = row[3]
        JD1 = row[4]
        JD2 = row[5]
        JB1_up = row[6]
        JB1_down = row[7]
        JB2_up = row[8]
        JB2_down = row[9]
        jb1_gap = row[10]
        jb2_gap = row[11]
        H1 = row[12]
        H2 = row[13]
        C1 = row[14]        
        C2 = row[15]
        A1 = row[16]
        A2 = row[17]
        LH1 = row[18]
        LH2_top = row[19]
        LH2_bottom = row[20]
        RH1 = row[21]
        RH2_top = row[22]        
        RH2_bottom = row[23]        
        U = row[24]        
        G = row[25]        
        UW = row[26]        
        SW = row[27]        
        K1_up = row[28]        
        K1_down = row[29]        
        K2_up = row[30]        
        K2_down = row[31]                
        Upper_size = row[32]
        Left_side_size = row[33]
        Right_side_size = row[34]                
        Angle = row[40]
        Top_pop1 = round(row[41] + 2, 1) #소수점 첫째자리까지만 나오게
        Top_pop2 = round(row[42] + 2, 1)         
        Kadapt =  row[52]   # K값 J값에 적용여부 넣기 1이면 넣기
        G_Setvalue = row[53]  # 손상민 소장 G값 강제로 적용 작지 52번째

        if(Kadapt==1):
            K1_up = 0
            K1_down = 0
            K2_up = 0
            K2_down = 0   

        OP = validate_or_default(OP)

        if(OP>0):          
            X1 = Abs_Xpos 
            Y1 = Abs_Ypos + JB1_up + K1_up +  SW - Vcut_rate * 5
            X2 = Abs_Xpos + LH2_top - 2
            Y2 = Y1
            X3 = X2            
            Y3 = Y2 - SW + Vcut_rate * 2             
            X4 = X3 + 2
            Y4 = Y3            
            X5 = X4
            Y5 = Y4 + (JB1_down - JB1_up)            
            X6 = X5 
            Y6 = Y5 + SW - Vcut_rate * 2               
            X7 = Abs_Xpos + LH2_top + LH2_bottom  
            Y7 = Abs_Ypos + JB1_down + SW - Vcut_rate * 5            
            X8 = X7 
            Y8 = Abs_Ypos + JB1_down - Vcut_rate * 3            
            X9 = Abs_Xpos + LH1 
            Y9 = Abs_Ypos            

            if(A1>0):       
                X10 = X9
                Y10 = Abs_Ypos - C1 + Vcut_rate * 2                
                X11 = X10
                Y11 = Y10 - A1 + Vcut_rate                 
                X12 = Abs_Xpos             
                Y12 = Abs_Ypos - C1 - A1 + Vcut_rate * 3                
                X13 = Abs_Xpos             
                Y13 = Abs_Ypos - C1 + Vcut_rate * 2                

            if(A1<1):       
                X10 = X9
                Y10 = Abs_Ypos - C1 + Vcut_rate                
                X11 = X10
                Y11 = Y10                
                X12 = Abs_Xpos             
                Y12 = Y11                
                X13 = Abs_Xpos             
                Y13 = Abs_Ypos - C1 + Vcut_rate * 1                

            X14 = Abs_Xpos 
            Y14 = Abs_Ypos

            X15 = Abs_Xpos 
            Y15 = Abs_Ypos + JB1_up + K1_up  - Vcut_rate * 3 

            prev_x, prev_y = X1, Y1  # 첫 번째 점으로 초기화
            for i in range(1, 15 + 1):
                curr_x, curr_y = eval(f'X{i}'), eval(f'Y{i}')
                line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                prev_x, prev_y = curr_x, curr_y
                # print(prev_x)
            # 마지막으로 첫번째 점과 연결
            line(doc, X1, Y1, prev_x, prev_y, layer="레이져")

            # 좌기둥 절곡선 layer 22 
            line(doc, X15, Y15, X3, Y3,  layer='22')
            line(doc, X5, Y5, X8, Y8,  layer='22')
            line(doc, X14, Y14, X9, Y9,  layer='22')
            if(A1>0):
                line(doc, X13, Y13, X10, Y10,  layer='22')         

            # 뒷날개쪽 치수선
            dim_linear(doc,  X1, Y1, X7, Y7, "", 120 + extract_abs(Y1,Y7),  direction="up", layer='dim')                           
            dim_linear(doc,  X1, Y1, X6, Y6, "", 60  + extract_abs(Y1,Y7),  direction="up", layer='dim')                           
            dim_linear(doc,  X6, Y6, X7, Y7, "", 60,  direction="up", layer='dim')                    

            # 도어두께 표시 상부
            dim_vertical_right(doc, X4,  Y4, X5, Y5, 100, 'dim', text_height=0.22, text_gap=0.01)                         

            # 기둥 우측 절곡치수선
            dim_vertical_right(doc, X7,  Y7, X8, Y8, 60  + extract_abs(X8,X9) , 'dim', text_height=0.22, text_gap=0.01)                 
            # JB
            dim_vertical_right(doc, X8,  Y8, X9, Y9, 60  + extract_abs(X8,X9) , 'dim', text_height=0.22, text_gap=0.01)                              

            dim_vertical_right(doc, X9,  Y9, X10, Y10, 60 + extract_abs(X8,X9) , 'dim', text_height=0.22, text_gap=0.01)                                          
            if(A1>0):
                dim_vertical_right(doc, X10,  Y10, X11, Y11,  120 + extract_abs(X8,X9), 'dim', text_height=0.22, text_gap=0.01)  
                # V컷 치수표현
                dim_vertical_left(doc, X9,  Y9, X11, Y11,   120 + extract_abs(X8,X9), 'dim', text_height=0.22, text_gap=0.01)  
            dim_vertical_right(doc, X7,  Y7, X11, Y11,   extract_abs(X9,X8) + 200 , 'dim', text_height=0.22, text_gap=0.01)      

            # 기둥 좌측 절곡치수선  
            dim_vertical_left(doc, X1, Y1, X15, Y15,  80, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X14, Y14, X15, Y15, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_left(doc, X14, Y14, X13, Y13,  80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_left(doc, X12, Y12, X13, Y13,  150 , 'dim', text_height=0.22, text_gap=0.01)
            dim_vertical_left(doc, X1, Y1, X12,  Y12,  230 , 'dim', text_height=0.22, text_gap=0.01)

            # 기둥 하부 치수선 2    
            dim_linear(doc,   X12, Y12, X11, Y11, "", 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   

            # LH1, LH2 치수크기 차이로 인해 설치
            if (LH1 > LH2_top+LH2_bottom) :
                dim_linear(doc,  X7, Y7, X9, Y9, "",  100 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)       
            if (LH1 < LH2_top+LH2_bottom) :
                dim_linear(doc,  X11, Y11, X8, Y8, "", 100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"       

            Xpos = Abs_Xpos + 200  
            Ypos = Abs_Ypos + JB1_up / 3.5
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-좌"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')        

            Xpos = Abs_Xpos + LH1/2 - 200
            Ypos = Abs_Ypos + JB1_up / 3.5 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-좌"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/2 - 200
            Ypos = Abs_Ypos + JB1_up / 3.5 + 30 - 40
            Textstr = f"Mat.Spec : {normal_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/1.2 -200
            Ypos = Abs_Ypos + JB1_up / 3.5 + 30
            Textstr = f"Size : {Left_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/1.2 - 200
            Ypos = Abs_Ypos + JB1_up / 3.5 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')       

            # 기둥 좌우 도면틀 넣기
            BasicXscale = 4487
            BasicYscale = 2770
            # 우측단면도가 있을 경우 감안해야 함. 
            TargetXscale = max(LH1, LH2_top+LH2_bottom) + 800*2

            TargetYscale = 250*4 + (JB1_up + C1 + A1 + K1_up + SW) * 2
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            # 우측 단면도가 있는 경우 도면틀 기본점
            insert_frame(X1 -  1110 , Abs_Ypos - (JB2_up+C2+A2+K2_up + SW + 1300 + C1 + A1 ) , frame_scale, "side_normal", workplace)                

            ##############################################################################################################################################################
            # 2UP 좌기둥 좌측 단면도    
            ##############################################################################################################################################################
            Section_Xpos = Abs_Xpos - 500
            Section_Ypos = Abs_Ypos

            X1 = Section_Xpos - C1 
            Y1 = Section_Ypos + A1
            X2 = X1
            Y2 = Y1 - A1
            X3 = X2 + C1
            Y3 = Y2
            X4 = X3
            Y4 = Y3 + JB1_up + K1_up
            X5 = X4
            Y5 = Y4 + (JB1_down - JB1_up)  
            X6 = X5 - SW
            Y6 = Y5
            X7 = X4 - SW
            Y7 = Y4
            
            line(doc, X1, Y1, X2, Y2, layer='0')  
            lineto(doc,  X3, Y3, layer='0')               
            lineto(doc,  X4, Y4, layer='0')               
            lineto(doc,  X5, Y5, layer='0')               
            lineto(doc,  X6, Y6, layer='0')               
            line(doc, X4, Y4, X7, Y7, layer='0')  
           
            if(A1>0):
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6')             
                dim_vertical_left(doc, X1, Y1, X2, Y2, 50, 'dim', text_height=0.22,  text_gap=0.07)                 
            if(C1>0):                
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6') 
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 80,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)                 

            if(A1>0 and C1>0):                
                dim_leader_line(doc, X3, Y3 , X3+50, Y3-70, "V-cut 2개소", layer="dim", style='0')
            else:
                if(C1>0):
                    dim_leader_line(doc, X3, Y3 , X3+50, Y3-70, "V-cut", layer="dim", style='0')

            # JB
            dim_vertical_right(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X5, Y5, X4, Y4, 80, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X3, Y3, X5, Y5, 160, 'dim', text_height=0.22, text_gap=0.07)  
            # 뒷날개
            if(SW>0):
               dim_linear(doc,   X6, Y6,  X5, Y5, "", 100,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)    
            
            ######################################################################################################################################################
            # 2UP 우기둥 전개도
            ######################################################################################################################################################

            Abs_Ypos -= 2000 
            R_Ypos = Abs_Ypos + 1550 - (C1 + C2 + A1 + A2)             

            if(OP>0):          
                if(A2>0):  
                    X1 = Abs_Xpos 
                    Y1 = R_Ypos + C2 + A2 - Vcut_rate * 3

                    X2 = Abs_Xpos + RH1
                    Y2 = Y1
                    
                    X3 = X2            
                    Y3 = Y2 - A2 + Vcut_rate

                    X4 = X3
                    Y4 = Y3 - C2 + Vcut_rate*2               

                    X5 = Abs_Xpos + RH2_top + RH2_bottom 
                    Y5 = Y4 - JB2_down + Vcut_rate * 3   

                    X6 = X5 
                    Y6 = Y5 - SW + Vcut_rate * 2
                    
                    X7 = Abs_Xpos + RH2_top 
                    Y7 = Y6
                    
                    X8 = X7 
                    Y8 = Y7 + SW - Vcut_rate * 2 
                    
                    X9 = X8
                    Y9 = Y8 + (JB2_down - JB2_up)

                    X10 = X9 - 2
                    Y10 = Y9
                                
                    X11 = X10
                    Y11 = Y10 - SW + Vcut_rate * 2
                    
                    X12 = Abs_Xpos             
                    Y12 = R_Ypos - JB2_up - SW + Vcut_rate * 5
                
                    X13 = Abs_Xpos             
                    Y13 = Y12 + SW - Vcut_rate * 2                              

                    X14 = Abs_Xpos 
                    Y14 = R_Ypos                   

                    X15 = Abs_Xpos 
                    Y15 = Y14 + C2 - Vcut_rate * 2        

                if(A2 < 1):  
                    X1 = Abs_Xpos 
                    Y1 = R_Ypos + C2 - Vcut_rate 

                    X2 = Abs_Xpos + RH1
                    Y2 = Y1
                    
                    X3 = X2            
                    Y3 = R_Ypos + C2 - Vcut_rate 

                    X4 = X3
                    Y4 = R_Ypos

                    X5 = Abs_Xpos + RH2_top + RH2_bottom 
                    Y5 = Y4 - JB2_down + Vcut_rate * 3  

                    X6 = X5 
                    Y6 = Y5 - SW + Vcut_rate * 2
                    
                    X7 = Abs_Xpos + RH2_top 
                    Y7 = Y6
                    
                    X8 = X7 
                    Y8 = Y7 + SW - Vcut_rate * 2 
                    
                    X9 = X8
                    Y9 = Y8 + (JB2_down - JB2_up)

                    X10 = X9 - 2
                    Y10 = Y9
                                
                    X11 = X10
                    Y11 = Y10 - SW + Vcut_rate * 2
                    
                    X12 = Abs_Xpos             
                    Y12 = R_Ypos - JB2_up - SW + Vcut_rate * 5
                
                    X13 = Abs_Xpos             
                    Y13 = Y12 + SW - Vcut_rate * 2                              

                    X14 = Abs_Xpos 
                    Y14 = R_Ypos                   

                    X15 = Abs_Xpos 
                    Y15 = Y14 + C2 - Vcut_rate          

                prev_x, prev_y = X1, Y1  # 첫 번째 점으로 초기화
                for i in range(1, 15 + 1):
                    curr_x, curr_y = eval(f'X{i}'), eval(f'Y{i}')
                    line(doc, prev_x, prev_y, curr_x, curr_y, layer="레이져")
                    prev_x, prev_y = curr_x, curr_y
                    # print(prev_x)
                # 마지막으로 첫번째 점과 연결
                line(doc, X1, Y1, prev_x, prev_y, layer="레이져")
                
                # 좌기둥 절곡선 layer 22 
                line(doc, X15, Y15, X3, Y3,  layer='22')
                line(doc, X14, Y14, X4, Y4,  layer='22')
                line(doc, X13, Y13, X10, Y10,  layer='22')
                line(doc, X8, Y8, X5, Y5,  layer='22')
                if(A2>0):
                    line(doc, X15, Y15, X3, Y3,  layer='22')         

                # 뒷날개쪽 치수선
                dim_linear(doc,  X12, Y12, X6, Y6, "", 120 + extract_abs(Y12,Y7),  direction="down", layer='dim')                           
                dim_linear(doc,  X12, Y12, X7, Y7, "", 60  + extract_abs(Y12,Y7),  direction="down", layer='dim')                           
                dim_linear(doc,  X7, Y7, X6, Y6, "", 60,  direction="down", layer='dim')                    

                # 도어두께 표시 상부
                dim_vertical_right(doc, X9,  Y9, X8, Y8, 100, 'dim', text_height=0.22, text_gap=0.01)                         

                # 기둥 우측 절곡치수선 C
                dim_vertical_right(doc, X3,  Y3, X4, Y4, 60  + extract_abs(X4,X5) , 'dim', text_height=0.22, text_gap=0.01)                 
                # JB
                dim_vertical_right(doc, X4,  Y4, X5, Y5, 60  + extract_abs(X4,X5) , 'dim', text_height=0.22, text_gap=0.01)                              
                dim_vertical_right(doc, X5,  Y5, X6, Y6, 60 + extract_abs(X4,X5) , 'dim', text_height=0.22, text_gap=0.01)                                          
                if(A1>0):
                    dim_vertical_right(doc, X2,  Y2, X3, Y3,  120 + extract_abs(X4,X5), 'dim', text_height=0.22, text_gap=0.01)                     
                dim_vertical_right(doc, X2,  Y2, X6, Y6,   extract_abs(X4,X5) + 200 , 'dim', text_height=0.22, text_gap=0.01)      

                # 기둥 좌측 절곡치수선  
                if(A1>0):
                    dim_vertical_left(doc, X1, Y1, X15, Y15,  150 , 'dim', text_height=0.22, text_gap=0.01)
                # C    
                dim_vertical_left(doc, X15, Y15, X14,  Y14,  80 , 'dim', text_height=0.22, text_gap=0.01)                                        
                # JB    
                dim_vertical_left(doc, X14, Y14, X13, Y13, 80 , 'dim', text_height=0.22, text_gap=0.01)
                # 뒷날개                      
                dim_vertical_left(doc, X13, Y13, X1, Y12,  80 , 'dim', text_height=0.22, text_gap=0.01)  
                # 좌측 전체 치수선
                dim_vertical_left(doc, X1, Y1, X12, Y12,  220 , 'dim', text_height=0.22, text_gap=0.01)  

                # 기둥 상부 치수선 2    
                dim_linear(doc,   X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)   

                # RH1, RH2 치수크기 차이로 인해 설치
                if (RH1 > RH2_top+RH2_bottom) :
                    dim_linear(doc,  X6, Y6,  X4, Y4, "",  100  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)       
                if (RH1 < RH2_top+RH2_bottom) :
                    dim_linear(doc,   X2, Y2, X5, Y5, "", 100   ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      

                # 전개도 문구        
                if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                    Pre_text = ""
                else:
                    Pre_text = str(Floor_mc) + "-"       

                Xpos = Abs_Xpos + 200  
                Ypos = R_Ypos - JB2_up / 3.5
                Textstr = f"코팅상"       
                draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
                Textstr = f"({Pre_text}{str(Floor_des)})-우"         
                draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')        

                Xpos = Abs_Xpos + RH1/2 - 200
                Ypos = R_Ypos - JB2_up / 3.5 + 30
                Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-우"       
                draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
                Xpos = Abs_Xpos + RH1/2 - 200
                Ypos = R_Ypos - JB2_up / 3.5 + 30 - 40
                Textstr = f"Mat.Spec : {normal_material} "       
                draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
                Xpos = Abs_Xpos + RH1/1.2 -200
                Ypos = R_Ypos - JB2_up / 3.5 + 30
                Textstr = f"Size : {Right_side_size} "       
                draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
                Xpos = Abs_Xpos + RH1/1.2 - 200
                Ypos = R_Ypos - JB2_up / 3.5 + 30 - 40
                Textstr = f"Quantity : 1 EA "       
                draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')       

                ##############################################################################################################################################################
                # 2UP 우기둥 좌측 단면도    
                ##############################################################################################################################################################
                Section_Xpos = Abs_Xpos - 500
                Section_Ypos = R_Ypos

                X1 = Section_Xpos - C2
                Y1 = Section_Ypos - A2
                X2 = X1
                Y2 = Y1 + A2
                X3 = X2 + C2
                Y3 = Y2
                X4 = X3
                Y4 = Y3 - JB2_up - K2_up
                X5 = X4
                Y5 = Y4 - (JB2_down - JB2_up)
                X6 = X5 - SW
                Y6 = Y5
                X7 = X4 - SW
                Y7 = Y4
                
                prev_x, prev_y = X1, Y1  # 첫 번째 점으로 초기화
                for i in range(1, 6 + 1):
                    curr_x, curr_y = eval(f'X{i}'), eval(f'Y{i}')
                    line(doc, prev_x, prev_y, curr_x, curr_y, layer="0")
                    prev_x, prev_y = curr_x, curr_y

                line(doc, X4, Y4, X7, Y7, layer='0')  
            
                if(A1>0):
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6')             
                    dim_vertical_left(doc, X1, Y1, X2, Y2, 50, 'dim', text_height=0.22,  text_gap=0.07)                 
                if(C1>0):                
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6') 
                    dim_linear(doc,   X2, Y2, X3, Y3,  "", 80,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)                 

                if(A1>0 and C1>0):                
                    dim_leader_line(doc, X3, Y3 , X3+50, Y3+70, "V-cut 2개소", layer="dim", style='0')
                else:
                    if(C1>0):
                        dim_leader_line(doc, X3, Y3 , X3+50, Y3+70, "V-cut", layer="dim", style='0')
                # JB
                dim_vertical_right(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                dim_vertical_right(doc, X5, Y5, X4, Y4, 80, 'dim', text_height=0.22, text_gap=0.07)  
                dim_vertical_right(doc, X3, Y3, X5, Y5, 160, 'dim', text_height=0.22, text_gap=0.07)  
                # 뒷날개
                if(SW>0):
                    dim_linear(doc,   X6, Y6,  X5, Y5, "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)    

            Abs_Ypos -= 2000 

#############################
# 4CO와이드 작도 사이드오픈
#############################
def execute_fourcowide():     
    # 시트 선택 (시트명을 지정)
    sheet_name = '4CO와이드제작'  # 원하는 시트명으로 변경
    sheet = workbook[sheet_name]

    # 2차원 배열을 저장할 리스트 생성
    data_2d = []

    # 1행부터 20행까지의 값을 2차원 배열에 저장
    for row_num in range(7, sheet.max_row + 1):  # 7행부터 20행까지 반복
        cell_value = sheet.cell(row=row_num, column=2).value  # 2열의 값 가져오기
        if cell_value is not None and cell_value != 'F':  # 값이 None이 아닌 경우에만 배열에 추가
            row_values =  [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, sheet.max_column + 1)]        
            data_2d.append(row_values)

    # 엑셀 파일 닫기
    workbook.close()

    if not data_2d:
        print(" 해당 데이터가 없습니다.")
        return 

    #################################################################### 
    # 4CO와이드  상판 전개도
    ####################################################################    
    Abs_Xpos = 52000
    Abs_Ypos = 0
    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        MH1 = row[2]
        MH2 = row[3]
        OP1 = row[4]
        OP2 = row[5]
        OP3 = row[6]
        R = row[7]
        JD1 = row[8]
        JD2 = row[9]
        JD3 = row[10]
        JB1_up = row[11]
        JB1_down = row[12]
        JB2_up = row[13]
        JB2_down = row[14]
        Top_gap = row[15]
        Side_gap = row[16]        
        H1 = row[17]
        H2 = row[18]
        C1 = row[19]        
        C2 = row[20]
        B1 = row[21]
        B2 = row[22]
        A1 = row[23]
        A2 = row[24]
        LH1 = row[25]
        LH2 = row[26]
        RH1 = row[27]
        RH2 = row[28]        
        U = row[29]        
        G = row[30]        
        UW = row[31]        
        SW = row[32]         
        K1_up = row[33]        
        K1_down = row[34]        
        K2_up = row[35]        
        K2_down = row[36]                
        SMH1 = row[37]                
        SMH2 = row[38]                
        G_check = row[39]
        Upper_size = row[40]
        Left_side_size = row[41]
        Right_side_size = row[42]
        E1 = row[43]
        E2 = row[44]
        HPI_height = row[45]
        SW1 = row[46]
        SW2 = row[47]
        SWAngle = row[new_func()]
        A1_var = row[49]
        A2_var = row[50]
        Angle = row[51]
        Kadapt = row[65]  # K값 적용여부 1 이면 K값 적용하지 말것

        if(Kadapt==1):
            K1_up = 0
            K2_up = 0                           

        if(K1_up>=10 and K2_up>=10 ):
            # 추영덕소장 E1, E2값 무시함 위의 조건이라면
            E1 = 0
            E2 = 0

        if(OP1>0):
            # 상판 기둥과 만나는 돌출값 B값과 연결되는 부분
            # Top_Pop1 = G-B1-1.5 
            # Top_Pop2 = G-B2-1.5 
            Top_Pop1 = G_check
            Top_Pop2 = G_check

            if(Top_Pop1<4):
                Top_Pop1 = 4
            if(Top_Pop2<4):
                Top_Pop2 = 4

            # 손상민소장의 경우 적용
            if(worker=="손상민" and U>2):
                draw_Uframe(Abs_Xpos, Abs_Ypos, U,G ,R, C1, C2, Floor_mc, Floor_des )  
            if(H1<1 or H2<1):
                U = 0
                G = 0                        

            SWAngle_radians = math.radians(SWAngle)
            if(SWAngle>0):
                SWAngle_var =  SW2 * math.cos(SWAngle_radians)        
            Angle_radians = math.radians(Angle)
            Right_Angle_radians = math.radians(Angle)

            if(Angle>0):
                Angle_cos = math.cos(Angle_radians)        
                Right_Angle_cos = math.cos(Right_Angle_radians)        
                Angle_sin = math.sin(Angle_radians)    
                Right_Angle_sin = math.sin(Right_Angle_radians)    

            if(U > 0):
                X1 = Abs_Xpos - H1
                Y1 = Abs_Ypos + MH1 + G + U - Vcut_rate*5
                X2 = Abs_Xpos  + H2 + R 
                Y2 = Abs_Ypos + MH2 + G + U - Vcut_rate*5
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Abs_Ypos + MH2 + G  - Vcut_rate * 4
                lineto(doc, X3, Y3, layer='레이져')          
                X4 = X3
                Y4 = Abs_Ypos + MH2 + G - Top_Pop2 - Vcut_rate * 3
                lineto(doc, X4, Y4, layer='레이져')          
                X5 = X4 - H2 + (G - Top_Pop2) * Right_Angle_sin
                Y5 = Y4
                lineto(doc, X5, Y5, layer='레이져')          
                X6 = Abs_Xpos + R 
                Y6 = Abs_Ypos + MH2 - Vcut_rate * 2
                lineto(doc,  X6, Y6, layer='레이져')
                X7 = X6
                Y7 = Abs_Ypos
                lineto(doc,  X7, Y7, layer='레이져')     

            if(U < 1):
                if(G>0):
                    X1 = Abs_Xpos - H1
                    Y1 = Abs_Ypos + MH1 + G - Vcut_rate*3
                    X2 = Abs_Xpos  + H2 + R 
                    Y2 = Y1
                    line(doc, X1, Y1, X2, Y2, layer='레이져')       
                    X3 = X2
                    Y3 = Y2
                    lineto(doc, X3, Y3, layer='레이져')          
                    X4 = X3
                    Y4 = Abs_Ypos + MH2 + G - Top_Pop2 - Vcut_rate * 3
                    lineto(doc, X4, Y4, layer='레이져')          
                    X5 = X4 - H2 + (G - Top_Pop2) * Right_Angle_sin
                    Y5 = Y4
                    lineto(doc, X5, Y5, layer='레이져')          
                    X6 = Abs_Xpos + R 
                    Y6 = Abs_Ypos + MH2 - Vcut_rate * 2
                    lineto(doc,  X6, Y6, layer='레이져')
                    X7 = X6
                    Y7 = Abs_Ypos
                    lineto(doc,  X7, Y7, layer='레이져')   
                else:
                    X1 = Abs_Xpos 
                    Y1 = Abs_Ypos + MH1 - Vcut_rate
                    X2 = Abs_Xpos + R 
                    Y2 = Abs_Ypos + MH2 - Vcut_rate
                    line(doc, X1, Y1, X2, Y2, layer='레이져')       
                    X3 = X2
                    Y3 = Y2            
                    X4 = X3
                    Y4 = Y3            
                    X5 = X4
                    Y5 = Y4            
                    X6 = X5
                    Y6 = Y5            
                    X7 = X6
                    Y7 = Abs_Ypos
                    lineto(doc,  X7, Y7, layer='레이져')            

            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                X8 = Abs_Xpos + (R - OP1 - OP2 - OP3 - 4)/2  + OP1 + OP2 + OP3 + 4      
                if(E2>0):
                    Y8 = Abs_Ypos - (JD3-E2) + Vcut_rate * 3
                else:
                    Y8 = Abs_Ypos - JD3 + Vcut_rate * 3
                lineto(doc,  X8, Y8, layer='레이져')
            else:
                X8 = Abs_Xpos + (R - OP1 - OP2 - OP3 - 4)/2  + OP1 + OP2 + OP3 + 4         
                Y8 = Abs_Ypos - JD3 + Vcut_rate * 3
                lineto(doc,  X8, Y8, layer='레이져')  
            
            X9 = X8
            if(E2>0):
                Y9 = Abs_Ypos - JD3 + Vcut_rate*3
            else:
                Y9 = Abs_Ypos - JD3 - K2_up + Vcut_rate*3
            lineto(doc,  X9, Y9, layer='레이져') 
            X10 = X9 
            if(E2>0):                
                Y10 = Abs_Ypos - JD3 - UW + Vcut_rate*5     
            else:
                Y10 = Abs_Ypos - JD3 - K2_up - UW + Vcut_rate*5     
            lineto(doc,  X10, Y10, layer='레이져')   
            X11 = Abs_Xpos + (R - OP1 - OP2 - OP3 - 4)/2  + OP1 + OP2 +  4 
            if(E2>0):                
                Y11 = Abs_Ypos - JD3 - UW + Vcut_rate*5     
            else:                    
                Y11 = Abs_Ypos - JD3 - K2_up - UW + Vcut_rate*5     
            lineto(doc,  X11, Y11, layer='레이져')      

            X12 = X11 
            if(E2>0):                
                Y12 =  Abs_Ypos - JD3 + Vcut_rate*3
            else:                    
                Y12 =  Abs_Ypos - JD3 - K2_up + Vcut_rate*3
            lineto(doc,  X12, Y12, layer='레이져')         
            X13 = X12 - 2 
            Y13 = Y12 
            lineto(doc,  X13, Y13, layer='레이져')                                      

            X14 = X13                
            Y14 = Y13 - ((JD2+K2_up) - ( JD3+K2_up))               
            lineto(doc,  X14, Y14, layer='레이져')         
                
            X15 = X14
            Y15 = Y14 -  UW + Vcut_rate*2
            lineto(doc,  X15, Y15, layer='레이져')                                                      
            X16 = Abs_Xpos + (R - OP1 - OP2 - OP3 - 4)/2  + OP1 + 2
            Y16 =  Y15
            lineto(doc,  X16, Y16, layer='레이져')                                                                   

            X17 = X16
            Y17 = Y14
            lineto(doc,  X17, Y17, layer='레이져')                                      
            X18 = X17                
            Y18 = Abs_Ypos - JD1 - K1_up + Vcut_rate*3
            lineto(doc,  X18, Y18, layer='레이져')                                                      
            X19 = X18 - 2
            Y19 = Y18
            lineto(doc,  X19, Y19, layer='레이져')     
                
            # 4co 추가된 부분   
            X20 = X19
            Y20 = Y19 -  UW + Vcut_rate*2
            lineto(doc,  X20, Y20, layer='레이져')                                                      
            X21 = Abs_Xpos + (R - OP1 - OP2 - OP3 - 4)/2  
            Y21 =  Abs_Ypos - JD1 - K1_up - UW  + Vcut_rate*5
            lineto(doc,  X21, Y21, layer='레이져')                                                                   

            X22 = X21
            Y22 = Abs_Ypos - JD1 - K1_up  + Vcut_rate*3
            lineto(doc,  X22, Y22, layer='레이져')                                      
            X23 = X22                
            Y23 =  Abs_Ypos - JD1  + Vcut_rate*3
            lineto(doc,  X23, Y23, layer='레이져')                                                      
            
            # 기준선 MH1과 JD1 만나는 점
            X24 = Abs_Xpos
            Y24 = Abs_Ypos
            lineto(doc,  X24, Y24, layer='레이져')

            if(U > 0):         
                X25 = Abs_Xpos
                Y25 = Abs_Ypos + MH1 - Vcut_rate * 2
                lineto(doc,  X25, Y25, layer='레이져')                
                X26 = Abs_Xpos - (G-Top_Pop1) * Angle_sin
                Y26 = Abs_Ypos + MH1 + G - Top_Pop1 - Vcut_rate * 3
                lineto(doc,  X26, Y26, layer='레이져')      
                X27 = Abs_Xpos - H1
                Y27 = Y26
                lineto(doc,  X27, Y27, layer='레이져')      
                X28 = X27
                Y28 = Abs_Ypos + MH1 + G  - Vcut_rate * 4
                lineto(doc,  X28, Y28, layer='레이져')      
                lineto(doc,  X1, Y1, layer='레이져')       

            if(U < 1):   
                if( G < 1):
                    X25 = Abs_Xpos
                    Y25 = Abs_Ypos + MH1 - Vcut_rate 
                    lineto(doc,  X1, Y1, layer='레이져')       
                    X28 = X27 = X26 = X25 
                    Y28 = Y23 = Y22 = Y21 
                else:
                    X25 = Abs_Xpos
                    Y25 = Abs_Ypos + MH1 - Vcut_rate * 2
                    lineto(doc,  X25, Y25, layer='레이져')                
                    X26 = Abs_Xpos - (G-Top_Pop1) * Angle_sin
                    Y26 = Abs_Ypos + MH1 + G - Top_Pop1 - Vcut_rate * 3
                    lineto(doc,  X26, Y26, layer='레이져')      
                    X27 = Abs_Xpos - H1
                    Y27 = Y27
                    lineto(doc,  X27, Y27, layer='레이져')      
                    X28 = X27
                    Y28 = Abs_Ypos + MH1 + G  - Vcut_rate * 3
                    lineto(doc,  X28, Y28, layer='레이져')
                    lineto(doc,  X1, Y1, layer='레이져')

        if args.opt11:
            drawcircle(doc, X21 + 12.5, Y21 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이
            drawcircle(doc, X10 - 12.5, Y10 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이

        # 절곡선
        if(U>0):
            line(doc, X28, Y28, X3, Y3,  layer='22')        
        if(G>0):            
            # 실선처리
            line(doc, X25, Y25, X6, Y6,  layer='HIDDEN2')

        line(doc, X24, Y24, X7, Y7,  layer='22')
        # 뒷날개쪽 3개
        line(doc, X22, Y22, X19, Y19,  layer='22')
        line(doc, X17, Y17, X14, Y14,  layer='22')
        line(doc, X12, Y12, X9, Y9,  layer='22')
        # 도어두께 표시하는 치수선
        dim_vertical_right(doc, X18, Y18, X17, Y17, 50, 'dim', text_height=0.22,  text_gap=0.07)  
        dim_vertical_right(doc, X17, Y17, X16, Y16, 150, 'dim', text_height=0.22,  text_gap=0.07)  
        dim_vertical_left(doc, X13, Y13, X14, Y14, 150, 'dim', text_height=0.22,  text_gap=0.07)                                

        # HPI 그리기
        if(HPI_height>0):
            draw_HPI(Abs_Xpos + (X7-X24)/2, Y24, R, HPI_height )     
        else:
            draw_NoHPI(Abs_Xpos + (X7-X24)/2, Y24 )                   

        #  전개도 상부 치수선
        dim_linear(doc,  X1, Y1, X2, Y2, "", 160,  direction="up", layer='dim')

        # H1, OP, H2 치수선
        if(G > 0):   
            dim_linear(doc,  X1, Y1, X25, Y25,  "", 60,  direction="up", layer='dim')
            dim_linear(doc,  X25, Y25, X6, Y6,   "",  extract_abs(Y1,Y5) + 80 + 25,  direction="up", layer='dim')
            dim_linear(doc,  X6, Y6, X2, Y2,  "",  extract_abs(Y1,Y5) + 80 + 25,  direction="up", layer='dim')

        # 우측 절곡치수선
        if(U>0):        
            dim_vertical_right(doc, X2, Y2, X3, Y3, 160 , 'dim', text_height=0.22, text_gap=0.07)  
        if(G>0):                    
            dim_vertical_right(doc, X6,  Y6, X3, Y3,  extract_abs(X2,X6) + 100 , 'dim', text_height=0.22, text_gap=0.07)  

        dim_vertical_right(doc, X6, Y6,  X7, Y7,  extract_abs(X2,X6) +  100, 'dim', text_height=0.22, text_gap=0.07)  
        dim_vertical_right(doc, X9, Y9,  X7, Y7,  extract_abs(X2,X6) + extract_abs(X9,X6) + 100, 'dim', text_height=0.22, text_gap=0.07)   
        dim_vertical_right(doc, X9, Y9, X10, Y10, extract_abs(X2,X6) + extract_abs(X9,X6) + 100, 'dim', text_height=0.22, text_gap=0.07)   # 뒷날개
        dim_vertical_right(doc, X2, Y2,  X15, Y15, extract_abs(X2,X6) + extract_abs(X9,X6) + 170, 'dim', text_height=0.22, text_gap=0.07)             

        # 좌측 절곡치수선        
        if(U>0):                    
            dim_vertical_left(doc, X1, Y1, X28, Y28,   120, 'dim', text_height=0.22, text_gap=0.07)  
        if(G>0):                        
            dim_vertical_left(doc, X25, Y25, X28, Y28, extract_abs(X24,X28) + 70 , 'dim', text_height=0.22, text_gap=0.07)     
        # MH                             
        dim_vertical_left(doc, X24, Y24, X25, Y25,  extract_abs(X24,X28) + 70 , 'dim', text_height=0.22, text_gap=0.07)      
        # JD      
        dim_vertical_left(doc, X24, Y24, X23, Y23,  extract_abs(X24,X28) + 70 , 'dim', text_height=0.22, text_gap=0.07)           
        # 뒷날개
        dim_vertical_left(doc, X22, Y22, X21, Y21,  extract_abs(X24,X28) +  extract_abs(X24,X21) + 70, 'dim', text_height=0.22, text_gap=0.07)                          
        # 좌측 전체
        dim_vertical_left(doc, X16, Y16, X1, Y1,   extract_abs(X1,X16) + 220, 'dim', text_height=0.22, text_gap=0.07)                          

        # 쫄대타입 테둘림유 상부 Vcut 치수 표기
        if(U>0 and G>0):   
            dim_vertical_left(doc, X1+H1+(OP1+OP2+OP3)/2, Y1, X1+H1+(OP1+OP2+OP3)/2, Y6, 100, 'dim', text_height=0.22,  text_gap=0.07)  
        
        # 하부 치수선 4개 노칭 1개 위방향
        dim_linear(doc,  X24, Y24, X21, Y21,  "",  extract_abs(Y21, Y24) + 160 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
        dim_linear(doc,  X16, Y16, X15, Y15,  "",  80 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
        dim_linear(doc,  X21, Y21, X20, Y20,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
        dim_linear(doc,  X11, Y11, X10, Y10,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)              
        dim_linear(doc,  X21, Y21, X10, Y10,  "",  160 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)              
        dim_linear(doc,  X10, Y10, X7, Y7,    "",  160 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)              

        # 노칭 2곳
        dim_linear(doc,  X19, Y19, X18, Y18,  "",  50  ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      
        dim_linear(doc,  X13, Y13, X12, Y12,  "",  50  ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      
                    
        # 상판 돌출부위 H1, H2 치수선
        if(G > 0):  
            dim_vertical_right(doc, X28, Y28, X26, Y26, 100, 'dim', text_height=0.22,  text_gap=0.07)
            dim_vertical_left(doc, X3, Y3, X5, Y5, 100, 'dim', text_height=0.22,  text_gap=0.07)

        # 기둥과의 접선 부위 aligned 표현 각도표현
        # 좌측
        dim_linear(doc, X24, Y24, X22, Y22,   "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) 
        dim_angular(doc, X23, Y23, X24, Y24, X22, Y22, X23, Y23 + 10,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다
        # 우측        
        dim_linear(doc, X9, Y9, X7, Y7,  "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) 
        dim_angular(doc, X9, Y9  , X8, Y8 + 10,  X8, Y8,  X7, Y7,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다

        # 상판 K1 K2 값 부분 치수선 
        if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
            dim_vertical_right(doc, X23, Y23, X22, Y22, 80, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_left(doc, X9, Y9, X8, Y8, 80, 'dim', text_height=0.22,  text_gap=0.07)  

        # JD 간격 치수선 그림
        dim_vertical_right(doc, X24, Y24, X22, Y22, 220, 'dim', text_height=0.22,  text_gap=0.07)  
        dim_vertical_right(doc, X1+H1+(OP1+OP2+OP3)/2 , Y7, X14, Y14, 220, 'dim', text_height=0.22,  text_gap=0.07)                  
        dim_vertical_left(doc, X7, Y7, X9, Y9, 220, 'dim', text_height=0.22,  text_gap=0.07)       

        # 문구 설정
        if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
            Pre_text = ""
        else:
            Pre_text = str(Floor_mc) + "-"
        # Textstr 생성
        Textstr = f"({Pre_text}{str(Floor_des)})"       

        X2 = Abs_Xpos + (H1 + R + H2)/2 - len(Textstr)*50/3
        Y2 = Abs_Ypos - JD1 / 1.8
        draw_Text(doc, X2, Y2, 50, str(Textstr), '레이져')

        # 문구            
        sizedes = Upper_size                  
        
        Textstr = f"Part Name : 상판({Pre_text}{str(Floor_des)})"    
        draw_Text(doc, X10+500, Y10 + 50 - 150, 28, str(Textstr), '0')
        Textstr = f"Mat.Spec : {widejamb_material}"    
        draw_Text(doc, X10+500, Y10 - 150, 28, str(Textstr), '0')
        Textstr = f"Size : {sizedes}"    
        draw_Text(doc, X10+500, Y10 - 50 - 150 , 28, str(Textstr), '0')
        Textstr = f"Quantity : 1 EA"    
        draw_Text(doc, X10+500, Y10 - 100 - 150 , 28, str(Textstr), '0')            

        ##################################################################
        # 4CO와이드 상판 좌측 단면도    (좌우가 다른 모양일때 단면도를 그려준다)
        #################################################################
        if(JD1 != JD2 or K1_up != K2_up or MH1 != MH2 or E1 != E2):
            Section_Xpos = Abs_Xpos - H1 - 400 - G           

            if(G > 0):
                X1 = Section_Xpos + G 
                Y1 = Abs_Ypos + MH1 + U
                X2 = X1
                Y2 = Y1 - U  
                X3 = X2 - G
                Y3 = Y2
                if(U>0):
                    line(doc, X1, Y1, X2, Y2, layer='0')                                    
                    dim_vertical_right(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)                         
                    lineto(doc,  X3, Y3, layer='0')    
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                else:
                    line(doc, X2, Y2, X3, Y3, layer='0')  
                dim_linear(doc, X3, Y3, X1, Y1, "", 100,  direction="up", layer='dim')                                
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                
                X4 = X3 
                Y4 = Y3 - MH1
                lineto(doc,  X4, Y4, layer='0')                                
            if(G < 1):                       
                X1 = Section_Xpos 
                Y1 = Abs_Ypos + MH1 
                X2 = X1
                Y2 = Y1
                X3 = X2
                Y3 = Y2                
                X4 = X3 
                Y4 = Y3 - MH1
                line(doc, X3, Y3, X4, Y4, layer='0')                      

            if(G > 1):
                drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
                if(U>0):                    
                    dim_leader_line(doc, X4, Y4 , X4 - 200 , Y4+100, "V-cut 3개", layer="dim", style='0')      
                else:            
                    dim_leader_line(doc, X4, Y4 , X4 - 200 , Y4+100, "V-cut 2개", layer="dim", style='0')                              
            if(G < 1):
                drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
                dim_leader_line(doc, X4, Y4 , X4 - 100 , Y4+100, "V-cut 1개", layer="dim", style='0')                              

            if(K1_up > 0):
                X5 = X4 - JD1
                Y5 = Y4 
                lineto(doc,  X5, Y5, layer='0')               
                X6 = X5 - K1_up
                Y6 = Y5 
                lineto(doc,  X6, Y6, layer='0')   
                X7 = X6 - (JD2-JD1)
                Y7 = Y6 
                lineto(doc,  X6, Y6, layer='0')   
                dim_linear(doc,  X4, Y4,  X5, Y5, "", 50, direction="down", layer='dim')     
                dim_linear(doc,  X6, Y6,  X5, Y5, "", 50, direction="down", layer='dim')     
                dim_linear(doc,  X7, Y7,  X6, Y6, "", 50, direction="down", layer='dim')     
                dim_linear(doc,  X4, Y4,  X7, Y7, "", 150, direction="down", layer='dim')    

            else:
                X5 = X4 - JD1
                Y5 = Y4 
                lineto(doc,  X5, Y5, layer='0')               
                X6 = X5 
                Y6 = Y5                              
                X7 = X6 - (JD2-JD1)
                Y7 = Y6              
                lineto(doc,  X7, Y7, layer='0')                                         
                dim_linear(doc, X7, Y7, X6, Y6,   "", 50, direction="down", layer='dim')     
                dim_linear(doc, X6, Y6, X4, Y4,   "", 50, direction="down", layer='dim')     
                dim_linear(doc, X4, Y4, X7, Y7,   "", 150, direction="down", layer='dim')    

            # drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
            X8 = X7 
            Y8 = Y7 + UW
            lineto(doc,  X8, Y8, layer='0')     
            dim_vertical_left(doc,  X7, Y7, X8, Y8, 50,'dim' ,text_height=0.22,  text_gap=0.07)            

            dim_vertical_left(doc,  X7, Y7, X3, Y3,  110,'dim' ,text_height=0.22,  text_gap=0.07)    
            
            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JD1 != JD2 or K1_up != K2_up ):
                rectangle(doc, X6+(JD1+K1_up)/2 -55, Y6-90, X6+(JD1+K1_up)/2 + 55, Y6-190, layer='0')               
                rectangle(doc, X4+250+50, Y6-40, X4+350+50, Y6 - JD1 + 40, layer='0')               
            if( MH1 != MH2 ):
                rectangle(doc, X6-70, Y6 + MH1/2 + 90 , X6 - 70 - 100 , Y6 + MH1/2 - 90, layer='0')               
                rectangle(doc, X4+230, Y6 + MH1/2 + 90  , X4+330,  Y6 + MH1/2 - 90 , layer='0')               
        
        ###########################################################################
        # 4CO와이드 상판 우측 단면도 기본적으로 다 그림
        ###########################################################################
        Section_Xpos = Abs_Xpos + R + H2 + 450 + G           

        if(G > 0):
            X1 = Section_Xpos - G 
            Y1 = Abs_Ypos + MH2 + U
            X2 = X1
            Y2 = Y1 - U                          
            X3 = X2 + G
            Y3 = Y2
            if(U>0):
                line(doc, X1, Y1, X2, Y2, layer='0')                                    
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                dim_vertical_left(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)                         
                lineto(doc,  X3, Y3, layer='0')    
            else:
                line(doc, X2, Y2, X3, Y3, layer='0')  

            lineto(doc,  X3, Y3, layer='0')    
            dim_linear(doc, X1, Y1, X3, Y3,  "", 100,  direction="up", layer='dim')                            
            drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                
            X4 = X3 
            Y4 = Y3 - MH2
            lineto(doc,  X4, Y4, layer='0')              

        if(G < 1):                       
            X1 = Section_Xpos 
            Y1 = Abs_Ypos + MH2
            X2 = X1
            Y2 = Y1
            X3 = X2
            Y3 = Y2                
            X4 = X3 
            Y4 = Y3 - MH2
            line(doc, X3, Y3, X4, Y4, layer='0')                      

        if(G > 1):
            drawcircle(doc, X4, Y4 , 5 , layer='0', color='6')             
            if(U>0):                
                dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 3개", layer="dim", style='0')      
            else:
                dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 2개", layer="dim", style='0')      
        if(G < 1):
            drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
            dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 1개", layer="dim", style='0')   

        if(K1_up > 0):
            X5 = X4 + JD3
            Y5 = Y4 
            lineto(doc,  X5, Y5, layer='0')               
            X6 = X5 + K2_up
            Y6 = Y5 
            lineto(doc,  X6, Y6, layer='0')   
            X7 = X6 + (JD2-JD3)
            Y7 = Y6 
            lineto(doc,  X6, Y6, layer='0')   
            dim_linear(doc,  X5, Y5, X4, Y4,  "", 50, direction="down", layer='dim')     
            dim_linear(doc,  X5, Y5, X6, Y6,  "", 50, direction="down", layer='dim')     
            dim_linear(doc,  X6, Y6, X7, Y7,  "", 50, direction="down", layer='dim')     
            dim_linear(doc,  X4, Y4, X7, Y7,   "", 150, direction="down", layer='dim')                               
        else:
            X5 = X4 + JD3
            Y5 = Y4 
            lineto(doc,  X5, Y5, layer='0')               
            X6 = X5 
            Y6 = Y5                              
            X7 = X6 + (JD2-JD3)
            Y7 = Y6              
            lineto(doc,  X7, Y7, layer='0')                                         
            dim_linear(doc,  X6, Y6, X7, Y7,  "", 50, direction="down", layer='dim')     
            dim_linear(doc,  X4, Y4, X6, Y6,  "", 50, direction="down", layer='dim')     
            dim_linear(doc,  X4, Y4, X7, Y7,  "", 150, direction="down", layer='dim')    

        # drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
        X8 = X7 
        Y8 = Y7 + UW
        lineto(doc,  X8, Y8, layer='0')     
        dim_vertical_right(doc,  X7, Y7, X8, Y8, 30,'dim' ,text_height=0.22,  text_gap=0.07)            
        dim_vertical_right(doc,   X6, Y6, X3, Y3,   130,'dim' ,text_height=0.22,  text_gap=0.07)    

        if(K1_up > 0):
            line(doc, X5, Y5, X5 , Y5 + UW , layer='22')               
            line(doc, X6, Y6, X6 , Y6 + UW , layer='22')                               
        else:
            line(doc, X6, Y6, X6 , Y6 + UW , layer='22')                   


        # 서로 치수 차이나는 곳 사각형으로 그려주기
        if(JD1 != JD2 or K1_up != K2_up ):
            rectangle(doc, X6-(JD1+K1_up)/2 + 55, Y6-90, X6-(JD1+K1_up)/2 - 55, Y6-190, layer='0')                               
            rectangle(doc, X4 - 320 - 50 , Y6 - 40, X4 - 420 - 50, Y6 - JD1 + 40, layer='0')        
        if( MH1 != MH2 ):
            rectangle(doc, X6 + 60, Y6 + MH2/2 + 90 , X6 + 60 + 100 , Y6 + MH2/2 - 90, layer='0')               
            rectangle(doc, X4 - 330, Y6 + MH2/2 + 90  , X4 - 430,  Y6 + MH2/2 - 90 , layer='0')                           

        # 상판 도면틀 넣기
        BasicXscale = 2560
        BasicYscale = 1550
        TargetXscale = 1000 + R + (JD1 + K1_up)*2 + 200
        # 상판 크게 만든것
        TargetYscale = 500 + JD2 + MH1 + G + U   
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale
        # print (f'스케일 : {frame_scale}')
        insert_frame(X1 - ( R + (JD2 + K1_up)*2 + 900 ),Y1-(850 + JD2 + MH1 + G + U) , frame_scale, "top", workplace)

        # 다음 도면 간격 계산
        Abs_Ypos -= 4000

    ###################################################################################################################################################################
    # 4CO와이드 좌기둥 전개도
    ###################################################################################################################################################################
    Abs_Xpos = 52000 + 5000
    Abs_Ypos = 1000    
    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        MH1 = row[2]
        MH2 = row[3]
        OP1 = row[4]
        OP2 = row[5]
        OP3 = row[6]
        R = row[7]
        JD1 = row[8]
        JD2 = row[9]
        JD3 = row[10]
        JB1_up = row[11]
        JB1_down = row[12]
        JB2_up = row[13]
        JB2_down = row[14]
        Top_gap = row[15]
        Side_gap = row[16]        
        H1 = row[17]
        H2 = row[18]
        C1 = row[19]        
        C2 = row[20]
        B1 = row[21]
        B2 = row[22]
        A1 = row[23]
        A2 = row[24]
        LH1 = row[25]
        LH2 = row[26]
        RH1 = row[27]
        RH2 = row[28]        
        U = row[29]        
        G = row[30]        
        UW = row[31]        
        SW = row[32]         
        K1_up = row[33]        
        K1_down = row[34]        
        K2_up = row[35]        
        K2_down = row[36]                
        SMH1 = row[37]                
        SMH2 = row[38]                
        G_check = row[39]
        Upper_size = row[40]
        Left_side_size = row[41]
        Right_side_size = row[42]
        E1 = row[43]
        E2 = row[44]
        HPI_height = row[45]
        SW1 = row[46]
        SW2 = row[47]
        SWAngle = row[48]
        A1_var = row[49]
        A2_var = row[50]
        Angle = row[51]
        Kadapt = row[65]  # K값 적용여부 1 이면 K값 적용하지 말것

        # 데이터 검증 및 기본값 설정
        SWAngle = validate_or_default(SWAngle)
        K1_up = validate_or_default(K1_up)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_down = validate_or_default(K2_down)
        OP1 = validate_or_default(OP1)
        Kadapt = validate_or_default(Kadapt)

        if(Kadapt==1):
            K1_up = 0            
            K1_down = 0            
            K2_up = 0            
            K2_down = 0            

                   
        if(K1_up>=10 and K2_up>=10 ):
            # 추영덕소장 E1, E2값 무시함 위의 조건이라면
            E1 = 0
            E2 = 0             

        if(OP1>0):            
            SWAngle_radians = math.radians(SWAngle)
            if(SWAngle>0):
                SWAngle_var =  SW2 * math.cos(SWAngle_radians)        
            Angle_radians = math.radians(Angle)
            if(Angle>0):
                Angle_cos = math.cos(Angle_radians)        
                Angle_sin = math.sin(Angle_radians)    

            if (SWAngle>1):  #SWAngle 0 이상인 경우            
                if(K1_up>0):      
                    if (SWAngle == 90):      
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 3 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3                
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4           
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')
                    else:
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 2 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3                
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4           
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                else:
                    if (SWAngle == 90):                            
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 3 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3          
                        Y4 =  Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate   
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4              
                    else:
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 2 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3          
                        Y4 =  Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate   
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4              

            if (SWAngle<1):  #SWAngle 0 이상인 경우
                if(K1_up>0):            
                    X1 = Abs_Xpos + SMH1
                    Y1 = Abs_Ypos + JB1_up + K1_up + SW - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + SMH1 + LH2 
                    Y2 = Abs_Ypos + JB1_down + K1_down + SW - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                    X3 = X2            
                    Y3 = Y2
                    X4 = X3
                    Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = X4               
                    Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                    lineto(doc,  X5, Y5, layer='레이져')                    
                else:
                    X1 = Abs_Xpos + SMH1
                    Y1 = Abs_Ypos + JB1_up + SW - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + SMH1 + LH2 
                    Y2 = Abs_Ypos + JB1_down + SW - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                    X3 = X2            
                    Y3 = Y2
                    X4 = X3             
                    Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = X4             
                    Y5 = Y4                

            X6 = Abs_Xpos +  SMH1 + LH1             
            Y6 = Abs_Ypos
            lineto(doc,  X6, Y6, layer='레이져')
            if(A1>0):     
                X7 = X6 
                Y7 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X7, Y7, layer='레이져')                  
                X8 = X7
                Y8 = Abs_Ypos - C1 - A1 + Vcut_rate * 3
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = Abs_Xpos             
                Y10 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X10, Y10, layer='레이져')

            if(A1<1):       
                X7 = X6 
                Y7 = Abs_Ypos - C1 + Vcut_rate 
                lineto(doc,  X7, Y7, layer='레이져')                
                X8 = Abs_Xpos + SMH1 + LH1  
                Y8 = Y7
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = X9
                Y10 = Y9            

            X11 = Abs_Xpos 
            Y11 = Abs_Ypos
            lineto(doc,  X11, Y11, layer='레이져')        

            if (SWAngle>0): 
                if(K1_up>0):  
                    if (SWAngle == 90):
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1  
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Abs_Ypos + JB1_up + K1_up - Bending_rate   - Vcut_rate   
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 3 - Vcut_rate
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1  
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                else:  
                    if (SWAngle == 90):                      
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1     
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Y14
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + SW1  - Bending_rate * 3 - Vcut_rate  
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                      
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1     
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Y14
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                      
                    
            if (SWAngle<1):  # 헤밍구조 아닌경우
                if(K1_up>0):  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + B1
                    lineto(doc,  X12, Y12, layer='레이져')                            
                    X13 = Abs_Xpos + SMH1
                    Y13 = Y12
                    lineto(doc,  X13, Y13, layer='레이져')                
                    X14 = X13
                    Y14 = Abs_Ypos +  JB1_up  - Vcut_rate
                    lineto(doc,  X14, Y14, layer='레이져')
                    X15 = X14
                    Y15 = Abs_Ypos + JB1_up + K1_up  - Bending_rate  - Vcut_rate  
                    lineto(doc,  X15, Y15, layer='레이져')
                    X16 = X15 
                    Y16 = Y15
                    lineto(doc,  X1, Y1, layer='레이져')
                else:  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + B1
                    lineto(doc,  X12, Y12, layer='레이져')                            
                    X13 = Abs_Xpos + SMH1
                    Y13 = Y12
                    lineto(doc,  X13, Y13, layer='레이져')                
                    X14 = X13
                    Y14 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate  
                    lineto(doc,  X14, Y14, layer='레이져')
                    X15 = X14
                    Y15 = Y14
                    lineto(doc,  X15, Y15, layer='레이져')
                    X16 = X15 
                    Y16 = Y15
                    lineto(doc,  X1, Y1, layer='레이져')      

            # 좌기둥 절곡선
            if(SW1>0):
                line(doc, X16, Y16, X3, Y3,  layer='22')
            line(doc, X15, Y15, X4, Y4,  layer='22')
            if(K1_up>0):
                line(doc, X14, Y14, X5, Y5,  layer='22')
            # JB    
            line(doc, X11, Y11, X6, Y6,  layer='22')
            if(A1>0):
                line(doc, X10, Y10, X7, Y7,  layer='22')                

            # 좌기둥 상부 전개도 치수선        
            dim_linear(doc,  X12, Y12, X1, Y1, "", extract_abs(Y12,Y1) + 100,  direction="up", layer='dim')                
            dim_linear(doc,  X1, Y1, X2, Y2, "", extract_abs(Y1,Y2) + 100,  direction="up", layer='dim')                
            dim_linear(doc,  X12, Y12, X2, Y2, "", extract_abs(Y12,Y2) + 150,  direction="up", layer='dim')                        

            # 우측 절곡치수선
            if(SW1>0):
                dim_vertical_right(doc, X2,  Y2, X3, Y3,  extract_abs(X2,X7) + 180, 'dim', text_height=0.22, text_gap=0.01)      
            dim_vertical_right(doc, X3, Y3, X4,  Y4,   extract_abs(X4,X7) +  80, 'dim', text_height=0.22, text_gap=0.01)      
            if(K1_up>0):
                dim_vertical_right(doc, X4,  Y4, X5, Y5,  extract_abs(X4,X7) + 130, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_right(doc, X6, Y6, X5,  Y5, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_right(doc,  X7, Y7, X6,  Y6, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_right(doc, X7,  Y7, X8, Y8, 120, 'dim', text_height=0.22, text_gap=0.01)  
                #Vcut A값 C값 전개치수선     
                dim_vertical_left(doc, X6, Y6, X8,  Y8,   extract_abs(X4,X7) + 80 , 'dim', text_height=0.22, text_gap=0.01)
            dim_vertical_right(doc, X2,  Y2, X8, Y8,   extract_abs(X2,X8) + 270 , 'dim', text_height=0.22, text_gap=0.01)      

            # 좌측 절곡치수선  
            if(SW1>0):                       
                dim_vertical_left(doc, X1, Y1, X16, Y16, SMH1 + 180, 'dim', text_height=0.22, text_gap=0.01)  
                dim_vertical_left(doc, X15, Y15, X16, Y16,  SMH1 + 80, 'dim', text_height=0.22, text_gap=0.01)  
            else:
                dim_vertical_left(doc, X1, Y1, X15, Y15, SMH1 + 80, 'dim', text_height=0.22, text_gap=0.01)                  
            if(K1_up>0):
                dim_vertical_left(doc, X15, Y15, X14, Y14, SMH1 + 130, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X11, Y11, X14, Y14, 80, 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_left(doc, X11, Y11, X10, Y10,  80, 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_left(doc, X9, Y9, X10, Y10, 160, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X1, Y1, X9,  Y9,  SMH1 + 270 , 'dim', text_height=0.22, text_gap=0.01) 

            if(B1>0):
                #B1 치수선     
                dim_vertical_left(doc, X13, Y13, X13,  Y11,   70 , 'dim', text_height=0.22, text_gap=0.01)      
                # JD 접선치수선
                dim_vertical_right(doc, X13, Y13, X15,  Y15,   50 , 'dim', text_height=0.22, text_gap=0.01)      

            # 기둥 하부 치수선 2    
            dim_linear(doc,   X13, Y13, X8, Y8, "", B1+ C1 + A1 + 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   
            dim_linear(doc,   X9, Y9, X8, Y8, "", 150,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   

            if LH1>LH2:
                dim_linear(doc,  X2, Y2, X6, Y6, "",  150 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)                   
            if LH1<LH2:    
                dim_linear(doc,  X4, Y4, X8, Y8, "", extract_abs(Y4, Y8) + 150 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            if args.opt7:
                drawcircle(doc, X1 + 12.5, Y1 - 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X2 - 12.5, Y2 - 5, 2.5 , layer='레이져') # 도장홀 5파이        

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"

            Xpos = Abs_Xpos + SMH1 + 400
            Ypos = Abs_Ypos + JB1_up / 2.3
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-좌"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')           

            Xpos = Abs_Xpos + SMH1 + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-좌"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + SMH1 + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Mat.Spec : {widejamb_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + SMH1 + LH1/1.2        
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Size : {Left_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + SMH1 + LH1/1.2        
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')            

            ########################################################################################################################################################################################
            # 와이드 좌기둥 좌측 단면도    
            ########################################################################################################################################################################################
            Section_Xpos = Abs_Xpos - 530
            Section_Ypos = Abs_Ypos + JB1_up + K1_up

            # 가정: SWAngle은 도 단위로 주어진 각도
            SWAngle_radians = math.radians(SWAngle)

            if(SWAngle>0):
                X1 = Section_Xpos - SW1 - SW2 * math.cos(SWAngle_radians)
                Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                X2 = Section_Xpos - SW1
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos
                
                line(doc, X2, Y2, X1, Y1, layer='0')               
                line(doc, X2, Y2, X3, Y3, layer='0')        
                dim_linear(doc, X1, Y1, X2, Y2,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # 각도 90도가 아니면 각도 표기
                if (SWAngle != 90):               
                    dim_angular(doc, X2 , Y2, X1, Y1, X2-30, Y2, X3-30, Y3,  200, direction="left" )
                    # help(ezdxf.layouts.Modelspace.add_angular_dim_2l)
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)   

            if(SWAngle<1):
                X1 = Section_Xpos - SW
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos            
                line(doc,  X1, Y1, X2, Y2, layer='0')                                                   
                dim_linear(doc, X1, Y1, X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

            if(K1_up>5):
                X4 = X3
                Y4 = Y3 - K1_up
                lineto(doc,  X4, Y4, layer='0')       
            else:
                X4 = X3
                Y4 = Y3

            # Angle은 도 단위로 주어진 각도
            Angle_radians = math.radians(Angle)

            X5 = X4 - JB1_up * math.sin(Angle_radians)                
            Y5 = Y4 - JB1_up * math.cos(Angle_radians)  
            lineto(doc,  X5, Y5, layer='0')    
            X6 = X5 - C1
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='0')   

            if(A1>0):
                X7 = X6
                Y7 = Y6 + A1
                lineto(doc,  X7, Y7, layer='0') 
                drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut 2개소", layer="dim", style='0')
                # A 치수
                dim_vertical_left(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
            if(A1<1):
                X7 = X6
                Y7 = Y6            
                drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut", layer="dim", style='0')

            # C값 치수선
            dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            # if(K1_up>5):
            dim_vertical_right(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                # if (Angle > 1):   # 직각이 아닌경우                        
            dim_angular(doc, X4, Y4, X5, Y5, X4, Y4, X3, Y3,  30, direction="left" )

            # JB사선 치수선
            dim_linear(doc, X4, Y4, X5, Y5,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
            # if (Angle > 1):   # 직각이 아닌경우            
            dim_angular(doc, X6, Y6, X5, Y5, X5, Y5, X4, Y4,  20, direction="left" )

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB1_up != JB1_down ):
                rectangle(doc, X5+60, Y5 + K1_up + (JB1_up)/2 - 75 , X5 + 60 + 100, Y5 + K1_up + (JB1_up)/2 + 75 , layer='0')
                rectangle(doc, X5+60+350, Y5 + K1_up + (JB1_up)/2 - 75, X5 + 60 + 100 + 350, Y5 + K1_up + (JB1_up)/2 + 75 , layer='0')
            if(K1_up != K1_down ):
                rectangle(doc, X4 + 50, Y4 - 10 , X4 + 50 + 100, Y4 + K1_up + 10 , layer='0')                         
                rectangle(doc, X4 + 50 + 300, Y4 - 10 , X4 + 50 + 100 + 300, Y4 + K1_up + 10 , layer='0')                         
                            
            ####################################################################################################################################################################################
            # 와이드 좌기둥 우측 단면도    
            ####################################################################################################################################################################################
            if( JB1_up != JB1_down or K1_up != K1_down ):      
                # print ('기둥좌측 하단 단면도 작도')  
                Section_Xpos = Abs_Xpos + SMH1 + LH1 + 430
                Section_Ypos = Abs_Ypos + JB1_up + K1_up

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(SWAngle>0):  
                    X1 = Section_Xpos + SW1 + SW2 * math.cos(SWAngle_radians)
                    Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                    X2 = Section_Xpos + SW1
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc, X2, Y2, X1, Y1, layer='0')                   
                    line(doc, X2, Y2, X3, Y3, layer='0')        
                    dim_linear(doc, X2, Y2, X1, Y1,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # 각도 90도가 아니면 각도 표기
                    if (SWAngle != 90):               
                        dim_angular(doc,  X3+5, Y3, X2+5, Y2, X1, Y1, X2 , Y2,  200, direction="right" )                    
                    dim_linear(doc, X3, Y3,  X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

                if(SWAngle<1):  
                    X1 = Section_Xpos + SW
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos 
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc,   X1, Y1, X2, Y2,  layer='0')                                   
                    dim_linear(doc,  X2, Y2, X1, Y1,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

                if(K1_down>5):
                    X4 = X3
                    Y4 = Y3 - K1_down
                    lineto(doc,  X4, Y4, layer='0')       
                else:
                    X4 = X3
                    Y4 = Y3

                # Angle은 도 단위로 주어진 각도
                Angle_radians = math.radians(Angle)

                X5 = X4 + JB1_down * math.sin(Angle_radians)                
                Y5 = Y4 - JB1_down * math.cos(Angle_radians)  
                lineto(doc,  X5, Y5, layer='0')    
                X6 = X5 + C1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')   

                if(A1>0):
                    X7 = X6
                    Y7 = Y6 + A1
                    lineto(doc,  X7, Y7, layer='0') 
                    drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                    drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                    dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut 2개소", layer="dim", style='0')
                    # A 치수
                    dim_vertical_right(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
                if(A1<1):
                    X7 = X6
                    Y7 = Y6            
                    drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                    dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut", layer="dim", style='0')

                # C값 치수선
                dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                if(K1_up>5):
                    dim_vertical_left(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                    # if (Angle > 1):   # 직각이 아닌경우                        
                    dim_angular(doc,  X3, Y3, X4 , Y4,  X5, Y5, X4, Y4,   15, direction="right" )

                # JB사선 치수선
                dim_linear(doc,  X5, Y5, X4, Y4,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
                # if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc,  X5, Y5, X4, Y4, X5, Y5, X6 , Y6,  15, direction="right" )  

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB1_up != JB1_down ):
                    rectangle(doc, X5-60,Y5 + (JB1_up+K1_up)/2 - 75 , X5 - 60 - 100,Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')               
                    rectangle(doc, X5 - 60 - 300, Y5 + (JB1_up+K1_up)/2 - 75 , X5 - 60 - 100 - 300, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')   

                if(K1_up != K1_down ):
                    rectangle(doc, X4 - 50, Y4 - 10 , X4 - 50 - 100, Y4 + K1_down + 10 , layer='0')                             
                    rectangle(doc, X4 - 50 - 220, Y4 - 10 , X4 - 50 - 100 - 220, Y4 + K1_down + 10 , layer='0')                             

            #########################################################################################################################################################################
            # 와이드 우기둥 전개도 및 단면도
            # 주의 사항 기둥의 우측은 X1의 선이 끊어지는 현상을 방지하기 위해서 좌기둥과 다르게 설계해야 함  (와이드형태만 해당됨)
            #########################################################################################################################################################################
                    
            Abs_Ypos -= 2000 
            R_Xpos = Abs_Xpos + SMH2
            R_Ypos = Abs_Ypos + 1100 - (C1 + C2 + A1 + A2)

            if(A2>0):            
                #  X1은 좌측 끝단으로 이동해서 점을 하나 다르게 해야 함
                X1 = R_Xpos  - SMH2
                Y1 = R_Ypos + C2 + A2 - Vcut_rate * 3
                X2 = R_Xpos + RH1 
                Y2 = Y1
                line(doc, X1, Y1, X2, Y2, layer='레이져')                    
                X3 = X2
                Y3 = R_Ypos + C2 - Vcut_rate*2
                lineto(doc,  X3, Y3, layer='레이져')            

            if(A2<1):
                X1 = R_Xpos  - SMH2
                Y1 = R_Ypos + C2  - Vcut_rate 
                X2 = R_Xpos + RH1 
                Y2 = Y1
                line(doc, X1, Y1, X2, Y2, layer='레이져')                    
                X3 = X2
                Y3 = Y2

            X4 = X3
            Y4 = R_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')

            X5 = R_Xpos + RH2 

            if (SWAngle>1):  #SWAngle 0 이상인 경우    헤밍구조인 경우        
                if(K2_up>0):   
                    if (SWAngle == 90):    
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = X6
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = R_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')
                    else:   
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = X6
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = R_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')

                if(K2_up<1):   #K값이 없는 경우
                    if (SWAngle == 90):   
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = X6
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11      
                    else:
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = X6
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11

            if (SWAngle<1):  #SWAngle 0     일반구조
                if(K2_up>0):   
                    Y5 = R_Ypos - JB2_down +  Vcut_rate
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = X5
                    Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                    lineto(doc,  X6, Y6, layer='레이져')
                    X7 = X6
                    Y7 = R_Ypos - JB2_down - K2_down - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = R_Xpos 
                    Y9 = R_Ypos - JB2_up - K2_up - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = R_Xpos 
                    Y10 = Y9                
                    X11 = R_Xpos 
                    Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = R_Xpos 
                    Y12 = R_Ypos - JB2_up +  Vcut_rate
                    lineto(doc,  X12, Y12, layer='레이져')
                if(K2_up<1):   #K값이 없는 경우
                    Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = X5
                    Y6 = Y5                
                    X7 = X6
                    Y7 = R_Ypos - JB2_down  - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = R_Xpos 
                    Y9 = R_Ypos - JB2_up  - SW  + Bending_rate * 2 + Vcut_rate
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = X9
                    Y10 = Y9                
                    X11 = R_Xpos 
                    Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = X11
                    Y12 = Y11                
            X13 = R_Xpos 
            Y13 = R_Ypos - B2
            lineto(doc,  X13, Y13, layer='레이져')
            X14 = R_Xpos - SMH2
            Y14 = Y13
            lineto(doc,  X14, Y14, layer='레이져')
            X15 = X14
            Y15 = R_Ypos
            lineto(doc,  X15, Y15, layer='레이져')
            if(A2>0):
                X16 = X15
                Y16 = R_Ypos + C2 - Vcut_rate*2
                lineto(doc,  X16, Y16, layer='레이져')
                X17 = X16
                Y17 = R_Ypos + C2 + A2 - Vcut_rate*3
                lineto(doc,  X17, Y17, layer='레이져')
            if(A2<1):
                X16 = X15 
                Y16 = R_Ypos + C2 - Vcut_rate
                X17 = X16
                Y17 = Y16
                lineto(doc,  X16, Y16, layer='레이져')
 
            # 절곡선1
            if(A2>1):        
                line(doc, X16, Y16, X3, Y3,  layer='22')
            # C2                
            line(doc, X15, Y15, X4, Y4,  layer='22')
            # K2
            if(K2_up>0):
                line(doc, X12, Y12, X5, Y5,  layer='22')
            line(doc, X11, Y11, X6, Y6,  layer='22')
            # 해밍구조
            if (SWAngle>0):   
                line(doc, X10, Y10, X7, Y7,  layer='22')

            # 하부 전개도 치수선
            dim_linear(doc,  X14, Y14, X9, Y9, "",  extract_abs(Y14,Y9) + 100,  direction="down", layer='dim')                
            dim_linear(doc,  X9, Y9, X8, Y8, "", 100,  direction="down", layer='dim')                
            dim_linear(doc,  X14, Y14, X8, Y8, "", extract_abs(Y14,Y9) + 150,  direction="down", layer='dim')                

            # 우측 절곡치수선
            if(A2 > 1):  
                dim_vertical_right(doc, X3, Y3, X2,  Y2, 130, 'dim', text_height=0.22,  text_gap=0.07)      
                #Vcut A값 C값 전개치수선     
                dim_vertical_left(doc, X2, Y2, X4,  Y4,   80 , 'dim', text_height=0.22, text_gap=0.01)                     
            if(C2 > 1):      
                dim_vertical_right(doc, X3, Y3, X4,  Y4, 80, 'dim', text_height=0.22,  text_gap=0.07)                      
            # JB    
            dim_vertical_right(doc, X5, Y5, X4,  Y4, extract_abs(X5,X4) + 80, 'dim', text_height=0.22,  text_gap=0.07) 
            if(K2_down > 0): 
                dim_vertical_right(doc, X6, Y6, X5,  Y5, extract_abs(X8,X5) + 130, 'dim', text_height=0.22,  text_gap=0.07)  
            # 뒷날개    
            dim_vertical_right(doc, X6, Y6, X7,  Y7, extract_abs(X5,X8) + 80, 'dim', text_height=0.22,  text_gap=0.07)  
            if (SWAngle > 0):   
                dim_vertical_right(doc, X7, Y7, X8,  Y8, extract_abs(X7,X8) + 180, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X2, Y2, X8,  Y8, extract_abs(X2,X8) + 230, 'dim', text_height=0.22,  text_gap=0.07)  

            # 좌측 절곡치수선
            if(A2 > 1):          
                dim_vertical_left(doc, X17, Y17, X16, Y16,  130, 'dim', text_height=0.22, text_gap=0.01)  
            if(C2 > 1):              
                dim_vertical_left(doc, X15, Y15, X16, Y16,  60, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X15, Y15, X12, Y12,  60, 'dim', text_height=0.22, text_gap=0.01)  
            if(K2_up > 0): 
                dim_vertical_left(doc, X12, Y12, X11, Y11,  SMH2 + 120, 'dim', text_height=0.22, text_gap=0.01)              
            # 뒷날개    
            dim_vertical_left(doc, X11, Y11, X10, Y10,  SMH2 + 60, 'dim', text_height=0.22, text_gap=0.01)  
            if (SWAngle > 0):               
                dim_vertical_left(doc, X10, Y10, X9, Y9,  SMH2 + 180, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X17, Y17, X9,  Y9,  240 , 'dim', text_height=0.22, text_gap=0.01)      

            # 전개도 상부 치수선    
            dim_linear(doc,  X17, Y17, X2, Y2,  "",   150,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)                      
            dim_linear(doc,  X13, Y13, X2, Y2,  "",   extract_abs(Y1,Y13) + 100,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)                      
            if RH1<RH2:
                dim_linear(doc,  X2, Y2, X5, Y5,  "",  150 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)       
            if RH1>RH2:    
                dim_linear(doc,  X8, Y8, X4, Y4, "",   150 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.01)          
            if(B2>0):
                dim_vertical_left(doc, X13, Y15, X13, Y13,  80, 'dim', text_height=0.22, text_gap=0.01)              
                dim_vertical_right(doc, X11, Y11, X13, Y13,  50, 'dim', text_height=0.22, text_gap=0.01)              
            if args.opt7:
                drawcircle(doc, X9 + 12.5, Y9 + 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X8 - 12.5, Y8 + 5, 2.5 , layer='레이져') # 도장홀 5파이            

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"       

            Xpos = R_Xpos + 400 
            Ypos = R_Ypos - JB2_up / 2
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-우"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')                       

            Xpos = R_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-우"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = R_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Mat.Spec : {widejamb_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = R_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Size : {Right_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = R_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')      

            # 좌우 도면틀 넣기
            BasicXscale = 4487
            BasicYscale = 2770
            TargetXscale = 800 + max(SMH1, SMH2) + max(LH1, LH2) + 900
            TargetYscale = 400*4 + (JB1_up + C1 + A1 + K1_up + SW) * 2
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale

            # print (f'스케일 : {frame_scale}')
            insert_frame(X1 -  950 ,Y1 - (JB2_up+C2+A2+K2_up+SW + 500 + C1 + A1 + 300) , frame_scale, "side", workplace)              
            
            ###################################
            # 와이드 우기둥 좌측 단면도    
            ###################################            
            Section_Xpos = R_Xpos - SMH2 - 500
            Section_Ypos = R_Ypos

            SWAngle_radians = math.radians(SWAngle)

            if(A2>0):
                X1 = Section_Xpos
                Y1 = Section_Ypos - A2
                X2 = Section_Xpos 
                Y2 = Section_Ypos 

                line(doc, X1, Y1, X2, Y2, layer='0')                               
                dim_vertical_left(doc, X1, Y1, X2, Y2, 50, 'dim', text_height=0.22, text_gap=0.07)  
                X3 = X2 + C2
                Y3 = Y2
                lineto(doc,  X3, Y3, layer='0')   
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # vcut 표기                
                # drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                # dim_leader_line(doc, X3, Y3 , X3+50, Y3+100, "V-cut 2개소", layer="dim", style='0')         

            if(A2<1):
                X1 = Section_Xpos
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos             
                X3 = X2 + C2
                Y3 = Y2
                line(doc, X1, Y1, X3, Y3, layer='0')               
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # # vcut 표기                            
                # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                # dim_leader_line(doc, X3, Y3 , X3+50, Y3+100, "V-cut", layer="dim", style='0')                                        

            X4 = X3 + JB2_up * math.sin(Angle_radians)                
            Y4 = Y3 - JB2_up * math.cos(Angle_radians)        
            lineto(doc,  X4, Y4, layer='0')	
            dim_linear(doc, X3, Y3, X4, Y4,  "", 50,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)     

            if(K2_up > 5):
                X5 = X4
                Y5 = Y4 - K2_up
                lineto(doc,  X5, Y5, layer='0')	
                dim_vertical_right(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
            else:
                X5 = X4
                Y5 = Y4
                
            if(SWAngle>0):
                X6 = X5 - SW1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                X7 = X6 - SW2 * math.cos(SWAngle_radians)
                Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                lineto(doc,  X7, Y7, layer='0')
                dim_linear(doc, X6, Y6, X7, Y7,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                dim_angular(doc, X5-5, Y5, X6-5, Y6,  X7, Y7, X6 , Y6,  100, direction="left" )

            if(SWAngle<1): # 헤밍구조가 아닌경우 일반형
                X6 = X5 - SW
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                X7 = X6
                Y7 = Y6                        

            # if (Angle > 1):   # 직각이 아닌경우            
            dim_angular( doc,   X3, Y3, X4, Y4, X3, Y3, X2 , Y2,    15, direction="left" )		                            
            if(K2_up > 5): # K2값
                dim_angular(doc,  X4, Y4, X5 , Y5, X4, Y4, X3, Y3,  5, direction="left" )    

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB2_up != JB2_down ):
                rectangle(doc, X3+60,  Y5 + (JB1_up+K1_up)/2 - 75 , X3 + 60 + 100, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')               
                rectangle(doc, X3+60+300, Y5 + (JB1_up+K1_up)/2 - 75 , X3 + 60 + 400,  Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')     

            if(K2_up != K2_down ):
                rectangle(doc, X5 + 50, Y5 - 10 , X5 + 50 + 100, Y5 + K2_up + 10 , layer='0')                         
                rectangle(doc, X5 + 60 + 250, Y5 - 10 , X5 + 60 + 350, Y5 + K2_up + 10 , layer='0')                            

            ###################################
            # 와이드 우기둥 우측 단면도    
            ###################################
                    
            if( JB2_up != JB2_down or K2_up != K2_down ):   

                Section_Xpos = R_Xpos + max(RH1,RH2) +  500
                Section_Ypos = R_Ypos

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(A2>0):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos - A2
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos 

                    line(doc, X1, Y1, X2, Y2, layer='0')
                    dim_vertical_right(doc, X2, Y2, X1, Y1,  50, 'dim', text_height=0.22, text_gap=0.07)
                    X3 = Section_Xpos
                    Y3 = Section_Ypos
                    lineto(doc,  X3, Y3, layer='0')
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # # vcut 표기                
                    # drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    # dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut 2개소", layer="dim", style='0')                                
                if(A2<1):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos                 
                    X3 = Section_Xpos
                    Y3 = Section_Ypos 
                    line(doc, X2, Y2, X3, Y3, layer='0')                               
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # vcut 표기                                
                    # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    # dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut", layer="dim", style='0')                                

                X4 = X3 - JB2_down * math.sin(Angle_radians)                
                Y4 = Y3 - JB2_down * math.cos(Angle_radians)        
                lineto(doc,  X4, Y4, layer='0')	
                dim_linear(doc, X4, Y4, X3, Y3,   "", 50,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)

                if(K2_down > 5):
                    X5 = X4
                    Y5 = Y4 - K2_down
                    lineto(doc,  X5, Y5, layer='0')	
                    dim_vertical_left(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
                else:
                    X5 = X4
                    Y5 = Y4
                    lineto(doc,  X5, Y5, layer='0')	

                if(SWAngle>0):
                    X6 = X5 + SW1
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                    X7 = X6 + SW2 * math.cos(SWAngle_radians)
                    Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                    lineto(doc,  X7, Y7, layer='0')
                    dim_linear(doc, X7, Y7, X6, Y6,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                    dim_angular(doc,  X7 , Y7, X6, Y6, X6 + 5 , Y6, X6, Y6,  150, direction="right" )
                if(SWAngle<1):
                    X6 = X5 + SW
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,"" , 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    X7 = X6
                    Y7 = Y6

                # if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc,  X3, Y3, X2-8 , Y2,  X3, Y3, X4, Y4,   25, direction="right" )		            
                    # if(K2_down>0):
                dim_angular(doc,   X4, Y4, X3 , Y3, X5, Y5,  X4, Y4, 30, direction="right" )    

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB2_up != JB2_down ):
                    rectangle(doc, X3 - 60, Y5 + (JB1_up+K1_up)/2 - 75 , X3 - 60 - 100, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')               
                    rectangle(doc, X3 - 60 - 200, Y5 + (JB1_up+K1_up)/2 - 75 , X3 - 60 - 100 - 200, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')      

                if(K2_up != K2_down ):
                    rectangle(doc, X5 - 50, Y5 - 10 , X5 - 50 - 100, Y5 + K2_down + 10 , layer='0')                             
                    rectangle(doc, X5 - 50 - 250, Y5 - 10 , X5 - 50 - 100 - 250, Y5 + K2_down + 10 , layer='0')                                        

            Abs_Ypos -= 2000 

def new_func():
    return 48   

##########################################################################################################################################
# 4CO멍텅구리 작도
##########################################################################################################################################
def execute_fourconormal():     

    # 시트 선택 (시트명을 지정)
    sheet_name = '4CO멍텅제작'  # 원하는 시트명으로 변경
    sheet = workbook[sheet_name]

    # 2차원 배열을 저장할 리스트 생성
    data_2d = []

    # 1행부터 20행까지의 값을 2차원 배열에 저장
    for row_num in range(7, sheet.max_row + 1):  # 7행부터 20행까지 반복
        cell_value = sheet.cell(row=row_num, column=2).value  # 2열의 값 가져오기
        if cell_value is not None and cell_value != 'F':  # 값이 None이 아닌 경우에만 배열에 추가
            row_values =  [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, sheet.max_column + 1)]        
            data_2d.append(row_values)

    # 엑셀 파일 닫기
    workbook.close()

    if not data_2d:
        print("해당 데이터가 없습니다.")
        return  # 함수를 빠져나감    

    ########################################################################################################################################################################
    # 4CO멍텅구리 상판 전개도
    ########################################################################################################################################################################

    Abs_Xpos = 63000
    Abs_Ypos = 0

    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        OP1 = row[2]
        OP2 = row[3]
        OP3 = row[4]
        R = row[5]
        JD1 = row[6]
        JD2 = row[7]
        JD3 = row[8]
        JB1_up = row[9]
        JB1_down = row[10]
        JB2_up = row[11]
        JB2_down = row[12]
        left_gap = row[13]
        right_gap = row[14]
        H1 = row[15]
        H2 = row[16]
        C1 = row[17]        
        C2 = row[18]
        A1 = row[19]
        A2 = row[20]
        LH1 = row[21]
        LH2 = row[22]
        RH1 = row[23]
        RH2 = row[24]        
        U = row[25]        
        G = row[26]        
        UW = row[27]        
        SW = row[28]        
        K1_up = row[29]        
        K1_down = row[30]        
        K2_up = row[31]        
        K2_down = row[32]                
        Upper_size = row[33]
        Left_side_size = row[34]
        Right_side_size = row[35]
        SW1 = row[36]
        SW2 = row[37]
        SWAngle = row[38]
        E1 = row[39]
        E2 = row[40]
        Angle = row[44]
        Top_pop1 = round(row[45] + 2, 1) #소수점 첫째자리까지만 나오게
        Top_pop2 = round(row[46] + 2, 1)         
        Kadapt =  row[56]   # K값 J값에 적용여부 넣기 1이면 넣기
        G_Setvalue = row[57]  # 손상민 소장 G값 강제로 적용 작지 52번째

        # 상판 E1, E2는 추영덕 소장이 K값을 상판에 표현하는 수치표현식인데, K값으로 환산해서 정리함
        #  K1_up = E1 - UW - 1.2

        # 데이터 검증 및 기본값 설정
        SWAngle = validate_or_default(SWAngle)
        K1_up = validate_or_default(K1_up)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_down = validate_or_default(K2_down)
        U = validate_or_default(U)
        G = validate_or_default(G)        
        G_Setvalue = validate_or_default(G_Setvalue)        
        OP1 = validate_or_default(OP1)
        OP2 = validate_or_default(OP2)
        OP3 = validate_or_default(OP3)
        Kadapt = validate_or_default(Kadapt)

        # print("작업소장 : ")
        # print(worker)
        # print("G_setvalue : ")
        # print(G_Setvalue)        
        if(worker=="손상민" and G_Setvalue > 0) :
            # 상부 돌출
            Top_pop1 = G_Setvalue
            Top_pop2 = G_Setvalue
            # print("G_setvalue : ")
            # print(G_Setvalue)
       
        if(Kadapt==1):
            K1_up = 0            
            K2_up = 0    
            if(E1>0):
                E1 = E1 + Vcut_rate
                E2 = E2 + Vcut_rate
        else:
            if(E1>0):                
                K1_up = E1 
                K2_up = E2                        
             
        if(K1_up>=10 and K2_up>=10 ):
            # 추영덕소장 E1, E2값 무시함 위의 조건이라면
            E1 = 0
            E2 = 0

        if(Top_pop1<4):
            Top_pop1 = 4
        if(Top_pop2<4):
            Top_pop2 = 4                

        if(OP1>0):             
            if(G > 0):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U + G - Vcut_rate*3
                X2 = Abs_Xpos + H1 + H2 + R 
                Y2 = Y1      
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Abs_Ypos + U - Vcut_rate * 2           
                line(doc, X2, Y2, X3, Y3, layer='레이져')          
            if(G < 1):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U  - Vcut_rate
                X2 = Abs_Xpos + H1 + H2 + R 
                Y2 = Y1      
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Y2

            X4 = X3
            Y4 = Abs_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')
            X5 = X4
            Y5 = Abs_Ypos - Top_pop2 + Vcut_rate        
            lineto(doc,  X5, Y5, layer='레이져')    
            X6 = X5 - H2
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='레이져')

            if(K1_up>0 or K2_up>0):
                X7 = Abs_Xpos + H1 + (R - OP1 - OP2 - OP3 - 4)/2  + OP1 + OP2 + OP3 + 4          
                Y7 = Abs_Ypos - JD3 + Vcut_rate                
            else:
                if(E1>0):
                    X7 = Abs_Xpos + H1 + (R-OP1-OP2-OP3-4)/2 + OP1+OP2+OP3+4
                    Y7 = Abs_Ypos - (JD3-E2) + Vcut_rate * 3                    
                else:
                    X7 = Abs_Xpos + H1 + (R-OP1-OP2-OP3-4)/2 + OP1+OP2+OP3+4 
                    Y7 = Abs_Ypos - JD3 + Vcut_rate * 3
            
            lineto(doc,  X7, Y7, layer='레이져')                      
            X8 = X7                
            Y8 = X7         

            X9 = X8
            if(E2>0):
                Y9 = Abs_Ypos - JD3 + Vcut_rate*3
            else:
                Y9 = Abs_Ypos - JD3 - K2_up + Vcut_rate*3
            lineto(doc,  X9, Y9, layer='레이져') 
            X10 = X9 
            if(E2>0):                
                Y10 = Abs_Ypos - JD3 - UW + Vcut_rate*5     
            else:
                Y10 = Abs_Ypos - JD3 - K2_up - UW + Vcut_rate*5     
            lineto(doc,  X10, Y10, layer='레이져')   
            X11 = Abs_Xpos +  H1 + (R - OP1 - OP2 - OP3 - 4)/2  + OP1 + OP2 +  4 
            if(E2>0):                
                Y11 = Abs_Ypos - JD3 - UW + Vcut_rate*5     
            else:                    
                Y11 = Abs_Ypos - JD3 - K2_up - UW + Vcut_rate*5     
            lineto(doc,  X11, Y11, layer='레이져')      

            X12 = X11 
            if(E2>0):                
                Y12 =  Abs_Ypos - JD3 + Vcut_rate*3
            else:                    
                Y12 =  Abs_Ypos - JD3 - K2_up + Vcut_rate*3
            lineto(doc,  X12, Y12, layer='레이져')         
            X13 = X12 - 2 
            Y13 = Y12 
            lineto(doc,  X13, Y13, layer='레이져')                                      

            X14 = X13                
            Y14 = Y13 - ((JD2+K2_up) - ( JD3+K2_up))               
            lineto(doc,  X14, Y14, layer='레이져')         
                
            X15 = X14
            Y15 = Y14 -  UW + Vcut_rate*2
            lineto(doc,  X15, Y15, layer='레이져')                                                      
            X16 = Abs_Xpos +  H1 + (R - OP1 - OP2 - OP3 - 4)/2  + OP1 + 2
            Y16 =  Y15
            lineto(doc,  X16, Y16, layer='레이져')                                                                   

            X17 = X16
            Y17 = Y14
            lineto(doc,  X17, Y17, layer='레이져')                                      
            X18 = X17                
            Y18 = Abs_Ypos - JD1 - K1_up + Vcut_rate*3
            lineto(doc,  X18, Y18, layer='레이져')                                                      
            X19 = X18 - 2
            Y19 = Y18
            lineto(doc,  X19, Y19, layer='레이져')     
                
            # 4co 추가된 부분   
            X20 = X19
            Y20 = Y19 -  UW + Vcut_rate*2
            lineto(doc,  X20, Y20, layer='레이져')                                                      
            X21 = Abs_Xpos +  H1 + (R - OP1 - OP2 - OP3 - 4)/2  
            Y21 =  Abs_Ypos - JD1 - K1_up - UW  + Vcut_rate*5
            lineto(doc,  X21, Y21, layer='레이져')                                                                   

            X22 = X21
            Y22 = Abs_Ypos - JD1 - K1_up  + Vcut_rate*3
            lineto(doc,  X22, Y22, layer='레이져')                                      
            X23 = X22                
            Y23 =  Abs_Ypos - JD1  + Vcut_rate*3
            lineto(doc,  X23, Y23, layer='레이져')                                                      
            
            # 기준선 
            X24 = X23     
            Y24 = Y23            

            X25 = Abs_Xpos + H1
            Y25 = Abs_Ypos - Top_pop2 + Vcut_rate 
            lineto(doc,  X25, Y25, layer='레이져')      
            X26 = Abs_Xpos
            Y26 = Y25
            lineto(doc,  X26, Y26, layer='레이져')      
            X27 = X26 
            Y27 = Abs_Ypos
            lineto(doc,  X27, Y27, layer='레이져')     
            if(G > 0):         
                X28 = X27 
                Y28 = Abs_Ypos + U - Vcut_rate * 2   
                lineto(doc,  X28, Y28, layer='레이져')      
                lineto(doc,  X1, Y1, layer='레이져')                  
            if(G < 1):         
                X28 = X27 
                Y28 = Abs_Ypos + U - Vcut_rate 
                lineto(doc,  X1, Y1, layer='레이져')               

            if args.opt15:
                drawcircle(doc, X10 + 12.5, Y10 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X9 - 12.5, Y9 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이

            # 상판 절곡선
            if(G>0):
                line(doc, X28, Y28, X3, Y3,  layer='22')        
            line(doc, X27, Y27, X4, Y4,  layer='22')            
    
            # 뒷날개쪽 3개
            line(doc, X22, Y22, X19, Y19,  layer='22')
            line(doc, X17, Y17, X14, Y14,  layer='22')
            line(doc, X12, Y12, X9, Y9,  layer='22')
            # 도어두께 표시하는 치수선
            dim_vertical_right(doc, X18, Y18, X17, Y17, 50, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X17, Y17, X16, Y16, 150, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_left(doc, X13, Y13, X14, Y14, 150, 'dim', text_height=0.22,  text_gap=0.07)                                       

            #  전개도 상부 치수선
            dim_linear(doc,  X1, Y1, X2, Y2, "", 130,  direction="up", layer='dim')

            # 상부 H1, OP, H2 치수선
            if(G > 0):   
                dim_linear(doc,  X1, Y1, X25, Y25,  "" , extract_abs(Y1,Y5) + 10 ,  direction="up", layer='dim')
                dim_linear(doc,  X25, Y25, X6, Y6,  "",  extract_abs(Y1,Y5) + U + G + 10 ,  direction="up", layer='dim')
                dim_linear(doc,  X6, Y6, X2, Y2,  "",  extract_abs(Y1,Y5)  + U + G + 10,  direction="up", layer='dim')

            # 우측 절곡치수선
            if(U>0):        
                dim_vertical_right(doc, X2, Y2, X3, Y3, 110 , 'dim', text_height=0.22, text_gap=0.07)  
            if(G>0):                    
                dim_vertical_right(doc, X4,  Y4, X3, Y3,  extract_abs(X2,X6) + 50 , 'dim', text_height=0.22, text_gap=0.07)  

            dim_vertical_right(doc, X4, Y4,  X9, Y9,  extract_abs(X2,X6) +  50, 'dim', text_height=0.22, text_gap=0.07)              
            dim_vertical_right(doc, X9, Y9, X10, Y10, extract_abs(X2,X6) + extract_abs(X5,X6) +  extract_abs(X9,X6) + 50, 'dim', text_height=0.22, text_gap=0.07)   # 뒷날개
            dim_vertical_right(doc, X2, Y2,  X15, Y15, extract_abs(X2,X7) +  100, 'dim', text_height=0.22, text_gap=0.07)      # 전체치수       

            # 좌측 절곡치수선        
            if(U>0):                    
                dim_vertical_left(doc, X1, Y1, X28, Y28,   120, 'dim', text_height=0.22, text_gap=0.07)  
            if(G>0):                        
                dim_vertical_left(doc, X27, Y27, X28, Y28, 70 , 'dim', text_height=0.22, text_gap=0.07)                 
            dim_vertical_left(doc, X22, Y22, X27, Y27,  extract_abs(X24,X28) + 70 , 'dim', text_height=0.22, text_gap=0.07)                  
            # 뒷날개
            dim_vertical_left(doc, X22, Y22, X21, Y21,  extract_abs(X24,X28) +  extract_abs(X24,X21) + 70, 'dim', text_height=0.22, text_gap=0.07)                          
            # 좌측 전체
            dim_vertical_left(doc, X16, Y16, X1, Y1,   extract_abs(X1,X16) + 220, 'dim', text_height=0.22, text_gap=0.07)                          

            # 쫄대타입 테둘림유 상부 Vcut 치수 표기
            if(U>0 and G>0):   
                dim_vertical_left(doc, X1+H1+(OP1+OP2+OP3)/2, Y1, X1+H1+(OP1+OP2+OP3)/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            
            # 하부 치수선 4개 노칭 1개 위방향
            dim_linear(doc,  X25, Y25, X21, Y21,  "",  extract_abs(Y21, Y25) + 160 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X16, Y16, X15, Y15,  "",  80 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X21, Y21, X20, Y20,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X11, Y11, X10, Y10,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)              
            dim_linear(doc,  X21, Y21, X10, Y10,  "",  160 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)              
            dim_linear(doc,  X10, Y10, X6, Y6,    "",  160 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)              

            # 노칭 2곳
            dim_linear(doc,  X19, Y19, X18, Y18,  "",  50  ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X13, Y13, X12, Y12,  "",  50  ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      
                        
            # 상판 돌출부위 H1, H2 치수선 좌,우
            if(G > 0):  
                dim_vertical_right(doc, X27, Y27, X25, Y25, 100, 'dim', text_height=0.22,  text_gap=0.07)
                dim_vertical_left(doc, X4, Y4, X6, Y6, 100, 'dim', text_height=0.22,  text_gap=0.07)

            # 기둥과의 접선 부위 aligned 표현 각도표현
            # 좌측
            dim_linear(doc, X25, Y25, X22, Y22,   "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) 
            dim_angular(doc, X23, Y23, X25, Y25, X21, Y21, X22, Y22 + 10,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다
            # 우측        
            dim_linear(doc, X9, Y9, X6, Y6,  "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) 
            dim_angular(doc, X10, Y10  , X7, Y7 + 10,  X7, Y7,  X6, Y6,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다

            # 상판 K1 K2 값 부분 치수선 
            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                dim_vertical_right(doc, X23, Y23, X22, Y22, 80, 'dim', text_height=0.22,  text_gap=0.07)  
                dim_vertical_left(doc, X9, Y9, X8, Y8, 80, 'dim', text_height=0.22,  text_gap=0.07)  

            # JD 간격 치수선 그림
            dim_vertical_right(doc, X27, Y27, X22, Y22, 220, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X1+H1+(OP1+OP2+OP3)/2 , Y4, X14, Y14, 220, 'dim', text_height=0.22,  text_gap=0.07)                  
            dim_vertical_left(doc, X4, Y4, X9, Y9, (OP1+OP2+OP3)/4 , 'dim', text_height=0.22,  text_gap=0.07)     

            # 상부 Vcut 치수 표기
            if(U>0 and G>0):   
               dim_vertical_left(doc, X1+H1+(OP1+OP2+OP3)/2, Y1,X1+H1+(OP1+OP2+OP3)/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  

            # 상판 좌우가 다를때 Vcut 상부에 치수선 표기
            # if((JD1 != JD2 or K1_up != K2_up) and G>0):
            #     dim_vertical_left(doc, X1 + R/2, Y1, X1 + R/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            # 문구 설정
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"        
            Textstr = f"({Pre_text}{str(Floor_des)})"       

            X2 = Abs_Xpos + (H1 + R + H2)/2 - len(Textstr)*50/2
            Y2 = Abs_Ypos - JD1/1.5
            draw_Text(doc, X2, Y2, 50, str(Textstr), '레이져')

            # 문구            
            sizedes = Upper_size                  
            
            Textstr = f"Part Name : 상판({Pre_text}{str(Floor_des)})"    
            draw_Text(doc, X9+150, Y9 + 50 - 270, 28, str(Textstr), '0')
            Textstr = f"Mat.Spec : {normal_material}"    
            draw_Text(doc, X9+150, Y9 - 270, 28, str(Textstr), '0')
            Textstr = f"Size : {sizedes}"    
            draw_Text(doc, X9+150, Y9 - 50 - 270 , 28, str(Textstr), '0')
            Textstr = f"Quantity : 1 EA"    
            draw_Text(doc, X9+150, Y9 - 100 - 270 , 28, str(Textstr), '0')                        

            # 도면틀 넣기
            BasicXscale = 2560
            BasicYscale = 1550
            TargetXscale = 800 + R + U * 2 + 200
            TargetYscale = 300 + JD2 + G + U + UW
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            insert_frame(X1 - (U + 580) , Y1 - ( 800 + JD2 + K1_up + G + U + UW) , frame_scale, "top_normal", workplace)        

            ###########################################################################################################################################################################
            # 멍텅구리 상판 좌측 단면도    (좌우가 다른 모양일때 단면도를 그려준다)
            ###########################################################################################################################################################################
            if(JD1 != JD2 or K1_up != K2_up):
                Section_Xpos = Abs_Xpos - H1 - 400           

                if(G > 0):
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos - G
                    X2 = Section_Xpos
                    Y2 = Abs_Ypos  

                    line(doc, X1, Y1, X2, Y2, layer='0')                    
                    dim_vertical_left(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)     
                    X3 = Section_Xpos + U
                    Y3 = Abs_Ypos  
                    lineto(doc,  X3, Y3, layer='0')    
                    dim_linear(doc,  X2, Y2, X3, Y3, "", 50,  direction="up", layer='dim')                
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                    dim_leader_line(doc, X3, Y3 , X2-130 , Y2-100, "V-cut 2개", layer="dim", style='0')                
                if(G < 1):
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos 
                    X2 = Section_Xpos
                    Y2 = Abs_Ypos                  
                    X3 = Section_Xpos + U
                    Y3 = Abs_Ypos   
                    line(doc, X2, Y2, X3, Y3, layer='0')                  
                    dim_linear(doc,  X2, Y2, X3, Y3, "", 50,  direction="up", layer='dim')    
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                    dim_leader_line(doc, X3, Y3 , X2-130 , Y2-100, "V-cut 1개", layer="dim", style='0')                                
                
                X4 = X3 
                Y4 = Y3 - JD1
                lineto(doc,  X4, Y4, layer='0')                   
                X5 = X4 
                Y5 = Y4 - K1_up
                lineto(doc,  X5, Y5, layer='0')   
                X6 = X5 
                Y6 = Y5 - (JD2 - JD1)
                lineto(doc,  X6, Y6, layer='0')                   
                if(K1_up > 0):
                    dim_vertical_right(doc,  X3, Y3, X4, Y4, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X4, Y4, X5, Y5, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X5, Y5, X6, Y6, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X3, Y3, X6, Y6, 110,'dim' ,text_height=0.22,  text_gap=0.07)                                     
                else:
                    dim_vertical_right(doc,  X3, Y3, X5, Y5, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X5, Y5, X6, Y6, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X3, Y3, X6, Y6, 110,'dim' ,text_height=0.22,  text_gap=0.07)     

                # drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')         
                X7 = X6 - UW
                Y7 = Y6
                lineto(doc,  X7, Y7, layer='0')   
                dim_linear(doc,  X7, Y7, X6, Y6,  "", 80,  direction="down", layer='dim')            
                
                if(K1_up > 0):                 
                    # K값 부분 라인 그리기
                    line(doc, X4, Y4, X4 - UW, Y4, layer='22')     
                line(doc, X5, Y5, X5 - UW, Y5, layer='22')         
                
                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JD1 != JD2 or K1_up != K2_up ):
                    rectangle(doc, X4+50, Y5+(JD1+K1_up)/2 - 75 , X4 + 50 + 70, Y5+(JD1+K1_up)/2 + 75, layer='0')               
                    rectangle(doc, X4+285,  Y5+(JD1+K1_up)/2 - 75 , X4+285+70, Y5+(JD1+K1_up)/2 + 75 , layer='0')                                                         
                
            ################################################################################################################################################################
            # 멍텅구리 상판 우측 단면도 
            ################################################################################################################################################################
            Section_Xpos = Abs_Xpos + R + H1 + H2 + 400

            if(G>0):
                X1 = Section_Xpos 
                Y1 = Abs_Ypos - G
                X2 = Section_Xpos
                Y2 = Abs_Ypos  

                line(doc, X1, Y1, X2, Y2, layer='0')                    
                dim_vertical_right(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)     
                X3 = Section_Xpos - U
                Y3 = Abs_Ypos  
                lineto(doc,  X3, Y3, layer='0')    
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim')
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                dim_leader_line(doc, X3, Y3 , X3+50 , Y2-100, "V-cut 2개", layer="dim", style='0')                

            if(G<1):
                X1 = Section_Xpos 
                Y1 = Abs_Ypos
                X2 = Section_Xpos
                Y2 = Abs_Ypos  
                X3 = Section_Xpos - U
                Y3 = Abs_Ypos  
                line(doc, X1, Y1, X3, Y3, layer='0')                            
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim')    
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                dim_leader_line(doc, X3, Y3 , X3 + 50 , Y2-100, "V-cut 1개", layer="dim", style='0')                            
                        
            X4 = X3 
            Y4 = Y3 - JD3
            lineto(doc,  X4, Y4, layer='0')                   
            X5 = X4 
            Y5 = Y4 - K1_up
            lineto(doc,  X5, Y5, layer='0')   
            X6 = X5 
            Y6 = Y5 - (JD2 - JD3)
            lineto(doc,  X6, Y6, layer='0')                   
            if(K1_up > 0):
                dim_vertical_left(doc,  X3, Y3, X4, Y4, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                dim_vertical_left(doc,  X4, Y4, X5, Y5, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                dim_vertical_left(doc,  X5, Y5, X6, Y6, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                dim_vertical_left(doc,  X3, Y3, X6, Y6, 110,'dim' ,text_height=0.22,  text_gap=0.07)                                     
            else:
                dim_vertical_left(doc,  X3, Y3, X5, Y5, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                dim_vertical_left(doc,  X5, Y5, X6, Y6, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                dim_vertical_left(doc,  X3, Y3, X6, Y6, 110,'dim' ,text_height=0.22,  text_gap=0.07)     

            # drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')         
            X7 = X6 + UW
            Y7 = Y6
            lineto(doc,  X7, Y7, layer='0')   
            dim_linear(doc,  X7, Y7, X6, Y6,  "", 80,  direction="down", layer='dim')    
            
            # drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
            
            if(K2_up > 0):                 
                # K값 부분 라인 그리기
                line(doc, X4, Y4, X4 + UW, Y4, layer='22')      
            line(doc, X5, Y5, X5 + UW, Y5, layer='22')   

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JD1 != JD2 or K1_up != K2_up ):
                rectangle(doc, X4-50, Y5+(JD2+K2_up)/2 - 75 , X4 - 50 - 70, Y5+(JD2+K2_up)/2 + 75, layer='0')               
                rectangle(doc, X4-285,  Y5+(JD2+K2_up)/2 - 75 , X4-285-70,  Y5+(JD2+K2_up)/2 + 75 , layer='0')                                                         
            
            # 다음 도면 간격 계산
            Abs_Ypos -= 4000

    Abs_Xpos = 63000 + 4000
    Abs_Ypos = 500

    ############################################################################################################################################################################
    # 4CO멍텅구리 좌기둥 전개도
    ############################################################################################################################################################################
    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        OP1 = row[2]
        OP2 = row[3]
        OP3 = row[4]
        R = row[5]
        JD1 = row[6]
        JD2 = row[7]
        JD3 = row[8]
        JB1_up = row[9]
        JB1_down = row[10]
        JB2_up = row[11]
        JB2_down = row[12]
        left_gap = row[13]
        right_gap = row[14]
        H1 = row[15]
        H2 = row[16]
        C1 = row[17]        
        C2 = row[18]
        A1 = row[19]
        A2 = row[20]
        LH1 = row[21]
        LH2 = row[22]
        RH1 = row[23]
        RH2 = row[24]        
        U = row[25]        
        G = row[26]        
        UW = row[27]        
        SW = row[28]        
        K1_up = row[29]        
        K1_down = row[30]        
        K2_up = row[31]        
        K2_down = row[32]                
        Upper_size = row[33]
        Left_side_size = row[34]
        Right_side_size = row[35]
        SW1 = row[36]
        SW2 = row[37]
        SWAngle = row[38]
        E1 = row[39]
        E2 = row[40]
        Angle = row[44]
        Top_pop1 = round(row[45] + 2, 1) #소수점 첫째자리까지만 나오게
        Top_pop2 = round(row[46] + 2, 1)         
        Kadapt =  row[56]   # K값 J값에 적용여부 넣기 1이면 넣기
        G_Setvalue = row[57]  # 손상민 소장 G값 강제로 적용 작지 52번째

        # 상판 E1, E2는 추영덕 소장이 K값을 상판에 표현하는 수치표현식인데, K값으로 환산해서 정리함
        #  K1_up = E1 - UW - 1.2

        # 데이터 검증 및 기본값 설정
        SWAngle = validate_or_default(SWAngle)
        K1_up = validate_or_default(K1_up)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_down = validate_or_default(K2_down)
        U = validate_or_default(U)
        G = validate_or_default(G)        
        G_Setvalue = validate_or_default(G_Setvalue)        
        OP1 = validate_or_default(OP1)
        OP2 = validate_or_default(OP2)
        OP3 = validate_or_default(OP3)
        Kadapt = validate_or_default(Kadapt)
      
        if(Kadapt==1):
            K1_up = 0
            K1_down = 0
            K2_up = 0
            K2_down = 0

        if(OP1>0):  
            if (SWAngle>1):  #SWAngle 0 이상인 경우            
                if(K1_up>0):     
                    if (SWAngle == 90):       
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 3 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = Abs_Xpos + LH2             
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                        lineto(doc,  X5, Y5, layer='레이져')                    
                    else:
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 2 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = Abs_Xpos + LH2             
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                        lineto(doc,  X5, Y5, layer='레이져')                    
                else:
                    if (SWAngle == 90):                     
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 3  - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4                
                    else:
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 2  - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4                

            if (SWAngle<1):  #SWAngle 0 이하인 경우
                if(K1_up>0):            
                    X1 = Abs_Xpos 
                    Y1 = Abs_Ypos + JB1_up + K1_up + SW - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + LH2 
                    Y2 = Abs_Ypos + JB1_down + K1_down + SW - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')        
                    dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                    X3 = X2            
                    Y3 = Y2
                    X4 = Abs_Xpos + LH2             
                    Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = Abs_Xpos + LH2             
                    Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                    lineto(doc,  X5, Y5, layer='레이져')                    
                else:
                    X1 = Abs_Xpos 
                    Y1 = Abs_Ypos + JB1_up + SW - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + LH2 
                    Y2 = Abs_Ypos + JB1_down + SW - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')        
                    dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                    X3 = X2            
                    Y3 = Y2
                    X4 = Abs_Xpos + LH2             
                    Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = X4             
                    Y5 = Y4                

            X6 = Abs_Xpos + LH1             
            Y6 = Abs_Ypos
            lineto(doc,  X6, Y6, layer='레이져')
            if(A1>0):       
                X7 = Abs_Xpos + LH1   
                Y7 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X7, Y7, layer='레이져')                
                X8 = Abs_Xpos + LH1  
                Y8 = Abs_Ypos - C1 - A1 + Vcut_rate * 3
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = Abs_Xpos             
                Y10 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X10, Y10, layer='레이져')

            if(A1<1):       
                X7 = Abs_Xpos + LH1   
                Y7 = Abs_Ypos - C1 + Vcut_rate 
                lineto(doc,  X7, Y7, layer='레이져')                
                X8 = Abs_Xpos + LH1  
                Y8 = Y7
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = X9
                Y10 = Y9            

            X11 = Abs_Xpos 
            Y11 = Abs_Ypos
            lineto(doc,  X11, Y11, layer='레이져')        

            if (SWAngle>0): 
                if(K1_up>0):  
                    if (SWAngle == 90):                    
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up  - Vcut_rate   
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = Abs_Xpos 
                        Y13 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X13, Y13, layer='레이져')
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 3 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up  - Vcut_rate   
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = Abs_Xpos 
                        Y13 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X13, Y13, layer='레이져')
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                else:  
                    if (SWAngle == 90):   
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate     
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = X12
                        Y13 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate                 
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 3 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                          
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate     
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = X12
                        Y13 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate                 
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                          
                    
            if (SWAngle<1): 
                if(K1_up>0):  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + JB1_up  - Vcut_rate   
                    lineto(doc,  X12, Y12, layer='레이져')
                    X13 = Abs_Xpos 
                    Y13 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                    lineto(doc,  X13, Y13, layer='레이져')
                    X14 = Abs_Xpos 
                    Y14 = Abs_Ypos + JB1_up + K1_up + SW  - Bending_rate * 2 - Vcut_rate  
                    lineto(doc,  X14, Y14, layer='레이져')  
                    lineto(doc,  X1, Y1, layer='레이져')    
                else:  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate 
                    lineto(doc,  X12, Y12, layer='레이져')
                    X13 = X12
                    Y13 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate 
                    X14 = X13
                    Y14 = Y13
                    lineto(doc,  X1, Y1, layer='레이져')   

            # 좌기둥 기준선 
            if(SW1>0):
                line(doc, X14, Y14, X3, Y3,  layer='22')
            line(doc, X13, Y13, X4, Y4,  layer='22')
            if(K1_up>0):
                line(doc, X12, Y12, X5, Y5,  layer='22')
            line(doc, X11, Y11, X6, Y6,  layer='22')
            if(A1>0):
                line(doc, X10, Y10, X7, Y7,  layer='22')                

            # 우측 절곡치수선
            if(SW1>0):
                dim_vertical_right(doc, X2,  Y2, X3, Y3,  extract_abs(X2,X3) + 160, 'dim', text_height=0.22, text_gap=0.01)      
            dim_vertical_right(doc, X3, Y3, X4,  Y4,   extract_abs(X3,X4) + 80, 'dim', text_height=0.22, text_gap=0.01)      
            if(K1_up>0):
                dim_vertical_right(doc, X4,  Y4, X5, Y5,  extract_abs(X4,X5) +  130, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_right(doc, X6, Y6, X5,  Y5, extract_abs(X6,X5) +  80, 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_right(doc,  X7, Y7, X6,  Y6,   extract_abs(X6,X7) + 80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_right(doc, X7,  Y7, X8, Y8,   extract_abs(X7,X8) + 150, 'dim', text_height=0.22, text_gap=0.01)  
                # V컷 치수표현
                dim_vertical_left(doc, X6,  Y6, X8, Y8,   extract_abs(X6,X8) + 100, 'dim', text_height=0.22, text_gap=0.01)  
            dim_vertical_right(doc, X2,  Y2, X8, Y8,   extract_abs(X2,X8) + 230 , 'dim', text_height=0.22, text_gap=0.01)      

            # 좌측 절곡치수선  
            if(SW1>0):                       
                dim_vertical_left(doc, X1, Y1, X14, Y14,  160, 'dim', text_height=0.22, text_gap=0.01)  
                dim_vertical_left(doc, X13, Y13, X14, Y14,   80 , 'dim', text_height=0.22, text_gap=0.01)  
            else:
                dim_vertical_left(doc, X1, Y1, X13, Y13,   80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(K1_up>0):
                dim_vertical_left(doc, X13, Y13, X12, Y12,  130 , 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X11, Y11, X12, Y12, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_left(doc, X11, Y11, X10, Y10,  80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_left(doc, X9, Y9, X10, Y10,  150 , 'dim', text_height=0.22, text_gap=0.01)
            dim_vertical_left(doc, X1, Y1, X9,  Y9,  230 , 'dim', text_height=0.22, text_gap=0.01)

            # 기둥 하부 치수선 2    
            dim_linear(doc,   X9, Y9, X8, Y8, "", 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   

            if LH1>LH2:
                dim_linear(doc,  X2, Y2, X6, Y6, "",  100 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)       
            if LH1<LH2:    
                dim_linear(doc,  X4, Y4, X8, Y8, "", extract_abs(Y4, Y8) + 100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            if args.opt15:
                drawcircle(doc, X1 + 12.5, Y1 - 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X2 - 12.5, Y2 - 5, 2.5 , layer='레이져') # 도장홀 5파이        

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"       

            Xpos = Abs_Xpos + 500 
            Ypos = Abs_Ypos + JB1_up / 2.3
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-좌"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')        

            Xpos = Abs_Xpos + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-좌"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Mat.Spec : {normal_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/1.2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Size : {Left_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/1.2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')       

            # 멍텅구리 기둥 좌우 도면틀 넣기
            BasicXscale = 4487
            BasicYscale = 2770
            # 우측단면도가 있을 경우 감안해야 함. 
            if( JB1_up != JB1_down or K1_up != K1_down  or JB2_up != JB2_down or K2_up != K2_down ):                                 
                TargetXscale = 550 + max(LH1, LH2) + 1000
            else:
                TargetXscale = 550 + max(LH1, LH2) + 550

            TargetYscale = 320*4 + (JB1_up + C1 + A1 + K1_up + SW) * 2
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            # 우측 단면도가 있는 경우 도면틀 기본점을 아래로 250 내림    
            if( JB1_up != JB1_down or K1_up != K1_down  or JB2_up != JB2_down or K2_up != K2_down ):                                 
                insert_frame(X1 -  710 , Abs_Ypos - (JB2_up+C2+A2+K2_up + SW + 1200 + C1 + A1 ) - 250 , frame_scale, "side_normal", workplace)                
            else:
                insert_frame(X1 -  710 , Abs_Ypos - (JB2_up+C2+A2+K2_up + SW + 1200 + C1 + A1 ) , frame_scale, "side_normal", workplace)                

            ##############################################################################################################################################################
            # 4CO멍텅구리 좌기둥 좌측 단면도    
            ##############################################################################################################################################################
            Section_Xpos = Abs_Xpos - 400
            Section_Ypos = Abs_Ypos + JB1_up + K1_up

            # 가정: SWAngle은 도 단위로 주어진 각도
            SWAngle_radians = math.radians(SWAngle)

            if(SWAngle>0):
                X1 = Section_Xpos - SW1 - SW2 * math.cos(SWAngle_radians)
                Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                X2 = Section_Xpos - SW1
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos
                
                line(doc, X2, Y2, X1, Y1, layer='0')               
                line(doc, X2, Y2, X3, Y3, layer='0')        
                dim_linear(doc, X1, Y1, X2, Y2,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # 각도 90도가 아니면 각도 표기
                if (SWAngle != 90):               
                    dim_angular(doc, X2 , Y2, X1, Y1, X2-30, Y2, X3-30, Y3,  160, direction="left" )
                    # help(ezdxf.layouts.Modelspace.add_angular_dim_2l)
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)   

            if(SWAngle<1):
                X1 = Section_Xpos - SW
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos            
                line(doc,  X1, Y1, X2, Y2, layer='0')                                                   
                dim_linear(doc, X1, Y1, X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

            if(K1_up>5):
                X4 = X3
                Y4 = Y3 - K1_up
                lineto(doc,  X4, Y4, layer='0')       
            else:
                X4 = X3
                Y4 = Y3

            # Angle은 도 단위로 주어진 각도
            Angle_radians = math.radians(Angle)

            X5 = X4 - JB1_up * math.sin(Angle_radians)                
            Y5 = Y4 - JB1_up * math.cos(Angle_radians)  
            lineto(doc,  X5, Y5, layer='0')    
            X6 = X5 - C1
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='0')   

            if(A1>0):
                X7 = X6
                Y7 = Y6 + A1
                lineto(doc,  X7, Y7, layer='0') 
                drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut 2개소", layer="dim", style='0')
                # A 치수
                dim_vertical_left(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
            if(A1<1):
                X7 = X6
                Y7 = Y6            
                drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut", layer="dim", style='0')

            # C값 치수선
            dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            if(K1_up>5):
                dim_vertical_right(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                # if (Angle > 1):   # 직각이 아닌경우                        
                dim_angular(doc, X4, Y4, X5, Y5, X4, Y4, X3, Y3,  30, direction="left" )

            # JB사선 치수선
            dim_linear(doc, X4, Y4, X5, Y5,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
            # if (Angle > 1):   # 직각이 아닌경우            
            dim_angular(doc, X6, Y6, X5, Y5, X5, Y5, X4, Y4,  20, direction="left" )
            

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB1_up != JB1_down ):
                rectangle(doc, X3+10, Y3 - JB2_up/2 - 75, X3 + 10 + 100, Y3 - JB1_up/2 + 75 , layer='0')               
                rectangle(doc, X3+60+200, Y3 - JB2_up/2 - 75 , X3 + 60 + 300, Y3 - JB1_up/2 + 75 , layer='0')               
            if(K1_up != K1_down ):
                rectangle(doc, X5 + 50, Y5 - 10 , X5 + 50 + 100, Y5 + K1_up + 10 , layer='0')                         
                rectangle(doc, X5 + 10 + 250, Y5 - 10 , X5 + 10 + 300, Y5 + K1_up + 10 , layer='0')                            

            #######################################################################################################################################################################
            # 4CO 멍텅구리 좌기둥 우측 단면도    
            #######################################################################################################################################################################
            
            if( JB1_up != JB1_down or K1_up != K1_down ):                  
                Section_Xpos = Abs_Xpos + LH1 + 400
                Section_Ypos = Abs_Ypos + JB1_up + K1_up

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(SWAngle>0):  
                    X1 = Section_Xpos + SW1 + SW2 * math.cos(SWAngle_radians)
                    Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                    X2 = Section_Xpos + SW1
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc, X2, Y2, X1, Y1, layer='0')                   
                    line(doc, X2, Y2, X3, Y3, layer='0')        
                    dim_linear(doc, X2, Y2, X1, Y1,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # 각도 90도가 아니면 각도 표기
                    if (SWAngle != 90):               
                        dim_angular(doc,  X3+5, Y3, X2+5, Y2, X1, Y1, X2 , Y2,  150, direction="right" )
                        # help(ezdxf.layouts.Modelspace.add_angular_dim_2l)
                    dim_linear(doc, X3, Y3,  X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)    

                if(SWAngle<1):  
                    X1 = Section_Xpos + SW
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos 
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc,   X1, Y1, X2, Y2,  layer='0')                                   
                    dim_linear(doc,  X2, Y2, X1, Y1,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

                if(K1_down>5):
                    X4 = X3
                    Y4 = Y3 - K1_down
                    lineto(doc,  X4, Y4, layer='0')       
                else:
                    X4 = X3
                    Y4 = Y3

                # Angle은 도 단위로 주어진 각도
                Angle_radians = math.radians(Angle)

                X5 = X4 + JB1_down * math.sin(Angle_radians)                
                Y5 = Y4 - JB1_down * math.cos(Angle_radians)  
                lineto(doc,  X5, Y5, layer='0')    
                X6 = X5 + C1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')   

                if(A1>0):
                    X7 = X6
                    Y7 = Y6 + A1
                    lineto(doc,  X7, Y7, layer='0') 
                    drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                    drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                    dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut 2개소", layer="dim", style='0')
                    # A 치수
                    dim_vertical_right(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
                if(A1<1):
                    X7 = X6
                    Y7 = Y6            
                    drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                    dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut", layer="dim", style='0')

                # C값 치수선
                dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                if(K1_up>5):
                    dim_vertical_left(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                    # if (Angle > 1):   # 직각이 아닌경우                        
                    dim_angular(doc,  X3, Y3, X4 , Y4,  X5, Y5, X4, Y4,   15, direction="right" )

                # JB사선 치수선
                dim_linear(doc,  X5, Y5, X4, Y4,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
                # if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc,  X5, Y5, X4, Y4, X5, Y5, X6 , Y6,  15, direction="right" )

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB1_up != JB1_down ):
                    rectangle(doc, X3-10, Y3 - JB1_up/2 - 75, X3 - 10 - 100, Y3 - JB1_up/2 + 75 , layer='0')               
                    rectangle(doc, X3 - 60 - 200, Y3 - JB1_up/2 - 75 , X3 - 60 - 300, Y3 - JB1_up/2 + 75 , layer='0')               
                if(K1_up != K1_down ):
                    rectangle(doc, X5 - 50, Y5 - 10 , X5 - 50 - 100, Y5 + K1_up + 10 , layer='0')                         
                    rectangle(doc, X5 - 10 + 250, Y5 - 10 , X5 - 10 - 300, Y5 + K1_up + 10 , layer='0')                              
            
            ######################################################################################################################################################
            # 멍텅구리 우기둥 전개도
            # 멍텅구리 우기둥 전개도
            ######################################################################################################################################################

            Abs_Ypos -= 2000 
            R_Ypos = Abs_Ypos + 1400 - (C1 + C2 + A1 + A2)

            if(A2>0):            
                X1 = Abs_Xpos 
                Y1 = R_Ypos + C2 + A2 - Vcut_rate * 3
                X2 = Abs_Xpos + RH1 
                Y2 = R_Ypos + C2 + A2 - Vcut_rate * 3
                line(doc, X1, Y1, X2, Y2, layer='레이져')        
                dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                X3 = Abs_Xpos + RH1 
                Y3 = R_Ypos + C2 - Vcut_rate * 2
                lineto(doc,  X3, Y3, layer='레이져')                
            if(A2<1):
                X1 = Abs_Xpos 
                Y1 = R_Ypos + C2  - Vcut_rate 
                X2 = Abs_Xpos + RH1 
                Y2 = Y1
                line(doc, X1, Y1, X2, Y2, layer='레이져')        
                dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                X3 = X2
                Y3 = Y2

            X4 = Abs_Xpos + RH1 
            Y4 = R_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')
            if (SWAngle>1):  #SWAngle 0 이상인 경우    헤밍구조인 경우        
                if(K2_up>0):   
                    if (SWAngle == 90):                       
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = Abs_Xpos + RH2 
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = Abs_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')
                    else:
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = Abs_Xpos + RH2 
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = Abs_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')
                if(K2_up<1):   #K값이 없는 경우
                    if (SWAngle == 90):                      
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11
                    else:
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11

            if (SWAngle<1):  #SWAngle 0     일반구조
                if(K2_up>0):   
                    X5 = Abs_Xpos + RH2 
                    Y5 = R_Ypos - JB2_down +  Vcut_rate
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = Abs_Xpos + RH2 
                    Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                    lineto(doc,  X6, Y6, layer='레이져')
                    X7 = Abs_Xpos + RH2             
                    Y7 = R_Ypos - JB2_down - K2_down - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = Abs_Xpos 
                    Y9 = R_Ypos - JB2_up - K2_up - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = Abs_Xpos 
                    Y10 = Y9                
                    X11 = Abs_Xpos 
                    Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = Abs_Xpos 
                    Y12 = R_Ypos - JB2_up +  Vcut_rate
                    lineto(doc,  X12, Y12, layer='레이져')
                if(K2_up<1):   #K값이 없는 경우
                    X5 = Abs_Xpos + RH2 
                    Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = X5
                    Y6 = Y5                
                    X7 = Abs_Xpos + RH2             
                    Y7 = R_Ypos - JB2_down  - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = Abs_Xpos 
                    Y9 = R_Ypos - JB2_up  - SW  + Bending_rate * 2 + Vcut_rate
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = X9
                    Y10 = Y9                
                    X11 = Abs_Xpos 
                    Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = X11
                    Y12 = Y11                

            X13 = Abs_Xpos 
            Y13 = R_Ypos
            lineto(doc,  X13, Y13, layer='레이져')

            if(A2>0):
                X14 = Abs_Xpos 
                Y14 = R_Ypos + C2 - Vcut_rate * 2
                lineto(doc,  X14, Y14, layer='레이져')
                lineto(doc,  X1, Y1, layer='레이져')
            if(A2<1):
                X14 = X1
                Y14 = Y1
                lineto(doc,  X14, Y14, layer='레이져')

            # 기둥 절곡선 1
            if(A2>1):        
                line(doc, X14, Y14, X3, Y3,  layer='22')
            # C2    
            line(doc, X13, Y13, X4, Y4,  layer='22')
            # JB2
            line(doc, X12, Y12, X5, Y5,  layer='22')
            if(K2_up>0):
                line(doc, X11, Y11, X6, Y6,  layer='22')
            # 해밍구조
            if (SWAngle>0):   
                line(doc, X10, Y10, X7, Y7,  layer='22')

            # 멍텅구리 우기둥 우측 절곡치수선
            if(A2>1):  
                dim_vertical_right(doc, X3, Y3, X2,  Y2,  150, 'dim', text_height=0.22,  text_gap=0.07)      
                dim_vertical_left(doc, X4, Y4, X2,  Y2, extract_abs(X2,X4) + 150, 'dim', text_height=0.22,  text_gap=0.07)      
            # C    
            dim_vertical_right(doc, X3, Y3, X4,  Y4, 60, 'dim', text_height=0.22,  text_gap=0.07)      
            dim_vertical_right(doc, X5, Y5, X4,  Y4, extract_abs(X6,X4) + 60, 'dim', text_height=0.22,  text_gap=0.07) 
            if(K2_down>0): 
                dim_vertical_right(doc, X6, Y6, X5,  Y5, extract_abs(X6,X4) + 110, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X6, Y6, X7,  Y7, extract_abs(X6,X4) + 60, 'dim', text_height=0.22,  text_gap=0.07)  
            if (SWAngle>0):   
                dim_vertical_right(doc, X7, Y7, X8,  Y8, extract_abs(X6,X4) + 200, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X2, Y2, X8,  Y8, extract_abs(X2,X8) + 250, 'dim', text_height=0.22,  text_gap=0.07)  

            # 멍텅구리 우기둥 좌측 절곡치수선
            if(A2>1):          
                dim_vertical_left(doc, X1, Y1, X14, Y14, 150 , 'dim', text_height=0.22, text_gap=0.01)  
            # C    
            dim_vertical_left(doc, X13, Y13, X14, Y14,   50 , 'dim', text_height=0.22, text_gap=0.01)  
            # JB
            dim_vertical_left(doc, X13, Y13, X12, Y12,  50 , 'dim', text_height=0.22, text_gap=0.01)              
            if(K2_up>0): 
                dim_vertical_left(doc, X11, Y11, X12, Y12,  extract_abs(X11,X12) + 110, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X11, Y11, X10, Y10,  extract_abs(X11,X10) + 50, 'dim', text_height=0.22, text_gap=0.01)  
            if (SWAngle>0):               
                dim_vertical_left(doc, X9, Y9, X10, Y10,  extract_abs(X9,X10) + 160, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X1, Y1, X9,  Y9,  extract_abs(X1,X9) + 230 , 'dim', text_height=0.22, text_gap=0.01)      

            # 멍텅구리 우기둥 전개도 하부 치수선    
            dim_linear(doc,  X9, Y9, X8, Y8,  "", 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.01)      
            if RH1<RH2:
                dim_linear(doc,  X2, Y2, X5, Y5,  "",  100 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)       
            if RH1>RH2:    
                dim_linear(doc,  X8, Y8, X4, Y4, "",  100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.01)      
    
            if args.opt15:
                drawcircle(doc, X21 + 12.5, Y21 + 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X10 - 12.5, Y10 + 5, 2.5 , layer='레이져') # 도장홀 5파이

            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"

            Xpos = Abs_Xpos + 500 
            Ypos = R_Ypos - JB2_up / 2
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-우"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')           

            Xpos = Abs_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-우"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Mat.Spec : {normal_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Size : {Right_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')      
            
            ####################################################################################################################################################################################
            # 멍텅구리 우기둥 좌측 단면도    
            ####################################################################################################################################################################################
            Section_Xpos = Abs_Xpos - 500        
            Section_Ypos = R_Ypos

            SWAngle_radians = math.radians(SWAngle)

            if(A2>0):
                X1 = Section_Xpos
                Y1 = Section_Ypos - A2
                X2 = Section_Xpos 
                Y2 = Section_Ypos 

                line(doc, X1, Y1, X2, Y2, layer='0')                               
                dim_vertical_left(doc, X1, Y1, X2, Y2, 50, 'dim', text_height=0.22, text_gap=0.07)  
                X3 = X2 + C2
                Y3 = Y2
                lineto(doc,  X3, Y3, layer='0')   
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # vcut 표기                
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                dim_leader_line(doc, X3, Y3 , X3 + 50, Y3 + 100, "V-cut 2개소", layer="dim", style='0')    

            if(A2<1):
                X1 = Section_Xpos
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos             
                X3 = X2 + C2
                Y3 = Y2
                line(doc, X1, Y1, X3, Y3, layer='0')               
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # vcut 표기                            
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                dim_leader_line(doc, X3, Y3 , X3+50, Y3+100, "V-cut", layer="dim", style='0')                                        

            X4 = X3 + JB2_up * math.sin(Angle_radians)                
            Y4 = Y3 - JB2_up * math.cos(Angle_radians)        
            lineto(doc,  X4, Y4, layer='0')	
            dim_linear(doc, X3, Y3, X4, Y4,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)     

            if(K2_up > 5):
                X5 = X4
                Y5 = Y4 - K2_up
                lineto(doc,  X5, Y5, layer='0')	
                dim_vertical_right(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
            else:
                X5 = X4
                Y5 = Y4
                
            if(SWAngle>0):
                X6 = X5 - SW1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                X7 = X6 - SW2 * math.cos(SWAngle_radians)
                Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                lineto(doc,  X7, Y7, layer='0')
                dim_linear(doc, X6, Y6, X7, Y7,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                dim_angular(doc, X5-5, Y5, X6-5, Y6,  X7, Y7, X6 , Y6,  100, direction="left" )

            if(SWAngle<1): # 헤밍구조가 아닌경우 일반형
                X6 = X5 - SW
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                X7 = X6
                Y7 = Y6                        

            # if (Angle > 1):   # 직각이 아닌경우            
            dim_angular( doc,   X3, Y3, X4, Y4, X3, Y3, X2 , Y2,    15, direction="left" )		                            
            if(K2_up > 5): # K2값
                dim_angular(doc,  X4, Y4, X5 , Y5, X4, Y4, X3, Y3,  5, direction="left" )    

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB2_up != JB2_down ):
                rectangle(doc, X3+60, Y3 - JB2_up/2 - 75, X3 + 60 + 100, Y3 - JB2_up/2 + 75 , layer='0')               
                rectangle(doc, X3+60+300, Y3 - JB2_up/2 - 75 , X3 + 60 + 400, Y3 - JB2_up/2 + 75 , layer='0')               
            if(K2_up != K2_down ):
                rectangle(doc, X5 + 50, Y5 - 10 , X5 + 50 + 100, Y5 + K2_up + 10 , layer='0')                         
                rectangle(doc, X5 + 10 + 250, Y5 - 10 , X5 + 10 + 300, Y5 + K2_up + 10 , layer='0')                            


            ####################################################################################################################################################################################
            # 멍텅구리 우기둥 우측 단면도    
            #################################################################################################################################################################################### 

            if( JB2_up != JB2_down or K2_up != K2_down ):   

                Section_Xpos = Abs_Xpos + max(RH1,RH2) + 500
                Section_Ypos = R_Ypos

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(A2>0):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos - A2
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos 

                    line(doc, X1, Y1, X2, Y2, layer='0')
                    dim_vertical_right(doc, X2, Y2, X1, Y1,  50, 'dim', text_height=0.22, text_gap=0.07)
                    X3 = Section_Xpos
                    Y3 = Section_Ypos
                    lineto(doc,  X3, Y3, layer='0')
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # vcut 표기                
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut 2개소", layer="dim", style='0')     

                if(A2<1):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos                 
                    X3 = Section_Xpos
                    Y3 = Section_Ypos 
                    line(doc, X2, Y2, X3, Y3, layer='0')                               
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # vcut 표기                                
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut", layer="dim", style='0')                                

                X4 = X3 - JB2_down * math.sin(Angle_radians)                
                Y4 = Y3 - JB2_down * math.cos(Angle_radians)        
                lineto(doc,  X4, Y4, layer='0')	
                dim_linear(doc, X4, Y4, X3, Y3,   "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)

                if(K2_down > 5):
                    X5 = X4
                    Y5 = Y4 - K2_down
                    lineto(doc,  X5, Y5, layer='0')	
                    dim_vertical_left(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
                else:
                    X5 = X4
                    Y5 = Y4
                    lineto(doc,  X5, Y5, layer='0')	

                if(SWAngle>0):
                    X6 = X5 + SW1
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                    X7 = X6 + SW2 * math.cos(SWAngle_radians)
                    Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                    lineto(doc,  X7, Y7, layer='0')
                    dim_linear(doc, X7, Y7, X6, Y6,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                    dim_angular(doc,  X7 , Y7, X6, Y6, X6 + 5 , Y6, X6, Y6,  150, direction="right" )
                if(SWAngle<1):
                    X6 = X5 + SW
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,"" , 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    X7 = X6
                    Y7 = Y6

                # if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc,  X3, Y3, X2-8 , Y2,  X3, Y3, X4, Y4,   25, direction="right")		            
                    # if(K2_down > 7): # K2값
                dim_angular(doc,  X4, Y4, X3-3 , Y3, X5, Y5,  X4, Y4, 30, direction="right")

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB2_up != JB2_down ):
                    rectangle(doc, X3 - 60, Y3 - JB2_up/2 - 75, X3 - 60 - 100, Y3 - JB2_up/2 + 75 , layer='0')               
                    rectangle(doc, X3 - 60 - 350, Y3 - JB2_up/2 - 75 , X3 - 60 - 450, Y3 - JB2_up/2 + 75 , layer='0')               
                if(K2_up != K2_down ):
                    rectangle(doc, X5 - 50, Y5 - 10 , X5 - 50 - 100, Y5 + K2_up + 10 , layer='0')                         
                    rectangle(doc, X5 - 10 + 350, Y5 - 10 , X5 - 10 - 400, Y5 + K2_up + 10 , layer='0')                              
                                    

            Abs_Ypos -= 2000 

#############################
# SO와이드 작도 사이드오픈
#############################
def execute_sowide():     
    # 변경 코멘트 2024 03 04 ********** JD가 큰쪽의 기둥 뒷날개 해밍구조 없애기 가림판과 결합되는 부위
    # 시트 선택 (시트명을 지정)
    sheet_name = 'SO와이드제작'  # 원하는 시트명으로 변경
    sheet = workbook[sheet_name]

    # 2차원 배열을 저장할 리스트 생성
    data_2d = []

    # 1행부터 20행까지의 값을 2차원 배열에 저장
    for row_num in range(7, sheet.max_row + 1):  # 7행부터 20행까지 반복
        cell_value = sheet.cell(row=row_num, column=2).value  # 2열의 값 가져오기
        if cell_value is not None and cell_value != 'F':  # 값이 None이 아닌 경우에만 배열에 추가
            row_values =  [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, sheet.max_column + 1)]        
            data_2d.append(row_values)

    # 엑셀 파일 닫기
    workbook.close()

    if not data_2d:
        print(" SO 와이드 해당 데이터가 없습니다.")
        return 

    #################################################################### 
    # SO와이드  상판 전개도
    ####################################################################    
    Abs_Xpos = 32000
    Abs_Ypos = 0
    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        MH1 = row[2]
        MH2 = row[3]
        OP1 = row[4]
        OP2 = row[5]
        R = row[6]
        JD1 = row[7]
        JD2 = row[8]
        JB1_up = row[9]
        JB1_down = row[10]
        JB2_up = row[11]
        JB2_down = row[12]
        Top_gap = row[13]
        Side_gap = row[14]        
        H1 = row[15]
        H2 = row[16]
        C1 = row[17]        
        C2 = row[18]
        B1 = row[19]
        B2 = row[20]
        A1 = row[21]
        A2 = row[22]
        LH1 = row[23]
        LH2 = row[24]
        RH1 = row[25]
        RH2 = row[26]        
        U = row[27]        
        G = row[28]        
        UW = row[29]        
        SW_Left = row[30]        
        SW_Right = row[31]        
        K1_up = row[32]        
        K1_down = row[33]        
        K2_up = row[34]        
        K2_down = row[35]                
        SMH1 = row[36]                
        SMH2 = row[37]                
        G_check = row[38]
        Upper_size = row[39]
        Left_side_size = row[40]
        Right_side_size = row[41]
        E1 = row[42]
        E2 = row[43]
        HPI_height = row[44]
        SW1 = row[45]
        SW2 = row[46]
        SWAngle = row[47]
        A1_var = row[48]
        A2_var = row[49]
        Angle = row[50]
        Right_Angle = row[51]
        Kadapt = row[65]  # K값 적용여부 1 이면 K값 적용하지 말것

        if(Kadapt==1):
            K1_up = 0
            K2_up = 0                           

        if(K1_up>=10 and K2_up>=10 ):
            # 추영덕소장 E1, E2값 무시함 위의 조건이라면
            E1 = 0
            E2 = 0

        if(OP1>0):
            # 상판 기둥과 만나는 돌출값 B값과 연결되는 부분
            # Top_Pop1 = G-B1-1.5 
            # Top_Pop2 = G-B2-1.5 
            Top_Pop1 = G_check
            Top_Pop2 = G_check

            if(Top_Pop1<4):
                Top_Pop1 = 4
            if(Top_Pop2<4):
                Top_Pop2 = 4

            # 손상민소장의 경우 적용
            if(worker=="손상민" and U>2):
                draw_Uframe(Abs_Xpos, Abs_Ypos, U,G ,R, C1, C2, Floor_mc, Floor_des )  
            if(H1<1 or H2<1):
                U = 0
                G = 0                        

            SWAngle_radians = math.radians(SWAngle)
            if(SWAngle>0):
                SWAngle_var =  SW2 * math.cos(SWAngle_radians)        
            Angle_radians = math.radians(Angle)
            Right_Angle_radians = math.radians(Right_Angle)

            if(Angle>0):
                Angle_cos = math.cos(Angle_radians)        
                Right_Angle_cos = math.cos(Right_Angle_radians)        
                Angle_sin = math.sin(Angle_radians)    
                Right_Angle_sin = math.sin(Right_Angle_radians)    

            if(U > 0):
                X1 = Abs_Xpos - H1
                Y1 = Abs_Ypos + MH1 + G + U - Vcut_rate*5
                X2 = Abs_Xpos  + H2 + R 
                Y2 = Abs_Ypos + MH2 + G + U - Vcut_rate*5
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Abs_Ypos + MH2 + G  - Vcut_rate * 4
                lineto(doc, X3, Y3, layer='레이져')          
                X4 = X3
                Y4 = Abs_Ypos + MH2 + G - Top_Pop2 - Vcut_rate * 3
                lineto(doc, X4, Y4, layer='레이져')          
                X5 = X4 - H2 + (G - Top_Pop2) * Right_Angle_sin
                Y5 = Y4
                lineto(doc, X5, Y5, layer='레이져')          
                X6 = Abs_Xpos + R 
                Y6 = Abs_Ypos + MH2 - Vcut_rate * 2
                lineto(doc,  X6, Y6, layer='레이져')
                X7 = X6
                Y7 = Abs_Ypos
                lineto(doc,  X7, Y7, layer='레이져')     

            if(U < 1):
                if(G>0):
                    X1 = Abs_Xpos - H1
                    Y1 = Abs_Ypos + MH1 + G - Vcut_rate*3
                    X2 = Abs_Xpos  + H2 + R 
                    Y2 = Y1
                    line(doc, X1, Y1, X2, Y2, layer='레이져')       
                    X3 = X2
                    Y3 = Y2
                    lineto(doc, X3, Y3, layer='레이져')          
                    X4 = X3
                    Y4 = Abs_Ypos + MH2 + G - Top_Pop2 - Vcut_rate * 3
                    lineto(doc, X4, Y4, layer='레이져')          
                    X5 = X4 - H2 + (G - Top_Pop2) * Right_Angle_sin
                    Y5 = Y4
                    lineto(doc, X5, Y5, layer='레이져')          
                    X6 = Abs_Xpos + R 
                    Y6 = Abs_Ypos + MH2 - Vcut_rate * 2
                    lineto(doc,  X6, Y6, layer='레이져')
                    X7 = X6
                    Y7 = Abs_Ypos
                    lineto(doc,  X7, Y7, layer='레이져')   
                else:
                    X1 = Abs_Xpos 
                    Y1 = Abs_Ypos + MH1 - Vcut_rate
                    X2 = Abs_Xpos + R 
                    Y2 = Abs_Ypos + MH2 - Vcut_rate
                    line(doc, X1, Y1, X2, Y2, layer='레이져')       
                    X3 = X2
                    Y3 = Y2            
                    X4 = X3
                    Y4 = Y3            
                    X5 = X4
                    Y5 = Y4            
                    X6 = X5
                    Y6 = Y5            
                    X7 = X6
                    Y7 = Abs_Ypos
                    lineto(doc,  X7, Y7, layer='레이져')            

            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                X8 = Abs_Xpos + (R - OP1 - OP2 - 2)/2  + OP1 + OP2 + 2      
                if(Kadapt==1 and E2>0):
                    Y8 = Abs_Ypos - (JD2-E2) + Vcut_rate 
                else:
                    Y8 = Abs_Ypos - JD2 + Vcut_rate 
                lineto(doc,  X8, Y8, layer='레이져')
            else:
                X8 = Abs_Xpos + (R - OP1 - OP2 - 2)/2  + OP1 + OP2 + 2      
                Y8 = Abs_Ypos - JD2 + Vcut_rate * 2
                lineto(doc,  X8, Y8, layer='레이져')  
            
            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                X9 = X8
                if(E2>0):
                    Y9 = Abs_Ypos - JD2 + Vcut_rate*3
                else:
                    Y9 = Abs_Ypos - JD2 - K2_up + Vcut_rate*3
                lineto(doc,  X9, Y9, layer='레이져') 
                X10 = X9 
                if(E2>0):                
                    Y10 = Abs_Ypos - JD2 - UW + Vcut_rate*5     
                else:
                    Y10 = Abs_Ypos - JD2 - K2_up - UW + Vcut_rate*5     
                lineto(doc,  X10, Y10, layer='레이져')   
                X11 = Abs_Xpos + (R - OP1 - OP2 - 2)/2  + OP1 +  2 
                if(E2>0):                
                    Y11 = Abs_Ypos - JD2 - UW + Vcut_rate*5     
                else:                    
                    Y11 = Abs_Ypos - JD2 - K2_up - UW + Vcut_rate*5     
                lineto(doc,  X11, Y11, layer='레이져')      

                # SO 좌 열림, 우 열림에 따라 다른값을 갖는다. SO는 이것이 핵심임
                # 좌측이 클때
                if(JD1+K1_up > JD2+K2_up):   
                    X12 = X11 
                    if(E2>0):                
                        Y12 =  Abs_Ypos - JD2 + Vcut_rate*3
                    else:                    
                        Y12 =  Abs_Ypos - JD2 - K2_up + Vcut_rate*3
                    lineto(doc,  X12, Y12, layer='레이져')                                              
                    X13 = X12 
                    Y13 = Y12 
                    lineto(doc,  X13, Y13, layer='레이져')                                      
                else:
                    X12 = X11 
                    if(E2>0):                
                        Y12 =  Abs_Ypos - JD2 + Vcut_rate*3
                    else:                    
                        Y12 =  Abs_Ypos - JD2 - K2_up + Vcut_rate*3
                    lineto(doc,  X12, Y12, layer='레이져')                                              
                    X13 = X12 
                    Y13 = Y12 + ( (JD2+K2_up) - (JD1+K1_up) )
                    lineto(doc,  X13, Y13, layer='레이져')                                            
                X14 = X13 - 2                
                Y14 =  Y13                
                lineto(doc,  X14, Y14, layer='레이져')         
                # SO 좌 열림, 우 열림에 따라 다른값을 갖는다. SO는 이것이 핵심임
                # 좌측이 클때
                if(JD1+K1_up > JD2+K2_up):     
                    X15 = X14
                    if(E1>0):                
                        Y15 =  Abs_Ypos - JD1 + Vcut_rate*3
                    else:                    
                        Y15 =  Abs_Ypos - JD1 - K1_up + Vcut_rate*3
                    lineto(doc,  X15, Y15, layer='레이져')                                                      
                    X16 = X15                
                    if(E1>0):                
                        Y16 =  Abs_Ypos - JD1 - UW + Vcut_rate*5   
                    else:                    
                        Y16 =  Abs_Ypos - JD1 - K1_up - UW + Vcut_rate*5
                    lineto(doc,  X16, Y16, layer='레이져')                                      
                else:
                    X15 = X14                
                    Y15 = Y14
                    lineto(doc,  X15, Y15, layer='레이져')                                                      
                    X16 = X15                
                    if(E1>0):                
                        Y16 =  Abs_Ypos - JD1 - UW + Vcut_rate*5   
                    else:                    
                        Y16 =  Abs_Ypos - JD1 - K1_up - UW + Vcut_rate*5
                    lineto(doc,  X16, Y16, layer='레이져')                                      

                X17 = Abs_Xpos + (R - OP1 - OP2 - 2)/2 
                Y17 = Y16 
                lineto(doc,  X17, Y17, layer='레이져')                                      
                X18 = X17
                if(E1>0):                
                    Y18 =  Abs_Ypos - JD1 + Vcut_rate*3
                else:                    
                    Y18 =  Abs_Ypos - JD1 - K1_up + Vcut_rate*3
                lineto(doc,  X18, Y18, layer='레이져')                                                      
                X19 = Abs_Xpos + (R - OP1 - OP2 - 2)/2 
                Y19 = Abs_Ypos - JD1 + Vcut_rate 
                lineto(doc,  X19, Y19, layer='레이져')     

            else:
                # E1 적용 공식인데, K1, K2에 대한 내용 삭제된 것임
                X9 = X8
                Y9 = Abs_Ypos - JD2 + Vcut_rate*3                
                lineto(doc,  X9, Y9, layer='레이져') 
                X10 = X9                 
                Y10 = Abs_Ypos - JD2 - UW + Vcut_rate*5
                lineto(doc,  X10, Y10, layer='레이져')   
                X11 = Abs_Xpos + (R - OP1 - OP2 - 2)/2  + OP1 +  2 
                Y11 = Abs_Ypos - JD2 - UW + Vcut_rate*5                     
                lineto(doc,  X11, Y11, layer='레이져')      
                # SO 좌 열림, 우 열림에 따라 다른값을 갖는다. SO는 이것이 핵심임
                # 좌측이 클때
                if(JD1+K1_up > JD2+K2_up):   
                    X12 = X11 
                    Y12 =  Abs_Ypos - JD2 + Vcut_rate*3                    
                    lineto(doc,  X12, Y12, layer='레이져')                                              
                    X13 = X12 
                    Y13 = Y12 
                    lineto(doc,  X13, Y13, layer='레이져')                                      
                else:
                    X12 = X11 
                    Y12 =  Abs_Ypos - JD2 + Vcut_rate*3                    
                    lineto(doc,  X12, Y12, layer='레이져')                                              
                    X13 = X12 
                    Y13 = Y12 + ( (JD2+K2_up) - (JD1+K1_up) )
                    lineto(doc,  X13, Y13, layer='레이져')                               
                X14 = X13 - 2                
                Y14 =  Y13                
                lineto(doc,  X14, Y14, layer='레이져')    

                # SO 좌 열림, 우 열림에 따라 다른값을 갖는다. SO는 이것이 핵심임
                # 좌측이 클때
                if(JD1+K1_up > JD2+K2_up):     
                    X15 = X14
                    Y15 =  Abs_Ypos - JD1 + Vcut_rate*3                    
                    lineto(doc,  X15, Y15, layer='레이져')                                                      
                    X16 = X15                
                    Y16 =  Abs_Ypos - JD1 - UW + Vcut_rate*5                       
                    lineto(doc,  X16, Y16, layer='레이져')                                      
                else:
                    X15 = X14                
                    Y15 = Y14
                    lineto(doc,  X15, Y15, layer='레이져')                                                      
                    X16 = X15                
                    Y16 =  Abs_Ypos - JD1 - UW + Vcut_rate*5   
                    lineto(doc,  X16, Y16, layer='레이져')    

                X17 = Abs_Xpos + (R - OP1 - OP2 - 2)/2 
                Y17 = Y16 
                lineto(doc,  X17, Y17, layer='레이져')                                      
                X18 = X17
                Y18 =  Abs_Ypos - JD1 + Vcut_rate*3
                lineto(doc,  X18, Y18, layer='레이져')                                                      
                X19 = Abs_Xpos + (R - OP1 - OP2 - 2)/2 
                Y19 = Abs_Ypos - JD1 + Vcut_rate*3 
                lineto(doc,  X19, Y19, layer='레이져')           
            
            # 기준선 MH1과 JD1 만나는 점
            X20 = Abs_Xpos
            Y20 = Abs_Ypos
            lineto(doc,  X20, Y20, layer='레이져')      
    
            if(U > 0):         
                X21 = Abs_Xpos
                Y21 = Abs_Ypos + MH1 - Vcut_rate * 2
                lineto(doc,  X21, Y21, layer='레이져')                
                X22 = Abs_Xpos - (G-Top_Pop1) * Angle_sin
                Y22 = Abs_Ypos + MH1 + G - Top_Pop1 - Vcut_rate * 3
                lineto(doc,  X22, Y22, layer='레이져')      
                X23 = Abs_Xpos - H1
                Y23 = Y22
                lineto(doc,  X23, Y23, layer='레이져')      
                X24 = X23
                Y24 = Abs_Ypos + MH1 + G  - Vcut_rate * 4
                lineto(doc,  X24, Y24, layer='레이져')      
                lineto(doc,  X1, Y1, layer='레이져')       

            if(U < 1):   
                if( G < 1):
                    X21 = Abs_Xpos
                    Y21 = Abs_Ypos + MH1 - Vcut_rate 
                    lineto(doc,  X1, Y1, layer='레이져')       
                    X24 = X23 = X22 = X21 
                    Y24 = Y23 = Y22 = Y21 
                else:
                    X21 = Abs_Xpos
                    Y21 = Abs_Ypos + MH1 - Vcut_rate * 2
                    lineto(doc,  X21, Y21, layer='레이져')                
                    X22 = Abs_Xpos - (G-Top_Pop1) * Angle_sin
                    Y22 = Abs_Ypos + MH1 + G - Top_Pop1 - Vcut_rate * 3
                    lineto(doc,  X22, Y22, layer='레이져')      
                    X23 = Abs_Xpos - H1
                    Y23 = Y22
                    lineto(doc,  X23, Y23, layer='레이져')      
                    X24 = X23
                    Y24 = Abs_Ypos + MH1 + G  - Vcut_rate * 3
                    lineto(doc,  X24, Y24, layer='레이져')     
                    lineto(doc,  X1, Y1, layer='레이져')                      

        if args.opt9:
            drawcircle(doc, X17 + 12.5, Y17 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이
            drawcircle(doc, X10 - 12.5, Y10 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이

        # 절곡선
        if(U>0):
            line(doc, X24, Y24, X3, Y3,  layer='22')        
        if(G>0):            
            # 실선처리
            line(doc, X21, Y21, X6, Y6,  layer='HIDDEN2')

        line(doc, X20, Y20, X7, Y7,  layer='22')

        # 좌측이 클때
        if(JD1+K1_up > JD2+K2_up):     
            line(doc, X18, Y18, X15, Y15,  layer='22')
            line(doc, X13, Y13, X9, Y9,  layer='22')
            # 도어두께 표시하는 치수선
            dim_vertical_right(doc, X16, Y16, X11, Y11, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_left(doc, X14, Y14, X15, Y15, 100, 'dim', text_height=0.22,  text_gap=0.07)                  
        else:
            line(doc, X18, Y18, X15, Y15,  layer='22')
            line(doc, X12, Y12, X9, Y9,  layer='22')            
            # 도어두께 표시하는 치수선
            dim_vertical_left(doc, X16, Y16, X11, Y11, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X13, Y13, X12, Y12, 100, 'dim', text_height=0.22,  text_gap=0.07)                  

        # HPI 그리기
        if(HPI_height>0):
            draw_HPI(Abs_Xpos + (X7-X20)/2, Y20, R, HPI_height )     
        else:
            draw_NoHPI(Abs_Xpos + (X7-X20)/2, Y20 )                   

        #  전개도 상부 치수선
        dim_linear(doc,  X1, Y1, X2, Y2, "", 160,  direction="up", layer='dim')

        # H1, OP, H2 치수선
        if(G > 0):   
            dim_linear(doc,  X1, Y1, X21, Y21,  "", 80,  direction="up", layer='dim')
            dim_linear(doc,  X21, Y21, X6, Y6, "",  extract_abs(Y5,Y1) + 80 + 17.4,  direction="up", layer='dim')
            dim_linear(doc,  X6, Y6, X2, Y2,  "",  extract_abs(Y1,Y5) + 80 + 17.4,  direction="up", layer='dim')

        # 우측 절곡치수선
        if(U>0):        
            dim_vertical_right(doc, X2, Y2, X3, Y3, 160 , 'dim', text_height=0.22, text_gap=0.07)  
        if(G>0):                    
            dim_vertical_right(doc, X6,  Y6, X3, Y3,  extract_abs(X2,X6) + 100 , 'dim', text_height=0.22, text_gap=0.07)  

        dim_vertical_right(doc, X6, Y6,  X7, Y7,  extract_abs(X2,X6) +  100, 'dim', text_height=0.22, text_gap=0.07)  
        dim_vertical_right(doc, X9, Y9,  X7, Y7,  extract_abs(X2,X6) + extract_abs(X9,X6) + 100, 'dim', text_height=0.22, text_gap=0.07)   
        dim_vertical_right(doc, X9, Y9, X10, Y10, extract_abs(X2,X6) + extract_abs(X9,X6) + 100, 'dim', text_height=0.22, text_gap=0.07)   # 뒷날개
        dim_vertical_right(doc, X2, Y2,  X10, Y10, extract_abs(X2,X6) + extract_abs(X9,X6) + 170, 'dim', text_height=0.22, text_gap=0.07)             

        # 좌측 절곡치수선
        if(JD1 != JD2 or K1_up != K2_up or MH1 != MH2 or E1 != E2):
            if(U>0):                    
                dim_vertical_left(doc, X1, Y1, X24, Y24,   120, 'dim', text_height=0.22, text_gap=0.07)  
            if(G>0):                        
                dim_vertical_left(doc, X21, Y21, X24, Y24,   extract_abs(X20,X24) + 70 , 'dim', text_height=0.22, text_gap=0.07)     
            # MH                             
            dim_vertical_left(doc, X20, Y20, X21, Y21,  extract_abs(X20,X24) + 70 , 'dim', text_height=0.22, text_gap=0.07)      
            # JD      
            dim_vertical_left(doc, X20, Y20, X18, Y18,   extract_abs(X20,X24) + 70 , 'dim', text_height=0.22, text_gap=0.07)           
            # 뒷날개
            dim_vertical_left(doc, X17, Y17, X18, Y18,  extract_abs(X20,X24) +  extract_abs(X15,X11) + 70, 'dim', text_height=0.22, text_gap=0.07)                          
            # 좌측 전체
            dim_vertical_left(doc, X17, Y17, X1, Y1,   extract_abs(X1,X17) + 220, 'dim', text_height=0.22, text_gap=0.07)                          
        else:    
            dim_vertical_left(doc, X17, Y17, X18, Y13, extract_abs(X17,X20) + 50, 'dim', text_height=0.22,  text_gap=0.07) 
            dim_vertical_left(doc, X17, Y17, X15, Y15, extract_abs(X17,X20) + 120, 'dim', text_height=0.22,  text_gap=0.07) 
            dim_vertical_left(doc, X17, Y17, X21, Y21, extract_abs(X17,X24) + 180 , 'dim', text_height=0.22,  text_gap=0.07) 
            dim_vertical_left(doc, X17, Y17, X24, Y24, extract_abs(X17,X24) + 240 , 'dim', text_height=0.22,  text_gap=0.07) 

        # 쫄대타입 상부 Vcut 치수 표기
        if(U>0 and G>0):   
            dim_vertical_left(doc, X1+H1+(OP1+OP2)/2, Y1, X1+H1+(OP1+OP2)/2, Y6, 100, 'dim', text_height=0.22,  text_gap=0.07)  

        # SO는 치수선이 좌측 열림, 우측열림에 따라 다르게 해야 한다.
        if(JD1+K1_up > JD2+K2_up):       
            # 하부 치수선 4개 노칭 1개 위방향
            dim_linear(doc,  X20, Y20, X17, Y17,  "",  extract_abs(Y17, Y20) + 160 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X17, Y17, X16, Y16,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X11, Y11, X10, Y10,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up) + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X13, Y13, X14, Y14,  "",  50  ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X17, Y17, X10, Y10,  "",  160 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)              
        else:
            # 하부 치수선 4개 노칭 1개 위방향
            dim_linear(doc,  X20, Y20, X17, Y17,  "",  extract_abs(Y17, Y20) + 180 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X17, Y17, X16, Y16,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up)  + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X11, Y11, X10, Y10,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X13, Y13, X14, Y14,  "",  120  ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X17, Y17, X10, Y10,  "",  180 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)              
            

        if(JD1+K1_up > JD2+K2_up) :
            dim_linear(doc,  X10, Y10, X7, Y7,    "",  160 + extract_abs(JD1+K1_up, JD2+K2_up) + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
        else:
            dim_linear(doc,  X10, Y10, X7, Y7,    "",  180 + extract_abs(JD1+K1_up, JD2+K2_up) - extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
        
        # 상판 돌출부위 H1, H2 치수선
        if(G > 0):  
            dim_vertical_right(doc, X24, Y24, X22, Y22, 100, 'dim', text_height=0.22,  text_gap=0.07)
            dim_vertical_left(doc, X3, Y3, X5, Y5, 100, 'dim', text_height=0.22,  text_gap=0.07)

        # 기둥과의 접선 부위 aligned 표현 각도표현
        # 좌측
        dim_linear(doc, X20, Y20, X18, Y18,   "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) 
        dim_angular(doc, X19, Y19, X20, Y20, X18, Y18, X19, Y19 + 10,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다
        # 우측        
        dim_linear(doc, X9, Y9, X7, Y7,  "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) 
        dim_angular(doc, X9, Y9  , X8, Y8 + 10,  X8, Y8,  X7, Y7,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다

        # 와이드 상판 K1 K2 값 부분 치수선 
        if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
            dim_vertical_right(doc, X18, Y18, X19, Y19, 80, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_left(doc, X9, Y9, X8, Y8, 80, 'dim', text_height=0.22,  text_gap=0.07)  


        # 문구 설정
        if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
            Pre_text = ""
        else:
            Pre_text = str(Floor_mc) + "-"
        # Textstr 생성
        Textstr = f"({Pre_text}{str(Floor_des)})"       

        X2 = Abs_Xpos + (H1 + R + H2)/2 - len(Textstr)*50/3
        Y2 = Abs_Ypos - JD1 / 1.8
        draw_Text(doc, X2, Y2, 50, str(Textstr), '레이져')

        # 문구            
        sizedes = Upper_size                  
        
        Textstr = f"Part Name : 상판({Pre_text}{str(Floor_des)})"    
        draw_Text(doc, X10+500, Y10 + 50 - 150, 28, str(Textstr), '0')
        Textstr = f"Mat.Spec : {widejamb_material}"    
        draw_Text(doc, X10+500, Y10 - 150, 28, str(Textstr), '0')
        Textstr = f"Size : {sizedes}"    
        draw_Text(doc, X10+500, Y10 - 50 - 150 , 28, str(Textstr), '0')
        Textstr = f"Quantity : 1 EA"    
        draw_Text(doc, X10+500, Y10 - 100 - 150 , 28, str(Textstr), '0')            

        ##################################################################
        # SO와이드 상판 좌측 단면도    (좌우가 다른 모양일때 단면도를 그려준다)
        #################################################################
        if(JD1 != JD2 or K1_up != K2_up or MH1 != MH2 or E1 != E2):
            Section_Xpos = Abs_Xpos - H1 - 400 - G           

            if(G > 0):
                X1 = Section_Xpos + G 
                Y1 = Abs_Ypos + MH1 + U
                X2 = X1
                Y2 = Y1 - U  
                X3 = X2 - G
                Y3 = Y2
                if(U>0):
                    line(doc, X1, Y1, X2, Y2, layer='0')                                    
                    dim_vertical_right(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)                         
                    lineto(doc,  X3, Y3, layer='0')    
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                else:
                    line(doc, X2, Y2, X3, Y3, layer='0')  
                dim_linear(doc, X3, Y3, X1, Y1, "", 100,  direction="up", layer='dim')                                
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                
                X4 = X3 
                Y4 = Y3 - MH1
                lineto(doc,  X4, Y4, layer='0')                                
            if(G < 1):                       
                X1 = Section_Xpos 
                Y1 = Abs_Ypos + MH1 
                X2 = X1
                Y2 = Y1
                X3 = X2
                Y3 = Y2                
                X4 = X3 
                Y4 = Y3 - MH1
                line(doc, X3, Y3, X4, Y4, layer='0')                      

            if(G > 1):
                drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
                if(U>0):                    
                    dim_leader_line(doc, X4, Y4 , X4 - 200 , Y4+100, "V-cut 3개", layer="dim", style='0')      
                else:            
                    dim_leader_line(doc, X4, Y4 , X4 - 200 , Y4+100, "V-cut 2개", layer="dim", style='0')                              
            if(G < 1):
                drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
                dim_leader_line(doc, X4, Y4 , X4 - 100 , Y4+100, "V-cut 1개", layer="dim", style='0')                              

            if(K1_up > 0):
                X5 = X4 - JD1
                Y5 = Y4 
                lineto(doc,  X5, Y5, layer='0')               
                X6 = X5 - K1_up
                Y6 = Y5 
                lineto(doc,  X6, Y6, layer='0')   
                dim_linear(doc,  X5, Y5, X4, Y4,  "", 80, direction="down", layer='dim')     
                dim_linear(doc,  X5, Y5, X6, Y6,  "", 80, direction="down", layer='dim')     
                dim_linear(doc,  X6, Y6, X4, Y4,  "", 150, direction="down", layer='dim')                  
            else:
                X5 = X4 - JD1
                Y5 = Y4 
                lineto(doc,  X5, Y5, layer='0')               
                X6 = X5 
                Y6 = Y5                                 
                dim_linear(doc,  X6, Y6, X4, Y4,  "", 150 , direction="down", layer='dim')     

            # drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
            X7 = X6 
            Y7 = Y6 + UW
            lineto(doc,  X7, Y7, layer='0')     
            dim_vertical_left(doc,  X6, Y6, X7, Y7, 50,'dim' ,text_height=0.22,  text_gap=0.07)            

            if(K1_up > 0):                 
                # K값 부분 라인 그리기
                line(doc, X5, Y5, X5 , Y5 - UW , layer='22')               

            dim_vertical_left(doc,  X6, Y6, X3, Y3,  110,'dim' ,text_height=0.22,  text_gap=0.07)    
            
            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JD1 != JD2 or K1_up != K2_up ):
                rectangle(doc, X6+(JD1+K1_up)/2 -55, Y6-90, X6+(JD1+K1_up)/2 + 55, Y6-190, layer='0')               
                rectangle(doc, X4+250, Y6-40, X4+350, Y6 - JD1 + 40, layer='0')               
            if( MH1 != MH2 ):
                rectangle(doc, X6-70, Y6 + MH1/2 + 90 , X6 - 70 - 100 , Y6 + MH1/2 - 90, layer='0')               
                rectangle(doc, X4+230, Y6 + MH1/2 + 90  , X4+330,  Y6 + MH1/2 - 90 , layer='0')               
        
        ###########################################################################
        # 와이드 상판 우측 단면도 기본적으로 다 그림
        ###########################################################################
        Section_Xpos = Abs_Xpos + R + H2 + 450 + G           

        if(G > 0):
            X1 = Section_Xpos - G 
            Y1 = Abs_Ypos + MH2 + U
            X2 = X1
            Y2 = Y1 - U                          
            X3 = X2 + G
            Y3 = Y2
            if(U>0):
                line(doc, X1, Y1, X2, Y2, layer='0')                                    
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                dim_vertical_left(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)                         
                lineto(doc,  X3, Y3, layer='0')    
            else:
                line(doc, X2, Y2, X3, Y3, layer='0')  

            lineto(doc,  X3, Y3, layer='0')    
            dim_linear(doc, X1, Y1, X3, Y3,  "", 100,  direction="up", layer='dim')                            
            drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                
            X4 = X3 
            Y4 = Y3 - MH2
            lineto(doc,  X4, Y4, layer='0')              

        if(G < 1):                       
            X1 = Section_Xpos 
            Y1 = Abs_Ypos + MH2
            X2 = X1
            Y2 = Y1
            X3 = X2
            Y3 = Y2                
            X4 = X3 
            Y4 = Y3 - MH2
            line(doc, X3, Y3, X4, Y4, layer='0')                      

        # if(G > 1):
        #     drawcircle(doc, X4, Y4 , 5 , layer='0', color='6')             
        #     if(U>0):                
        #         dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 3개", layer="dim", style='0')      
        #     else:
        #         dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 2개", layer="dim", style='0')      
        # if(G < 1):
        #     drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
        #     dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 1개", layer="dim", style='0')      

        if(K2_up > 0):
            X5 = X4 + JD2
            Y5 = Y4 
            lineto(doc,  X5, Y5, layer='0')               
            X6 = X5 + K2_up
            Y6 = Y5 
            lineto(doc,  X6, Y6, layer='0')   
            dim_linear(doc,  X4, Y4, X5, Y5, "", 80, direction="down", layer='dim')     
            dim_linear(doc,  X5, Y5, X6, Y6, "", 80, direction="down", layer='dim')     
            dim_linear(doc,  X4, Y4, X6, Y6, "", 150, direction="down", layer='dim')                  
        else:
            X5 = X4 + JD2
            Y5 = Y4 
            lineto(doc,  X5, Y5, layer='0')               
            X6 = X5 
            Y6 = Y5                                 
            dim_linear(doc,  X4, Y4, X6, Y6, "", 150 , direction="down", layer='dim')     

        # drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
        X7 = X6 
        Y7 = Y6 + UW
        lineto(doc,  X7, Y7, layer='0')     
        dim_vertical_right(doc,  X6, Y6, X7, Y7, 50,'dim' ,text_height=0.22,  text_gap=0.07)            

        if(K2_up > 0):                 
            # K값 부분 라인 그리기
            line(doc, X5, Y5, X5 , Y5 - UW , layer='22')        

        dim_vertical_right(doc,   X6, Y6, X3, Y3,   110,'dim' ,text_height=0.22,  text_gap=0.07)                           

        # 서로 치수 차이나는 곳 사각형으로 그려주기
        if(JD1 != JD2 or K1_up != K2_up ):
            rectangle(doc, X6-(JD1+K1_up)/2 + 55, Y6-90, X6-(JD1+K1_up)/2 - 55, Y6-190, layer='0')                               
            rectangle(doc, X4 - 320, Y6 - 40, X4 - 420, Y6 - JD1 + 40, layer='0')        
        if( MH1 != MH2 ):
            rectangle(doc, X6 + 60, Y6 + MH2/2 + 90 , X6 + 60 + 100 , Y6 + MH2/2 - 90, layer='0')               
            rectangle(doc, X4 - 330, Y6 + MH2/2 + 90  , X4 - 430,  Y6 + MH2/2 - 90 , layer='0')                           

        # 상판 도면틀 넣기
        BasicXscale = 2560
        BasicYscale = 1550
        TargetXscale = 1000 + R + (JD1 + K1_up)*2 + 200
        TargetYscale = 500 + JD1 + MH1 + G + U
        if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
            frame_scale = TargetXscale/BasicXscale
        else:
            frame_scale = TargetYscale/BasicYscale
        # print (f'스케일 : {frame_scale}')
        insert_frame(X1 - ( R + (JD1 + K1_up)*2 + 900 ),Y1-(850 + JD1 + MH1 + G + U) , frame_scale, "top", workplace)

        # 다음 도면 간격 계산
        Abs_Ypos -= 4000

    ###################################################################################################################################################################
    # SO와이드 좌기둥 전개도
    ###################################################################################################################################################################
    Abs_Xpos = 32000 + 5000
    Abs_Ypos = 1000    
    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        MH1 = row[2]
        MH2 = row[3]
        OP1 = row[4]
        OP2 = row[5]
        R = row[6]
        JD1 = row[7]
        JD2 = row[8]
        JB1_up = row[9]
        JB1_down = row[10]
        JB2_up = row[11]
        JB2_down = row[12]
        Top_gap = row[13]
        Side_gap = row[14]        
        H1 = row[15]
        H2 = row[16]
        C1 = row[17]        
        C2 = row[18]
        B1 = row[19]
        B2 = row[20]
        A1 = row[21]
        A2 = row[22]
        LH1 = row[23]
        LH2 = row[24]
        RH1 = row[25]
        RH2 = row[26]        
        U = row[27]        
        G = row[28]        
        UW = row[29]        
        SW_Left = row[30]        
        SW_Right = row[31]         
        K1_up = row[32]        
        K1_down = row[33]        
        K2_up = row[34]        
        K2_down = row[35]                
        SMH1 = row[36]                
        SMH2 = row[37]                
        G_check = row[38]
        Upper_size = row[39]
        Left_side_size = row[40]
        Right_side_size = row[41]
        E1 = row[42]
        E2 = row[43]
        HPI_height = row[44]
        Original_SW1 = row[45] # 변수명 변경 사이드오픈은 이렇게 적용해야 함
        Original_SW2 = row[46] # 변수명 변경 사이드오픈은 이렇게 적용해야 함
        Original_SWAngle = row[47] # 변수명 변경 사이드오픈은 이렇게 적용해야 함
        A1_var = row[48]
        A2_var = row[49]
        Angle = row[50]
        Right_Angle = row[51]
        Kadapt = row[65]  # K값 적용여부 1 이면 K값 적용하지 말것    

        # 데이터 검증 및 기본값 설정
        SWAngle = validate_or_default(SWAngle)
        K1_up = validate_or_default(K1_up)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_down = validate_or_default(K2_down)
        OP1 = validate_or_default(OP1)
        Kadapt = validate_or_default(Kadapt)

        # 변경 코멘트 2024 03 04 ********** JD가 큰쪽의 기둥 뒷날개 해밍구조 없애기 가림판과 결합되는 부위
        # JD1이 크면 가림판과 결합되는 부위로 SWAngle에 0 적용

        # 우측기둥에서 다시 판단하게 만들려고 Original_SWAngle 변수 만듬
        SW1 = Original_SW1
        SW2 = Original_SW2        
        SWAngle = Original_SWAngle
        if(JD1 + K1_up > JD2 + K2_up):
            SWAngle =  0
            SW1 = 0
            SW2 = 0

        if(Kadapt==1):
            K1_up = 0            
            K1_down = 0            
            K2_up = 0            
            K2_down = 0            

        if(OP1>0):            
            SWAngle_radians = math.radians(SWAngle)
            if(SWAngle>0):
                SWAngle_var =  SW2 * math.cos(SWAngle_radians)        
            Angle_radians = math.radians(Angle)
            if(Angle>0):
                Angle_cos = math.cos(Angle_radians)        
                Angle_sin = math.sin(Angle_radians)    

            if (SWAngle>1):  #SWAngle 0 이상인 경우            
                if(K1_up>0):      
                    if (SWAngle == 90):      
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 3 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3                
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4           
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')
                    else:
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 2 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3                
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4           
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                else:
                    if (SWAngle == 90):                            
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 3 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3          
                        Y4 =  Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate   
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4              
                    else:
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 2 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3          
                        Y4 =  Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate   
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4              

            if (SWAngle<1):  #SWAngle 0 이상인 경우
                if(K1_up>0):            
                    X1 = Abs_Xpos + SMH1
                    Y1 = Abs_Ypos + JB1_up + K1_up + SW_Left - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + SMH1 + LH2 
                    Y2 = Abs_Ypos + JB1_down + K1_down + SW_Left - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                    X3 = X2            
                    Y3 = Y2
                    X4 = X3
                    Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = X4               
                    Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                    lineto(doc,  X5, Y5, layer='레이져')                    
                else:
                    X1 = Abs_Xpos + SMH1
                    Y1 = Abs_Ypos + JB1_up + SW_Left - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + SMH1 + LH2 
                    Y2 = Abs_Ypos + JB1_down + SW_Left - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                    X3 = X2            
                    Y3 = Y2
                    X4 = X3             
                    Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = X4             
                    Y5 = Y4                

            X6 = Abs_Xpos +  SMH1 + LH1             
            Y6 = Abs_Ypos
            lineto(doc,  X6, Y6, layer='레이져')
            if(A1>0):     
                X7 = X6 
                Y7 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X7, Y7, layer='레이져')                  
                X8 = X7
                Y8 = Abs_Ypos - C1 - A1 + Vcut_rate * 3
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = Abs_Xpos             
                Y10 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X10, Y10, layer='레이져')

            if(A1<1):       
                X7 = X6 
                Y7 = Abs_Ypos - C1 + Vcut_rate 
                lineto(doc,  X7, Y7, layer='레이져')                
                X8 = Abs_Xpos + SMH1 + LH1  
                Y8 = Y7
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = X9
                Y10 = Y9            

            X11 = Abs_Xpos 
            Y11 = Abs_Ypos
            lineto(doc,  X11, Y11, layer='레이져')        

            if (SWAngle>0): 
                if(K1_up>0):  
                    if (SWAngle == 90):
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1  
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Abs_Ypos + JB1_up + K1_up - Bending_rate   - Vcut_rate   
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 3 - Vcut_rate
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1  
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                else:  
                    if (SWAngle == 90):                      
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1     
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Y14
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + SW1  - Bending_rate * 3 - Vcut_rate  
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                      
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1     
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Y14
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                      
                    
            if (SWAngle<1):  # 헤밍구조 아닌경우
                if(K1_up>0):  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + B1
                    lineto(doc,  X12, Y12, layer='레이져')                            
                    X13 = Abs_Xpos + SMH1
                    Y13 = Y12
                    lineto(doc,  X13, Y13, layer='레이져')                
                    X14 = X13
                    Y14 = Abs_Ypos +  JB1_up  - Vcut_rate
                    lineto(doc,  X14, Y14, layer='레이져')
                    X15 = X14
                    Y15 = Abs_Ypos + JB1_up + K1_up  - Bending_rate  - Vcut_rate  
                    lineto(doc,  X15, Y15, layer='레이져')
                    X16 = X15 
                    Y16 = Y15
                    lineto(doc,  X1, Y1, layer='레이져')
                else:  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + B1
                    lineto(doc,  X12, Y12, layer='레이져')                            
                    X13 = Abs_Xpos + SMH1
                    Y13 = Y12
                    lineto(doc,  X13, Y13, layer='레이져')                
                    X14 = X13
                    Y14 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate  
                    lineto(doc,  X14, Y14, layer='레이져')
                    X15 = X14
                    Y15 = Y14
                    lineto(doc,  X15, Y15, layer='레이져')
                    X16 = X15 
                    Y16 = Y15
                    lineto(doc,  X1, Y1, layer='레이져')      

            # 좌기둥 절곡선
            if(SW1>0):
                line(doc, X16, Y16, X3, Y3,  layer='22')
            line(doc, X15, Y15, X4, Y4,  layer='22')
            if(K1_up>0):
                line(doc, X14, Y14, X5, Y5,  layer='22')
            # JB    
            line(doc, X11, Y11, X6, Y6,  layer='22')
            if(A1>0):
                line(doc, X10, Y10, X7, Y7,  layer='22')                

            # 좌기둥 상부 전개도 치수선        
            dim_linear(doc,  X12, Y12, X1, Y1, "", extract_abs(Y12,Y1) + 100,  direction="up", layer='dim')                
            dim_linear(doc,  X1, Y1, X2, Y2, "", extract_abs(Y1,Y2) + 100,  direction="up", layer='dim')                
            dim_linear(doc,  X12, Y12, X2, Y2, "", extract_abs(Y12,Y2) + 150,  direction="up", layer='dim')                        

            # 우측 절곡치수선
            if(SW1>0):
                dim_vertical_right(doc, X2,  Y2, X3, Y3,  extract_abs(X2,X7) + 180, 'dim', text_height=0.22, text_gap=0.01)      
            dim_vertical_right(doc, X3, Y3, X4,  Y4,   extract_abs(X4,X7) +  80, 'dim', text_height=0.22, text_gap=0.01)      
            if(K1_up>0):
                dim_vertical_right(doc, X4,  Y4, X5, Y5,  extract_abs(X4,X7) + 130, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_right(doc, X6, Y6, X5,  Y5, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_right(doc,  X7, Y7, X6,  Y6, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_right(doc, X7,  Y7, X8, Y8, 120, 'dim', text_height=0.22, text_gap=0.01)  
                #Vcut A값 C값 전개치수선     
                dim_vertical_left(doc, X6, Y6, X8,  Y8,   extract_abs(X4,X7) + 80 , 'dim', text_height=0.22, text_gap=0.01)
            dim_vertical_right(doc, X2,  Y2, X8, Y8,   extract_abs(X2,X8) + 270 , 'dim', text_height=0.22, text_gap=0.01)      

            # 좌측 절곡치수선  
            if(SW1>0):                       
                dim_vertical_left(doc, X1, Y1, X16, Y16, SMH1 + 180, 'dim', text_height=0.22, text_gap=0.01)  
                dim_vertical_left(doc, X15, Y15, X16, Y16,  SMH1 + 80, 'dim', text_height=0.22, text_gap=0.01)  
            else:
                dim_vertical_left(doc, X1, Y1, X15, Y15, SMH1 + 80, 'dim', text_height=0.22, text_gap=0.01)                  
            if(K1_up>0):
                dim_vertical_left(doc, X15, Y15, X14, Y14, SMH1 + 130, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X11, Y11, X14, Y14, 80, 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_left(doc, X11, Y11, X10, Y10,  80, 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_left(doc, X9, Y9, X10, Y10, 160, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X1, Y1, X9,  Y9,  SMH1 + 270 , 'dim', text_height=0.22, text_gap=0.01) 

            if(B1>0):
                #B1 치수선     
                dim_vertical_left(doc, X13, Y13, X13,  Y11,   70 , 'dim', text_height=0.22, text_gap=0.01)      
                # JD 접선치수선
                dim_vertical_right(doc, X13, Y13, X15,  Y15,   50 , 'dim', text_height=0.22, text_gap=0.01)      

            # 기둥 하부 치수선 2    
            dim_linear(doc,   X13, Y13, X8, Y8, "", B1+ C1 + A1 + 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   
            dim_linear(doc,   X9, Y9, X8, Y8, "", 150,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   

            if LH1>LH2:
                dim_linear(doc,  X2, Y2, X6, Y6, "",  150 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)                   
            if LH1<LH2:    
                dim_linear(doc,  X4, Y4, X8, Y8, "", extract_abs(Y4, Y8) + 150 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            if args.opt9:
                drawcircle(doc, X1 + 12.5, Y1 - 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X2 - 12.5, Y2 - 5, 2.5 , layer='레이져') # 도장홀 5파이        

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"

            Xpos = Abs_Xpos + SMH1 + 400
            Ypos = Abs_Ypos + JB1_up / 2.3
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-좌"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')           

            Xpos = Abs_Xpos + SMH1 + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-좌"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + SMH1 + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Mat.Spec : {widejamb_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + SMH1 + LH1/1.2        
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Size : {Left_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + SMH1 + LH1/1.2        
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')            

            ########################################################################################################################################################################################
            # SO와이드 좌기둥 좌측 단면도    
            ########################################################################################################################################################################################
            Section_Xpos = Abs_Xpos - 530
            Section_Ypos = Abs_Ypos + JB1_up + K1_up

            # 가정: SWAngle은 도 단위로 주어진 각도
            SWAngle_radians = math.radians(SWAngle)

            if(SWAngle>0):
                X1 = Section_Xpos - SW1 - SW2 * math.cos(SWAngle_radians)
                Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                X2 = Section_Xpos - SW1
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos
                
                line(doc, X2, Y2, X1, Y1, layer='0')               
                line(doc, X2, Y2, X3, Y3, layer='0')        
                dim_linear(doc, X1, Y1, X2, Y2,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # 각도 90도가 아니면 각도 표기
                if (SWAngle != 90):               
                    dim_angular(doc, X2 , Y2, X1, Y1, X2-30, Y2, X3-30, Y3,  200, direction="left" )
                    # help(ezdxf.layouts.Modelspace.add_angular_dim_2l)
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)   

            if(SWAngle<1):
                X1 = Section_Xpos - SW_Left
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos            
                line(doc,  X1, Y1, X2, Y2, layer='0')                                                   
                dim_linear(doc, X1, Y1, X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

            if(K1_up>5):
                X4 = X3
                Y4 = Y3 - K1_up
                lineto(doc,  X4, Y4, layer='0')       
            else:
                X4 = X3
                Y4 = Y3

            # Angle은 도 단위로 주어진 각도
            Angle_radians = math.radians(Angle)

            X5 = X4 - JB1_up * math.sin(Angle_radians)                
            Y5 = Y4 - JB1_up * math.cos(Angle_radians)  
            lineto(doc,  X5, Y5, layer='0')    
            X6 = X5 - C1
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='0')   

            if(A1>0):
                X7 = X6
                Y7 = Y6 + A1
                lineto(doc,  X7, Y7, layer='0') 
                drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut 2개소", layer="dim", style='0')
                # A 치수
                dim_vertical_left(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
            if(A1<1):
                X7 = X6
                Y7 = Y6            
                drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut", layer="dim", style='0')

            # C값 치수선
            dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            if(K1_up>5):
                dim_vertical_right(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                # if (Angle > 1):   # 직각이 아닌경우                        
                dim_angular(doc, X4, Y4, X5, Y5, X4, Y4, X3, Y3,  30, direction="left" )

            # JB사선 치수선
            dim_linear(doc, X4, Y4, X5, Y5,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
            if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc, X6, Y6, X5, Y5, X5, Y5, X4, Y4,  20, direction="left" )

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB1_up != JB1_down ):
                rectangle(doc, X5+60, Y5 + K1_up + (JB1_up)/2 - 75 , X5 + 60 + 100, Y5 + K1_up + (JB1_up)/2 + 75 , layer='0')
                rectangle(doc, X5+60+350, Y5 + K1_up + (JB1_up)/2 - 75, X5 + 60 + 100 + 350, Y5 + K1_up + (JB1_up)/2 + 75 , layer='0')
            if(K1_up != K1_down ):
                rectangle(doc, X4 + 50, Y4 - 10 , X4 + 50 + 100, Y4 + K1_up + 10 , layer='0')                         
                rectangle(doc, X4 + 50 + 300, Y4 - 10 , X4 + 50 + 100 + 300, Y4 + K1_up + 10 , layer='0')                         
                            
            ####################################################################################################################################################################################
            # SO와이드 좌기둥 우측 단면도    
            ####################################################################################################################################################################################
            if( JB1_up != JB1_down or K1_up != K1_down ):      
                # print ('기둥좌측 하단 단면도 작도')  
                Section_Xpos = Abs_Xpos + SMH1 + LH1 + 430
                Section_Ypos = Abs_Ypos + JB1_up + K1_up

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(SWAngle>0):  
                    X1 = Section_Xpos + SW1 + SW2 * math.cos(SWAngle_radians)
                    Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                    X2 = Section_Xpos + SW1
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc, X2, Y2, X1, Y1, layer='0')                   
                    line(doc, X2, Y2, X3, Y3, layer='0')        
                    dim_linear(doc, X2, Y2, X1, Y1,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # 각도 90도가 아니면 각도 표기
                    if (SWAngle != 90):               
                        dim_angular(doc,  X3+5, Y3, X2+5, Y2, X1, Y1, X2 , Y2,  200, direction="right" )                    
                    dim_linear(doc, X3, Y3,  X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

                if(SWAngle<1):  
                    X1 = Section_Xpos + SW_Left
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos 
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc,   X1, Y1, X2, Y2,  layer='0')                                   
                    dim_linear(doc,  X2, Y2, X1, Y1,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

                if(K1_down>5):
                    X4 = X3
                    Y4 = Y3 - K1_down
                    lineto(doc,  X4, Y4, layer='0')       
                else:
                    X4 = X3
                    Y4 = Y3

                # Angle은 도 단위로 주어진 각도
                Angle_radians = math.radians(Right_Angle)

                X5 = X4 + JB1_down * math.sin(Angle_radians)                
                Y5 = Y4 - JB1_down * math.cos(Angle_radians)  
                lineto(doc,  X5, Y5, layer='0')    
                X6 = X5 + C1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')   

                if(A1>0):
                    X7 = X6
                    Y7 = Y6 + A1
                    lineto(doc,  X7, Y7, layer='0') 
                    drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                    drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                    dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut 2개소", layer="dim", style='0')
                    # A 치수
                    dim_vertical_right(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
                if(A1<1):
                    X7 = X6
                    Y7 = Y6            
                    drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                    dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut", layer="dim", style='0')

                # C값 치수선
                dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                if(K1_up>5):
                    dim_vertical_left(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                    if (Right_Angle > 1):   # 직각이 아닌경우                        
                        dim_angular(doc,  X3, Y3, X4 , Y4,  X5, Y5, X4, Y4,   15, direction="right" )

                # JB사선 치수선
                dim_linear(doc,  X5, Y5, X4, Y4,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
                if (Right_Angle > 1):   # 직각이 아닌경우            
                    dim_angular(doc,  X5, Y5, X4, Y4, X5, Y5, X6 , Y6,  15, direction="right" )  

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB1_up != JB1_down ):
                    rectangle(doc, X5-60,Y5 + (JB1_up+K1_up)/2 - 75 , X5 - 60 - 100,Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')               
                    rectangle(doc, X5 - 60 - 300, Y5 + (JB1_up+K1_up)/2 - 75 , X5 - 60 - 100 - 300, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')   

                if(K1_up != K1_down ):
                    rectangle(doc, X4 - 50, Y4 - 10 , X4 - 50 - 100, Y4 + K1_down + 10 , layer='0')                             
                    rectangle(doc, X4 - 50 - 220, Y4 - 10 , X4 - 50 - 100 - 220, Y4 + K1_down + 10 , layer='0')                             

            #########################################################################################################################################################################
            # 우기둥 전개도 및 단면도
            # 주의 사항 기둥의 우측은 X1의 선이 끊어지는 현상을 방지하기 위해서 좌기둥과 다르게 설계해야 함  (와이드형태만 해당됨)
            #########################################################################################################################################################################

            Abs_Ypos -= 2000 
            R_Xpos = Abs_Xpos + SMH2
            R_Ypos = Abs_Ypos + 1100 - (C1 + C2 + A1 + A2)

            # 변경 코멘트 2024 03 04 ********** JD가 큰쪽의 기둥 뒷날개 해밍구조 없애기 가림판과 결합되는 부위
            # JD1이 크면 가림판과 결합되는 부위로 SWAngle에 0 적용

            # 우측기둥에서 다시 판단하게 만들려고 Original_SWAngle 변수 만듬
            SW1 = Original_SW1
            SW2 = Original_SW2        
            SWAngle = Original_SWAngle            
            if(JD1 + K1_up < JD2 + K2_up):
                SWAngle =  0
                SW1 = 0
                SW2 = 0

            if(A2>0):            
                #  X1은 좌측 끝단으로 이동해서 점을 하나 다르게 해야 함
                X1 = R_Xpos  - SMH2
                Y1 = R_Ypos + C2 + A2 - Vcut_rate * 3
                X2 = R_Xpos + RH1 
                Y2 = Y1
                line(doc, X1, Y1, X2, Y2, layer='레이져')                    
                X3 = X2
                Y3 = R_Ypos + C2 - Vcut_rate*2
                lineto(doc,  X3, Y3, layer='레이져')            

            if(A2<1):
                X1 = R_Xpos  - SMH2
                Y1 = R_Ypos + C2  - Vcut_rate 
                X2 = R_Xpos + RH1 
                Y2 = Y1
                line(doc, X1, Y1, X2, Y2, layer='레이져')                    
                X3 = X2
                Y3 = Y2

            X4 = X3
            Y4 = R_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')

            X5 = R_Xpos + RH2 

            if (SWAngle>1):  #SWAngle 0 이상인 경우    헤밍구조인 경우        
                if(K2_up>0):   
                    if (SWAngle == 90):    
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = X6
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = R_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')
                    else:   
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = X6
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = R_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')

                if(K2_up<1):   #K값이 없는 경우
                    if (SWAngle == 90):   
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = X6
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11      
                    else:
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = X6
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11

            if (SWAngle<1):  #SWAngle 0     일반구조
                if(K2_up>0):   
                    Y5 = R_Ypos - JB2_down +  Vcut_rate
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = X5
                    Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                    lineto(doc,  X6, Y6, layer='레이져')
                    X7 = X6
                    Y7 = R_Ypos - JB2_down - K2_down - SW_Right  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = R_Xpos 
                    Y9 = R_Ypos - JB2_up - K2_up - SW_Right  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = R_Xpos 
                    Y10 = Y9                
                    X11 = R_Xpos 
                    Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = R_Xpos 
                    Y12 = R_Ypos - JB2_up +  Vcut_rate
                    lineto(doc,  X12, Y12, layer='레이져')
                if(K2_up<1):   #K값이 없는 경우
                    Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = X5
                    Y6 = Y5                
                    X7 = X6
                    Y7 = R_Ypos - JB2_down  - SW_Right  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = R_Xpos 
                    Y9 = R_Ypos - JB2_up  - SW_Right  + Bending_rate * 2 + Vcut_rate
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = X9
                    Y10 = Y9                
                    X11 = R_Xpos 
                    Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = X11
                    Y12 = Y11                
            X13 = R_Xpos 
            Y13 = R_Ypos - B2
            lineto(doc,  X13, Y13, layer='레이져')
            X14 = R_Xpos - SMH2
            Y14 = Y13
            lineto(doc,  X14, Y14, layer='레이져')
            X15 = X14
            Y15 = R_Ypos
            lineto(doc,  X15, Y15, layer='레이져')
            if(A2>0):
                X16 = X15
                Y16 = R_Ypos + C2 - Vcut_rate*2
                lineto(doc,  X16, Y16, layer='레이져')
                X17 = X16
                Y17 = R_Ypos + C2 + A2 - Vcut_rate*3
                lineto(doc,  X17, Y17, layer='레이져')
            if(A2<1):
                X16 = X15 
                Y16 = R_Ypos + C2 - Vcut_rate
                X17 = X16
                Y17 = Y16
                lineto(doc,  X16, Y16, layer='레이져')
 
            # 절곡선1
            if(A2>1):        
                line(doc, X16, Y16, X3, Y3,  layer='22')
            # C2                
            line(doc, X15, Y15, X4, Y4,  layer='22')
            # K2
            if(K2_up>0):
                line(doc, X12, Y12, X5, Y5,  layer='22')
            line(doc, X11, Y11, X6, Y6,  layer='22')
            # 해밍구조
            if (SWAngle>0):   
                line(doc, X10, Y10, X7, Y7,  layer='22')

            # 하부 전개도 치수선
            dim_linear(doc,  X14, Y14, X9, Y9, "",  extract_abs(Y14,Y9) + 100,  direction="down", layer='dim')                
            dim_linear(doc,  X9, Y9, X8, Y8, "", 100,  direction="down", layer='dim')                
            dim_linear(doc,  X14, Y14, X8, Y8, "", extract_abs(Y14,Y9) + 150,  direction="down", layer='dim')                

            # 우측 절곡치수선
            if(A2 > 1):  
                dim_vertical_right(doc, X3, Y3, X2,  Y2, 130, 'dim', text_height=0.22,  text_gap=0.07)      
                #Vcut A값 C값 전개치수선     
                dim_vertical_left(doc, X2, Y2, X4,  Y4,   80 , 'dim', text_height=0.22, text_gap=0.01)                     
            if(C2 > 1):      
                dim_vertical_right(doc, X3, Y3, X4,  Y4, 80, 'dim', text_height=0.22,  text_gap=0.07)                      
            # JB    
            dim_vertical_right(doc, X5, Y5, X4,  Y4, extract_abs(X5,X4) + 80, 'dim', text_height=0.22,  text_gap=0.07) 
            if(K2_down > 0): 
                dim_vertical_right(doc, X6, Y6, X5,  Y5, extract_abs(X8,X5) + 130, 'dim', text_height=0.22,  text_gap=0.07)  
            # 뒷날개    
            dim_vertical_right(doc, X6, Y6, X7,  Y7, extract_abs(X5,X8) + 80, 'dim', text_height=0.22,  text_gap=0.07)  
            if (SWAngle > 0):   
                dim_vertical_right(doc, X7, Y7, X8,  Y8, extract_abs(X7,X8) + 180, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X2, Y2, X8,  Y8, extract_abs(X2,X8) + 230, 'dim', text_height=0.22,  text_gap=0.07)  

            # 좌측 절곡치수선
            if(A2 > 1):          
                dim_vertical_left(doc, X17, Y17, X16, Y16,  130, 'dim', text_height=0.22, text_gap=0.01)  
            if(C2 > 1):              
                dim_vertical_left(doc, X15, Y15, X16, Y16,  60, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X15, Y15, X12, Y12,  60, 'dim', text_height=0.22, text_gap=0.01)  
            if(K2_up > 0): 
                dim_vertical_left(doc, X12, Y12, X11, Y11,  SMH2 + 120, 'dim', text_height=0.22, text_gap=0.01)              
            # 뒷날개    
            dim_vertical_left(doc, X11, Y11, X10, Y10,  SMH2 + 60, 'dim', text_height=0.22, text_gap=0.01)  
            if (SWAngle > 0):               
                dim_vertical_left(doc, X10, Y10, X9, Y9,  SMH2 + 180, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X17, Y17, X9,  Y9,  240 , 'dim', text_height=0.22, text_gap=0.01)      

            # 전개도 상부 치수선    
            dim_linear(doc,  X17, Y17, X2, Y2,  "",   150,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)                      
            dim_linear(doc,  X13, Y13, X2, Y2,  "",   extract_abs(Y1,Y13) + 100,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)                      
            if RH1<RH2:
                dim_linear(doc,  X2, Y2, X5, Y5,  "",  150 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)       
            if RH1>RH2:    
                dim_linear(doc,  X8, Y8, X4, Y4, "",   150 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.01)          
            if(B2>0):
                dim_vertical_left(doc, X13, Y15, X13, Y13,  80, 'dim', text_height=0.22, text_gap=0.01)              
                dim_vertical_right(doc, X11, Y11, X13, Y13,  50, 'dim', text_height=0.22, text_gap=0.01)      

            if args.opt9:
                drawcircle(doc, X9 + 12.5, Y9 + 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X8 - 12.5, Y8 + 5, 2.5 , layer='레이져') # 도장홀 5파이            

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"       

            Xpos = R_Xpos + 400 
            Ypos = R_Ypos - JB2_up / 2
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-우"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')                       

            Xpos = R_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-우"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = R_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Mat.Spec : {widejamb_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = R_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Size : {Right_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = R_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')      

            # 좌우 도면틀 넣기
            BasicXscale = 4487
            BasicYscale = 2770
            TargetXscale = 800 + max(SMH1, SMH2) + max(LH1, LH2) + 900
            TargetYscale = 400*4 + (JB1_up + C1 + A1 + K1_up + SW_Right) * 2
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale

            # print (f'스케일 : {frame_scale}')
            insert_frame(X1 -  950 ,Y1 - (JB2_up + C2 + A2 + K2_up + SW_Right + 500 + C1 + A1 + 300) , frame_scale, "side", workplace)              
            
            ###################################
            # 우기둥 좌측 단면도    
            ###################################            
            Section_Xpos = R_Xpos - SMH2 - 500
            Section_Ypos = R_Ypos

            SWAngle_radians = math.radians(SWAngle)

            if(A2>0):
                X1 = Section_Xpos
                Y1 = Section_Ypos - A2
                X2 = Section_Xpos 
                Y2 = Section_Ypos 

                line(doc, X1, Y1, X2, Y2, layer='0')                               
                dim_vertical_left(doc, X1, Y1, X2, Y2, 50, 'dim', text_height=0.22, text_gap=0.07)  
                X3 = X2 + C2
                Y3 = Y2
                lineto(doc,  X3, Y3, layer='0')   
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # vcut 표기                
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                dim_leader_line(doc, X3, Y3 , X3+50, Y3+100, "V-cut 2개소", layer="dim", style='0')         

            if(A2<1):
                X1 = Section_Xpos
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos             
                X3 = X2 + C2
                Y3 = Y2
                line(doc, X1, Y1, X3, Y3, layer='0')               
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # vcut 표기                            
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                dim_leader_line(doc, X3, Y3 , X3+50, Y3+100, "V-cut", layer="dim", style='0')                                        

            X4 = X3 + JB2_up * math.sin(Angle_radians)                
            Y4 = Y3 - JB2_up * math.cos(Angle_radians)        
            lineto(doc,  X4, Y4, layer='0')	
            dim_linear(doc, X3, Y3, X4, Y4,  "", 50,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)     

            if(K2_up > 5):
                X5 = X4
                Y5 = Y4 - K2_up
                lineto(doc,  X5, Y5, layer='0')	
                dim_vertical_right(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
            else:
                X5 = X4
                Y5 = Y4
                
            if(SWAngle>0):
                X6 = X5 - SW1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                X7 = X6 - SW2 * math.cos(SWAngle_radians)
                Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                lineto(doc,  X7, Y7, layer='0')
                dim_linear(doc, X6, Y6, X7, Y7,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                dim_angular(doc, X5-5, Y5, X6-5, Y6,  X7, Y7, X6 , Y6,  100, direction="left" )

            if(SWAngle<1): # 헤밍구조가 아닌경우 일반형
                X6 = X5 - SW_Right
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                X7 = X6
                Y7 = Y6                        

            # if (Right_Angle > 1):   # 직각이 아닌경우            
            dim_angular( doc,   X3, Y3, X4, Y4, X3, Y3, X2 , Y2,    15, direction="left" )		                            
            if(K2_up > 5): # K2값
                dim_angular(doc,  X4, Y4, X5 , Y5, X4, Y4, X3, Y3,  5, direction="left" )    

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB2_up != JB2_down ):
                rectangle(doc, X3+60,  Y5 + (JB1_up+K1_up)/2 - 75 , X3 + 60 + 100, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')               
                rectangle(doc, X3+60+300, Y5 + (JB1_up+K1_up)/2 - 75 , X3 + 60 + 400,  Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')     

            if(K2_up != K2_down ):
                rectangle(doc, X5 + 50, Y5 - 10 , X5 + 50 + 100, Y5 + K2_up + 10 , layer='0')                         
                rectangle(doc, X5 + 60 + 250, Y5 - 10 , X5 + 60 + 350, Y5 + K2_up + 10 , layer='0')                            

            ###################################
            # 와이드 우기둥 우측 단면도    
            ###################################
                    
            if( JB2_up != JB2_down or K2_up != K2_down ):   

                Section_Xpos = R_Xpos + max(RH1,RH2) +  500
                Section_Ypos = R_Ypos

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(A2>0):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos - A2
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos 

                    line(doc, X1, Y1, X2, Y2, layer='0')
                    dim_vertical_right(doc, X2, Y2, X1, Y1,  50, 'dim', text_height=0.22, text_gap=0.07)
                    X3 = Section_Xpos
                    Y3 = Section_Ypos
                    lineto(doc,  X3, Y3, layer='0')
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # vcut 표기                
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut 2개소", layer="dim", style='0')                                
                if(A2<1):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos                 
                    X3 = Section_Xpos
                    Y3 = Section_Ypos 
                    line(doc, X2, Y2, X3, Y3, layer='0')                               
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # vcut 표기                                
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut", layer="dim", style='0')                                

                X4 = X3 - JB2_down * math.sin(Angle_radians)                
                Y4 = Y3 - JB2_down * math.cos(Angle_radians)        
                lineto(doc,  X4, Y4, layer='0')	
                dim_linear(doc, X4, Y4, X3, Y3,   "", 50,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)

                if(K2_down > 5):
                    X5 = X4
                    Y5 = Y4 - K2_down
                    lineto(doc,  X5, Y5, layer='0')	
                    dim_vertical_left(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
                else:
                    X5 = X4
                    Y5 = Y4
                    lineto(doc,  X5, Y5, layer='0')	

                if(SWAngle>0):
                    X6 = X5 + SW1
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                    X7 = X6 + SW2 * math.cos(SWAngle_radians)
                    Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                    lineto(doc,  X7, Y7, layer='0')
                    dim_linear(doc, X7, Y7, X6, Y6,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                    dim_angular(doc,  X7 , Y7, X6, Y6, X6 + 5 , Y6, X6, Y6,  150, direction="right" )
                if(SWAngle<1):
                    X6 = X5 + SW_Right
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,"" , 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    X7 = X6
                    Y7 = Y6

                if (Right_Angle > 1):   # 직각이 아닌경우            
                    dim_angular(doc,  X3, Y3, X2-8 , Y2,  X3, Y3, X4, Y4,   25, direction="right" )		            
                    if(K2_down>0):
                        dim_angular(doc,   X4, Y4, X3 , Y3, X5, Y5,  X4, Y4, 30, direction="right" )    

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB2_up != JB2_down ):
                    rectangle(doc, X3 - 60, Y5 + (JB1_up+K1_up)/2 - 75 , X3 - 60 - 100, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')               
                    rectangle(doc, X3 - 60 - 200, Y5 + (JB1_up+K1_up)/2 - 75 , X3 - 60 - 100 - 200, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')      

                if(K2_up != K2_down ):
                    rectangle(doc, X5 - 50, Y5 - 10 , X5 - 50 - 100, Y5 + K2_down + 10 , layer='0')                             
                    rectangle(doc, X5 - 50 - 250, Y5 - 10 , X5 - 50 - 100 - 250, Y5 + K2_down + 10 , layer='0')                                        

            Abs_Ypos -= 2000    

##########################################################################################################################################
# 와이드 작도
##########################################################################################################################################
def execute_wide():     

    # 시트 선택 (시트명을 지정)
    sheet_name = '와이드제작'  # 원하는 시트명으로 변경
    sheet = workbook[sheet_name]

    # 2차원 배열을 저장할 리스트 생성
    data_2d = []

    # 1행부터 20행까지의 값을 2차원 배열에 저장
    for row_num in range(7, sheet.max_row + 1):  # 7행부터 20행까지 반복
        cell_value = sheet.cell(row=row_num, column=2).value  # 2열의 값 가져오기
        if cell_value is not None and cell_value != 'F':  # 값이 None이 아닌 경우에만 배열에 추가
            row_values =  [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, sheet.max_column + 1)]        
            data_2d.append(row_values)

    # 엑셀 파일 닫기
    workbook.close()

    if not data_2d:
        print("해당 데이터가 없습니다.")
        return  # 함수를 빠져나감    

    #######################################################################################################################################################################################
    # 와이드 상판 전개도
    #######################################################################################################################################################################################    
    Abs_Xpos = 2000
    Abs_Ypos = 0
    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        MH1 = row[2]
        MH2 = row[3]
        OP = row[4]
        R = row[5]
        JD1 = row[6]
        JD2 = row[7]
        JB1_up = row[8]
        JB1_down = row[9]
        JB2_up = row[10]
        JB2_down = row[11]
        H1 = row[12]
        H2 = row[13]
        C1 = row[14]        
        C2 = row[15]
        B1 = row[16]
        B2 = row[17]
        A1 = row[18]
        A2 = row[19]
        LH1 = row[20]
        LH2 = row[21]
        RH1 = row[22]
        RH2 = row[23]        
        U = row[24]        
        G = row[25]        
        UW = row[26]        
        SW = row[27]        
        K1_up = row[28]        
        K1_down = row[29]        
        K2_up = row[30]        
        K2_down = row[31]                
        SMH1 = row[32]                
        SMH2 = row[33]                
        G_check = row[34]
        Upper_size = row[35]
        Left_side_size = row[36]
        Right_side_size = row[37]
        E1 = row[38]
        E2 = row[39]
        HPI_height = row[40]
        SW1 = row[41]
        SW2 = row[42]
        SWAngle = row[43]
        A1_var = row[44]
        A2_var = row[45]
        Angle = row[46]
        Kadapt = row[60]  # K값 적용여부 1 이면 K값 적용하지 말것

        # 데이터 검증 및 기본값 설정
        SWAngle = validate_or_default(SWAngle)
        K1_up = validate_or_default(K1_up)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_down = validate_or_default(K2_down)
        U = validate_or_default(U)
        G = validate_or_default(G)
        HPI_height = validate_or_default(HPI_height)
        OP = validate_or_default(OP)
        Kadapt = validate_or_default(Kadapt)

        if(Kadapt==1):
            K1_up = 0
            K2_up = 0                           

        if(K1_up>=10 and K2_up>=10 ):
            # 추영덕소장 E1, E2값 무시함 위의 조건이라면
            E1 = 0
            E2 = 0

        if(OP>0):
            # 상판 기둥과 만나는 돌출값 B값과 연결되는 부분
            # Top_Pop1 = G-B1-1.5 
            # Top_Pop2 = G-B2-1.5 
            Top_Pop1 = G_check
            Top_Pop2 = G_check

            if(Top_Pop1<4):
                Top_Pop1 = 4
            if(Top_Pop2<4):
                Top_Pop2 = 4

            # 손상민소장의 경우 적용
            if(worker=="손상민" and U>2):
                draw_Uframe(Abs_Xpos, Abs_Ypos, U,G ,R, C1, C2, Floor_mc, Floor_des )  
            if(H1<1 or H2<1):
                U = 0
                G = 0        
                

            SWAngle_radians = math.radians(SWAngle)
            if(SWAngle>0):
                SWAngle_var =  SW2 * math.cos(SWAngle_radians)        
            Angle_radians = math.radians(Angle)
            if(Angle>-1):
                Angle_cos = math.cos(Angle_radians)        
                Angle_sin = math.sin(Angle_radians)     
        
            if(U > 0):
                X1 = Abs_Xpos - H1
                Y1 = Abs_Ypos + MH1 + G + U - Vcut_rate*5
                X2 = Abs_Xpos  + H2 + R 
                Y2 = Abs_Ypos + MH2 + G + U - Vcut_rate*5
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Abs_Ypos + MH2 + G  - Vcut_rate * 4
                lineto(doc, X3, Y3, layer='레이져')          
                X4 = X3
                Y4 = Abs_Ypos + MH2 + G - Top_Pop2 - Vcut_rate * 3
                lineto(doc, X4, Y4, layer='레이져')          
                X5 = X4 - H2 + (G - Top_Pop2) * Angle_sin
                Y5 = Y4
                lineto(doc, X5, Y5, layer='레이져')          
                X6 = Abs_Xpos + R 
                Y6 = Abs_Ypos + MH2 - Vcut_rate * 2
                lineto(doc,  X6, Y6, layer='레이져')
                X7 = X6
                Y7 = Abs_Ypos
                lineto(doc,  X7, Y7, layer='레이져')     

            if(U < 1):
                if(G>0):
                    X1 = Abs_Xpos - H1
                    Y1 = Abs_Ypos + MH1 + G - Vcut_rate*3
                    X2 = Abs_Xpos  + H2 + R 
                    Y2 = Y1
                    line(doc, X1, Y1, X2, Y2, layer='레이져')       
                    X3 = X2
                    Y3 = Y2
                    lineto(doc, X3, Y3, layer='레이져')          
                    X4 = X3
                    Y4 = Abs_Ypos + MH2 + G - Top_Pop2 - Vcut_rate * 3
                    lineto(doc, X4, Y4, layer='레이져')          
                    X5 = X4 - H2 + (G - Top_Pop2) * Angle_sin
                    Y5 = Y4
                    lineto(doc, X5, Y5, layer='레이져')          
                    X6 = Abs_Xpos + R 
                    Y6 = Abs_Ypos + MH2 - Vcut_rate * 2
                    lineto(doc,  X6, Y6, layer='레이져')
                    X7 = X6
                    Y7 = Abs_Ypos
                    lineto(doc,  X7, Y7, layer='레이져')   
                else:
                    X1 = Abs_Xpos 
                    Y1 = Abs_Ypos + MH1 - Vcut_rate
                    X2 = Abs_Xpos + R 
                    Y2 = Abs_Ypos + MH2 - Vcut_rate
                    line(doc, X1, Y1, X2, Y2, layer='레이져')       
                    X3 = X2
                    Y3 = Y2            
                    X4 = X3
                    Y4 = Y3            
                    X5 = X4
                    Y5 = Y4            
                    X6 = X5
                    Y6 = Y5            
                    X7 = X6
                    Y7 = Abs_Ypos
                    lineto(doc,  X7, Y7, layer='레이져')            

            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                X8 = Abs_Xpos + (R-OP)/2 + OP          
                if(Kadapt==1 and E2>0):
                    Y8 = Abs_Ypos - (JD2-E2) + Vcut_rate 
                else:
                    Y8 = Abs_Ypos - JD2 + Vcut_rate 
                lineto(doc,  X8, Y8, layer='레이져')
            else:
                X8 = Abs_Xpos + (R-OP)/2 + OP    
                Y8 = Abs_Ypos - JD2 + Vcut_rate * 2
                lineto(doc,  X8, Y8, layer='레이져')  
            
            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                X9 = X8
                if(E2>0):
                    Y9 = Abs_Ypos - JD2 + Vcut_rate*2
                else:
                    Y9 = Abs_Ypos - JD2 - K2_up + Vcut_rate*2
                lineto(doc,  X9, Y9, layer='레이져') 
                X10 = X9 
                if(E2>0):                
                    Y10 = Abs_Ypos - JD2 - UW + Vcut_rate*3     
                else:
                    Y10 = Abs_Ypos - JD2 - K2_up - UW + Vcut_rate*3     
                lineto(doc,  X10, Y10, layer='레이져')   
                X11 = Abs_Xpos  + (R-OP)/2 
                if(E1>0):                
                    Y11 = Abs_Ypos - JD1 - UW + Vcut_rate*3     
                else:                    
                    Y11 = Abs_Ypos - JD1 - K1_up - UW + Vcut_rate*3     
                lineto(doc,  X11, Y11, layer='레이져')      
                X12 = X11 
                if(E1>0):                
                    Y12 =  Abs_Ypos - JD1 + Vcut_rate*2
                else:                    
                    Y12 =  Abs_Ypos - JD1 - K1_up + Vcut_rate*2
                lineto(doc,  X12, Y12, layer='레이져')                                      
                X13 = X12 
                if(Kadapt==1 and E1>0):
                    Y13 =  Abs_Ypos - (JD1-E1) + Vcut_rate
                else:                
                    Y13 =  Abs_Ypos - JD1 + Vcut_rate
                lineto(doc,  X13, Y13, layer='레이져')                                      
            else:
                X9 = X8
                Y9 = Abs_Ypos - JD2  + Vcut_rate*2
                lineto(doc,  X9, Y9, layer='레이져') 
                X10 = X9 
                Y10 = Abs_Ypos - JD2 -  UW + Vcut_rate*3  
                lineto(doc,  X10, Y10, layer='레이져')   
                X11 = Abs_Xpos + (R-OP)/2 
                Y11 = Abs_Ypos - JD1 - UW + Vcut_rate*3     
                lineto(doc,  X11, Y11, layer='레이져')      
                X12 = X11 
                Y12 =  Abs_Ypos - JD1  + Vcut_rate*2
                lineto(doc,  X12, Y12, layer='레이져')                                      
                X13 = X12 
                Y13 =  Abs_Ypos - JD1 + Vcut_rate*2
                lineto(doc,  X13, Y13, layer='레이져')               
            
            X14 = Abs_Xpos
            Y14 = Abs_Ypos
            lineto(doc,  X14, Y14, layer='레이져')      
    
            if(U > 0):         
                X15 = Abs_Xpos
                Y15 = Abs_Ypos + MH1 - Vcut_rate * 2
                lineto(doc,  X15, Y15, layer='레이져')                
                X16 = Abs_Xpos - (G-Top_Pop1) * Angle_sin
                Y16 = Abs_Ypos + MH1 + G - Top_Pop1 - Vcut_rate * 3
                lineto(doc,  X16, Y16, layer='레이져')      
                X17 = Abs_Xpos - H1
                Y17 = Y16
                lineto(doc,  X17, Y17, layer='레이져')      
                X18 = X17
                Y18 = Abs_Ypos + MH1 + G  - Vcut_rate * 4
                lineto(doc,  X18, Y18, layer='레이져')      
                lineto(doc,  X1, Y1, layer='레이져')       

            if(U < 1):   
                if( G < 1):
                    X15 = Abs_Xpos
                    Y15 = Abs_Ypos + MH1 - Vcut_rate 
                    lineto(doc,  X1, Y1, layer='레이져')       
                    X16 = X15
                    Y16 = Y15                    
                    X17 = X16
                    Y17 = Y16
                    X18 = X17
                    Y18 = Y17
                else:
                    X15 = Abs_Xpos
                    Y15 = Abs_Ypos + MH1 - Vcut_rate * 2
                    lineto(doc,  X15, Y15, layer='레이져')                
                    X16 = Abs_Xpos - (G-Top_Pop1) * Angle_sin
                    Y16 = Abs_Ypos + MH1 + G - Top_Pop1 - Vcut_rate * 3
                    lineto(doc,  X16, Y16, layer='레이져')      
                    X17 = Abs_Xpos - H1
                    Y17 = Y16
                    lineto(doc,  X17, Y17, layer='레이져')      
                    X18 = X17
                    Y18 = Abs_Ypos + MH1 + G  - Vcut_rate * 3
                    lineto(doc,  X18, Y18, layer='레이져')      
                    lineto(doc,  X1, Y1, layer='레이져')                

            if args.opt7:
                drawcircle(doc, X11 + 12.5, Y11 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X10 - 12.5, Y10 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이

            # 절곡선
            if(U>0):
                line(doc, X18, Y18, X3, Y3,  layer='22')        
            if(G>0):            
                # 실선처리
                line(doc, X15, Y15, X6, Y6,  layer='HIDDEN2')

            line(doc, X14, Y14, X7, Y7,  layer='22')
            line(doc, X12, Y12, X9, Y9,  layer='22')

            # HPI 그리기
            if(HPI_height>0):
                draw_HPI(Abs_Xpos + (X7-X14)/2, Y14, R, HPI_height )     
            else:
                draw_NoHPI(Abs_Xpos + (X7-X14)/2, Y14 )        

            #  전개도 상부 치수선
            dim_linear(doc,  X1, Y1, X2, Y2, "", 120,  direction="up", layer='dim')

            if(G > 0):   
                dim_linear(doc,  X1, Y1, X15, Y15,  "", 50,  direction="up", layer='dim')
                dim_linear(doc,  X15, Y15, X6, Y6, "",  extract_abs(Y6,Y1) + 50,  direction="up", layer='dim')
                dim_linear(doc,  X6, Y6, X2, Y2,  "",  extract_abs(Y2,Y6) + 50,  direction="up", layer='dim')
    
            # 우측 절곡치수선
            if(U>0):        
                dim_vertical_right(doc, X2, Y2, X3, Y3, 160 , 'dim', text_height=0.22, text_gap=0.07)  
            if(G>0):                    
                dim_vertical_right(doc, X6,  Y6, X3, Y3,  extract_abs(X2,X6) + 100 , 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X6, Y6,  X7, Y7,    extract_abs(X2,X6) +  100, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X9, Y9,  X7, Y7,   extract_abs(X2,X6) + extract_abs(X9,X6) + 100, 'dim', text_height=0.22, text_gap=0.07)   
            dim_vertical_right(doc, X9, Y9, X10, Y10,  extract_abs(X2,X6) + extract_abs(X9,X6) + 100, 'dim', text_height=0.22, text_gap=0.07)   # 뒷날개
            dim_vertical_right(doc, X2, Y2,  X10, Y10, extract_abs(X2,X6) + extract_abs(X9,X6) + 170, 'dim', text_height=0.22, text_gap=0.07)             

            # 좌측 절곡치수선(테둘림이 없는 경우 V컷 치수 변경)
            if(JD1 != JD2 or K1_up != K2_up or MH1 != MH2 or E1 != E2):
                if(U>0):                    
                    dim_vertical_left(doc, X1, Y1, X18, Y18,   120, 'dim', text_height=0.22, text_gap=0.07)  
                if(G>0):                        
                    dim_vertical_left(doc, X15, Y15, X18, Y18,   extract_abs(X14,X18) + 70 , 'dim', text_height=0.22, text_gap=0.07)     
                # MH                             
                dim_vertical_left(doc, X14, Y14, X15, Y15,  extract_abs(X14,X18) + 70 , 'dim', text_height=0.22, text_gap=0.07)      
                # JD      
                dim_vertical_left(doc, X14, Y14, X12, Y12,   extract_abs(X14,X18) + 70 , 'dim', text_height=0.22, text_gap=0.07)           
                # 뒷날개
                dim_vertical_left(doc, X11, Y11, X12, Y12,  extract_abs(X14,X18) +  extract_abs(X14,X11) + 70, 'dim', text_height=0.22, text_gap=0.07)                          
                # 좌측 전체
                dim_vertical_left(doc, X11, Y11, X1, Y1,   extract_abs(X1,X11) + 220, 'dim', text_height=0.22, text_gap=0.07)                          
            else:    
                dim_vertical_left(doc, X11, Y11, X12, Y12, extract_abs(X11,X14) + 50, 'dim', text_height=0.22,  text_gap=0.07) 
                dim_vertical_left(doc, X11, Y11, X14, Y14, extract_abs(X11,X14) + 120, 'dim', text_height=0.22,  text_gap=0.07) 
                if(G>0):  
                    dim_vertical_left(doc, X11, Y11, X15, Y15, extract_abs(X11,X18) + 180 , 'dim', text_height=0.22,  text_gap=0.07) 
                    dim_vertical_left(doc, X11, Y11, X18, Y18, extract_abs(X11,X18) + 240 , 'dim', text_height=0.22,  text_gap=0.07) 

            # 쫄대타입 상부 Vcut 치수 표기
            if(U>0 and G>0):   
               dim_vertical_left(doc, X1+H1+OP/2, Y1, X1+H1+OP/2, Y6, 100, 'dim', text_height=0.22,  text_gap=0.07)  
               
            # 하부 치수선 3개
            dim_linear(doc,  X14, Y14, X11, Y11,  "",  extract_abs(Y11, Y14) + 100 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X11, Y11, X10, Y10,  "",  100 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            if(JD1+K1_up > JD2+K2_up) :
                dim_linear(doc,  X10, Y10, X7, Y7,    "",  100 + extract_abs(JD1+K1_up, JD2+K2_up) + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            else:
                dim_linear(doc,  X10, Y10, X7, Y7,    "",  100 + extract_abs(JD1+K1_up, JD2+K2_up) - extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            
            # 상판 돌출부위 H1, H2 치수선
            if(G > 0):  
                dim_vertical_right(doc, X18, Y18, X16, Y16, 100, 'dim', text_height=0.22,  text_gap=0.07)
                dim_vertical_left(doc, X3, Y3, X5, Y5, 100, 'dim', text_height=0.22,  text_gap=0.07)

            # 기둥과의 접선 부위 aligned 표현 각도표현
            # 좌측
            dim_linear(doc, X14, Y14, X12, Y12,   "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) 
            if not abs(Angle) <= 1:# 각도가 1도 이내가 아니면
                dim_angular(doc, X13, Y13, X14, Y14, X12, Y12, X13, Y13 + 10,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다
            # 우측        
            dim_linear(doc, X9, Y9, X7, Y7,  "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) 
            if not abs(Angle) <= 1:# 각도가 1도 이내가 아니면            
                dim_angular(doc, X9, Y9  , X8, Y8 + 10,  X8, Y8,  X7, Y7,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다

            # 와이드 상판 K1 K2 값 부분 치수선 
            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                dim_vertical_right(doc, X12, Y12, X13, Y13, 80, 'dim', text_height=0.22,  text_gap=0.07)  
                dim_vertical_left(doc, X9, Y9, X8, Y8, 80, 'dim', text_height=0.22,  text_gap=0.07)  

            # 문구 설정
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"
            # Textstr 생성
            Textstr = f"({Pre_text}{str(Floor_des)})"       

            X2 = Abs_Xpos + (H1 + R + H2)/2 - len(Textstr)*50/3
            Y2 = Abs_Ypos - JD1/1.5
            draw_Text(doc, X2, Y2, 50, str(Textstr), '레이져')

            # 문구            
            sizedes = Upper_size                  
            
            Textstr = f"Part Name : 상판({Pre_text}{str(Floor_des)})"    
            draw_Text(doc, X10+500, Y10 + 50 - 150, 28, str(Textstr), '0')
            Textstr = f"Mat.Spec : {widejamb_material}"    
            draw_Text(doc, X10+500, Y10 - 150, 28, str(Textstr), '0')
            Textstr = f"Size : {sizedes}"    
            draw_Text(doc, X10+500, Y10 - 50 - 150 , 28, str(Textstr), '0')
            Textstr = f"Quantity : 1 EA"    
            draw_Text(doc, X10+500, Y10 - 100 - 150 , 28, str(Textstr), '0')            

            ##################################################################
            # 와이드 상판 좌측 단면도    (좌우가 다른 모양일때 단면도를 그려준다)
            #################################################################
            if(JD1 != JD2 or K1_up != K2_up or MH1 != MH2 or E1 != E2):
                Section_Xpos = Abs_Xpos - H1 - 400 - G           

                if(G > 0):
                    X1 = Section_Xpos + G 
                    Y1 = Abs_Ypos + MH1 + U
                    X2 = X1
                    Y2 = Y1 - U  
                    X3 = X2 - G
                    Y3 = Y2
                    if(U>0):
                        line(doc, X1, Y1, X2, Y2, layer='0')                                    
                        dim_vertical_right(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)                         
                        lineto(doc,  X3, Y3, layer='0')    
                        drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    else:
                        line(doc, X2, Y2, X3, Y3, layer='0')  
                    dim_linear(doc, X3, Y3, X1, Y1, "", 100,  direction="up", layer='dim')                                
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                
                    
                    X4 = X3 
                    Y4 = Y3 - MH1
                    lineto(doc,  X4, Y4, layer='0')                                
                if(G < 1):                       
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos + MH1 
                    X2 = X1
                    Y2 = Y1
                    X3 = X2
                    Y3 = Y2                
                    X4 = X3 
                    Y4 = Y3 - MH1
                    line(doc, X3, Y3, X4, Y4, layer='0')                      
                    
                    

                # if(G > 1):
                #     drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
                #     if(U>0):                    
                #         dim_leader_line(doc, X4, Y4 , X4 - 200 , Y4+100, "V-cut 4개", layer="dim", style='0')      
                #     else:            
                #         dim_leader_line(doc, X4, Y4 , X4 - 200 , Y4+100, "V-cut 3개", layer="dim", style='0')                              
                # if(G < 1):
                #     drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
                #     dim_leader_line(doc, X4, Y4 , X4 - 100 , Y4+100, "V-cut 2개", layer="dim", style='0')                              

                if(K1_up > 0):
                    X5 = X4 - JD1
                    Y5 = Y4 
                    lineto(doc,  X5, Y5, layer='0')               
                    X6 = X5 - K1_up
                    Y6 = Y5 
                    lineto(doc,  X6, Y6, layer='0')   
                    dim_linear(doc,  X5, Y5, X4, Y4,  "", 80, direction="down", layer='dim')     
                    dim_linear(doc,  X5, Y5, X6, Y6,  "", 80, direction="down", layer='dim')     
                    dim_linear(doc,  X6, Y6, X4, Y4,  "", 150, direction="down", layer='dim')                  
                else:
                    X5 = X4 - JD1
                    Y5 = Y4 
                    lineto(doc,  X5, Y5, layer='0')               
                    X6 = X5 
                    Y6 = Y5                                 
                    dim_linear(doc,  X6, Y6, X4, Y4,  "", 150 , direction="down", layer='dim')     

                # drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                X7 = X6 
                Y7 = Y6 + UW
                lineto(doc,  X7, Y7, layer='0')     
                dim_vertical_left(doc,  X6, Y6, X7, Y7, 50,'dim' ,text_height=0.22,  text_gap=0.07)            

                if(K1_up > 0):                 
                    # K값 부분 라인 그리기
                    line(doc, X5, Y5, X5 , Y5 - UW , layer='22')               

                dim_vertical_left(doc,  X6, Y6, X3, Y3,  110,'dim' ,text_height=0.22,  text_gap=0.07)    
                
                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JD1 != JD2 or K1_up != K2_up ):
                    rectangle(doc, X6+(JD1+K1_up)/2 -55, Y6-90, X6+(JD1+K1_up)/2 + 55, Y6-190, layer='0')               
                    rectangle(doc, X4+250, Y6-40, X4+350, Y6 - JD1 + 40, layer='0')               
                if( MH1 != MH2 ):
                    rectangle(doc, X6-70, Y6 + MH1/2 + 90 , X6 - 70 - 100 , Y6 + MH1/2 - 90, layer='0')               
                    rectangle(doc, X4+230, Y6 + MH1/2 + 90  , X4+330,  Y6 + MH1/2 - 90 , layer='0')               
            
            ###########################################################################
            # 와이드 상판 우측 단면도 기본적으로 다 그림
            ###########################################################################
            Section_Xpos = Abs_Xpos + R + H2 + 450 + G           

            if(G > 0):
                X1 = Section_Xpos - G 
                Y1 = Abs_Ypos + MH2 + U
                X2 = X1
                Y2 = Y1 - U                          
                X3 = X2 + G
                Y3 = Y2
                if(U>0):
                    line(doc, X1, Y1, X2, Y2, layer='0')                                    
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') #상판 쫄대 V-CUT
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6') #상판 쫄대 V-CUT
                    
                    dim_vertical_left(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)                         
                    lineto(doc,  X3, Y3, layer='0')    
                else:
                    line(doc, X2, Y2, X3, Y3, layer='0')  

                lineto(doc,  X3, Y3, layer='0')    
                dim_linear(doc, X1, Y1, X3, Y3,  "", 100,  direction="up", layer='dim')                            
                # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                
                X4 = X3 
                Y4 = Y3 - MH2
                lineto(doc,  X4, Y4, layer='0')              

            if(G < 1):                       
                X1 = Section_Xpos 
                Y1 = Abs_Ypos + MH2
                X2 = X1
                Y2 = Y1
                X3 = X2
                Y3 = Y2                
                X4 = X3 
                Y4 = Y3 - MH2
                line(doc, X3, Y3, X4, Y4, layer='0')                      
                
            
            if(G > 1):
                # drawcircle(doc, X4, Y4 , 5 , layer='0', color='6')             
                if(U>0):                
                    dim_leader_line(doc, X4, Y2 , X4 + 100 , Y2+100, "V-cut 2개", layer="dim", style='0')      #상판 쫄대 V-CUT
                else:
                    dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 3개", layer="dim", style='0')      #상판 쫄대 V-CUT
            # if(G < 1):
            #     drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
            #     dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 2개", layer="dim", style='0')          

            # if(G > 1):
            #     drawcircle(doc, X4, Y4 , 5 , layer='0', color='6')             
            #     if(U>0):                
            #         dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 4개", layer="dim", style='0')      
            #     else:
            #         dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 3개", layer="dim", style='0')      
            # if(G < 1):
            #     drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
            #     dim_leader_line(doc, X4, Y4 , X4 + 100 , Y4+100, "V-cut 2개", layer="dim", style='0')      

            if(K2_up > 0):
                X5 = X4 + JD2
                Y5 = Y4 
                lineto(doc,  X5, Y5, layer='0')               
                X6 = X5 + K2_up
                Y6 = Y5 
                lineto(doc,  X6, Y6, layer='0')   
                dim_linear(doc,  X4, Y4, X5, Y5, "", 80, direction="down", layer='dim')     
                dim_linear(doc,  X5, Y5, X6, Y6, "", 80, direction="down", layer='dim')     
                dim_linear(doc,  X4, Y4, X6, Y6, "", 150, direction="down", layer='dim')                  
            else:
                X5 = X4 + JD2
                Y5 = Y4 
                lineto(doc,  X5, Y5, layer='0')               
                X6 = X5 
                Y6 = Y5                                 
                dim_linear(doc,  X4, Y4, X6, Y6, "", 150 , direction="down", layer='dim')     

            #drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
            X7 = X6 
            Y7 = Y6 + UW
            lineto(doc,  X7, Y7, layer='0')     
            dim_vertical_right(doc,  X6, Y6, X7, Y7, 50,'dim' ,text_height=0.22,  text_gap=0.07)            

            if(K2_up > 0):                 
                # K값 부분 라인 그리기
                line(doc, X5, Y5, X5 , Y5 - UW , layer='22')        

            dim_vertical_right(doc,   X6, Y6, X3, Y3,   110,'dim' ,text_height=0.22,  text_gap=0.07)                           

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JD1 != JD2 or K1_up != K2_up ):
                rectangle(doc, X6-(JD1+K1_up)/2 + 55, Y6-90, X6-(JD1+K1_up)/2 - 55, Y6-190, layer='0')                               
                rectangle(doc, X4 - 320, Y6 - 40, X4 - 420, Y6 - JD1 + 40, layer='0')        
            if( MH1 != MH2 ):
                rectangle(doc, X6 + 60, Y6 + MH2/2 + 90 , X6 + 60 + 100 , Y6 + MH2/2 - 90, layer='0')               
                rectangle(doc, X4 - 330, Y6 + MH2/2 + 90  , X4 - 430,  Y6 + MH2/2 - 90 , layer='0')                           

            # 상판 도면틀 넣기
            BasicXscale = 2560
            BasicYscale = 1550
            TargetXscale = 1000 + R + (JD1 + K1_up)*2 + 200
            TargetYscale = 500 + JD1 + MH1 + G + U
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            insert_frame(X1 - ( R + (JD1 + K1_up)*2 + 900 ),Y1-(850 + JD1 + MH1 + G + U) , frame_scale, "top", workplace)
    
            # 다음 도면 간격 계산
            Abs_Ypos -= 4000

    Abs_Xpos = 2000 + 5000
    Abs_Ypos = 1000

    ###################################################################################################################################################################
    # 와이드 좌기둥 전개도
    ###################################################################################################################################################################
    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        MH1 = row[2]
        MH2 = row[3]
        OP = row[4]
        R = row[5]
        JD1 = row[6]
        JD2 = row[7]
        JB1_up = row[8]
        JB1_down = row[9]
        JB2_up = row[10]
        JB2_down = row[11]
        H1 = row[12]
        H2 = row[13]
        C1 = row[14]        
        C2 = row[15]
        B1 = row[16]
        B2 = row[17]
        A1 = row[18]
        A2 = row[19]
        LH1 = row[20]
        LH2 = row[21]
        RH1 = row[22]
        RH2 = row[23]        
        U = row[24]        
        G = row[25]        
        UW = row[26]        
        SW = row[27]        
        K1_up = row[28]        
        K1_down = row[29]        
        K2_up = row[30]        
        K2_down = row[31]                
        SMH1 = row[32]                
        SMH2 = row[33]                
        G_check = row[34]
        Upper_size = row[35]
        Left_side_size = row[36]
        Right_side_size = row[37]
        E1 = row[38]
        E2 = row[39]
        HPI_height = row[40]
        SW1 = row[41]
        SW2 = row[42]
        SWAngle = row[43]
        A1_var = row[44]
        A2_var = row[45]
        Angle = row[46]
        Kadapt = row[60]  # K값 적용여부 1 이면 K값 적용하지 말것

        # 데이터 검증 및 기본값 설정
        SWAngle = validate_or_default(SWAngle)
        K1_up = validate_or_default(K1_up)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_down = validate_or_default(K2_down)
        U = validate_or_default(U)
        G = validate_or_default(G)
        HPI_height = validate_or_default(HPI_height)
        OP = validate_or_default(OP)
        Kadapt = validate_or_default(Kadapt)
        
        
         
        if(Kadapt==1):
            K1_up = 0            
            K1_down = 0            
            K2_up = 0            
            K2_down = 0            

        if(OP>0):            
            SWAngle_radians = math.radians(SWAngle)
            if(SWAngle>0):
                SWAngle_var =  SW2 * math.cos(SWAngle_radians)        
            Angle_radians = math.radians(Angle)
            if(Angle>0):
                Angle_cos = math.cos(Angle_radians)        
                Angle_sin = math.sin(Angle_radians)    
            
            if (SWAngle>1):  #SWAngle 0 이상인 경우            
                if(K1_up>0):      
                    if (SWAngle == 90):      
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 3 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3                
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4           
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')
                    else:
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 2 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3                
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4           
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                else:
                    if (SWAngle == 90):                            
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 3 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3          
                        Y4 =  Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate   
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4              
                    else:
                        X1 = Abs_Xpos + SMH1
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + SMH1 + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 2 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = X3          
                        Y4 =  Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate   
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4              

            if (SWAngle<1):  #SWAngle 0 이상인 경우
                if(K1_up>0):            
                    X1 = Abs_Xpos + SMH1
                    Y1 = Abs_Ypos + JB1_up + K1_up + SW - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + SMH1 + LH2 
                    Y2 = Abs_Ypos + JB1_down + K1_down + SW - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                    X3 = X2            
                    Y3 = Y2
                    X4 = X3
                    Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = X4               
                    Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                    lineto(doc,  X5, Y5, layer='레이져')                    
                else:
                    X1 = Abs_Xpos + SMH1
                    Y1 = Abs_Ypos + JB1_up + SW - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + SMH1 + LH2 
                    Y2 = Abs_Ypos + JB1_down + SW - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')                        
                    X3 = X2            
                    Y3 = Y2
                    X4 = X3             
                    Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = X4             
                    Y5 = Y4                

            X6 = Abs_Xpos +  SMH1 + LH1             
            Y6 = Abs_Ypos
            lineto(doc,  X6, Y6, layer='레이져')
            if(A1>0):     
                X7 = X6 
                Y7 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X7, Y7, layer='레이져')                  
                X8 = X7
                Y8 = Abs_Ypos - C1 - A1 + Vcut_rate * 3
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = Abs_Xpos             
                Y10 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X10, Y10, layer='레이져')

            if(A1<1):       
                X7 = X6 
                Y7 = Abs_Ypos - C1 + Vcut_rate 
                lineto(doc,  X7, Y7, layer='레이져')                
                X8 = Abs_Xpos + SMH1 + LH1  
                Y8 = Y7
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = X9
                Y10 = Y9            

            X11 = Abs_Xpos 
            Y11 = Abs_Ypos
            lineto(doc,  X11, Y11, layer='레이져')        

            if (SWAngle>0): 
                if(K1_up>0):  
                    if (SWAngle == 90):
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1  
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Abs_Ypos + JB1_up + K1_up - Bending_rate   - Vcut_rate   
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 3 - Vcut_rate
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1  
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                else:  
                    if (SWAngle == 90):                      
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1     
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Y14
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + SW1  - Bending_rate * 3 - Vcut_rate  
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                      
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + B1     
                        lineto(doc,  X12, Y12, layer='레이져')                            
                        X13 = Abs_Xpos + SMH1
                        Y13 = Y12
                        lineto(doc,  X13, Y13, layer='레이져')                
                        X14 = X13
                        Y14 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X14, Y14, layer='레이져')
                        X15 = X14
                        Y15 = Y14
                        lineto(doc,  X15, Y15, layer='레이져')
                        X16 = X15 
                        Y16 = Abs_Ypos + JB1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X16, Y16, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                      
                    
            if (SWAngle<1):  # 헤밍구조 아닌경우
                if(K1_up>0):  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + B1
                    lineto(doc,  X12, Y12, layer='레이져')                            
                    X13 = Abs_Xpos + SMH1
                    Y13 = Y12
                    lineto(doc,  X13, Y13, layer='레이져')                
                    X14 = X13
                    Y14 = Abs_Ypos +  JB1_up  - Vcut_rate
                    lineto(doc,  X14, Y14, layer='레이져')
                    X15 = X14
                    Y15 = Abs_Ypos + JB1_up + K1_up  - Bending_rate  - Vcut_rate  
                    lineto(doc,  X15, Y15, layer='레이져')
                    X16 = X15 
                    Y16 = Y15
                    lineto(doc,  X1, Y1, layer='레이져')
                else:  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + B1
                    lineto(doc,  X12, Y12, layer='레이져')                            
                    X13 = Abs_Xpos + SMH1
                    Y13 = Y12
                    lineto(doc,  X13, Y13, layer='레이져')                
                    X14 = X13
                    Y14 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate  
                    lineto(doc,  X14, Y14, layer='레이져')
                    X15 = X14
                    Y15 = Y14
                    lineto(doc,  X15, Y15, layer='레이져')
                    X16 = X15 
                    Y16 = Y15
                    lineto(doc,  X1, Y1, layer='레이져')      

            # 좌기둥 절곡선
            if(SW1>0):
                line(doc, X16, Y16, X3, Y3,  layer='22')
            line(doc, X15, Y15, X4, Y4,  layer='22')
            if(K1_up>0):
                line(doc, X14, Y14, X5, Y5,  layer='22')
            # JB    
            line(doc, X11, Y11, X6, Y6,  layer='22')
            if(A1>0):
                line(doc, X10, Y10, X7, Y7,  layer='22')                

            # 좌기둥 상부 전개도 치수선        
            dim_linear(doc,  X12, Y12, X1, Y1, "", extract_abs(Y12,Y1) + 100,  direction="up", layer='dim')                
            dim_linear(doc,  X1, Y1, X2, Y2, "", extract_abs(Y1,Y2) + 100,  direction="up", layer='dim')                
            dim_linear(doc,  X12, Y12, X2, Y2, "", extract_abs(Y12,Y2) + 150,  direction="up", layer='dim')                        

            # 우측 절곡치수선
            if(SW1>0):
                dim_vertical_right(doc, X2,  Y2, X3, Y3,  extract_abs(X2,X7) + 180, 'dim', text_height=0.22, text_gap=0.01)      
            dim_vertical_right(doc, X3, Y3, X4,  Y4,   extract_abs(X4,X7) +  80, 'dim', text_height=0.22, text_gap=0.01)      
            if(K1_up>0):
                dim_vertical_right(doc, X4,  Y4, X5, Y5,  extract_abs(X4,X7) + 130, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_right(doc, X6, Y6, X5,  Y5, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_right(doc,  X7, Y7, X6,  Y6, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_right(doc, X7,  Y7, X8, Y8, 120, 'dim', text_height=0.22, text_gap=0.01)  
                #Vcut A값 C값 전개치수선     
                dim_vertical_left(doc, X6, Y6, X8,  Y8,   extract_abs(X4,X7) + 80 , 'dim', text_height=0.22, text_gap=0.01)
            dim_vertical_right(doc, X2,  Y2, X8, Y8,   extract_abs(X2,X8) + 270 , 'dim', text_height=0.22, text_gap=0.01)      

            # 좌측 절곡치수선  
            if(SW1>0):                       
                dim_vertical_left(doc, X1, Y1, X16, Y16, SMH1 + 180, 'dim', text_height=0.22, text_gap=0.01)  
                dim_vertical_left(doc, X15, Y15, X16, Y16,  SMH1 + 80, 'dim', text_height=0.22, text_gap=0.01)  
            else:
                dim_vertical_left(doc, X1, Y1, X15, Y15, SMH1 + 80, 'dim', text_height=0.22, text_gap=0.01)                  
            if(K1_up>0):
                dim_vertical_left(doc, X15, Y15, X14, Y14, SMH1 + 130, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X11, Y11, X14, Y14, 80, 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_left(doc, X11, Y11, X10, Y10,  80, 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_left(doc, X9, Y9, X10, Y10, 160, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X1, Y1, X9,  Y9,  SMH1 + 270 , 'dim', text_height=0.22, text_gap=0.01) 

            if(B1>0):
                #B1 치수선     
                dim_vertical_left(doc, X13, Y13, X13,  Y11,   70 , 'dim', text_height=0.22, text_gap=0.01)      
                # JD 접선치수선
                dim_vertical_right(doc, X13, Y13, X15,  Y15,   50 , 'dim', text_height=0.22, text_gap=0.01)      

            # 기둥 하부 치수선 2    
            dim_linear(doc,   X13, Y13, X8, Y8, "", B1+ C1 + A1 + 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   
            dim_linear(doc,   X9, Y9, X8, Y8, "", 150,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   

            if LH1>LH2:
                dim_linear(doc,  X2, Y2, X6, Y6, "",  150 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)                   
            if LH1<LH2:    
                dim_linear(doc,  X4, Y4, X8, Y8, "", extract_abs(Y4, Y8) + 150 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            if args.opt7:
                drawcircle(doc, X1 + 12.5, Y1 - 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X2 - 12.5, Y2 - 5, 2.5 , layer='레이져') # 도장홀 5파이        

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"

            Xpos = Abs_Xpos + SMH1 + 400
            Ypos = Abs_Ypos + JB1_up / 2.3
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-좌"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')           

            Xpos = Abs_Xpos + SMH1 + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-좌"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + SMH1 + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Mat.Spec : {widejamb_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + SMH1 + LH1/1.2        
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Size : {Left_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + SMH1 + LH1/1.2        
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')            

            ########################################################################################################################################################################################
            # 와이드 좌기둥 좌측 단면도    
            ########################################################################################################################################################################################
            Section_Xpos = Abs_Xpos - 530
            Section_Ypos = Abs_Ypos + JB1_up + K1_up

            # 가정: SWAngle은 도 단위로 주어진 각도
            SWAngle_radians = math.radians(SWAngle)

            if(SWAngle>0):
                X1 = Section_Xpos - SW1 - SW2 * math.cos(SWAngle_radians)
                Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                X2 = Section_Xpos - SW1
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos
                
                line(doc, X2, Y2, X1, Y1, layer='0')               
                line(doc, X2, Y2, X3, Y3, layer='0')        
                dim_linear(doc, X1, Y1, X2, Y2,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # 각도 90도가 아니면 각도 표기
                if (SWAngle != 90):               
                    dim_angular(doc, X2 , Y2, X1, Y1, X2-30, Y2, X3-30, Y3,  200, direction="left" )
                    # help(ezdxf.layouts.Modelspace.add_angular_dim_2l)
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)   

            if(SWAngle<1):
                X1 = Section_Xpos - SW
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos            
                line(doc,  X1, Y1, X2, Y2, layer='0')                                                   
                dim_linear(doc, X1, Y1, X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

            if(K1_up>5):
                X4 = X3
                Y4 = Y3 - K1_up
                lineto(doc,  X4, Y4, layer='0')       
            else:
                X4 = X3
                Y4 = Y3

            # Angle은 도 단위로 주어진 각도
            Angle_radians = math.radians(Angle)

            X5 = X4 - JB1_up * math.sin(Angle_radians)                
            Y5 = Y4 - JB1_up * math.cos(Angle_radians)  
            lineto(doc,  X5, Y5, layer='0')    
            X6 = X5 - C1
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='0')   

            if(A1>0):
                X7 = X6
                Y7 = Y6 + A1
                lineto(doc,  X7, Y7, layer='0') 
            #     drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
            #     drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
            #     dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut 2개소", layer="dim", style='0')
            #     # A 치수
                dim_vertical_left(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
            # if(A1<1):
            #     X7 = X6
            #     Y7 = Y6            
            #     drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
            #     dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut", layer="dim", style='0')

            # C값 치수선
            dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            if(K1_up>5):
                dim_vertical_right(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                # if (Angle > 1):   # 직각이 아닌경우                        
                dim_angular(doc, X4, Y4, X5, Y5, X4, Y4, X3, Y3,  30, direction="left" )

            # JB사선 치수선
            dim_linear(doc, X4, Y4, X5, Y5,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
            # if (Angle > 1):   # 직각이 아닌경우            
            dim_angular(doc, X6, Y6, X5, Y5, X5, Y5, X4, Y4,  20, direction="left" )

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB1_up != JB1_down ):
                rectangle(doc, X5+60, Y5 + K1_up + (JB1_up)/2 - 75 , X5 + 60 + 100, Y5 + K1_up + (JB1_up)/2 + 75 , layer='0')
                rectangle(doc, X5+60+350, Y5 + K1_up + (JB1_up)/2 - 75, X5 + 60 + 100 + 350, Y5 + K1_up + (JB1_up)/2 + 75 , layer='0')
            if(K1_up != K1_down ):
                rectangle(doc, X4 + 50, Y4 - 10 , X4 + 50 + 100, Y4 + K1_up + 10 , layer='0')                         
                rectangle(doc, X4 + 50 + 300, Y4 - 10 , X4 + 50 + 100 + 300, Y4 + K1_up + 10 , layer='0')                         
                            
            ####################################################################################################################################################################################
            # 와이드 좌기둥 우측 단면도    
            ####################################################################################################################################################################################
            if( JB1_up != JB1_down or K1_up != K1_down ):      
                # print ('기둥좌측 하단 단면도 작도')  
                Section_Xpos = Abs_Xpos + SMH1 + LH1 + 430
                Section_Ypos = Abs_Ypos + JB1_up + K1_up

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(SWAngle>0):  
                    X1 = Section_Xpos + SW1 + SW2 * math.cos(SWAngle_radians)
                    Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                    X2 = Section_Xpos + SW1
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc, X2, Y2, X1, Y1, layer='0')                   
                    line(doc, X2, Y2, X3, Y3, layer='0')        
                    dim_linear(doc, X2, Y2, X1, Y1,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # 각도 90도가 아니면 각도 표기
                    if (SWAngle != 90):               
                        dim_angular(doc,  X3+5, Y3, X2+5, Y2, X1, Y1, X2 , Y2,  200, direction="right" )                    
                    dim_linear(doc, X3, Y3,  X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

                if(SWAngle<1):  
                    X1 = Section_Xpos + SW
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos 
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc,   X1, Y1, X2, Y2,  layer='0')                                   
                    dim_linear(doc,  X2, Y2, X1, Y1,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

                if(K1_down>5):
                    X4 = X3
                    Y4 = Y3 - K1_down
                    lineto(doc,  X4, Y4, layer='0')       
                else:
                    X4 = X3
                    Y4 = Y3

                # Angle은 도 단위로 주어진 각도
                Angle_radians = math.radians(Angle)

                X5 = X4 + JB1_down * math.sin(Angle_radians)                
                Y5 = Y4 - JB1_down * math.cos(Angle_radians)  
                lineto(doc,  X5, Y5, layer='0')    
                X6 = X5 + C1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')   

                if(A1>0):
                    X7 = X6
                    Y7 = Y6 + A1
                    lineto(doc,  X7, Y7, layer='0') 
                #     drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                #     drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                #     dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut 2개소", layer="dim", style='0')
                #     # A 치수
                    dim_vertical_right(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
                # if(A1<1):
                #     X7 = X6
                #     Y7 = Y6            
                #     drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                #     dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut", layer="dim", style='0')

                # C값 치수선
                dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                if(K1_up>5):
                    dim_vertical_left(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                    # if (Angle > 1):   # 직각이 아닌경우                        
                    dim_angular(doc,  X3, Y3, X4 , Y4,  X5, Y5, X4, Y4,   15, direction="right" )

                # JB사선 치수선
                dim_linear(doc,  X5, Y5, X4, Y4,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
                # if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc,  X5, Y5, X4, Y4, X5, Y5, X6 , Y6,  15, direction="right" )  

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB1_up != JB1_down ):
                    rectangle(doc, X5-60,Y5 + (JB1_up+K1_up)/2 - 75 , X5 - 60 - 100,Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')               
                    rectangle(doc, X5 - 60 - 300, Y5 + (JB1_up+K1_up)/2 - 75 , X5 - 60 - 100 - 300, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')   

                if(K1_up != K1_down ):
                    rectangle(doc, X4 - 50, Y4 - 10 , X4 - 50 - 100, Y4 + K1_down + 10 , layer='0')                             
                    rectangle(doc, X4 - 50 - 220, Y4 - 10 , X4 - 50 - 100 - 220, Y4 + K1_down + 10 , layer='0')                             

            #########################################################################################################################################################################
            # 와이드 우기둥 전개도 및 단면도
            # 주의 사항 기둥의 우측은 X1의 선이 끊어지는 현상을 방지하기 위해서 좌기둥과 다르게 설계해야 함  (와이드형태만 해당됨)
            #########################################################################################################################################################################
                    
            Abs_Ypos -= 2000 
            R_Xpos = Abs_Xpos + SMH2
            R_Ypos = Abs_Ypos + 1100 - (C1 + C2 + A1 + A2)

            if(A2>0):            
                #  X1은 좌측 끝단으로 이동해서 점을 하나 다르게 해야 함
                X1 = R_Xpos  - SMH2
                Y1 = R_Ypos + C2 + A2 - Vcut_rate * 3
                X2 = R_Xpos + RH1 
                Y2 = Y1
                line(doc, X1, Y1, X2, Y2, layer='레이져')                    
                X3 = X2
                Y3 = R_Ypos + C2 - Vcut_rate*2
                lineto(doc,  X3, Y3, layer='레이져')            

            if(A2<1):
                X1 = R_Xpos  - SMH2
                Y1 = R_Ypos + C2  - Vcut_rate 
                X2 = R_Xpos + RH1 
                Y2 = Y1
                line(doc, X1, Y1, X2, Y2, layer='레이져')                    
                X3 = X2
                Y3 = Y2

            X4 = X3
            Y4 = R_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')

            X5 = R_Xpos + RH2 

            if (SWAngle>1):  #SWAngle 0 이상인 경우    헤밍구조인 경우        
                if(K2_up>0):   
                    if (SWAngle == 90):    
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = X6
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = R_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')
                    else:   
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = X6
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = R_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')

                if(K2_up<1):   #K값이 없는 경우
                    if (SWAngle == 90):   
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = X6
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11      
                    else:
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = X6
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = X7
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = R_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = R_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = R_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11

            if (SWAngle<1):  #SWAngle 0     일반구조
                if(K2_up>0):   
                    Y5 = R_Ypos - JB2_down +  Vcut_rate
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = X5
                    Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                    lineto(doc,  X6, Y6, layer='레이져')
                    X7 = X6
                    Y7 = R_Ypos - JB2_down - K2_down - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = R_Xpos 
                    Y9 = R_Ypos - JB2_up - K2_up - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = R_Xpos 
                    Y10 = Y9                
                    X11 = R_Xpos 
                    Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = R_Xpos 
                    Y12 = R_Ypos - JB2_up +  Vcut_rate
                    lineto(doc,  X12, Y12, layer='레이져')
                if(K2_up<1):   #K값이 없는 경우
                    Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = X5
                    Y6 = Y5                
                    X7 = X6
                    Y7 = R_Ypos - JB2_down  - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = R_Xpos 
                    Y9 = R_Ypos - JB2_up  - SW  + Bending_rate * 2 + Vcut_rate
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = X9
                    Y10 = Y9                
                    X11 = R_Xpos 
                    Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = X11
                    Y12 = Y11                
            X13 = R_Xpos 
            Y13 = R_Ypos - B2
            lineto(doc,  X13, Y13, layer='레이져')
            X14 = R_Xpos - SMH2
            Y14 = Y13
            lineto(doc,  X14, Y14, layer='레이져')
            X15 = X14
            Y15 = R_Ypos
            lineto(doc,  X15, Y15, layer='레이져')
            if(A2>0):
                X16 = X15
                Y16 = R_Ypos + C2 - Vcut_rate*2
                lineto(doc,  X16, Y16, layer='레이져')
                X17 = X16
                Y17 = R_Ypos + C2 + A2 - Vcut_rate*3
                lineto(doc,  X17, Y17, layer='레이져')
            if(A2<1):
                X16 = X15 
                Y16 = R_Ypos + C2 - Vcut_rate
                X17 = X16
                Y17 = Y16
                lineto(doc,  X16, Y16, layer='레이져')
 
            # 절곡선1
            if(A2>1):        
                line(doc, X16, Y16, X3, Y3,  layer='22')
            # C2                
            line(doc, X15, Y15, X4, Y4,  layer='22')
            # K2
            if(K2_up>0):
                line(doc, X12, Y12, X5, Y5,  layer='22')
            line(doc, X11, Y11, X6, Y6,  layer='22')
            # 해밍구조
            if (SWAngle>0):   
                line(doc, X10, Y10, X7, Y7,  layer='22')

            # 하부 전개도 치수선
            dim_linear(doc,  X14, Y14, X9, Y9, "",  extract_abs(Y14,Y9) + 100,  direction="down", layer='dim')                
            dim_linear(doc,  X9, Y9, X8, Y8, "", 100,  direction="down", layer='dim')                
            dim_linear(doc,  X14, Y14, X8, Y8, "", extract_abs(Y14,Y9) + 150,  direction="down", layer='dim')                

            # 우측 절곡치수선
            if(A2 > 1):  
                dim_vertical_right(doc, X3, Y3, X2,  Y2, 130, 'dim', text_height=0.22,  text_gap=0.07)      
                #Vcut A값 C값 전개치수선     
                dim_vertical_left(doc, X2, Y2, X4,  Y4,   80 , 'dim', text_height=0.22, text_gap=0.01)                     
            if(C2 > 1):      
                dim_vertical_right(doc, X3, Y3, X4,  Y4, 80, 'dim', text_height=0.22,  text_gap=0.07)                      
            # JB    
            dim_vertical_right(doc, X5, Y5, X4,  Y4, extract_abs(X5,X4) + 80, 'dim', text_height=0.22,  text_gap=0.07) 
            if(K2_down > 0): 
                dim_vertical_right(doc, X6, Y6, X5,  Y5, extract_abs(X8,X5) + 130, 'dim', text_height=0.22,  text_gap=0.07)  
            # 뒷날개    
            dim_vertical_right(doc, X6, Y6, X7,  Y7, extract_abs(X5,X8) + 80, 'dim', text_height=0.22,  text_gap=0.07)  
            if (SWAngle > 0):   
                dim_vertical_right(doc, X7, Y7, X8,  Y8, extract_abs(X7,X8) + 180, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X2, Y2, X8,  Y8, extract_abs(X2,X8) + 230, 'dim', text_height=0.22,  text_gap=0.07)  

            # 좌측 절곡치수선
            if(A2 > 1):          
                dim_vertical_left(doc, X17, Y17, X16, Y16,  130, 'dim', text_height=0.22, text_gap=0.01)  
            if(C2 > 1):              
                dim_vertical_left(doc, X15, Y15, X16, Y16,  60, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X15, Y15, X12, Y12,  60, 'dim', text_height=0.22, text_gap=0.01)  
            if(K2_up > 0): 
                dim_vertical_left(doc, X12, Y12, X11, Y11,  SMH2 + 120, 'dim', text_height=0.22, text_gap=0.01)              
            # 뒷날개    
            dim_vertical_left(doc, X11, Y11, X10, Y10,  SMH2 + 60, 'dim', text_height=0.22, text_gap=0.01)  
            if (SWAngle > 0):               
                dim_vertical_left(doc, X10, Y10, X9, Y9,  SMH2 + 180, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X17, Y17, X9,  Y9,  240 , 'dim', text_height=0.22, text_gap=0.01)      

            # 전개도 상부 치수선    
            dim_linear(doc,  X17, Y17, X2, Y2,  "",   150,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)                      
            dim_linear(doc,  X13, Y13, X2, Y2,  "",   extract_abs(Y1,Y13) + 100,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)                      
            if RH1<RH2:
                dim_linear(doc,  X2, Y2, X5, Y5,  "",  150 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)       
            if RH1>RH2:    
                dim_linear(doc,  X8, Y8, X4, Y4, "",   150 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.01)          
            if(B2>0):
                dim_vertical_left(doc, X13, Y15, X13, Y13,  80, 'dim', text_height=0.22, text_gap=0.01)              
                dim_vertical_right(doc, X11, Y11, X13, Y13,  50, 'dim', text_height=0.22, text_gap=0.01)              
            if args.opt7:
                drawcircle(doc, X9 + 12.5, Y9 + 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X8 - 12.5, Y8 + 5, 2.5 , layer='레이져') # 도장홀 5파이            

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"       

            Xpos = R_Xpos + 400 
            Ypos = R_Ypos - JB2_up / 2
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-우"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')                       

            Xpos = R_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-우"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = R_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Mat.Spec : {widejamb_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = R_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Size : {Right_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = R_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')      

            # 좌우 도면틀 넣기
            BasicXscale = 4487
            BasicYscale = 2770
            TargetXscale = 800 + max(SMH1, SMH2) + max(LH1, LH2) + 900
            TargetYscale = 400*4 + (JB1_up + C1 + A1 + K1_up + SW) * 2
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale

            # print (f'스케일 : {frame_scale}')
            insert_frame(X1 -  950 ,Y1 - (JB2_up+C2+A2+K2_up+SW + 500 + C1 + A1 + 300) , frame_scale, "side", workplace)              
            
            ###################################
            # 와이드 우기둥 좌측 단면도    
            ###################################            
            Section_Xpos = R_Xpos - SMH2 - 500
            Section_Ypos = R_Ypos

            SWAngle_radians = math.radians(SWAngle)

            if(A2>0):
                X1 = Section_Xpos
                Y1 = Section_Ypos - A2
                X2 = Section_Xpos 
                Y2 = Section_Ypos 

                line(doc, X1, Y1, X2, Y2, layer='0')                               
                dim_vertical_left(doc, X1, Y1, X2, Y2, 50, 'dim', text_height=0.22, text_gap=0.07)  
                X3 = X2 + C2
                Y3 = Y2
                lineto(doc,  X3, Y3, layer='0')   
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # vcut 표기                
                # drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                # dim_leader_line(doc, X3, Y3 , X3+50, Y3+100, "V-cut 2개소", layer="dim", style='0')         

            if(A2<1):
                X1 = Section_Xpos
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos             
                X3 = X2 + C2
                Y3 = Y2
                line(doc, X1, Y1, X3, Y3, layer='0')               
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # # vcut 표기                            
                # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                # dim_leader_line(doc, X3, Y3 , X3+50, Y3+100, "V-cut", layer="dim", style='0')                                        

            X4 = X3 + JB2_up * math.sin(Angle_radians)                
            Y4 = Y3 - JB2_up * math.cos(Angle_radians)        
            lineto(doc,  X4, Y4, layer='0')	
            dim_linear(doc, X3, Y3, X4, Y4,  "", 50,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)     

            if(K2_up > 5):
                X5 = X4
                Y5 = Y4 - K2_up
                lineto(doc,  X5, Y5, layer='0')	
                dim_vertical_right(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
            else:
                X5 = X4
                Y5 = Y4
                
            if(SWAngle>0):
                X6 = X5 - SW1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                X7 = X6 - SW2 * math.cos(SWAngle_radians)
                Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                lineto(doc,  X7, Y7, layer='0')
                dim_linear(doc, X6, Y6, X7, Y7,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                dim_angular(doc, X5-5, Y5, X6-5, Y6,  X7, Y7, X6 , Y6,  100, direction="left" )

            if(SWAngle<1): # 헤밍구조가 아닌경우 일반형
                X6 = X5 - SW
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                X7 = X6
                Y7 = Y6                        

            # if (Angle > 1):   # 직각이 아닌경우            
            dim_angular( doc,   X3, Y3, X4, Y4, X3, Y3, X2 , Y2,    15, direction="left" )		                            
            if(K2_up > 5): # K2값
                dim_angular(doc,  X4, Y4, X5 , Y5, X4, Y4, X3, Y3,  5, direction="left" )    

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB2_up != JB2_down ):
                rectangle(doc, X3+60,  Y5 + (JB1_up+K1_up)/2 - 75 , X3 + 60 + 100, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')               
                rectangle(doc, X3+60+300, Y5 + (JB1_up+K1_up)/2 - 75 , X3 + 60 + 400,  Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')     

            if(K2_up != K2_down ):
                rectangle(doc, X5 + 50, Y5 - 10 , X5 + 50 + 100, Y5 + K2_up + 10 , layer='0')                         
                rectangle(doc, X5 + 60 + 250, Y5 - 10 , X5 + 60 + 350, Y5 + K2_up + 10 , layer='0')                            

            ###################################
            # 와이드 우기둥 우측 단면도    
            ###################################
                    
            if( JB2_up != JB2_down or K2_up != K2_down ):   

                Section_Xpos = R_Xpos + max(RH1,RH2) +  500
                Section_Ypos = R_Ypos

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(A2>0):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos - A2
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos 

                    line(doc, X1, Y1, X2, Y2, layer='0')
                    dim_vertical_right(doc, X2, Y2, X1, Y1,  50, 'dim', text_height=0.22, text_gap=0.07)
                    X3 = Section_Xpos
                    Y3 = Section_Ypos
                    lineto(doc,  X3, Y3, layer='0')
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # vcut 표기                
                    # drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    # dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut 2개소", layer="dim", style='0')                                
                if(A2<1):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos                 
                    X3 = Section_Xpos
                    Y3 = Section_Ypos 
                    line(doc, X2, Y2, X3, Y3, layer='0')                               
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # vcut 표기                                
                    # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    # dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut", layer="dim", style='0')                                

                X4 = X3 - JB2_down * math.sin(Angle_radians)                
                Y4 = Y3 - JB2_down * math.cos(Angle_radians)        
                lineto(doc,  X4, Y4, layer='0')	
                dim_linear(doc, X4, Y4, X3, Y3,   "", 50,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)

                if(K2_down > 5):
                    X5 = X4
                    Y5 = Y4 - K2_down
                    lineto(doc,  X5, Y5, layer='0')	
                    dim_vertical_left(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
                else:
                    X5 = X4
                    Y5 = Y4
                    lineto(doc,  X5, Y5, layer='0')	

                if(SWAngle>0):
                    X6 = X5 + SW1
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                    X7 = X6 + SW2 * math.cos(SWAngle_radians)
                    Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                    lineto(doc,  X7, Y7, layer='0')
                    dim_linear(doc, X7, Y7, X6, Y6,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                    dim_angular(doc,  X7 , Y7, X6, Y6, X6 + 5 , Y6, X6, Y6,  150, direction="right" )
                if(SWAngle<1):
                    X6 = X5 + SW
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,"" , 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    X7 = X6
                    Y7 = Y6

                # if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc,  X3, Y3, X2-8 , Y2,  X3, Y3, X4, Y4,   25, direction="right" )		            
                    # if(K2_down>0):
                dim_angular(doc,   X4, Y4, X3 , Y3, X5, Y5,  X4, Y4, 30, direction="right" )    

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB2_up != JB2_down ):
                    rectangle(doc, X3 - 60, Y5 + (JB1_up+K1_up)/2 - 75 , X3 - 60 - 100, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')               
                    rectangle(doc, X3 - 60 - 200, Y5 + (JB1_up+K1_up)/2 - 75 , X3 - 60 - 100 - 200, Y5 + (JB1_up+K1_up)/2 + 75 , layer='0')      

                if(K2_up != K2_down ):
                    rectangle(doc, X5 - 50, Y5 - 10 , X5 - 50 - 100, Y5 + K2_down + 10 , layer='0')                             
                    rectangle(doc, X5 - 50 - 250, Y5 - 10 , X5 - 50 - 100 - 250, Y5 + K2_down + 10 , layer='0')                                        

            Abs_Ypos -= 2000    

##########################################################################################################################################
# 멍텅구리 작도 기본
##########################################################################################################################################
def execute_normal():     

    # 시트 선택 (시트명을 지정)
    sheet_name = '멍텅제작'  # 원하는 시트명으로 변경
    sheet = workbook[sheet_name]

    # 2차원 배열을 저장할 리스트 생성
    data_2d = []

    # 1행부터 20행까지의 값을 2차원 배열에 저장
    for row_num in range(7, sheet.max_row + 1):  # 7행부터 20행까지 반복
        cell_value = sheet.cell(row=row_num, column=2).value  # 2열의 값 가져오기
        if cell_value is not None and cell_value != 'F':  # 값이 None이 아닌 경우에만 배열에 추가
            row_values =  [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, sheet.max_column + 1)]        
            data_2d.append(row_values)

    # 엑셀 파일 닫기
    workbook.close()

    if not data_2d:
        print("해당 데이터가 없습니다.")
        return  # 함수를 빠져나감    

    ########################################################################################################################################################################
    # 멍텅구리 상판 전개도
    ########################################################################################################################################################################

    Abs_Xpos = 13000
    Abs_Ypos = 0

    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        OP = row[2]
        R = row[3]
        JD1 = row[4]
        JD2 = row[5]
        JB1_up = row[6]
        JB1_down = row[7]
        JB2_up = row[8]
        JB2_down = row[9]
        H1 = row[10]
        H2 = row[11]
        C1 = row[12]        
        C2 = row[13]
        A1 = row[14]
        A2 = row[15]
        LH1 = row[16]
        LH2 = row[17]
        RH1 = row[18]
        RH2 = row[19]        
        U = row[20]        
        G = row[21]        
        UW = row[22]        
        SW = row[23]        
        K1_up = row[24]        
        K1_down = row[25]        
        K2_up = row[26]        
        K2_down = row[27]                
        Upper_size = row[28]
        Left_side_size = row[29]
        Right_side_size = row[30]
        SW1 = row[31]
        SW2 = row[32]
        SWAngle = row[33]
        E1 = row[34]
        E2 = row[35]
        G_Setvalue = row[52]  # 손상민 소장 G값 강제로 적용 작지 52번째
        Angle = row[39]
        Top_pop1 = round(row[40] + 2, 1) #소수점 첫째자리까지만 나오게
        Top_pop2 = round(row[41] + 2, 1)         
        Kadapt =  row[51]   # K값 J값에 적용여부 넣기 1이면 넣기

        # 상판 E1, E2는 추영덕 소장이 K값을 상판에 표현하는 수치표현식인데, K값으로 환산해서 정리함
        #  K1_up = E1 - UW - 1.2

        # 데이터 검증 및 기본값 설정
        SWAngle = validate_or_default(SWAngle)
        K1_up = validate_or_default(K1_up)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_down = validate_or_default(K2_down)
        U = validate_or_default(U)
        G = validate_or_default(G)        
        G_Setvalue = validate_or_default(G_Setvalue)        
        OP = validate_or_default(OP)
        Kadapt = validate_or_default(Kadapt)


        # print("작업소장 : ")
        # print(worker)
        # print("G_setvalue : ")
        # print(G_Setvalue)        
        if(worker=="손상민" and G_Setvalue > 0) :
            # 상부 돌출
            Top_pop1 = G_Setvalue
            Top_pop2 = G_Setvalue
            # print("G_setvalue : ")
            # print(G_Setvalue)
       
        if(Kadapt==1):
            K1_up = 0            
            K2_up = 0    
            if(E1>0):
                E1 = E1 + Vcut_rate
                E2 = E2 + Vcut_rate
        else:
            if(E1>0):                
                K1_up = E1 
                K2_up = E2                        
             
        if(K1_up>=10 and K2_up>=10 ):
            # 추영덕소장 E1, E2값 무시함 위의 조건이라면
            E1 = 0
            E2 = 0

        if(Top_pop1<4):
            Top_pop1 = 4
        if(Top_pop2<4):
            Top_pop2 = 4                

        if(OP>0):             
            if(G > 0):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U + G - Vcut_rate*3
                X2 = Abs_Xpos + H1 + H2 + R 
                Y2 = Y1      
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Abs_Ypos + U - Vcut_rate * 2           
                line(doc, X2, Y2, X3, Y3, layer='레이져')          
            if(G < 1):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U  - Vcut_rate
                X2 = Abs_Xpos + H1 + H2 + R 
                Y2 = Y1      
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Y2

            X4 = X3
            Y4 = Abs_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')
            X5 = X4
            Y5 = Abs_Ypos - Top_pop2 + Vcut_rate        
            lineto(doc,  X5, Y5, layer='레이져')    
            X6 = X5 - H2
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='레이져')

            if(K1_up>0 or K2_up>0):
                X7 = Abs_Xpos + H1 + (R-OP)/2 + OP                
                Y7 = Abs_Ypos - JD2 + Vcut_rate
                lineto(doc,  X7, Y7, layer='레이져')
            else:
                if(E1>0):
                    X7 = Abs_Xpos + H1 + (R-OP)/2 + OP    
                    Y7 = Abs_Ypos - (JD2-E2) + Vcut_rate * 2
                    lineto(doc,  X7, Y7, layer='레이져')  
                else:
                    X7 = Abs_Xpos + H1 + (R-OP)/2 + OP    
                    Y7 = Abs_Ypos - JD2 + Vcut_rate * 2
                    lineto(doc,  X7, Y7, layer='레이져')  
                
            if(K1_up>0 or K2_up>0 or (Kadapt==1 and E1>0)):
                X8 = X7
                Y8 = Abs_Ypos - JD2 - K2_up + Vcut_rate*2
                lineto(doc,  X8, Y8, layer='레이져') 
                X9 = X8 
                Y9 = Abs_Ypos - JD2 - K2_up - UW + Vcut_rate*3     
                lineto(doc,  X9, Y9, layer='레이져')   
                X10 = Abs_Xpos + H1 + (R-OP)/2 
                Y10 = Abs_Ypos - JD1 - K1_up - UW + Vcut_rate*3
                lineto(doc,  X10, Y10, layer='레이져')      
                X11 = X10 
                Y11 =  Abs_Ypos - JD1 - K1_up + Vcut_rate*2
                lineto(doc,  X11, Y11, layer='레이져')                                      
            else:
                X8 = X7
                Y8 = Y7
                X9 = X8
                Y9 = Abs_Ypos - JD2 - UW + Vcut_rate*3            
                lineto(doc,  X9, Y9, layer='레이져')   
                X10 = Abs_Xpos + H1 + (R-OP)/2 
                Y10 = Abs_Ypos - JD1 - UW + Vcut_rate*3
                lineto(doc,  X10, Y10, layer='레이져')      
                X11 = X10                 
                Y11 =  Abs_Ypos - JD1 + Vcut_rate*2
                lineto(doc,  X11, Y11, layer='레이져')                                

            X12 = X11     
            if(K1_up>0 or K2_up>0):
                Y12 = Abs_Ypos - JD1 + Vcut_rate 
            else:  
                if(Kadapt==1 and E1>0):
                    Y12 = Abs_Ypos - (JD1-E1) + Vcut_rate 
                else:   
                    Y12 = Abs_Ypos - JD1 + Vcut_rate * 2
            
            lineto(doc,  X12, Y12, layer='레이져')                   
            X13 = Abs_Xpos + H1
            Y13 = Abs_Ypos - Top_pop2 + Vcut_rate 
            lineto(doc,  X13, Y13, layer='레이져')      
            X14 = Abs_Xpos
            Y14 = Y13
            lineto(doc,  X14, Y14, layer='레이져')      
            X15 = X14 
            Y15 = Abs_Ypos
            lineto(doc,  X15, Y15, layer='레이져')     
            if(G > 0):         
                X16 = X15 
                Y16 = Abs_Ypos + U - Vcut_rate * 2   
                lineto(doc,  X16, Y16, layer='레이져')      
                lineto(doc,  X1, Y1, layer='레이져')                  
            if(G < 1):         
                X16 = X15 
                Y16 = Abs_Ypos + U - Vcut_rate 
                lineto(doc,  X1, Y1, layer='레이져')                  

            if args.opt5:
                drawcircle(doc, X10 + 12.5, Y10 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X9 - 12.5, Y9 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이

            # 상판 절곡선
            if(G>0):
                line(doc, X16, Y16, X3, Y3,  layer='22')        
            line(doc, X15, Y15, X4, Y4,  layer='22')
            line(doc, X11, Y11, X8, Y8,  layer='22')
    
            # 상판 우측 절곡치수선
            if(G>0):        
                dim_vertical_right(doc, X2, Y2, X3, Y3,  100, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X4, Y4, X3, Y3,  50, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X4, Y4, X8, Y8,  50, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc,  X9, Y9, X8, Y8, extract_abs(X8,X5) + 50, 'dim', text_height=0.22, text_gap=0.07)   # 뒷날개
            dim_vertical_right(doc, X2, Y2,  X9, Y9,  extract_abs(X9,X2) + 100, 'dim', text_height=0.22, text_gap=0.07)             

            # 상판 좌측 절곡치수선 (Vcut자리)
            if(JD1 != JD2 or K1_up != K2_up):
                if(G>0):                    
                    dim_vertical_left(doc, X1, Y1, X16, Y16,  extract_abs(X1,X16) + 100, 'dim', text_height=0.22, text_gap=0.07)  
                dim_vertical_left(doc, X15, Y15, X16, Y16,  extract_abs(X15,X16) + 50, 'dim', text_height=0.22, text_gap=0.07)                              
                dim_vertical_left(doc, X11, Y11, X15, Y15,  extract_abs(X11,X15) + 50 , 'dim', text_height=0.22, text_gap=0.07)              
                dim_vertical_left(doc, X10, Y10, X11, Y11,  H1 + (R-OP)/2 + extract_abs(X10,X11) + 50 , 'dim', text_height=0.22, text_gap=0.07)              
                dim_vertical_left(doc, X10, Y10, X1, Y1,  extract_abs(X10,X1) + 200, 'dim', text_height=0.22, text_gap=0.07)                          
            else:    
                dim_vertical_left(doc, X10, Y10, X11, Y11,   H1 + (R-OP)/2  + 10, 'dim', text_height=0.22,  text_gap=0.07) 
                dim_vertical_left(doc, X10, Y10, X15, Y15,   H1 + (R-OP)/2  + 100 , 'dim', text_height=0.22,  text_gap=0.07) 
                dim_vertical_left(doc, X10, Y10, X16, Y16,   H1 + (R-OP)/2  + 150, 'dim', text_height=0.22,  text_gap=0.07) 

            # 상판 하부 치수선 3개
            dim_linear(doc,  X9, Y9, X6, Y6, "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X10, Y10, X9, Y9,  "", 100 + extract_abs(JD1+K1_up, JD2+K2_up),  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            dim_linear(doc,  X13, Y13, X10, Y10,  "",  extract_abs(Y6,Y9) + 100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            # 상판 돌출부위 3.4 치수선
            dim_vertical_right(doc, X15, Y15, X13, Y13, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_left(doc, X4, Y4, X6, Y6, 100, 'dim', text_height=0.22,  text_gap=0.07)  

            # 기둥과의 접선 부위 aligned 표현 각도표현
            dim_linear(doc, X13, Y13, X11, Y11,   "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) # 좌측
            dim_angular(doc, X12, Y12, X13, Y13, X11, Y11, X12, Y12 + 10,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다        
            dim_linear(doc, X8, Y8, X6, Y6,  "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) # 좌측
            dim_angular(doc, X8, Y8  , X7, Y7 + 10,  X7, Y7,  X6, Y6,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다

            # 상판 K1 K2 값 부분 치수선 
            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                dim_vertical_right(doc, X11, Y11, X12, Y12, 80, 'dim', text_height=0.22,  text_gap=0.07)  
                dim_vertical_left(doc, X8, Y8, X7, Y7, 80, 'dim', text_height=0.22,  text_gap=0.07)  

            # 전개도 상부 치수선
            dim_linear(doc,  X1, Y1, X2, Y2, "", 120,  direction="up", layer='dim')
            dim_linear(doc,  X13, Y13, X6, Y6,  "", extract_abs(Y1,Y13) + 50,  direction="up", layer='dim')
            dim_linear(doc,  X1, Y1, X13, Y13, "",  50,  direction="up", layer='dim')
            dim_linear(doc,  X6, Y6, X2, Y2,  "",  extract_abs(Y2,Y6) + 50,  direction="up", layer='dim')

            # 상부 Vcut 치수 표기
            if(U>0 and G>0):   
               dim_vertical_left(doc, X1+H1+OP/2, Y1, X1+H1+OP/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  

            # 상판 좌우가 다를때 Vcut 상부에 치수선 표기
            # if((JD1 != JD2 or K1_up != K2_up) and G>0):
            #     dim_vertical_left(doc, X1 + R/2, Y1, X1 + R/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            # 문구 설정
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"        
            Textstr = f"({Pre_text}{str(Floor_des)})"       

            X2 = Abs_Xpos + (H1 + R + H2)/2 - len(Textstr)*50/2
            Y2 = Abs_Ypos - JD1/1.5
            draw_Text(doc, X2, Y2, 50, str(Textstr), '레이져')

            # 문구            
            sizedes = Upper_size                  
            
            Textstr = f"Part Name : 상판({Pre_text}{str(Floor_des)})"    
            draw_Text(doc, X9+150, Y9 + 50 - 220, 28, str(Textstr), '0')
            Textstr = f"Mat.Spec : {normal_material}"    
            draw_Text(doc, X9+150, Y9 - 220, 28, str(Textstr), '0')
            Textstr = f"Size : {sizedes}"    
            draw_Text(doc, X9+150, Y9 - 50 - 220 , 28, str(Textstr), '0')
            Textstr = f"Quantity : 1 EA"    
            draw_Text(doc, X9+150, Y9 - 100 - 220 , 28, str(Textstr), '0')                        

            # 도면틀 넣기
            BasicXscale = 2560
            BasicYscale = 1550
            TargetXscale = 800 + R + U * 2 + 200
            TargetYscale = 300 + JD1 + G + U + UW
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            insert_frame(X1 - (U + 580) ,Y1- ( 700 + JD1 + K1_up + G + U + UW) , frame_scale, "top_normal", workplace)        

            ###########################################################################################################################################################################
            # 멍텅구리 상판 좌측 단면도    (좌우가 다른 모양일때 단면도를 그려준다)
            ###########################################################################################################################################################################
            if(JD1 != JD2 or K1_up != K2_up):
                Section_Xpos = Abs_Xpos - H1 - 400           

                if(G > 0):
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos - G
                    X2 = Section_Xpos
                    Y2 = Abs_Ypos  

                    line(doc, X1, Y1, X2, Y2, layer='0')                    
                    dim_vertical_left(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)     
                    X3 = Section_Xpos + U
                    Y3 = Abs_Ypos  
                    lineto(doc,  X3, Y3, layer='0')    
                    dim_linear(doc,  X2, Y2, X3, Y3, "", 50,  direction="up", layer='dim')                
                    # drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                    # dim_leader_line(doc, X3, Y3 , X2-130 , Y2-100, "V-cut 3개", layer="dim", style='0')                
                if(G < 1):
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos 
                    X2 = Section_Xpos
                    Y2 = Abs_Ypos                  
                    X3 = Section_Xpos + U
                    Y3 = Abs_Ypos   
                    line(doc, X2, Y2, X3, Y3, layer='0')                  
                    dim_linear(doc,  X2, Y2, X3, Y3, "", 50,  direction="up", layer='dim')    
                    # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                    # dim_leader_line(doc, X3, Y3 , X2-130 , Y2-100, "V-cut 2개", layer="dim", style='0')                                
                
                X4 = X3 
                Y4 = Y3 - JD1
                lineto(doc,  X4, Y4, layer='0')                   
                if(K1_up > 0):
                    X5 = X4 
                    Y5 = Y4 - K1_up
                    lineto(doc,  X5, Y5, layer='0')   
                    dim_vertical_right(doc,  X3, Y3, X4, Y4, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X4, Y4, X5, Y5, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X3, Y3, X5, Y5, 110,'dim' ,text_height=0.22,  text_gap=0.07)                                     
                else:
                    dim_vertical_right(doc,  X3, Y3, X4, Y4, 110,'dim' ,text_height=0.22,  text_gap=0.07)     
                    X5 = X4 
                    Y5 = Y4

                # drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')         
                X6 = X5 - UW
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')   
                dim_linear(doc,  X6, Y6, X5, Y5,  "", 80,  direction="down", layer='dim')            
                
                if(K1_up > 0):                 
                    # K값 부분 라인 그리기
                    line(doc, X4, Y4, X4 - UW, Y4, layer='22')     
                
                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JD1 != JD2 or K1_up != K2_up ):
                    rectangle(doc, X4+50, Y5+(JD1+K1_up)/2 - 75 , X4 + 50 + 100, Y5+(JD1+K1_up)/2 + 75, layer='0')               
                    rectangle(doc, X4+285,  Y5+(JD1+K1_up)/2 - 75 , X4+285+100, Y5+(JD1+K1_up)/2 + 75 , layer='0')                                                         
                
            ################################################################################################################################################################
            # 멍텅구리 상판 우측 단면도 
            ################################################################################################################################################################
            Section_Xpos = Abs_Xpos + R + H1 + H2 + 400

            if(G>0):
                X1 = Section_Xpos 
                Y1 = Abs_Ypos - G
                X2 = Section_Xpos
                Y2 = Abs_Ypos  

                line(doc, X1, Y1, X2, Y2, layer='0')                    
                dim_vertical_right(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)     
                X3 = Section_Xpos - U
                Y3 = Abs_Ypos  
                lineto(doc,  X3, Y3, layer='0')    
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim')
                # drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                # dim_leader_line(doc, X3, Y3 , X3+50 , Y2-100, "V-cut 3개", layer="dim", style='0')                

            if(G<1):
                X1 = Section_Xpos 
                Y1 = Abs_Ypos
                X2 = Section_Xpos
                Y2 = Abs_Ypos  
                X3 = Section_Xpos - U
                Y3 = Abs_Ypos  
                line(doc, X1, Y1, X3, Y3, layer='0')                            
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim')    
                # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                # dim_leader_line(doc, X3, Y3 , X3 + 50 , Y2-100, "V-cut 2개", layer="dim", style='0')                            
                
            X4 = X3 
            Y4 = Y3 - JD2
            lineto(doc,  X4, Y4, layer='0')   
            if(K1_up > 0):
                X5 = X4 
                Y5 = Y4 - K2_up
                lineto(doc,  X5, Y5, layer='0')   
                dim_vertical_left(doc,  X3, Y3, X4, Y4, 60,'dim' ,text_height=0.22,  text_gap=0.07)                     
                dim_vertical_left(doc,  X4, Y4, X5, Y5, 60,'dim' ,text_height=0.22,  text_gap=0.07)     
                dim_vertical_left(doc,  X3, Y3, X5, Y5, 110,'dim' ,text_height=0.22,  text_gap=0.07)                                     
            else:
                dim_vertical_left(doc,  X3, Y3, X4, Y4, 110,'dim' ,text_height=0.22,  text_gap=0.07)                     
                X5 = X4 
                Y5 = Y4

            X6 = X5 + UW
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='0')   
            dim_linear(doc,  X5, Y5, X6, Y6,   "", 80,  direction="down", layer='dim')
            
            # drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
            
            if(K2_up > 0):                 
                # K값 부분 라인 그리기
                line(doc, X4, Y4, X4 + UW, Y4, layer='22')      

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JD1 != JD2 or K1_up != K2_up ):
                rectangle(doc, X4-50, Y5+(JD2+K2_up)/2 - 75 , X4 - 50 - 100, Y5+(JD2+K2_up)/2 + 75, layer='0')               
                rectangle(doc, X4-285,  Y5+(JD2+K2_up)/2 - 75 , X4-285-100,  Y5+(JD2+K2_up)/2 + 75 , layer='0')                                                         
            
            # 다음 도면 간격 계산
            Abs_Ypos -= 4000

    Abs_Xpos = 13000 + 4000
    Abs_Ypos = 500

    ############################################################################################################################################################################
    # 멍텅구리 좌기둥 전개도
    ############################################################################################################################################################################
    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        OP = row[2]
        R = row[3]
        JD1 = row[4]
        JD2 = row[5]
        JB1_up = row[6]
        JB1_down = row[7]
        JB2_up = row[8]
        JB2_down = row[9]
        H1 = row[10]
        H2 = row[11]
        C1 = row[12]        
        C2 = row[13]
        A1 = row[14]
        A2 = row[15]
        LH1 = row[16]
        LH2 = row[17]
        RH1 = row[18]
        RH2 = row[19]        
        U = row[20]        
        G = row[21]        
        UW = row[22]        
        SW = row[23]        
        K1_up = row[24]        
        K1_down = row[25]        
        K2_up = row[26]        
        K2_down = row[27]                
        Upper_size = row[28]
        Left_side_size = row[29]
        Right_side_size = row[30]
        SW1 = row[31]
        SW2 = row[32]
        SWAngle = row[33]
        E1 = row[34]
        E2 = row[35]
        Angle = round(row[39] + 0.6, 1)
        Top_pop1 = round(row[40] + 0.6, 1) #소수점 첫째자리까지만 나오게
        Top_pop2 = round(row[41] + 0.6, 1)
        Kadapt =  row[51]   # K값 J값에 적용여부 넣기 1이면 넣기

        if(Kadapt==1):
            K1_up = 0
            K1_down = 0
            K2_up = 0
            K2_down = 0

        OP = validate_or_default(OP)

        if(OP>0):  
            if (SWAngle>1):  #SWAngle 0 이상인 경우            
                if(K1_up>0):     
                    if (SWAngle == 90):       
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 3 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = Abs_Xpos + LH2             
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                        lineto(doc,  X5, Y5, layer='레이져')                    
                    else:
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 2 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = Abs_Xpos + LH2             
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                        lineto(doc,  X5, Y5, layer='레이져')                    
                else:
                    if (SWAngle == 90):                     
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 3  - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4                
                    else:
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 2  - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4                

            if (SWAngle<1):  #SWAngle 0 이상인 경우
                if(K1_up>0):            
                    X1 = Abs_Xpos 
                    Y1 = Abs_Ypos + JB1_up + K1_up + SW - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + LH2 
                    Y2 = Abs_Ypos + JB1_down + K1_down + SW - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')        
                    dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                    X3 = X2            
                    Y3 = Y2
                    X4 = Abs_Xpos + LH2             
                    Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = Abs_Xpos + LH2             
                    Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                    lineto(doc,  X5, Y5, layer='레이져')                    
                else:
                    X1 = Abs_Xpos 
                    Y1 = Abs_Ypos + JB1_up + SW - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + LH2 
                    Y2 = Abs_Ypos + JB1_down + SW - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')        
                    dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                    X3 = X2            
                    Y3 = Y2
                    X4 = Abs_Xpos + LH2             
                    Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = X4             
                    Y5 = Y4                

            X6 = Abs_Xpos + LH1             
            Y6 = Abs_Ypos
            lineto(doc,  X6, Y6, layer='레이져')
            if(A1>0):       
                X7 = Abs_Xpos + LH1   
                Y7 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X7, Y7, layer='레이져')                
                X8 = Abs_Xpos + LH1  
                Y8 = Abs_Ypos - C1 - A1 + Vcut_rate * 3
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = Abs_Xpos             
                Y10 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X10, Y10, layer='레이져')

            if(A1<1):       
                X7 = Abs_Xpos + LH1   
                Y7 = Abs_Ypos - C1 + Vcut_rate 
                lineto(doc,  X7, Y7, layer='레이져')                
                X8 = Abs_Xpos + LH1  
                Y8 = Y7
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = X9
                Y10 = Y9            

            X11 = Abs_Xpos 
            Y11 = Abs_Ypos
            lineto(doc,  X11, Y11, layer='레이져')        

            if (SWAngle>0): 
                if(K1_up>0):  
                    if (SWAngle == 90):                    
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up  - Vcut_rate   
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = Abs_Xpos 
                        Y13 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X13, Y13, layer='레이져')
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 3 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up  - Vcut_rate   
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = Abs_Xpos 
                        Y13 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X13, Y13, layer='레이져')
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                else:  
                    if (SWAngle == 90):   
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate     
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = X12
                        Y13 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate                 
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 3 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                          
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate     
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = X12
                        Y13 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate                 
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                          
                    
            if (SWAngle<1): 
                if(K1_up>0):  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + JB1_up  - Vcut_rate   
                    lineto(doc,  X12, Y12, layer='레이져')
                    X13 = Abs_Xpos 
                    Y13 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                    lineto(doc,  X13, Y13, layer='레이져')
                    X14 = Abs_Xpos 
                    Y14 = Abs_Ypos + JB1_up + K1_up + SW  - Bending_rate * 2 - Vcut_rate  
                    lineto(doc,  X14, Y14, layer='레이져')  
                    lineto(doc,  X1, Y1, layer='레이져')    
                else:  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate 
                    lineto(doc,  X12, Y12, layer='레이져')
                    X13 = X12
                    Y13 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate 
                    X14 = X13
                    Y14 = Y13
                    lineto(doc,  X1, Y1, layer='레이져')   

            # 좌기둥 기준선 
            if(SW1>0):
                line(doc, X14, Y14, X3, Y3,  layer='22')
            line(doc, X13, Y13, X4, Y4,  layer='22')
            if(K1_up>0):
                line(doc, X12, Y12, X5, Y5,  layer='22')
            line(doc, X11, Y11, X6, Y6,  layer='22')
            if(A1>0):
                line(doc, X10, Y10, X7, Y7,  layer='22')                

            # 우측 절곡치수선
            if(SW1>0):
                dim_vertical_right(doc, X2,  Y2, X3, Y3,  extract_abs(X2,X3) + 160, 'dim', text_height=0.22, text_gap=0.01)      
            dim_vertical_right(doc, X3, Y3, X4,  Y4,   extract_abs(X3,X4) + 80, 'dim', text_height=0.22, text_gap=0.01)      
            if(K1_up>0):
                dim_vertical_right(doc, X4,  Y4, X5, Y5,  extract_abs(X4,X5) +  130, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_right(doc, X6, Y6, X5,  Y5, extract_abs(X6,X5) +  80, 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_right(doc,  X7, Y7, X6,  Y6,   extract_abs(X6,X7) + 80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_right(doc, X7,  Y7, X8, Y8,   extract_abs(X7,X8) + 150, 'dim', text_height=0.22, text_gap=0.01)  
                # V컷 치수표현
                dim_vertical_left(doc, X6,  Y6, X8, Y8,   extract_abs(X6,X8) + 100, 'dim', text_height=0.22, text_gap=0.01)  
            dim_vertical_right(doc, X2,  Y2, X8, Y8,   extract_abs(X2,X8) + 230 , 'dim', text_height=0.22, text_gap=0.01)      

            # 좌측 절곡치수선  
            if(SW1>0):                       
                dim_vertical_left(doc, X1, Y1, X14, Y14,  160, 'dim', text_height=0.22, text_gap=0.01)  
                dim_vertical_left(doc, X13, Y13, X14, Y14,   80 , 'dim', text_height=0.22, text_gap=0.01)  
            else:
                dim_vertical_left(doc, X1, Y1, X13, Y13,   80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(K1_up>0):
                dim_vertical_left(doc, X13, Y13, X12, Y12,  130 , 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X11, Y11, X12, Y12, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_left(doc, X11, Y11, X10, Y10,  80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_left(doc, X9, Y9, X10, Y10,  150 , 'dim', text_height=0.22, text_gap=0.01)
            dim_vertical_left(doc, X1, Y1, X9,  Y9,  230 , 'dim', text_height=0.22, text_gap=0.01)

            # 기둥 하부 치수선 2    
            dim_linear(doc,   X9, Y9, X8, Y8, "", 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   

            if LH1>LH2:
                dim_linear(doc,  X2, Y2, X6, Y6, "",  100 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)       
            if LH1<LH2:    
                dim_linear(doc,  X4, Y4, X8, Y8, "", extract_abs(Y4, Y8) + 100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            if args.opt5:
                drawcircle(doc, X1 + 12.5, Y1 - 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X2 - 12.5, Y2 - 5, 2.5 , layer='레이져') # 도장홀 5파이        

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"       

            Xpos = Abs_Xpos + 500 
            Ypos = Abs_Ypos + JB1_up / 2.3
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-좌"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')        

            Xpos = Abs_Xpos + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-좌"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Mat.Spec : {normal_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/1.2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Size : {Left_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/1.2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')       

            # 멍텅구리 기둥 좌우 도면틀 넣기
            BasicXscale = 4487
            BasicYscale = 2770
            # 우측단면도가 있을 경우 감안해야 함. 
            if( JB1_up != JB1_down or K1_up != K1_down  or JB2_up != JB2_down or K2_up != K2_down ):                                 
                TargetXscale = 550 + max(LH1, LH2) + 1000
            else:
                TargetXscale = 550 + max(LH1, LH2) + 550

            TargetYscale = 320*4 + (JB1_up + C1 + A1 + K1_up + SW) * 2
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            # 우측 단면도가 있는 경우 도면틀 기본점을 아래로 250 내림    
            if( JB1_up != JB1_down or K1_up != K1_down  or JB2_up != JB2_down or K2_up != K2_down ):                                 
                insert_frame(X1 -  710 , Abs_Ypos - (JB2_up+C2+A2+K2_up + SW + 1200 + C1 + A1 ) - 250 , frame_scale, "side_normal", workplace)                
            else:
                insert_frame(X1 -  710 , Abs_Ypos - (JB2_up+C2+A2+K2_up + SW + 1200 + C1 + A1 ) , frame_scale, "side_normal", workplace)                

            ##############################################################################################################################################################
            # 멍텅구리 좌기둥 좌측 단면도    
            ##############################################################################################################################################################
            Section_Xpos = Abs_Xpos - 400
            Section_Ypos = Abs_Ypos + JB1_up + K1_up

            # 가정: SWAngle은 도 단위로 주어진 각도
            SWAngle_radians = math.radians(SWAngle)

            if(SWAngle>0):
                X1 = Section_Xpos - SW1 - SW2 * math.cos(SWAngle_radians)
                Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                X2 = Section_Xpos - SW1
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos
                
                line(doc, X2, Y2, X1, Y1, layer='0')               
                line(doc, X2, Y2, X3, Y3, layer='0')        
                dim_linear(doc, X1, Y1, X2, Y2,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # 각도 90도가 아니면 각도 표기
                if (SWAngle != 90):               
                    dim_angular(doc, X2 , Y2, X1, Y1, X2-30, Y2, X3-30, Y3,  160, direction="left" )
                    # help(ezdxf.layouts.Modelspace.add_angular_dim_2l)
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)   

            if(SWAngle<1):
                X1 = Section_Xpos - SW
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos            
                line(doc,  X1, Y1, X2, Y2, layer='0')                                                   
                dim_linear(doc, X1, Y1, X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

            if(K1_up>5):
                X4 = X3
                Y4 = Y3 - K1_up
                lineto(doc,  X4, Y4, layer='0')       
            else:
                X4 = X3
                Y4 = Y3

            # Angle은 도 단위로 주어진 각도
            Angle_radians = math.radians(Angle)

            X5 = X4 - JB1_up * math.sin(Angle_radians)                
            Y5 = Y4 - JB1_up * math.cos(Angle_radians)  
            lineto(doc,  X5, Y5, layer='0')    
            X6 = X5 - C1
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='0')   

            if(A1>0):
                X7 = X6
                Y7 = Y6 + A1
                lineto(doc,  X7, Y7, layer='0') 
                # drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                # drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                # dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut 2개소", layer="dim", style='0')
                # A 치수
                dim_vertical_left(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
            if(A1<1):
                X7 = X6
                Y7 = Y6            
                # drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                # dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut", layer="dim", style='0')

            # C값 치수선
            dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            if(K1_up>5):
                dim_vertical_right(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                # if (Angle > 1):   # 직각이 아닌경우                        
                dim_angular(doc, X4, Y4, X5, Y5, X4, Y4, X3, Y3,  30, direction="left" )

            # JB사선 치수선
            dim_linear(doc, X4, Y4, X5, Y5,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
            # if (Angle > 1):   # 직각이 아닌경우            
            dim_angular(doc, X6, Y6, X5, Y5, X5, Y5, X4, Y4,  20, direction="left" )
            

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB1_up != JB1_down ):
                rectangle(doc, X3+10, Y3 - JB2_up/2 - 75, X3 + 10 + 100, Y3 - JB1_up/2 + 75 , layer='0')               
                rectangle(doc, X3+60+200, Y3 - JB2_up/2 - 75 , X3 + 60 + 300, Y3 - JB1_up/2 + 75 , layer='0')               
            if(K1_up != K1_down ):
                rectangle(doc, X5 + 50, Y5 - 10 , X5 + 50 + 100, Y5 + K1_up + 10 , layer='0')                         
                rectangle(doc, X5 + 10 + 250, Y5 - 10 , X5 + 10 + 300, Y5 + K1_up + 10 , layer='0')                            

            #######################################################################################################################################################################
            # 멍텅구리 좌기둥 우측 단면도    
            #######################################################################################################################################################################
            
            if( JB1_up != JB1_down or K1_up != K1_down ):                  
                Section_Xpos = Abs_Xpos + LH1 + 400
                Section_Ypos = Abs_Ypos + JB1_up + K1_up

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(SWAngle>0):  
                    X1 = Section_Xpos + SW1 + SW2 * math.cos(SWAngle_radians)
                    Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                    X2 = Section_Xpos + SW1
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc, X2, Y2, X1, Y1, layer='0')                   
                    line(doc, X2, Y2, X3, Y3, layer='0')        
                    dim_linear(doc, X2, Y2, X1, Y1,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # 각도 90도가 아니면 각도 표기
                    if (SWAngle != 90):               
                        dim_angular(doc,  X3+5, Y3, X2+5, Y2, X1, Y1, X2 , Y2,  150, direction="right" )
                        # help(ezdxf.layouts.Modelspace.add_angular_dim_2l)
                    dim_linear(doc, X3, Y3,  X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)    

                if(SWAngle<1):  
                    X1 = Section_Xpos + SW
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos 
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc,   X1, Y1, X2, Y2,  layer='0')                                   
                    dim_linear(doc,  X2, Y2, X1, Y1,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

                if(K1_down>5):
                    X4 = X3
                    Y4 = Y3 - K1_down
                    lineto(doc,  X4, Y4, layer='0')       
                else:
                    X4 = X3
                    Y4 = Y3

                # Angle은 도 단위로 주어진 각도
                Angle_radians = math.radians(Angle)

                X5 = X4 + JB1_down * math.sin(Angle_radians)                
                Y5 = Y4 - JB1_down * math.cos(Angle_radians)  
                lineto(doc,  X5, Y5, layer='0')    
                X6 = X5 + C1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')   

                # if(A1>0):
                #     X7 = X6
                #     Y7 = Y6 + A1
                #     lineto(doc,  X7, Y7, layer='0') 
                #     drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                #     drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                #     dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut 2개소", layer="dim", style='0')
                #     # A 치수
                #     dim_vertical_right(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
                # if(A1<1):
                #     X7 = X6
                #     Y7 = Y6            
                #     drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                #     dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut", layer="dim", style='0')

                # C값 치수선
                dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                if(K1_up>5):
                    dim_vertical_left(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                    # if (Angle > 1):   # 직각이 아닌경우                        
                    dim_angular(doc,  X3, Y3, X4 , Y4,  X5, Y5, X4, Y4,   15, direction="right" )

                # JB사선 치수선
                dim_linear(doc,  X5, Y5, X4, Y4,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
                # if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc,  X5, Y5, X4, Y4, X5, Y5, X6 , Y6,  15, direction="right" )

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB1_up != JB1_down ):
                    rectangle(doc, X3-10, Y3 - JB1_up/2 - 75, X3 - 10 - 100, Y3 - JB1_up/2 + 75 , layer='0')               
                    rectangle(doc, X3 - 60 - 200, Y3 - JB1_up/2 - 75 , X3 - 60 - 300, Y3 - JB1_up/2 + 75 , layer='0')               
                if(K1_up != K1_down ):
                    rectangle(doc, X5 - 50, Y5 - 10 , X5 - 50 - 100, Y5 + K1_up + 10 , layer='0')                         
                    rectangle(doc, X5 - 10 + 250, Y5 - 10 , X5 - 10 - 300, Y5 + K1_up + 10 , layer='0')                              
            
            ######################################################################################################################################################
            # 멍텅구리 우기둥 전개도
            # 멍텅구리 우기둥 전개도
            ######################################################################################################################################################

            Abs_Ypos -= 2000 
            R_Ypos = Abs_Ypos + 1400 - (C1 + C2 + A1 + A2)

            if(A2>0):            
                X1 = Abs_Xpos 
                Y1 = R_Ypos + C2 + A2 - Vcut_rate * 3
                X2 = Abs_Xpos + RH1 
                Y2 = R_Ypos + C2 + A2 - Vcut_rate * 3
                line(doc, X1, Y1, X2, Y2, layer='레이져')        
                dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                X3 = Abs_Xpos + RH1 
                Y3 = R_Ypos + C2 - Vcut_rate * 2
                lineto(doc,  X3, Y3, layer='레이져')                
            if(A2<1):
                X1 = Abs_Xpos 
                Y1 = R_Ypos + C2  - Vcut_rate 
                X2 = Abs_Xpos + RH1 
                Y2 = Y1
                line(doc, X1, Y1, X2, Y2, layer='레이져')        
                dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                X3 = X2
                Y3 = Y2

            X4 = Abs_Xpos + RH1 
            Y4 = R_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')
            if (SWAngle>1):  #SWAngle 0 이상인 경우    헤밍구조인 경우        
                if(K2_up>0):   
                    if (SWAngle == 90):                       
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = Abs_Xpos + RH2 
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = Abs_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')
                    else:
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = Abs_Xpos + RH2 
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = Abs_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')
                if(K2_up<1):   #K값이 없는 경우
                    if (SWAngle == 90):                      
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11
                    else:
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11

            if (SWAngle<1):  #SWAngle 0     일반구조
                if(K2_up>0):   
                    X5 = Abs_Xpos + RH2 
                    Y5 = R_Ypos - JB2_down +  Vcut_rate
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = Abs_Xpos + RH2 
                    Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                    lineto(doc,  X6, Y6, layer='레이져')
                    X7 = Abs_Xpos + RH2             
                    Y7 = R_Ypos - JB2_down - K2_down - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = Abs_Xpos 
                    Y9 = R_Ypos - JB2_up - K2_up - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = Abs_Xpos 
                    Y10 = Y9                
                    X11 = Abs_Xpos 
                    Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = Abs_Xpos 
                    Y12 = R_Ypos - JB2_up +  Vcut_rate
                    lineto(doc,  X12, Y12, layer='레이져')
                if(K2_up<1):   #K값이 없는 경우
                    X5 = Abs_Xpos + RH2 
                    Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = X5
                    Y6 = Y5                
                    X7 = Abs_Xpos + RH2             
                    Y7 = R_Ypos - JB2_down  - SW  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = Abs_Xpos 
                    Y9 = R_Ypos - JB2_up  - SW  + Bending_rate * 2 + Vcut_rate
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = X9
                    Y10 = Y9                
                    X11 = Abs_Xpos 
                    Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = X11
                    Y12 = Y11                

            X13 = Abs_Xpos 
            Y13 = R_Ypos
            lineto(doc,  X13, Y13, layer='레이져')

            if(A2>0):
                X14 = Abs_Xpos 
                Y14 = R_Ypos + C2 - Vcut_rate * 2
                lineto(doc,  X14, Y14, layer='레이져')
                lineto(doc,  X1, Y1, layer='레이져')
            if(A2<1):
                X14 = X1
                Y14 = Y1
                lineto(doc,  X14, Y14, layer='레이져')

            # 기둥 절곡선 1
            if(A2>1):        
                line(doc, X14, Y14, X3, Y3,  layer='22')
            # C2    
            line(doc, X13, Y13, X4, Y4,  layer='22')
            # JB2
            line(doc, X12, Y12, X5, Y5,  layer='22')
            if(K2_up>0):
                line(doc, X11, Y11, X6, Y6,  layer='22')
            # 해밍구조
            if (SWAngle>0):   
                line(doc, X10, Y10, X7, Y7,  layer='22')

            # 멍텅구리 우기둥 우측 절곡치수선
            if(A2>1):  
                dim_vertical_right(doc, X3, Y3, X2,  Y2,  150, 'dim', text_height=0.22,  text_gap=0.07)      
                dim_vertical_left(doc, X4, Y4, X2,  Y2, extract_abs(X2,X4) + 150, 'dim', text_height=0.22,  text_gap=0.07)      
            # C    
            dim_vertical_right(doc, X3, Y3, X4,  Y4, 60, 'dim', text_height=0.22,  text_gap=0.07)      
            dim_vertical_right(doc, X5, Y5, X4,  Y4, extract_abs(X6,X4) + 60, 'dim', text_height=0.22,  text_gap=0.07) 
            if(K2_down>0): 
                dim_vertical_right(doc, X6, Y6, X5,  Y5, extract_abs(X6,X4) + 110, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X6, Y6, X7,  Y7, extract_abs(X6,X4) + 60, 'dim', text_height=0.22,  text_gap=0.07)  
            if (SWAngle>0):   
                dim_vertical_right(doc, X7, Y7, X8,  Y8, extract_abs(X6,X4) + 200, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X2, Y2, X8,  Y8, extract_abs(X2,X8) + 250, 'dim', text_height=0.22,  text_gap=0.07)  

            # 멍텅구리 우기둥 좌측 절곡치수선
            if(A2>1):          
                dim_vertical_left(doc, X1, Y1, X14, Y14, 150 , 'dim', text_height=0.22, text_gap=0.01)  
            # C    
            dim_vertical_left(doc, X13, Y13, X14, Y14,   50 , 'dim', text_height=0.22, text_gap=0.01)  
            # JB
            dim_vertical_left(doc, X13, Y13, X12, Y12,  50 , 'dim', text_height=0.22, text_gap=0.01)              
            if(K2_up>0): 
                dim_vertical_left(doc, X11, Y11, X12, Y12,  extract_abs(X11,X12) + 110, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X11, Y11, X10, Y10,  extract_abs(X11,X10) + 50, 'dim', text_height=0.22, text_gap=0.01)  
            if (SWAngle>0):               
                dim_vertical_left(doc, X9, Y9, X10, Y10,  extract_abs(X9,X10) + 160, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X1, Y1, X9,  Y9,  extract_abs(X1,X9) + 230 , 'dim', text_height=0.22, text_gap=0.01)      

            # 멍텅구리 우기둥 전개도 하부 치수선    
            dim_linear(doc,  X9, Y9, X8, Y8,  "", 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.01)      
            if RH1<RH2:
                dim_linear(doc,  X2, Y2, X5, Y5,  "",  100 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)       
            if RH1>RH2:    
                dim_linear(doc,  X8, Y8, X4, Y4, "",  100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.01)      
    
            if args.opt5:
                drawcircle(doc, X9 + 12.5, Y9 + 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X8 - 12.5, Y8 + 5, 2.5 , layer='레이져') # 도장홀 5파이

            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"

            Xpos = Abs_Xpos + 500 
            Ypos = R_Ypos - JB2_up / 2
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-우"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')           

            Xpos = Abs_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-우"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Mat.Spec : {normal_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Size : {Right_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')      
            
            ####################################################################################################################################################################################
            # 멍텅구리 우기둥 좌측 단면도    
            ####################################################################################################################################################################################
            Section_Xpos = Abs_Xpos - 500        
            Section_Ypos = R_Ypos

            SWAngle_radians = math.radians(SWAngle)

            if(A2>0):
                X1 = Section_Xpos
                Y1 = Section_Ypos - A2
                X2 = Section_Xpos 
                Y2 = Section_Ypos 

                line(doc, X1, Y1, X2, Y2, layer='0')                               
                dim_vertical_left(doc, X1, Y1, X2, Y2, 50, 'dim', text_height=0.22, text_gap=0.07)  
                X3 = X2 + C2
                Y3 = Y2
                lineto(doc,  X3, Y3, layer='0')   
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # vcut 표기                
                # drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                # dim_leader_line(doc, X3, Y3 , X3 + 50, Y3 + 100, "V-cut 2개소", layer="dim", style='0')    

            if(A2<1):
                X1 = Section_Xpos
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos             
                X3 = X2 + C2
                Y3 = Y2
                line(doc, X1, Y1, X3, Y3, layer='0')               
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # # vcut 표기                            
                # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                # dim_leader_line(doc, X3, Y3 , X3+50, Y3+100, "V-cut", layer="dim", style='0')                                        

            X4 = X3 + JB2_up * math.sin(Angle_radians)                
            Y4 = Y3 - JB2_up * math.cos(Angle_radians)        
            lineto(doc,  X4, Y4, layer='0')	
            dim_linear(doc, X3, Y3, X4, Y4,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)     

            if(K2_up > 5):
                X5 = X4
                Y5 = Y4 - K2_up
                lineto(doc,  X5, Y5, layer='0')	
                dim_vertical_right(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
            else:
                X5 = X4
                Y5 = Y4
                
            if(SWAngle>0):
                X6 = X5 - SW1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                X7 = X6 - SW2 * math.cos(SWAngle_radians)
                Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                lineto(doc,  X7, Y7, layer='0')
                dim_linear(doc, X6, Y6, X7, Y7,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                dim_angular(doc, X5-5, Y5, X6-5, Y6,  X7, Y7, X6 , Y6,  100, direction="left" )

            if(SWAngle<1): # 헤밍구조가 아닌경우 일반형
                X6 = X5 - SW
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                X7 = X6
                Y7 = Y6                        

            # if (Angle > 1):   # 직각이 아닌경우            
            dim_angular( doc,   X3, Y3, X4, Y4, X3, Y3, X2 , Y2,    15, direction="left" )		                            
            if(K2_up > 5): # K2값
                dim_angular(doc,  X4, Y4, X5 , Y5, X4, Y4, X3, Y3,  5, direction="left" )    

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB2_up != JB2_down ):
                rectangle(doc, X3+60, Y3 - JB2_up/2 - 75, X3 + 60 + 100, Y3 - JB2_up/2 + 75 , layer='0')               
                rectangle(doc, X3+60+300, Y3 - JB2_up/2 - 75 , X3 + 60 + 400, Y3 - JB2_up/2 + 75 , layer='0')               
            if(K2_up != K2_down ):
                rectangle(doc, X5 + 50, Y5 - 10 , X5 + 50 + 100, Y5 + K2_up + 10 , layer='0')                         
                rectangle(doc, X5 + 10 + 250, Y5 - 10 , X5 + 10 + 300, Y5 + K2_up + 10 , layer='0')                            


            ####################################################################################################################################################################################
            # 멍텅구리 우기둥 우측 단면도    
            #################################################################################################################################################################################### 

            if( JB2_up != JB2_down or K2_up != K2_down ):   

                Section_Xpos = Abs_Xpos + max(RH1,RH2) + 500
                Section_Ypos = R_Ypos

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(A2>0):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos - A2
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos 

                    line(doc, X1, Y1, X2, Y2, layer='0')
                    dim_vertical_right(doc, X2, Y2, X1, Y1,  50, 'dim', text_height=0.22, text_gap=0.07)
                    X3 = Section_Xpos
                    Y3 = Section_Ypos
                    lineto(doc,  X3, Y3, layer='0')
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # vcut 표기                
                    # drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    # dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut 2개소", layer="dim", style='0')     

                if(A2<1):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos                 
                    X3 = Section_Xpos
                    Y3 = Section_Ypos 
                    line(doc, X2, Y2, X3, Y3, layer='0')                               
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # # vcut 표기                                
                    # drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    # dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut", layer="dim", style='0')                                

                X4 = X3 - JB2_down * math.sin(Angle_radians)                
                Y4 = Y3 - JB2_down * math.cos(Angle_radians)        
                lineto(doc,  X4, Y4, layer='0')	
                dim_linear(doc, X4, Y4, X3, Y3,   "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)

                if(K2_down > 5):
                    X5 = X4
                    Y5 = Y4 - K2_down
                    lineto(doc,  X5, Y5, layer='0')	
                    dim_vertical_left(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
                else:
                    X5 = X4
                    Y5 = Y4
                    lineto(doc,  X5, Y5, layer='0')	

                if(SWAngle>0):
                    X6 = X5 + SW1
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                    X7 = X6 + SW2 * math.cos(SWAngle_radians)
                    Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                    lineto(doc,  X7, Y7, layer='0')
                    dim_linear(doc, X7, Y7, X6, Y6,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                    dim_angular(doc,  X7 , Y7, X6, Y6, X6 + 5 , Y6, X6, Y6,  150, direction="right" )
                if(SWAngle<1):
                    X6 = X5 + SW
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,"" , 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    X7 = X6
                    Y7 = Y6

                # if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc,  X3, Y3, X2-8 , Y2,  X3, Y3, X4, Y4,   25, direction="right")		            
                # if(K2_down>0):
                dim_angular(doc,   X4, Y4, X3 , Y3, X5, Y5,  X4, Y4, 30, direction="right")

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB2_up != JB2_down ):
                    rectangle(doc, X3 - 60, Y3 - JB2_up/2 - 75, X3 - 60 - 100, Y3 - JB2_up/2 + 75 , layer='0')               
                    rectangle(doc, X3 - 60 - 350, Y3 - JB2_up/2 - 75 , X3 - 60 - 450, Y3 - JB2_up/2 + 75 , layer='0')               
                if(K2_up != K2_down ):
                    rectangle(doc, X5 - 50, Y5 - 10 , X5 - 50 - 100, Y5 + K2_up + 10 , layer='0')                         
                    rectangle(doc, X5 - 10 + 350, Y5 - 10 , X5 - 10 - 400, Y5 + K2_up + 10 , layer='0')                              
                                    

            Abs_Ypos -= 2000 

####################################################################################################################################################################################
# 쪽쟘 자동작도
####################################################################################################################################################################################
def execute_narrow():     

    # 시트 선택 (시트명을 지정)
    sheet_name = '쪽쟘제작'  # 원하는 시트명으로 변경
    sheet = workbook[sheet_name]

    # 2차원 배열을 저장할 리스트 생성
    data_2d = []

    narrow_Vcut_rate = Vcut_rate
    narrow_Bending_rate = Bending_rate	

    # 1행부터 20행까지의 값을 2차원 배열에 저장
    for row_num in range(7, sheet.max_row + 1):  # 7행부터 20행까지 반복
        cell_value = sheet.cell(row=row_num, column=2).value  # 2열의 값 가져오기
        if cell_value is not None and cell_value != 'F':  # 값이 None이 아닌 경우에만 배열에 추가
            row_values =  [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, sheet.max_column + 10)]        
            data_2d.append(row_values)

    # 엑셀 파일 닫기
    workbook.close()   

#########################################
# 쪽쟘 상판 전개도
#########################################

    Abs_Xpos = 23000
    Abs_Ypos = 100

    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]
        size = row[2]
        OP = row[3]
        JD = row[4]
        JB = row[5]
        Left_side_size = row[6]
        Right_side_size = row[7]
        H1 = row[8]
        H2 = row[9]
        C1_up = row[10]
        C1_down = row[11]
        C2_up = row[12]
        C2_down = row[13]
        LH1 = row[14]
        LH2 = row[15]
        RH1 = row[16]
        RH2 = row[17]
        U1 = row[18]
        U2 = row[19]    
        UW = row[20]
        SW = row[21]
        K1 = row[22]
        K2 = row[23]
        SW1 = row[24]
        SW2 = row[25]
        SWAngle = row[26]
        SideBoxcut_Left_A = row[34]  # 좌기둥 JD방향         
        SideBoxcut_Left_B = row[35]  # 좌기둥 높이 방향 
        SideBoxcut_Right_A = row[36]  # 우기둥 JD방향         
        SideBoxcut_Right_B = row[37]  # 우기둥 높이 방향 

        SideBoxcut_Left_A = validate_or_default(SideBoxcut_Left_A)
        SideBoxcut_Left_B = validate_or_default(SideBoxcut_Left_B)
        SideBoxcut_Right_A = validate_or_default(SideBoxcut_Right_A)
        SideBoxcut_Right_B = validate_or_default(SideBoxcut_Right_B)
        OP = validate_or_default(OP)

        if args.opt1_1: # 상판 Vcut 없음
            narrow_Vcut_rate = narrow_Bending_rate

        # 데이터가 실제 그리는 것인지 판단하는 OP문장 추가
        if(OP>0):
            
            # print(f"Row {idx + 1}: Floor_mc = {Floor_mc}, Floor_des = {Floor_des}, size = {size}, OP = {OP}, JD = {JD}, JB = {JB}, Left_side_size = {Left_side_size}, Right_side_size = {Right_side_size}, H1 = {H1}, H2 = {H2}, C1_up = {C1_up}, C1_down = {C1_down}, C2_up = {C2_up}, C2_down = {C2_down}, LH1 = {LH1}, LH2 = {LH2}, RH1 = {RH1}, RH2 = {RH2}, U1 = {U1}, U2 = {U2}, UW = {UW}, SW = {SW}, K1 = {K1}, K2 = {K2}, SW1 = {SW1}, SW2 = {SW2}, SWAngle = {SWAngle}")
            if(U1 > 1 and U2 > 1):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U1 - narrow_Vcut_rate
                X2 = Abs_Xpos + H1 + H2 + OP 
                Y2 = Abs_Ypos + U2 - narrow_Vcut_rate  

                if args.opt3: # 추영덕소장 상판5R 적용형태 그리기            
                    line(doc, X1 + 5, Y1, X2 - 5, Y2, layer='레이져')        
                else:
                    line(doc, X1, Y1, X2, Y2, layer='레이져')       

                dim_linear(doc,  X1, Y1, X2, Y2, "", 50,  direction="up", layer='dim')
    
                X3 = Abs_Xpos + H1 + H2 + OP 
                Y3 = Abs_Ypos 
                # 치수선을 그리는 함수 호출    
                if args.opt3: # 추영덕소장 상판5R 적용형태 그리기            
                    line(doc, X2, Y2 - 5, X3, Y3, layer='레이져')        
                else:
                    line(doc, X2, Y2, X3, Y3, layer='레이져')    
                X4 = Abs_Xpos + H1 + H2 + OP
                Y4 = Abs_Ypos - args.narrow_TopPopup  + narrow_Vcut_rate
                lineto(doc,  X4, Y4, layer='레이져')
                X5 = Abs_Xpos + H1 +  OP
                Y5 = Abs_Ypos - args.narrow_TopPopup + narrow_Vcut_rate
                lineto(doc,  X5, Y5, layer='레이져')
                X6 = X5
                Y6 = Abs_Ypos - JD  + narrow_Vcut_rate + narrow_Bending_rate
                lineto(doc,  X6, Y6, layer='레이져')
                X7 = X6
                Y7 = Abs_Ypos - JD - UW  + narrow_Vcut_rate + narrow_Bending_rate*2
                lineto(doc,  X7, Y7, layer='레이져')    
                X8 = Abs_Xpos + H1
                Y8 = Abs_Ypos - JD - UW  + narrow_Vcut_rate + narrow_Bending_rate*2
                lineto(doc,  X8, Y8, layer='레이져')    
                X9 = X8
                Y9 = Abs_Ypos - JD  + narrow_Vcut_rate + narrow_Bending_rate
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = X9
                Y10 = Abs_Ypos - args.narrow_TopPopup + narrow_Vcut_rate
                lineto(doc,  X10, Y10, layer='레이져')
                X11 = Abs_Xpos
                Y11 = Y10
                lineto(doc,  X11, Y11, layer='레이져')
                X12 = Abs_Xpos
                Y12 = Abs_Ypos
                lineto(doc,  X12, Y12, layer='레이져')                                        
            else:
                # U값 없는 경우
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos 
                X2 = Abs_Xpos + H1 + H2 + OP 
                Y2 = Abs_Ypos    
                line(doc, X1, Y1, X2, Y2, layer='레이져')              
                dim_linear(doc,  X1, Y1, X2, Y2, "", 50,  direction="up", layer='dim')
                X3 = X2
                Y3 = Y2                    
                X4 = X3
                Y4 = Y3                
                X5 = Abs_Xpos + H1 + H2 + OP 
                Y5 = Y4                
                X6 = X5
                Y6 = Abs_Ypos - JD  + narrow_Bending_rate
                lineto(doc,  X6, Y6, layer='레이져')
                X7 = X6
                Y7 = Abs_Ypos - JD - UW  + narrow_Bending_rate*2
                lineto(doc,  X7, Y7, layer='레이져')    
                X8 = Abs_Xpos + H1
                Y8 = Abs_Ypos - JD - UW  + narrow_Bending_rate*2
                lineto(doc,  X8, Y8, layer='레이져')    
                X9 = X8
                Y9 = Abs_Ypos - JD  + narrow_Bending_rate
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = X9
                Y10 = Abs_Ypos
                lineto(doc,  X10, Y10, layer='레이져')
                X11 = Abs_Xpos
                Y11 = Y10                
                X12 = Abs_Xpos
                Y12 = Abs_Ypos                

            if(U1 > 1 and U2 > 1):        
                if args.opt3: # 추영덕소장 상판5R 적용형태 그리기            
                    line(doc, X12, Y12, X1, Y1 - 5, layer='레이져')        
                    add_90_degree_fillet(doc,  (X1 + 5, Y1 - 5), (X1 + 5, Y1), (X1 + 5,  Y1 - 5), (X1, Y1 - 5 ), 5)
                    add_90_degree_fillet(doc,  (X2 - 5, Y2 - 5), (X2, Y2 - 5), (X2 - 5, Y2 - 5), (X2 - 5, Y2),  5)
                else:
                    lineto(doc,  X1, Y1, layer='레이져')      

            if args.opt2:
                drawcircle(doc, X8 + 12.5, Y8 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X7 - 12.5, Y7 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이

            # 상판 전개도 절곡선
            if(U1 > 1 and U2 > 1):             
                line(doc, X12, Y12, X3, Y3,  layer='22')
                # vcut 지시선
                if(narrow_Vcut_rate != narrow_Bending_rate):
                    dim_leader_line(doc, X12+100, Y12, X12+100 + 50,Y12 - 10, "V-cut", layer="dim", style='0')        

            # 상판 뒷날개쪽 절곡선
            line(doc, X9, Y9, X6, Y6,  layer='22')

            # 상판 우측 전개도 치수선
            if(U1 > 1 and U2 > 1):             
                dim_vertical_right(doc,  X2, Y2, X3, Y3,  70, 'dim', text_height=0.22,  text_gap=0.07)             
            dim_vertical_right(doc, X3, Y3, X6, Y6,   extract_abs(X3,X6) +  120, 'dim', text_height=0.22,  text_gap=0.07)             
            dim_vertical_right(doc, X7, Y7, X6, Y6,   extract_abs(X2,X7) +  70, 'dim', text_height=0.22,  text_gap=0.07)     
            dim_vertical_right(doc, X2, Y2,  X7, Y7,   extract_abs(X2,X7) + 220, 'dim', text_height=0.22,  text_gap=0.07)    

            # 상판 좌측 전개도 치수선
            if(U1 > 1 and U2 > 1):             
                dim_vertical_left(doc, X12,  Y12,  X1, Y1,  40, 'dim', text_height=0.22, text_gap=0.01)  
                # 상판 돌출부위 3.4 치수선 만들기            
                dim_vertical_left(doc, X5, Y5,  X3, Y3, 50, 'dim', text_height=0.22,  text_gap=0.07) 
                dim_vertical_right(doc, X12, Y12,  X10, Y10, 50, 'dim', text_height=0.22,  text_gap=0.07)                 

            dim_vertical_left(doc, X12, Y12, X9, Y9,  120 , 'dim', text_height=0.22, text_gap=0.01)      
            dim_vertical_left(doc, X8, Y8, X9, Y9,  extract_abs(X1,X8) +  40, 'dim', text_height=0.22, text_gap=0.01)      
            dim_vertical_left(doc, X1, Y1,  X8, Y8,  extract_abs(X1,X8) + 180 , 'dim', text_height=0.22, text_gap=0.01)     

            # 상판 하부 치수선 3개
            if(H1>1.2):
                dim_linear(doc,  X11, Y11, X8, Y8, "", extract_abs(Y11,Y8) + 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)                  
            dim_linear(doc,  X8, Y8, X7, Y7, "",  extract_abs(Y8,Y7) + 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            if(H2>1.2):               
                dim_linear(doc,  X7, Y7, X4, Y4, "",  100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            

            # 문구 설정        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"
            # Textstr 생성
            Textstr = f"{Pre_text}{str(Floor_des)}"       

            X2 = Abs_Xpos + (H1 + OP + H2)/2 - len(Textstr)*40/2
            U = max(U1,U2)
            Y2 = Abs_Ypos - (JD + U + UW )/2
            draw_Text(doc, X2, Y2, 40, str(Textstr), '레이져')

            ###########################################################################################################
            # 쪽쟘 상판 좌측 단면도
            ###########################################################################################################

            Section_Xpos = Abs_Xpos - 400

            X1 = Section_Xpos - U1
            Y1 = Abs_Ypos 
            X2 = Section_Xpos 
            Y2 = Abs_Ypos

            X3 = Section_Xpos 
            Y3 = Abs_Ypos - JD    
            X4 = Section_Xpos - UW
            Y4 = Abs_Ypos - JD          

            # vcut 표기 
            if(U1 > 1 and U2 > 1):                    
                if(narrow_Vcut_rate != narrow_Bending_rate):    # Vcut없는 경우도 적용할 수 있도록 로직개발함          
                    dim_leader_line(doc, X3, Y2 , X2 + 50, Y2 + 50, "V-cut", layer="dim", style='0')                                
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6')   
                dim_linear(doc,  X1, Y1, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

            line(doc, X1, Y1, X2, Y2, layer='0')                    
            lineto(doc,  X3, Y3, layer='0')    
            lineto(doc,  X4, Y4, layer='0')           
            dim_vertical_right(doc, X2, Y2, X3,  Y3, 80, 'dim', text_height=0.22,  text_gap=0.07)                  
            dim_linear(doc,  X4, Y4, X3, Y3,  "", 80,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            
            ###########################################################################################################
            # 쪽쟘 상판 우측 단면도
            ###########################################################################################################
              
            if(U1!=U2):        
                Section_Xpos = Abs_Xpos + OP + H2 + 400

                X1 = Section_Xpos + U2
                Y1 = Abs_Ypos 
                X2 = Section_Xpos 
                Y2 = Abs_Ypos

                # 좌측단면도 U1값
                line(doc, X1, Y1, X2, Y2, layer='0')        
                dim_linear(doc,  X2, Y2,  X1, Y1, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                X3 = Section_Xpos 
                Y3 = Abs_Ypos - JD    
                lineto(doc,  X3, Y3, layer='0')    
                X4 = Section_Xpos + UW
                Y4 = Abs_Ypos - JD  
                lineto(doc,  X4, Y4, layer='0')                  

                # vcut 지시선
                if(narrow_Vcut_rate != narrow_Bending_rate):
                    dim_leader_line(doc, X2, Y2 , X2-120, Y2+100, "V-cut", layer="dim", style='0')
                    # vcut 표기 2개소
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 

                # 우측 절곡치수선
                dim_vertical_left(doc, X2, Y2, X3,  Y3, 50, 'dim', text_height=0.22,  text_gap=0.07)  
                        
                #쪽쟘 하부 치수선 1개
                dim_linear(doc, X3, Y3, X4, Y4,   "뒷날개", 80,  direction="down", layer='dim' ,text_height=0.22,  text_gap=0.07)      
                
            Abs_Ypos -= 1000

    ########################################################
    # 쪽쟘 좌기둥 전개도
    ########################################################
    Abs_Xpos = 26000
    Abs_Ypos = 100

    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]
        size = row[2]
        OP = row[3]
        JD = row[4]
        JB = row[5]
        Left_side_size = row[6]
        Right_side_size = row[7]
        H1 = row[8]
        H2 = row[9]
        C1_up = row[10]
        C1_down = row[11]
        C2_up = row[12]
        C2_down = row[13]
        LH1 = row[14]
        LH2 = row[15]
        RH1 = row[16]
        RH2 = row[17]
        U1 = row[18]
        U2 = row[19]    
        UW = row[20]
        SW = row[21]
        K1 = row[22]
        K2 = row[23]
        SW1 = row[24]
        SW2 = row[25]
        SWAngle = row[26]
        SideBoxcut_Left_A = row[34]  # 좌기둥 JD방향         
        SideBoxcut_Left_B = row[35]  # 좌기둥 높이 방향 
        SideBoxcut_Right_A = row[36]  # 우기둥 JD방향         
        SideBoxcut_Right_B = row[37]  # 우기둥 높이 방향 

        SideBoxcut_Left_A = validate_or_default(SideBoxcut_Left_A)
        SideBoxcut_Left_B = validate_or_default(SideBoxcut_Left_B)
        SideBoxcut_Right_A = validate_or_default(SideBoxcut_Right_A)
        SideBoxcut_Right_B = validate_or_default(SideBoxcut_Right_B)        
        OP = validate_or_default(OP)        

        if(C1_up<1):
            side_narrow_Vcut_rate = 0
        else:
            side_narrow_Vcut_rate = narrow_Vcut_rate

        # 데이터가 실제 그리는 것인지 판단하는 OP문장 추가
        if(OP>0):
            # 헤밍구조, 일반구조 동시 좌기둥 부분처리
            if (SWAngle>1):  # SWAngle 0 이상인 경우 90도 추영덕소장 모양이 포함됨        
                X1 = Abs_Xpos 
                if (SWAngle == 90):
                    if(C1_up > 0):  
                        Y1 = Abs_Ypos + JB + SW1 + SW2 - narrow_Bending_rate * 4 - side_narrow_Vcut_rate
                    else:
                        Y1 = Abs_Ypos + JB + SW1 + SW2 - narrow_Bending_rate * 4 
                else:
                    if(C1_up > 0):  
                        Y1 = Abs_Ypos + JB + SW1 + SW2 - narrow_Bending_rate * 2 - side_narrow_Vcut_rate    
                    else:
                        Y1 = Abs_Ypos + JB + SW1 + SW2 - narrow_Bending_rate * 2
                X2 = Abs_Xpos + LH2 
                if (SWAngle == 90):
                    if(C1_down > 0): 
                        Y2 = Abs_Ypos + JB + SW1 + SW2 - narrow_Bending_rate * 4 - side_narrow_Vcut_rate
                    else:
                        Y2 = Abs_Ypos + JB + SW1 + SW2 - narrow_Bending_rate * 4
                else:
                    if(C1_down > 0): 
                        Y2 = Abs_Ypos + JB + SW1 + SW2 - narrow_Bending_rate * 2 - side_narrow_Vcut_rate    
                    else:
                        Y2 = Abs_Ypos + JB + SW1 + SW2 - narrow_Bending_rate * 2 

                line(doc, X1, Y1, X2, Y2, layer='레이져')        
                dim_linear(doc,  X1, Y1, X2, Y2, "", 50,  direction="up", layer='dim')    
                X3 = Abs_Xpos + LH2 
                if (SWAngle == 90):
                    if(C1_down > 0): 
                        Y3 = Abs_Ypos + JB + SW1  - narrow_Bending_rate * 3 - side_narrow_Vcut_rate
                    else:
                        Y3 = Abs_Ypos + JB + SW1  - narrow_Bending_rate * 3
                else:
                    if(C1_down > 0): 
                        Y3 = Abs_Ypos + JB + SW1  - narrow_Bending_rate * 2 - side_narrow_Vcut_rate    
                    else:
                        Y3 = Abs_Ypos + JB + SW1  - narrow_Bending_rate * 2 
                        
                lineto(doc,  X3, Y3, layer='레이져')    

            if (SWAngle<1):  # SWAngle 0 경우
                X1 = Abs_Xpos 
                if(C1_down > 0):   
                    Y1 = Abs_Ypos + JB + SW - narrow_Bending_rate * 2 - side_narrow_Vcut_rate    
                else:
                    Y1 = Abs_Ypos + JB + SW - narrow_Bending_rate * 2 
                X2 = Abs_Xpos + LH2 
                Y2 = Y1

                line(doc, X1, Y1, X2, Y2, layer='레이져')        
                dim_linear(doc,  X1, Y1, X2, Y2, "", 50,  direction="up", layer='dim')    
                X3 = X2            
                Y3 = Y2
                lineto(doc,  X3, Y3, layer='레이져')    
            X4 = Abs_Xpos + LH2 
            if(C1_down > 0):            
                Y4 = Abs_Ypos + JB  - narrow_Bending_rate  - side_narrow_Vcut_rate                        
            else:            
                Y4 = Abs_Ypos + JB  - narrow_Bending_rate

            if(args.narrow_sillcut>0):
                moveY = Y4 - args.narrow_sillcut
                lineto(doc,  X4, moveY, layer='레이져')
            else:
                lineto(doc,  X4, Y4, layer='레이져')
            X5 = Abs_Xpos + LH1             
            Y5 = Abs_Ypos                     
            if(C1_down > 0):
                lineto(doc,  X5, Y5, layer='레이져')   
                # A,B값을 주고 사각 C값 따임하는 부분 구현함
                if(SideBoxcut_Left_A>0):
                    X6 = Abs_Xpos + LH1            
                    if(C1_down > 0):    
                        Y6 = Abs_Ypos - C1_down + side_narrow_Vcut_rate                 
                    else:
                        Y6 = Abs_Ypos 
                    X7 = Abs_Xpos 
                    if(C1_down > 0):    
                        Y7 = Abs_Ypos - C1_up + side_narrow_Vcut_rate 
                    else:
                        Y7 = Abs_Ypos 
                    if(C1_down > 0):       
                        lineto(doc, Abs_Xpos + LH1 ,Abs_Ypos - C1_down + side_narrow_Vcut_rate + SideBoxcut_Left_A , layer='레이져')
                        lineto(doc, Abs_Xpos + LH1 - SideBoxcut_Left_B  ,Abs_Ypos - C1_down + side_narrow_Vcut_rate + SideBoxcut_Left_A , layer='레이져')
                        lineto(doc, Abs_Xpos + LH1 - SideBoxcut_Left_B  ,Abs_Ypos - C1_down + side_narrow_Vcut_rate , layer='레이져')
                    else:
                        lineto(doc, Abs_Xpos + LH1 ,Abs_Ypos + SideBoxcut_Left_A , layer='레이져')
                        lineto(doc, Abs_Xpos + LH1 - SideBoxcut_Left_B  ,Abs_Ypos  + SideBoxcut_Left_A , layer='레이져')
                        lineto(doc, Abs_Xpos + LH1 - SideBoxcut_Left_B  ,Abs_Ypos  , layer='레이져')                        

                else:
                    X6 = Abs_Xpos + LH1      
                    if(C1_down > 0):          
                        Y6 = Abs_Ypos - C1_down + side_narrow_Vcut_rate 
                    else:
                        Y6 = Abs_Ypos 

                    lineto(doc,  X6, Y6, layer='레이져')
                    X7 = Abs_Xpos 
                    if(C1_up > 0):          
                        Y7 = Abs_Ypos - C1_up + side_narrow_Vcut_rate                 
                    else:
                        Y7 = Abs_Ypos 
            else:
                lineto(doc,  X5, Y5, layer='레이져')   
                X6 = X5
                Y6 = Y5
                X7 = Abs_Xpos 
                Y7 = Abs_Ypos 

            lineto(doc,  X7, Y7, layer='레이져')                    
            
            X8 = Abs_Xpos 
            Y8 = Abs_Ypos             
            lineto(doc,  X8, Y8, layer='레이져')
            X9 = Abs_Xpos         
            if(C1_up > 0):            
                Y9 = Abs_Ypos + JB  - narrow_Bending_rate  - side_narrow_Vcut_rate            
            else:            
                Y9 = Abs_Ypos + JB  - narrow_Bending_rate            
            lineto(doc,  X9, Y9, layer='레이져')

            X10 = Abs_Xpos 
            if (SWAngle>1):  # SWAngle 0 이상인 경우 90도 추영덕소장 모양이 포함됨    
                if (SWAngle == 90):
                    if(C1_up > 0):            
                        Y10 = Abs_Ypos + JB + SW1  - narrow_Bending_rate * 3 - side_narrow_Vcut_rate
                    else: 
                        Y10 = Abs_Ypos + JB + SW1  - narrow_Bending_rate * 3 
                else:
                    if(C1_up > 0):            
                        Y10 = Abs_Ypos + JB + SW1  - narrow_Bending_rate * 2 - side_narrow_Vcut_rate    
                    else:
                        Y10 = Abs_Ypos + JB + SW1  - narrow_Bending_rate * 2

                lineto(doc,  X10, Y10, layer='레이져')
            else:
                if(C1_up > 0):  
                    Y10 = Abs_Ypos + JB + SW - narrow_Bending_rate * 2 - side_narrow_Vcut_rate   
                else:
                    Y10 = Abs_Ypos + JB + SW - narrow_Bending_rate * 2 
            
            lineto(doc,  X1, Y1, layer='레이져')
            
            # 쪽쟘 좌기둥 절곡선
            if(SW1>0):
                line(doc, X10, Y10, X3, Y3,  layer='22')
            line(doc, X9, Y9, X4, Y4,  layer='22')                    
            if(C1_up>0):
                line(doc, X8, Y8, X5, Y5,  layer='22')            
                if(narrow_Vcut_rate != narrow_Bending_rate):    # Vcut없는 경우도 적용할 수 있도록 로직개발함     
                    # vcut 지시선
                    dim_leader_line(doc, X8+150, Y8, X8+250, Y8+30,  "V-cut", layer="dim", style='0')        
                    
            # 쪽쟘 좌기둥 우측 절곡치수선
            if(SW1>0): # 헤밍구조        
                dim_vertical_right(doc, X3, Y3, X2,  Y2, 170 , 'dim', text_height=0.22,  text_gap=0.07)      
            dim_vertical_right(doc, X3, Y3, X4,  Y4, extract_abs(X3,X4) + 100, 'dim', text_height=0.22,  text_gap=0.07)      
            dim_vertical_right(doc, X5, Y5, X4,  Y4, 60, 'dim', text_height=0.22,  text_gap=0.07)  
            if(C1_down>0):
                dim_vertical_right(doc, X6, Y6, X5,  Y5, extract_abs(X5,X6) + 100, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X6, Y6, X2,  Y2, 60 + 170 , 'dim', text_height=0.22,  text_gap=0.07)      

            # 좌기둥 좌측 절곡치수선        
            if(SW1>0): # 헤밍구조         
                dim_vertical_left(doc, X1, Y1, X10,  Y10, 170, 'dim', text_height=0.22, text_gap=0.01)  
            dim_vertical_left(doc, X10, Y10, X9, Y9, 80, 'dim', text_height=0.22, text_gap=0.01)  
            dim_vertical_left(doc, X9, Y9, X8,  Y8, 30, 'dim', text_height=0.22, text_gap=0.01)      
            if(C1_down>0):        
                dim_vertical_left(doc, X8, Y8, X7,  Y7, 80, 'dim', text_height=0.22, text_gap=0.01)                                          
            dim_vertical_left(doc, X1, Y1, X7,  Y7, 60 + 170 , 'dim', text_height=0.22, text_gap=0.01)      

            #기둥 하부 치수선 2    
            dim_linear(doc,  X6, Y6, X7, Y7, "", 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            if LH1>LH2:
                dim_linear(doc,  X2, Y2, X5, Y5, "",  50 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)       
            if LH1<LH2:    
                dim_linear(doc,  X4, Y4, X6, Y6, "", extract_abs(Y4,Y6) + 100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            if args.opt2:
                drawcircle(doc, X1 + 12.5, Y1 - 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X2 - 12.5, Y2 - 5, 2.5 , layer='레이져') # 도장홀 5파이        

            # 좌기둥 
            # 전개도에 문구            
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"
            # Textstr 생성
            Textstr = f"({Pre_text}{str(Floor_des)}) 좌"       

            X2 = Abs_Xpos + (LH2)/1.1 - len(Textstr)*30/2 - 300
            
            Y2 = Abs_Ypos + (JB + C1_up + SW )/2  - 30/2          
            
            draw_Text(doc, X2, Y2, 30, str(Textstr), '레이져')

            X2 = Abs_Xpos + 50 
            Y2 = Abs_Ypos + (JB + C1_up + SW )/2  - 20/2          
            Textstr = f"(상부)"       
            draw_Text(doc, X2, Y2, 20, str(Textstr), '0')

            #############################################################################################################
            # 쪽쟘 좌기둥 좌측 단면도
            #############################################################################################################

            Section_Xpos = Abs_Xpos - 400
            Section_Ypos = Abs_Ypos + JB

            # 가정: SWAngle은 도 단위로 주어진 각도
            SWAngle_radians = math.radians(SWAngle)

            if (SWAngle>1):  # SWAngle 0 이상인 경우 90도 추영덕소장 모양이 포함됨
                X1 = Section_Xpos - SW1 - SW2 * math.cos(SWAngle_radians)
                Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                X2 = Section_Xpos - SW1
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos            
                dim_linear(doc, X1, Y1, X2, Y2,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)                   
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)                  
                line(doc, X2, Y2, X1, Y1, layer='0')                
                line(doc, X2, Y2, X3, Y3, layer='0')   
                
            if (SWAngle<1):
                X1 = Section_Xpos - SW
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos            
                line(doc,  X1, Y1, X2, Y2, layer='0')                
                dim_linear(doc, X1, Y1, X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)                  
        
            # 각도 90도가 아니면 각도 표기
            if (SWAngle > 0 and SWAngle != 90):
                dim_angular(doc, X2 , Y2, X1, Y1, X2-30, Y2, X3-30, Y3,  50, direction="left" )

            X4 = Section_Xpos 
            Y4 = Section_Ypos - JB   
            lineto(doc,  X4, Y4, layer='0')    

            dim_vertical_right(doc, X3, Y3, X4,  Y4, 50, 'dim', text_height=0.22,  text_gap=0.07)          
            if(C1_up > 0):
                X5 = Section_Xpos - C1_up
                Y5 = Section_Ypos - JB 
                lineto(doc,  X5, Y5, layer='0')   
                #쪽쟘 하부 치수선 1개
                dim_linear(doc, X5, Y5, X4, Y4, "", 80,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                if(narrow_Vcut_rate != narrow_Bending_rate):    # Vcut없는 경우도 적용할 수 있도록 로직개발함     
                    # vcut 표기 1개소            
                    drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
                    # vcut 지시선
                    dim_leader_line(doc, X4, Y4 , X4 + 50, Y4 - 50, "V-cut", layer="dim", style='0')            
            
            #############################################################################################################
            # 쪽쟘 좌기둥 우측 단면도
            #############################################################################################################
            if(C1_up!=C1_down):        
                Section_Xpos = Abs_Xpos + LH1 + 400
                if (SWAngle>1):  # SWAngle 0 이상인 경우 90도 추영덕소장 모양이 포함됨                                    
                    X1 = Section_Xpos + SW1 + SW2 * math.cos(SWAngle_radians)
                    Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)
                    X2 = Section_Xpos + SW1
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc, X2, Y2, X1, Y1, layer='0')                   
                    line(doc, X2, Y2, X3, Y3, layer='0')        
                    dim_linear(doc, X2, Y2, X1, Y1,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                if (SWAngle<1):
                    X1 = Section_Xpos + SW
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos 
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos 
                    line(doc,  X1, Y1, X2, Y2, layer='0')                
                    dim_linear(doc, X2, Y2, X1, Y1,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)                  

                # 각도 90도가 아니면 각도 표기
                if (SWAngle > 0 and SWAngle != 90):              
                    dim_angular(doc, X2 + 30 , Y2, X3 + 30 , Y3, X2 , Y2, X1, Y1,   80, direction="right" )                                  

                X4 = Section_Xpos 
                Y4 = Section_Ypos - JB   
                lineto(doc,  X4, Y4, layer='0')    

                if(C1_down > 0):            
                    X5 = Section_Xpos + C1_down
                    Y5 = Section_Ypos - JB 
                    lineto(doc,  X5, Y5, layer='0')   
                    if(narrow_Vcut_rate != narrow_Bending_rate):    # Vcut없는 경우도 적용할 수 있도록 로직개발함     
                        # vcut 표기                
                        drawcircle(doc, X4, Y4 , 5 , layer='0', color='6') 
                        # vcut 지시선
                        dim_leader_line(doc, X4, Y4 , X4-130, Y4-50, "V-cut", layer="dim", style='0')
                    # 우측 절곡치수선
                    dim_vertical_left(doc, X3, Y3, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)                      
                    #쪽쟘 하부 치수선 1개
                    dim_linear(doc,  X5, Y5, X4, Y4,  "", 80,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)
                
            ###############################################################################################################################################################################
            # 쪽쟘 우기둥 전개도 및 단면도 (해밍구조 포함된 통합버전)
            ###############################################################################################################################################################################                
            Abs_Ypos -= 400

            if(C2_up<1):
                side_narrow_Vcut_rate = 0
            else:
                side_narrow_Vcut_rate = narrow_Vcut_rate            

            if (C2_up>0):  
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + C2_up - side_narrow_Vcut_rate   

                if(SideBoxcut_Right_A > 0):
                    X2 = Abs_Xpos + RH1
                    Y2 = Abs_Ypos + C2_down - side_narrow_Vcut_rate                       
                    line(doc, X1, Y1, Abs_Xpos + RH1 - SideBoxcut_Right_B, Y2, layer='레이져')        
                    lineto(doc,  Abs_Xpos + RH1 - SideBoxcut_Right_B, Y2 - SideBoxcut_Right_A, layer='레이져')                     
                    lineto(doc,  Abs_Xpos + RH1, Y2 - SideBoxcut_Right_A, layer='레이져')                     
                    dim_linear(doc,  X1, Y1, X2, Y2, "", 50,  direction="up", layer='dim')    
                    X3 = Abs_Xpos + RH1
                    Y3 = Abs_Ypos 
                    lineto(doc,  X3, Y3, layer='레이져')                     

                else:
                    X2 = Abs_Xpos + RH1
                    Y2 = Abs_Ypos + C2_down - side_narrow_Vcut_rate       
                    line(doc, X1, Y1, X2, Y2, layer='레이져')        
                    dim_linear(doc,  X1, Y1, X2, Y2, "", 50,  direction="up", layer='dim')    
                    X3 = Abs_Xpos + RH1
                    Y3 = Abs_Ypos 
                    lineto(doc,  X3, Y3, layer='레이져')         

            if (C2_up<1):  
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos 
                X2 = Abs_Xpos + RH1
                Y2 = Abs_Ypos 
                line(doc, X1, Y1, X2, Y2, layer='레이져')        
                dim_linear(doc,  X1, Y1, X2, Y2, "", 50,  direction="up", layer='dim')    
                X3 = X2
                Y3 = Y2

            if (C2_down>0):                         
                X4 = Abs_Xpos + RH2 
                Y4 = Abs_Ypos - JB + narrow_Bending_rate  + side_narrow_Vcut_rate            
            if (C2_down<1):             
                X4 = Abs_Xpos + RH2 
                Y4 = Abs_Ypos - JB + narrow_Bending_rate              

            if(args.narrow_sillcut>0):
                moveY = Y4 + args.narrow_sillcut
                lineto(doc,  X4, moveY, layer='레이져')
            else:
                lineto(doc,  X4, Y4, layer='레이져')            

            X5 = Abs_Xpos + RH2 
            if (SWAngle > 0 ):
                if (SWAngle > 0 and SWAngle == 90):
                    if (C2_down>0):   
                        Y5 = Abs_Ypos - JB - SW1  + narrow_Bending_rate * 3 + side_narrow_Vcut_rate
                    else:
                        Y5 = Abs_Ypos - JB - SW1  + narrow_Bending_rate * 3 
                else:
                    if (C2_down>0):  
                        Y5 = Abs_Ypos - JB - SW1  + narrow_Bending_rate * 2 + side_narrow_Vcut_rate   
                    else:
                        Y5 = Abs_Ypos - JB - SW1  + narrow_Bending_rate * 2 
                lineto(doc,  X5, Y5, layer='레이져')    
            else:
                if (C2_down>0):  
                    Y5 = Abs_Ypos  - JB - SW  + narrow_Bending_rate * 2 + side_narrow_Vcut_rate  
                else:
                    Y5 = Abs_Ypos  - JB - SW  + narrow_Bending_rate * 2 
                lineto(doc,  X5, Y5, layer='레이져')   

            X6 = Abs_Xpos + RH2 
            if (SWAngle > 0 ):        
                if (SWAngle > 0 and SWAngle == 90):
                    if (C2_down>0):  
                        Y6 = Abs_Ypos - JB - SW1 - SW2 + narrow_Bending_rate * 4 + side_narrow_Vcut_rate
                    else:
                        Y6 = Abs_Ypos - JB - SW1 - SW2 + narrow_Bending_rate * 4 
                else:
                    if (C2_down>0):  
                        Y6 = Abs_Ypos - JB - SW1 - SW2 + narrow_Bending_rate * 2 + side_narrow_Vcut_rate                           
                    else:
                        Y6 = Abs_Ypos - JB - SW1 - SW2 + narrow_Bending_rate * 2
                lineto(doc,  X6, Y6, layer='레이져')    
            else:
                Y6 = Y5
            
            X7 = Abs_Xpos 
            if (SWAngle > 0 ):  
                if (SWAngle > 0 and SWAngle == 90):
                    Y7 = Abs_Ypos - JB - SW1 - SW2 + narrow_Bending_rate * 4 + side_narrow_Vcut_rate
                else:
                    Y7 = Abs_Ypos - JB - SW1 - SW2 + narrow_Bending_rate * 2 + side_narrow_Vcut_rate                                       
            else:
                Y7 = Abs_Ypos - JB - SW + narrow_Bending_rate * 2 + side_narrow_Vcut_rate

            lineto(doc,  X7, Y7, layer='레이져')
            X8 = Abs_Xpos 
            if (SWAngle > 0 ):  
                if (SWAngle == 90):
                    Y8 = Abs_Ypos - JB - SW1  + narrow_Bending_rate * 3 + side_narrow_Vcut_rate
                else:
                    Y8 = Abs_Ypos - JB - SW1  + narrow_Bending_rate * 2 + side_narrow_Vcut_rate   
            else:
                Y8 = Abs_Ypos  - JB - SW  + narrow_Bending_rate * 2 + side_narrow_Vcut_rate

            lineto(doc,  X8, Y8, layer='레이져')
            X9 = Abs_Xpos 
            Y9 = Abs_Ypos - JB + narrow_Bending_rate  + side_narrow_Vcut_rate
            lineto(doc,  X9, Y9, layer='레이져')
            X10 = Abs_Xpos 
            Y10 = Abs_Ypos 
            lineto(doc,  X10, Y10, layer='레이져')
            if(C2_up > 0): 
                lineto(doc,  X1, Y1, layer='레이져')        
            
            # 쪽쟘 우기둥  절곡선
            if(C2_down > 0): 
                line(doc, X10, Y10, X3, Y3,  layer='22')                                    
                if(narrow_Vcut_rate != narrow_Bending_rate):    # Vcut없는 경우도 적용할 수 있도록 로직개발함     
                    # vcut 지시선
                    dim_leader_line(doc, X10+200, Y10, X10+200 + 50,Y10 - 40, "V-cut", layer="dim", style='0')                
            line(doc, X9, Y9, X4, Y4,  layer='22')            
            if(SW1>0): # 헤밍구조    
                line(doc, X8, Y8, X5, Y5,  layer='22')

            # 쪽쟘 우기둥 우측 절곡치수선
            if(C2_down>0):  
                dim_vertical_right(doc, X3, Y3, X2,  Y2, 60 + 50, 'dim', text_height=0.22,  text_gap=0.07)      
            dim_vertical_right(doc, X3, Y3, X4,  Y4, 60, 'dim', text_height=0.22,  text_gap=0.07)      
            dim_vertical_right(doc, X5, Y5, X4,  Y4, 60 + 50, 'dim', text_height=0.22,  text_gap=0.07)  
            if(SW1>0): # 헤밍구조              
                dim_vertical_right(doc, X6, Y6, X5,  Y5, 170, 'dim', text_height=0.22,  text_gap=0.07)              
            dim_vertical_right(doc, X6, Y6, X2,  Y2, 60 + 170 , 'dim', text_height=0.22,  text_gap=0.07)      

            # 쪽쟘 우기둥 좌측 절곡치수선
            if(C2_up>0):          
                dim_vertical_left(doc, X1, Y1, X10,  Y10, 80, 'dim', text_height=0.22, text_gap=0.01)  
            dim_vertical_left(doc, X10, Y10, X9, Y9, 30, 'dim', text_height=0.22, text_gap=0.01)  
            dim_vertical_left(doc, X9, Y9, X8,  Y8, 80, 'dim', text_height=0.22, text_gap=0.01)                  
            if(SW1>0): # 헤밍구조                      
                dim_vertical_left(doc, X8, Y8, X7,  Y7, 170, 'dim', text_height=0.22, text_gap=0.01)                  
            dim_vertical_left(doc, X1, Y1, X7,  Y7, 60 + 170 , 'dim', text_height=0.22, text_gap=0.01)      

            #기둥 하부 치수선 2    
            dim_linear(doc,  X6, Y6, X7, Y7, "", 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
            if RH1<RH2:
                dim_linear(doc,   X2, Y2, X5, Y5, "",  45 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)       
            if RH1>RH2:    
                dim_linear(doc,  X3, Y3, X6, Y6, "", extract_abs(Y3,Y6) + 100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            if args.opt2:
                drawcircle(doc, X7 + 12.5, Y7 + 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X6 - 12.5, Y6 + 5, 2.5 , layer='레이져') # 도장홀 5파이
            
            # 전개도 문구
            # Floor_mc 값이 없거나 0이거나 공백인 경우
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"
            # Textstr 생성
            Textstr = f"({Pre_text}{str(Floor_des)}) 우 "       

            X2 = Abs_Xpos + (RH2)/1.1 - len(Textstr)*30/2 - 300
            Y2 = Abs_Ypos + C2_up - (JB + C2_up + SW )/2  - 30/2      
            draw_Text(doc, X2, Y2, 30, str(Textstr), '레이져')

            X2 = Abs_Xpos + 50 
            Y2 = Abs_Ypos  + C2_up  - (JB + C2_up + SW )/2  - 20/2      
            Textstr = f"(상부)"       
            draw_Text(doc, X2, Y2, 20, str(Textstr), '0')

            ###########################################################################################################################################################################
            # 쪽쟘 우기둥 좌측 단면도    
            ###########################################################################################################################################################################

            Section_Xpos = Abs_Xpos - 400
            Section_Ypos = Abs_Ypos 
        
            if(C2_up>0):
                X1 = Section_Xpos - C2_up
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos

                # 좌측 단면도 
                line(doc, X1, Y1, X2, Y2, layer='0')        
                dim_linear(doc,  X1, Y1, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                X3 = Section_Xpos 
                Y3 = Section_Ypos - JB   
                lineto(doc,  X3, Y3, layer='0')    
                if(narrow_Vcut_rate != narrow_Bending_rate):    # Vcut없는 경우도 적용할 수 있도록 로직개발함     
                    # vcut 표기 1개소            
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    # vcut 지시선
                    dim_leader_line(doc, X2, Y2 , X2+50, Y2+50, "V-cut", layer="dim", style='0')

            if(C2_up<1):
                X1 = Section_Xpos 
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos
                X3 = Section_Xpos 
                Y3 = Section_Ypos - JB   
                line(doc, X1, Y1, X3, Y3, layer='0')      
                
            if(SW1>0): # 헤밍구조                
                X4 = Section_Xpos - SW1
                Y4 = Section_Ypos - JB 
                lineto(doc,  X4, Y4, layer='0')   
                X5 = Section_Xpos - SW1 - SW2 * math.cos(SWAngle_radians)
                Y5 = Section_Ypos - JB + SW2 * math.sin(SWAngle_radians)    
                lineto(doc,  X5, Y5, layer='0')   
                # 10미리 aligned
                dim_linear(doc,  X4, Y4, X5, Y5,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)    

            if(SW1<1): # 헤밍구조 아닌경우
                X4 = Section_Xpos - SW
                Y4 = Section_Ypos - JB 
                lineto(doc,  X4, Y4, layer='0')   
                X5 = X4
                Y5 = Y4
                lineto(doc,  X5, Y5, layer='0')       

            if (SWAngle> 0 and SWAngle != 90):
                dim_angular(doc,  X4 - 30, Y4, X3 - 30, Y3, X5, Y5, X4, Y4,  50, direction="left" )

            # 우측 절곡치수선
            dim_vertical_right(doc, X2, Y2, X3,  Y3, 50, 'dim', text_height=0.22,  text_gap=0.07)  

            # 뒷날개 치수선 1개
            dim_linear(doc,  X3, Y3, X4, Y4, "", 50,  direction="down", layer='dim' ,text_height=0.22,  text_gap=0.07)      
            
            # side 우측 단면도    
            if(C2_up!=C2_down):             
                Section_Xpos = Abs_Xpos + RH1 + 400
                if(C2_down>0):
                    X1 = Section_Xpos + C2_down
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos 
                    Y2 = Section_Ypos
                    # 우기둥 우측 단면도 U1값
                    line(doc, X1, Y1, X2, Y2, layer='0')        
                    dim_linear(doc,  X2, Y2,  X1, Y1, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos - JB   
                    lineto(doc,  X3, Y3, layer='0')
                    if(narrow_Vcut_rate != narrow_Bending_rate):    # Vcut없는 경우도 적용할 수 있도록 로직개발함     
                        # vcut 표기                
                        drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                        # vcut 지시선
                        dim_leader_line(doc, X2, Y2 , X2-130, Y2+50, "V-cut", layer="dim", style='0')                

                if(C2_down<1):
                    X1 = Section_Xpos
                    Y1 = Section_Ypos
                    X2 = Section_Xpos
                    Y2 = Section_Ypos
                    X3 = Section_Xpos
                    Y3 = Section_Ypos - JB
                    lineto(doc,  X3, Y3, layer='0')

                if(SW1>0): # 헤밍구조      
                    X4 = Section_Xpos + SW1
                    Y4 = Section_Ypos - JB 
                    lineto(doc,  X4, Y4, layer='0')   
                    X5 = Section_Xpos + SW1 + SW2 * math.cos(SWAngle_radians)
                    Y5 = Section_Ypos - JB + SW2 * math.sin(SWAngle_radians)    
                    lineto(doc,  X5, Y5, layer='0')   
                    # 10미리 aligned
                    dim_linear(doc,   X5, Y5, X4, Y4, "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)                      

                    # 각도 90도가 아니면 각도 표기
                    if (SWAngle != 90):               
                        dim_angular(doc,   X4 , Y4, X5, Y5, X3 + 40 , Y3, X4 + 40, Y4 , 60, direction="right" )
                    # 우측 절곡치수선
                    dim_vertical_left(doc, X2, Y2, X3,  Y3, 50, 'dim', text_height=0.22,  text_gap=0.07)  
                            
                    #쪽쟘 하부 치수선 1개
                    dim_linear(doc,   X4, Y4, X3, Y3, "", 50,  direction="down", layer='dim' ,text_height=0.22,  text_gap=0.07)     

                if(SW1<1): # 헤밍구조      
                    X4 = Section_Xpos + SW
                    Y4 = Section_Ypos - JB 
                    lineto(doc,  X4, Y4, layer='0')   
                    X5 = X4
                    Y5 = X4
                    # 우측 절곡치수선
                    dim_vertical_left(doc, X2, Y2, X3,  Y3, 50, 'dim', text_height=0.22,  text_gap=0.07)                          
                    #쪽쟘 하부 치수선 1개
                    dim_linear(doc,   X4, Y4, X3, Y3, "", 50,  direction="down", layer='dim' ,text_height=0.22,  text_gap=0.07)      
            
            Abs_Ypos -= 600
 
#############################
# SO멍텅구리
#############################    
def execute_sonormal():     
    # 시트 선택 (시트명을 지정)
    sheet_name = 'SO멍텅제작'  # 원하는 시트명으로 변경
    sheet = workbook[sheet_name]

    # 2차원 배열을 저장할 리스트 생성
    data_2d = []

    # 1행부터 20행까지의 값을 2차원 배열에 저장
    for row_num in range(7, sheet.max_row + 1):  # 7행부터 20행까지 반복
        cell_value = sheet.cell(row=row_num, column=2).value  # 2열의 값 가져오기
        if cell_value is not None and cell_value != 'F':  # 값이 None이 아닌 경우에만 배열에 추가
            row_values =  [sheet.cell(row=row_num, column=col_num).value for col_num in range(1, sheet.max_column + 1)]        
            data_2d.append(row_values)

    # 엑셀 파일 닫기
    workbook.close()

    if not data_2d:
        print("해당 데이터가 없습니다.")
        return  # 함수를 빠져나감    
    

    ########################################################################################################################################################################
    # SO멍텅구리 상판 전개도
    ########################################################################################################################################################################

    Abs_Xpos = 43000
    Abs_Ypos = 0

    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        OP1 = row[2]
        OP2 = row[3]
        R = row[4]
        JD1 = row[5]
        JD2 = row[6]
        JB1_up = row[7]
        JB1_down = row[8]
        JB2_up = row[9]
        JB2_down = row[10]
        Top_gap = row[11]
        Side_gap = row[12]
        H1 = row[13]
        H2 = row[14]
        C1 = row[15]        
        C2 = row[16]
        A1 = row[17]
        A2 = row[18]
        LH1 = row[19]
        LH2 = row[20]
        RH1 = row[21]
        RH2 = row[22]        
        U = row[23]        
        G = row[24]        
        UW = row[25]        
        SW_Left = row[26]        
        SW_Right = row[27]        
        K1_up = row[28]        
        K1_down = row[29]        
        K2_up = row[30]        
        K2_down = row[31]                
        Upper_size = row[32]
        Left_side_size = row[33]
        Right_side_size = row[34]
        SW1 = row[35]
        SW2 = row[36]
        SWAngle = row[37]
        E1 = row[38]
        E2 = row[39]        
        Angle = row[43]
        Right_Angle = row[44]
        Top_pop1 = round(row[45] + 2, 1) #소수점 첫째자리까지만 나오게
        Top_pop2 = round(row[46] + 2, 1)         
        Kadapt =  row[56]   # K값 J값에 적용여부 넣기 1이면 넣기 '늘치J에 포함여부'
        G_Setvalue = row[57]  # 손상민 소장 G값 강제로 적용 작지 52번째

        # 상판 E1, E2는 추영덕 소장이 K값을 상판에 표현하는 수치표현식인데, K값으로 환산해서 정리함
        #  K1_up = E1 - UW - 1.2

        # 데이터 검증 및 기본값 설정
        SWAngle = validate_or_default(SWAngle)
        K1_up = validate_or_default(K1_up)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_down = validate_or_default(K2_down)
        U = validate_or_default(U)
        G = validate_or_default(G)        
        G_Setvalue = validate_or_default(G_Setvalue)        
        OP1 = validate_or_default(OP1)
        OP2 = validate_or_default(OP2)
        Kadapt = validate_or_default(Kadapt)

        # print("작업소장 : ")
        # print(worker)
        # print("G_setvalue : ")
        # print(G_Setvalue)        
        if(worker=="손상민" and G_Setvalue > 0) :
            # 상부 돌출
            Top_pop1 = G_Setvalue
            Top_pop2 = G_Setvalue
            # print("G_setvalue : ")
            # print(G_Setvalue)
       
        if(Kadapt==1):
            K1_up = 0            
            K2_up = 0    
            if(E1>0):
                E1 = E1 + Vcut_rate
                E2 = E2 + Vcut_rate
        else:
            if(E1>0):                
                K1_up = E1 
                K2_up = E2                        
             
        if(K1_up>=10 and K2_up>=10 ):
            # 추영덕소장 E1, E2값 무시함 위의 조건이라면
            E1 = 0
            E2 = 0

        if(Top_pop1<4):
            Top_pop1 = 4
        if(Top_pop2<4):
            Top_pop2 = 4                

        if(OP1>0):             
            if(G > 0):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U + G - Vcut_rate*3
                X2 = Abs_Xpos + H1 + H2 + R 
                Y2 = Y1      
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Abs_Ypos + U - Vcut_rate * 2           
                line(doc, X2, Y2, X3, Y3, layer='레이져')          
            if(G < 1):
                X1 = Abs_Xpos 
                Y1 = Abs_Ypos + U  - Vcut_rate
                X2 = Abs_Xpos + H1 + H2 + R 
                Y2 = Y1      
                line(doc, X1, Y1, X2, Y2, layer='레이져')       
                X3 = X2
                Y3 = Y2

            X4 = X3
            Y4 = Abs_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')
            X5 = X4
            Y5 = Abs_Ypos - Top_pop2 + Vcut_rate        
            lineto(doc,  X5, Y5, layer='레이져')    
            X6 = X5 - H2
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='레이져')

            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                X7 = Abs_Xpos + (R - OP1 - OP2 - 2)/2  + OP1 + OP2 + 2 + H1    
                if(Kadapt==1 and E2>0):
                    Y7 = Abs_Ypos - (JD2-E2) + Vcut_rate*3 
                else:
                    Y7 = Abs_Ypos - JD2 + Vcut_rate*3 
                lineto(doc,  X7, Y7, layer='레이져')
            else:
                X7 = Abs_Xpos + (R - OP1 - OP2 - 2)/2  + OP1 + OP2 + 2 + H1        
                Y7 = Abs_Ypos - JD2 + Vcut_rate * 3
                lineto(doc,  X7, Y7, layer='레이져')  
            
            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                X8 = X7
                if(E2>0):
                    Y8 = Abs_Ypos - JD2 + Vcut_rate*3
                else:
                    Y8 = Abs_Ypos - JD2 - K2_up + Vcut_rate*3
                lineto(doc,  X8, Y8, layer='레이져') 
                X9 = X8 
                if(E2>0):                
                    Y9 = Abs_Ypos - JD2 - UW + Vcut_rate*5     
                else:
                    Y9 = Abs_Ypos - JD2 - K2_up - UW + Vcut_rate*5     
                lineto(doc,  X9, Y9, layer='레이져')   
                X10 = Abs_Xpos + (R - OP1 - OP2 - 2)/2  + OP1 +  2  + H1    
                if(E2>0):                
                    Y10 = Abs_Ypos - JD2 - UW + Vcut_rate*5     
                else:                    
                    Y10 = Abs_Ypos - JD2 - K2_up - UW + Vcut_rate*5     
                lineto(doc,  X10, Y10, layer='레이져')      

                # SO 좌 열림, 우 열림에 따라 다른값을 갖는다. SO는 이것이 핵심임
                # 좌측이 클때
                if(JD1+K1_up > JD2+K2_up):   
                    X11 = X10 
                    if(E2>0):                
                        Y11 =  Abs_Ypos - JD2 + Vcut_rate*3
                    else:                    
                        Y11 =  Abs_Ypos - JD2 - K2_up + Vcut_rate*3
                    lineto(doc,  X11, Y11, layer='레이져')                                              
                    X12 = X11 
                    Y12 = Y11 
                    lineto(doc,  X12, Y12, layer='레이져')                                      
                else:
                    X11 = X10 
                    if(E2>0):                
                        Y11 =  Abs_Ypos - JD2 + Vcut_rate*3
                    else:                    
                        Y11 =  Abs_Ypos - JD2 - K2_up + Vcut_rate*3
                    lineto(doc,  X11, Y11, layer='레이져')                                              
                    X12 = X11 
                    Y12 = Y11 + ( (JD2+K2_up) - (JD1+K1_up) )
                    lineto(doc,  X12, Y12, layer='레이져')                                            
                X13 = X12 - 2                
                Y13 =  Y12                
                lineto(doc,  X13, Y13, layer='레이져')         
                # SO 좌 열림, 우 열림에 따라 다른값을 갖는다. SO는 이것이 핵심임
                # 좌측이 클때
                if(JD1+K1_up > JD2+K2_up):     
                    X14 = X13
                    if(E1>0):                
                        Y14 =  Abs_Ypos - JD1 + Vcut_rate*3
                    else:                    
                        Y14 =  Abs_Ypos - JD1 - K1_up + Vcut_rate*3
                    lineto(doc,  X14, Y14, layer='레이져')                                                      
                    X15 = X14                
                    if(E1>0):                
                        Y15 =  Abs_Ypos - JD1 - UW + Vcut_rate*5   
                    else:                    
                        Y15 =  Abs_Ypos - JD1 - K1_up - UW + Vcut_rate*5
                    lineto(doc,  X15, Y15, layer='레이져')                                      
                else:
                    X14 = X13                
                    Y14 = Y13
                    lineto(doc,  X14, Y14, layer='레이져')                                                      
                    X15 = X14                
                    if(E1>0):                
                        Y15 =  Abs_Ypos - JD1 - UW + Vcut_rate*5   
                    else:                    
                        Y15 =  Abs_Ypos - JD1 - K1_up - UW + Vcut_rate*5
                    lineto(doc,  X15, Y15, layer='레이져')                                      

                X16 = Abs_Xpos + (R - OP1 - OP2 - 2)/2  + H1    
                Y16 = Y15 
                lineto(doc,  X16, Y16, layer='레이져')                                      
                X17 = X16
                if(E1>0):                
                    Y17 =  Abs_Ypos - JD1 + Vcut_rate*3
                else:                    
                    Y17 =  Abs_Ypos - JD1 - K1_up + Vcut_rate*3
                lineto(doc,  X17, Y17, layer='레이져')     

            else:
                # E1 적용 공식인데, K1, K2에 대한 내용 삭제된 것임
                X8 = X7
                Y8 = Abs_Ypos - JD2 + Vcut_rate*3                
                lineto(doc,  X8, Y8, layer='레이져') 
                X9 = X8                 
                Y9 = Abs_Ypos - JD2 - UW + Vcut_rate*5
                lineto(doc,  X9, Y9, layer='레이져')   
                X10 = Abs_Xpos + (R - OP1 - OP2 - 2)/2  + OP1 +  2  + H1    
                Y10 = Abs_Ypos - JD2 - UW + Vcut_rate*5                     
                lineto(doc,  X10, Y10, layer='레이져')      
                # SO 좌 열림, 우 열림에 따라 다른값을 갖는다. SO는 이것이 핵심임
                # 좌측이 클때
                if(JD1+K1_up > JD2+K2_up):   
                    X11 = X10 
                    Y11 =  Abs_Ypos - JD2 + Vcut_rate*3                    
                    lineto(doc,  X11, Y11, layer='레이져')                                              
                    X12 = X11 
                    Y12 = Y11 
                    lineto(doc,  X12, Y12, layer='레이져')
                else:
                    X11 = X10 
                    Y11 =  Abs_Ypos - JD2 + Vcut_rate*3                    
                    lineto(doc,  X11, Y11, layer='레이져')                                              
                    X12 = X11 
                    Y12 = Y11 + ( (JD2+K2_up) - (JD1+K1_up) )
                    lineto(doc,  X12, Y12, layer='레이져')                               
                X13 = X12 - 2                
                Y13 = Y12                
                lineto(doc,  X13, Y13, layer='레이져')    

                # SO 좌 열림, 우 열림에 따라 다른값을 갖는다. SO는 이것이 핵심임
                # 좌측이 클때
                if(JD1+K1_up > JD2+K2_up):     
                    X14 = X13
                    Y14 =  Abs_Ypos - JD1 + Vcut_rate*3                    
                    lineto(doc,  X14, Y14, layer='레이져')                                                      
                    X15 = X14                
                    Y15 =  Abs_Ypos - JD1 - UW + Vcut_rate*5                       
                    lineto(doc,  X15, Y15, layer='레이져')                                      
                else:
                    X14 = X13                
                    Y14 = Y13
                    lineto(doc,  X14, Y14, layer='레이져')                                                      
                    X15 = X14                
                    Y15 =  Abs_Ypos - JD1 - UW + Vcut_rate*5   
                    lineto(doc,  X15, Y15, layer='레이져')    

                X16 = Abs_Xpos + (R - OP1 - OP2 - 2)/2  + H1    
                Y16 = Y15 
                lineto(doc,  X16, Y16, layer='레이져')                                      
                X17 = X16
                Y17 =  Abs_Ypos - JD1 + Vcut_rate*3
                lineto(doc,  X17, Y17, layer='레이져')                                                      

            X18 = X17     
            if(K1_up>0 or K2_up>0):
                Y18 = Abs_Ypos - JD1 + Vcut_rate * 3
            else:  
                if(Kadapt==1 and E1>0):
                    Y18 = Abs_Ypos - (JD1-E1) + Vcut_rate * 3
                else:   
                    Y18 = Abs_Ypos - JD1 + Vcut_rate * 3
            
            lineto(doc,  X18, Y18, layer='레이져')                   
            X19 = Abs_Xpos + H1
            Y19 = Abs_Ypos - Top_pop2 + Vcut_rate 
            lineto(doc,  X19, Y19, layer='레이져')      
            X20 = Abs_Xpos
            Y20 = Y19
            lineto(doc,  X20, Y20, layer='레이져')      
            X21 = X20 
            Y21 = Abs_Ypos
            lineto(doc,  X21, Y21, layer='레이져')     
            if(G > 0):         
                X22 = X21 
                Y22 = Abs_Ypos + U - Vcut_rate * 2   
                lineto(doc,  X22, Y22, layer='레이져')      
                lineto(doc,  X1, Y1, layer='레이져')                  
            if(G < 1):         
                X22 = X21 
                Y22 = Abs_Ypos + U - Vcut_rate 
                lineto(doc,  X1, Y1, layer='레이져')                  

            if args.opt11:
                drawcircle(doc, X16 + 12.5, Y16 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X9 - 12.5, Y9 + 7.5, 2.5 , layer='레이져') # 도장홀 5파이

            # 상판 절곡선
            if(G>0):
                line(doc, X22, Y22, X3, Y3,  layer='22')        
            line(doc, X21, Y21, X4, Y4,  layer='22')            

            # 노칭부분 치수선 및 절곡선 좌측이 클때
            if( JD1 + K1_up > JD2 + K2_up ):     
                line(doc, X17, Y17, X14, Y14,  layer='22')
                line(doc, X12, Y12, X8, Y8,  layer='22')
                # 도어두께 표시하는 치수선
                dim_vertical_right(doc, X15, Y15, X10, Y10, 100, 'dim', text_height=0.22,  text_gap=0.07)  
                dim_vertical_left(doc, X13, Y13, X14, Y14, 100, 'dim', text_height=0.22,  text_gap=0.07)                  
            else:
                line(doc, X17, Y17, X14, Y14,  layer='22')
                line(doc, X11, Y11, X8, Y8,  layer='22')            
                # 도어두께 표시하는 치수선
                dim_vertical_left(doc, X15, Y15, X10, Y10, 100, 'dim', text_height=0.22,  text_gap=0.07)  
                dim_vertical_right(doc, X12, Y12, X11, Y11, 100, 'dim', text_height=0.22,  text_gap=0.07)                
    
            # 상판 우측 절곡치수선
            if(G>0):        
                dim_vertical_right(doc, X2, Y2, X3, Y3,  100, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X4, Y4, X3, Y3,  50, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc, X4, Y4, X8, Y8,  50, 'dim', text_height=0.22, text_gap=0.07)  
            dim_vertical_right(doc,  X9, Y9, X8, Y8, extract_abs(X8,X5) + 50, 'dim', text_height=0.22, text_gap=0.07)   # 뒷날개
            dim_vertical_right(doc, X2, Y2,  X9, Y9,  extract_abs(X9,X2) + 100, 'dim', text_height=0.22, text_gap=0.07)             

            # 상판 좌측 절곡치수선
            if(JD1 != JD2 or K1_up != K2_up):
                if(G>0):                    
                    dim_vertical_left(doc, X1, Y1, X22, Y22,  extract_abs(X1,X22) + 100, 'dim', text_height=0.22, text_gap=0.07)  
                dim_vertical_left(doc, X21, Y21, X22, Y22,  extract_abs(X21,X22) + 50, 'dim', text_height=0.22, text_gap=0.07)                              
                dim_vertical_left(doc, X17, Y17, X21, Y21,  extract_abs(X17,X21) + 50 , 'dim', text_height=0.22, text_gap=0.07)              
                dim_vertical_left(doc, X16, Y16, X17, Y17,  H1 + (R-OP1-OP2-2)/2 + extract_abs(X16,X17) + 50 , 'dim', text_height=0.22, text_gap=0.07)              
                dim_vertical_left(doc, X16, Y16, X1, Y1,  extract_abs(X16,X1) + 200, 'dim', text_height=0.22, text_gap=0.07)                          
            else:    
                dim_vertical_left(doc, X16, Y16, X17, Y17,   H1 + (R-OP1-OP2-2)/2  + 10, 'dim', text_height=0.22,  text_gap=0.07) 
                dim_vertical_left(doc, X16, Y16, X21, Y21,   H1 + (R-OP1-OP2-2)/2  + 100 , 'dim', text_height=0.22,  text_gap=0.07) 
                dim_vertical_left(doc, X16, Y16, X22, Y22,   H1 + (R-OP1-OP2-2)/2  + 150, 'dim', text_height=0.22,  text_gap=0.07) 

            # SO는 치수선이 좌측 열림, 우측열림에 따라 다르게 해야 한다.
            if(JD1+K1_up > JD2+K2_up):       
                # 하부 치수선 4개 노칭 1개 위방향
                dim_linear(doc,  X19, Y19, X16, Y16,  "",  extract_abs(Y16, Y19) + 150 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
                dim_linear(doc,  X16, Y16, X15, Y15,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
                dim_linear(doc,  X10, Y10, X9, Y9,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up) + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
                dim_linear(doc,  X12, Y12, X13, Y13,  "",  50  ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      
                dim_linear(doc,  X16, Y16, X9, Y9,  "",  150 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)              
            else:
                # 하부 치수선 4개 노칭 1개 위방향
                dim_linear(doc,  X19, Y19, X16, Y16,  "",  extract_abs(Y16, Y19) + 180 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
                dim_linear(doc,  X16, Y16, X15, Y15,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up)  + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
                dim_linear(doc,  X10, Y10, X9, Y9,  "",  80 + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      
                dim_linear(doc,  X12, Y12, X13, Y13,  "",  110  ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)      
                dim_linear(doc,  X16, Y16, X9, Y9,  "",  180 + extract_abs(JD1+K1_up, JD2+K2_up)  ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            # 하부 우측 치수선
            if(JD1+K1_up > JD2+K2_up) :
                dim_linear(doc,  X9, Y9, X6, Y6,    "",  160 + extract_abs(JD1+K1_up, JD2+K2_up) + extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            else:
                dim_linear(doc,  X9, Y9, X6, Y6,    "",  180 + extract_abs(JD1+K1_up, JD2+K2_up) - extract_abs(JD1+K1_up, JD2+K2_up) ,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

            # 상판 돌출부위  기본 3.4 치수선
            dim_vertical_right(doc, X21, Y21, X19, Y19, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_left(doc, X4, Y4, X6, Y6, 100, 'dim', text_height=0.22,  text_gap=0.07)  

            # 기둥과의 접선 부위 aligned 표현 각도표현
            dim_linear(doc, X19, Y19, X17, Y17,   "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) # 좌측
            dim_angular(doc, X18, Y18, X19, Y19, X17, Y17, X18, Y18 + 10,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다        
            dim_linear(doc, X8, Y8, X6, Y6,  "", 60,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07) # 좌측
            dim_angular(doc, X8, Y8  , X7, Y7 + 10,  X7, Y7,  X6, Y6,  100, direction="up" )   # up,down, left, right각도를 표시할때는 기존치수선 연장선에 해줘야 제대로 값을 얻는다

            # 상판 K1 K2 값 부분 치수선 
            if(K1_up > 3 or K2_up > 3 or E1> 0 or E2> 0):
                dim_vertical_right(doc, X17, Y17, X18, Y18, 80, 'dim', text_height=0.22,  text_gap=0.07)  
                dim_vertical_left(doc, X8, Y8, X7, Y7, 80, 'dim', text_height=0.22,  text_gap=0.07)  

            # 전개도 상부 치수선
            dim_linear(doc,  X1, Y1, X2, Y2, "", 160,  direction="up", layer='dim')
            dim_linear(doc,  X19, Y19, X6, Y6,  "", extract_abs(Y1,Y19) + 80,  direction="up", layer='dim')
            dim_linear(doc,  X1, Y1, X19, Y19, "",  80,  direction="up", layer='dim')
            dim_linear(doc,  X6, Y6, X2, Y2,  "",  extract_abs(Y2,Y6) + 80,  direction="up", layer='dim')

            # 상부 Vcut 치수 표기
            if(U>0 and G>0):   
               dim_vertical_left(doc, X1+H1+(OP1+OP2+2)/2, Y1, X1+H1+(OP1+OP2+2)/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  

            # 상판 좌우가 다를때 Vcut 상부에 치수선 표기
            # if((JD1 != JD2 or K1_up != K2_up) and G>0):
            #     dim_vertical_left(doc, X1 + R/2, Y1, X1 + R/2, Y4, 100, 'dim', text_height=0.22,  text_gap=0.07)  
            # 문구 설정
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"        
            Textstr = f"({Pre_text}{str(Floor_des)})"       

            X2 = Abs_Xpos + (H1 + R + H2)/2 - len(Textstr)*50/2
            Y2 = Abs_Ypos - JD1/1.5
            draw_Text(doc, X2, Y2, 50, str(Textstr), '레이져')

            # 문구            
            sizedes = Upper_size                  
            
            Textstr = f"Part Name : 상판({Pre_text}{str(Floor_des)})"    
            draw_Text(doc, X9+150, Y9 + 50 - 220, 28, str(Textstr), '0')
            Textstr = f"Mat.Spec : {normal_material}"    
            draw_Text(doc, X9+150, Y9 - 220, 28, str(Textstr), '0')
            Textstr = f"Size : {sizedes}"    
            draw_Text(doc, X9+150, Y9 - 50 - 220 , 28, str(Textstr), '0')
            Textstr = f"Quantity : 1 EA"    
            draw_Text(doc, X9+150, Y9 - 100 - 220 , 28, str(Textstr), '0')                        

            # 도면틀 넣기
            BasicXscale = 2560
            BasicYscale = 1550
            TargetXscale = 800 + R + U * 2 + 200
            TargetYscale = 300 + JD1 + G + U + UW
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            insert_frame(X1 - (U + 580) ,Y1- ( 600 + JD1 + K1_up + G + U + UW) , frame_scale, "top_normal", workplace)        

            ###########################################################################################################################################################################
            # SO멍텅구리 상판 좌측 단면도    (좌우가 다른 모양일때 단면도를 그려준다)
            ###########################################################################################################################################################################
            if(JD1 != JD2 or K1_up != K2_up):
                Section_Xpos = Abs_Xpos - H1 - 400           

                if(G > 0):
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos - G
                    X2 = Section_Xpos
                    Y2 = Abs_Ypos  

                    line(doc, X1, Y1, X2, Y2, layer='0')                    
                    dim_vertical_left(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)     
                    X3 = Section_Xpos + U
                    Y3 = Abs_Ypos  
                    lineto(doc,  X3, Y3, layer='0')    
                    dim_linear(doc,  X2, Y2, X3, Y3, "", 50,  direction="up", layer='dim')                
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                    dim_leader_line(doc, X3, Y3 , X2-130 , Y2-100, "V-cut 2개", layer="dim", style='0')                
                if(G < 1):
                    X1 = Section_Xpos 
                    Y1 = Abs_Ypos 
                    X2 = Section_Xpos
                    Y2 = Abs_Ypos                  
                    X3 = Section_Xpos + U
                    Y3 = Abs_Ypos   
                    line(doc, X2, Y2, X3, Y3, layer='0')                  
                    dim_linear(doc,  X2, Y2, X3, Y3, "", 50,  direction="up", layer='dim')    
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                    dim_leader_line(doc, X3, Y3 , X2-130 , Y2-100, "V-cut 1개", layer="dim", style='0')                                
                
                X4 = X3 
                Y4 = Y3 - JD1
                lineto(doc,  X4, Y4, layer='0')                   
                if(K1_up > 0):
                    X5 = X4 
                    Y5 = Y4 - K1_up
                    lineto(doc,  X5, Y5, layer='0')   
                    dim_vertical_right(doc,  X3, Y3, X4, Y4, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X4, Y4, X5, Y5, 50,'dim' ,text_height=0.22,  text_gap=0.07)     
                    dim_vertical_right(doc,  X3, Y3, X5, Y5, 110,'dim' ,text_height=0.22,  text_gap=0.07)                                     
                else:
                    dim_vertical_right(doc,  X3, Y3, X4, Y4, 110,'dim' ,text_height=0.22,  text_gap=0.07)     
                    X5 = X4 
                    Y5 = Y4

                # drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')         
                X6 = X5 - UW
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')   
                dim_linear(doc,  X6, Y6, X5, Y5,  "", 80,  direction="down", layer='dim')            
                
                if(K1_up > 0):                 
                    # K값 부분 라인 그리기
                    line(doc, X4, Y4, X4 - UW, Y4, layer='22')     
                
                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JD1 != JD2 or K1_up != K2_up ):
                    rectangle(doc, X4+50, Y5+(JD1+K1_up)/2 - 75 , X4 + 50 + 100, Y5+(JD1+K1_up)/2 + 75, layer='0')               
                    rectangle(doc, X4+285,  Y5+(JD1+K1_up)/2 - 75 , X4+285+100, Y5+(JD1+K1_up)/2 + 75 , layer='0')                                                         
                
            ################################################################################################################################################################
            # 멍텅구리 상판 우측 단면도 
            ################################################################################################################################################################
            Section_Xpos = Abs_Xpos + R + H1 + H2 + 400

            if(G>0):
                X1 = Section_Xpos 
                Y1 = Abs_Ypos - G
                X2 = Section_Xpos
                Y2 = Abs_Ypos  

                line(doc, X1, Y1, X2, Y2, layer='0')                    
                dim_vertical_right(doc,  X1, Y1, X2, Y2, 50, 'dim' ,text_height=0.22,  text_gap=0.07)     
                X3 = Section_Xpos - U
                Y3 = Abs_Ypos  
                lineto(doc,  X3, Y3, layer='0')    
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim')
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                dim_leader_line(doc, X3, Y3 , X3+50 , Y2-100, "V-cut 2개", layer="dim", style='0')                

            if(G<1):
                X1 = Section_Xpos 
                Y1 = Abs_Ypos
                X2 = Section_Xpos
                Y2 = Abs_Ypos  
                X3 = Section_Xpos - U
                Y3 = Abs_Ypos  
                line(doc, X1, Y1, X3, Y3, layer='0')                            
                dim_linear(doc,  X3, Y3, X2, Y2,  "", 50,  direction="up", layer='dim')    
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')                         
                dim_leader_line(doc, X3, Y3 , X3 + 50 , Y2-100, "V-cut 1개", layer="dim", style='0')                            
                
            X4 = X3 
            Y4 = Y3 - JD2
            lineto(doc,  X4, Y4, layer='0')   
            if(K1_up > 0):
                X5 = X4 
                Y5 = Y4 - K2_up
                lineto(doc,  X5, Y5, layer='0')   
                dim_vertical_left(doc,  X3, Y3, X4, Y4, 60,'dim' ,text_height=0.22,  text_gap=0.07)                     
                dim_vertical_left(doc,  X4, Y4, X5, Y5, 60,'dim' ,text_height=0.22,  text_gap=0.07)     
                dim_vertical_left(doc,  X3, Y3, X5, Y5, 110,'dim' ,text_height=0.22,  text_gap=0.07)                                     
            else:
                dim_vertical_left(doc,  X3, Y3, X4, Y4, 110,'dim' ,text_height=0.22,  text_gap=0.07)                     
                X5 = X4 
                Y5 = Y4

            X6 = X5 + UW
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='0')   
            dim_linear(doc,  X5, Y5, X6, Y6,   "", 80,  direction="down", layer='dim')
            
            # drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
            
            if(K2_up > 0):                 
                # K값 부분 라인 그리기
                line(doc, X4, Y4, X4 + UW, Y4, layer='22')      

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JD1 != JD2 or K1_up != K2_up ):
                rectangle(doc, X4-50, Y5+(JD2+K2_up)/2 - 75 , X4 - 50 - 100, Y5+(JD2+K2_up)/2 + 75, layer='0')               
                rectangle(doc, X4-285,  Y5+(JD2+K2_up)/2 - 75 , X4-285-100,  Y5+(JD2+K2_up)/2 + 75 , layer='0')                                                         
            
            # 다음 도면 간격 계산
            Abs_Ypos -= 4000

    Abs_Xpos = 43000 + 4000
    Abs_Ypos = 500

    ############################################################################################################################################################################
    # SO멍텅구리 좌기둥 전개도
    ############################################################################################################################################################################
  
    for idx, row in enumerate(data_2d): 
        Floor_mc = row[0]
        Floor_des = row[1]        
        OP1 = row[2]
        OP2 = row[3]
        R = row[4]
        JD1 = row[5]
        JD2 = row[6]
        JB1_up = row[7]
        JB1_down = row[8]
        JB2_up = row[9]
        JB2_down = row[10]
        Top_gap = row[11]
        Side_gap = row[12]
        H1 = row[13]
        H2 = row[14]
        C1 = row[15]        
        C2 = row[16]
        A1 = row[17]
        A2 = row[18]
        LH1 = row[19]
        LH2 = row[20]
        RH1 = row[21]
        RH2 = row[22]        
        U = row[23]        
        G = row[24]        
        UW = row[25]        
        SW_Left = row[26]        
        SW_Right = row[27]        
        K1_up = row[28]        
        K1_down = row[29]        
        K2_up = row[30]        
        K2_down = row[31]                
        Upper_size = row[32]
        Left_side_size = row[33]
        Right_side_size = row[34]
        Original_SW1 = row[35] # 변수명 변경 사이드오픈은 이렇게 적용해야 함
        Original_SW2 = row[36] # 변수명 변경 사이드오픈은 이렇게 적용해야 함
        Original_SWAngle = row[37] # 변수명 변경 사이드오픈은 이렇게 적용해야 함
        E1 = row[38]
        E2 = row[39]        
        Angle = row[43]
        Right_Angle = row[44]
        Top_pop1 = round(row[45] + 2, 1) #소수점 첫째자리까지만 나오게
        Top_pop2 = round(row[46] + 2, 1)         
        Kadapt =  row[56]   # K값 J값에 적용여부 넣기 1이면 넣기 '늘치J에 포함여부'
        G_Setvalue = row[57]  # 손상민 소장 G값 강제로 적용 작지 52번째

        # 상판 E1, E2는 추영덕 소장이 K값을 상판에 표현하는 수치표현식인데, K값으로 환산해서 정리함
        #  K1_up = E1 - UW - 1.2

        # 데이터 검증 및 기본값 설정
        SWAngle = validate_or_default(SWAngle)
        K1_up = validate_or_default(K1_up)
        K1_down = validate_or_default(K1_down)
        K2_up = validate_or_default(K2_up)
        K2_down = validate_or_default(K2_down)
        U = validate_or_default(U)
        G = validate_or_default(G)        
        G_Setvalue = validate_or_default(G_Setvalue)        
        OP1 = validate_or_default(OP1)
        OP2 = validate_or_default(OP2)
        Kadapt = validate_or_default(Kadapt)

        SW1 = Original_SW1
        SW2 = Original_SW2        
        SWAngle = Original_SWAngle
        if(JD1 + K1_up > JD2 + K2_up):
            SWAngle =  0
            SW1 = 0
            SW2 = 0        

        if(Kadapt==1):
            K1_up = 0
            K1_down = 0
            K2_up = 0
            K2_down = 0

        if(OP1>0):  
            if (SWAngle>1):  #SWAngle 0 이상인 경우            
                if(K1_up>0):     
                    if (SWAngle == 90):       
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 3 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = Abs_Xpos + LH2             
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                        lineto(doc,  X5, Y5, layer='레이져')                    
                    else:
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + K1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + K1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + K1_down + SW1  - Bending_rate * 2 - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = Abs_Xpos + LH2             
                        Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                        lineto(doc,  X5, Y5, layer='레이져')                    
                else:
                    if (SWAngle == 90):                     
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 4 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 3  - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4                
                    else:
                        X1 = Abs_Xpos 
                        Y1 = Abs_Ypos + JB1_up + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        X2 = Abs_Xpos + LH2 
                        Y2 = Abs_Ypos + JB1_down + SW1 + SW2 - Bending_rate * 2 - Vcut_rate    
                        line(doc, X1, Y1, X2, Y2, layer='레이져')        
                        dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                        X3 = X2            
                        Y3 = Abs_Ypos + JB1_down + SW1  - Bending_rate * 2  - Vcut_rate      
                        lineto(doc,  X3, Y3, layer='레이져')    
                        X4 = Abs_Xpos + LH2             
                        Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                        lineto(doc,  X4, Y4, layer='레이져')
                        X5 = X4             
                        Y5 = Y4                

            if (SWAngle<1):  #SWAngle 0 이상인 경우
                if(K1_up>0):            
                    X1 = Abs_Xpos 
                    Y1 = Abs_Ypos + JB1_up + K1_up + SW_Left - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + LH2 
                    Y2 = Abs_Ypos + JB1_down + K1_down + SW_Left - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')        
                    dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                    X3 = X2            
                    Y3 = Y2
                    X4 = Abs_Xpos + LH2             
                    Y4 = Abs_Ypos + JB1_down + K1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = Abs_Xpos + LH2             
                    Y5 = Abs_Ypos + JB1_down  - Vcut_rate                
                    lineto(doc,  X5, Y5, layer='레이져')                    
                else:
                    X1 = Abs_Xpos 
                    Y1 = Abs_Ypos + JB1_up + SW_Left - Bending_rate * 2 - Vcut_rate    
                    X2 = Abs_Xpos + LH2 
                    Y2 = Abs_Ypos + JB1_down + SW_Left - Bending_rate * 2 - Vcut_rate    
                    line(doc, X1, Y1, X2, Y2, layer='레이져')        
                    dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                    X3 = X2            
                    Y3 = Y2
                    X4 = Abs_Xpos + LH2             
                    Y4 = Abs_Ypos + JB1_down - Bending_rate  - Vcut_rate    
                    lineto(doc,  X4, Y4, layer='레이져')
                    X5 = X4             
                    Y5 = Y4                

            X6 = Abs_Xpos + LH1             
            Y6 = Abs_Ypos
            lineto(doc,  X6, Y6, layer='레이져')
            if(A1>0):       
                X7 = Abs_Xpos + LH1   
                Y7 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X7, Y7, layer='레이져')                
                X8 = Abs_Xpos + LH1  
                Y8 = Abs_Ypos - C1 - A1 + Vcut_rate * 3
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = Abs_Xpos             
                Y10 = Abs_Ypos - C1 + Vcut_rate * 2
                lineto(doc,  X10, Y10, layer='레이져')

            if(A1<1):       
                X7 = Abs_Xpos + LH1   
                Y7 = Abs_Ypos - C1 + Vcut_rate 
                lineto(doc,  X7, Y7, layer='레이져')                
                X8 = Abs_Xpos + LH1  
                Y8 = Y7
                lineto(doc,  X8, Y8, layer='레이져')
                X9 = Abs_Xpos             
                Y9 = Y8
                lineto(doc,  X9, Y9, layer='레이져')
                X10 = X9
                Y10 = Y9            

            X11 = Abs_Xpos 
            Y11 = Abs_Ypos
            lineto(doc,  X11, Y11, layer='레이져')        

            if (SWAngle>0): 
                if(K1_up>0):  
                    if (SWAngle == 90):                    
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up  - Vcut_rate   
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = Abs_Xpos 
                        Y13 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X13, Y13, layer='레이져')
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 3 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up  - Vcut_rate   
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = Abs_Xpos 
                        Y13 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                        lineto(doc,  X13, Y13, layer='레이져')
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')
                else:  
                    if (SWAngle == 90):   
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate     
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = X12
                        Y13 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate                 
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 3 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                          
                    else:
                        X12 = Abs_Xpos 
                        Y12 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate     
                        lineto(doc,  X12, Y12, layer='레이져')                
                        X13 = X12
                        Y13 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate                 
                        X14 = Abs_Xpos 
                        Y14 = Abs_Ypos + JB1_up + K1_up + SW1  - Bending_rate * 2 - Vcut_rate  
                        lineto(doc,  X14, Y14, layer='레이져')
                        lineto(doc,  X1, Y1, layer='레이져')                                                          
                    
            if (SWAngle<1): 
                if(K1_up>0):  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + JB1_up  - Vcut_rate   
                    lineto(doc,  X12, Y12, layer='레이져')
                    X13 = Abs_Xpos 
                    Y13 = Abs_Ypos + JB1_up + K1_up - Bending_rate  - Vcut_rate   
                    lineto(doc,  X13, Y13, layer='레이져')
                    X14 = Abs_Xpos 
                    Y14 = Abs_Ypos + JB1_up + K1_up + SW_Left  - Bending_rate * 2 - Vcut_rate  
                    lineto(doc,  X14, Y14, layer='레이져')  
                    lineto(doc,  X1, Y1, layer='레이져')    
                else:  
                    X12 = Abs_Xpos 
                    Y12 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate 
                    lineto(doc,  X12, Y12, layer='레이져')
                    X13 = X12
                    Y13 = Abs_Ypos + JB1_up - Bending_rate  - Vcut_rate 
                    X14 = X13
                    Y14 = Y13
                    lineto(doc,  X1, Y1, layer='레이져')   

            # 좌기둥 기준선 
            if(SW1>0):
                line(doc, X14, Y14, X3, Y3,  layer='22')
            line(doc, X13, Y13, X4, Y4,  layer='22')
            if(K1_up>0):
                line(doc, X12, Y12, X5, Y5,  layer='22')
            line(doc, X11, Y11, X6, Y6,  layer='22')
            if(A1>0):
                line(doc, X10, Y10, X7, Y7,  layer='22')                

            # 우측 절곡치수선
            if(SW1>0):
                dim_vertical_right(doc, X2,  Y2, X3, Y3,  extract_abs(X2,X3) + 160, 'dim', text_height=0.22, text_gap=0.01)      
            dim_vertical_right(doc, X3, Y3, X4,  Y4,   extract_abs(X3,X4) + 80, 'dim', text_height=0.22, text_gap=0.01)      
            if(K1_up>0):
                dim_vertical_right(doc, X4,  Y4, X5, Y5,  extract_abs(X4,X5) +  130, 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_right(doc, X6, Y6, X5,  Y5, extract_abs(X6,X5) +  80, 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_right(doc,  X7, Y7, X6,  Y6,   extract_abs(X6,X7) + 80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_right(doc, X7,  Y7, X8, Y8,   extract_abs(X7,X8) + 150, 'dim', text_height=0.22, text_gap=0.01)  
                # V컷 치수표현
                dim_vertical_left(doc, X6,  Y6, X8, Y8,   extract_abs(X6,X8) + 100, 'dim', text_height=0.22, text_gap=0.01)  
            dim_vertical_right(doc, X2,  Y2, X8, Y8,   extract_abs(X2,X8) + 230 , 'dim', text_height=0.22, text_gap=0.01)      

            # 좌측 절곡치수선  
            if(SW1>0):                       
                dim_vertical_left(doc, X1, Y1, X14, Y14,  160, 'dim', text_height=0.22, text_gap=0.01)  
                dim_vertical_left(doc, X13, Y13, X14, Y14,   80 , 'dim', text_height=0.22, text_gap=0.01)  
            else:
                dim_vertical_left(doc, X1, Y1, X13, Y13,   80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(K1_up>0):
                dim_vertical_left(doc, X13, Y13, X12, Y12,  130 , 'dim', text_height=0.22, text_gap=0.01)  
            # JB    
            dim_vertical_left(doc, X11, Y11, X12, Y12, 80 , 'dim', text_height=0.22, text_gap=0.01)  
            # C
            dim_vertical_left(doc, X11, Y11, X10, Y10,  80 , 'dim', text_height=0.22, text_gap=0.01)  
            if(A1>0):
                dim_vertical_left(doc, X9, Y9, X10, Y10,  150 , 'dim', text_height=0.22, text_gap=0.01)
            dim_vertical_left(doc, X1, Y1, X9,  Y9,  230 , 'dim', text_height=0.22, text_gap=0.01)

            # 기둥 하부 치수선 2    
            dim_linear(doc,   X9, Y9, X8, Y8, "", 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)   

            if LH1>LH2:
                dim_linear(doc,  X2, Y2, X6, Y6, "",  100 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.07)       
            if LH1<LH2:    
                dim_linear(doc,  X4, Y4, X8, Y8, "", extract_abs(Y4, Y8) + 100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.07)      

            if args.opt11:
                drawcircle(doc, X1 + 12.5, Y1 - 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X2 - 12.5, Y2 - 5, 2.5 , layer='레이져') # 도장홀 5파이        

            # 전개도 문구        
            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"       

            Xpos = Abs_Xpos + 500 
            Ypos = Abs_Ypos + JB1_up / 2.3
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-좌"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')        

            Xpos = Abs_Xpos + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-좌"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Mat.Spec : {normal_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/1.2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30
            Textstr = f"Size : {Left_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + LH1/1.2
            Ypos = Abs_Ypos + JB1_up / 2.3 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')       

            # 멍텅구리 기둥 좌우 도면틀 넣기
            BasicXscale = 4487
            BasicYscale = 2770
            # 우측단면도가 있을 경우 감안해야 함. 
            if( JB1_up != JB1_down or K1_up != K1_down  or JB2_up != JB2_down or K2_up != K2_down ):                                 
                TargetXscale = 550 + max(LH1, LH2) + 1000
            else:
                TargetXscale = 550 + max(LH1, LH2) + 550

            TargetYscale = 320*4 + (JB1_up + C1 + A1 + K1_up + SW_Left) * 2
            if(TargetXscale/BasicXscale > TargetYscale/BasicYscale ):
                frame_scale = TargetXscale/BasicXscale
            else:
                frame_scale = TargetYscale/BasicYscale
            # print (f'스케일 : {frame_scale}')
            # 우측 단면도가 있는 경우 도면틀 기본점을 아래로 250 내림    
            if( JB1_up != JB1_down or K1_up != K1_down  or JB2_up != JB2_down or K2_up != K2_down ):                                 
                insert_frame(X1 -  710 , Abs_Ypos - (JB2_up+C2+A2+K2_up + SW_Left + 1200 + C1 + A1 ) - 250 , frame_scale, "side_normal", workplace)                
            else:
                insert_frame(X1 -  710 , Abs_Ypos - (JB2_up+C2+A2+K2_up + SW_Left + 1200 + C1 + A1 ) , frame_scale, "side_normal", workplace)                

            ##############################################################################################################################################################
            # 좌기둥 좌측 단면도    
            ##############################################################################################################################################################
            Section_Xpos = Abs_Xpos - 400
            Section_Ypos = Abs_Ypos + JB1_up + K1_up

            # 가정: SWAngle은 도 단위로 주어진 각도
            SWAngle_radians = math.radians(SWAngle)

            if(SWAngle>0):
                X1 = Section_Xpos - SW1 - SW2 * math.cos(SWAngle_radians)
                Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                X2 = Section_Xpos - SW1
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos
                
                line(doc, X2, Y2, X1, Y1, layer='0')               
                line(doc, X2, Y2, X3, Y3, layer='0')        
                dim_linear(doc, X1, Y1, X2, Y2,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # 각도 90도가 아니면 각도 표기
                if (SWAngle != 90):               
                    dim_angular(doc, X2 , Y2, X1, Y1, X2-30, Y2, X3-30, Y3,  160, direction="left" )
                    # help(ezdxf.layouts.Modelspace.add_angular_dim_2l)
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)   

            if(SWAngle<1):
                X1 = Section_Xpos - SW_Left
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos 
                X3 = Section_Xpos 
                Y3 = Section_Ypos            
                line(doc,  X1, Y1, X2, Y2, layer='0')                                                   
                dim_linear(doc, X1, Y1, X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

            if(K1_up>5):
                X4 = X3
                Y4 = Y3 - K1_up
                lineto(doc,  X4, Y4, layer='0')       
            else:
                X4 = X3
                Y4 = Y3

            # Angle은 도 단위로 주어진 각도
            Angle_radians = math.radians(Angle)

            X5 = X4 - JB1_up * math.sin(Angle_radians)                
            Y5 = Y4 - JB1_up * math.cos(Angle_radians)  
            lineto(doc,  X5, Y5, layer='0')    
            X6 = X5 - C1
            Y6 = Y5
            lineto(doc,  X6, Y6, layer='0')   

            if(A1>0):
                X7 = X6
                Y7 = Y6 + A1
                lineto(doc,  X7, Y7, layer='0') 
                drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut 2개소", layer="dim", style='0')
                # A 치수
                dim_vertical_left(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
            if(A1<1):
                X7 = X6
                Y7 = Y6            
                drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                dim_leader_line(doc, X5, Y5 , X5+50, Y5-70, "V-cut", layer="dim", style='0')

            # C값 치수선
            dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
            if(K1_up>5):
                dim_vertical_right(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                # if (Angle > 1):   # 직각이 아닌경우                        
                dim_angular(doc, X4, Y4, X5, Y5, X4, Y4, X3, Y3,  30, direction="left" )

            # JB사선 치수선
            dim_linear(doc, X4, Y4, X5, Y5,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
            # if (Angle > 1):   # 직각이 아닌경우            
            dim_angular(doc, X6, Y6, X5, Y5, X5, Y5, X4, Y4,  20, direction="left" )
            

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB1_up != JB1_down ):
                rectangle(doc, X3+10, Y3 - JB2_up/2 - 75, X3 + 10 + 100, Y3 - JB1_up/2 + 75 , layer='0')               
                rectangle(doc, X3+60+200, Y3 - JB2_up/2 - 75 , X3 + 60 + 300, Y3 - JB1_up/2 + 75 , layer='0')               
            if(K1_up != K1_down ):
                rectangle(doc, X5 + 50, Y5 - 10 , X5 + 50 + 100, Y5 + K1_up + 10 , layer='0')                         
                rectangle(doc, X5 + 10 + 250, Y5 - 10 , X5 + 10 + 300, Y5 + K1_up + 10 , layer='0')                            

            #######################################################################################################################################################################
            # 좌기둥 우측 단면도    
            #######################################################################################################################################################################
            
            if( JB1_up != JB1_down or K1_up != K1_down ):                  
                Section_Xpos = Abs_Xpos + LH1 + 400
                Section_Ypos = Abs_Ypos + JB1_up + K1_up

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(SWAngle>0):  
                    X1 = Section_Xpos + SW1 + SW2 * math.cos(SWAngle_radians)
                    Y1 = Section_Ypos - SW2 * math.sin(SWAngle_radians)

                    X2 = Section_Xpos + SW1
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc, X2, Y2, X1, Y1, layer='0')                   
                    line(doc, X2, Y2, X3, Y3, layer='0')        
                    dim_linear(doc, X2, Y2, X1, Y1,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # 각도 90도가 아니면 각도 표기
                    if (SWAngle != 90):               
                        dim_angular(doc,  X3+5, Y3, X2+5, Y2, X1, Y1, X2 , Y2,  150, direction="right" )
                        # help(ezdxf.layouts.Modelspace.add_angular_dim_2l)
                    dim_linear(doc, X3, Y3,  X2, Y2,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)    

                if(SWAngle<1):  
                    X1 = Section_Xpos + SW_Left
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos 
                    Y2 = Section_Ypos 
                    X3 = Section_Xpos 
                    Y3 = Section_Ypos
                    line(doc,   X1, Y1, X2, Y2,  layer='0')                                   
                    dim_linear(doc,  X2, Y2, X1, Y1,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      

                if(K1_down>5):
                    X4 = X3
                    Y4 = Y3 - K1_down
                    lineto(doc,  X4, Y4, layer='0')       
                else:
                    X4 = X3
                    Y4 = Y3

                # Angle은 도 단위로 주어진 각도
                Angle_radians = math.radians(Angle)

                X5 = X4 + JB1_down * math.sin(Angle_radians)                
                Y5 = Y4 - JB1_down * math.cos(Angle_radians)  
                lineto(doc,  X5, Y5, layer='0')    
                X6 = X5 + C1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')   

                if(A1>0):
                    X7 = X6
                    Y7 = Y6 + A1
                    lineto(doc,  X7, Y7, layer='0') 
                    drawcircle(doc, X5, Y5 , 5 , layer='0', color='6') 
                    drawcircle(doc, X6, Y6 , 5 , layer='0', color='6') 
                    dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut 2개소", layer="dim", style='0')
                    # A 치수
                    dim_vertical_right(doc, X6, Y6, X7, Y7, 50, 'dim', text_height=0.22,  text_gap=0.07) 
                if(A1<1):
                    X7 = X6
                    Y7 = Y6            
                    drawcircle(doc, X5, Y5 , 5 , layer='0', color='6')             
                    dim_leader_line(doc, X5, Y5 , X5-100, Y5-120,  "V-cut", layer="dim", style='0')

                # C값 치수선
                dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                if(K1_up>5):
                    dim_vertical_left(doc, X4, Y4, X3, Y3, 80, 'dim', text_height=0.22, text_gap=0.07)  
                    # if (Angle > 1):   # 직각이 아닌경우                        
                    dim_angular(doc,  X3, Y3, X4 , Y4,  X5, Y5, X4, Y4,   15, direction="right" )

                # JB사선 치수선
                dim_linear(doc,  X5, Y5, X4, Y4,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)       
                # if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc,  X5, Y5, X4, Y4, X5, Y5, X6 , Y6,  15, direction="right" )

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB1_up != JB1_down ):
                    rectangle(doc, X3-10, Y3 - JB1_up/2 - 75, X3 - 10 - 100, Y3 - JB1_up/2 + 75 , layer='0')               
                    rectangle(doc, X3 - 60 - 200, Y3 - JB1_up/2 - 75 , X3 - 60 - 300, Y3 - JB1_up/2 + 75 , layer='0')               
                if(K1_up != K1_down ):
                    rectangle(doc, X5 - 50, Y5 - 10 , X5 - 50 - 100, Y5 + K1_up + 10 , layer='0')                         
                    rectangle(doc, X5 - 10 + 250, Y5 - 10 , X5 - 10 - 300, Y5 + K1_up + 10 , layer='0')                              
            
            ######################################################################################################################################################
            # SO멍텅구리 우기둥 전개도
            ######################################################################################################################################################

            Abs_Ypos -= 2000 
            R_Ypos = Abs_Ypos + 1400 - (C1 + C2 + A1 + A2)

            # 변경 코멘트 2024 03 04 ********** JD가 큰쪽의 기둥 뒷날개 해밍구조 없애기 가림판과 결합되는 부위
            # JD1이 크면 가림판과 결합되는 부위로 SWAngle에 0 적용

            # 우측기둥에서 다시 판단하게 만들려고 Original_SWAngle 변수 만듬
            SW1 = Original_SW1
            SW2 = Original_SW2        
            SWAngle = Original_SWAngle
            # 기둥의 전개도가 큰 쪽은 모깍기가 안되는 것이 원칙이다. 가림판과 결합되는 부분이다.
            if(JD1 + K1_up < JD2 + K2_up):
                SWAngle =  0
                SW1 = 0
                SW2 = 0            

            if(A2>0):            
                X1 = Abs_Xpos 
                Y1 = R_Ypos + C2 + A2 - Vcut_rate * 3
                X2 = Abs_Xpos + RH1 
                Y2 = R_Ypos + C2 + A2 - Vcut_rate * 3
                line(doc, X1, Y1, X2, Y2, layer='레이져')        
                dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                X3 = Abs_Xpos + RH1 
                Y3 = R_Ypos + C2 - Vcut_rate * 2
                lineto(doc,  X3, Y3, layer='레이져')                
            if(A2<1):
                X1 = Abs_Xpos 
                Y1 = R_Ypos + C2  - Vcut_rate 
                X2 = Abs_Xpos + RH1 
                Y2 = Y1
                line(doc, X1, Y1, X2, Y2, layer='레이져')        
                dim_linear(doc,  X1, Y1, X2, Y2, "", 100,  direction="up", layer='dim')    
                X3 = X2
                Y3 = Y2

            X4 = Abs_Xpos + RH1 
            Y4 = R_Ypos 
            lineto(doc,  X4, Y4, layer='레이져')
            if (SWAngle>1):  #SWAngle 0 이상인 경우    헤밍구조인 경우        
                if(K2_up>0):   
                    if (SWAngle == 90):                       
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = Abs_Xpos + RH2 
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = Abs_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')
                    else:
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down +  Vcut_rate
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = Abs_Xpos + RH2 
                        Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                        lineto(doc,  X6, Y6, layer='레이져')
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down - K2_down - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down - K2_down - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up - K2_up - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - K2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = Abs_Xpos 
                        Y12 = R_Ypos - JB2_up +  Vcut_rate
                        lineto(doc,  X12, Y12, layer='레이져')
                if(K2_up<1):   #K값이 없는 경우
                    if (SWAngle == 90):                      
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 3 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 4 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 3 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11
                    else:
                        X5 = Abs_Xpos + RH2 
                        Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                        lineto(doc,  X5, Y5, layer='레이져')    
                        X6 = X5
                        Y6 = Y5                
                        X7 = Abs_Xpos + RH2             
                        Y7 = R_Ypos - JB2_down  - SW1  + Bending_rate * 2 + Vcut_rate                      
                        lineto(doc,  X7, Y7, layer='레이져')            
                        X8 = Abs_Xpos + RH2             
                        Y8 = R_Ypos - JB2_down  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X8, Y8, layer='레이져')
                        X9 = Abs_Xpos 
                        Y9 = R_Ypos - JB2_up  - SW1 - SW2 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X9, Y9, layer='레이져')
                        X10 = Abs_Xpos 
                        Y10 = R_Ypos - JB2_up - SW1 + Bending_rate * 2 + Vcut_rate
                        lineto(doc,  X10, Y10, layer='레이져')
                        X11 = Abs_Xpos 
                        Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                        lineto(doc,  X11, Y11, layer='레이져')
                        X12 = X11
                        Y12 = Y11

            if (SWAngle<1):  #SWAngle 0     일반구조
                if(K2_up>0):   
                    X5 = Abs_Xpos + RH2 
                    Y5 = R_Ypos - JB2_down +  Vcut_rate
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = Abs_Xpos + RH2 
                    Y6 = R_Ypos - JB2_down - K2_down  + Bending_rate + Vcut_rate                         
                    lineto(doc,  X6, Y6, layer='레이져')
                    X7 = Abs_Xpos + RH2             
                    Y7 = R_Ypos - JB2_down - K2_down - SW_Right  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = Abs_Xpos 
                    Y9 = R_Ypos - JB2_up - K2_up - SW_Right  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = Abs_Xpos 
                    Y10 = Y9                
                    X11 = Abs_Xpos 
                    Y11 = R_Ypos - JB2_up - K2_up + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = Abs_Xpos 
                    Y12 = R_Ypos - JB2_up +  Vcut_rate
                    lineto(doc,  X12, Y12, layer='레이져')
                if(K2_up<1):   #K값이 없는 경우
                    X5 = Abs_Xpos + RH2 
                    Y5 = R_Ypos - JB2_down + Bending_rate + Vcut_rate                         
                    lineto(doc,  X5, Y5, layer='레이져')    
                    X6 = X5
                    Y6 = Y5                
                    X7 = Abs_Xpos + RH2             
                    Y7 = R_Ypos - JB2_down  - SW_Right  + Bending_rate * 2 + Vcut_rate                      
                    lineto(doc,  X7, Y7, layer='레이져')            
                    X8 = X7
                    Y8 = Y7
                    lineto(doc,  X8, Y8, layer='레이져')
                    X9 = Abs_Xpos 
                    Y9 = R_Ypos - JB2_up  - SW_Right  + Bending_rate * 2 + Vcut_rate
                    lineto(doc,  X9, Y9, layer='레이져')
                    X10 = X9
                    Y10 = Y9                
                    X11 = Abs_Xpos 
                    Y11 = R_Ypos - JB2_up  + Bending_rate + Vcut_rate   
                    lineto(doc,  X11, Y11, layer='레이져')
                    X12 = X11
                    Y12 = Y11                

            X13 = Abs_Xpos 
            Y13 = R_Ypos
            lineto(doc,  X13, Y13, layer='레이져')

            if(A2>0):
                X14 = Abs_Xpos 
                Y14 = R_Ypos + C2 - Vcut_rate * 2
                lineto(doc,  X14, Y14, layer='레이져')
                lineto(doc,  X1, Y1, layer='레이져')
            if(A2<1):
                X14 = X1
                Y14 = Y1
                lineto(doc,  X14, Y14, layer='레이져')

            # 기둥 절곡선 1
            if(A2>1):        
                line(doc, X14, Y14, X3, Y3,  layer='22')
            # C2    
            line(doc, X13, Y13, X4, Y4,  layer='22')
            # JB2
            line(doc, X12, Y12, X5, Y5,  layer='22')
            if(K2_up>0):
                line(doc, X11, Y11, X6, Y6,  layer='22')
            # 해밍구조
            if (SWAngle>0):   
                line(doc, X10, Y10, X7, Y7,  layer='22')

            # 멍텅구리 우기둥 우측 절곡치수선
            if(A2>1):  
                dim_vertical_right(doc, X3, Y3, X2,  Y2,  150, 'dim', text_height=0.22,  text_gap=0.07)      
                dim_vertical_left(doc, X4, Y4, X2,  Y2, extract_abs(X2,X4) + 150, 'dim', text_height=0.22,  text_gap=0.07)      
            # C    
            dim_vertical_right(doc, X3, Y3, X4,  Y4, 60, 'dim', text_height=0.22,  text_gap=0.07)      
            dim_vertical_right(doc, X5, Y5, X4,  Y4, extract_abs(X6,X4) + 60, 'dim', text_height=0.22,  text_gap=0.07) 
            if(K2_down>0): 
                dim_vertical_right(doc, X6, Y6, X5,  Y5, extract_abs(X6,X4) + 110, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X6, Y6, X7,  Y7, extract_abs(X6,X4) + 60, 'dim', text_height=0.22,  text_gap=0.07)  
            if (SWAngle>0):   
                dim_vertical_right(doc, X7, Y7, X8,  Y8, extract_abs(X6,X4) + 200, 'dim', text_height=0.22,  text_gap=0.07)  
            dim_vertical_right(doc, X2, Y2, X8,  Y8, extract_abs(X2,X8) + 250, 'dim', text_height=0.22,  text_gap=0.07)  

            # 멍텅구리 우기둥 좌측 절곡치수선
            if(A2>1):          
                dim_vertical_left(doc, X1, Y1, X14, Y14, 150 , 'dim', text_height=0.22, text_gap=0.01)  
            # C    
            dim_vertical_left(doc, X13, Y13, X14, Y14,   50 , 'dim', text_height=0.22, text_gap=0.01)  
            # JB
            dim_vertical_left(doc, X13, Y13, X12, Y12,  50 , 'dim', text_height=0.22, text_gap=0.01)              
            if(K2_up>0): 
                dim_vertical_left(doc, X11, Y11, X12, Y12,  extract_abs(X11,X12) + 110, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X11, Y11, X10, Y10,  extract_abs(X11,X10) + 50, 'dim', text_height=0.22, text_gap=0.01)  
            if (SWAngle>0):               
                dim_vertical_left(doc, X9, Y9, X10, Y10,  extract_abs(X9,X10) + 160, 'dim', text_height=0.22, text_gap=0.01)              
            dim_vertical_left(doc, X1, Y1, X9,  Y9,  extract_abs(X1,X9) + 230 , 'dim', text_height=0.22, text_gap=0.01)      

            # 멍텅구리 우기둥 전개도 하부 치수선    
            dim_linear(doc,  X9, Y9, X8, Y8,  "", 100,  direction="down", layer='dim',text_height=0.22, text_gap=0.01)      
            if RH1<RH2:
                dim_linear(doc,  X2, Y2, X5, Y5,  "",  100 ,  direction="up", layer='dim',text_height=0.22, text_gap=0.01)       
            if RH1>RH2:    
                dim_linear(doc,  X8, Y8, X4, Y4, "",  100 ,  direction="down", layer='dim',text_height=0.22, text_gap=0.01)      
    
            if args.opt11:
                drawcircle(doc, X9 + 12.5, Y9 + 5, 2.5 , layer='레이져') # 도장홀 5파이
                drawcircle(doc, X8 - 12.5, Y8 + 5, 2.5 , layer='레이져') # 도장홀 5파이

            if Floor_mc is None or Floor_mc == 0 or (isinstance(Floor_mc, str) and Floor_mc.isspace()):
                Pre_text = ""
            else:
                Pre_text = str(Floor_mc) + "-"

            Xpos = Abs_Xpos + 500 
            Ypos = R_Ypos - JB2_up / 2
            Textstr = f"코팅상"       
            draw_Text(doc, Xpos, Ypos, 40, str(Textstr), '0')         
            Textstr = f"({Pre_text}{str(Floor_des)})-우"         
            draw_Text(doc, Xpos+200, Ypos, 40, str(Textstr), '레이져')           

            Xpos = Abs_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Part Name : Side Jamb({Pre_text}{str(Floor_des)})-우"       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + RH1/2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Mat.Spec : {normal_material} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30
            Textstr = f"Size : {Right_side_size} "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')
            Xpos = Abs_Xpos + RH1/1.2
            Ypos = R_Ypos - JB2_up / 2 + 30 - 40
            Textstr = f"Quantity : 1 EA "       
            draw_Text(doc, Xpos, Ypos, 22, str(Textstr), '0')      
            
            ####################################################################################################################################################################################
            # 멍텅구리 우기둥 좌측 단면도    
            ####################################################################################################################################################################################
            Section_Xpos = Abs_Xpos - 500        
            Section_Ypos = R_Ypos

            SWAngle_radians = math.radians(SWAngle)

            if(A2>0):
                X1 = Section_Xpos
                Y1 = Section_Ypos - A2
                X2 = Section_Xpos 
                Y2 = Section_Ypos 

                line(doc, X1, Y1, X2, Y2, layer='0')                               
                dim_vertical_left(doc, X1, Y1, X2, Y2, 50, 'dim', text_height=0.22, text_gap=0.07)  
                X3 = X2 + C2
                Y3 = Y2
                lineto(doc,  X3, Y3, layer='0')   
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # vcut 표기                
                drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                dim_leader_line(doc, X3, Y3 , X3 + 50, Y3 + 100, "V-cut 2개소", layer="dim", style='0')    

            if(A2<1):
                X1 = Section_Xpos
                Y1 = Section_Ypos 
                X2 = Section_Xpos 
                Y2 = Section_Ypos             
                X3 = X2 + C2
                Y3 = Y2
                line(doc, X1, Y1, X3, Y3, layer='0')               
                dim_linear(doc, X2, Y2, X3, Y3,  "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                # vcut 표기                            
                drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')         
                dim_leader_line(doc, X3, Y3 , X3+50, Y3+100, "V-cut", layer="dim", style='0')                                        

            X4 = X3 + JB2_up * math.sin(Angle_radians)                
            Y4 = Y3 - JB2_up * math.cos(Angle_radians)        
            lineto(doc,  X4, Y4, layer='0')	
            dim_linear(doc, X3, Y3, X4, Y4,  "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)     

            if(K2_up > 5):
                X5 = X4
                Y5 = Y4 - K2_up
                lineto(doc,  X5, Y5, layer='0')	
                dim_vertical_right(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
            else:
                X5 = X4
                Y5 = Y4
                
            if(SWAngle>0):
                X6 = X5 - SW1
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                X7 = X6 - SW2 * math.cos(SWAngle_radians)
                Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                lineto(doc,  X7, Y7, layer='0')
                dim_linear(doc, X6, Y6, X7, Y7,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                dim_angular(doc, X5-5, Y5, X6-5, Y6,  X7, Y7, X6 , Y6,  100, direction="left" )

            if(SWAngle<1): # 헤밍구조가 아닌경우 일반형
                X6 = X5 - SW_Right
                Y6 = Y5
                lineto(doc,  X6, Y6, layer='0')
                dim_linear(doc, X6, Y6, X5, Y5,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                X7 = X6
                Y7 = Y6                        

            # if (Angle > 1):   # 직각이 아닌경우            
            dim_angular( doc,   X3, Y3, X4, Y4, X3, Y3, X2 , Y2,    15, direction="left" )		                            
            if(K2_up > 5): # K2값
                dim_angular(doc,  X4, Y4, X5 , Y5, X4, Y4, X3, Y3,  5, direction="left" )    

            # 서로 치수 차이나는 곳 사각형으로 그려주기
            if(JB2_up != JB2_down ):
                rectangle(doc, X3+60, Y3 - JB2_up/2 - 75, X3 + 60 + 100, Y3 - JB2_up/2 + 75 , layer='0')               
                rectangle(doc, X3+60+300, Y3 - JB2_up/2 - 75 , X3 + 60 + 400, Y3 - JB2_up/2 + 75 , layer='0')               
            if(K2_up != K2_down ):
                rectangle(doc, X5 + 50, Y5 - 10 , X5 + 50 + 100, Y5 + K2_up + 10 , layer='0')                         
                rectangle(doc, X5 + 10 + 250, Y5 - 10 , X5 + 10 + 300, Y5 + K2_up + 10 , layer='0')                            


            ####################################################################################################################################################################################
            # 멍텅구리 우기둥 우측 단면도    
            #################################################################################################################################################################################### 

            if( JB2_up != JB2_down or K2_up != K2_down ):   

                Section_Xpos = Abs_Xpos + max(RH1,RH2) + 500
                Section_Ypos = R_Ypos

                # 가정: SWAngle은 도 단위로 주어진 각도
                SWAngle_radians = math.radians(SWAngle)

                if(A2>0):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos - A2
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos 

                    line(doc, X1, Y1, X2, Y2, layer='0')
                    dim_vertical_right(doc, X2, Y2, X1, Y1,  50, 'dim', text_height=0.22, text_gap=0.07)
                    X3 = Section_Xpos
                    Y3 = Section_Ypos
                    lineto(doc,  X3, Y3, layer='0')
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # vcut 표기                
                    drawcircle(doc, X2, Y2 , 5 , layer='0', color='6') 
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut 2개소", layer="dim", style='0')     

                if(A2<1):
                    X1 = Section_Xpos + C2
                    Y1 = Section_Ypos 
                    X2 = Section_Xpos + C2
                    Y2 = Section_Ypos                 
                    X3 = Section_Xpos
                    Y3 = Section_Ypos 
                    line(doc, X2, Y2, X3, Y3, layer='0')                               
                    dim_linear(doc,  X3, Y3, X2, Y2, "", 50,  direction="up", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    # vcut 표기                                
                    drawcircle(doc, X3, Y3 , 5 , layer='0', color='6')             
                    dim_leader_line(doc, X3, Y3 , X3-100, Y3+120, "V-cut", layer="dim", style='0')                                

                X4 = X3 - JB2_down * math.sin(Angle_radians)                
                Y4 = Y3 - JB2_down * math.cos(Angle_radians)        
                lineto(doc,  X4, Y4, layer='0')	
                dim_linear(doc, X4, Y4, X3, Y3,   "", 80,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)

                if(K2_down > 5):
                    X5 = X4
                    Y5 = Y4 - K2_down
                    lineto(doc,  X5, Y5, layer='0')	
                    dim_vertical_left(doc, X5, Y5, X4, Y4, 50, 'dim', text_height=0.22, text_gap=0.07)  
                else:
                    X5 = X4
                    Y5 = Y4
                    lineto(doc,  X5, Y5, layer='0')	

                if(SWAngle>0):
                    X6 = X5 + SW1
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,  "", 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      

                    X7 = X6 + SW2 * math.cos(SWAngle_radians)
                    Y7 = Y6 + SW2 * math.sin(SWAngle_radians)
                    lineto(doc,  X7, Y7, layer='0')
                    dim_linear(doc, X7, Y7, X6, Y6,  "", 140,  direction="aligned", layer='dim' ,text_height=0.22, text_gap=0.07)               
                    dim_angular(doc,  X7 , Y7, X6, Y6, X6 + 5 , Y6, X6, Y6,  150, direction="right" )
                if(SWAngle<1):
                    X6 = X5 + SW_Right
                    Y6 = Y5
                    lineto(doc,  X6, Y6, layer='0')
                    dim_linear(doc, X5, Y5, X6, Y6,"" , 100,  direction="down", layer='dim' ,text_height=0.22, text_gap=0.07)      
                    X7 = X6
                    Y7 = Y6

                # if (Angle > 1):   # 직각이 아닌경우            
                dim_angular(doc,  X3, Y3, X2-8 , Y2,  X3, Y3, X4, Y4,   25, direction="right")		            
                    # if(K2_down>0):
                dim_angular(doc,   X4, Y4, X3 , Y3, X5, Y5,  X4, Y4, 30, direction="right")

                # 서로 치수 차이나는 곳 사각형으로 그려주기
                if(JB2_up != JB2_down ):
                    rectangle(doc, X3 - 60, Y3 - JB2_up/2 - 75, X3 - 60 - 100, Y3 - JB2_up/2 + 75 , layer='0')               
                    rectangle(doc, X3 - 60 - 350, Y3 - JB2_up/2 - 75 , X3 - 60 - 450, Y3 - JB2_up/2 + 75 , layer='0')               
                if(K2_up != K2_down ):
                    rectangle(doc, X5 - 50, Y5 - 10 , X5 - 50 - 100, Y5 + K2_up + 10 , layer='0')                         
                    rectangle(doc, X5 - 10 + 350, Y5 - 10 , X5 - 10 - 400, Y5 + K2_up + 10 , layer='0')                              
                                    

            Abs_Ypos -= 2000 

# @Gooey(program_name='Jamb cladding 자동작도 프로그램 ver01', tabbed_groups=True, navigation='Tabbed')
@Gooey(encoding='utf-8', program_name='Jamb cladding 자동작도', tabbed_groups=True, navigation='Tabbed', show_success_modal=False,  default_size=(800, 600))
def main():
    global args  #수정할때는 global 선언이 필요하다. 단순히 읽기만 할때는 필요없다.    
    global start_time

    saved_disk_id = load_env_settings()
    current_disk_id = get_current_disk_id()

    if saved_disk_id is None or saved_disk_id != current_disk_id:
        print("라이센스를 구매해 주세요.")

        args = parse_arguments_settings()                    

        if args.config:
            if args.password == '123456':  # '1234'는 원하는 비밀번호로 변경            
                print("라이센스가 설정되었습니다. 다시 시작해 주세요.")
                envsettings() 
                sys.exit()           
            else:
                print("잘못된 비밀번호입니다.")
                print("라이센스를 구매해 주세요.")
                sys.exit()

    else:
        print("환영합니다! 프로그램을 실행합니다.")

        args = parse_arguments()    
        # 숫자로 변수를 선언해 준다.
        args.narrow_sillcut = int(args.narrow_sillcut)  # Convert to integer
        args.narrow_TopPopup = float(args.narrow_TopPopup)  
        # 프로그램 시작 시간 기록
        start_time = time.time()

        print("설정상태 : ")
        print(args)

        # time.sleep(100)
        if args.config:
            if args.password == '1234':  # '1234'는 원하는 비밀번호로 변경            
                print("라이센스가 설정되었습니다. 다시 시작해 주세요.")
                envsettings()            
            else:
                print("잘못된 비밀번호입니다.")
                print("라이센스를 구매해 주세요.")
                sys.exit()

        if not args.config:
            if args.opt1 or args.opt2 or args.opt3:
                execute_narrow()            
            if args.opt4 or args.opt5:
                execute_normal()            
            if args.opt6 or args.opt7:
                execute_wide()
            if args.opt8 or args.opt9:
                execute_sowide()
            if args.opt10 or args.opt11:
                execute_sonormal()            
            if args.opt12 or args.opt13:
                execute_fourcowide()
            if args.opt14 or args.opt15:
                execute_fourconormal()
            if args.opt16 :
                execute_2up()
            if args.opt17 :
                execute_3up()
        
        # HPI bracket 그려주기
        if(HPI_count > 0 ):
            draw_HPI_bracket()

        # display_message()    
        end_time = time.time()

        # 실행 시간 계산
        execution_time = end_time - start_time

        # 시간 형식으로 변환
        hours, remainder = divmod(execution_time, 3600)
        minutes, seconds = divmod(remainder, 60)

        # 결과를 포맷팅하는 부분 수정
        parts = []
        if hours:
            parts.append(f"{int(hours)}h")
        if minutes:
            parts.append(f"{int(minutes)}m")
        parts.append(f"{int(seconds) + 3 }초")  # 초를 '초'로 표시

        formatted_execution_time = " ".join(parts)

        print(f"실행 시간: {formatted_execution_time}")

        # 생성된 파일 이름으로 DXF 파일 저장
        file_path = save_file(workplace)
        doc.saveas(file_path)
        print(f" 저장 파일명: '{file_name}' 저장 완료!")               

def parse_arguments_settings():
    parser = GooeyParser()

    settings = parser.add_argument_group('설정')
    settings.add_argument('--config', action='store_true', default=True,  help='라이센스')
    settings.add_argument('--password', widget='PasswordField', gooey_options={'visible': True, 'show_label': True}, help='비밀번호')

    return parser.parse_args()

def parse_arguments():
    parser = GooeyParser()

    group1 = parser.add_argument_group('쪽쟘')
    # group1.add_argument('--opt1', action='store_true',  help='기본')
    # group1.add_argument('--opt2', action='store_true',  help='도장홀')
    # group1.add_argument('--opt3', action='store_true',  help='쪽쟘 상판끝 라운드(추영덕소장)')    
    group1.add_argument('--opt1', action='store_true', default=True, help='기본')
    group1.add_argument('--opt1_1', action='store_true', default=False, help='상판/기둥 Vcut없음')
    group1.add_argument('--opt2', action='store_true', default=False, help='도장홀')
    group1.add_argument('--opt3', action='store_true', default=False, help='쪽쟘 상판끝 라운드(추영덕소장)')
    group1.add_argument('--narrow_sillcut', widget='TextField', default='0', gooey_options={'validator': {'test': 'user_input.isdigit()', 'message': 'Please enter a number'}}, help='씰간격')
    group1.add_argument('--narrow_TopPopup', widget='TextField', default='4', gooey_options={'validator': {'test': 'user_input.isdigit()', 'message': 'UD방향 밑부분 돌출 기본4mm'}}, help='상판밑 UD방향 돌출값')

    group2 = parser.add_argument_group('멍텅(막판무)')
    # group2.add_argument('--opt4', action='store_true',  help='멍텅구리')
    # group2.add_argument('--opt5', action='store_true', help='도장홀')
    group2.add_argument('--opt4', action='store_true', default=True,  help='멍텅구리')
    group2.add_argument('--opt5', action='store_true', default=False, help='도장홀')

    group3 = parser.add_argument_group('와이드(막판유)')
    group3.add_argument('--opt6', action='store_true', default=True,  help='와이드(광폭)')
    group3.add_argument('--opt7', action='store_true', default=False, help='도장홀')

    group4 = parser.add_argument_group('SO(와이드)')    
    group4.add_argument('--opt8', action='store_true', default=True, help='SO와이드(광폭)')    
    group4.add_argument('--opt9', action='store_true', default=False, help='도장홀')    

    group5 = parser.add_argument_group('SO(멍텅)')    
    group5.add_argument('--opt10', action='store_true', default=True, help='SO멍텅구리')    
    group5.add_argument('--opt11', action='store_true', default=False, help='도장홀')    

    group6 = parser.add_argument_group('4CO(와이드)')    
    group6.add_argument('--opt12', action='store_true', default=True, help='4CO(와이드)')    
    group6.add_argument('--opt13', action='store_true', default=False, help='도장홀')    

    group7 = parser.add_argument_group('4CO(멍텅)')    
    group7.add_argument('--opt14', action='store_true', default=True, help='4CO(멍텅구리)')    
    group7.add_argument('--opt15', action='store_true', default=False, help='도장홀')    

    group8 = parser.add_argument_group('2UP')    
    group8.add_argument('--opt16', action='store_true', default=True, help='2UP(멍텅구리)')    

    group9 = parser.add_argument_group('3UP')    
    group9.add_argument('--opt17', action='store_true', default=True, help='3UP(멍텅구리)')    

    settings = parser.add_argument_group('설정')
    settings.add_argument('--config', action='store_true', help='라이센스')
    settings.add_argument('--password', widget='PasswordField', gooey_options={'visible': True, 'show_label': True}, help='비밀번호')

    return parser.parse_args()

def envsettings():
    # 하드디스크 고유번호를 가져오는 코드 (시스템에 따라 다를 수 있음)
    disk_id = os.popen('wmic diskdrive get serialnumber').read().strip()
    data = {"DiskID": disk_id}
    with open(license_file_path, 'w', encoding='utf-8') as file:
        json.dump(data, file)

    # print("환경설정이 저장되었습니다.")
        
if __name__ == '__main__':
    main()
    sys.exit()